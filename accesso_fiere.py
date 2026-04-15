#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════╗
║   ACCESSO FIERE v1  –  Gestionale Allestitori Fieristici  ║
║   Gestionale per allestitori fieristici: Fiere, Personale,              ║
║   Veicoli, Documenti, Preventivi, Report mensile          ║
╚══════════════════════════════════════════════════════════╝
Installa:  pip install flask openpyxl
Avvio:     python gestionale.py
"""

from flask import (Flask, render_template_string, request,
                   redirect, url_for, session, jsonify, flash, Response, send_file)
from datetime import datetime, date, timedelta
import sqlite3, hashlib, os, json, smtplib, io, zipfile, base64, mimetypes, threading, time
from functools import wraps
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.utils import secure_filename

# Lock cross-process per SQLite (funziona con gunicorn multi-worker)
_db_write_lock = threading.Lock()

def _db_commit(conn):
    """Commit con retry automatico su database locked."""
    for attempt in range(10):
        try:
            conn.commit()
            return
        except sqlite3.OperationalError as e:
            if 'locked' in str(e) and attempt < 9:
                time.sleep(0.2 * (attempt + 1))
            else:
                raise

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    import anthropic
    AI_OK = True
except ImportError:
    AI_OK = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'accesso-fiere-secret-2025-changeme')
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
# Su Railway usa /data (volume persistente), in locale usa la cartella del file
if os.path.isdir('/data'):
    DATA_DIR = '/data'
else:
    DATA_DIR = BASE_DIR

# ── MASTER DB (SaaS) ─────────────────────────────────────────
# Contiene aziende, abbonamenti, piani — separato dai DB tenant
MASTER_DB = os.path.join(DATA_DIR, 'accesso_fiere_master.db')

def get_master_db():
    conn = sqlite3.connect(MASTER_DB, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn

def init_master_db():
    db = get_master_db()
    db.executescript("""
    CREATE TABLE IF NOT EXISTS aziende (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        email_admin TEXT UNIQUE NOT NULL,
        password_admin TEXT NOT NULL,
        piano TEXT DEFAULT 'base',
        stato TEXT DEFAULT 'trial',
        stripe_customer_id TEXT,
        stripe_subscription_id TEXT,
        data_scadenza TEXT,
        trial_fino_al TEXT,
        creato_il TEXT DEFAULT (datetime('now')),
        slug TEXT UNIQUE
    );
    CREATE TABLE IF NOT EXISTS piani (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        stripe_price_id TEXT,
        prezzo_mensile REAL,
        max_dipendenti INTEGER DEFAULT 10,
        descrizione TEXT
    );
    """)
    # Inserisci piani di default
    existing = db.execute("SELECT COUNT(*) FROM piani").fetchone()[0]
    if not existing:
        db.execute("INSERT INTO piani (nome,prezzo_mensile,max_dipendenti,descrizione) VALUES (?,?,?,?)",
                   ('Base', 29.0, 5, 'Fino a 5 operatori - ideale per piccole squadre'))
        db.execute("INSERT INTO piani (nome,prezzo_mensile,max_dipendenti,descrizione) VALUES (?,?,?,?)",
                   ('Professional', 59.0, 20, 'Fino a 20 operatori - per squadre medie'))
        db.execute("INSERT INTO piani (nome,prezzo_mensile,max_dipendenti,descrizione) VALUES (?,?,?,?)",
                   ('Enterprise', 99.0, 999, 'Dipendenti illimitati'))
    db.commit(); db.close()

# ── TENANT DB ────────────────────────────────────────────────
# Ogni azienda ha il suo DB isolato
def get_tenant_db_path(azienda_id):
    p = os.path.join(DATA_DIR, 'tenants')
    os.makedirs(p, exist_ok=True)
    return os.path.join(p, f'accesso_fiere_{azienda_id}.db')

# DB corrente = dipende dall'azienda in sessione (o DB legacy se non SaaS)
DB_LEGACY = os.path.join(DATA_DIR, 'accesso_fiere.db')

def get_current_db_path():
    """Restituisce il path DB corretto per il tenant corrente."""
    azienda_id = session.get('azienda_id')
    if azienda_id:
        return get_tenant_db_path(azienda_id)
    # Fallback al DB legacy (uso diretto senza SaaS)
    return DB_LEGACY

# Ridefinisco DB come property dinamica
DB = DB_LEGACY  # mantenuto per compatibilità con codice esistente

UPLOAD_DIR         = os.path.join(DATA_DIR, 'uploads_dipendenti')
UPLOAD_DIR_VEICOLI = os.path.join(DATA_DIR, 'uploads_veicoli')
UPLOAD_DIR_EVENTI  = os.path.join(DATA_DIR, 'uploads_eventi')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR_VEICOLI, exist_ok=True)
os.makedirs(UPLOAD_DIR_EVENTI,  exist_ok=True)
ALLOWED_EXT = {'pdf','png','jpg','jpeg','gif','webp','bmp'}
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20 MB max

def allowed_file(fn):
    return '.' in fn and fn.rsplit('.',1)[1].lower() in ALLOWED_EXT

def get_upload_path(uid):
    p = os.path.join(UPLOAD_DIR, str(uid))
    os.makedirs(p, exist_ok=True)
    return p

def get_upload_path_veicolo(vid):
    p = os.path.join(UPLOAD_DIR_VEICOLI, str(vid))
    os.makedirs(p, exist_ok=True)
    return p

def get_upload_path_evento(eid):
    p = os.path.join(UPLOAD_DIR_EVENTI, str(eid))
    os.makedirs(p, exist_ok=True)
    return p
    return p


def analizza_documento_ai(file_path, nome_file, tipo_doc_hint='', uid=None):
    """Analizza documento dipendente con AI e restituisce dati estratti."""
    api_key = get_setting('anthropic_api_key', '')
    if not api_key:
        return {}
    try:
        import anthropic, base64 as _b64
        client = anthropic.Anthropic(api_key=api_key)
        with open(file_path, 'rb') as fh:
            pdf_data = _b64.standard_b64encode(fh.read()).decode()
        cb = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_data}}
        prompt = """Analizza questo documento di un dipendente e rispondi SOLO con JSON valido, senza markdown.
Estrai: tipo_doc (una tra: Carta d'identità, Passaporto, Patente, Permesso di soggiorno, Codice fiscale, Contratto, Certificato medico, Attestato formazione, DURC, Altro), data_scadenza (YYYY-MM-DD se presente), nome_cognome (nome e cognome se visibili), note (breve descrizione 1 frase).
Formato esatto: {"tipo_doc":"","data_scadenza":"","nome_cognome":"","note":""}
Se un campo non esiste metti "". Date in formato YYYY-MM-DD."""
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=512,
            messages=[{"role": "user", "content": [cb, {"type": "text", "text": prompt}]}])
        raw = resp.content[0].text.strip().strip('`').lstrip('json').strip()
        return json.loads(raw)
    except Exception as e:
        return {'_error': str(e)}

# ─────────────────────────────────────────────
#  DATABASE
# ─────────────────────────────────────────────
def get_db():
    db_path = get_current_db_path()
    conn = sqlite3.connect(db_path, timeout=60, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA busy_timeout=60000")
    return conn

def safe_commit(conn):
    import time
    for i in range(10):
        try:
            conn.commit()
            return
        except sqlite3.OperationalError as e:
            if 'locked' in str(e):
                time.sleep(0.3 * (i + 1))
            else:
                raise
    conn.commit()  # ultimo tentativo, se fallisce lascia propagare

def init_db():
    db = get_db()  # già con WAL e timeout 30s
    db.executescript("""
    CREATE TABLE IF NOT EXISTS utenti (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL, cognome TEXT NOT NULL,
        email TEXT UNIQUE NOT NULL, password TEXT NOT NULL,
        ruolo TEXT DEFAULT 'dipendente',
        titolo TEXT, telefono TEXT, data_assunzione TEXT,
        attivo INTEGER DEFAULT 1,
        creato_il TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS cantieri (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        indirizzo TEXT,
        attivo INTEGER DEFAULT 1,
        creato_il TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS presenze (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        utente_id INTEGER NOT NULL,
        data TEXT NOT NULL,
        ora_entrata TEXT, ora_uscita TEXT, ore_totali REAL,
        cantiere_id INTEGER,
        note TEXT,
        FOREIGN KEY(utente_id) REFERENCES utenti(id),
        FOREIGN KEY(cantiere_id) REFERENCES cantieri(id)
    );
    CREATE TABLE IF NOT EXISTS richieste_presenze (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        utente_id INTEGER NOT NULL,
        data TEXT NOT NULL,
        ora_entrata TEXT NOT NULL, ora_uscita TEXT NOT NULL,
        ore_totali REAL,
        cantiere_id INTEGER,
        note TEXT,
        stato TEXT DEFAULT 'in_attesa',
        nota_admin TEXT,
        creato_il TEXT DEFAULT (datetime('now')),
        gestito_il TEXT,
        FOREIGN KEY(utente_id) REFERENCES utenti(id)
    );
    CREATE TABLE IF NOT EXISTS ferie_permessi (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        utente_id INTEGER NOT NULL,
        tipo TEXT NOT NULL,
        data_inizio TEXT NOT NULL,
        data_fine TEXT NOT NULL,
        giorni INTEGER,
        motivo TEXT,
        stato TEXT DEFAULT 'in_attesa',
        nota_admin TEXT,
        creato_il TEXT DEFAULT (datetime('now')),
        gestito_il TEXT,
        FOREIGN KEY(utente_id) REFERENCES utenti(id)
    );
    CREATE TABLE IF NOT EXISTS documenti (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titolo TEXT NOT NULL, descrizione TEXT, categoria TEXT,
        data_scadenza TEXT, creato_da INTEGER, assegnato_a INTEGER,
        creato_il TEXT DEFAULT (datetime('now')),
        FOREIGN KEY(creato_da) REFERENCES utenti(id),
        FOREIGN KEY(assegnato_a) REFERENCES utenti(id)
    );
    CREATE TABLE IF NOT EXISTS eventi (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titolo TEXT NOT NULL, descrizione TEXT,
        data_inizio TEXT NOT NULL, data_fine TEXT,
        colore TEXT DEFAULT '#e63946',
        tutto_il_giorno INTEGER DEFAULT 1,
        creato_da INTEGER,
        FOREIGN KEY(creato_da) REFERENCES utenti(id)
    );
    CREATE TABLE IF NOT EXISTS impostazioni (
        chiave TEXT PRIMARY KEY,
        valore TEXT
    );
    """)
    # ── Migrazione automatica colonne mancanti ──
    migrations = [
        "ALTER TABLE presenze ADD COLUMN cantiere_id INTEGER",
        "ALTER TABLE richieste_presenze ADD COLUMN cantiere_id INTEGER",
        "ALTER TABLE utenti ADD COLUMN mansione TEXT",
        "ALTER TABLE documenti ADD COLUMN doc_dipendente_id INTEGER",
        "ALTER TABLE cantieri ADD COLUMN ente_organizzatore TEXT",
        "ALTER TABLE cantieri ADD COLUMN citta TEXT",
        "ALTER TABLE cantieri ADD COLUMN padiglione TEXT",
        "ALTER TABLE cantieri ADD COLUMN numero_stand TEXT",
        "ALTER TABLE cantieri ADD COLUMN superficie_mq REAL",
        "ALTER TABLE cantieri ADD COLUMN data_inizio TEXT",
        "ALTER TABLE cantieri ADD COLUMN data_fine TEXT",
        "ALTER TABLE cantieri ADD COLUMN responsabile TEXT",
        "ALTER TABLE cantieri ADD COLUMN tipo_allestimento TEXT DEFAULT 'standard'",
        "ALTER TABLE cantieri ADD COLUMN note_logistica TEXT",
        """CREATE TABLE IF NOT EXISTS veicoli (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            targa TEXT NOT NULL,
            marca TEXT,
            modello TEXT,
            tipo TEXT DEFAULT 'Auto',
            anno INTEGER,
            colore TEXT,
            note TEXT,
            attivo INTEGER DEFAULT 1,
            creato_il TEXT DEFAULT (datetime('now')),
            scad_assicurazione TEXT,
            scad_revisione TEXT,
            scad_bollo TEXT,
            scad_tagliando TEXT,
            note_assicurazione TEXT,
            note_revisione TEXT,
            note_bollo TEXT,
            note_tagliando TEXT
        )""",
        "ALTER TABLE veicoli ADD COLUMN scad_assicurazione TEXT",
        "ALTER TABLE veicoli ADD COLUMN scad_revisione TEXT",
        "ALTER TABLE veicoli ADD COLUMN scad_bollo TEXT",
        "ALTER TABLE veicoli ADD COLUMN scad_tagliando TEXT",
        "ALTER TABLE veicoli ADD COLUMN note_assicurazione TEXT",
        "ALTER TABLE veicoli ADD COLUMN note_revisione TEXT",
        "ALTER TABLE veicoli ADD COLUMN note_bollo TEXT",
        "ALTER TABLE veicoli ADD COLUMN note_tagliando TEXT",
        """CREATE TABLE IF NOT EXISTS documenti_veicoli (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            veicolo_id INTEGER NOT NULL,
            nome_originale TEXT NOT NULL,
            nome_file TEXT NOT NULL,
            tipo_doc TEXT DEFAULT 'Altro',
            dimensione INTEGER,
            note TEXT,
            data_scadenza TEXT,
            caricato_il TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(veicolo_id) REFERENCES veicoli(id)
        )""",
        """CREATE TABLE IF NOT EXISTS documenti_dipendente (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            utente_id INTEGER NOT NULL,
            nome_originale TEXT NOT NULL,
            nome_file TEXT NOT NULL,
            tipo_doc TEXT DEFAULT 'Generico',
            dimensione INTEGER,
            note TEXT,
            data_scadenza TEXT,
            caricato_il TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(utente_id) REFERENCES utenti(id)
        )""",
        """CREATE TABLE IF NOT EXISTS pos_documenti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titolo TEXT NOT NULL,
            cantiere_nome TEXT,
            cantiere_indirizzo TEXT,
            committente TEXT,
            responsabile_lavori TEXT,
            data_inizio TEXT,
            data_fine TEXT,
            data_documento TEXT,
            num_lavoratori INTEGER DEFAULT 0,
            personale_json TEXT,
            opere TEXT,
            note_cantiere TEXT,
            creato_il TEXT DEFAULT (datetime('now')),
            creato_da INTEGER
        )""",
        "ALTER TABLE impostazioni ADD COLUMN valore2 TEXT",
        "INSERT OR IGNORE INTO impostazioni (chiave,valore) VALUES ('sede_legale','')",
        "INSERT OR IGNORE INTO impostazioni (chiave,valore) VALUES ('partita_iva','')",
        "INSERT OR IGNORE INTO impostazioni (chiave,valore) VALUES ('codice_inail','')",
        "INSERT OR IGNORE INTO impostazioni (chiave,valore) VALUES ('codice_inps','')",
        "INSERT OR IGNORE INTO impostazioni (chiave,valore) VALUES ('cciaa','')",
        "INSERT OR IGNORE INTO impostazioni (chiave,valore) VALUES ('rspp','')",
        "INSERT OR IGNORE INTO impostazioni (chiave,valore) VALUES ('medico_competente','')",
        "INSERT OR IGNORE INTO impostazioni (chiave,valore) VALUES ('attivita_azienda','')",
        """CREATE TABLE IF NOT EXISTS clienti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            partita_iva TEXT,
            indirizzo TEXT,
            email TEXT,
            telefono TEXT,
            note TEXT,
            creato_il TEXT DEFAULT (datetime('now'))
        )""",
        """CREATE TABLE IF NOT EXISTS fatture (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            cliente_id INTEGER,
            cliente_nome TEXT,
            data_emissione TEXT,
            data_scadenza TEXT,
            importo_totale REAL DEFAULT 0,
            imponibile REAL DEFAULT 0,
            iva_perc REAL DEFAULT 22,
            iva_importo REAL DEFAULT 0,
            descrizione TEXT,
            note TEXT,
            condizioni_pagamento TEXT DEFAULT '30gg',
            stato TEXT DEFAULT 'da_pagare',
            file_nome TEXT,
            creato_il TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(cliente_id) REFERENCES clienti(id)
        )""",
        """CREATE TABLE IF NOT EXISTS rate_fattura (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fattura_id INTEGER NOT NULL,
            numero_rata INTEGER DEFAULT 1,
            importo REAL DEFAULT 0,
            data_scadenza TEXT,
            data_pagamento TEXT,
            stato TEXT DEFAULT 'da_pagare',
            note TEXT,
            FOREIGN KEY(fattura_id) REFERENCES fatture(id)
        )""",
        "ALTER TABLE fatture ADD COLUMN file_path TEXT",
        "ALTER TABLE fatture ADD COLUMN preventivo_id INTEGER",
        "ALTER TABLE fatture ADD COLUMN imponibile_lordo REAL DEFAULT 0",
        "ALTER TABLE fatture ADD COLUMN sconto_importo REAL DEFAULT 0",
        "ALTER TABLE clienti ADD COLUMN codice_fiscale TEXT",
        "ALTER TABLE clienti ADD COLUMN citta TEXT",
        "ALTER TABLE clienti ADD COLUMN cap TEXT",
        "ALTER TABLE clienti ADD COLUMN paese TEXT DEFAULT 'Italia'",
        "ALTER TABLE clienti ADD COLUMN sito_web TEXT",
        "ALTER TABLE clienti ADD COLUMN referente TEXT",
        "ALTER TABLE utenti ADD COLUMN paga_oraria_netta REAL DEFAULT 0",
        "ALTER TABLE utenti ADD COLUMN paga_base_netta REAL DEFAULT 0",
        "ALTER TABLE utenti ADD COLUMN ore_soglia_mensile REAL DEFAULT 0",
        "ALTER TABLE utenti ADD COLUMN google_id TEXT",
        "ALTER TABLE utenti ADD COLUMN avatar_url TEXT",
        "ALTER TABLE presenze ADD COLUMN evento_id INTEGER",
        """CREATE TABLE IF NOT EXISTS spese_evento (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evento_id INTEGER NOT NULL,
            categoria TEXT NOT NULL,
            descrizione TEXT,
            importo REAL DEFAULT 0,
            data TEXT,
            nota TEXT,
            creato_da INTEGER,
            creato_il TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(evento_id) REFERENCES eventi(id)
        )""",
        """CREATE TABLE IF NOT EXISTS eventi_dipendenti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evento_id INTEGER NOT NULL,
            dipendente_id INTEGER NOT NULL,
            ruolo TEXT DEFAULT 'generico',
            FOREIGN KEY(evento_id) REFERENCES eventi(id),
            FOREIGN KEY(dipendente_id) REFERENCES utenti(id)
        )""",
        "ALTER TABLE eventi ADD COLUMN tipo TEXT DEFAULT 'lavoro'",
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_ev_dip_unique ON eventi_dipendenti(evento_id, dipendente_id)",
        "ALTER TABLE eventi ADD COLUMN luogo TEXT",
        """CREATE TABLE IF NOT EXISTS eventi_veicoli (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evento_id INTEGER NOT NULL,
            veicolo_id INTEGER NOT NULL,
            note TEXT,
            FOREIGN KEY(evento_id) REFERENCES eventi(id),
            FOREIGN KEY(veicolo_id) REFERENCES veicoli(id)
        )""",
        "ALTER TABLE eventi ADD COLUMN cantiere_id INTEGER",
        """CREATE TABLE IF NOT EXISTS eventi_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            evento_id INTEGER NOT NULL,
            categoria TEXT NOT NULL DEFAULT 'disegno',
            nome_originale TEXT NOT NULL,
            nome_file TEXT NOT NULL,
            dimensione INTEGER,
            note TEXT,
            caricato_da INTEGER,
            caricato_il TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(evento_id) REFERENCES eventi(id),
            FOREIGN KEY(caricato_da) REFERENCES utenti(id)
        )""",
        """CREATE TABLE IF NOT EXISTS preventivi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            data_emissione TEXT,
            validita TEXT DEFAULT '30 giorni',
            stato TEXT DEFAULT 'bozza',
            cliente_id INTEGER,
            cliente_nome TEXT,
            cliente_piva TEXT,
            cliente_indirizzo TEXT,
            oggetto TEXT,
            descrizione_servizio TEXT,
            data_inizio_lavoro TEXT,
            data_fine_lavoro TEXT,
            luogo_lavoro TEXT,
            iva INTEGER DEFAULT 22,
            imponibile REAL DEFAULT 0,
            totale_finale REAL DEFAULT 0,
            note TEXT,
            pagamento TEXT,
            creato_il TEXT DEFAULT (datetime('now'))
        )""",
        "ALTER TABLE preventivi ADD COLUMN data_inizio_lavoro TEXT",
        "ALTER TABLE preventivi_voci ADD COLUMN data_inizio TEXT",
        "ALTER TABLE preventivi_voci ADD COLUMN data_fine TEXT",
        "ALTER TABLE preventivi_voci ADD COLUMN importo_modificato REAL",
        "ALTER TABLE preventivi_voci ADD COLUMN sconto_riga REAL DEFAULT 0",
        "ALTER TABLE preventivi ADD COLUMN data_fine_lavoro TEXT",
        "ALTER TABLE preventivi ADD COLUMN luogo_lavoro TEXT",
        "ALTER TABLE eventi ADD COLUMN preventivo_id INTEGER",
        """CREATE TABLE IF NOT EXISTS preventivi_voci (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            preventivo_id INTEGER NOT NULL,
            descrizione TEXT,
            quantita REAL,
            unita_misura TEXT,
            prezzo_unitario REAL,
            totale REAL,
            ordine INTEGER DEFAULT 0,
            FOREIGN KEY(preventivo_id) REFERENCES preventivi(id)
        )""",
        """CREATE TABLE IF NOT EXISTS documenti_azienda (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titolo TEXT NOT NULL,
            categoria TEXT,
            note TEXT,
            data_emissione TEXT,
            data_scadenza TEXT,
            avviso_giorni INTEGER DEFAULT 30,
            file_nome TEXT,
            file_path TEXT,
            creato_il TEXT DEFAULT (datetime('now')),
            aggiornato_il TEXT DEFAULT (datetime('now'))
        )""",
        """CREATE TABLE IF NOT EXISTS spese_rimborso (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            utente_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            categoria TEXT NOT NULL,
            descrizione TEXT,
            importo REAL NOT NULL DEFAULT 0,
            veicolo_id INTEGER,
            foto_nome TEXT,
            foto_path TEXT,
            stato TEXT DEFAULT 'in_attesa',
            note_admin TEXT,
            creato_il TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(utente_id) REFERENCES utenti(id),
            FOREIGN KEY(veicolo_id) REFERENCES veicoli(id)
        )""",
        "ALTER TABLE eventi_files ADD COLUMN dipendente_id INTEGER REFERENCES utenti(id)",
        "ALTER TABLE presenze ADD COLUMN nome_jolly TEXT",
        "ALTER TABLE presenze ADD COLUMN cognome_jolly TEXT",
        "ALTER TABLE presenze ADD COLUMN evento_id INTEGER",
        "ALTER TABLE eventi_dipendenti ADD COLUMN data_da TEXT",
        "ALTER TABLE eventi_dipendenti ADD COLUMN data_a TEXT",
        "ALTER TABLE eventi_dipendenti ADD COLUMN data_da TEXT",
        "ALTER TABLE eventi_dipendenti ADD COLUMN data_a TEXT",
        # Rimuove NOT NULL da utente_id per supportare i jolly (SQLite richiede recreate)
        """CREATE TABLE IF NOT EXISTS presenze_v2 (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            utente_id INTEGER,
            data TEXT NOT NULL,
            ora_entrata TEXT, ora_uscita TEXT, ore_totali REAL,
            cantiere_id INTEGER,
            note TEXT,
            evento_id INTEGER,
            nome_jolly TEXT,
            cognome_jolly TEXT,
            FOREIGN KEY(utente_id) REFERENCES utenti(id),
            FOREIGN KEY(cantiere_id) REFERENCES cantieri(id)
        )""",
        """INSERT OR IGNORE INTO presenze_v2 (id,utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note,evento_id,nome_jolly,cognome_jolly)
           SELECT id,utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note,
                  CASE WHEN EXISTS(SELECT 1 FROM pragma_table_info('presenze') WHERE name='evento_id') THEN evento_id ELSE NULL END,
                  CASE WHEN EXISTS(SELECT 1 FROM pragma_table_info('presenze') WHERE name='nome_jolly') THEN nome_jolly ELSE NULL END,
                  CASE WHEN EXISTS(SELECT 1 FROM pragma_table_info('presenze') WHERE name='cognome_jolly') THEN cognome_jolly ELSE NULL END
           FROM presenze""",
        "DROP TABLE IF EXISTS presenze_old",
        "ALTER TABLE presenze RENAME TO presenze_old",
        "ALTER TABLE presenze_v2 RENAME TO presenze",
        "DROP TABLE IF EXISTS presenze_old",
        """CREATE TABLE IF NOT EXISTS detrazioni_dipendente (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            utente_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'altro',
            descrizione TEXT,
            importo REAL NOT NULL,
            mese_competenza TEXT,
            note TEXT,
            creato_il TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(utente_id) REFERENCES utenti(id)
        )""",
        """CREATE TABLE IF NOT EXISTS jolly_paghe (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_jolly TEXT NOT NULL,
            cognome_jolly TEXT NOT NULL,
            paga_oraria REAL DEFAULT 0,
            aggiornato_il TEXT DEFAULT (datetime('now')),
            UNIQUE(nome_jolly, cognome_jolly)
        )""",
    ]
    for sql in migrations:
        try: db.execute(sql)
        except: pass
    safe_commit(db)

    # Seed
    pw_admin = hashlib.sha256('admin123'.encode()).hexdigest()
    pw_emp   = hashlib.sha256('pass123'.encode()).hexdigest()
    try:
        db.execute("INSERT INTO utenti (nome,cognome,email,password,ruolo,titolo,data_assunzione) VALUES (?,?,?,?,?,?,?)",
                   ('Admin','Sistema','admin@accessofiere.it',pw_admin,'admin','Resp. Allestimenti','2024-01-01'))
        db.execute("INSERT INTO utenti (nome,cognome,email,password,ruolo,titolo,data_assunzione) VALUES (?,?,?,?,?,?,?)",
                   ('Marco','Rossi','marco@accessofiere.it',pw_emp,'dipendente','Capo Montaggio','2024-03-01'))
        db.execute("INSERT INTO utenti (nome,cognome,email,password,ruolo,titolo,data_assunzione) VALUES (?,?,?,?,?,?,?)",
                   ('Elena','Bianchi','elena@accessofiere.it',pw_emp,'dipendente','Allestitorice','2024-02-15'))
        db.execute("INSERT INTO cantieri (nome,indirizzo,ente_organizzatore,citta,padiglione,superficie_mq) VALUES (?,?,?,?,?,?)",('EICMA 2025','Fiera Milano, Rho','Fiera Milano S.p.A.','Milano','Pad. 18',1200))
        db.execute("INSERT INTO cantieri (nome,indirizzo,ente_organizzatore,citta,padiglione,superficie_mq) VALUES (?,?,?,?,?,?)",('SANA 2025','BolognaFiere, Via della Fiera 20','BolognaFiere S.p.A.','Bologna','Pad. 30',800))
        db.execute("INSERT INTO cantieri (nome,indirizzo,ente_organizzatore,citta,padiglione,superficie_mq) VALUES (?,?,?,?,?,?)",('HOST 2025','Fiera Milano, Rho','Fiera Milano S.p.A.','Milano','Pad. 02',600))
        db.execute("INSERT INTO documenti (titolo,categoria,data_scadenza,creato_da) VALUES (?,?,?,?)",
                   ('Contratto CCNL Edilizia 2024','Contratti','2025-12-31',1))
        db.execute("INSERT INTO eventi (titolo,descrizione,data_inizio,data_fine,colore,creato_da) VALUES (?,?,?,?,?,?)",
                   ('Brief Allestimento EICMA','Riunione pre-allestimento','2026-03-10','2026-03-10','#e63946',1))
        db.execute("INSERT INTO impostazioni (chiave,valore) VALUES (?,?)",('azienda','Accesso Fiere'))
        db.execute("INSERT INTO impostazioni (chiave,valore) VALUES (?,?)",('email_notifiche',''))
        db.execute("INSERT INTO impostazioni (chiave,valore) VALUES (?,?)",('smtp_host','smtp.gmail.com'))
        db.execute("INSERT INTO impostazioni (chiave,valore) VALUES (?,?)",('smtp_port','587'))
        db.execute("INSERT INTO impostazioni (chiave,valore) VALUES (?,?)",('smtp_user',''))
        db.execute("INSERT INTO impostazioni (chiave,valore) VALUES (?,?)",('smtp_pass',''))
        db.execute("INSERT INTO impostazioni (chiave,valore) VALUES (?,?)",('app_url',''))
        safe_commit(db)
    except: pass
    db.close()

def get_setting(chiave, default=''):
    db = get_db()
    r = db.execute("SELECT valore FROM impostazioni WHERE chiave=?", (chiave,)).fetchone()
    db.close()
    return r['valore'] if r else default

def get_base_url():
    """Restituisce l'URL base del sito (usato nei link delle email)."""
    try:
        from flask import request as _req
        return _req.host_url.rstrip('/')
    except:
        # fallback se chiamato fuori dal contesto di una request (es. thread)
        url = get_setting('app_url', '')
        return url.rstrip('/') if url else 'http://localhost:5000'

_ensure_done = False
_ensure_lock = threading.Lock()

def ensure_columns():
    """Aggiunge le colonne mancanti. Idempotente."""
    import time
    for attempt in range(5):
        try:
            db = get_db()
            # Ricrea presenze senza NOT NULL su utente_id se necessario
            cols = {c['name']: c for c in db.execute("PRAGMA table_info(presenze)").fetchall()}
            if cols.get('utente_id') and cols['utente_id']['notnull'] == 1:
                has_jolly  = 'nome_jolly'   in cols
                has_evento = 'evento_id'    in cols
                nj = 'nome_jolly'   if has_jolly  else 'NULL'
                cj = 'cognome_jolly' if has_jolly else 'NULL'
                ev = 'evento_id'    if has_evento else 'NULL'
                db.executescript(f"""
                    CREATE TABLE IF NOT EXISTS presenze_v2 (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        utente_id INTEGER, data TEXT NOT NULL,
                        ora_entrata TEXT, ora_uscita TEXT, ore_totali REAL,
                        cantiere_id INTEGER, note TEXT, evento_id INTEGER,
                        nome_jolly TEXT, cognome_jolly TEXT
                    );
                    INSERT OR IGNORE INTO presenze_v2
                        (id,utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note,evento_id,nome_jolly,cognome_jolly)
                        SELECT id,utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note,{ev},{nj},{cj} FROM presenze;
                    ALTER TABLE presenze RENAME TO presenze_old;
                    ALTER TABLE presenze_v2 RENAME TO presenze;
                    DROP TABLE IF EXISTS presenze_old;
                """)
            # Colonne aggiuntive
            for sql in [
                "ALTER TABLE presenze ADD COLUMN nome_jolly TEXT",
                "ALTER TABLE presenze ADD COLUMN cognome_jolly TEXT",
                "ALTER TABLE presenze ADD COLUMN evento_id INTEGER",
                "ALTER TABLE eventi_files ADD COLUMN dipendente_id INTEGER",
                "ALTER TABLE eventi_dipendenti ADD COLUMN data_da TEXT",
                "ALTER TABLE eventi_dipendenti ADD COLUMN data_a TEXT",
            ]:
                try: db.execute(sql)
                except: pass
            safe_commit(db)
            db.close()
            return
        except sqlite3.OperationalError as e:
            if 'locked' in str(e) and attempt < 4:
                time.sleep(1)
            else:
                print(f'[ensure_columns] {e}')
                return

# ─────────────────────────────────────────────
#  AUTH
# ─────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def d(*a,**k):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        # Dipendenti non-admin: blocca accesso a pagine admin
        admin_pages = {'dashboard','dipendenti','presenze','ferie','cantieri',
                       'documenti','scadenze','calendario','report','richieste','impostazioni','fatturazione','preventivi','clienti','cedolini'}
        if session.get('ruolo') != 'admin' and f.__name__ in admin_pages:
            return redirect(url_for('mobile'))
        return f(*a,**k)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a,**k):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('ruolo') != 'admin':
            flash('Accesso riservato agli amministratori.','error')
            return redirect(url_for('mobile'))
        return f(*a,**k)
    return d

def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

# ─────────────────────────────────────────────
#  EMAIL
# ─────────────────────────────────────────────
def send_email(to, subject, body):
    """
    Invia email tramite Brevo HTTP API (funziona su Railway).
    Configura in Impostazioni: smtp_user = API Key Brevo, email_notifiche = mittente verificato.
    Fallback SMTP se smtp_host non e' brevo.
    """
    import traceback, urllib.request
    api_key  = get_setting('smtp_pass', '')   # riutilizziamo smtp_pass per la API key
    mittente = get_setting('smtp_user', '')    # email mittente verificata su Brevo
    metodo   = get_setting('smtp_host', 'smtp.gmail.com')

    if not mittente or not api_key or not to:
        print(f"[EMAIL] Config mancante: mittente={mittente!r} api_key={'ok' if api_key else 'VUOTA'} to={to!r}")
        return False

    print(f"[EMAIL] Invio a={to} via={metodo} mittente={mittente}")

    # ── Brevo / Sendinblue HTTP API (funziona su Railway) ──
    if 'brevo' in metodo or 'sendinblue' in metodo or metodo.strip() == 'brevo':
        try:
            import json as _json
            payload = _json.dumps({
                "sender":  {"email": mittente},
                "to":      [{"email": to}],
                "subject": subject,
                "htmlContent": body,
            }).encode('utf-8')
            req = urllib.request.Request(
                'https://api.brevo.com/v3/smtp/email',
                data=payload,
                headers={
                    'api-key':      api_key,
                    'Content-Type': 'application/json',
                    'Accept':       'application/json',
                },
                method='POST'
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                print(f"[EMAIL] Brevo OK status={resp.status} a={to}")
                return True
        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8', errors='replace')
            print(f"[EMAIL] Brevo HTTP {e.code}: {err_body}")
            return False
        except Exception as e:
            print(f"[EMAIL] Brevo errore: {e}")
            traceback.print_exc()
            return False

    # ── Fallback SMTP (per ambienti non Railway) ──
    try:
        host = metodo
        port = int(get_setting('smtp_port', '587'))
        msg  = MIMEMultipart('alternative')
        msg['From']    = mittente
        msg['To']      = to
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html', 'utf-8'))
        if port == 465:
            import ssl as _ssl
            s = smtplib.SMTP_SSL(host, port, context=_ssl.create_default_context())
        else:
            s = smtplib.SMTP(host, port, timeout=15)
            s.ehlo(); s.starttls(); s.ehlo()
        s.login(mittente, api_key)
        s.sendmail(mittente, to, msg.as_string())
        s.quit()
        print(f"[EMAIL] SMTP OK a={to}")
        return True
    except Exception as e:
        print(f"[EMAIL] SMTP errore: {e}")
        traceback.print_exc()
        return False

# ─────────────────────────────────────────────
#  NOTIFICHE SCADENZE — eseguita 1 volta/giorno
# ─────────────────────────────────────────────
def check_scadenze_email():
    """Controlla tutti i documenti in scadenza e manda 1 sola email riassuntiva all'admin."""
    email_admin = get_setting('email_notifiche', '')
    if not email_admin:
        return
    oggi = date.today().isoformat()
    db = get_db()
    righe = []

    # ── Documenti dipendenti (tabella documenti con assegnato_a) ──
    try:
        docs_dip = db.execute("""
            SELECT d.titolo, d.data_scadenza, u.nome, u.cognome,
                   CAST(julianday(d.data_scadenza)-julianday('now') AS INTEGER) as days_left
            FROM documenti d JOIN utenti u ON u.id=d.assegnato_a
            WHERE d.data_scadenza IS NOT NULL
              AND julianday(d.data_scadenza)-julianday('now') BETWEEN 0 AND 30
            ORDER BY d.data_scadenza""").fetchall()
        for d in docs_dip:
            gg = d['days_left']
            label = f"<span style='color:#dc2626'>⚠️ Scade oggi</span>" if gg == 0 else f"tra <b>{gg} giorni</b>"
            righe.append(f"<tr><td style='padding:6px 12px'>📋 Doc. dipendente</td>"
                         f"<td style='padding:6px 12px'><b>{d['titolo']}</b> — {d['nome']} {d['cognome']}</td>"
                         f"<td style='padding:6px 12px;font-family:monospace'>{d['data_scadenza']}</td>"
                         f"<td style='padding:6px 12px'>{label}</td></tr>")
    except: pass

    # ── Documenti caricati dalla sezione Dipendenti ──
    try:
        docs_dip2 = db.execute("""
            SELECT dd.nome_originale as titolo, dd.tipo_doc, dd.data_scadenza, u.nome, u.cognome,
                   CAST(julianday(dd.data_scadenza)-julianday('now') AS INTEGER) as days_left
            FROM documenti_dipendente dd JOIN utenti u ON u.id=dd.utente_id
            WHERE dd.data_scadenza IS NOT NULL
              AND julianday(dd.data_scadenza)-julianday('now') BETWEEN 0 AND 30
            ORDER BY dd.data_scadenza""").fetchall()
        for d in docs_dip2:
            gg = d['days_left']
            label = f"<span style='color:#dc2626'>⚠️ Scade oggi</span>" if gg == 0 else f"tra <b>{gg} giorni</b>"
            righe.append(f"<tr><td style='padding:6px 12px'>📄 {d['tipo_doc']}</td>"
                         f"<td style='padding:6px 12px'><b>{d['titolo']}</b> — {d['nome']} {d['cognome']}</td>"
                         f"<td style='padding:6px 12px;font-family:monospace'>{d['data_scadenza']}</td>"
                         f"<td style='padding:6px 12px'>{label}</td></tr>")
    except: pass

    # ── Documenti azienda ──
    try:
        docs_az = db.execute("""
            SELECT titolo, data_scadenza, avviso_giorni,
                   CAST(julianday(data_scadenza)-julianday('now') AS INTEGER) as days_left
            FROM documenti_azienda
            WHERE data_scadenza IS NOT NULL
              AND CAST(julianday(data_scadenza)-julianday('now') AS INTEGER) BETWEEN 0 AND COALESCE(avviso_giorni,30)
            ORDER BY data_scadenza""").fetchall()
        for d in docs_az:
            gg = d['days_left']
            label = f"<span style='color:#dc2626'>⚠️ Scade oggi</span>" if gg == 0 else f"tra <b>{gg} giorni</b>"
            righe.append(f"<tr><td style='padding:6px 12px'>🏢 Doc. azienda</td>"
                         f"<td style='padding:6px 12px'><b>{d['titolo']}</b></td>"
                         f"<td style='padding:6px 12px;font-family:monospace'>{d['data_scadenza']}</td>"
                         f"<td style='padding:6px 12px'>{label}</td></tr>")
    except: pass

    # ── Scadenze veicoli ──
    try:
        veicoli_sc = db.execute("""
            SELECT targa, marca, modello, revisione, assicurazione, bollo,
                   MIN(CASE WHEN revisione IS NOT NULL THEN CAST(julianday(revisione)-julianday('now') AS INTEGER) END,
                       CASE WHEN assicurazione IS NOT NULL THEN CAST(julianday(assicurazione)-julianday('now') AS INTEGER) END,
                       CASE WHEN bollo IS NOT NULL THEN CAST(julianday(bollo)-julianday('now') AS INTEGER) END) as min_days
            FROM veicoli WHERE attivo=1""").fetchall()
        for v in veicoli_sc:
            for campo, etichetta in [('revisione','Revisione'),('assicurazione','Assicurazione'),('bollo','Bollo')]:
                sc = v[campo]
                if not sc: continue
                try:
                    gg = (date.fromisoformat(sc) - date.today()).days
                except: continue
                if 0 <= gg <= 30:
                    label = f"<span style='color:#dc2626'>⚠️ Scade oggi</span>" if gg == 0 else f"tra <b>{gg} giorni</b>"
                    righe.append(f"<tr><td style='padding:6px 12px'>🚗 Veicolo</td>"
                                 f"<td style='padding:6px 12px'><b>{etichetta}</b> — {v['targa']} {v['marca']} {v['modello']}</td>"
                                 f"<td style='padding:6px 12px;font-family:monospace'>{sc}</td>"
                                 f"<td style='padding:6px 12px'>{label}</td></tr>")
    except: pass

    db.close()
    if not righe:
        return

    base = get_base_url()
    corpo = (f"<h2 style='color:#0f172a'>⚠️ Riepilogo scadenze — {oggi}</h2>"
             f"<p>Hai <b>{len(righe)} scadenza/e</b> nei prossimi 30 giorni:</p>"
             f"<table style='border-collapse:collapse;font-family:sans-serif;width:100%'>"
             f"<thead><tr style='background:#0f172a;color:#fff'>"
             f"<th style='padding:8px 12px;text-align:left'>Tipo</th>"
             f"<th style='padding:8px 12px;text-align:left'>Documento</th>"
             f"<th style='padding:8px 12px;text-align:left'>Scadenza</th>"
             f"<th style='padding:8px 12px;text-align:left'>Quando</th>"
             f"</tr></thead><tbody>{''.join(righe)}</tbody></table>"
             f"<p style='margin-top:20px'>"
             f"<a href='{base}/documenti-azienda' style='background:#0f172a;color:#fff;padding:10px 18px;border-radius:8px;text-decoration:none;font-weight:700;margin-right:10px'>Documenti azienda →</a>"
             f"<a href='{base}/documenti' style='background:#2563eb;color:#fff;padding:10px 18px;border-radius:8px;text-decoration:none;font-weight:700'>Documenti dipendenti →</a>"
             f"</p>")

    import threading
    threading.Thread(target=lambda: send_email(email_admin,
        f'[ACCESSO FIERE] {len(righe)} scadenze nei prossimi 30 giorni — {oggi}', corpo), daemon=True).start()


# ─────────────────────────────────────────────
#  RENDER HELPER
# ─────────────────────────────────────────────
def render_page(tmpl, **ctx):
    if session.get('ruolo') == 'admin':
        db = get_db()
        ctx['notifiche_count'] = db.execute(
            "SELECT COUNT(*) FROM richieste_presenze WHERE stato='in_attesa'"
        ).fetchone()[0] + db.execute(
            "SELECT COUNT(*) FROM ferie_permessi WHERE stato='in_attesa'"
        ).fetchone()[0]
        try:
            ctx['spese_attesa'] = db.execute(
                "SELECT COUNT(*) FROM spese_rimborso WHERE stato='in_attesa'"
            ).fetchone()[0]
        except:
            ctx['spese_attesa'] = 0
        db.close()
        # ── Controllo scadenze 1 volta al giorno ──
        oggi = date.today().isoformat()
        if session.get('_scadenze_check') != oggi:
            session['_scadenze_check'] = oggi
            import threading
            threading.Thread(target=check_scadenze_email, daemon=True).start()
    else:
        ctx['notifiche_count'] = 0
        ctx['spese_attesa'] = 0
    ctx.setdefault('azienda_nome', get_setting('azienda','Accesso Fiere'))
    full = BASE.replace('    {% block content %}{% endblock %}', tmpl)
    return render_template_string(full, **ctx)

# ══════════════════════════════════════════════════════════
#  BASE TEMPLATE
# ══════════════════════════════════════════════════════════
BASE = """<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ page_title }} | Accesso Fiere — Gestionale Allestitori</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.0/css/all.min.css">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root{--bg:#f4f5f7;--sidebar:#0f172a;--sidebar-hover:#1e293b;--accent:#0f4c81;--accent2:#00b4d8;--white:#fff;--text:#1e293b;--text-light:#64748b;--border:#e2e8f0;--card:#fff;--success:#22c55e;--warning:#f59e0b;--danger:#ef4444;--radius:12px;--shadow:0 1px 3px rgba(0,0,0,.07),0 4px 16px rgba(0,0,0,.05)}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);display:flex;min-height:100vh}
.sidebar{width:240px;background:var(--sidebar);display:flex;flex-direction:column;position:fixed;top:0;left:0;height:100vh;overflow-y:auto;z-index:100}
.sidebar-logo{padding:16px 18px 14px;border-bottom:1px solid rgba(255,255,255,.07);display:flex;align-items:center;gap:10px}
.sidebar-logo img{height:36px;width:auto;object-fit:contain;flex-shrink:0;border-radius:5px}
.sidebar-logo .company{font-size:14px;font-weight:700;color:#fff}
.sidebar-logo .sub{font-size:10.5px;color:rgba(255,255,255,.4);margin-top:2px}
nav{flex:1;padding:12px 0}
.nav-section{font-size:10px;font-weight:600;color:rgba(255,255,255,.3);text-transform:uppercase;letter-spacing:.8px;padding:16px 20px 6px}
nav a{display:flex;align-items:center;gap:10px;padding:9px 20px;color:rgba(255,255,255,.7);text-decoration:none;font-size:13.5px;font-weight:500;border-left:3px solid transparent;transition:all .18s}
nav a:hover{background:var(--sidebar-hover);color:#fff}
nav a.active{background:rgba(15,76,129,.15);color:#fff;border-left-color:var(--accent)}
nav a i{width:16px;text-align:center;font-size:13px}
.sidebar-user{padding:14px 20px;border-top:1px solid rgba(255,255,255,.07);display:flex;align-items:center;gap:10px}
.avatar{width:32px;height:32px;border-radius:50%;background:var(--accent);display:flex;align-items:center;justify-content:center;color:#fff;font-size:12px;font-weight:700;flex-shrink:0}
.sidebar-user .uname{font-size:12.5px;font-weight:600;color:#fff;line-height:1.2}
.sidebar-user .urole{font-size:11px;color:rgba(255,255,255,.4)}
.main{margin-left:240px;flex:1;display:flex;flex-direction:column;min-height:100vh}
.topbar{background:var(--white);border-bottom:1px solid var(--border);padding:0 28px;height:60px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:50}
.topbar h1{font-size:18px;font-weight:700}
.topbar-actions{display:flex;gap:10px;align-items:center}
.content{padding:28px;flex:1}
.card{background:var(--card);border-radius:var(--radius);box-shadow:var(--shadow);border:1px solid var(--border);overflow:hidden}
.card-header{padding:18px 22px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between}
.card-header h3{font-size:14px;font-weight:600}
.card-body{padding:22px}
.stats-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:16px;margin-bottom:24px}
.stat-card{background:var(--white);border-radius:var(--radius);padding:20px;box-shadow:var(--shadow);border:1px solid var(--border);display:flex;align-items:center;gap:14px}
.stat-icon{width:46px;height:46px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:19px;flex-shrink:0}
.stat-icon.red{background:#e8f0fe;color:var(--accent)}
.stat-icon.blue{background:#e3f2fd;color:var(--accent2)}
.stat-icon.green{background:#dcfce7;color:var(--success)}
.stat-icon.amber{background:#fef3c7;color:var(--warning)}
.stat-icon.purple{background:#ede9fe;color:#7c3aed}
.stat-val{font-size:24px;font-weight:700;line-height:1}
.stat-lbl{font-size:12px;color:var(--text-light);margin-top:3px}
.table-wrap{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:13.5px}
th{font-size:11px;font-weight:600;color:var(--text-light);text-transform:uppercase;letter-spacing:.5px;padding:10px 14px;background:#f8fafc;border-bottom:1px solid var(--border);text-align:left}
td{padding:12px 14px;border-bottom:1px solid var(--border);vertical-align:middle}
tr:last-child td{border-bottom:none}
tr:hover td{background:#fafafa}
.avatar-sm{width:28px;height:28px;border-radius:50%;background:var(--accent2);display:inline-flex;align-items:center;justify-content:center;color:#fff;font-size:11px;font-weight:700;margin-right:8px;flex-shrink:0}
.badge{display:inline-block;padding:3px 10px;border-radius:99px;font-size:11px;font-weight:600}
.badge-green{background:#dcfce7;color:#16a34a}
.badge-red{background:#fee2e2;color:#dc2626}
.badge-amber{background:#fef3c7;color:#92400e}
.badge-blue{background:#dbeafe;color:#1d4ed8}
.badge-gray{background:#f1f5f9;color:#475569}
.badge-purple{background:#ede9fe;color:#6d28d9}
.btn{display:inline-flex;align-items:center;gap:7px;padding:8px 16px;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;border:none;text-decoration:none;transition:all .18s;font-family:inherit}
.btn-primary{background:var(--accent);color:#fff}.btn-primary:hover{background:#0a3660}
.btn-secondary{background:var(--white);color:var(--text);border:1px solid var(--border)}.btn-secondary:hover{background:#f1f5f9}
.btn-blue{background:var(--accent2);color:#fff}.btn-blue:hover{background:#1565c0}
.btn-green{background:var(--success);color:#fff}.btn-green:hover{background:#16a34a}
.btn-sm{padding:5px 12px;font-size:12px}
.btn-danger{background:#fee2e2;color:#dc2626}.btn-danger:hover{background:#fecaca}
.form-group{margin-bottom:16px}
label{display:block;font-size:12.5px;font-weight:600;color:var(--text-light);margin-bottom:6px}
input,select,textarea{width:100%;padding:9px 12px;border:1.5px solid var(--border);border-radius:8px;font-size:13.5px;font-family:inherit;color:var(--text);transition:border .18s;background:#fff}
input:focus,select:focus,textarea:focus{outline:none;border-color:var(--accent)}
textarea{resize:vertical;min-height:80px}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.form-row-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px}
.form-row-4{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:16px}
.alert{padding:12px 16px;border-radius:8px;font-size:13px;margin-bottom:16px;display:flex;align-items:center;gap:8px}
.alert-success{background:#dcfce7;color:#16a34a;border:1px solid #bbf7d0}
.alert-error{background:#fee2e2;color:#dc2626;border:1px solid #fecaca}
.alert-info{background:#dbeafe;color:#1d4ed8;border:1px solid #bfdbfe}
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:20px}
.grid-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px}
.empty-state{text-align:center;padding:40px 20px;color:var(--text-light)}
.empty-state i{font-size:36px;margin-bottom:10px;opacity:.4;display:block}
.tag{display:inline-block;background:#f1f5f9;color:var(--text-light);font-size:11px;padding:2px 8px;border-radius:6px;margin-right:4px}
.cal-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:4px}
.cal-header-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:4px;margin-bottom:4px}
.cal-dow{text-align:center;font-size:11px;font-weight:700;color:var(--text-light);padding:6px 0;text-transform:uppercase}
.cal-day{min-height:80px;background:#f8fafc;border-radius:8px;padding:6px;border:1px solid var(--border)}
.cal-day.today{border-color:var(--accent);background:#e8f0fe}
.cal-day.other-month{opacity:.4}
.cal-day-num{font-size:12px;font-weight:600;color:var(--text-light);margin-bottom:3px}
.cal-day.today .cal-day-num{color:var(--accent)}
.cal-event{border-radius:4px;padding:2px 5px;font-size:10px;font-weight:600;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.cal-nav{display:flex;align-items:center;gap:12px}
.cal-nav span{font-size:15px;font-weight:700;min-width:160px;text-align:center}
.clock-card{background:linear-gradient(135deg,#0f172a,#0f4c81);color:#fff;border-radius:var(--radius);padding:28px;text-align:center}
.clock-time{font-family:'DM Mono',monospace;font-size:42px;font-weight:500;letter-spacing:2px;margin:12px 0}
.scadenza-bar{display:flex;align-items:center;gap:10px;padding:11px 0;border-bottom:1px solid var(--border)}
.scadenza-bar:last-child{border-bottom:none}
.notif-badge{background:var(--accent);color:#fff;font-size:10px;font-weight:700;padding:1px 6px;border-radius:99px;margin-left:auto}
.progress-bar{background:#e2e8f0;border-radius:99px;height:6px;overflow:hidden;margin-top:4px}
.progress-fill{background:linear-gradient(90deg,#2563eb,#22c55e);height:100%;border-radius:99px;transition:.3s}
@media(max-width:900px){.grid-2,.grid-3,.form-row,.form-row-3,.form-row-4{grid-template-columns:1fr}.sidebar{width:200px}.main{margin-left:200px}}
</style>
</head>
<body>
<aside class="sidebar">
  <div class="sidebar-logo">
    <img src="data:image/png;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCABDALADASIAAhEBAxEB/8QAHQAAAQQDAQEAAAAAAAAAAAAAAAEGBwgDBAUJAv/EAEcQAAEDAwMCBAQBBgYTAAAAAAECAwQABREGEiEHMRMiQVEIFGGBcRYyN5GysxUjUoKEoRckJSc1QkNEVWNkcnN0dYOi0fH/xAAaAQACAwEBAAAAAAAAAAAAAAAAAgEDBAUG/8QAKxEAAgIBAwIGAQQDAAAAAAAAAQIAAxEEEjETIQVBUWGBoTMUMjSxweHw/9oADAMBAAIRAxEAPwCwtJS1kKU/KJXjzFwjP0wK6c83tmE96K5t5ucu2x50lNoXLjx4njFxLyQchQyNp54GTmmUx1UYfebYZsb63HFBKE+MOSTgDtV9dNlgJAmS7WUUMFsOCfYyR60bnebRa34rFyukKE7LX4cZEh9Lanl8eVAJ8x5HA9xWkzP1MtafF0uIyFf47s9vA+wBJ+wqD/iJcuV66jdPGUWh+Mm33YOLW6tICwVNrJT7gBpR9z7Z4qtwQO3eaKrK3YBsjPqCPL3lijSUybL1Jsc1QanBy3u5xlfmQf5w7fcCnjFksSmEvxnm3mlfmrQoKB+4qx63T9wlVOqpu71tmZaQ1DvVbrVL0Z1JhaPtulHb+7IiodKI7xS9vWpQCUpCVZ4SD963ZfUvXMS3vT5XR67tR2WlOuKVdI/lSkZJI79hVPUUHE3ChyAfWSrRVbYHxW2+dNYhRdDz3H33EtNIE5A3KUQAOU+5p/T+pGvokRyS50avakNpKlBu4suKx9EpyT9hQLlPEk6ewciSoa+airoh1cf6ks6hBsKbdMtRR4cUyNynAoKwCopGDlGO3Ga4mvOvVw0JOYh6q6eToLshBcZ2XJp1K0g4JBSMd/TvU9VcZgKHLbcd5OFBqFen/XG7a9clJ0r06mzExNvjrcubLSW92dvKhznB4Ge1Snqi4Xm32sSLJYjepW8AxhLQwduDlW5fHHHH1plcMMiK1TIcNOtRVd718UMSy3aVabroa5Rp0R1TT7RmNnYsHBGQMH8RUjdMdf37W7EO5p0PIttkloWpue9cGlE4yBhsDdgkYz96hbUY4BjNp7FGSJIVIajTqf1Kvug2Zdym6Eky7Gw4lAuDNxa827ABLeNyeTimRpv4l2tS3yNZLFoS4y7hKUUss/Otp3EAk8kYAwCcmpNyA4JgunsYZA7SwNFc3Tcy6T7S3JvFoNomKJ3xTJS/sAPHnTwcjn710qtEpIx2m0a2Aha4CdqeA4rJJwBwPWsTSezhW2kDtu5/qrYddSYaVeZ0hw4Ln4D0FY2PfAjqBg5mjc2ULsN1b8z5MJwFDfGeO24jj9RqELE6zDvcBsPw4JMlseHDT47x8w7uE4GfXar+bU1X91xembylSvL8g55RwO3sKrfZ2fmLvDYDi2y4+hG9BwpOVAZH1rpaGsur5M8545d07ato7/7/AO85ZlTqMEIbHPdSzuP/AKqquvNKXDrnre9X7TV3fjQNPOCCyVneiQtAKlKYwQASffAPlO7nAcvXJnVmnIsS1WbWMyXPv8kQbdBKXPFwoYUsrDmMJz32+o/Gur0+6f6/0ZbYFls1ztbMSPcwt9xUVR8drwlgrUA6NwKiny8HODnAwee6KchTxPR0WuoDMME8ef8Aj4jEtVzZvlmi3ZhLyNyfl5CHzl1t9oBKw4f5RwFn/frp2u53G2PeNb5j0dfrsVgH8R2P3rX1zp676J6nIdmSbau0axfy443CcSyxOSDtAT4uQXM992CVHjy1uQ48WXd4UNTny6HFobfcKdgSSrBUApSsAD3Ue1drRajqV7XHcczxfjXhx0up6tTdnOR6+/3NrSd2h23qNcNb3eK5cLlOjtxy5lI8BCQAdicY5AGeff3NSZqDVlivWhr6iJOQl5VskYZd8i8+ErgA9/tmms/ozT6LvFiJ1K2pl1pxa3d7fkKSnA7453H9VNDXUONYZkxq3y0TmmGPES5kEKO3OOKRqNPZnZ2l9XiHiGnK9XDDIHIz9Sr+gQRruwZ/0nH/AHqa9Lx2rz1seqI161vYy/p62syFXKPh9jchWfFTycHB+9ehQ7VxqFVSQpzPa6p3cKXXafgxn2PQNvsvUq8a0t8hTK7vGS1JhpbAbLiSD4uf5Rxzx6k+tV/+PQf3W0ofdiT+03VsTVTvj1/wrpP/AIEn9pupvAFZxI0rFrgTOp8BeBadVn/Xxv2XKs0cYqnPwlaBtGtLff3rpNu8ZUR1hLYgzVMA7gvO7b37VOq+hulCnBvWrMf9adooLbBgRtUENpyfqVF+Igbetmqh/txP/iKub8PP6FdK8f5gn9o1SHrJaY9i6oX+0RHZDrEWWW0LkOFxwgAfnKPc1d74ef0K6V/5AftGk035DLtZ+Fficr4rsf2Cb/8A0f8Aft1Vr4VB/f409/SP3DlWl+K79BOoP+x+/bqrfwp/p40//SP3DlF35li6X+O3z/UvrRRRW4TnzarPhSoKdqSf409h9BWuaa9n1eq4ayuemEW9bT9uQlbzpeGxQUAU7eMnuPbHNZTGXgzp6ouEeLZLxFcDpkKt6yhtLSiTuISOw9zUC2mNMh3WJLfhSktMvoWs+CrgBQJ9KmnSGrlaiut3gIguxjanzHfUp7cFLBI8uByPKeTj04pzFSiMFSiPxrXptSalIAzmcrX+GfqrFZmxt9vn1kHdL1p6gdab7rmeHPl7K2INljONLT4SFFQU6cgAKVgnHcBfPpU4U37tqC22e8xbMwyqRdbiorbjtYBIA5cWo9kgDv34wAcVhf1axbr/AA7LfIi7e/OyIjwWHGXlDjZu4IVyOCPUc1lXCzrsWfHaYerWj4+udC3CwOlLb7iPEiPH/Ivp5QrPpzwfoTUG6Ouk9yPbb3dLe6m4xJSo9wYcSU7pLCgHDnHZXBJ7bioelWdNcDVdyjaVsE69pt4eabWHn0IUEqOcJKhkYJ7e3qavqs6T7/LzmLW6YaukVcEEEH0kev6/S5dos/8AJ5KQw0434ficK3lPP5vpt/rpmdRLmu+G4z24Bjb4qkhpPm7II9hUwxNcxVXKxQp1sdjfw6wHoTiVBxPIB2r7FJwR7jnvTvKE5/NH6q0rq0wdq/c5r+E3sR1LMjOeB5TzU0vFuFs1Na7lJtk4MRZjLyyI6vzUrCj6ewr0O0bq/Turor0nT1xExDBSl3+KWgoJGQCFAH0rp3qWq32qTOSz43y7SnSjdtyEjJwffArnaHv6NT2Bm9sQ1RWHyoNpWsFR2qKSTjgcg1zaq+mcZnpL7usASOJ2ZT7UaM7IfWENNIK1qPZKQMk/qqm/xcantWur1YvyUXIujUFh0POtRnAkKWpOAMpGfzfSrlq7c02LtqVy1W+fcpVsEaBDWpAdffCC7tOMpTgnk8D3p7V3LjMSizY2cZMrX8Ier7Hohm/wtVvvWpc1xhcdT0ZzavaFhQyEnHcd6tHqnUtk0xahdL7OTDiFYbDikKVlRBIACQSexrHo28u6jsbd1etT1vaeOWW31AqWj0VgdgfT9ddpSUkYIBqakKrgGF1gd9xE87+rzj+pepmoL5aYE5+DLmrWw58ssb0dgcEZGcZ5q1/w16207K0Fp7Sfzqmr5GiqbchusLQrylRJBIweOeDUgN6mYmXyZZrPHM6TBA+aWV7GmlHsgqwSVcHgA4wcmksOqINzvcuyPMOQbtDG5yM9gkpOPOhQ4UnkfXntSV17GznmW239RNpXiRX8VOtdPSOn160ZDluSr644ygxWmHFFGHELJKgnaOB7+tV06CS/yS6sWW/X6LNiW+Op0PPGMtQQFNLSDgDOMqHar/hCQcgDJps671X+SrUd520yJzLylJ/tdQK07UlSiUn0CQTn6UWVZbeTxCq/anTC8+86um75atR2lq62WYmXCdJCHUpIBIODwQD3FdGm2rVUV232yfZ2m7hFuL6WG1od2bFqBOFDBxjBz6j2pxpyUjPf1rSpzMpGJtGo5tOlJx6oX6/XS0Jct89tCI5LiCpJSlIyQDxnacd+4zipHpMVmIzBWK5xGD0u07dLBetSOzYAjRrjMMiLtcSrajKsJUAeDhQ9x9aftB70U6jAxBmLHJjI1TpW4K13bNaWdTT0iK0WJER1W0OtkKGUq9FeY9+DxyKTVmmp2sL5ZHZbBt9vtb5kuBxaVPPL4wkbSQlPHJznntT4oqNgjCwjHtENNnqdbJ960TcLRbWPGkykpQgFYSlPmBJJJ7cemacx70lMRkYigkHIkWWvR2obDKs1+tcduXKYhtxJsCY+FYCUgFTDhzsyc8cD6elSbEccejocejrjuEeZtZSSn7pJH9dZqQ96hVC8SWctzOdqVl+Rp+fGisl556O40hAIGSpJA5JAxzXE6T2i42DRMOz3SOGZMYubtqwpKgpalAgg/X1xTsopsd8yNx27YGo1vtgv+pNYtydQWxbmn4Ct0OCy+2fHWOy3cqH6vt75ko180EZgrFT2mhZ3bg946psJEJtKwmO2FhSikAeZRBwOc8Dtj61vntRQaaRGJYNN3PS2sbzcoDQuFsvDgedbSsJeYdyo8bsBSfMfUHtWa06XlPdRJWs7jtjqMcR4sZKtygnGCpZHGTzwM/jTzxS1AQRy5hTe1JHnv32yvxoC32Ib7jj6g4gcKaWgAAkZOVA/hThpDTEZig4keS9AuwNXwLtp5/5e3KmpkTrfnDe4AjxED0PPI/8AlSEKKKlVA4ksxbGZtUUUVnixD3pKKKeEKKKKIRD3oooohCvk96KKkQhRRRUwga+aKKkQhQaKKmESiiipEIUhooohEoooqRCf/9k=" alt="Logo">
    <div>
      <div class="company">{{ azienda_nome }}</div>
      <div class="sub">Gestionale Allestitori</div>
    </div>
  </div>
  <nav>   <div class="nav-section">Personale di Cantiere</div>
    <a href="/dashboard" class="{{ 'active' if active=='dashboard' }}"><i class="fa fa-house"></i> Dashboard</a>
    <a href="/dipendenti" class="{{ 'active' if active=='dipendenti' }}"><i class="fa fa-users"></i> Dipendenti</a>
    <a href="/presenze" class="{{ 'active' if active=='presenze' }}"><i class="fa fa-clock"></i> Presenze</a>
    <a href="/ferie" class="{{ 'active' if active=='ferie' }}"><i class="fa fa-umbrella-beach"></i> Ferie & Permessi</a>
    <div class="nav-section">Fiere & Allestimenti</div>
    <a href="/cantieri" class="{{ 'active' if active=='cantieri' }}"><i class="fa fa-store"></i> Fiere & Stand</a>
    <a href="/documenti" class="{{ 'active' if active=='documenti' }}"><i class="fa fa-file-alt"></i> Documenti</a>
    <a href="/scadenze" class="{{ 'active' if active=='scadenze' }}" style="padding-left:32px;font-size:13px;opacity:.85"><i class="fa fa-calendar-exclamation" style="font-size:11px"></i> Scadenze dipendenti</a>
    <a href="/documenti-azienda" class="{{ 'active' if active=='documenti_azienda' }}"><i class="fa fa-building-columns"></i> Documenti Azienda</a>
    <a href="/veicoli" class="{{ 'active' if active=='veicoli' }}"><i class="fa fa-truck"></i> Mezzi & Veicoli</a>
    <a href="/pos" class="{{ 'active' if active=='pos' }}"><i class="fa fa-file-shield"></i> PSAF</a>
    <a href="/clienti" class="{{ 'active' if active=='clienti' }}"><i class="fa fa-address-book"></i> Clienti</a>
    <a href="/fatturazione" class="{{ 'active' if active=='fatturazione' }}"><i class="fa fa-file-invoice-dollar"></i> Fatturazione</a>
    <a href="/preventivi" class="{{ 'active' if active=='preventivi' }}"><i class="fa fa-file-invoice"></i> Preventivi</a>
    <a href="/calendario" class="{{ 'active' if active=='calendario' }}"><i class="fa fa-calendar"></i> Calendario</a>
    <a href="/cedolini" class="{{ 'active' if active=='cedolini' }}"><i class="fa fa-money-bill-wave"></i> Cedolini</a>
    {% if session.ruolo=='admin' %}
    <div class="nav-section">Admin</div>
    <a href="/admin/richieste" class="{{ 'active' if active=='richieste' }}">
      <i class="fa fa-bell"></i> Richieste
      {% if notifiche_count > 0 %}<span class="notif-badge">{{ notifiche_count }}</span>{% endif %}
    </a>
    <a href="/admin/spese" class="{{ 'active' if active=='admin_spese' }}"><i class="fa fa-receipt"></i> Spese rimborso
      {% if spese_attesa > 0 %}<span class="notif-badge" style="background:#f59e0b">{{ spese_attesa }}</span>{% endif %}
    </a>
    <a href="/admin/report" class="{{ 'active' if active=='report' }}"><i class="fa fa-chart-bar"></i> Report</a>
    <a href="/admin/impostazioni" class="{{ 'active' if active=='impostazioni' }}"><i class="fa fa-gear"></i> Impostazioni</a>
    {% endif %}
  </nav>
  <div class="sidebar-user">
    <div class="avatar">{{ (session.nome or 'U')[0] }}{{ (session.cognome or '')[0] }}</div>
    <div>
      <div class="uname">{{ session.nome }} {{ session.cognome }}</div>
      <div class="urole">{{ 'Amministratore' if session.ruolo=='admin' else 'Dipendente' }}</div>
    </div>
  </div>
</aside>
<main class="main">
  <div class="topbar">
    <h1>{{ page_title }}</h1>
    <div class="topbar-actions">
      {% if session.ruolo=='admin' %}
        {% if active=='dipendenti' %}<a href="/dipendenti/nuovo" class="btn btn-primary btn-sm"><i class="fa fa-plus"></i> Aggiungi persona</a>
        {% elif active=='cantieri' %}<a href="/cantieri/nuovo" class="btn btn-primary btn-sm"><i class="fa fa-plus"></i> Nuova fiera</a>
        {% elif active=='documenti' %}<a href="/documenti/nuovo" class="btn btn-primary btn-sm"><i class="fa fa-plus"></i> Nuovo documento</a>
        {% elif active=='calendario' %}<a href="/calendario/nuovo" class="btn btn-primary btn-sm"><i class="fa fa-plus"></i> Nuovo evento</a>
        {% elif active=='presenze' %}<div style="display:flex;gap:8px"><a href="#inserisci" onclick="document.getElementById('form-manuale').style.display='block'" class="btn btn-primary btn-sm"><i class="fa fa-plus"></i> Inserisci presenza</a><a href="/presenze/import-ai" class="btn btn-sm" style="background:#7c3aed;color:#fff"><i class="fa fa-robot"></i> Import AI</a></div>
        {% endif %}
      {% endif %}
      <a href="/logout" class="btn btn-secondary btn-sm"><i class="fa fa-sign-out-alt"></i> Esci</a>
    </div>
  </div>
  <div class="content">
    {% with msgs = get_flashed_messages(with_categories=true) %}
      {% for cat, msg in msgs %}
        <div class="alert alert-{{ 'success' if cat=='success' else 'error' if cat=='error' else 'info' }}">
          <i class="fa fa-{{ 'check-circle' if cat=='success' else 'exclamation-circle' }}"></i> {{ msg }}
        </div>
      {% endfor %}
    {% endwith %}
    {% block content %}{% endblock %}
  </div>
</main>
<script>
function updateClock(){var e=document.getElementById('live-clock');if(e)e.textContent=new Date().toLocaleTimeString('it-IT');}
setInterval(updateClock,1000);updateClock();
</script>
</body></html>"""

# ══════════════════════════════════════════════════════════
#  LOGIN
# ══════════════════════════════════════════════════════════
LOGIN_TMPL = """<!DOCTYPE html>
<html lang="{{ t.get('dir','ltr') == 'rtl' and 'ar' or 'it' }}" dir="{{ t.dir }}">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Accesso</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;600;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.0/css/all.min.css">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;min-height:100vh;display:flex;background:#f4f5f7}
.left{width:420px;background:#0f172a;display:flex;flex-direction:column;justify-content:center;align-items:center;padding:48px;color:#fff;flex-shrink:0}
.left .logo{font-size:28px;font-weight:700;margin-bottom:6px}
.left .tagline{font-size:14px;opacity:.5;text-align:center;max-width:260px;line-height:1.6}
.features{margin-top:40px;width:100%}
.feature{display:flex;align-items:center;gap:12px;padding:10px 0;border-bottom:1px solid rgba(255,255,255,.07);font-size:13px;color:rgba(255,255,255,.6)}
.feature i{color:#0f4c81;font-size:14px;width:16px}
.right{flex:1;display:flex;align-items:center;justify-content:center;padding:40px}
.login-box{background:#fff;border-radius:16px;padding:40px;width:100%;max-width:400px;box-shadow:0 4px 40px rgba(0,0,0,.08)}
.login-box h2{font-size:22px;font-weight:700;margin-bottom:6px}
.login-box p{font-size:13px;color:#64748b;margin-bottom:20px}
.form-group{margin-bottom:18px}
label{display:block;font-size:12.5px;font-weight:600;color:#64748b;margin-bottom:6px}
input,select{width:100%;padding:10px 13px;border:1.5px solid #e2e8f0;border-radius:9px;font-size:14px;font-family:inherit;transition:border .18s;background:#fff;color:#1e293b}
input:focus,select:focus{outline:none;border-color:#0f4c81}
.btn{width:100%;padding:11px;background:#0f4c81;color:#fff;border:none;border-radius:9px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit}
.btn:hover{background:#c62535}
.alert{padding:10px 14px;background:#fee2e2;color:#dc2626;border-radius:8px;font-size:13px;margin-bottom:16px;border:1px solid #fecaca}
.lang-bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:20px}
.lang-btn{flex:1;min-width:70px;padding:8px 6px;border:1.5px solid #e2e8f0;border-radius:8px;background:#f8fafc;font-size:13px;font-weight:600;cursor:pointer;text-align:center;transition:.15s;color:#475569}
.lang-btn.active{border-color:#0f4c81;background:#fff1f2;color:#0f4c81}
@media(max-width:700px){.left{display:none}.right{padding:20px}}
{% if t.dir == 'rtl' %}body,input,select,button{font-family:'DM Sans',Arial,sans-serif}{% endif %}
</style></head>
<body>
<div class="left">
  <div class="logo">🏢 Accesso Fiere</div>
  <div class="tagline">Gestionale Allestitori</div>
  <div class="features">
    <div class="feature"><i class="fa fa-users"></i> Dipendenti & Cantieri</div>
    <div class="feature"><i class="fa fa-clock"></i> Rilevazione presenze</div>
    <div class="feature"><i class="fa fa-umbrella-beach"></i> Ferie e permessi</div>
    <div class="feature"><i class="fa fa-chart-bar"></i> Report e grafici</div>
  </div>
</div>
<div class="right">
  <div class="login-box">
    <!-- Selettore lingua -->
    <div class="form-group">
      <label>{{ t.login_lang }}</label>
      <div class="lang-bar">
        {% for code, l in langs.items() %}
        <form method="POST" action="/set-lang" style="flex:1;min-width:70px">
          <input type="hidden" name="lang" value="{{ code }}">
          <input type="hidden" name="next" value="/login">
          <button type="submit" class="lang-btn {{ 'active' if current_lang == code }}">
            {{ l.flag }} {{ l.name }}
          </button>
        </form>
        {% endfor %}
      </div>
    </div>

    <h2>{{ t.login_title }}</h2>
    <p>{{ t.login_sub }}</p>
    {% if error %}<div class="alert"><i class="fa fa-exclamation-circle"></i> {{ error }}</div>{% endif %}
    <form method="POST">
      <div class="form-group">
        <label>{{ t.login_email }}</label>
        <input type="email" name="email" placeholder="nome@azienda.it" required dir="ltr">
      </div>
      <div class="form-group">
        <label>{{ t.login_pass }}</label>
        <input type="password" name="password" placeholder="••••••••" required dir="ltr">
      </div>
      <button class="btn" type="submit">{{ t.login_btn }} <i class="fa fa-arrow-right" style="margin-{{ 'right' if t.dir=='rtl' else 'left' }}:6px"></i></button>
    </form>
    <div style="display:flex;align-items:center;gap:10px;margin:18px 0">
      <div style="flex:1;height:1px;background:#e2e8f0"></div>
      <span style="font-size:12px;color:#94a3b8">oppure</span>
      <div style="flex:1;height:1px;background:#e2e8f0"></div>
    </div>
    <a href="/auth/google" style="display:flex;align-items:center;justify-content:center;gap:10px;width:100%;padding:11px;background:#fff;border:1.5px solid #e2e8f0;border-radius:9px;font-size:14px;font-weight:600;color:#1e293b;text-decoration:none;transition:.15s" onmouseover="this.style.borderColor='#4285f4';this.style.background='#f8faff'" onmouseout="this.style.borderColor='#e2e8f0';this.style.background='#fff'">
      <svg width="18" height="18" viewBox="0 0 48 48"><path fill="#4285F4" d="M45.12 24.5c0-1.56-.14-3.06-.4-4.5H24v8.51h11.84c-.51 2.75-2.06 5.08-4.39 6.64v5.52h7.11c4.16-3.83 6.56-9.47 6.56-16.17z"/><path fill="#34A853" d="M24 46c5.94 0 10.92-1.97 14.56-5.33l-7.11-5.52c-1.97 1.32-4.49 2.1-7.45 2.1-5.73 0-10.58-3.87-12.31-9.07H4.34v5.7C7.96 41.07 15.4 46 24 46z"/><path fill="#FBBC05" d="M11.69 28.18C11.25 26.86 11 25.45 11 24s.25-2.86.69-4.18v-5.7H4.34C2.85 17.09 2 20.45 2 24c0 3.55.85 6.91 2.34 9.88l7.35-5.7z"/><path fill="#EA4335" d="M24 10.75c3.23 0 6.13 1.11 8.41 3.29l6.31-6.31C34.91 4.18 29.93 2 24 2 15.4 2 7.96 6.93 4.34 14.12l7.35 5.7c1.73-5.2 6.58-9.07 12.31-9.07z"/></svg>
      Accedi con Google
    </a>
</div>
</body></html>"""

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('ruolo') == 'admin':
        return redirect(url_for('dashboard'))
    return redirect(url_for('mobile'))

@app.route('/login', methods=['GET','POST'])
def login():
    t = get_lang()
    lang_ctx = dict(t=t, langs=LANGS, current_lang=session.get('lang','it'))
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        pw = hash_pw(request.form.get('password',''))

        # Controlla se l'email corrisponde a un'azienda SaaS (admin dell'azienda)
        mdb = get_master_db()
        az = mdb.execute("SELECT * FROM aziende WHERE email_admin=? AND password_admin=?", (email, pw)).fetchone()
        mdb.close()
        if az:
            az = dict(az)
            # Verifica stato abbonamento
            if az['stato'] == 'sospeso':
                return render_template_string(LOGIN_TMPL, error='Abbonamento sospeso. Contatta il supporto.', **lang_ctx)
            # Inizializza DB tenant se non esiste
            db_path = get_tenant_db_path(az['id'])
            if not os.path.exists(db_path):
                session['azienda_id'] = az['id']
                init_db()  # crea schema sul DB tenant
            session.update({
                'user_id': 0,  # user_id speciale per il super-admin dell'azienda
                'nome': 'Admin',
                'cognome': az['nome'],
                'ruolo': 'admin',
                'email': email,
                'azienda_id': az['id'],
                'azienda_nome': az['nome'],
                'is_saas': True,
            })
            # Assicura che esista un utente admin nel DB tenant
            session_backup = dict(session)
            db = get_db()
            admin_tenant = db.execute("SELECT id FROM utenti WHERE ruolo='admin' LIMIT 1").fetchone()
            if not admin_tenant:
                db.execute(
                    "INSERT INTO utenti (nome,cognome,email,password,ruolo,titolo) VALUES (?,?,?,?,?,?)",
                    ('Admin', az['nome'], email, pw, 'admin', 'Amministratore'))
                safe_commit(db)
                admin_tenant = db.execute("SELECT id FROM utenti WHERE ruolo='admin' LIMIT 1").fetchone()
            db.close()
            session['user_id'] = admin_tenant['id']
            return redirect(url_for('dashboard'))

        # Login normale nel DB tenant corrente (o legacy)
        db = get_db()
        u = db.execute("SELECT * FROM utenti WHERE email=? AND password=? AND attivo=1",(email,pw)).fetchone()
        db.close()
        if u:
            session.update({'user_id':u['id'],'nome':u['nome'],'cognome':u['cognome'],'ruolo':u['ruolo'],'email':u['email']})
            if u['ruolo'] == 'admin':
                return redirect(url_for('dashboard'))
            else:
                return redirect(url_for('mobile'))
        return render_template_string(LOGIN_TMPL, error=t['login_error'], **lang_ctx)
    return render_template_string(LOGIN_TMPL, error=None, **lang_ctx)

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))


# ══════════════════════════════════════════════════════════
#  GOOGLE OAUTH 2.0
# ══════════════════════════════════════════════════════════

GOOGLE_AUTH_URL      = 'https://accounts.google.com/o/oauth2/v2/auth'
GOOGLE_TOKEN_URL     = 'https://oauth2.googleapis.com/token'
GOOGLE_USERINFO_URL  = 'https://www.googleapis.com/oauth2/v3/userinfo'
GOOGLE_SCOPES        = 'openid email profile'

def _google_redirect_uri():
    # Priorità: variabile ambiente > impostazione DB > fallback automatico
    env_uri = os.environ.get('GOOGLE_REDIRECT_URI', '')
    if env_uri:
        return env_uri
    db_uri = get_setting('google_redirect_uri', '')
    if db_uri:
        return db_uri
    # Fallback: forza https
    base = request.host_url.rstrip('/')
    if base.startswith('http://'):
        base = 'https://' + base[7:]
    return base + '/auth/google/callback'

@app.route('/auth/google/debug')
def auth_google_debug():
    uri = _google_redirect_uri()
    env_uri = os.environ.get('GOOGLE_REDIRECT_URI', '(non impostata)')
    db_uri = get_setting('google_redirect_uri', '(non impostata)')
    host_url = request.host_url
    return f"""<pre>
URI CHE IL CODICE MANDA A GOOGLE:
{uri}

Variabile GOOGLE_REDIRECT_URI:
{env_uri}

Impostazione DB google_redirect_uri:
{db_uri}

request.host_url grezzo:
{host_url}

COSA DEVI METTERE IN GOOGLE CONSOLE (copia esatto):
{uri}
</pre>"""


@app.route('/auth/google')
def auth_google():
    client_id = get_setting('google_client_id', '')
    if not client_id:
        flash('Google OAuth non configurato. Contatta l\'amministratore.', 'error')
        return redirect(url_for('login'))
    import secrets as _sec, urllib.parse as _up
    state = _sec.token_urlsafe(16)
    session['oauth_state'] = state
    params = {
        'client_id':     client_id,
        'redirect_uri':  _google_redirect_uri(),
        'response_type': 'code',
        'scope':         GOOGLE_SCOPES,
        'state':         state,
        'access_type':   'online',
        'prompt':        'select_account',
    }
    return redirect(GOOGLE_AUTH_URL + '?' + _up.urlencode(params))


@app.route('/auth/google/callback')
def auth_google_callback():
    import urllib.request as _ur, urllib.parse as _up, json as _j
    error = request.args.get('error')
    if error:
        flash(f'Accesso Google negato: {error}', 'error')
        return redirect(url_for('login'))

    if request.args.get('state') != session.pop('oauth_state', None):
        flash('Stato OAuth non valido.', 'error')
        return redirect(url_for('login'))

    code = request.args.get('code')
    client_id     = get_setting('google_client_id', '')
    client_secret = get_setting('google_client_secret', '')

    # Scambia il codice per un token
    token_data = _up.urlencode({
        'code':          code,
        'client_id':     client_id,
        'client_secret': client_secret,
        'redirect_uri':  _google_redirect_uri(),
        'grant_type':    'authorization_code',
    }).encode()
    try:
        req = _ur.Request(GOOGLE_TOKEN_URL, data=token_data,
                          headers={'Content-Type': 'application/x-www-form-urlencoded'})
        with _ur.urlopen(req, timeout=10) as r:
            token_resp = _j.loads(r.read())
    except Exception as ex:
        flash(f'Errore token Google: {ex}', 'error')
        return redirect(url_for('login'))

    access_token = token_resp.get('access_token')
    if not access_token:
        flash('Token Google non ricevuto.', 'error')
        return redirect(url_for('login'))

    # Ottieni info utente
    try:
        req2 = _ur.Request(GOOGLE_USERINFO_URL,
                           headers={'Authorization': f'Bearer {access_token}'})
        with _ur.urlopen(req2, timeout=10) as r:
            user_info = _j.loads(r.read())
    except Exception as ex:
        flash(f'Errore userinfo Google: {ex}', 'error')
        return redirect(url_for('login'))

    google_id  = user_info.get('sub', '')
    email      = user_info.get('email', '').lower()
    nome_g     = user_info.get('given_name', '')
    cognome_g  = user_info.get('family_name', '')
    avatar_url = user_info.get('picture', '')

    db = get_db()
    # Cerca per google_id o email
    u = db.execute("SELECT * FROM utenti WHERE google_id=? OR email=?", (google_id, email)).fetchone()

    if u:
        # Utente esistente — aggiorna google_id e avatar se mancanti
        db.execute("UPDATE utenti SET google_id=?, avatar_url=? WHERE id=?",
                   (google_id, avatar_url, u['id']))
        safe_commit(db); db.close()
        session.update({'user_id': u['id'], 'nome': u['nome'], 'cognome': u['cognome'],
                        'ruolo': u['ruolo'], 'email': u['email']})
        return redirect(url_for('dashboard') if u['ruolo'] == 'admin' else url_for('mobile'))
    else:
        # Nuovo utente — salva dati temporanei e mostra form completamento
        db.close()
        session['google_reg'] = {
            'google_id':  google_id,
            'email':      email,
            'nome':       nome_g,
            'cognome':    cognome_g,
            'avatar_url': avatar_url,
        }
        return redirect(url_for('auth_google_completa'))


@app.route('/auth/google/completa', methods=['GET', 'POST'])
def auth_google_completa():
    reg = session.get('google_reg')
    if not reg:
        return redirect(url_for('login'))

    if request.method == 'POST':
        nome     = request.form.get('nome', '').strip()
        cognome  = request.form.get('cognome', '').strip()
        telefono = request.form.get('telefono', '').strip()
        mansione = request.form.get('mansione', '').strip()
        data_ass = request.form.get('data_assunzione', '').strip()

        if not nome or not cognome:
            return render_template_string(_COMPLETA_TMPL, reg=reg, error='Nome e cognome sono obbligatori.')

        db = get_db()
        # Password casuale (non usata — login solo via Google)
        import secrets as _sec
        pw_random = hash_pw(_sec.token_hex(32))
        try:
            cur = db.execute("""INSERT INTO utenti
                (nome, cognome, email, password, ruolo, telefono, mansione,
                 data_assunzione, google_id, avatar_url, attivo)
                VALUES (?,?,?,?,?,?,?,?,?,?,1)""",
                (nome, cognome, reg['email'], pw_random, 'dipendente',
                 telefono, mansione, data_ass or None,
                 reg['google_id'], reg['avatar_url']))
            uid = cur.lastrowid
            safe_commit(db)
        except Exception as ex:
            db.close()
            return render_template_string(_COMPLETA_TMPL, reg=reg, error=f'Errore: {ex}')
        db.close()
        session.pop('google_reg', None)
        session.update({'user_id': uid, 'nome': nome, 'cognome': cognome,
                        'ruolo': 'dipendente', 'email': reg['email']})
        flash(f'Benvenuto {nome}! Profilo creato con successo.', 'success')
        return redirect(url_for('mobile'))

    return render_template_string(_COMPLETA_TMPL, reg=reg, error=None)


_COMPLETA_TMPL = """<!DOCTYPE html>
<html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Completa profilo</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;background:#f4f5f7;padding:20px}
.box{background:#fff;border-radius:16px;padding:36px;width:100%;max-width:440px;box-shadow:0 4px 40px rgba(0,0,0,.1)}
.avatar{width:68px;height:68px;border-radius:50%;object-fit:cover;border:3px solid #e2e8f0}
.avatar-ph{width:68px;height:68px;border-radius:50%;background:#0f172a;display:flex;align-items:center;justify-content:center;color:#fff;font-size:24px;font-weight:700}
h2{font-size:20px;font-weight:700;margin:14px 0 4px}
p{font-size:13px;color:#64748b;margin-bottom:20px}
label{display:block;font-size:12px;font-weight:600;color:#64748b;margin-bottom:5px}
input,select{width:100%;padding:10px 12px;border:1.5px solid #e2e8f0;border-radius:9px;font-size:14px;font-family:inherit;margin-bottom:14px}
input:focus{outline:none;border-color:#2563eb}
.btn{width:100%;padding:12px;background:#2563eb;color:#fff;border:none;border-radius:9px;font-size:15px;font-weight:700;cursor:pointer;font-family:inherit}
.alert{padding:10px;background:#fee2e2;color:#dc2626;border-radius:8px;font-size:13px;margin-bottom:14px}
.badge{background:#f0fdf4;color:#15803d;border:1px solid #86efac;padding:3px 10px;border-radius:99px;font-size:12px;font-weight:600}
</style></head>
<body>
<div class="box">
  <div style="display:flex;align-items:center;gap:14px;margin-bottom:6px">
    {% if reg.avatar_url %}
    <img src="{{ reg.avatar_url }}" class="avatar" alt="">
    {% else %}
    <div class="avatar-ph">{{ (reg.nome[0] if reg.nome else '?') }}</div>
    {% endif %}
    <div>
      <h2>Benvenuto! 👋</h2>
      <span class="badge">✓ Google {{ reg.email }}</span>
    </div>
  </div>
  <p>Completa il tuo profilo per accedere al gestionale.</p>
  {% if error %}<div class="alert">{{ error }}</div>{% endif %}
  <form method="POST">
    <label>Nome *</label>
    <input name="nome" required value="{{ reg.nome }}">
    <label>Cognome *</label>
    <input name="cognome" required value="{{ reg.cognome }}">
    <label>Telefono</label>
    <input name="telefono" type="tel" placeholder="+39 333...">
    <label>Mansione</label>
    <input name="mansione" placeholder="Es. Allestitorе, Montaggio stand...">
    <label>Data assunzione</label>
    <input name="data_assunzione" type="date">
    <button class="btn" type="submit">Crea profilo e accedi →</button>
  </form>
</div>
</body></html>"""

# ══════════════════════════════════════════════════════════
#  TRADUZIONI MULTILINGUA (dipendenti mobile)
# ══════════════════════════════════════════════════════════
LANGS = {
    'it': {
        'dir': 'ltr', 'flag': '🇮🇹', 'name': 'Italiano',
        'login_title': 'Bentornato 👋',
        'login_sub': 'Accedi al gestionale aziendale',
        'login_email': 'Email',
        'login_pass': 'Password',
        'login_btn': 'Accedi',
        'login_error': 'Email o password errati.',
        'login_lang': 'Lingua',
        'logout': 'Esci',
        'day_worked': 'Giorno lavorato',
        'site': 'Cantiere',
        'select_site': '— Seleziona fiera —',
        'hours_worked': 'Ore lavorate',
        'total_hours': 'Ore totali',
        'break': 'Pausa',
        'no_break': 'Nessuna',
        'notes': 'Note (opzionale)',
        'notes_ph': 'Es. lavoro straordinario, materiali usati...',
        'submit': 'Invia ore',
        'sending': 'Invio in corso...',
        'recent': 'Ultime registrazioni',
        'my_hours': 'Le mie ore',
        'my_hours_sub': 'Storico, filtri per mese e report PDF',
        'desktop': 'Versione desktop',
        'unknown_site': 'Cantiere non specificato',
        'waiting': 'In attesa',
        'approved': 'Approvata',
        'rejected': 'Rifiutata',
        'net_hours': 'Ore nette',
        'back': 'Indietro',
        'filter_month': 'Filtra per mese',
        'all_months': '— Tutti i mesi —',
        'total_approved': 'ore totali approvate',
        'days_worked': 'giorni lavorati',
        'pending': 'in attesa',
        'download_pdf': 'Scarica report PDF',
        'timbrature': 'Timbrature',
        'no_timbrature': 'Nessuna timbratura',
        'months': ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno','Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre'],
        'days': ['Domenica','Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato'],
        # Calendario lavori
        'work_calendar': 'Calendario Lavori',
        'work_calendar_sub': 'I tuoi servizi confermati',
        'your_role': 'IL TUO RUOLO',
        'full_team': 'Team completo',
        'drawings': 'Disegni / Progetti',
        'passes': 'I tuoi Pass',
        'no_events': 'Nessun lavoro assegnato per',
        'duration': 'durata',
        'start': 'inizio',
        'end': 'fine',
        'days_unit': 'gg',
        'download': 'Scarica',
        'back_btn': 'Torna',
        'assigned_work': 'LAVORI ASSEGNATI',
        'ciao':'Ciao','day_label':'Giorno lavorato','break_label':'Pausa (ore)','break_hint':'Es. 1h','note_ph':'Es. lavoro straordinario, materiali usati...','submit_hours':'Invia ore','sending':'Invio in corso...','last_entries':'Ultime registrazioni','desktop_ver':'Versione desktop','no_site':'Cantiere non specificato','net_hours_lbl':'nette','break_lbl':'Pausa','work_hours':'Ore lavorate','hours_suffix':'h','my_hours_title':'Le mie ore','filter_by_month':'Filtra per mese','all_months_opt':'— Tutti i mesi —','approved_total':'ore totali approvate','days_worked_lbl':'giorni lavorati','pending_lbl':'in attesa','download_pdf_btn':'Scarica report PDF','status_waiting':'In attesa','status_approved':'Approvata','status_rejected':'Rifiutata','net_h':'Ore nette','no_entries':'Nessuna timbratura','work_calendar_btn':'Calendario Lavori','spese_btn':'Spese Rimborsabili','spese_title':'Spese Rimborsabili','spese_new':'Nuova spesa','spese_date':'Data','spese_cat':'Categoria','spese_amount':'Importo (€)','spese_desc':'Descrizione','spese_vehicle':'Veicolo aziendale (opzionale)','spese_no_vehicle':'– nessun veicolo –','spese_photo_label':'Foto scontrino / fattura','spese_photo_req':'obbligatoria','spese_photo_hint':'Tocca per scattare una foto o caricare dalla galleria','spese_submit':'Invia spesa','spese_sending':'Invio in corso...','spese_history':'Ultime spese','spese_waiting':'⏳ In attesa','spese_approved':'✓ Approvata','spese_rejected':'✗ Rifiutata','spese_note_admin':'Nota:','spese_error_photo':'La foto dello scontrino è obbligatoria.','spese_success':'Spesa inviata!','report_title':'Report Ore Lavorate','report_period':'Periodo','report_days':'Giorni lavorati','report_total_hours':'Ore totali','report_date':'Data','report_site':'Cantiere','report_hours':'Ore','report_notes':'Note','report_total_row':'TOTALE','report_spese_title':'Spese Rimborsabili Approvate','report_spese_cat':'Categoria','report_spese_desc':'Descrizione','report_spese_vehicle':'Veicolo','report_spese_amount':'Importo','report_spese_total':'TOTALE RIMBORSI','report_footer':'Report generato automaticamente da','all_months_label':'Tutti i mesi',
    },
    'en': {
        'dir': 'ltr', 'flag': '🇬🇧', 'name': 'English',
        'login_title': 'Welcome back 👋',
        'login_sub': 'Sign in to the company portal',
        'login_email': 'Email',
        'login_pass': 'Password',
        'login_btn': 'Sign in',
        'login_error': 'Wrong email or password.',
        'login_lang': 'Language',
        'logout': 'Logout',
        'day_worked': 'Work day',
        'site': 'Site',
        'select_site': '— Select site —',
        'hours_worked': 'Hours worked',
        'total_hours': 'Total hours',
        'break': 'Break',
        'no_break': 'No break',
        'notes': 'Notes (optional)',
        'notes_ph': 'E.g. overtime, materials used...',
        'submit': 'Submit hours',
        'sending': 'Sending...',
        'recent': 'Recent entries',
        'my_hours': 'My hours',
        'my_hours_sub': 'History, monthly filters and PDF report',
        'desktop': 'Desktop version',
        'unknown_site': 'Site not specified',
        'waiting': 'Pending',
        'approved': 'Approved',
        'rejected': 'Rejected',
        'net_hours': 'Net hours',
        'back': 'Back',
        'filter_month': 'Filter by month',
        'all_months': '— All months —',
        'total_approved': 'total approved hours',
        'days_worked': 'days worked',
        'pending': 'pending',
        'download_pdf': 'Download PDF report',
        'timbrature': 'Time entries',
        'no_timbrature': 'No entries',
        'months': ['','January','February','March','April','May','June','July','August','September','October','November','December'],
        'days': ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'],
        # Work calendar
        'work_calendar': 'Work Calendar',
        'work_calendar_sub': 'Your confirmed assignments',
        'your_role': 'YOUR ROLE',
        'full_team': 'Full team',
        'drawings': 'Drawings / Plans',
        'passes': 'Your Passes',
        'no_events': 'No work assigned for',
        'duration': 'duration',
        'start': 'start',
        'end': 'end',
        'days_unit': 'd',
        'download': 'Download',
        'back_btn': 'Back',
        'assigned_work': 'ASSIGNED WORK',
        'ciao':'Hello','day_label':'Work day','break_label':'Break (hours)','break_hint':'E.g. 1h','note_ph':'E.g. overtime, materials used...','submit_hours':'Submit hours','sending':'Sending...','last_entries':'Recent entries','desktop_ver':'Desktop version','no_site':'Site not specified','net_hours_lbl':'net','break_lbl':'Break','work_hours':'Hours worked','hours_suffix':'h','my_hours_title':'My hours','filter_by_month':'Filter by month','all_months_opt':'— All months —','approved_total':'total approved hours','days_worked_lbl':'days worked','pending_lbl':'pending','download_pdf_btn':'Download PDF report','status_waiting':'Pending','status_approved':'Approved','status_rejected':'Rejected','net_h':'Net hours','no_entries':'No entries','work_calendar_btn':'Work Calendar','spese_btn':'Reimbursable Expenses','spese_title':'Reimbursable Expenses','spese_new':'New expense','spese_date':'Date','spese_cat':'Category','spese_amount':'Amount (€)','spese_desc':'Description','spese_vehicle':'Company vehicle (optional)','spese_no_vehicle':'– no vehicle –','spese_photo_label':'Receipt / invoice photo','spese_photo_req':'required','spese_photo_hint':'Tap to take a photo or upload from gallery','spese_submit':'Submit expense','spese_sending':'Submitting...','spese_history':'Recent expenses','spese_waiting':'⏳ Pending','spese_approved':'✓ Approved','spese_rejected':'✗ Rejected','spese_note_admin':'Note:','spese_error_photo':'Receipt photo is required.','spese_success':'Expense submitted!','report_title':'Hours Worked Report','report_period':'Period','report_days':'Days worked','report_total_hours':'Total hours','report_date':'Date','report_site':'Site','report_hours':'Hours','report_notes':'Notes','report_total_row':'TOTAL','report_spese_title':'Approved Reimbursable Expenses','report_spese_cat':'Category','report_spese_desc':'Description','report_spese_vehicle':'Vehicle','report_spese_amount':'Amount','report_spese_total':'TOTAL REIMBURSEMENTS','report_footer':'Report automatically generated by','all_months_label':'All months',
    },
    'fr': {
        'dir': 'ltr', 'flag': '🇫🇷', 'name': 'Français',
        'login_title': 'Bon retour 👋',
        'login_sub': "Connectez-vous au portail de l'entreprise",
        'login_email': 'E-mail',
        'login_pass': 'Mot de passe',
        'login_btn': 'Se connecter',
        'login_error': 'E-mail ou mot de passe incorrect.',
        'login_lang': 'Langue',
        'logout': 'Déconnexion',
        'day_worked': 'Jour travaillé',
        'site': 'Chantier',
        'select_site': '— Sélectionner chantier —',
        'hours_worked': 'Heures travaillées',
        'total_hours': 'Total heures',
        'break': 'Pause',
        'no_break': 'Aucune',
        'notes': 'Notes (facultatif)',
        'notes_ph': 'Ex. heures supplémentaires, matériaux...',
        'submit': 'Envoyer les heures',
        'sending': 'Envoi en cours...',
        'recent': 'Dernières saisies',
        'my_hours': 'Mes heures',
        'my_hours_sub': 'Historique, filtres par mois et rapport PDF',
        'desktop': 'Version bureau',
        'unknown_site': 'Chantier non spécifié',
        'waiting': 'En attente',
        'approved': 'Approuvée',
        'rejected': 'Refusée',
        'net_hours': 'Heures nettes',
        'back': 'Retour',
        'filter_month': 'Filtrer par mois',
        'all_months': '— Tous les mois —',
        'total_approved': 'heures totales approuvées',
        'days_worked': 'jours travaillés',
        'pending': 'en attente',
        'download_pdf': 'Télécharger rapport PDF',
        'timbrature': 'Saisies de temps',
        'no_timbrature': 'Aucune saisie',
        'months': ['','Janvier','Février','Mars','Avril','Mai','Juin','Juillet','Août','Septembre','Octobre','Novembre','Décembre'],
        'days': ['Dimanche','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'],
        # Calendrier des travaux
        'work_calendar': 'Calendrier des Travaux',
        'work_calendar_sub': 'Vos missions confirmées',
        'your_role': 'VOTRE RÔLE',
        'full_team': 'Équipe complète',
        'drawings': 'Plans / Projets',
        'passes': 'Vos Badges',
        'no_events': 'Aucun travail assigné pour',
        'duration': 'durée',
        'start': 'début',
        'end': 'fin',
        'days_unit': 'j',
        'download': 'Télécharger',
        'back_btn': 'Retour',
        'assigned_work': 'MISSIONS ASSIGNÉES',
        'ciao':'Bonjour','day_label':'Jour travaillé','break_label':'Pause (heures)','break_hint':'Ex. 1h','note_ph':'Ex. heures supplémentaires, matériaux...','submit_hours':'Envoyer les heures','sending':'Envoi...','last_entries':'Dernières saisies','desktop_ver':'Version bureau','no_site':'Chantier non spécifié','net_hours_lbl':'nettes','break_lbl':'Pause','work_hours':'Heures travaillées','hours_suffix':'h','my_hours_title':'Mes heures','filter_by_month':'Filtrer par mois','all_months_opt':'— Tous les mois —','approved_total':'heures approuvées','days_worked_lbl':'jours travaillés','pending_lbl':'en attente','download_pdf_btn':'Télécharger rapport PDF','status_waiting':'En attente','status_approved':'Approuvée','status_rejected':'Refusée','net_h':'Heures nettes','no_entries':'Aucune saisie','work_calendar_btn':'Calendrier Travaux','spese_btn':'Frais Remboursables','spese_title':'Frais Remboursables','spese_new':'Nouveau frais','spese_date':'Date','spese_cat':'Catégorie','spese_amount':'Montant (€)','spese_desc':'Description','spese_vehicle':'Véhicule de société (optionnel)','spese_no_vehicle':'– aucun véhicule –','spese_photo_label':'Photo reçu / facture','spese_photo_req':'obligatoire','spese_photo_hint':'Toucher pour prendre une photo ou importer de la galerie','spese_submit':'Envoyer la dépense','spese_sending':'Envoi...','spese_history':'Dernières dépenses','spese_waiting':'⏳ En attente','spese_approved':'✓ Approuvé','spese_rejected':'✗ Refusé','spese_note_admin':'Note:','spese_error_photo':'La photo du reçu est obligatoire.','spese_success':'Dépense envoyée!','report_title':'Rapport Heures Travaillées','report_period':'Période','report_days':'Jours travaillés','report_total_hours':'Total heures','report_date':'Date','report_site':'Chantier','report_hours':'Heures','report_notes':'Notes','report_total_row':'TOTAL','report_spese_title':'Frais Remboursés Approuvés','report_spese_cat':'Catégorie','report_spese_desc':'Description','report_spese_vehicle':'Véhicule','report_spese_amount':'Montant','report_spese_total':'TOTAL REMBOURSEMENTS','report_footer':'Rapport généré automatiquement par','all_months_label':'Tous les mois',
    },
    'ar': {
        'dir': 'rtl', 'flag': '🇲🇦', 'name': 'العربية',
        'login_title': 'مرحباً بعودتك 👋',
        'login_sub': 'سجّل الدخول إلى بوابة الشركة',
        'login_email': 'البريد الإلكتروني',
        'login_pass': 'كلمة المرور',
        'login_btn': 'دخول',
        'login_error': 'البريد الإلكتروني أو كلمة المرور غير صحيحة.',
        'login_lang': 'اللغة',
        'logout': 'خروج',
        'day_worked': 'يوم العمل',
        'site': 'الورشة',
        'select_site': '— اختر الورشة —',
        'hours_worked': 'ساعات العمل',
        'total_hours': 'إجمالي الساعات',
        'break': 'استراحة',
        'no_break': 'بدون استراحة',
        'notes': 'ملاحظات (اختياري)',
        'notes_ph': 'مثال: ساعات إضافية، مواد مستخدمة...',
        'submit': 'إرسال الساعات',
        'sending': 'جاري الإرسال...',
        'recent': 'آخر التسجيلات',
        'my_hours': 'ساعاتي',
        'my_hours_sub': 'السجل والفلترة الشهرية وتقرير PDF',
        'desktop': 'النسخة المكتبية',
        'unknown_site': 'ورشة غير محددة',
        'waiting': 'قيد الانتظار',
        'approved': 'موافق عليها',
        'rejected': 'مرفوضة',
        'net_hours': 'الساعات الصافية',
        'back': 'رجوع',
        'filter_month': 'تصفية حسب الشهر',
        'all_months': '— كل الأشهر —',
        'total_approved': 'ساعات معتمدة',
        'days_worked': 'أيام عمل',
        'pending': 'قيد الانتظار',
        'download_pdf': 'تحميل تقرير PDF',
        'timbrature': 'سجل الحضور',
        'no_timbrature': 'لا توجد تسجيلات',
        'months': ['','يناير','فبراير','مارس','أبريل','مايو','يونيو','يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر'],
        'days': ['الأحد','الاثنين','الثلاثاء','الأربعاء','الخميس','الجمعة','السبت'],
        # تقويم الأعمال
        'work_calendar': 'جدول الأعمال',
        'work_calendar_sub': 'مهامك المؤكدة',
        'your_role': 'دورك في هذا العمل',
        'full_team': 'الفريق الكامل',
        'drawings': 'الرسومات / المشاريع',
        'passes': 'تصاريح الدخول',
        'no_events': 'لا توجد أعمال مخصصة لشهر',
        'duration': 'المدة',
        'start': 'البداية',
        'end': 'النهاية',
        'days_unit': 'يوم',
        'download': 'تحميل',
        'back_btn': 'رجوع',
        'assigned_work': 'الأعمال المخصصة',
        'ciao':'مرحباً','day_label':'يوم العمل','break_label':'الاستراحة (ساعات)','break_hint':'مثال: 1h','note_ph':'مثال: ساعات إضافية، مواد مستخدمة...','submit_hours':'إرسال الساعات','sending':'جاري الإرسال...','last_entries':'آخر التسجيلات','desktop_ver':'النسخة المكتبية','no_site':'ورشة غير محددة','net_hours_lbl':'صافي','break_lbl':'استراحة','work_hours':'ساعات العمل','hours_suffix':'س','my_hours_title':'ساعاتي','filter_by_month':'تصفية حسب الشهر','all_months_opt':'— كل الأشهر —','approved_total':'ساعات معتمدة','days_worked_lbl':'أيام عمل','pending_lbl':'قيد الانتظار','download_pdf_btn':'تحميل تقرير PDF','status_waiting':'قيد الانتظار','status_approved':'موافق عليها','status_rejected':'مرفوضة','net_h':'الساعات الصافية','no_entries':'لا توجد تسجيلات','work_calendar_btn':'جدول الأعمال','spese_btn':'المصاريف القابلة للاسترداد','spese_title':'المصاريف القابلة للاسترداد','spese_new':'مصروف جديد','spese_date':'التاريخ','spese_cat':'الفئة','spese_amount':'المبلغ (€)','spese_desc':'الوصف','spese_vehicle':'مركبة الشركة (اختياري)','spese_no_vehicle':'– بدون مركبة –','spese_photo_label':'صورة الإيصال / الفاتورة','spese_photo_req':'إلزامي','spese_photo_hint':'اضغط لالتقاط صورة أو التحميل من المعرض','spese_submit':'إرسال المصروف','spese_sending':'جاري الإرسال...','spese_history':'آخر المصاريف','spese_waiting':'⏳ قيد الانتظار','spese_approved':'✓ موافق عليه','spese_rejected':'✗ مرفوض','spese_note_admin':'ملاحظة:','spese_error_photo':'صورة الإيصال إلزامية.','spese_success':'تم إرسال المصروف!','report_title':'تقرير ساعات العمل','report_period':'الفترة','report_days':'أيام العمل','report_total_hours':'إجمالي الساعات','report_date':'التاريخ','report_site':'الورشة','report_hours':'الساعات','report_notes':'ملاحظات','report_total_row':'المجموع','report_spese_title':'المصاريف المعتمدة القابلة للاسترداد','report_spese_cat':'الفئة','report_spese_desc':'الوصف','report_spese_vehicle':'المركبة','report_spese_amount':'المبلغ','report_spese_total':'إجمالي التعويضات','report_footer':'تم إنشاء التقرير تلقائياً بواسطة','all_months_label':'كل الأشهر',
    },
}

def get_lang():
    """Restituisce il dizionario traduzioni per la lingua corrente."""
    lang = session.get('lang', 'it')
    return LANGS.get(lang, LANGS['it'])

@app.route('/set-lang', methods=['POST'])
def set_lang():
    lang = request.form.get('lang', 'it')
    if lang in LANGS:
        session['lang'] = lang
    # Redirect back
    next_url = request.form.get('next', url_for('login'))
    return redirect(next_url)



# ══════════════════════════════════════════════════════════
#  DASHBOARD con GRAFICI
# ══════════════════════════════════════════════════════════
DASH_TMPL = """
<div class="stats-grid">
  <div class="stat-card"><div class="stat-icon blue"><i class="fa fa-users"></i></div><div><div class="stat-val">{{ s.dip }}</div><div class="stat-lbl">Dipendenti attivi</div></div></div>
  <div class="stat-card"><div class="stat-icon green"><i class="fa fa-clock"></i></div><div><div class="stat-val">{{ s.presenti }}</div><div class="stat-lbl">Presenti oggi</div></div></div>
  <div class="stat-card"><div class="stat-icon amber"><i class="fa fa-umbrella-beach"></i></div><div><div class="stat-val">{{ s.ferie }}</div><div class="stat-lbl">Ferie in attesa</div></div></div>
  <div class="stat-card"><div class="stat-icon red"><i class="fa fa-bell"></i></div><div><div class="stat-val">{{ s.richieste }}</div><div class="stat-lbl">Richieste in attesa</div></div></div>
  <div class="stat-card"><div class="stat-icon purple"><i class="fa fa-store"></i></div><div><div class="stat-val">{{ s.cantieri }}</div><div class="stat-lbl">Fiere attive</div></div></div>
</div>

<div class="grid-2" style="margin-bottom:20px">
  <!-- Grafico ore settimanali -->
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-chart-bar" style="color:var(--accent2);margin-right:8px"></i>Ore lavorate – ultimi 7 giorni</h3></div>
    <div class="card-body"><canvas id="chartOre" height="200"></canvas></div>
  </div>
  <!-- Grafico presenze per cantiere -->
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-hard-hat" style="color:var(--warning);margin-right:8px"></i>Presenze per cantiere (mese corrente)</h3></div>
    <div class="card-body"><canvas id="chartCantieri" height="200"></canvas></div>
  </div>
</div>

<div class="grid-2">
  <!-- Presenze oggi -->
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-clock" style="color:var(--accent);margin-right:8px"></i>Presenze oggi</h3><a href="/presenze" class="btn btn-secondary btn-sm">Vedi tutte</a></div>
    <div class="table-wrap">
    {% if presenze_oggi %}
    <table><thead><tr><th>Dipendente</th><th>Fiera</th><th>Entrata</th><th>Stato</th></tr></thead><tbody>
    {% for p in presenze_oggi %}
    <tr>
      <td><span class="avatar-sm">{{ p.nome[0] }}{{ p.cognome[0] }}</span>{{ p.nome }} {{ p.cognome }}</td>
      <td><span class="tag">{{ p.cantiere_nome or '–' }}</span></td>
      <td style="font-family:monospace;color:var(--success)">{{ p.ora_entrata }}</td>
      <td>{% if p.ora_uscita %}<span class="badge badge-gray">Uscito</span>{% else %}<span class="badge badge-green">● In sede</span>{% endif %}</td>
    </tr>{% endfor %}</tbody></table>
    {% else %}<div class="empty-state"><i class="fa fa-calendar-day"></i><p>Nessuna presenza oggi</p></div>{% endif %}
    </div>
  </div>
  <!-- Scadenze + ferie -->
  <div>
    <div class="card" style="margin-bottom:16px">
      <div class="card-header"><h3><i class="fa fa-calendar-exclamation" style="color:var(--warning);margin-right:8px"></i>Scadenze imminenti</h3></div>
      <div class="card-body" style="padding:14px 22px">
      {% for d in scadenze %}
      <div class="scadenza-bar">
        <div style="flex:1"><div style="font-size:13.5px;font-weight:600">{{ d.titolo }}</div>
        <div style="font-size:12px;color:var(--text-light)">{{ d.data_scadenza }}{% if d.nome %} · {{ d.nome }} {{ d.cognome }}{% endif %}</div></div>
        {% if d.days_left <= 7 %}<span class="badge badge-red">{{ d.days_left }}gg</span>
        {% elif d.days_left <= 30 %}<span class="badge badge-amber">{{ d.days_left }}gg</span>
        {% else %}<span class="badge badge-green">OK</span>{% endif %}
      </div>{% else %}<div class="empty-state" style="padding:20px"><i class="fa fa-check-circle"></i><p>Nessuna scadenza</p></div>{% endfor %}
      </div>
    </div>
    <div class="card">
      <div class="card-header"><h3><i class="fa fa-umbrella-beach" style="color:#7c3aed;margin-right:8px"></i>Ferie in attesa</h3><a href="/ferie" class="btn btn-secondary btn-sm">Gestisci</a></div>
      <div class="table-wrap">
      {% if ferie_attesa %}
      <table><thead><tr><th>Dipendente</th><th>Tipo</th><th>Dal</th><th>Al</th></tr></thead><tbody>
      {% for f in ferie_attesa %}
      <tr><td><span class="avatar-sm">{{ f.nome[0] }}{{ f.cognome[0] }}</span>{{ f.nome }} {{ f.cognome }}</td>
      <td><span class="badge badge-purple">{{ f.tipo }}</span></td>
      <td>{{ f.data_inizio }}</td><td>{{ f.data_fine }}</td></tr>{% endfor %}
      </tbody></table>
      {% else %}<div class="empty-state" style="padding:20px"><i class="fa fa-check-circle"></i><p>Nessuna richiesta</p></div>{% endif %}
      </div>
    </div>
  </div>
</div>

<script>
// Grafico ore settimanali
const oreData = {{ ore_settimana | tojson }};
new Chart(document.getElementById('chartOre'), {
  type: 'bar',
  data: {
    labels: oreData.map(d => d.data),
    datasets: [{
      label: 'Ore lavorate',
      data: oreData.map(d => d.ore),
      backgroundColor: '#2196F320',
      borderColor: '#2196F3',
      borderWidth: 2, borderRadius: 6
    }]
  },
  options: { responsive:true, plugins:{legend:{display:false}}, scales:{y:{beginAtZero:true}} }
});
// Grafico cantieri
const cantieriData = {{ presenze_cantiere | tojson }};
new Chart(document.getElementById('chartCantieri'), {
  type: 'doughnut',
  data: {
    labels: cantieriData.map(d => d.nome),
    datasets: [{
      data: cantieriData.map(d => d.count),
      backgroundColor: ['#e63946','#2196F3','#22c55e','#f59e0b','#8b5cf6','#ec4899'],
      borderWidth: 2, borderColor: '#fff'
    }]
  },
  options: { responsive:true, plugins:{legend:{position:'bottom'}} }
});
</script>"""

@app.route('/dashboard')
@login_required
def dashboard():
    db = get_db()
    today = date.today().isoformat()
    mese  = date.today().strftime('%Y-%m')

    s = {
        'dip':      db.execute("SELECT COUNT(*) FROM utenti WHERE attivo=1 AND ruolo='dipendente'").fetchone()[0],
        'presenti': db.execute("SELECT COUNT(*) FROM presenze WHERE data=? AND ora_uscita IS NULL",(today,)).fetchone()[0],
        'ferie':    db.execute("SELECT COUNT(*) FROM ferie_permessi WHERE stato='in_attesa'").fetchone()[0],
        'richieste':db.execute("SELECT COUNT(*) FROM richieste_presenze WHERE stato='in_attesa'").fetchone()[0],
        'cantieri': db.execute("SELECT COUNT(*) FROM cantieri WHERE attivo=1").fetchone()[0],
    }

    presenze_oggi = db.execute("""SELECT p.*,u.nome,u.cognome,c.nome as cantiere_nome
        FROM presenze p JOIN utenti u ON u.id=p.utente_id
        LEFT JOIN cantieri c ON c.id=p.cantiere_id WHERE p.data=?""",(today,)).fetchall()

    # Ore ultimi 7 giorni (tutti i dipendenti, solo admin)
    ore_settimana = []
    for i in range(6,-1,-1):
        d = (date.today() - timedelta(days=i)).isoformat()
        ore = db.execute("SELECT COALESCE(SUM(ore_totali),0) FROM presenze WHERE data=?",(d,)).fetchone()[0]
        ore_settimana.append({'data': d[5:], 'ore': round(ore,1)})

    # Presenze per cantiere questo mese
    presenze_cantiere = db.execute("""SELECT COALESCE(c.nome,'Non specificato') as nome, COUNT(*) as count
        FROM presenze p LEFT JOIN cantieri c ON c.id=p.cantiere_id
        WHERE p.data LIKE ? GROUP BY p.cantiere_id""",(f'{mese}%',)).fetchall()

    scadenze_raw = db.execute("""
        SELECT titolo, data_scadenza, NULL as assegnato_a, NULL as nome, NULL as cognome,
               CAST(julianday(data_scadenza)-julianday('now') AS INTEGER) as days_left
        FROM documenti WHERE data_scadenza >= date('now')
        UNION ALL
        SELECT dd.nome_originale as titolo, dd.data_scadenza,
               dd.utente_id as assegnato_a, u.nome, u.cognome,
               CAST(julianday(dd.data_scadenza)-julianday('now') AS INTEGER) as days_left
        FROM documenti_dipendente dd JOIN utenti u ON u.id=dd.utente_id
        WHERE dd.data_scadenza >= date('now')
        ORDER BY data_scadenza LIMIT 5""").fetchall()

    ferie_attesa = db.execute("""SELECT f.*,u.nome,u.cognome FROM ferie_permessi f
        JOIN utenti u ON u.id=f.utente_id WHERE f.stato='in_attesa' LIMIT 5""").fetchall()

    db.close()
    return render_page(DASH_TMPL, page_title='Dashboard', active='dashboard',
        s=s, presenze_oggi=presenze_oggi,
        ore_settimana=ore_settimana,
        presenze_cantiere=[dict(r) for r in presenze_cantiere],
        scadenze=[dict(r) for r in scadenze_raw],
        ferie_attesa=ferie_attesa)

# ══════════════════════════════════════════════════════════
#  CANTIERI
# ══════════════════════════════════════════════════════════
CANTIERI_TMPL = """
<div style="display:flex;justify-content:flex-end;margin-bottom:20px">
  <a href="/cantieri/nuovo" class="btn btn-primary"><i class="fa fa-plus"></i> Nuova fiera</a>
</div>
<div class="card">
  <div class="table-wrap">
  <table>
    <thead><tr>
      <th>Fiera / Evento</th>
      <th>Ente Organizzatore</th>
      <th>Città</th>
      <th>Padiglione / Stand</th>
      <th>Superficie</th>
      <th>Date</th>
      <th>Tipo Allestimento</th>
      <th>Stato</th>
      <th>Azioni</th>
    </tr></thead>
    <tbody>
    {% for c in cantieri %}
    <tr>
      <td>
        <i class="fa fa-store" style="color:var(--accent);margin-right:8px"></i>
        <strong>{{ c.nome }}</strong>
        {% if c.indirizzo %}<div style="font-size:11px;color:var(--text-light);margin-top:2px">{{ c.indirizzo }}</div>{% endif %}
      </td>
      <td style="color:var(--text-light);font-size:13px">{{ c.ente_organizzatore or '–' }}</td>
      <td style="font-size:13px">{{ c.citta or '–' }}</td>
      <td>
        {% if c.padiglione %}<span class="tag">{{ c.padiglione }}</span>{% endif %}
        {% if c.numero_stand %}<span class="tag">Stand {{ c.numero_stand }}</span>{% endif %}
        {% if not c.padiglione and not c.numero_stand %}–{% endif %}
      </td>
      <td style="font-size:13px">{% if c.superficie_mq %}{{ c.superficie_mq|int }} m²{% else %}–{% endif %}</td>
      <td style="font-size:11px;font-family:monospace">
        {% if c.data_inizio %}{{ c.data_inizio }}{% endif %}
        {% if c.data_inizio and c.data_fine %} →<br>{% endif %}
        {% if c.data_fine %}{{ c.data_fine }}{% endif %}
        {% if not c.data_inizio and not c.data_fine %}–{% endif %}
      </td>
      <td>
        {% set tipi = {'standard':'Standard','doppio_piano':'Doppio piano','open_space':'Open space','isola':'Isola','peninsula':'Peninsula'} %}
        <span class="tag">{{ tipi.get(c.tipo_allestimento, c.tipo_allestimento or 'Standard') }}</span>
      </td>
      <td>{% if c.attivo %}<span class="badge badge-green">● Attivo</span>{% else %}<span class="badge badge-gray">Archiviato</span>{% endif %}</td>
      <td>
        <a href="/cantieri/{{ c.id }}/modifica" class="btn btn-secondary btn-sm"><i class="fa fa-pen"></i></a>
        <a href="/cantieri/{{ c.id }}/toggle" class="btn btn-sm {{ 'btn-danger' if c.attivo else 'btn-green' }}" title="{{ 'Archivia' if c.attivo else 'Riattiva' }}">
          {{ 'Archivia' if c.attivo else 'Riattiva' }}</a>
      </td>
    </tr>{% else %}
    <tr><td colspan="9"><div class="empty-state"><i class="fa fa-store"></i><p>Nessuna fiera registrata. <a href="/cantieri/nuovo">Aggiungi la prima fiera</a>.</p></div></td></tr>
    {% endfor %}
    </tbody>
  </table>
  </div>
</div>"""

CANTIERE_FORM_TMPL = """
<div class="card" style="max-width:700px;margin:0 auto">
  <div class="card-header">
    <h3><i class="fa fa-store" style="color:var(--accent)"></i> {{ 'Modifica' if c else 'Nuova' }} Fiera / Stand</h3>
  </div>
  <div class="card-body">
  <form method="POST">
    <div class="form-group">
      <label>Nome fiera / evento *</label>
      <input name="nome" value="{{ c.nome if c else '' }}" required placeholder="es. EICMA 2025, HOST Milano, SANA Bologna">
    </div>
    <div class="form-row">
      <div class="form-group">
        <label>Ente organizzatore</label>
        <input name="ente_organizzatore" value="{{ c.ente_organizzatore if c else '' }}" placeholder="es. Fiera Milano S.p.A.">
      </div>
      <div class="form-group">
        <label>Città</label>
        <input name="citta" value="{{ c.citta if c else '' }}" placeholder="es. Milano, Bologna, Roma">
      </div>
    </div>
    <div class="form-group">
      <label>Indirizzo sede fieristica</label>
      <input name="indirizzo" value="{{ c.indirizzo if c else '' }}" placeholder="es. Fiera Milano, S.S. del Sempione 28, Rho (MI)">
    </div>
    <div class="form-row">
      <div class="form-group">
        <label>Padiglione</label>
        <input name="padiglione" value="{{ c.padiglione if c else '' }}" placeholder="es. Pad. 18, Hall B">
      </div>
      <div class="form-group">
        <label>Numero stand</label>
        <input name="numero_stand" value="{{ c.numero_stand if c else '' }}" placeholder="es. A14, B-245">
      </div>
    </div>
    <div class="form-row">
      <div class="form-group">
        <label>Superficie stand (m²)</label>
        <input name="superficie_mq" type="number" min="0" step="0.5" value="{{ c.superficie_mq if c else '' }}" placeholder="es. 120">
      </div>
      <div class="form-group">
        <label>Tipo allestimento</label>
        <select name="tipo_allestimento">
          <option value="standard" {{ 'selected' if (c and c.tipo_allestimento=='standard') or not c }}>Standard (lineare)</option>
          <option value="doppio_piano" {{ 'selected' if c and c.tipo_allestimento=='doppio_piano' }}>Doppio piano</option>
          <option value="open_space" {{ 'selected' if c and c.tipo_allestimento=='open_space' }}>Open space</option>
          <option value="isola" {{ 'selected' if c and c.tipo_allestimento=='isola' }}>Isola (4 lati aperti)</option>
          <option value="peninsula" {{ 'selected' if c and c.tipo_allestimento=='peninsula' }}>Peninsula (3 lati aperti)</option>
        </select>
      </div>
    </div>
    <div class="form-row">
      <div class="form-group">
        <label>Data inizio allestimento</label>
        <input name="data_inizio" type="date" value="{{ c.data_inizio if c else '' }}">
      </div>
      <div class="form-group">
        <label>Data fine smontaggio</label>
        <input name="data_fine" type="date" value="{{ c.data_fine if c else '' }}">
      </div>
    </div>
    <div class="form-group">
      <label>Responsabile cantiere</label>
      <input name="responsabile" value="{{ c.responsabile if c else '' }}" placeholder="Nome del capo montaggio / responsabile">
    </div>
    <div class="form-group">
      <label>Note logistica / trasporti</label>
      <textarea name="note_logistica" placeholder="es. Camion da 75 q, accesso dal cancello Sud, varco B...">{{ c.note_logistica if c else '' }}</textarea>
    </div>
    <div style="display:flex;gap:10px;justify-content:flex-end">
      <a href="/cantieri" class="btn btn-secondary">Annulla</a>
      <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva fiera</button>
    </div>
  </form>
  </div>
</div>"""

@app.route('/cantieri')
@login_required
def cantieri():
    db = get_db()
    cs = db.execute("SELECT * FROM cantieri ORDER BY attivo DESC, data_inizio DESC, nome").fetchall()
    db.close()
    return render_page(CANTIERI_TMPL, page_title='Fiere & Stand', active='cantieri', cantieri=cs)

@app.route('/cantieri/nuovo', methods=['GET','POST'])
@admin_required
def cantiere_nuovo():
    if request.method == 'POST':
        db = get_db()
        db.execute("""INSERT INTO cantieri
            (nome, indirizzo, ente_organizzatore, citta, padiglione, numero_stand,
             superficie_mq, tipo_allestimento, data_inizio, data_fine, responsabile, note_logistica)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (request.form['nome'], request.form.get('indirizzo',''),
             request.form.get('ente_organizzatore',''), request.form.get('citta',''),
             request.form.get('padiglione',''), request.form.get('numero_stand',''),
             request.form.get('superficie_mq') or None, request.form.get('tipo_allestimento','standard'),
             request.form.get('data_inizio') or None, request.form.get('data_fine') or None,
             request.form.get('responsabile',''), request.form.get('note_logistica','')))
        safe_commit(db); db.close()
        flash('Fiera aggiunta con successo!','success')
        return redirect(url_for('cantieri'))
    return render_page(CANTIERE_FORM_TMPL, page_title='Nuova Fiera', active='cantieri', c=None)

@app.route('/cantieri/<int:cid>/modifica', methods=['GET','POST'])
@admin_required
def cantiere_modifica(cid):
    db = get_db()
    c = db.execute("SELECT * FROM cantieri WHERE id=?",(cid,)).fetchone()
    if request.method == 'POST':
        db.execute("""UPDATE cantieri SET
            nome=?, indirizzo=?, ente_organizzatore=?, citta=?, padiglione=?, numero_stand=?,
            superficie_mq=?, tipo_allestimento=?, data_inizio=?, data_fine=?, responsabile=?, note_logistica=?
            WHERE id=?""",
            (request.form['nome'], request.form.get('indirizzo',''),
             request.form.get('ente_organizzatore',''), request.form.get('citta',''),
             request.form.get('padiglione',''), request.form.get('numero_stand',''),
             request.form.get('superficie_mq') or None, request.form.get('tipo_allestimento','standard'),
             request.form.get('data_inizio') or None, request.form.get('data_fine') or None,
             request.form.get('responsabile',''), request.form.get('note_logistica',''), cid))
        safe_commit(db); db.close()
        flash('Fiera aggiornata con successo.','success')
        return redirect(url_for('cantieri'))
    db.close()
    return render_page(CANTIERE_FORM_TMPL, page_title='Modifica Fiera', active='cantieri', c=c)

@app.route('/cantieri/<int:cid>/toggle')
@admin_required
def cantiere_toggle(cid):
    db = get_db()
    c = db.execute("SELECT attivo FROM cantieri WHERE id=?",(cid,)).fetchone()
    db.execute("UPDATE cantieri SET attivo=? WHERE id=?", (0 if c['attivo'] else 1, cid))
    safe_commit(db); db.close()
    flash('Fiera aggiornata.','success')
    return redirect(url_for('cantieri'))

# ══════════════════════════════════════════════════════════
#  PRESENZE  (con cantiere)
# ══════════════════════════════════════════════════════════
PRES_TMPL = """
{% if session.ruolo=='admin' %}
<div class="card" style="margin-bottom:20px;border-left:4px solid var(--accent2)" id="inserisci">
  <div class="card-header" style="cursor:pointer" onclick="var f=document.getElementById('form-manuale');f.style.display=f.style.display=='none'?'block':'none'">
    <h3><i class="fa fa-pen-to-square" style="color:var(--accent2);margin-right:8px"></i>Inserisci presenza manualmente</h3>
    <span style="font-size:12px;color:var(--text-light)">clicca per aprire/chiudere</span>
  </div>
  <div id="form-manuale" style="display:none">
  <div class="card-body">
    <form method="POST" action="/presenze/admin-inserisci" id="form-ins-pres">
      <!-- Modalità -->
      <div style="display:flex;gap:8px;margin-bottom:16px">
        <label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:8px 16px;border-radius:8px;border:2px solid var(--accent);background:var(--accent);color:#fff;font-size:13px">
          <input type="radio" name="modalita" value="ore" checked onchange="toggleModalita()" style="display:none"> ⏱ Ore totali
        </label>
        <label style="display:flex;align-items:center;gap:6px;cursor:pointer;padding:8px 16px;border-radius:8px;border:2px solid var(--accent);color:var(--accent);font-size:13px" id="lbl-orari">
          <input type="radio" name="modalita" value="orari" onchange="toggleModalita()" style="display:none"> 🕐 Entrata / Uscita
        </label>
      </div>
      <div class="form-row-4">
        <div class="form-group"><label>Dipendente *</label>
          <select name="utente_id" required id="sel-utente-ins" onchange="toggleJolly(this.value)">
            <option value="">— scegli —</option>
            {% for u in tutti_utenti %}<option value="{{ u.id }}">{{ u.nome }} {{ u.cognome }}</option>{% endfor %}
            <option value="jolly" style="color:#f59e0b;font-weight:700">👷 Jolly (esterno / nome manuale)</option>
          </select></div>
        <div class="form-group"><label>Fiera</label>
          <select name="cantiere_id"><option value="">— nessuno —</option>
          {% for c in cantieri %}<option value="{{ c.id }}">{{ c.nome }}</option>{% endfor %}
          </select></div>
        <div class="form-group"><label>Data *</label><input type="date" name="data" required max="{{ oggi_iso }}"></div>
        <div class="form-group" id="grp-ore-tot">
          <label>Ore lavorate *</label>
          <input type="number" name="ore_dirette" step="0.5" min="0.5" max="24" placeholder="Es. 9" style="font-size:18px;font-weight:700;color:var(--accent)">
        </div>
      </div>
      <!-- Campi jolly (nascosti di default) -->
      <div id="grp-jolly" style="display:none;background:#fefce8;border:1px solid #fde047;border-radius:10px;padding:14px;margin-bottom:12px">
        <div style="font-size:12px;font-weight:700;color:#92400e;margin-bottom:10px">👷 Dati lavoratore esterno</div>
        <div class="form-row">
          <div class="form-group" style="margin-bottom:0"><label>Nome *</label><input name="nome_jolly" id="inp-nome-jolly" placeholder="Mario"></div>
          <div class="form-group" style="margin-bottom:0"><label>Cognome *</label><input name="cognome_jolly" id="inp-cognome-jolly" placeholder="Rossi"></div>
        </div>
      </div>
      <div class="form-row" id="grp-orari" style="display:none">
        <div class="form-group"><label>Entrata</label><input type="time" name="ora_entrata"></div>
        <div class="form-group"><label>Uscita</label><input type="time" name="ora_uscita"></div>
      </div>
      <div class="form-group"><label>Note</label><input name="note" placeholder="Opzionale"></div>
      <button type="submit" class="btn btn-blue"><i class="fa fa-save"></i> Salva presenza</button>
    </form>
  </div>
  </div>
</div>
<script>
function toggleModalita() {
  var isOrari = document.querySelector('[name=modalita]:checked').value === 'orari';
  document.getElementById('grp-ore-tot').style.display  = isOrari ? 'none' : '';
  document.getElementById('grp-orari').style.display    = isOrari ? '' : 'none';
  var lblOre   = document.querySelector('[name=modalita][value=ore]').parentElement;
  var lblOrari = document.getElementById('lbl-orari');
  if (isOrari) {
    lblOrari.style.background = 'var(--accent)'; lblOrari.style.color = '#fff';
    lblOre.style.background = ''; lblOre.style.color = 'var(--accent)';
  } else {
    lblOre.style.background = 'var(--accent)'; lblOre.style.color = '#fff';
    lblOrari.style.background = ''; lblOrari.style.color = 'var(--accent)';
  }
}
function toggleJolly(val) {
  var isJolly = val === 'jolly';
  document.getElementById('grp-jolly').style.display = isJolly ? 'block' : 'none';
  document.getElementById('inp-nome-jolly').required = isJolly;
  document.getElementById('inp-cognome-jolly').required = isJolly;
}
</script>

<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3>Filtra</h3></div>
  <div class="card-body">
  <form method="GET" style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap" id="form-filtro-pres">
    <div class="form-group" style="margin:0;flex:1;min-width:140px"><label>Dipendente</label>
      <select name="uid"><option value="">Tutti</option>
      {% for u in tutti_utenti %}<option value="{{ u.id }}" {{ 'selected' if filtro_uid==u.id|string }}>{{ u.nome }} {{ u.cognome }}</option>{% endfor %}
      </select></div>
    <div class="form-group" style="margin:0;min-width:120px"><label>Fiera</label>
      <select name="cid"><option value="">Tutti</option>
      {% for c in cantieri %}<option value="{{ c.id }}" {{ 'selected' if filtro_cid==c.id|string }}>{{ c.nome }}</option>{% endfor %}
      </select></div>
    <div class="form-group" style="margin:0;min-width:130px"><label>Mese rapido</label>
      <select id="sel-mese" onchange="applicaMese(this.value)" style="height:40px">
        <option value="">– seleziona –</option>
        {% for m in mesi_disponibili %}
        <option value="{{ m.valore }}" {{ 'selected' if filtro_mese==m.valore }}>{{ m.etichetta }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="form-group" style="margin:0"><label>Da</label><input type="date" name="da" id="inp-da" value="{{ filtro_da }}"></div>
    <div class="form-group" style="margin:0"><label>A</label><input type="date" name="a" id="inp-a" value="{{ filtro_a }}"></div>
    <button type="submit" class="btn btn-primary" style="height:40px">Filtra</button>
    <a href="/presenze" class="btn btn-secondary" style="height:40px">Reset</a>
    <a href="/presenze/export?uid={{ filtro_uid }}&da={{ filtro_da }}&a={{ filtro_a }}&cid={{ filtro_cid }}" class="btn btn-green" style="height:40px"><i class="fa fa-file-excel"></i> Excel</a>
  </form>
  </div>
</div>
<script>
function applicaMese(val) {
  if (!val) return;
  const [y,m] = val.split('-');
  const last = new Date(parseInt(y), parseInt(m), 0).getDate();
  document.getElementById('inp-da').value = val + '-01';
  document.getElementById('inp-a').value  = val + '-' + String(last).padStart(2,'0');
  document.getElementById('form-filtro-pres').submit();
}
</script>

{% else %}
<div class="grid-2" style="margin-bottom:24px">
  <div class="clock-card">
    <div style="font-size:13px;opacity:.6;text-transform:uppercase;letter-spacing:1px">Orario attuale</div>
    <div class="clock-time" id="live-clock">--:--:--</div>
    <div style="font-size:13px;opacity:.7;margin-bottom:16px">{{ oggi }}</div>
    {% if not presenza_oggi %}
    <form method="POST" action="/presenze/entrata">
      <div class="form-group" style="margin-bottom:12px">
        <select name="cantiere_id" style="background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.2);color:#fff;border-radius:8px;padding:8px 12px;width:100%;font-size:13px">
          <option value="">📍 Seleziona fiera (opzionale)</option>
          {% for c in cantieri %}<option value="{{ c.id }}">{{ c.nome }}</option>{% endfor %}
        </select>
      </div>
      <button class="btn" style="background:var(--success);color:#fff;width:100%;font-size:14px;padding:12px"><i class="fa fa-sign-in-alt"></i> Registra Entrata</button>
    </form>
    {% elif not presenza_oggi.ora_uscita %}
    <div style="background:rgba(34,197,94,.15);border-radius:8px;padding:10px;margin-bottom:12px">
      <div style="color:#22c55e;font-size:13px">● In sede dalle <strong>{{ presenza_oggi.ora_entrata }}</strong></div>
      {% if presenza_oggi.cantiere_nome %}<div style="color:rgba(255,255,255,.6);font-size:12px;margin-top:4px">📍 {{ presenza_oggi.cantiere_nome }}</div>{% endif %}
    </div>
    <form method="POST" action="/presenze/uscita">
      <button class="btn" style="background:var(--accent);color:#fff;width:100%;font-size:14px;padding:12px"><i class="fa fa-sign-out-alt"></i> Registra Uscita</button>
    </form>
    {% else %}
    <div style="background:rgba(255,255,255,.1);border-radius:8px;padding:10px;text-align:center">
      <div style="font-size:13px;opacity:.7">Turno completato ✅</div>
      <div style="font-size:18px;font-weight:700;margin-top:4px">{{ presenza_oggi.ora_entrata }} → {{ presenza_oggi.ora_uscita }}</div>
      {% if presenza_oggi.ore_totali %}<div style="font-size:12px;opacity:.6">{{ "%.1f"|format(presenza_oggi.ore_totali) }} ore</div>{% endif %}
    </div>
    {% endif %}
  </div>
  <div class="card">
    <div class="card-header"><h3>Riepilogo mese</h3></div>
    <div class="card-body">
      <div class="stats-grid" style="grid-template-columns:1fr 1fr;gap:12px">
        <div class="stat-card" style="padding:14px"><div class="stat-icon green"><i class="fa fa-calendar-check"></i></div><div><div class="stat-val">{{ mese_stats.giorni }}</div><div class="stat-lbl">Giorni</div></div></div>
        <div class="stat-card" style="padding:14px"><div class="stat-icon blue"><i class="fa fa-clock"></i></div><div><div class="stat-val">{{ "%.0f"|format(mese_stats.ore) }}</div><div class="stat-lbl">Ore totali</div></div></div>
      </div>
    </div>
  </div>
</div>
{% endif %}

<div class="card">
  <div class="card-header" style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px">
    <h3>Storico presenze</h3>
    {% if session.ruolo=='admin' %}
    <div id="bulk-bar" style="display:none;align-items:center;gap:8px">
      <span id="bulk-count" style="font-size:13px;color:var(--text-light)">0 selezionate</span>
      <button onclick="apriBulkModifica()" class="btn btn-primary btn-sm"><i class="fa fa-pen-to-square"></i> Modifica selezionate</button>
      <button onclick="bulkElimina()" class="btn btn-danger btn-sm"><i class="fa fa-trash"></i> Elimina selezionate</button>
      <button onclick="deselezionaTutte()" class="btn btn-secondary btn-sm"><i class="fa fa-xmark"></i></button>
    </div>
    {% endif %}
  </div>
  <div class="table-wrap">
  <table id="tbl-presenze">
    <thead><tr>
      {% if session.ruolo=='admin' %}
      <th style="width:36px"><input type="checkbox" id="chk-all" onchange="toggleAll(this)" title="Seleziona tutte"></th>
      <th>Dipendente</th>
      {% endif %}
      <th>Data</th><th>Fiera</th><th>Entrata</th><th>Uscita</th><th>Ore</th><th>Note</th>
      {% if session.ruolo=='admin' %}<th></th>{% endif %}
    </tr></thead>
    <tbody>
    {% for p in presenze %}
    <tr id="row-{{ p.id }}">
      {% if session.ruolo=='admin' %}
      <td><input type="checkbox" class="chk-pres" value="{{ p.id }}"
          data-pid="{{ p.id }}" data-data="{{ p.data }}"
          data-oe="{{ p.ora_entrata or '' }}" data-ou="{{ p.ora_uscita or '' }}"
          data-ore="{{ p.ore_totali or '' }}" data-cid="{{ p.cantiere_id or '' }}"
          data-note="{{ p.note or '' }}" data-nome="{{ p.nome }} {{ p.cognome }}"
          data-jolly="{{ '1' if p.nome_jolly else '0' }}"
          onchange="aggiornaSelzione()"></td>
      <td>
        <span class="avatar-sm" style="{% if not p.utente_id or p.nome_jolly %}background:linear-gradient(135deg,#d97706,#f59e0b){% endif %}">
          {{ p.nome[0] }}{{ p.cognome[0] }}
        </span>
        {{ p.nome }} {{ p.cognome }}
        {% if p.nome_jolly %}<span style="font-size:10px;background:#fef3c7;color:#92400e;border-radius:4px;padding:1px 6px;margin-left:4px;font-weight:700">👷 JOLLY</span>{% endif %}
      </td>
      {% endif %}
      <td>{{ p.data }}</td>
      <td>{% if p.cantiere_nome %}<span class="tag"><i class="fa fa-store"></i> {{ p.cantiere_nome }}</span>{% else %}–{% endif %}</td>
      <td style="color:var(--success);font-family:monospace">{{ p.ora_entrata or '–' }}</td>
      <td style="font-family:monospace">{{ p.ora_uscita or '–' }}</td>
      <td><strong>{{ "%.1f"|format(p.ore_totali) if p.ore_totali else '–' }}</strong></td>
      <td style="color:var(--text-light);font-size:12px">{{ p.note or '' }}</td>
      {% if session.ruolo=='admin' %}<td style="display:flex;gap:4px">
        <button onclick="apriModifica({{ p.id }},'{{ p.data }}','{{ p.ora_entrata or '' }}','{{ p.ora_uscita or '' }}','{{ p.ore_totali or '' }}','{{ p.cantiere_id or '' }}','{{ p.note or '' }}','{{ p.nome }} {{ p.cognome }}')" class="btn btn-secondary btn-sm" title="Modifica"{% if p.nome_jolly %} style="opacity:.4;cursor:not-allowed" onclick="return false"{% endif %}><i class="fa fa-pen"></i></button>
        <a href="/presenze/{{ p.id }}/elimina" class="btn btn-danger btn-sm" onclick="return confirm('Eliminare?')" title="Elimina"><i class="fa fa-trash"></i></a>
      </td>{% endif %}
    </tr>{% else %}
    <tr><td colspan="9"><div class="empty-state"><i class="fa fa-history"></i><p>Nessuna presenza trovata</p></div></td></tr>
    {% endfor %}
    </tbody>
  </table>
  </div>
</div>

{% if session.ruolo=='admin' %}
<!-- Modal modifica singola (esistente) -->
<div id="modal-mod-pres" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:1000;align-items:center;justify-content:center">
  <div class="card" style="width:460px;max-width:95vw;max-height:90vh;overflow-y:auto">
    <div class="card-header">
      <h3 id="mod-pres-dip" style="font-size:15px"></h3>
      <button onclick="document.getElementById('modal-mod-pres').style.display='none'" style="background:none;border:none;font-size:22px;cursor:pointer;color:var(--text-light)">×</button>
    </div>
    <form method="POST" action="/presenze/modifica" id="form-mod-pres">
      <div class="card-body">
        <input type="hidden" name="pid" id="mod-pid">
        <div class="form-group"><label>Data</label><input type="date" name="data" id="mod-data"></div>
        <div style="display:flex;gap:8px;margin-bottom:12px">
          <label style="flex:1;display:flex;align-items:center;gap:6px;cursor:pointer;padding:8px 14px;border-radius:8px;border:2px solid var(--accent);background:var(--accent);color:#fff;font-size:13px;justify-content:center" id="mlbl-ore">
            <input type="radio" name="mod_modalita" value="ore" checked onchange="toggleModModalita()" style="display:none"> ⏱ Ore totali
          </label>
          <label style="flex:1;display:flex;align-items:center;gap:6px;cursor:pointer;padding:8px 14px;border-radius:8px;border:2px solid var(--accent);color:var(--accent);font-size:13px;justify-content:center" id="mlbl-orari">
            <input type="radio" name="mod_modalita" value="orari" onchange="toggleModModalita()" style="display:none"> 🕐 Entrata/Uscita
          </label>
        </div>
        <div id="mgrp-ore"><div class="form-group"><label>Ore totali</label><input type="number" step="0.5" name="ore_dirette" id="mod-ore" min="0.5" max="24"></div></div>
        <div id="mgrp-orari" style="display:none">
          <div class="form-row">
            <div class="form-group"><label>Entrata</label><input type="time" name="ora_entrata" id="mod-oe"></div>
            <div class="form-group"><label>Uscita</label><input type="time" name="ora_uscita" id="mod-ou"></div>
          </div>
        </div>
        <div class="form-group"><label>Fiera</label>
          <select name="cantiere_id" id="mod-cant">
            <option value="">— nessuno —</option>
            {% for c in cantieri %}<option value="{{ c.id }}">{{ c.nome }}</option>{% endfor %}
          </select>
        </div>
        <div class="form-group"><label>Note</label><input name="note" id="mod-note"></div>
      </div>
      <div style="display:flex;gap:10px;justify-content:flex-end;padding:0 20px 16px">
        <button type="button" onclick="document.getElementById('modal-mod-pres').style.display='none'" class="btn btn-secondary">Annulla</button>
        <button type="submit" class="btn btn-blue"><i class="fa fa-save"></i> Salva</button>
      </div>
    </form>
  </div>
</div>

<!-- Modal modifica BULK -->
<div id="modal-bulk" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:1001;align-items:center;justify-content:center">
  <div class="card" style="width:520px;max-width:95vw;max-height:92vh;overflow-y:auto">
    <div class="card-header" style="background:#0f172a;color:#fff">
      <h3 style="color:#fff"><i class="fa fa-pen-to-square"></i> Modifica multipla — <span id="bulk-titolo"></span></h3>
      <button onclick="document.getElementById('modal-bulk').style.display='none'" style="background:none;border:none;font-size:22px;cursor:pointer;color:rgba(255,255,255,.5)">×</button>
    </div>
    <div class="card-body">
      <div style="background:#fef3c7;border-radius:8px;padding:10px 14px;font-size:12px;color:#92400e;margin-bottom:16px">
        <i class="fa fa-info-circle"></i> Lascia vuoti i campi che non vuoi modificare. Solo i campi compilati verranno aggiornati.
      </div>

      <!-- Lista presenze selezionate -->
      <div id="bulk-lista" style="margin-bottom:16px;max-height:160px;overflow-y:auto;border:1px solid var(--border);border-radius:8px;padding:8px"></div>

      <form id="form-bulk" method="POST" action="/presenze/bulk-modifica">
        <input type="hidden" name="ids" id="bulk-ids">

        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px">
          <div class="form-group" style="margin:0">
            <label style="display:flex;align-items:center;gap:6px">
              <input type="checkbox" id="chk-campo-cantiere" onchange="toggleCampo('cantiere')"> Modifica cantiere
            </label>
            <select name="cantiere_id" id="bulk-cantiere" disabled style="margin-top:6px">
              <option value="">— nessuno —</option>
              {% for c in cantieri %}<option value="{{ c.id }}">{{ c.nome }}</option>{% endfor %}
            </select>
          </div>
          <div class="form-group" style="margin:0">
            <label style="display:flex;align-items:center;gap:6px">
              <input type="checkbox" id="chk-campo-data" onchange="toggleCampo('data')"> Modifica data
            </label>
            <input type="date" name="data_bulk" id="bulk-data" disabled style="margin-top:6px">
          </div>
        </div>

        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:12px">
          <div class="form-group" style="margin:0">
            <label style="display:flex;align-items:center;gap:6px">
              <input type="checkbox" id="chk-campo-ore" onchange="toggleCampo('ore')"> Ore totali
            </label>
            <input type="number" step="0.5" min="0.5" max="24" name="ore_bulk" id="bulk-ore" disabled placeholder="es. 8" style="margin-top:6px">
          </div>
          <div class="form-group" style="margin:0">
            <label style="display:flex;align-items:center;gap:6px">
              <input type="checkbox" id="chk-campo-oe" onchange="toggleCampo('oe')"> Ora entrata
            </label>
            <input type="time" name="oe_bulk" id="bulk-oe" disabled style="margin-top:6px">
          </div>
          <div class="form-group" style="margin:0">
            <label style="display:flex;align-items:center;gap:6px">
              <input type="checkbox" id="chk-campo-ou" onchange="toggleCampo('ou')"> Ora uscita
            </label>
            <input type="time" name="ou_bulk" id="bulk-ou" disabled style="margin-top:6px">
          </div>
        </div>

        <div class="form-group">
          <label style="display:flex;align-items:center;gap:6px">
            <input type="checkbox" id="chk-campo-note" onchange="toggleCampo('note')"> Modifica note
          </label>
          <input type="text" name="note_bulk" id="bulk-note" disabled placeholder="Note da applicare a tutte" style="margin-top:6px">
        </div>

        <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:8px">
          <button type="button" onclick="document.getElementById('modal-bulk').style.display='none'" class="btn btn-secondary">Annulla</button>
          <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Applica a tutte</button>
        </div>
      </form>
    </div>
  </div>
</div>

<script>
// ── Selezione multipla ─────────────────────────────
function toggleAll(cb) {
  document.querySelectorAll('.chk-pres').forEach(c => c.checked = cb.checked);
  aggiornaSelzione();
}
function deselezionaTutte() {
  document.querySelectorAll('.chk-pres').forEach(c => c.checked = false);
  document.getElementById('chk-all').checked = false;
  aggiornaSelzione();
}
function aggiornaSelzione() {
  var sel = document.querySelectorAll('.chk-pres:checked');
  var n = sel.length;
  var bar = document.getElementById('bulk-bar');
  bar.style.display = n > 0 ? 'flex' : 'none';
  document.getElementById('bulk-count').textContent = n + ' selezionat' + (n===1?'a':'e');
  // Aggiorna indeterminate su chk-all
  var tot = document.querySelectorAll('.chk-pres').length;
  var all = document.getElementById('chk-all');
  all.indeterminate = n > 0 && n < tot;
  all.checked = n === tot && tot > 0;
  // Evidenzia righe selezionate
  document.querySelectorAll('.chk-pres').forEach(c => {
    document.getElementById('row-' + c.value).style.background = c.checked ? '#eff6ff' : '';
  });
}

// ── Modifica bulk ──────────────────────────────────
function apriBulkModifica() {
  var sel = Array.from(document.querySelectorAll('.chk-pres:checked'));
  if (!sel.length) return;
  var ids = sel.map(c => c.value);
  document.getElementById('bulk-ids').value = ids.join(',');
  document.getElementById('bulk-titolo').textContent = ids.length + ' timbratur' + (ids.length===1?'a':'e');

  // Mostra lista presenze selezionate
  var lista = document.getElementById('bulk-lista');
  lista.innerHTML = sel.map(c => {
    var jolly = c.dataset.jolly === '1';
    return '<div style="display:flex;justify-content:space-between;align-items:center;padding:6px 4px;border-bottom:1px solid #f1f5f9">' +
      '<span style="font-size:13px;font-weight:600">' + c.dataset.nome + (jolly?' <span style="font-size:10px;color:#92400e;background:#fef3c7;border-radius:4px;padding:1px 5px">JOLLY</span>':'') + '</span>' +
      '<span style="font-size:12px;color:#64748b;font-family:monospace">' + c.dataset.data + ' · ' + (c.dataset.ore||'–') + 'h</span>' +
      '</div>';
  }).join('');

  // Reset campi
  ['cantiere','data','ore','oe','ou','note'].forEach(f => {
    document.getElementById('chk-campo-' + f).checked = false;
    toggleCampo(f);
  });

  document.getElementById('modal-bulk').style.display = 'flex';
}

function toggleCampo(campo) {
  var map = {cantiere:'bulk-cantiere', data:'bulk-data', ore:'bulk-ore', oe:'bulk-oe', ou:'bulk-ou', note:'bulk-note'};
  var el = document.getElementById(map[campo]);
  var chk = document.getElementById('chk-campo-' + campo);
  el.disabled = !chk.checked;
  el.style.opacity = chk.checked ? '1' : '.4';
}

function bulkElimina() {
  var sel = Array.from(document.querySelectorAll('.chk-pres:checked'));
  if (!sel.length) return;
  if (!confirm('Eliminare ' + sel.length + ' timbratur' + (sel.length===1?'a':'e') + '? Operazione irreversibile.')) return;
  var ids = sel.map(c => c.value).join(',');
  var f = document.createElement('form');
  f.method = 'POST'; f.action = '/presenze/bulk-elimina';
  var inp = document.createElement('input');
  inp.type = 'hidden'; inp.name = 'ids'; inp.value = ids;
  f.appendChild(inp); document.body.appendChild(f); f.submit();
}

// ── Modifica singola ───────────────────────────────
function apriModifica(pid, data, oe, ou, ore, cid, note, nome) {
  document.getElementById('mod-pid').value   = pid;
  document.getElementById('mod-data').value  = data;
  document.getElementById('mod-note').value  = note;
  document.getElementById('mod-pres-dip').textContent = '👤 ' + nome;
  var sel = document.getElementById('mod-cant');
  for (var i=0;i<sel.options.length;i++) sel.options[i].selected = (sel.options[i].value == cid);
  if (oe && ou) {
    document.querySelector('[name=mod_modalita][value=orari]').checked = true;
    document.getElementById('mod-oe').value = oe.substring(0,5);
    document.getElementById('mod-ou').value = ou.substring(0,5);
    toggleModModalita();
  } else {
    document.querySelector('[name=mod_modalita][value=ore]').checked = true;
    document.getElementById('mod-ore').value = ore ? parseFloat(ore).toFixed(1) : '';
    toggleModModalita();
  }
  document.getElementById('modal-mod-pres').style.display = 'flex';
}
function toggleModModalita() {
  var isOrari = document.querySelector('[name=mod_modalita]:checked').value === 'orari';
  document.getElementById('mgrp-ore').style.display   = isOrari ? 'none' : '';
  document.getElementById('mgrp-orari').style.display = isOrari ? '' : 'none';
  document.getElementById('mlbl-ore').style.background   = isOrari ? '' : 'var(--accent)';
  document.getElementById('mlbl-ore').style.color        = isOrari ? 'var(--accent)' : '#fff';
  document.getElementById('mlbl-orari').style.background = isOrari ? 'var(--accent)' : '';
  document.getElementById('mlbl-orari').style.color      = isOrari ? '#fff' : 'var(--accent)';
}

// Inizializza stati disabled
['cantiere','data','ore','oe','ou','note'].forEach(f => { toggleCampo(f); });
</script>
{% endif %}

{% if session.ruolo=='admin' and spese_filtrate %}
<div class="card" style="margin-top:20px;border-left:4px solid #10b981">
  <div class="card-header" style="display:flex;justify-content:space-between;align-items:center">
    <h3><i class="fa fa-receipt" style="color:#10b981"></i> Spese Rimborsabili Approvate
      {% if filtro_uid %}&nbsp;–&nbsp;{{ spese_filtrate[0].nome }} {{ spese_filtrate[0].cognome }}{% endif %}
    </h3>
    <span style="background:#f0fdf4;color:#15803d;border:1px solid #86efac;border-radius:8px;padding:5px 14px;font-size:13px;font-weight:700">
      Totale: € {{ "%.2f"|format(spese_filtrate|sum(attribute='importo')) }}
    </span>
  </div>
  <div class="table-wrap">
  <table>
    <thead><tr>
      {% if not filtro_uid %}<th>Dipendente</th>{% endif %}
      <th>Data</th><th>Categoria</th><th>Descrizione</th><th>Veicolo</th>
      <th style="text-align:right">Importo</th><th>Scontrino</th>
    </tr></thead>
    <tbody>
    {% for s in spese_filtrate %}
    <tr>
      {% if not filtro_uid %}<td><strong>{{ s.nome }} {{ s.cognome }}</strong></td>{% endif %}
      <td style="font-family:monospace">{{ s.data }}</td>
      <td><span class="tag">{{ s.categoria }}</span></td>
      <td style="color:var(--text-light);font-size:13px">{{ s.descrizione or '–' }}</td>
      <td style="font-size:12px">{{ s.veicolo_targa or '–' }}</td>
      <td style="text-align:right;font-weight:800;color:#059669">€ {{ "%.2f"|format(s.importo) }}</td>
      <td>
        {% if s.foto_nome %}
        <a href="/admin/spese/foto/{{ s.foto_nome }}" target="_blank"
           style="background:#eff6ff;color:#2563eb;border:1px solid #bfdbfe;border-radius:6px;padding:4px 10px;font-size:12px;text-decoration:none;display:inline-flex;align-items:center;gap:4px">
          <i class="fa fa-image"></i> Vedi
        </a>
        {% else %}–{% endif %}
      </td>
    </tr>
    {% endfor %}
    <tr style="background:#f0fdf4;font-weight:700">
      <td colspan="{{ 5 if filtro_uid else 6 }}" style="text-align:right;font-size:13px">TOTALE RIMBORSI</td>
      <td style="text-align:right;font-size:15px;color:#059669">€ {{ "%.2f"|format(spese_filtrate|sum(attribute='importo')) }}</td>
      <td></td>
    </tr>
    </tbody>
  </table>
  </div>
</div>
{% endif %}
{% if session.ruolo != 'admin' %}
<div class="card" style="margin-top:20px">
  <div class="card-header"><h3><i class="fa fa-paper-plane" style="color:var(--accent2);margin-right:8px"></i>Invia richiesta ore mancanti</h3></div>
  <div class="card-body">
    <form method="POST" action="/presenze/richiesta">
      <div class="form-row-4">
        <div class="form-group"><label>Data *</label><input type="date" name="data" required max="{{ oggi_iso }}"></div>
        <div class="form-group"><label>Fiera</label>
          <select name="cantiere_id"><option value="">— nessuno —</option>
          {% for c in cantieri %}<option value="{{ c.id }}">{{ c.nome }}</option>{% endfor %}
          </select></div>
        <div class="form-group"><label>Entrata *</label><input type="time" name="ora_entrata" required></div>
        <div class="form-group"><label>Uscita *</label><input type="time" name="ora_uscita" required></div>
      </div>
      <div class="form-group"><label>Motivo *</label><input name="note" required placeholder="es. Ho dimenticato di timbrare..."></div>
      <button type="submit" class="btn btn-blue"><i class="fa fa-paper-plane"></i> Invia richiesta</button>
    </form>
  </div>
</div>
{% if mie_richieste %}
<div class="card" style="margin-top:16px">
  <div class="card-header"><h3>Le mie richieste</h3></div>
  <div class="table-wrap"><table>
    <thead><tr><th>Data</th><th>Fiera</th><th>Entrata</th><th>Uscita</th><th>Motivo</th><th>Stato</th><th>Risposta</th></tr></thead>
    <tbody>{% for r in mie_richieste %}
    <tr>
      <td>{{ r.data }}</td>
      <td>{% if r.cantiere_nome %}<span class="tag">{{ r.cantiere_nome }}</span>{% else %}–{% endif %}</td>
      <td style="font-family:monospace;color:var(--success)">{{ r.ora_entrata }}</td>
      <td style="font-family:monospace">{{ r.ora_uscita }}</td>
      <td style="color:var(--text-light)">{{ r.note or '–' }}</td>
      <td>{% if r.stato=='in_attesa' %}<span class="badge badge-amber">⏳</span>
          {% elif r.stato=='approvata' %}<span class="badge badge-green">✅</span>
          {% else %}<span class="badge badge-red">❌</span>{% endif %}</td>
      <td style="font-size:12px;color:var(--text-light)">{{ r.nota_admin or '–' }}</td>
    </tr>{% endfor %}</tbody>
  </table></div>
</div>{% endif %}
{% endif %}"""

@app.route('/presenze')
@login_required
def presenze():
    db  = get_db()
    uid = session['user_id']
    today = date.today().isoformat()
    mese  = date.today().strftime('%Y-%m')

    pres_oggi = db.execute("""SELECT p.*,c.nome as cantiere_nome FROM presenze p
        LEFT JOIN cantieri c ON c.id=p.cantiere_id WHERE p.utente_id=? AND p.data=?""",(uid,today)).fetchone()

    stats = db.execute("SELECT COUNT(*) as giorni, COALESCE(SUM(ore_totali),0) as ore FROM presenze WHERE utente_id=? AND data LIKE ?",
                       (uid,f'{mese}%')).fetchone()

    filtro_uid  = request.args.get('uid','')
    filtro_da   = request.args.get('da','')
    filtro_a    = request.args.get('a','')
    filtro_cid  = request.args.get('cid','')
    filtro_mese = request.args.get('mese','')

    # Se arriva filtro_mese senza da/a, derivali
    if filtro_mese and not filtro_da and not filtro_a:
        import calendar as _cal
        try:
            y, m_n = map(int, filtro_mese.split('-'))
            last = _cal.monthrange(y, m_n)[1]
            filtro_da = f"{filtro_mese}-01"
            filtro_a  = f"{filtro_mese}-{last:02d}"
        except:
            pass

    q = """SELECT p.*,
               COALESCE(p.nome_jolly, u.nome) as nome,
               COALESCE(p.cognome_jolly, u.cognome) as cognome,
               c.nome as cantiere_nome, p.cantiere_id
           FROM presenze p
           LEFT JOIN utenti u ON u.id=p.utente_id
           LEFT JOIN cantieri c ON c.id=p.cantiere_id
           WHERE 1=1"""
    params = []
    if session['ruolo'] != 'admin': q += " AND p.utente_id=?"; params.append(uid)
    elif filtro_uid: q += " AND p.utente_id=?"; params.append(int(filtro_uid))
    if filtro_cid: q += " AND p.cantiere_id=?"; params.append(int(filtro_cid))
    if filtro_da: q += " AND p.data>=?"; params.append(filtro_da)
    if filtro_a:  q += " AND p.data<=?"; params.append(filtro_a)
    q += " ORDER BY p.data DESC, p.ora_entrata DESC LIMIT 200"

    presenze_list = db.execute(q, params).fetchall()
    tutti_utenti  = db.execute("SELECT id,nome,cognome FROM utenti WHERE attivo=1 ORDER BY cognome").fetchall()
    cantieri_list = db.execute("SELECT * FROM cantieri WHERE attivo=1 ORDER BY nome").fetchall()
    mie_richieste = db.execute("""SELECT r.*,c.nome as cantiere_nome FROM richieste_presenze r
        LEFT JOIN cantieri c ON c.id=r.cantiere_id WHERE r.utente_id=? ORDER BY r.creato_il DESC LIMIT 20""",(uid,)).fetchall()

    # Spese approvate con gli stessi filtri (solo per admin)
    spese_filtrate = []
    if session['ruolo'] == 'admin':
        sq = """SELECT s.*, u.nome, u.cognome, v.targa as veicolo_targa
                FROM spese_rimborso s JOIN utenti u ON u.id=s.utente_id
                LEFT JOIN veicoli v ON v.id=s.veicolo_id
                WHERE s.stato='approvata'"""
        sp = []
        if filtro_uid: sq += " AND s.utente_id=?"; sp.append(int(filtro_uid))
        if filtro_da:  sq += " AND s.data>=?"; sp.append(filtro_da)
        if filtro_a:   sq += " AND s.data<=?"; sp.append(filtro_a)
        sq += " ORDER BY s.data DESC"
        try:
            spese_filtrate = [dict(r) for r in db.execute(sq, sp).fetchall()]
        except:
            spese_filtrate = []

    # Mesi disponibili per il selettore rapido (ultimi 18 mesi)
    from datetime import date as _d2
    mesi_nomi = ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
                 'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']
    mesi_disponibili = []
    oggi_d = _d2.today()
    for i in range(18):
        m_off = oggi_d.month - i
        y_off = oggi_d.year
        while m_off <= 0: m_off += 12; y_off -= 1
        valore = f"{y_off}-{m_off:02d}"
        mesi_disponibili.append({'valore': valore, 'etichetta': f"{mesi_nomi[m_off]} {y_off}"})

    db.close()

    return render_page(PRES_TMPL, page_title='Presenze', active='presenze',
        oggi=date.today().strftime('%A %d %B %Y'), oggi_iso=today,
        presenza_oggi=pres_oggi,
        mese_stats={'giorni':stats['giorni'],'ore':stats['ore']},
        presenze=presenze_list, tutti_utenti=tutti_utenti, cantieri=cantieri_list,
        filtro_uid=filtro_uid, filtro_da=filtro_da, filtro_a=filtro_a, filtro_cid=filtro_cid,
        filtro_mese=filtro_mese, mesi_disponibili=mesi_disponibili,
        spese_filtrate=spese_filtrate,
        mie_richieste=mie_richieste)

@app.route('/presenze/entrata', methods=['POST'])
@login_required
def presenza_entrata():
    uid = session['user_id']; today = date.today().isoformat(); now = datetime.now().strftime('%H:%M:%S')
    cid = request.form.get('cantiere_id') or None
    db = get_db()
    if not db.execute("SELECT id FROM presenze WHERE utente_id=? AND data=?",(uid,today)).fetchone():
        db.execute("INSERT INTO presenze (utente_id,data,ora_entrata,cantiere_id) VALUES (?,?,?,?)",(uid,today,now,cid))
        safe_commit(db); flash(f'Entrata alle {now} ✅','success')
    else: flash("Entrata già registrata oggi.",'error')
    db.close(); return redirect(url_for('presenze'))

@app.route('/presenze/uscita', methods=['POST'])
@login_required
def presenza_uscita():
    uid = session['user_id']; today = date.today().isoformat(); now = datetime.now().strftime('%H:%M:%S')
    db = get_db()
    p = db.execute("SELECT * FROM presenze WHERE utente_id=? AND data=?",(uid,today)).fetchone()
    if p and not p['ora_uscita']:
        try:
            diff = datetime.strptime(now,'%H:%M:%S') - datetime.strptime(p['ora_entrata'],'%H:%M:%S')
            ore  = round(diff.total_seconds()/3600,2)
        except: ore = None
        db.execute("UPDATE presenze SET ora_uscita=?,ore_totali=? WHERE id=?",(now,ore,p['id']))
        safe_commit(db); flash(f'Uscita alle {now} — {ore:.1f}h ✅','success')
    db.close(); return redirect(url_for('presenze'))

@app.route('/presenze/admin-inserisci', methods=['POST'])
@admin_required
def admin_inserisci_presenza():
    ensure_columns()  # garantisce nome_jolly/cognome_jolly esistano
    uid_dest = request.form.get('utente_id'); data = request.form.get('data','')
    cid  = request.form.get('cantiere_id') or None
    note = request.form.get('note','') or 'Inserita da admin'
    modalita = request.form.get('modalita','ore')
    is_jolly = (uid_dest == 'jolly')

    if not uid_dest or not data:
        flash('Compila dipendente e data.','error'); return redirect(url_for('presenze'))

    if is_jolly:
        nome_jolly    = (request.form.get('nome_jolly','') or '').strip()
        cognome_jolly = (request.form.get('cognome_jolly','') or '').strip()
        if not nome_jolly or not cognome_jolly:
            flash('Inserisci nome e cognome del lavoratore esterno.','error')
            return redirect(url_for('presenze'))
        uid_dest = None  # nessun utente_id reale
    else:
        nome_jolly = cognome_jolly = None

    if modalita == 'orari':
        oe = request.form.get('ora_entrata','')
        ou = request.form.get('ora_uscita','')
        if not oe or not ou:
            flash('Inserisci entrata e uscita.','error'); return redirect(url_for('presenze'))
        try:
            diff = datetime.strptime(ou,'%H:%M') - datetime.strptime(oe,'%H:%M')
            ore  = round(diff.total_seconds()/3600,2)
            if ore <= 0: raise ValueError
        except:
            flash("Orari non validi.",'error'); return redirect(url_for('presenze'))
    else:
        ore_s = request.form.get('ore_dirette','')
        try:
            ore = round(float(ore_s),2)
            if ore <= 0: raise ValueError
        except:
            flash("Inserisci un numero di ore valido.",'error'); return redirect(url_for('presenze'))
        oe = '00:00'; ou = '00:00'

    db = get_db()

    if is_jolly:
        # Jolly: inserimento diretto senza controllo duplicato (nessun utente_id)
        db.execute(
            "INSERT INTO presenze (utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note,nome_jolly,cognome_jolly) VALUES (?,?,?,?,?,?,?,?,?)",
            (None, data, oe, ou, ore, cid, note, nome_jolly, cognome_jolly))
        flash(f'Presenza jolly {nome_jolly} {cognome_jolly} del {data} inserita!','success')
    else:
        # L'admin può inserire più presenze nello stesso giorno (cantieri diversi)
        ex = db.execute("SELECT id FROM presenze WHERE utente_id=? AND data=? AND cantiere_id IS ? AND (cantiere_id=? OR (cantiere_id IS NULL AND ? IS NULL))",
                        (uid_dest, data, cid, cid, cid)).fetchone()
        if ex:
            db.execute("UPDATE presenze SET ora_entrata=?,ora_uscita=?,ore_totali=?,note=? WHERE id=?",
                       (oe,ou,ore,note,ex['id']))
            flash(f'Presenza del {data} (stesso cantiere) aggiornata!','success')
        else:
            db.execute("INSERT INTO presenze (utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note) VALUES (?,?,?,?,?,?,?)",
                       (uid_dest,data,oe,ou,ore,cid,note))
            # Sincronizza anche in richieste_presenze (per mostrarlo nel report ore del dipendente)
            gia_req = db.execute(
                "SELECT id FROM richieste_presenze WHERE utente_id=? AND data=? AND cantiere_id IS ?",
                (uid_dest, data, cid)).fetchone()
            if not gia_req:
                db.execute(
                    "INSERT INTO richieste_presenze (utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note,stato) VALUES (?,?,?,?,?,?,?,?)",
                    (uid_dest, data, oe, ou, ore, cid, note, 'approvata'))
            else:
                db.execute(
                    "UPDATE richieste_presenze SET ora_entrata=?,ora_uscita=?,ore_totali=?,stato='approvata',note=? WHERE id=?",
                    (oe, ou, ore, note, gia_req['id']))
            flash(f'Presenza del {data} inserita!','success')

    safe_commit(db); db.close(); return redirect(url_for('presenze'))

@app.route('/presenze/modifica', methods=['POST'])
@admin_required
def presenza_modifica():
    pid  = int(request.form.get('pid',0))
    data = request.form.get('data','')
    cid  = request.form.get('cantiere_id') or None
    note = request.form.get('note','')
    modalita = request.form.get('mod_modalita','ore')

    if modalita == 'orari':
        oe = request.form.get('ora_entrata','')
        ou = request.form.get('ora_uscita','')
        if not oe or not ou:
            flash('Inserisci entrata e uscita.','error'); return redirect(url_for('presenze'))
        try:
            diff = datetime.strptime(ou,'%H:%M') - datetime.strptime(oe,'%H:%M')
            ore  = round(diff.total_seconds()/3600,2)
            if ore <= 0: raise ValueError
        except:
            flash('Orari non validi.','error'); return redirect(url_for('presenze'))
    else:
        ore_s = request.form.get('ore_dirette','')
        try:
            ore = round(float(ore_s),2)
            if ore <= 0: raise ValueError
        except:
            flash('Ore non valide.','error'); return redirect(url_for('presenze'))
        oe = '00:00'; ou = '00:00'

    db = get_db()
    db.execute(
        "UPDATE presenze SET data=?,ora_entrata=?,ora_uscita=?,ore_totali=?,cantiere_id=?,note=? WHERE id=?",
        (data, oe, ou, ore, cid, note, pid))
    # Aggiorna anche richieste_presenze se esiste
    db.execute(
        "UPDATE richieste_presenze SET data=?,ora_entrata=?,ora_uscita=?,ore_totali=?,cantiere_id=?,note=? WHERE id=(SELECT id FROM richieste_presenze WHERE utente_id=(SELECT utente_id FROM presenze WHERE id=?) AND data=(SELECT data FROM presenze WHERE id=?) LIMIT 1)",
        (data, oe, ou, ore, cid, note, pid, pid))
    safe_commit(db); db.close()
    flash('Presenza aggiornata!','success')
    return redirect(url_for('presenze'))


@app.route('/presenze/<int:pid>/elimina')
@admin_required
def presenza_elimina(pid):
    db = get_db()
    db.execute("DELETE FROM presenze WHERE id=?", (pid,))
    safe_commit(db); db.close()
    flash('Presenza eliminata.','success'); return redirect(url_for('presenze'))


@app.route('/presenze/bulk-modifica', methods=['POST'])
@admin_required
def presenze_bulk_modifica():
    ids_raw = request.form.get('ids','')
    try:
        ids = [int(x) for x in ids_raw.split(',') if x.strip()]
    except:
        flash('Nessuna presenza selezionata.','error')
        return redirect(url_for('presenze'))
    if not ids:
        flash('Nessuna presenza selezionata.','error')
        return redirect(url_for('presenze'))

    # Raccogli solo i campi che l'utente ha deciso di modificare
    aggiornamenti = {}

    # Cantiere
    if request.form.get('cantiere_id') is not None and 'cantiere_id' in request.form:
        cid = request.form.get('cantiere_id') or None
        aggiornamenti['cantiere_id'] = cid

    # Data
    data_bulk = request.form.get('data_bulk','').strip()
    if data_bulk:
        aggiornamenti['data'] = data_bulk

    # Ore
    ore_bulk = request.form.get('ore_bulk','').strip()
    if ore_bulk:
        try:
            ore = round(float(ore_bulk), 2)
            if ore > 0:
                aggiornamenti['ore_totali'] = ore
        except: pass

    # Entrata
    oe_bulk = request.form.get('oe_bulk','').strip()
    if oe_bulk:
        aggiornamenti['ora_entrata'] = oe_bulk

    # Uscita
    ou_bulk = request.form.get('ou_bulk','').strip()
    if ou_bulk:
        aggiornamenti['ora_uscita'] = ou_bulk
        # Ricalcola ore se abbiamo entrata e uscita
        if oe_bulk and ou_bulk:
            try:
                diff = datetime.strptime(ou_bulk,'%H:%M') - datetime.strptime(oe_bulk,'%H:%M')
                ore_calc = round(diff.total_seconds()/3600, 2)
                if ore_calc > 0:
                    aggiornamenti['ore_totali'] = ore_calc
            except: pass

    # Note
    note_bulk = request.form.get('note_bulk','')
    if 'note_bulk' in request.form and request.form.get('note_bulk') != '':
        aggiornamenti['note'] = note_bulk

    if not aggiornamenti:
        flash('Nessun campo selezionato da modificare.','error')
        return redirect(url_for('presenze'))

    db = get_db()
    set_clause = ', '.join(f"{k}=?" for k in aggiornamenti)
    vals = list(aggiornamenti.values())
    n_ok = 0
    for pid in ids:
        try:
            db.execute(f"UPDATE presenze SET {set_clause} WHERE id=?", vals + [pid])
            n_ok += 1
        except: pass
    safe_commit(db); db.close()
    flash(f'{n_ok} presenze aggiornate con successo!','success')
    return redirect(url_for('presenze'))


@app.route('/presenze/bulk-elimina', methods=['POST'])
@admin_required
def presenze_bulk_elimina():
    ids_raw = request.form.get('ids','')
    try:
        ids = [int(x) for x in ids_raw.split(',') if x.strip()]
    except:
        flash('Nessuna presenza selezionata.','error')
        return redirect(url_for('presenze'))
    if not ids:
        flash('Nessuna presenza selezionata.','error')
        return redirect(url_for('presenze'))
    db = get_db()
    placeholders = ','.join('?' for _ in ids)
    db.execute(f"DELETE FROM presenze WHERE id IN ({placeholders})", ids)
    safe_commit(db); db.close()
    flash(f'{len(ids)} presenze eliminate.','success')
    return redirect(url_for('presenze'))


# ══════════════════════════════════════════════════════════
#  IMPORT ORE STORICO — AI
# ══════════════════════════════════════════════════════════

IMPORT_ORE_TMPL = """
<div style="margin-bottom:18px">
  <a href="/presenze" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Presenze</a>
</div>
<div class="card" style="max-width:860px;margin:0 auto">
  <div class="card-header">
    <h3><i class="fa fa-robot" style="color:var(--accent2)"></i> Import ore storiche con AI</h3>
    <div style="color:var(--text-light);font-size:13px;margin-top:2px">Carica un file (PDF/immagine/testo) con le timbrature di un mese passato — l'AI estrae e inserisce automaticamente le ore per il dipendente selezionato.</div>
  </div>
  <div class="card-body">

    <!-- Step 1: seleziona dipendente + upload -->
    <div id="step1">
      <div class="form-row" style="margin-bottom:16px">
        <div class="form-group">
          <label>Dipendente *</label>
          <select id="sel_dip" style="width:100%">
            <option value="">— Seleziona dipendente —</option>
            {% for d in dipendenti %}
            <option value="{{ d.id }}">{{ d.nome }} {{ d.cognome }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="form-group">
          <label>Mese di riferimento *</label>
          <input type="month" id="inp_mese" value="{{ mese_default }}">
        </div>
      </div>
      <div class="form-group" style="margin-bottom:16px">
        <label>File timbrature (PDF, immagine, testo) *</label>
        <input type="file" id="inp_file" accept=".pdf,.png,.jpg,.jpeg,.txt,.csv" style="width:100%">
      </div>
      <div id="err_step1" style="color:#dc2626;font-size:13px;display:none;margin-bottom:8px"></div>
      <button class="btn btn-primary" onclick="analizzaFile()"><i class="fa fa-magic"></i> Analizza con AI</button>
    </div>

    <!-- Loader -->
    <div id="loader" style="display:none;text-align:center;padding:40px">
      <div style="font-size:32px;margin-bottom:12px">🤖</div>
      <div style="font-weight:700;color:var(--accent2);font-size:16px">L'AI sta analizzando il file...</div>
      <div style="color:var(--text-light);margin-top:6px;font-size:13px">Estrazione date e ore in corso</div>
    </div>

    <!-- Step 2: anteprima risultati -->
    <div id="step2" style="display:none">
      <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:10px;padding:14px;margin-bottom:18px">
        <strong style="color:#15803d"><i class="fa fa-check-circle"></i> AI ha estratto <span id="n_righe">0</span> giorni lavorativi</strong>
        <span style="color:var(--text-light);font-size:13px;margin-left:8px">— Verifica i dati prima di confermare</span>
      </div>
      <div style="overflow-x:auto;margin-bottom:16px">
        <table style="width:100%;border-collapse:collapse" id="tab_preview">
          <thead>
            <tr style="background:#0f172a;color:#fff">
              <th style="padding:8px;text-align:center;font-size:12px;width:30px">✓</th>
              <th style="padding:8px;text-align:left;font-size:12px">Data</th>
              <th style="padding:8px;text-align:center;font-size:12px">Ore</th>
              <th style="padding:8px;text-align:center;font-size:12px">Entrata</th>
              <th style="padding:8px;text-align:center;font-size:12px">Uscita</th>
              <th style="padding:8px;text-align:left;font-size:12px">Note AI</th>
            </tr>
          </thead>
          <tbody id="tbody_prev"></tbody>
        </table>
      </div>
      <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap">
        <button class="btn btn-primary" onclick="confermaSalva()"><i class="fa fa-save"></i> Salva ore selezionate</button>
        <button class="btn btn-secondary" onclick="resetForm()"><i class="fa fa-redo"></i> Rianalizza</button>
        <span id="msg_salva" style="font-size:13px;color:#15803d;display:none"><i class="fa fa-check"></i> Salvato!</span>
      </div>
    </div>

  </div>
</div>

<script>
var _righeAI = [];
var _dipId = '';
var _mese  = '';

function analizzaFile() {
  var dip  = document.getElementById('sel_dip').value;
  var mese = document.getElementById('inp_mese').value;
  var file = document.getElementById('inp_file').files[0];
  var err  = document.getElementById('err_step1');
  if (!dip || !mese || !file) { err.textContent='Seleziona dipendente, mese e file.'; err.style.display=''; return; }
  err.style.display='none';
  _dipId = dip; _mese = mese;
  document.getElementById('step1').style.display='none';
  document.getElementById('loader').style.display='';

  var reader = new FileReader();
  reader.onload = function(e) {
    var b64 = e.target.result.split(',')[1];
    var mime = file.type || 'application/octet-stream';
    fetch('/presenze/import-ai/analizza', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({file_b64: b64, mime: mime, filename: file.name, mese: mese, dip_id: dip})
    }).then(r=>r.json()).then(function(data) {
      document.getElementById('loader').style.display='none';
      if (data.error) { document.getElementById('step1').style.display=''; document.getElementById('err_step1').textContent=data.error; document.getElementById('err_step1').style.display=''; return; }
      _righeAI = data.righe;
      document.getElementById('n_righe').textContent = _righeAI.length;
      var tbody = document.getElementById('tbody_prev');
      tbody.innerHTML='';
      _righeAI.forEach(function(r,i) {
        var tr = document.createElement('tr');
        tr.style.borderBottom='1px solid #e2e8f0';
        var bg = i%2==0 ? '#fff' : '#f8fafc';
        tr.style.background = bg;
        tr.innerHTML = '<td style="padding:6px;text-align:center"><input type="checkbox" checked id="chk_'+i+'"></td>'+
          '<td style="padding:6px;font-weight:600">'+r.data+'</td>'+
          '<td style="padding:6px;text-align:center"><input type="number" step="0.5" min="0" max="24" value="'+r.ore+'" id="ore_'+i+'" style="width:60px;text-align:center;border:1px solid #e2e8f0;border-radius:4px;padding:2px"></td>'+
          '<td style="padding:6px;text-align:center;font-size:12px;color:var(--text-light)">'+( r.entrata||'—')+'</td>'+
          '<td style="padding:6px;text-align:center;font-size:12px;color:var(--text-light)">'+(r.uscita||'—')+'</td>'+
          '<td style="padding:6px;font-size:12px;color:var(--text-light)">'+( r.nota||'')+'</td>';
        tbody.appendChild(tr);
      });
      document.getElementById('step2').style.display='';
    }).catch(function(ex) {
      document.getElementById('loader').style.display='none';
      document.getElementById('step1').style.display='';
      document.getElementById('err_step1').textContent='Errore di connessione: '+ex;
      document.getElementById('err_step1').style.display='';
    });
  };
  reader.readAsDataURL(file);
}

function confermaSalva() {
  var righe_sel = [];
  _righeAI.forEach(function(r,i) {
    if (document.getElementById('chk_'+i) && document.getElementById('chk_'+i).checked) {
      var ore_val = parseFloat(document.getElementById('ore_'+i).value)||0;
      if (ore_val > 0) righe_sel.push({data: r.data, ore: ore_val, entrata: r.entrata||'00:00', uscita: r.uscita||'00:00'});
    }
  });
  if (!righe_sel.length) { alert('Nessuna riga selezionata.'); return; }
  fetch('/presenze/import-ai/salva', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({dip_id: _dipId, righe: righe_sel})
  }).then(r=>r.json()).then(function(d) {
    document.getElementById('msg_salva').textContent = '✓ Salvate '+d.saved+' presenze!';
    document.getElementById('msg_salva').style.display='';
    setTimeout(function(){ window.location.href='/presenze'; }, 1800);
  });
}

function resetForm() {
  _righeAI=[]; _dipId=''; _mese='';
  document.getElementById('step2').style.display='none';
  document.getElementById('step1').style.display='';
  document.getElementById('msg_salva').style.display='none';
  document.getElementById('inp_file').value='';
}
</script>
"""


@app.route('/presenze/import-ai')
@admin_required
def presenze_import_ai():
    db = get_db()
    dipendenti = db.execute("SELECT id,nome,cognome FROM utenti WHERE ruolo='dipendente' AND attivo=1 ORDER BY cognome,nome").fetchall()
    db.close()
    from datetime import date as _date
    mese_default = _date.today().strftime('%Y-%m')
    return render_page(IMPORT_ORE_TMPL, page_title='Import ore AI', active='presenze',
                       dipendenti=dipendenti, mese_default=mese_default)


@app.route('/presenze/import-ai/analizza', methods=['POST'])
@admin_required
def presenze_import_analizza():
    import base64, json as _json
    data = request.get_json()
    file_b64 = data.get('file_b64','')
    mime     = data.get('mime','application/octet-stream')
    filename = data.get('filename','')
    mese     = data.get('mese','')

    api_key = get_setting('anthropic_api_key','')
    if not api_key:
        return jsonify({'error':'API key Anthropic non configurata. Vai in Impostazioni → AI.'})

    # Prepara il contenuto per Claude
    import urllib.request as _urlreq
    prompt = f"""Sei un assistente HR. Analizza il documento allegato che contiene le timbrature o le ore di lavoro per il mese {mese}.
Estrai per ogni giorno lavorativo: data (formato YYYY-MM-DD), ore totali lavorate, ora entrata (HH:MM se presente), ora uscita (HH:MM se presente).
Ignora giorni festivi, sabati, domeniche se non lavorati.
Rispondi SOLO con un JSON array senza markdown, esempio:
[{{"data":"2026-02-03","ore":8.0,"entrata":"08:00","uscita":"16:00","nota":""}},{{"data":"2026-02-04","ore":7.5,"entrata":"08:30","uscita":"16:00","nota":"uscita anticipata"}}]
Se non trovi ore per un giorno, non includerlo. Se l'ora non è presente metti null."""

    # Costruisci messaggio con il file
    file_bytes = base64.b64decode(file_b64)

    if mime.startswith('image/'):
        content_part = {"type":"image","source":{"type":"base64","media_type":mime,"data":file_b64}}
    elif mime == 'application/pdf':
        content_part = {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":file_b64}}
    else:
        # Testo semplice
        try:
            testo = file_bytes.decode('utf-8','replace')
        except:
            testo = file_bytes.decode('latin-1','replace')
        content_part = {"type":"text","text":f"File: {filename}\n\n{testo}"}

    payload = _json.dumps({
        "model": "claude-sonnet-4-20250514",
        "max_tokens": 4096,
        "messages": [{"role":"user","content":[content_part,{"type":"text","text":prompt}]}]
    }).encode()

    req = _urlreq.Request('https://api.anthropic.com/v1/messages',
        data=payload,
        headers={
            'Content-Type':'application/json',
            'x-api-key': api_key,
            'anthropic-version':'2023-06-01'
        }, method='POST')
    try:
        with _urlreq.urlopen(req, timeout=60) as resp:
            result = _json.loads(resp.read())
        text = result['content'][0]['text'].strip()
        # Pulisci eventuali markdown
        if text.startswith('```'): text = text.split('\n',1)[1]
        if text.endswith('```'): text = text.rsplit('```',1)[0]
        righe = _json.loads(text.strip())
        # Valida e normalizza
        out = []
        for r in righe:
            if not r.get('data'): continue
            ore_v = float(r.get('ore') or 0)
            if ore_v <= 0: continue
            out.append({'data': r['data'], 'ore': round(ore_v,2),
                        'entrata': r.get('entrata') or '', 'uscita': r.get('uscita') or '',
                        'nota': r.get('nota') or ''})
        out.sort(key=lambda x: x['data'])
        return jsonify({'righe': out})
    except Exception as ex:
        return jsonify({'error': f'Errore AI: {ex}'})


@app.route('/presenze/import-ai/salva', methods=['POST'])
@admin_required
def presenze_import_salva():
    data = request.get_json()
    dip_id = int(data.get('dip_id',0))
    righe  = data.get('righe',[])
    db = get_db()
    saved = 0
    for r in righe:
        d    = r.get('data','')
        ore  = float(r.get('ore',0))
        oe   = r.get('entrata','00:00') or '00:00'
        ou   = r.get('uscita','00:00')  or '00:00'
        if not d or ore <= 0: continue
        ex = db.execute("SELECT id FROM presenze WHERE utente_id=? AND data=? AND cantiere_id IS NULL",
                        (dip_id, d)).fetchone()
        if ex:
            db.execute("UPDATE presenze SET ora_entrata=?,ora_uscita=?,ore_totali=?,note=? WHERE id=?",
                       (oe, ou, ore, 'Import storico AI', ex['id']))
        else:
            db.execute("INSERT INTO presenze (utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note) VALUES (?,?,?,?,?,?,?)",
                       (dip_id, d, oe, ou, ore, None, 'Import storico AI'))
        # Aggiorna anche richieste_presenze per la webapp dipendente
        ex2 = db.execute("SELECT id FROM richieste_presenze WHERE utente_id=? AND data=? AND cantiere_id IS NULL",
                         (dip_id, d)).fetchone()
        if ex2:
            db.execute("UPDATE richieste_presenze SET ora_entrata=?,ora_uscita=?,ore_totali=?,stato='approvata',note=? WHERE id=?",
                       (oe, ou, ore, 'Import storico AI', ex2['id']))
        else:
            db.execute("INSERT INTO richieste_presenze (utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note,stato) VALUES (?,?,?,?,?,?,?,?)",
                       (dip_id, d, oe, ou, ore, None, 'Import storico AI', 'approvata'))
        saved += 1
    safe_commit(db); db.close()
    return jsonify({'saved': saved})


@app.route('/presenze/richiesta', methods=['POST'])
@login_required
def invia_richiesta_presenza():
    data = request.form.get('data',''); oe = request.form.get('ora_entrata','')
    ou   = request.form.get('ora_uscita',''); note = request.form.get('note','')
    cid  = request.form.get('cantiere_id') or None
    if not data or not oe or not ou: flash('Compila tutti i campi.','error'); return redirect(url_for('presenze'))
    if data >= date.today().isoformat(): flash('Solo giorni passati.','error'); return redirect(url_for('presenze'))
    try:
        diff = datetime.strptime(ou,'%H:%M') - datetime.strptime(oe,'%H:%M')
        ore  = round(diff.total_seconds()/3600,2)
        if ore <= 0: raise ValueError
    except: flash("Orari non validi.",'error'); return redirect(url_for('presenze'))
    db = get_db()
    if db.execute("SELECT id FROM richieste_presenze WHERE utente_id=? AND data=? AND cantiere_id IS ? AND stato='in_attesa'",
                  (session['user_id'], data, int(cid) if cid else None)).fetchone():
        flash('Hai già una richiesta in attesa per questo giorno e cantiere.','error'); db.close(); return redirect(url_for('presenze'))
    db.execute("INSERT INTO richieste_presenze (utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note) VALUES (?,?,?,?,?,?,?)",
               (session['user_id'],data,oe,ou,ore,cid,note))
    safe_commit(db)
    db.close()
    # Email notifica in background
    email_admin = get_setting('email_notifiche','')
    _nome_d = f"{session.get('nome','')} {session.get('cognome','')}".strip()
    if email_admin:
        import threading
        def _send_desk():
            try:
                send_email(email_admin, f'[ACCESSO FIERE] Nuova richiesta da {_nome_d}',
                    f'<p><b>{_nome_d}</b> ha inviato una richiesta di presenza per il <b>{data}</b>.<br>'
                    f'Ore: {oe}–{ou} ({ore:.1f}h)<br>Motivo: {note}</p>'
                    f'<p><a href="/admin/richieste">Vai alle richieste</a></p>')
            except: pass
        threading.Thread(target=_send_desk, daemon=True).start()
    flash('Richiesta inviata!','success'); return redirect(url_for('presenze'))

# ── Export Excel ──────────────────────────────────────────
@app.route('/presenze/export')
@admin_required
def export_presenze():
    if not EXCEL_OK:
        flash('Installa openpyxl: pip install openpyxl','error')
        return redirect(url_for('presenze'))
    uid_f = request.args.get('uid',''); da = request.args.get('da','')
    a_    = request.args.get('a','');   cid_f = request.args.get('cid','')
    db = get_db()
    q = """SELECT p.*,
               COALESCE(p.nome_jolly, u.nome) as nome,
               COALESCE(p.cognome_jolly, u.cognome) as cognome,
               c.nome as cantiere_nome
           FROM presenze p
           LEFT JOIN utenti u ON u.id=p.utente_id
           LEFT JOIN cantieri c ON c.id=p.cantiere_id
           WHERE 1=1"""
    params = []
    if uid_f: q += " AND p.utente_id=?"; params.append(int(uid_f))
    if cid_f: q += " AND p.cantiere_id=?"; params.append(int(cid_f))
    if da:    q += " AND p.data>=?"; params.append(da)
    if a_:    q += " AND p.data<=?"; params.append(a_)
    q += " ORDER BY p.data, COALESCE(p.cognome_jolly, u.cognome)"
    rows = db.execute(q, params).fetchall()

    # Spese approvate con gli stessi filtri utente/date
    sq = """SELECT s.*, u.nome, u.cognome, v.targa as veicolo_targa
            FROM spese_rimborso s JOIN utenti u ON u.id=s.utente_id
            LEFT JOIN veicoli v ON v.id=s.veicolo_id
            WHERE s.stato='approvata'"""
    sp = []
    if uid_f: sq += " AND s.utente_id=?"; sp.append(int(uid_f))
    if da:    sq += " AND s.data>=?"; sp.append(da)
    if a_:    sq += " AND s.data<=?"; sp.append(a_)
    sq += " ORDER BY s.data, u.cognome"
    spese = db.execute(sq, sp).fetchall()
    db.close()

    wb = openpyxl.Workbook()

    # ── Foglio 1: Presenze ──────────────────────────────────
    ws = wb.active; ws.title = "Presenze"
    hdr_fill = PatternFill("solid", fgColor="0F172A")
    hdr_font = Font(color="FFFFFF", bold=True)
    headers = ['Dipendente','Data','Cantiere','Entrata','Uscita','Ore','Note']
    for i,h in enumerate(headers,1):
        c = ws.cell(1,i,h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20; ws.column_dimensions['G'].width = 30

    for row_i, r in enumerate(rows, 2):
        ws.cell(row_i,1,f"{r['nome']} {r['cognome']}")
        ws.cell(row_i,2,r['data']); ws.cell(row_i,3,r['cantiere_nome'] or '')
        ws.cell(row_i,4,r['ora_entrata'] or ''); ws.cell(row_i,5,r['ora_uscita'] or '')
        ws.cell(row_i,6, round(r['ore_totali'],1) if r['ore_totali'] else '')
        ws.cell(row_i,7,r['note'] or '')

    ws.cell(len(rows)+2, 5, 'TOTALE').font = Font(bold=True)
    ws.cell(len(rows)+2, 6, f'=SUM(F2:F{len(rows)+1})').font = Font(bold=True)

    # ── Foglio 2: Spese Rimborsabili ───────────────────────
    ws2 = wb.create_sheet("Spese Rimborsabili")
    grn_fill = PatternFill("solid", fgColor="065F46")
    headers2 = ['Dipendente','Data','Categoria','Descrizione','Veicolo','Importo €','Stato']
    for i,h in enumerate(headers2,1):
        c = ws2.cell(1,i,h); c.font = hdr_font; c.fill = grn_fill; c.alignment = Alignment(horizontal='center')
    ws2.column_dimensions['A'].width = 22; ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 18; ws2.column_dimensions['D'].width = 28
    ws2.column_dimensions['E'].width = 14; ws2.column_dimensions['F'].width = 12

    for row_i, s in enumerate(spese, 2):
        ws2.cell(row_i,1,f"{s['nome']} {s['cognome']}")
        ws2.cell(row_i,2,s['data'])
        ws2.cell(row_i,3,s['categoria'])
        ws2.cell(row_i,4,s['descrizione'] or '')
        ws2.cell(row_i,5,s['veicolo_targa'] or '')
        amt_cell = ws2.cell(row_i,6,round(s['importo'],2) if s['importo'] else 0)
        amt_cell.number_format = '#,##0.00 €'
        ws2.cell(row_i,7,'Approvata')

    if spese:
        tot_row = len(spese)+2
        ws2.cell(tot_row,5,'TOTALE RIMBORSI').font = Font(bold=True)
        tc = ws2.cell(tot_row,6,f'=SUM(F2:F{len(spese)+1})')
        tc.font = Font(bold=True); tc.number_format = '#,##0.00 €'
        tc.fill = PatternFill("solid", fgColor="D1FAE5")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"presenze_{date.today().isoformat()}.xlsx"
    return Response(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})

# ══════════════════════════════════════════════════════════
#  FERIE & PERMESSI
# ══════════════════════════════════════════════════════════
FERIE_TMPL = """
{% if session.ruolo != 'admin' %}
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-paper-plane" style="color:#7c3aed;margin-right:8px"></i>Nuova richiesta</h3></div>
  <div class="card-body">
    <form method="POST" action="/ferie/richiesta">
      <div class="form-row-4">
        <div class="form-group"><label>Tipo *</label>
          <select name="tipo" required>
            <option value="Ferie">🌴 Ferie</option>
            <option value="Permesso">⏰ Permesso</option>
            <option value="Malattia">🏥 Malattia</option>
            <option value="Permesso studio">📚 Permesso studio</option>
          </select></div>
        <div class="form-group"><label>Dal *</label><input type="date" name="data_inizio" required></div>
        <div class="form-group"><label>Al *</label><input type="date" name="data_fine" required></div>
        <div class="form-group"><label>Motivo</label><input name="motivo" placeholder="Opzionale"></div>
      </div>
      <button type="submit" class="btn btn-blue"><i class="fa fa-paper-plane"></i> Invia richiesta</button>
    </form>
  </div>
</div>
{% endif %}

{% if session.ruolo=='admin' %}
<div class="stats-grid" style="grid-template-columns:repeat(4,1fr)">
  <div class="stat-card"><div class="stat-icon amber"><i class="fa fa-clock"></i></div><div><div class="stat-val">{{ stats.attesa }}</div><div class="stat-lbl">In attesa</div></div></div>
  <div class="stat-card"><div class="stat-icon green"><i class="fa fa-check"></i></div><div><div class="stat-val">{{ stats.approvate }}</div><div class="stat-lbl">Approvate</div></div></div>
  <div class="stat-card"><div class="stat-icon red"><i class="fa fa-times"></i></div><div><div class="stat-val">{{ stats.rifiutate }}</div><div class="stat-lbl">Rifiutate</div></div></div>
  <div class="stat-card"><div class="stat-icon purple"><i class="fa fa-calendar-check"></i></div><div><div class="stat-val">{{ stats.giorni_mese }}</div><div class="stat-lbl">Giorni off questo mese</div></div></div>
</div>

{% if richieste_attesa %}
<div class="card" style="margin-bottom:20px;border-left:4px solid var(--warning)">
  <div class="card-header"><h3><i class="fa fa-bell" style="color:var(--warning);margin-right:8px"></i>Da approvare</h3></div>
  <div class="table-wrap"><table>
    <thead><tr><th>Dipendente</th><th>Tipo</th><th>Dal</th><th>Al</th><th>Giorni</th><th>Motivo</th><th>Azioni</th></tr></thead>
    <tbody>{% for r in richieste_attesa %}
    <tr style="background:#fffbf0">
      <td><span class="avatar-sm">{{ r.nome[0] }}{{ r.cognome[0] }}</span><strong>{{ r.nome }} {{ r.cognome }}</strong></td>
      <td><span class="badge badge-purple">{{ r.tipo }}</span></td>
      <td>{{ r.data_inizio }}</td><td>{{ r.data_fine }}</td>
      <td><strong>{{ r.giorni }}</strong></td>
      <td style="color:var(--text-light)">{{ r.motivo or '–' }}</td>
      <td>
        <form method="POST" action="/ferie/{{ r.id }}/gestisci" style="display:flex;gap:6px;align-items:center">
          <input name="nota_admin" placeholder="Nota..." style="padding:5px 8px;font-size:12px;width:120px;border:1px solid var(--border);border-radius:6px">
          <button name="azione" value="approva" class="btn btn-sm" style="background:#dcfce7;color:#16a34a;border:1px solid #bbf7d0"><i class="fa fa-check"></i> Approva</button>
          <button name="azione" value="rifiuta" class="btn btn-sm btn-danger"><i class="fa fa-times"></i> Rifiuta</button>
        </form>
      </td>
    </tr>{% endfor %}
    </tbody>
  </table></div>
</div>{% endif %}
{% endif %}

<div class="card">
  <div class="card-header"><h3>Storico richieste</h3>
    <div style="display:flex;gap:8px">
      {% if session.ruolo=='admin' %}
      <select onchange="window.location='/ferie?uid='+this.value" style="font-size:12px;padding:4px 8px;border:1px solid var(--border);border-radius:6px">
        <option value="">Tutti i dipendenti</option>
        {% for u in tutti_utenti %}<option value="{{ u.id }}" {{ 'selected' if filtro_uid==u.id|string }}>{{ u.nome }} {{ u.cognome }}</option>{% endfor %}
      </select>
      {% endif %}
    </div>
  </div>
  <div class="table-wrap"><table>
    <thead><tr>{% if session.ruolo=='admin' %}<th>Dipendente</th>{% endif %}<th>Tipo</th><th>Dal</th><th>Al</th><th>Giorni</th><th>Motivo</th><th>Stato</th><th>Risposta admin</th></tr></thead>
    <tbody>{% for r in storico %}
    <tr>
      {% if session.ruolo=='admin' %}<td><span class="avatar-sm">{{ r.nome[0] }}{{ r.cognome[0] }}</span>{{ r.nome }} {{ r.cognome }}</td>{% endif %}
      <td><span class="badge badge-purple">{{ r.tipo }}</span></td>
      <td>{{ r.data_inizio }}</td><td>{{ r.data_fine }}</td><td>{{ r.giorni }}</td>
      <td style="color:var(--text-light)">{{ r.motivo or '–' }}</td>
      <td>{% if r.stato=='in_attesa' %}<span class="badge badge-amber">⏳ In attesa</span>
          {% elif r.stato=='approvata' %}<span class="badge badge-green">✅ Approvata</span>
          {% else %}<span class="badge badge-red">❌ Rifiutata</span>{% endif %}</td>
      <td style="font-size:12px;color:var(--text-light)">{{ r.nota_admin or '–' }}</td>
    </tr>{% else %}
    <tr><td colspan="8"><div class="empty-state"><i class="fa fa-umbrella-beach"></i><p>Nessuna richiesta</p></div></td></tr>
    {% endfor %}</tbody>
  </table></div>
</div>"""

@app.route('/ferie')
@login_required
def ferie():
    db  = get_db()
    uid = session['user_id']
    filtro_uid = request.args.get('uid','')
    mese = date.today().strftime('%Y-%m')

    base_q = "SELECT f.*,u.nome,u.cognome FROM ferie_permessi f JOIN utenti u ON u.id=f.utente_id WHERE 1=1"
    if session['ruolo'] != 'admin':
        richieste_attesa = []
        storico = db.execute(base_q + " AND f.utente_id=? ORDER BY f.data_inizio DESC",(uid,)).fetchall()
        stats   = {}
    else:
        richieste_attesa = db.execute(base_q + " AND f.stato='in_attesa' ORDER BY f.creato_il").fetchall()
        q2 = base_q + (" AND f.utente_id=?" if filtro_uid else "") + " AND f.stato!='in_attesa' ORDER BY f.data_inizio DESC LIMIT 100"
        storico = db.execute(q2, [int(filtro_uid)] if filtro_uid else []).fetchall()
        stats = {
            'attesa':    db.execute("SELECT COUNT(*) FROM ferie_permessi WHERE stato='in_attesa'").fetchone()[0],
            'approvate': db.execute("SELECT COUNT(*) FROM ferie_permessi WHERE stato='approvata'").fetchone()[0],
            'rifiutate': db.execute("SELECT COUNT(*) FROM ferie_permessi WHERE stato='rifiutata'").fetchone()[0],
            'giorni_mese': db.execute("SELECT COALESCE(SUM(giorni),0) FROM ferie_permessi WHERE stato='approvata' AND data_inizio LIKE ?",(f'{mese}%',)).fetchone()[0],
        }

    tutti_utenti = db.execute("SELECT id,nome,cognome FROM utenti WHERE attivo=1 ORDER BY cognome").fetchall() if session['ruolo']=='admin' else []
    db.close()
    return render_page(FERIE_TMPL, page_title='Ferie & Permessi', active='ferie',
        richieste_attesa=richieste_attesa, storico=storico, stats=stats,
        tutti_utenti=tutti_utenti, filtro_uid=filtro_uid)

@app.route('/ferie/richiesta', methods=['POST'])
@login_required
def ferie_richiesta():
    tipo = request.form.get('tipo','Ferie')
    d_in = request.form.get('data_inizio',''); d_fi = request.form.get('data_fine','')
    motivo = request.form.get('motivo','')
    if not d_in or not d_fi: flash('Date obbligatorie.','error'); return redirect(url_for('ferie'))
    try:
        diff = datetime.strptime(d_fi,'%Y-%m-%d') - datetime.strptime(d_in,'%Y-%m-%d')
        giorni = diff.days + 1
        if giorni <= 0: raise ValueError
    except: flash("La data fine deve essere dopo la data inizio.",'error'); return redirect(url_for('ferie'))
    db = get_db()
    db.execute("INSERT INTO ferie_permessi (utente_id,tipo,data_inizio,data_fine,giorni,motivo) VALUES (?,?,?,?,?,?)",
               (session['user_id'],tipo,d_in,d_fi,giorni,motivo))
    safe_commit(db)
    email_admin = get_setting('email_notifiche','')
    if email_admin:
        send_email(email_admin, f'[ACCESSO FIERE] Richiesta {tipo} da {session["nome"]} {session["cognome"]}',
            f'<p><b>{session["nome"]} {session["cognome"]}</b> ha richiesto <b>{tipo}</b> dal {d_in} al {d_fi} ({giorni} giorni).</p>')
    db.close(); flash('Richiesta inviata!','success'); return redirect(url_for('ferie'))

@app.route('/ferie/<int:fid>/gestisci', methods=['POST'])
@admin_required
def ferie_gestisci(fid):
    azione = request.form.get('azione'); nota = request.form.get('nota_admin','')
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    stato = 'approvata' if azione == 'approva' else 'rifiutata'
    db = get_db()
    db.execute("UPDATE ferie_permessi SET stato=?,nota_admin=?,gestito_il=? WHERE id=?",
               (stato, nota or stato.capitalize(), now_str, fid))
    safe_commit(db); db.close()
    flash(f'{"✅ Approvata" if stato=="approvata" else "❌ Rifiutata"}!','success')
    return redirect(url_for('ferie'))

# ══════════════════════════════════════════════════════════
#  RICHIESTE PRESENZE (admin panel)
# ══════════════════════════════════════════════════════════
RIC_TMPL = """
{% if pending > 0 %}<div class="alert alert-info"><i class="fa fa-bell"></i> {{ pending }} richieste presenze in attesa</div>{% endif %}

{% if richieste_attesa %}
<div class="card" style="margin-bottom:20px;border-left:4px solid var(--warning)">
  <div class="card-header"><h3><i class="fa fa-clock" style="color:var(--warning);margin-right:8px"></i>Presenze da approvare ({{ pending }})</h3></div>
  <div class="table-wrap"><table>
    <thead><tr><th>Dipendente</th><th>Data</th><th>Fiera</th><th>Ore</th><th>Note</th><th>Azioni</th></tr></thead>
    <tbody>{% for r in richieste_attesa %}
    <tr style="background:#fffbf0">
      <td><span class="avatar-sm">{{ r.nome[0] }}{{ r.cognome[0] }}</span><strong>{{ r.nome }} {{ r.cognome }}</strong></td>
      <td style="font-family:monospace">{{ r.data }}</td>
      <td>{% if r.cantiere_nome %}<span class="tag">{{ r.cantiere_nome }}</span>{% else %}<span style="color:var(--text-light)">–</span>{% endif %}</td>
      <td><strong style="font-size:15px">{{ "%.1f"|format(r.ore_totali) if r.ore_totali else '–' }}h</strong></td>
      <td style="color:var(--text-light);font-size:12px;max-width:160px">{{ r.note or '–' }}</td>
      <td>
        <div style="display:flex;gap:6px;align-items:center;flex-wrap:wrap">
          <button onclick="apriModifica({{ r.id }},'{{ r.data }}',{{ r.ore_totali or 0 }},'{{ r.cantiere_id or '' }}','{{ r.note or '' }}')"
            class="btn btn-sm btn-secondary"><i class="fa fa-pen"></i> Modifica</button>
          <form method="POST" action="/admin/richieste/{{ r.id }}/gestisci" style="display:flex;gap:6px">
            <input name="nota_admin" placeholder="Nota..." style="padding:5px 8px;font-size:12px;width:100px;border:1px solid var(--border);border-radius:6px">
            <button name="azione" value="approva" class="btn btn-sm" style="background:#dcfce7;color:#16a34a;border:1px solid #bbf7d0" title="Approva"><i class="fa fa-check"></i></button>
            <button name="azione" value="rifiuta" class="btn btn-sm btn-danger" title="Rifiuta"><i class="fa fa-times"></i></button>
          </form>
        </div>
      </td>
    </tr>{% endfor %}</tbody>
  </table></div>
</div>{% endif %}

<div class="card">
  <div class="card-header"><h3>Storico richieste presenze</h3></div>
  <div class="table-wrap"><table>
    <thead><tr><th>Dipendente</th><th>Data</th><th>Fiera</th><th>Ore</th><th>Stato</th><th>Nota admin</th></tr></thead>
    <tbody>{% for r in storico %}
    <tr>
      <td><span class="avatar-sm">{{ r.nome[0] }}{{ r.cognome[0] }}</span>{{ r.nome }} {{ r.cognome }}</td>
      <td style="font-family:monospace">{{ r.data }}</td>
      <td>{% if r.cantiere_nome %}<span class="tag">{{ r.cantiere_nome }}</span>{% else %}–{% endif %}</td>
      <td><strong>{{ "%.1f"|format(r.ore_totali) if r.ore_totali else '–' }}h</strong></td>
      <td>{% if r.stato=='in_attesa' %}<span class="badge badge-amber">⏳ In attesa</span>
          {% elif r.stato=='approvata' %}<span class="badge badge-green">✅ Approvata</span>
          {% else %}<span class="badge badge-red">❌ Rifiutata</span>{% endif %}</td>
      <td style="font-size:12px;color:var(--text-light)">{{ r.nota_admin or '–' }}</td>
    </tr>{% else %}
    <tr><td colspan="6"><div class="empty-state"><i class="fa fa-inbox"></i><p>Nessuna richiesta</p></div></td></tr>
    {% endfor %}</tbody>
  </table></div>
</div>

<!-- Modal modifica richiesta -->
<div id="modal-ric" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9999;align-items:center;justify-content:center">
  <div class="card" style="width:460px;max-width:95vw">
    <div class="card-header">
      <h3><i class="fa fa-pen" style="color:var(--accent2)"></i> Modifica e approva</h3>
      <button onclick="document.getElementById('modal-ric').style.display='none'" class="btn btn-secondary btn-sm"><i class="fa fa-times"></i></button>
    </div>
    <div class="card-body">
      <form method="POST" id="form-modifica-ric">
        <input type="hidden" name="azione" value="modifica_approva">
        <div class="form-row">
          <div class="form-group">
            <label>Data</label>
            <input type="date" name="data_mod" id="mod-data">
          </div>
          <div class="form-group">
            <label>Ore nette</label>
            <input type="number" name="ore_mod" id="mod-ore" min="0" max="24" step="0.5" placeholder="8">
          </div>
        </div>
        <div class="form-group">
          <label>Fiera</label>
          <select name="cantiere_mod" id="mod-cantiere">
            <option value="">— Nessuno —</option>
            {% for c in cantieri %}
            <option value="{{ c.id }}">{{ c.nome }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="form-group">
          <label>Nota admin</label>
          <input name="nota_admin" id="mod-nota" placeholder="Es. corretta da 8h a 7.5h">
        </div>
        <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:16px">
          <button type="button" onclick="document.getElementById('modal-ric').style.display='none'" class="btn btn-secondary">Annulla</button>
          <button type="submit" class="btn btn-primary"><i class="fa fa-check"></i> Modifica e approva</button>
        </div>
      </form>
    </div>
  </div>
</div>
<script>
function apriModifica(id, data, ore, cantiere_id, note) {
  document.getElementById('form-modifica-ric').action = '/admin/richieste/' + id + '/gestisci';
  document.getElementById('mod-data').value = data;
  document.getElementById('mod-ore').value = ore;
  document.getElementById('mod-cantiere').value = cantiere_id;
  document.getElementById('mod-nota').value = '';
  document.getElementById('modal-ric').style.display = 'flex';
}
</script>
"""

@app.route('/admin/richieste')
@admin_required
def admin_richieste():
    db = get_db()
    attesa = db.execute("""SELECT r.*,u.nome,u.cognome,c.nome as cantiere_nome FROM richieste_presenze r
        JOIN utenti u ON u.id=r.utente_id LEFT JOIN cantieri c ON c.id=r.cantiere_id
        WHERE r.stato='in_attesa' ORDER BY r.data DESC""").fetchall()
    storico = db.execute("""SELECT r.*,u.nome,u.cognome,c.nome as cantiere_nome FROM richieste_presenze r
        JOIN utenti u ON u.id=r.utente_id LEFT JOIN cantieri c ON c.id=r.cantiere_id
        WHERE r.stato!='in_attesa' ORDER BY r.gestito_il DESC LIMIT 50""").fetchall()
    cantieri = db.execute("SELECT id, nome FROM cantieri WHERE attivo=1 ORDER BY nome").fetchall()
    db.close()
    return render_page(RIC_TMPL, page_title='Richieste Presenze', active='richieste',
        richieste_attesa=attesa, storico=storico, pending=len(attesa), cantieri=cantieri)

@app.route('/admin/richieste/<int:rid>/gestisci', methods=['POST'])
@admin_required
def gestisci_richiesta(rid):
    azione = request.form.get('azione')
    nota   = request.form.get('nota_admin', '')
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    db = get_db()
    r = db.execute("SELECT * FROM richieste_presenze WHERE id=?", (rid,)).fetchone()
    if not r:
        flash('Non trovata.', 'error'); db.close()
        return redirect(url_for('admin_richieste'))

    if azione == 'modifica_approva':
        # Leggi valori modificati
        data_mod     = request.form.get('data_mod') or r['data']
        ore_mod      = float(request.form.get('ore_mod') or r['ore_totali'] or 0)
        cantiere_mod = request.form.get('cantiere_mod') or r['cantiere_id']
        cantiere_mod = int(cantiere_mod) if cantiere_mod else None
        # Aggiorna la richiesta con i nuovi valori e approvala
        db.execute("""UPDATE richieste_presenze
                      SET data=?, ore_totali=?, cantiere_id=?, stato='approvata',
                          nota_admin=?, gestito_il=? WHERE id=?""",
                   (data_mod, ore_mod, cantiere_mod, nota or 'Modificata e approvata', now_str, rid))
        # Registra nelle presenze con i valori corretti
        note_p = f"Approvata (modificata da admin){': ' + nota if nota else ''}"
        ex = db.execute(
            "SELECT id FROM presenze WHERE utente_id=? AND data=? AND cantiere_id IS ?",
            (r['utente_id'], data_mod, cantiere_mod)).fetchone()
        if ex:
            db.execute("UPDATE presenze SET ore_totali=?, note=? WHERE id=?",
                       (ore_mod, note_p, ex['id']))
        else:
            db.execute("INSERT INTO presenze (utente_id,data,ore_totali,cantiere_id,note) VALUES (?,?,?,?,?)",
                       (r['utente_id'], data_mod, ore_mod, cantiere_mod, note_p))
        flash(f'✅ Modificata ({ore_mod}h) e approvata!', 'success')

    elif azione == 'approva':
        db.execute("UPDATE richieste_presenze SET stato='approvata', nota_admin=?, gestito_il=? WHERE id=?",
                   (nota or 'Approvata', now_str, rid))
        note_p = f"Approvata{': ' + nota if nota else ''}"
        # Controlla duplicato esatto (stesso giorno + stesso cantiere) per evitare doppioni
        ex = db.execute(
            "SELECT id FROM presenze WHERE utente_id=? AND data=? AND cantiere_id IS ?",
            (r['utente_id'], r['data'], r['cantiere_id'])).fetchone()
        if ex:
            db.execute("UPDATE presenze SET ore_totali=?, note=? WHERE id=?",
                       (r['ore_totali'], note_p, ex['id']))
        else:
            db.execute("INSERT INTO presenze (utente_id,data,ore_totali,cantiere_id,note) VALUES (?,?,?,?,?)",
                       (r['utente_id'], r['data'], r['ore_totali'], r['cantiere_id'], note_p))
        flash('✅ Approvata e presenza registrata!', 'success')

    else:  # rifiuta
        db.execute("UPDATE richieste_presenze SET stato='rifiutata', nota_admin=?, gestito_il=? WHERE id=?",
                   (nota or 'Rifiutata', now_str, rid))
        flash('❌ Rifiutata.', 'success')

    safe_commit(db); db.close()
    return redirect(url_for('admin_richieste'))

# ══════════════════════════════════════════════════════════
#  REPORT MENSILE
# ══════════════════════════════════════════════════════════
REPORT_TMPL = """
<!-- Filtri -->
<div class="card" style="margin-bottom:24px">
  <div class="card-header"><h3><i class="fa fa-sliders" style="color:var(--accent2);margin-right:8px"></i>Parametri Report</h3></div>
  <div class="card-body">
    <form method="GET" style="display:flex;gap:14px;align-items:flex-end;flex-wrap:wrap">
      <div class="form-group" style="margin:0"><label>Mese *</label>
        <input type="month" name="mese" value="{{ mese_sel }}" required style="width:160px"></div>
      <div class="form-group" style="margin:0;min-width:180px"><label>Vista per</label>
        <select name="vista" onchange="this.form.submit()">
          <option value="dipendente" {{ 'selected' if vista=='dipendente' }}>👤 Dipendente</option>
          <option value="cantiere"   {{ 'selected' if vista=='cantiere'   }}>🏗️ Cantiere</option>
        </select></div>
      {% if vista=='dipendente' %}
      <div class="form-group" style="margin:0;min-width:180px"><label>Dipendente</label>
        <select name="uid"><option value="">Tutti</option>
        {% for u in tutti_utenti %}<option value="{{ u.id }}" {{ 'selected' if uid_sel==u.id|string }}>{{ u.nome }} {{ u.cognome }}</option>{% endfor %}
        </select></div>
      {% else %}
      <div class="form-group" style="margin:0;min-width:180px"><label>Fiera</label>
        <select name="cid"><option value="">Tutti</option>
        {% for c in tutti_cantieri %}<option value="{{ c.id }}" {{ 'selected' if cid_sel==c.id|string }}>{{ c.nome }}</option>{% endfor %}
        </select></div>
      {% endif %}
      <button type="submit" class="btn btn-primary" style="height:40px"><i class="fa fa-search"></i> Genera</button>
      {% if report_data %}
      <a href="/admin/report/export?mese={{ mese_sel }}&uid={{ uid_sel }}&cid={{ cid_sel }}&vista={{ vista }}" class="btn btn-green" style="height:40px">
        <i class="fa fa-file-excel"></i> Scarica Excel
      </a>
      {% endif %}
    </form>
  </div>
</div>

{% if report_data %}
<!-- KPI -->
<div class="stats-grid" style="grid-template-columns:repeat(5,1fr);margin-bottom:24px">
  <div class="stat-card"><div class="stat-icon blue"><i class="fa fa-{{ 'users' if vista=='dipendente' else 'hard-hat' }}"></i></div><div><div class="stat-val">{{ report_data|length }}</div><div class="stat-lbl">{{ 'Dipendenti' if vista=='dipendente' else 'Cantieri' }}</div></div></div>
  <div class="stat-card"><div class="stat-icon green"><i class="fa fa-calendar-check"></i></div><div><div class="stat-val">{{ totali.giorni }}</div><div class="stat-lbl">Giorni totali</div></div></div>
  <div class="stat-card"><div class="stat-icon amber"><i class="fa fa-clock"></i></div><div><div class="stat-val">{{ "%.0f"|format(totali.ore) }}h</div><div class="stat-lbl">Ore totali</div></div></div>
  <div class="stat-card"><div class="stat-icon red"><i class="fa fa-clock-rotate-left"></i></div><div><div class="stat-val">{{ "%.1f"|format(totali.ore/totali.giorni) if totali.giorni else '0' }}h</div><div class="stat-lbl">Media ore/giorno</div></div></div>
  <div class="stat-card" style="border-left:3px solid #10b981"><div class="stat-icon green"><i class="fa fa-receipt"></i></div><div><div class="stat-val">€ {{ "%.0f"|format(totali.spese) }}</div><div class="stat-lbl">Rimborsi approvati</div></div></div>
</div>

<!-- Grafici -->
<div class="grid-2" style="margin-bottom:24px">
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-chart-bar" style="color:var(--accent2);margin-right:8px"></i>
      Ore lavorate per {{ 'dipendente' if vista=='dipendente' else 'cantiere' }}
    </h3></div>
    <div class="card-body"><canvas id="chartOreReport" height="260"></canvas></div>
  </div>
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-chart-pie" style="color:var(--warning);margin-right:8px"></i>
      Distribuzione ore {{ mese_sel }}
    </h3></div>
    <div class="card-body"><canvas id="chartPieReport" height="260"></canvas></div>
  </div>
</div>

<!-- Tabella -->
<div class="card">
  <div class="card-header">
    <h3>Dettaglio {{ mese_sel }}</h3>
    <span style="font-size:12px;color:var(--text-light)">Vista per {{ 'dipendente' if vista=='dipendente' else 'cantiere' }}</span>
  </div>
  <div class="table-wrap"><table>
    <thead>
    {% if vista=='dipendente' %}
    <tr><th>Dipendente</th><th>Giorni lavorati</th><th>Ore totali</th><th>Media/giorno</th><th>Ferie/Permessi</th><th>Cantiere principale</th><th>Straordinari est.</th><th style="color:#10b981">Rimborsi €</th><th style="color:#dc2626">Detrazioni €</th></tr>
    {% else %}
    <tr><th>Fiera</th><th>Giorni totali</th><th>Ore totali</th><th>Media/giorno</th><th>Dipendenti distinti</th><th>Dipendente top</th></tr>
    {% endif %}
    </thead>
    <tbody>{% for r in report_data %}
    {% if vista=='dipendente' %}
    <tr style="{% if r.is_jolly %}background:#fffbeb{% endif %}">
      <td>
        <span class="avatar-sm" style="{% if r.is_jolly %}background:linear-gradient(135deg,#d97706,#f59e0b){% endif %}">{{ r.nome[0] }}{{ r.cognome[0] }}</span>
        <strong>{{ r.nome }} {{ r.cognome }}</strong>
        {% if r.is_jolly %}<span style="font-size:10px;background:#fef3c7;color:#92400e;border-radius:4px;padding:1px 6px;margin-left:4px;font-weight:700">👷 JOLLY</span>{% endif %}
      </td>
      <td><strong>{{ r.giorni }}</strong></td>
      <td><strong style="color:var(--accent2)">{{ "%.1f"|format(r.ore) }}h</strong></td>
      <td>{{ "%.1f"|format(r.ore/r.giorni) if r.giorni else '–' }}h</td>
      <td>{% if r.ferie %}<span class="badge badge-purple">{{ r.ferie }} gg</span>{% else %}<span style="color:var(--text-light)">–</span>{% endif %}</td>
      <td>{% if r.cantiere %}<span class="tag"><i class="fa fa-store"></i> {{ r.cantiere }}</span>{% else %}<span style="color:var(--text-light)">–</span>{% endif %}</td>
      <td>{% set std = r.ore - (r.giorni * 8) %}
        {% if std > 0 %}<span class="badge badge-red">+{{ "%.1f"|format(std) }}h</span>
        {% else %}<span class="badge badge-green">In orario</span>{% endif %}</td>
      <td style="text-align:right">
        {% if r.spese > 0 and r.uid %}
        <a href="/admin/spese?uid={{ r.uid }}&stato=approvata" style="color:#059669;font-weight:800;text-decoration:none">€ {{ "%.2f"|format(r.spese) }}</a>
        {% else %}<span style="color:var(--text-light)">–</span>{% endif %}
      </td>
      <td style="text-align:right">
        {% if not r.is_jolly and r.detrazioni > 0 and r.uid %}
        <a href="/admin/detrazioni?uid={{ r.uid }}&mese={{ mese_sel }}"
           style="color:#dc2626;font-weight:800;text-decoration:none"
           title="{% for x in r.detr_dettaglio %}{{ x.descrizione }}: €{{ '%.2f'|format(x.importo) }}&#10;{% endfor %}">
          − € {{ "%.2f"|format(r.detrazioni) }}
        </a>
        {% else %}<span style="color:var(--text-light)">–</span>{% endif %}
      </td>
    </tr>
    {% else %}
    <tr>
      <td><i class="fa fa-hard-hat" style="color:var(--warning);margin-right:8px"></i><strong>{{ r.nome }}</strong></td>
      <td><strong>{{ r.giorni }}</strong></td>
      <td><strong style="color:var(--accent2)">{{ "%.1f"|format(r.ore) }}h</strong></td>
      <td>{{ "%.1f"|format(r.ore/r.giorni) if r.giorni else '–' }}h</td>
      <td><span class="badge badge-blue">{{ r.n_dipendenti }} dip.</span></td>
      <td>{% if r.top_dip %}<span class="avatar-sm" style="font-size:9px">{{ r.top_dip[:2] }}</span>{{ r.top_dip }}{% else %}–{% endif %}</td>
    </tr>
    {% endif %}
    {% endfor %}</tbody>
  </table></div>
</div>

{% if vista=='dipendente' and spese_report %}
<div class="card" style="margin-top:20px;border-left:4px solid #10b981">
  <div class="card-header" style="display:flex;justify-content:space-between;align-items:center">
    <h3><i class="fa fa-receipt" style="color:#10b981"></i> Spese Rimborsabili Approvate — {{ mese_sel }}</h3>
    <span style="background:#f0fdf4;color:#15803d;border:1px solid #86efac;border-radius:8px;padding:5px 16px;font-weight:700">
      Totale: € {{ "%.2f"|format(totali.spese) }}
    </span>
  </div>
  <div class="table-wrap"><table>
    <thead><tr>
      {% if not uid_sel %}<th>Dipendente</th>{% endif %}
      <th>Data</th><th>Categoria</th><th>Descrizione</th><th>Veicolo</th>
      <th style="text-align:right">Importo</th><th>Scontrino</th>
    </tr></thead>
    <tbody>
    {% for s in spese_report %}
    <tr>
      {% if not uid_sel %}<td><strong>{{ s.nome }} {{ s.cognome }}</strong></td>{% endif %}
      <td style="font-family:monospace">{{ s.data }}</td>
      <td><span class="tag">{{ s.categoria }}</span></td>
      <td style="color:var(--text-light);font-size:13px">{{ s.descrizione or '–' }}</td>
      <td style="font-size:12px">{{ s.veicolo_targa or '–' }}</td>
      <td style="text-align:right;font-weight:800;color:#059669">€ {{ "%.2f"|format(s.importo) }}</td>
      <td>
        {% if s.foto_nome %}
        <a href="/admin/spese/foto/{{ s.foto_nome }}" target="_blank"
           style="background:#eff6ff;color:#2563eb;border:1px solid #bfdbfe;border-radius:6px;padding:4px 10px;font-size:12px;text-decoration:none;display:inline-flex;align-items:center;gap:4px">
          <i class="fa fa-image"></i> Vedi
        </a>
        {% else %}–{% endif %}
      </td>
    </tr>
    {% endfor %}
    <tr style="background:#f0fdf4;font-weight:700">
      <td colspan="{{ 4 if uid_sel else 5 }}" style="text-align:right">TOTALE RIMBORSI</td>
      <td style="text-align:right;font-size:15px;color:#059669">€ {{ "%.2f"|format(totali.spese) }}</td>
      <td></td>
    </tr>
    </tbody>
  </table></div>
</div>
{% endif %}

{% if vista=='dipendente' and totali.detrazioni > 0 %}
<div class="card" style="margin-top:20px;border-left:4px solid #dc2626">
  <div class="card-header" style="display:flex;justify-content:space-between;align-items:center">
    <h3><i class="fa fa-minus-circle" style="color:#dc2626"></i> Detrazioni & Anticipi — {{ mese_sel }}</h3>
    <span style="background:#fef2f2;color:#dc2626;border:1px solid #fecaca;border-radius:8px;padding:5px 16px;font-weight:700">
      Totale: − € {{ "%.2f"|format(totali.detrazioni) }}
    </span>
  </div>
  <div class="table-wrap"><table>
    <thead><tr>
      {% if not uid_sel %}<th>Dipendente</th>{% endif %}
      <th>Tipo</th><th>Descrizione</th><th style="text-align:right">Importo</th>
    </tr></thead>
    <tbody>
    {% for r in report_data %}{% if not r.is_jolly and r.detrazioni > 0 %}
      {% for x in r.detr_dettaglio %}
      <tr>
        {% if not uid_sel %}<td><strong>{{ r.nome }} {{ r.cognome }}</strong></td>{% endif %}
        <td><span style="background:#fef2f2;color:#dc2626;padding:2px 8px;border-radius:99px;font-size:12px;font-weight:600">{{ x.tipo }}</span></td>
        <td style="color:var(--text-light);font-size:13px">{{ x.descrizione or '–' }}</td>
        <td style="text-align:right;font-weight:800;color:#dc2626">− € {{ "%.2f"|format(x.importo) }}</td>
      </tr>
      {% endfor %}
    {% endif %}{% endfor %}
    <tr style="background:#fef2f2;font-weight:700">
      <td colspan="{{ 2 if uid_sel else 3 }}" style="text-align:right">TOTALE DETRAZIONI</td>
      <td style="text-align:right;font-size:15px;color:#dc2626">− € {{ "%.2f"|format(totali.detrazioni) }}</td>
    </tr>
    </tbody>
  </table></div>
</div>
{% endif %}

<script>
const COLORS = ['#e63946','#2196F3','#22c55e','#f59e0b','#8b5cf6','#ec4899','#06b6d4','#84cc16','#f97316','#6366f1'];
const labels  = {{ report_data | map(attribute='label') | list | tojson }};
const oreData = {{ report_data | map(attribute='ore') | list | tojson }};

// Bar chart
new Chart(document.getElementById('chartOreReport'), {
  type: 'bar',
  data: {
    labels: labels,
    datasets: [{
      label: 'Ore lavorate',
      data: oreData,
      backgroundColor: COLORS.map(c => c + '30'),
      borderColor: COLORS,
      borderWidth: 2,
      borderRadius: 8,
    }]
  },
  options: {
    responsive: true,
    plugins: { legend: { display: false } },
    scales: {
      y: { beginAtZero: true, grid: { color: '#f1f5f9' }, ticks: { callback: v => v + 'h' } },
      x: { grid: { display: false } }
    }
  }
});

// Pie chart
new Chart(document.getElementById('chartPieReport'), {
  type: 'doughnut',
  data: {
    labels: labels,
    datasets: [{
      data: oreData,
      backgroundColor: COLORS,
      borderColor: '#fff',
      borderWidth: 3,
      hoverOffset: 8
    }]
  },
  options: {
    responsive: true,
    cutout: '55%',
    plugins: {
      legend: { position: 'bottom', labels: { padding: 14, font: { size: 12 } } },
      tooltip: { callbacks: { label: ctx => ` ${ctx.label}: ${ctx.raw.toFixed(1)}h` } }
    }
  }
});
</script>
{% endif %}"""

@app.route('/admin/report')
@admin_required
def admin_report():
    mese_sel = request.args.get('mese', date.today().strftime('%Y-%m'))
    uid_sel  = request.args.get('uid', '')
    cid_sel  = request.args.get('cid', '')
    vista    = request.args.get('vista', 'dipendente')

    db = get_db()
    tutti_utenti   = db.execute("SELECT id,nome,cognome FROM utenti WHERE attivo=1 AND ruolo='dipendente' ORDER BY cognome").fetchall()
    tutti_cantieri = db.execute("SELECT id,nome FROM cantieri WHERE attivo=1 ORDER BY nome").fetchall()

    report_data = []
    totali = {'giorni': 0, 'ore': 0.0, 'ferie': 0, 'spese': 0.0, 'detrazioni': 0.0}

    if vista == 'dipendente':
        q_users = "SELECT * FROM utenti WHERE attivo=1 AND ruolo='dipendente'"
        if uid_sel: q_users += f" AND id={int(uid_sel)}"
        users = db.execute(q_users + " ORDER BY cognome").fetchall()
        for u in users:
            presenze = db.execute("SELECT * FROM presenze WHERE utente_id=? AND data LIKE ?",
                                  (u['id'], f'{mese_sel}%')).fetchall()
            giorni = len(presenze)
            ore    = sum(p['ore_totali'] or 0 for p in presenze)
            ferie  = db.execute(
                "SELECT COALESCE(SUM(giorni),0) FROM ferie_permessi WHERE utente_id=? AND stato='approvata' AND data_inizio LIKE ?",
                (u['id'], f'{mese_sel}%')).fetchone()[0]
            c_row  = db.execute(
                """SELECT c.nome FROM presenze p JOIN cantieri c ON c.id=p.cantiere_id
                   WHERE p.utente_id=? AND p.data LIKE ? GROUP BY p.cantiere_id ORDER BY COUNT(*) DESC LIMIT 1""",
                (u['id'], f'{mese_sel}%')).fetchone()
            try:
                spese_dip = db.execute(
                    "SELECT COALESCE(SUM(importo),0) FROM spese_rimborso WHERE utente_id=? AND stato='approvata' AND substr(data,1,7)=?",
                    (u['id'], mese_sel)).fetchone()[0] or 0.0
            except:
                spese_dip = 0.0
            try:
                detr_dip = db.execute(
                    "SELECT COALESCE(SUM(importo),0) FROM detrazioni_dipendente WHERE utente_id=? AND mese_competenza=?",
                    (u['id'], mese_sel)).fetchone()[0] or 0.0
                detr_dettaglio = db.execute(
                    "SELECT tipo,descrizione,importo FROM detrazioni_dipendente WHERE utente_id=? AND mese_competenza=? ORDER BY data",
                    (u['id'], mese_sel)).fetchall()
            except:
                detr_dip = 0.0; detr_dettaglio = []
            if giorni > 0 or ferie > 0 or spese_dip > 0 or detr_dip > 0:
                nome_completo = f"{u['nome']} {u['cognome']}"
                report_data.append({
                    'uid': u['id'],
                    'nome': u['nome'], 'cognome': u['cognome'],
                    'label': nome_completo,
                    'giorni': giorni, 'ore': round(ore, 1),
                    'ferie': ferie,
                    'cantiere': c_row['nome'] if c_row else None,
                    'spese': round(spese_dip, 2),
                    'detrazioni': round(detr_dip, 2),
                    'detr_dettaglio': [dict(x) for x in detr_dettaglio],
                    'is_jolly': False,
                })
                totali['giorni'] += giorni; totali['ore'] += ore
                totali['ferie'] += ferie; totali['spese'] += spese_dip
                totali['detrazioni'] += detr_dip

        # Jolly: raggruppa per nome+cognome nel mese
        if not uid_sel:
            jolly_rows = db.execute("""
                SELECT nome_jolly, cognome_jolly,
                       COUNT(DISTINCT data) as giorni,
                       COALESCE(SUM(ore_totali),0) as ore,
                       MAX(cantieri.nome) as cantiere
                FROM presenze
                LEFT JOIN cantieri ON cantieri.id=presenze.cantiere_id
                WHERE nome_jolly IS NOT NULL AND data LIKE ?
                GROUP BY nome_jolly, cognome_jolly
                ORDER BY cognome_jolly
            """, (f'{mese_sel}%',)).fetchall()
            for j in jolly_rows:
                if j['ore'] > 0 or j['giorni'] > 0:
                    report_data.append({
                        'uid': None,
                        'nome': j['nome_jolly'], 'cognome': j['cognome_jolly'],
                        'label': f"{j['nome_jolly']} {j['cognome_jolly']}",
                        'giorni': j['giorni'], 'ore': round(float(j['ore']), 1),
                        'ferie': 0, 'cantiere': j['cantiere'],
                        'spese': 0.0, 'detrazioni': 0.0, 'detr_dettaglio': [],
                        'is_jolly': True,
                    })
                    totali['giorni'] += j['giorni']; totali['ore'] += float(j['ore'])

    else:  # vista cantiere
        q_cant = "SELECT * FROM cantieri WHERE attivo=1"
        if cid_sel: q_cant += f" AND id={int(cid_sel)}"
        cantieri_list = db.execute(q_cant + " ORDER BY nome").fetchall()
        for c in cantieri_list:
            rows = db.execute(
                """SELECT p.*,
                       COALESCE(p.nome_jolly, u.nome) as nome,
                       COALESCE(p.cognome_jolly, u.cognome) as cognome
                   FROM presenze p
                   LEFT JOIN utenti u ON u.id=p.utente_id
                   WHERE p.cantiere_id=? AND p.data LIKE ?""",
                (c['id'], f'{mese_sel}%')).fetchall()
            giorni = len(rows)
            ore    = sum(r['ore_totali'] or 0 for r in rows)
            dips   = len(set(r['utente_id'] for r in rows if r['utente_id']))
            # dipendente con più ore
            dip_ore = {}
            for r in rows:
                k = f"{r['nome']} {r['cognome']}"
                dip_ore[k] = dip_ore.get(k, 0) + (r['ore_totali'] or 0)
            top_dip = max(dip_ore, key=dip_ore.get) if dip_ore else None
            if giorni > 0:
                report_data.append({
                    'nome': c['nome'], 'label': c['nome'],
                    'giorni': giorni, 'ore': round(ore, 1),
                    'n_dipendenti': dips, 'top_dip': top_dip,
                    'ferie': 0
                })
                totali['giorni'] += giorni; totali['ore'] += ore

    db.close()
    totali['ore'] = round(totali['ore'], 1)
    totali['spese'] = round(totali['spese'], 2)

    # Dettaglio spese per la sezione tabella (solo vista dipendente)
    spese_report = []
    if vista == 'dipendente':
        db2 = get_db()
        sq = """SELECT s.*, u.nome, u.cognome, v.targa as veicolo_targa
                FROM spese_rimborso s JOIN utenti u ON u.id=s.utente_id
                LEFT JOIN veicoli v ON v.id=s.veicolo_id
                WHERE s.stato='approvata' AND substr(s.data,1,7)=?"""
        sp = [mese_sel]
        if uid_sel: sq += " AND s.utente_id=?"; sp.append(int(uid_sel))
        sq += " ORDER BY s.data, u.cognome"
        try:
            spese_report = [dict(r) for r in db2.execute(sq, sp).fetchall()]
        except:
            spese_report = []
        db2.close()

    return render_page(REPORT_TMPL, page_title='Report Mensile', active='report',
        tutti_utenti=tutti_utenti, tutti_cantieri=tutti_cantieri,
        mese_sel=mese_sel, uid_sel=uid_sel, cid_sel=cid_sel, vista=vista,
        report_data=report_data, totali=totali, spese_report=spese_report)


@app.route('/admin/report/export')
@admin_required
def export_report():
    if not EXCEL_OK:
        flash('Installa openpyxl: pip install openpyxl', 'error')
        return redirect(url_for('admin_report'))

    mese_sel = request.args.get('mese', date.today().strftime('%Y-%m'))
    uid_sel  = request.args.get('uid', '')
    cid_sel  = request.args.get('cid', '')
    vista    = request.args.get('vista', 'dipendente')

    db = get_db()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {mese_sel}"

    # Stili
    hdr_fill  = PatternFill("solid", fgColor="0F172A")
    hdr_font  = Font(color="FFFFFF", bold=True, size=11)
    tot_fill  = PatternFill("solid", fgColor="E2E8F0")
    tot_font  = Font(bold=True, size=11)
    center    = Alignment(horizontal='center', vertical='center')
    num_fmt   = '0.0'

    # Titolo
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"REPORT MENSILE – {mese_sel.upper()} – {'per Dipendente' if vista=='dipendente' else 'per Cantiere'}"
    title_cell.font  = Font(bold=True, size=14, color="0F172A")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    if vista == 'dipendente':
        headers = ['Dipendente','Giorni lavorati','Ore totali','Media ore/giorno','Ferie/Permessi gg','Cantiere principale','Straordinari est.','Paga €/h','Totale lordo €','Rimborsi €','Detrazioni €','Descrizione detrazioni','Totale da pagare €']
        col_widths = [25,16,14,18,18,22,16,12,16,14,14,35,18]
        q_users = "SELECT * FROM utenti WHERE attivo=1 AND ruolo='dipendente'"
        if uid_sel: q_users += f" AND id={int(uid_sel)}"
        users = db.execute(q_users + " ORDER BY cognome").fetchall()

        for i,h in enumerate(headers,1):
            c = ws.cell(2,i,h); c.font=hdr_font; c.fill=hdr_fill; c.alignment=center
            ws.column_dimensions[ws.cell(2,i).column_letter].width = col_widths[i-1]
        ws.row_dimensions[2].height = 22

        row_i = 3
        tot_giorni = tot_ore = tot_ferie = tot_spese_xl = tot_detr_xl = tot_lordo = tot_netto = 0
        for u in users:
            presenze = db.execute("SELECT * FROM presenze WHERE utente_id=? AND data LIKE ?",
                                  (u['id'],f'{mese_sel}%')).fetchall()
            giorni = len(presenze); ore = round(sum(p['ore_totali'] or 0 for p in presenze),1)
            ferie  = db.execute("SELECT COALESCE(SUM(giorni),0) FROM ferie_permessi WHERE utente_id=? AND stato='approvata' AND data_inizio LIKE ?",
                                (u['id'],f'{mese_sel}%')).fetchone()[0]
            c_row  = db.execute("SELECT c.nome FROM presenze p JOIN cantieri c ON c.id=p.cantiere_id WHERE p.utente_id=? AND p.data LIKE ? GROUP BY p.cantiere_id ORDER BY COUNT(*) DESC LIMIT 1",
                                (u['id'],f'{mese_sel}%')).fetchone()
            spese_xl = round(db.execute("SELECT COALESCE(SUM(importo),0) FROM spese_rimborso WHERE utente_id=? AND stato='approvata' AND substr(data,1,7)=?",
                             (u['id'],mese_sel)).fetchone()[0] or 0, 2)
            detr_rows = db.execute("SELECT tipo,descrizione,importo FROM detrazioni_dipendente WHERE utente_id=? AND mese_competenza=? ORDER BY data",
                                   (u['id'],mese_sel)).fetchall()
            detr_xl = round(sum(d['importo'] for d in detr_rows), 2)
            detr_desc = '; '.join(f"{d['descrizione'] or d['tipo']} (€{d['importo']:.2f})" for d in detr_rows) if detr_rows else ''

            # Calcolo paga
            paga_oraria = float(u['paga_oraria_netta'] or 0)
            paga_base   = float(u['paga_base_netta'] or 0)
            ore_soglia  = float(u['ore_soglia_mensile'] or 0)
            if paga_base > 0:
                ore_extra = max(0, ore - ore_soglia)
                totale_lordo = round(paga_base + ore_extra * paga_oraria, 2)
            else:
                totale_lordo = round(ore * paga_oraria, 2)
            totale_netto = round(totale_lordo + spese_xl - detr_xl, 2)

            if giorni > 0 or ferie > 0 or spese_xl > 0 or detr_xl > 0:
                straord = round(ore - (giorni * 8), 1)
                ws.cell(row_i,1,f"{u['nome']} {u['cognome']}")
                ws.cell(row_i,2,giorni).alignment = center
                c_ore = ws.cell(row_i,3,ore); c_ore.number_format=num_fmt; c_ore.alignment=center
                media = round(ore/giorni,1) if giorni else 0
                c_med = ws.cell(row_i,4,media); c_med.number_format=num_fmt; c_med.alignment=center
                ws.cell(row_i,5,ferie).alignment = center
                ws.cell(row_i,6,c_row['nome'] if c_row else '').alignment = center
                c_str = ws.cell(row_i,7,straord); c_str.number_format=num_fmt; c_str.alignment=center
                if straord > 0: c_str.font = Font(bold=True, color="DC2626")
                # Paga oraria
                c_ph = ws.cell(row_i,8, paga_oraria if paga_oraria else ''); c_ph.number_format='€ #,##0.00'; c_ph.alignment=center
                if not paga_oraria and not paga_base: c_ph.value = '⚠ n/d'; c_ph.font = Font(color="DC2626")
                # Totale lordo
                c_tl = ws.cell(row_i,9, totale_lordo if (paga_oraria or paga_base) else ''); c_tl.number_format='€ #,##0.00'; c_tl.alignment=center
                if paga_oraria or paga_base: c_tl.font = Font(bold=True, color="10B981")
                # Rimborsi
                c_sp = ws.cell(row_i,10, spese_xl if spese_xl else ''); c_sp.number_format='€ #,##0.00'; c_sp.alignment=center
                if spese_xl > 0: c_sp.font = Font(bold=True, color="059669")
                # Detrazioni
                c_dt = ws.cell(row_i,11, -detr_xl if detr_xl else ''); c_dt.number_format='€ #,##0.00'; c_dt.alignment=center
                if detr_xl > 0: c_dt.font = Font(bold=True, color="DC2626")
                # Descrizione detrazioni
                ws.cell(row_i,12, detr_desc)
                # Totale da pagare
                c_tn = ws.cell(row_i,13, totale_netto if (paga_oraria or paga_base) else '')
                c_tn.number_format='€ #,##0.00'; c_tn.alignment=center
                if paga_oraria or paga_base:
                    c_tn.font = Font(bold=True, size=11, color="059669" if totale_netto >= 0 else "DC2626")
                    c_tn.fill = PatternFill("solid", fgColor="F0FDF4" if totale_netto >= 0 else "FEF2F2")
                tot_giorni+=giorni; tot_ore+=ore; tot_ferie+=ferie
                tot_spese_xl+=spese_xl; tot_detr_xl+=detr_xl
                tot_lordo+=totale_lordo; tot_netto+=totale_netto
                row_i+=1

        # Jolly nel foglio Excel — solo se non è filtrato su un singolo dipendente
        if not uid_sel:
            jolly_rows = db.execute("""
                SELECT nome_jolly, cognome_jolly,
                       COUNT(DISTINCT data) as giorni,
                       COALESCE(SUM(ore_totali),0) as ore,
                       MAX(cantieri.nome) as cantiere
                FROM presenze
                LEFT JOIN cantieri ON cantieri.id=presenze.cantiere_id
                WHERE nome_jolly IS NOT NULL AND data LIKE ?
                GROUP BY nome_jolly, cognome_jolly ORDER BY cognome_jolly
            """, (f'{mese_sel}%',)).fetchall()
            jolly_fill = PatternFill("solid", fgColor="FEF3C7")
            for j in jolly_rows:
                giorni_j = j['giorni']; ore_j = round(float(j['ore']),1)
                if giorni_j == 0: continue
                straord_j = round(ore_j - (giorni_j * 8), 1)
                c1 = ws.cell(row_i,1,f"👷 {j['nome_jolly']} {j['cognome_jolly']} (Jolly)")
                c1.fill = jolly_fill
                ws.cell(row_i,2,giorni_j).alignment = center; ws.cell(row_i,2).fill = jolly_fill
                c_ore = ws.cell(row_i,3,ore_j); c_ore.number_format=num_fmt; c_ore.alignment=center; c_ore.fill=jolly_fill
                media_j = round(ore_j/giorni_j,1) if giorni_j else 0
                c_med = ws.cell(row_i,4,media_j); c_med.number_format=num_fmt; c_med.alignment=center; c_med.fill=jolly_fill
                ws.cell(row_i,5,0).alignment=center; ws.cell(row_i,5).fill=jolly_fill
                ws.cell(row_i,6,j['cantiere'] or '').alignment=center; ws.cell(row_i,6).fill=jolly_fill
                c_str = ws.cell(row_i,7,straord_j); c_str.number_format=num_fmt; c_str.alignment=center; c_str.fill=jolly_fill
                for col in [8,9,10,11,12,13]:
                    ws.cell(row_i,col,'–').fill=jolly_fill; ws.cell(row_i,col).alignment=center
                tot_giorni+=giorni_j; tot_ore+=ore_j; row_i+=1

        # Riga totali
        ws.cell(row_i,1,'TOTALE').font=tot_font; ws.cell(row_i,1).fill=tot_fill
        ws.cell(row_i,2,tot_giorni).font=tot_font; ws.cell(row_i,2).fill=tot_fill; ws.cell(row_i,2).alignment=center
        tc=ws.cell(row_i,3,round(tot_ore,1)); tc.font=tot_font; tc.fill=tot_fill; tc.number_format=num_fmt; tc.alignment=center
        ws.cell(row_i,5,tot_ferie).font=tot_font; ws.cell(row_i,5).fill=tot_fill; ws.cell(row_i,5).alignment=center
        c_tlt=ws.cell(row_i,9,round(tot_lordo,2)); c_tlt.font=Font(bold=True,color="10B981"); c_tlt.fill=tot_fill; c_tlt.number_format='€ #,##0.00'; c_tlt.alignment=center
        c_spt=ws.cell(row_i,10,round(tot_spese_xl,2)); c_spt.font=Font(bold=True,color="059669"); c_spt.fill=tot_fill; c_spt.number_format='€ #,##0.00'; c_spt.alignment=center
        c_dtt=ws.cell(row_i,11,-round(tot_detr_xl,2) if tot_detr_xl else ''); c_dtt.font=Font(bold=True,color="DC2626"); c_dtt.fill=tot_fill; c_dtt.number_format='€ #,##0.00'; c_dtt.alignment=center
        c_tnt=ws.cell(row_i,13,round(tot_netto,2)); c_tnt.font=Font(bold=True,size=12,color="059669"); c_tnt.fill=PatternFill("solid",fgColor="D1FAE5"); c_tnt.number_format='€ #,##0.00'; c_tnt.alignment=center

    else:  # cantiere
        headers = ['Cantiere','Giorni totali','Ore totali','Media ore/giorno','Dipendenti distinti','Dipendente top']
        col_widths = [25,16,14,18,20,25]
        q_cant = "SELECT * FROM cantieri WHERE attivo=1"
        if cid_sel: q_cant += f" AND id={int(cid_sel)}"
        cantieri_list = db.execute(q_cant+" ORDER BY nome").fetchall()

        for i,h in enumerate(headers,1):
            c=ws.cell(2,i,h); c.font=hdr_font; c.fill=hdr_fill; c.alignment=center
            ws.column_dimensions[ws.cell(2,i).column_letter].width=col_widths[i-1]
        ws.row_dimensions[2].height=22

        row_i=3; tot_giorni=tot_ore=0
        for c in cantieri_list:
            rows=db.execute("""SELECT p.*,
                                   COALESCE(p.nome_jolly, u.nome) as nome,
                                   COALESCE(p.cognome_jolly, u.cognome) as cognome
                               FROM presenze p
                               LEFT JOIN utenti u ON u.id=p.utente_id
                               WHERE p.cantiere_id=? AND p.data LIKE ?""",
                            (c['id'],f'{mese_sel}%')).fetchall()
            giorni=len(rows); ore=round(sum(r['ore_totali'] or 0 for r in rows),1)
            dips=len(set(r['utente_id'] for r in rows if r['utente_id']))
            dip_ore={}
            for r in rows:
                k=f"{r['nome']} {r['cognome']}"; dip_ore[k]=dip_ore.get(k,0)+(r['ore_totali'] or 0)
            top=max(dip_ore,key=dip_ore.get) if dip_ore else ''
            if giorni>0:
                ws.cell(row_i,1,c['nome'])
                ws.cell(row_i,2,giorni).alignment=center
                co=ws.cell(row_i,3,ore); co.number_format=num_fmt; co.alignment=center
                cm=ws.cell(row_i,4,round(ore/giorni,1) if giorni else 0); cm.number_format=num_fmt; cm.alignment=center
                ws.cell(row_i,5,dips).alignment=center
                ws.cell(row_i,6,top)
                tot_giorni+=giorni; tot_ore+=ore; row_i+=1

        ws.cell(row_i,1,'TOTALE').font=tot_font; ws.cell(row_i,1).fill=tot_fill
        ws.cell(row_i,2,tot_giorni).font=tot_font; ws.cell(row_i,2).fill=tot_fill; ws.cell(row_i,2).alignment=center
        tc=ws.cell(row_i,3,round(tot_ore,1)); tc.font=tot_font; tc.fill=tot_fill; tc.number_format=num_fmt; tc.alignment=center

    db.close()

    # Aggiungi un secondo foglio con dettaglio giornaliero
    ws2 = wb.create_sheet("Dettaglio presenze")
    det_headers = ['Data','Dipendente','Cantiere','Entrata','Uscita','Ore']
    for i,h in enumerate(det_headers,1):
        c=ws2.cell(1,i,h); c.font=hdr_font; c.fill=hdr_fill; c.alignment=center
    ws2.column_dimensions['A'].width=14; ws2.column_dimensions['B'].width=22; ws2.column_dimensions['C'].width=20
    db2=get_db()
    q2="""SELECT p.*,
              COALESCE(p.nome_jolly, u.nome) as nome,
              COALESCE(p.cognome_jolly, u.cognome) as cognome,
              c.nome as cnome
          FROM presenze p
          LEFT JOIN utenti u ON u.id=p.utente_id
          LEFT JOIN cantieri c ON c.id=p.cantiere_id
          WHERE p.data LIKE ?"""
    params2=[f'{mese_sel}%']
    if uid_sel and vista=='dipendente':
        q2+=" AND p.utente_id=?"; params2.append(int(uid_sel))
    elif uid_sel and vista=='dipendente':
        pass  # already handled
    else:
        pass  # tutti: jolly inclusi
    # Se filtra per dipendente specifico, escludi jolly (utente_id IS NULL)
    if uid_sel and vista=='dipendente':
        q2+=" AND p.nome_jolly IS NULL"
    if cid_sel and vista=='cantiere':   q2+=" AND p.cantiere_id=?"; params2.append(int(cid_sel))
    q2+=" ORDER BY p.data, COALESCE(p.cognome_jolly, u.cognome)"
    det_rows=db2.execute(q2,params2).fetchall(); db2.close()
    for ri,r in enumerate(det_rows,2):
        ws2.cell(ri,1,r['data']); ws2.cell(ri,2,f"{r['nome']} {r['cognome']}")
        ws2.cell(ri,3,r['cnome'] or ''); ws2.cell(ri,4,r['ora_entrata'] or '')
        ws2.cell(ri,5,r['ora_uscita'] or '')
        oc=ws2.cell(ri,6,round(r['ore_totali'],1) if r['ore_totali'] else ''); oc.number_format=num_fmt

    # ── Foglio 3: Spese Rimborsabili (solo vista dipendente) ──
    if vista == 'dipendente':
        grn_fill = PatternFill("solid", fgColor="065F46")
        ws3 = wb.create_sheet("Spese Rimborsabili")
        sp_headers = ['Dipendente','Data','Categoria','Descrizione','Veicolo','Importo €']
        for i,h in enumerate(sp_headers,1):
            c=ws3.cell(1,i,h); c.font=Font(color="FFFFFF",bold=True); c.fill=grn_fill; c.alignment=center
        ws3.column_dimensions['A'].width=22; ws3.column_dimensions['B'].width=14
        ws3.column_dimensions['C'].width=18; ws3.column_dimensions['D'].width=28
        ws3.column_dimensions['E'].width=14; ws3.column_dimensions['F'].width=14
        db3 = get_db()
        sq3 = """SELECT s.*, u.nome, u.cognome, v.targa as veicolo_targa
                 FROM spese_rimborso s JOIN utenti u ON u.id=s.utente_id
                 LEFT JOIN veicoli v ON v.id=s.veicolo_id
                 WHERE s.stato='approvata' AND substr(s.data,1,7)=?"""
        sp3 = [mese_sel]
        if uid_sel: sq3 += " AND s.utente_id=?"; sp3.append(int(uid_sel))
        sq3 += " ORDER BY s.data, u.cognome"
        try:
            sp_rows = db3.execute(sq3, sp3).fetchall()
        except:
            sp_rows = []
        db3.close()
        tot_sp = 0.0
        for ri, s in enumerate(sp_rows, 2):
            ws3.cell(ri,1,f"{s['nome']} {s['cognome']}")
            ws3.cell(ri,2,s['data'])
            ws3.cell(ri,3,s['categoria'])
            ws3.cell(ri,4,s['descrizione'] or '')
            ws3.cell(ri,5,s['veicolo_targa'] or '')
            ac = ws3.cell(ri,6, round(s['importo'],2) if s['importo'] else 0)
            ac.number_format = '#,##0.00 €'; ac.alignment = Alignment(horizontal='right')
            tot_sp += s['importo'] or 0
        if sp_rows:
            tr = len(sp_rows)+2
            ws3.cell(tr,5,'TOTALE RIMBORSI').font = Font(bold=True)
            tc3 = ws3.cell(tr,6,round(tot_sp,2))
            tc3.font = Font(bold=True); tc3.number_format = '#,##0.00 €'
            tc3.fill = PatternFill("solid", fgColor="D1FAE5")
            tc3.alignment = Alignment(horizontal='right')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"report_{vista}_{mese_sel}.xlsx"
    return Response(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fname}',
                 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})

# ══════════════════════════════════════════════════════════
#  DIPENDENTI
# ══════════════════════════════════════════════════════════
DIP_TMPL = """
<div class="card">
  <div class="table-wrap"><table>
    <thead><tr><th>Dipendente</th><th>Titolo</th><th>Email</th><th>Assunto il</th><th>Documenti</th><th>Stato</th>{% if session.ruolo=='admin' %}<th>Azioni</th>{% endif %}</tr></thead>
    <tbody>{% for d in dipendenti %}
    <tr>
      <td><div style="display:flex;align-items:center"><span class="avatar-sm">{{ d.nome[0] }}{{ d.cognome[0] }}</span><div><div style="font-weight:600">{{ d.nome }} {{ d.cognome }}</div><div style="font-size:11px;color:var(--text-light)">{{ d.ruolo }}</div></div></div></td>
      <td>{{ d.mansione or d.titolo or '–' }}</td>
      <td style="color:var(--accent2)">{{ d.email }}</td>
      <td>{{ d.data_assunzione or '–' }}</td>
      <td>
        <a href="/dipendenti/{{ d.id }}/documenti" class="btn btn-secondary btn-sm" style="gap:5px">
          <i class="fa fa-folder-open"></i>
          {% if d.n_docs > 0 %}<span class="badge badge-blue" style="margin-left:2px">{{ d.n_docs }}</span>{% endif %}
        </a>
      </td>
      <td>{% if d.attivo %}<span class="badge badge-green">● Attivo</span>{% else %}<span class="badge badge-red">Inattivo</span>{% endif %}</td>
      {% if session.ruolo=='admin' %}<td style="display:flex;gap:6px">
        <a href="/dipendenti/{{ d.id }}/modifica" class="btn btn-secondary btn-sm"><i class="fa fa-pen"></i></a>
        <a href="/dipendenti/{{ d.id }}/elimina" class="btn btn-danger btn-sm" onclick="return confirm('Rimuovere?')"><i class="fa fa-trash"></i></a>
      </td>{% endif %}
    </tr>{% endfor %}</tbody>
  </table></div>
</div>"""

DIP_FORM_TMPL = """
<div class="card" style="max-width:600px;margin:0 auto">
  <div class="card-header"><h3>{{ 'Modifica' if dip else 'Nuovo' }} Dipendente</h3></div>
  <div class="card-body">
  <form method="POST">
    <div class="form-row"><div class="form-group"><label>Nome *</label><input name="nome" value="{{ dip.nome if dip else '' }}" required></div>
    <div class="form-group"><label>Cognome *</label><input name="cognome" value="{{ dip.cognome if dip else '' }}" required></div></div>
    <div class="form-group"><label>Email *</label><input type="email" name="email" value="{{ dip.email if dip else '' }}" required></div>
    {% if not dip %}<div class="form-group"><label>Password *</label><input type="password" name="password" required></div>{% endif %}
    <div class="form-row">
      <div class="form-group"><label>Mansione / Qualifica</label><input name="mansione" value="{{ dip.mansione if dip else '' }}" placeholder="Es. Operaio, Muratore, Elettricista..."></div>
      <div class="form-group"><label>Telefono</label><input name="telefono" value="{{ dip.telefono if dip else '' }}"></div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Data assunzione</label><input type="date" name="data_assunzione" value="{{ dip.data_assunzione if dip else '' }}"></div>
      <div class="form-group"><label>Ruolo</label><select name="ruolo">
        <option value="dipendente" {{ 'selected' if dip and dip.ruolo=='dipendente' else '' }}>Dipendente</option>
        <option value="admin" {{ 'selected' if dip and dip.ruolo=='admin' else '' }}>Admin</option>
      </select></div>
    </div>
    <div style="display:flex;gap:10px;justify-content:flex-end">
      <a href="/dipendenti" class="btn btn-secondary">Annulla</a>
      <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva</button>
    </div>
  </form></div>
</div>"""

@app.route('/dipendenti')
@login_required
def dipendenti():
    db=get_db()
    dips_raw=db.execute("SELECT * FROM utenti WHERE attivo=1 ORDER BY cognome,nome").fetchall()
    dips=[]
    for d in dips_raw:
        n=db.execute("SELECT COUNT(*) FROM documenti_dipendente WHERE utente_id=?",(d['id'],)).fetchone()[0]
        row=dict(d); row['n_docs']=n; dips.append(row)
    db.close()
    return render_page(DIP_TMPL,page_title='Dipendenti',active='dipendenti',dipendenti=dips)

@app.route('/dipendenti/nuovo',methods=['GET','POST'])
@admin_required
def dipendente_nuovo():
    if request.method=='POST':
        db=get_db()
        try:
            db.execute("INSERT INTO utenti (nome,cognome,email,password,mansione,telefono,data_assunzione,ruolo) VALUES (?,?,?,?,?,?,?,?)",
                (request.form['nome'],request.form['cognome'],request.form['email'].lower(),hash_pw(request.form['password']),
                 request.form.get('mansione'),request.form.get('telefono'),request.form.get('data_assunzione'),request.form.get('ruolo','dipendente')))
            safe_commit(db);flash('Dipendente aggiunto!','success')
        except Exception as e:
            if 'UNIQUE' in str(e):
                flash('Esiste già un dipendente con questa email. Usane una diversa.','error')
            else:
                flash(f'Errore: {e}','error')
        finally:db.close()
        return redirect(url_for('dipendenti'))
    return render_page(DIP_FORM_TMPL,page_title='Nuovo Dipendente',active='dipendenti',dip=None)

@app.route('/dipendenti/<int:uid>/modifica',methods=['GET','POST'])
@admin_required
def dipendente_modifica(uid):
    db=get_db();dip=db.execute("SELECT * FROM utenti WHERE id=?",(uid,)).fetchone()
    if request.method=='POST':
        db.execute("UPDATE utenti SET nome=?,cognome=?,email=?,mansione=?,telefono=?,data_assunzione=?,ruolo=? WHERE id=?",
            (request.form['nome'],request.form['cognome'],request.form['email'].lower(),
             request.form.get('mansione'),request.form.get('telefono'),request.form.get('data_assunzione'),
             request.form.get('ruolo','dipendente'),uid))
        safe_commit(db);db.close();flash('Aggiornato.','success');return redirect(url_for('dipendenti'))
    db.close()
    return render_page(DIP_FORM_TMPL,page_title='Modifica Dipendente',active='dipendenti',dip=dip)

@app.route('/dipendenti/<int:uid>/elimina')
@admin_required
def dipendente_elimina(uid):
    db=get_db();db.execute("UPDATE utenti SET attivo=0 WHERE id=?",(uid,));safe_commit(db);db.close()
    flash('Rimosso.','success');return redirect(url_for('dipendenti'))

# ══════════════════════════════════════════════════════════
#  DOCUMENTI & SCADENZE
# ══════════════════════════════════════════════════════════
DOC_TMPL = """
<!-- Filtri -->
<div class="card" style="margin-bottom:20px">
  <div class="card-body">
    <form method="GET" style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap">
      <div class="form-group" style="margin:0;min-width:180px">
        <label>Dipendente</label>
        <select name="uid">
          <option value="">Tutti i dipendenti</option>
          {% for u in tutti_utenti %}<option value="{{ u.id }}" {{ 'selected' if filtro_uid==u.id|string }}>{{ u.nome }} {{ u.cognome }}</option>{% endfor %}
        </select>
      </div>
      <div class="form-group" style="margin:0;min-width:160px">
        <label>Categoria</label>
        <select name="cat">
          <option value="">Tutte le categorie</option>
          {% for c in categorie %}<option value="{{ c }}" {{ 'selected' if filtro_cat==c }}>{{ c }}</option>{% endfor %}
        </select>
      </div>
      <div class="form-group" style="margin:0;min-width:140px">
        <label>Stato</label>
        <select name="stato">
          <option value="">Tutti</option>
          <option value="scaduto" {{ 'selected' if filtro_stato=='scaduto' }}>⚠ Scaduti</option>
          <option value="presto" {{ 'selected' if filtro_stato=='presto' }}>🔶 In scadenza (30gg)</option>
          <option value="ok" {{ 'selected' if filtro_stato=='ok' }}>✅ Validi</option>
        </select>
      </div>
      <button type="submit" class="btn btn-primary" style="height:40px"><i class="fa fa-search"></i> Filtra</button>
      <a href="/documenti" class="btn btn-secondary" style="height:40px">Reset</a>
    </form>
  </div>
</div>

<!-- Tabella con selezione -->
<div class="card">
  <div class="card-header">
    <h3><i class="fa fa-file-alt" style="color:var(--accent2);margin-right:8px"></i>
      {{ documenti|length }} documento{{ 'i' if documenti|length != 1 else '' }}
    </h3>
    <div style="display:flex;gap:8px" id="bulk-actions" class="hidden">
      <button onclick="scaricaSelezionati()" class="btn btn-green btn-sm"><i class="fa fa-download"></i> Scarica selezionati (ZIP)</button>
      <span id="n-sel" style="font-size:12px;color:var(--text-light);align-self:center"></span>
    </div>
  </div>
  {% if documenti %}
  <form id="bulk-form" action="/documenti/scarica-zip" method="POST">
  <div class="table-wrap"><table>
    <thead><tr>
      <th style="width:36px"><input type="checkbox" id="sel-all" onchange="toggleAll(this)" title="Seleziona tutti"></th>
      <th>Titolo</th><th>Categoria</th><th>Dipendente</th><th>Scadenza</th><th>Stato</th>
      {% if session.ruolo=='admin' %}<th>Azioni</th>{% endif %}
    </tr></thead>
    <tbody>{% for d in documenti %}
    <tr id="row-{{ d.id }}">
      <td><input type="checkbox" name="doc_ids" value="{{ d.id }}" class="doc-check" onchange="aggiornaContatore()"></td>
      <td>
        <i class="fa fa-file-alt" style="color:var(--accent);margin-right:8px"></i>
        <strong>{{ d.titolo }}</strong>
        {% if d.sorgente == 'dipendente' %}
          <span style="font-size:10px;background:#eff6ff;color:#2563eb;border-radius:4px;padding:1px 6px;margin-left:4px">📎 file</span>
        {% endif %}
        {% if d.descrizione %}<div style="font-size:11px;color:var(--text-light)">{{ d.descrizione[:80] }}</div>{% endif %}
      </td>
      <td><span class="tag">{{ d.categoria or 'Generale' }}</span></td>
      <td>
        {% if d.dest_nome %}
          <span class="avatar-sm" style="font-size:9px">{{ d.dest_nome[:2] }}</span>{{ d.dest_nome }}
        {% else %}<span style="color:var(--text-light)">Tutti</span>{% endif %}
      </td>
      <td style="font-family:monospace;font-size:12px">{{ d.data_scadenza or '–' }}</td>
      <td>
        {% if not d.data_scadenza %}<span class="badge badge-gray">–</span>
        {% elif d.days_left < 0 %}<span class="badge badge-red">Scaduto</span>
        {% elif d.days_left <= 7 %}<span class="badge badge-red">⚠ {{ d.days_left }}gg</span>
        {% elif d.days_left <= 30 %}<span class="badge badge-amber">{{ d.days_left }}gg</span>
        {% else %}<span class="badge badge-green">OK</span>{% endif %}
      </td>
      {% if session.ruolo=='admin' %}<td style="display:flex;gap:6px">
        {% if d.sorgente == 'dipendente' %}
          <a href="/dipendenti/{{ d.assegnato_a }}/documenti/{{ d.id }}/scarica" class="btn btn-secondary btn-sm" title="Scarica"><i class="fa fa-download"></i></a>
          <a href="/dipendenti/{{ d.assegnato_a }}/documenti" class="btn btn-secondary btn-sm" title="Vai al dipendente"><i class="fa fa-user"></i></a>
          <a href="/dipendenti/{{ d.assegnato_a }}/documenti/{{ d.id }}/elimina" class="btn btn-danger btn-sm" onclick="return confirm('Eliminare?')" title="Elimina"><i class="fa fa-trash"></i></a>
        {% else %}
          <a href="/documenti/{{ d.id }}/modifica" class="btn btn-secondary btn-sm"><i class="fa fa-pen"></i></a>
          <a href="/documenti/{{ d.id }}/elimina" class="btn btn-danger btn-sm" onclick="return confirm('Eliminare?')"><i class="fa fa-trash"></i></a>
        {% endif %}
      </td>{% endif %}
    </tr>{% endfor %}</tbody>
  </table></div>
  </form>
  {% else %}
  <div class="empty-state"><i class="fa fa-folder-open"></i><p>Nessun documento trovato</p></div>
  {% endif %}
</div>

<script>
function toggleAll(cb) {
  document.querySelectorAll('.doc-check').forEach(c => c.checked = cb.checked);
  aggiornaContatore();
}
function aggiornaContatore() {
  const sel = document.querySelectorAll('.doc-check:checked').length;
  const tot = document.querySelectorAll('.doc-check').length;
  document.getElementById('sel-all').indeterminate = sel > 0 && sel < tot;
  document.getElementById('sel-all').checked = sel === tot && tot > 0;
  const ba = document.getElementById('bulk-actions');
  const ns = document.getElementById('n-sel');
  if (sel > 0) { ba.classList.remove('hidden'); ns.textContent = sel + ' selezionati'; }
  else { ba.classList.add('hidden'); }
}
function scaricaSelezionati() {
  document.getElementById('bulk-form').submit();
}
</script>
<style>.hidden{display:none!important}</style>
"""

DOC_FORM_TMPL = """
<div class="card" style="max-width:600px;margin:0 auto">
  <div class="card-header"><h3>{{ 'Modifica' if doc else 'Nuovo' }} Documento</h3></div>
  <div class="card-body"><form method="POST">
    <div class="form-group"><label>Titolo *</label><input name="titolo" value="{{ doc.titolo if doc else '' }}" required></div>
    <div class="form-group"><label>Descrizione</label><textarea name="descrizione">{{ doc.descrizione if doc else '' }}</textarea></div>
    <div class="form-row">
      <div class="form-group"><label>Categoria</label><input name="categoria" value="{{ doc.categoria if doc else '' }}" placeholder="es. Contratti"></div>
      <div class="form-group"><label>Scadenza</label><input type="date" name="data_scadenza" value="{{ doc.data_scadenza if doc else '' }}"></div>
    </div>
    <div class="form-group"><label>Assegnato a</label>
      <select name="assegnato_a"><option value="">Tutti</option>
      {% for u in utenti %}<option value="{{ u.id }}" {{ 'selected' if doc and doc.assegnato_a==u.id }}>{{ u.nome }} {{ u.cognome }}</option>{% endfor %}
      </select></div>
    <div style="display:flex;gap:10px;justify-content:flex-end">
      <a href="/documenti" class="btn btn-secondary">Annulla</a>
      <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva</button>
    </div>
  </form></div>
</div>"""

@app.route('/documenti')
@login_required
def documenti():
    db = get_db()
    uid = session['user_id']
    cat         = request.args.get('cat','')
    filtro_uid  = request.args.get('uid','')
    filtro_stato= request.args.get('stato','')

    # ── Tabella "documenti" (senza file fisico) ──
    q = """SELECT d.id, d.titolo, d.categoria, d.data_scadenza,
               u.nome||' '||u.cognome as dest_nome, d.assegnato_a,
               CAST(julianday(d.data_scadenza)-julianday('now') AS INTEGER) as days_left,
               NULL as nome_file, 'documenti' as sorgente
           FROM documenti d LEFT JOIN utenti u ON u.id=d.assegnato_a WHERE 1=1"""
    params = []
    if session['ruolo'] != 'admin':
        q += " AND (d.assegnato_a IS NULL OR d.assegnato_a=?)"; params.append(uid)
    if cat:        q += " AND d.categoria=?";    params.append(cat)
    if filtro_uid: q += " AND d.assegnato_a=?";  params.append(int(filtro_uid))

    docs1 = [dict(d) for d in db.execute(q, params).fetchall()]

    # ── Tabella "documenti_dipendente" (con file fisico, caricati da Dipendenti) ──
    q2 = """SELECT dd.id, dd.nome_originale as titolo, dd.tipo_doc as categoria,
                dd.data_scadenza,
                dd.note as descrizione,
                u.nome||' '||u.cognome as dest_nome, dd.utente_id as assegnato_a,
                CAST(julianday(dd.data_scadenza)-julianday('now') AS INTEGER) as days_left,
                dd.nome_file, 'dipendente' as sorgente
            FROM documenti_dipendente dd JOIN utenti u ON u.id=dd.utente_id WHERE 1=1"""
    params2 = []
    if session['ruolo'] != 'admin':
        q2 += " AND dd.utente_id=?"; params2.append(uid)
    if filtro_uid: q2 += " AND dd.utente_id=?";  params2.append(int(filtro_uid))

    docs2 = [dict(d) for d in db.execute(q2, params2).fetchall()]

    cats = ['UNILAV','Visita medica','Corso/Attestato','Documento identità','Contratto','Altro']
    tutti_utenti = db.execute("SELECT id,nome,cognome FROM utenti WHERE attivo=1 ORDER BY cognome").fetchall()
    db.close()

    docs_list = docs1 + docs2
    for d in docs_list:
        if d['days_left'] is None: d['days_left'] = 9999

    # Filtro categoria (solo docs1 ha categorie definite, docs2 usa tipo_doc)
    if cat:
        docs_list = [d for d in docs_list if d.get('categoria') == cat]

    # Filtro stato
    if filtro_stato == 'scaduto':
        docs_list = [d for d in docs_list if d['data_scadenza'] and d['days_left'] < 0]
    elif filtro_stato == 'presto':
        docs_list = [d for d in docs_list if d['data_scadenza'] and 0 <= d['days_left'] <= 30]
    elif filtro_stato == 'ok':
        docs_list = [d for d in docs_list if not d['data_scadenza'] or d['days_left'] > 30]

    docs_list.sort(key=lambda d: d['data_scadenza'] or '9999-99-99')

    return render_page(DOC_TMPL, page_title='Documenti', active='documenti',
        documenti=docs_list, categorie=cats,
        filtro_cat=cat, filtro_uid=filtro_uid, filtro_stato=filtro_stato,
        tutti_utenti=tutti_utenti)

@app.route('/documenti/nuovo',methods=['GET','POST'])
@admin_required
def documento_nuovo():
    if request.method=='POST':
        db=get_db(); a=request.form.get('assegnato_a') or None
        db.execute("INSERT INTO documenti (titolo,descrizione,categoria,data_scadenza,creato_da,assegnato_a) VALUES (?,?,?,?,?,?)",
            (request.form['titolo'],request.form.get('descrizione'),request.form.get('categoria'),
             request.form.get('data_scadenza') or None,session['user_id'],int(a) if a else None))
        safe_commit(db);db.close();flash('Documento aggiunto!','success');return redirect(url_for('documenti'))
    db=get_db();utenti=db.execute("SELECT id,nome,cognome FROM utenti WHERE attivo=1").fetchall();db.close()
    return render_page(DOC_FORM_TMPL,page_title='Nuovo Documento',active='documenti',doc=None,utenti=utenti)

@app.route('/documenti/<int:did>/modifica',methods=['GET','POST'])
@admin_required
def documento_modifica(did):
    db=get_db();doc=db.execute("SELECT * FROM documenti WHERE id=?",(did,)).fetchone()
    if request.method=='POST':
        a=request.form.get('assegnato_a') or None
        db.execute("UPDATE documenti SET titolo=?,descrizione=?,categoria=?,data_scadenza=?,assegnato_a=? WHERE id=?",
            (request.form['titolo'],request.form.get('descrizione'),request.form.get('categoria'),
             request.form.get('data_scadenza') or None,int(a) if a else None,did))
        safe_commit(db);db.close();flash('Aggiornato.','success');return redirect(url_for('documenti'))
    utenti=db.execute("SELECT id,nome,cognome FROM utenti WHERE attivo=1").fetchall();db.close()
    return render_page(DOC_FORM_TMPL,page_title='Modifica Documento',active='documenti',doc=doc,utenti=utenti)

@app.route('/documenti/<int:did>/elimina')
@admin_required
def documento_elimina(did):
    db=get_db();db.execute("DELETE FROM documenti WHERE id=?",(did,));safe_commit(db);db.close()
    flash('Eliminato.','success');return redirect(url_for('documenti'))

@app.route('/documenti/scarica-zip', methods=['POST'])
@login_required
def documenti_scarica_zip():
    doc_ids = request.form.getlist('doc_ids')
    if not doc_ids:
        flash('Nessun documento selezionato.', 'error')
        return redirect(url_for('documenti'))

    db  = get_db()
    buf = io.BytesIO()
    aggiunti = 0
    path_aggiunti = set()  # evita file duplicati nel ZIP

    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for did in doc_ids:
            doc = db.execute(
                "SELECT d.*,u.nome||' '||u.cognome as dest_nome FROM documenti d LEFT JOIN utenti u ON u.id=d.assegnato_a WHERE d.id=?",
                (int(did),)).fetchone()
            if not doc or not doc['assegnato_a']: continue

            file_doc = None
            # Cerca tramite collegamento diretto
            if doc['doc_dipendente_id']:
                file_doc = db.execute(
                    "SELECT * FROM documenti_dipendente WHERE id=?",
                    (doc['doc_dipendente_id'],)).fetchone()

            # Fallback: cerca per categoria → tipo_doc
            if not file_doc:
                cat = (doc['categoria'] or '').upper()
                tipo_like = {
                    'UNILAV': '%UNILAV%',
                    'VISITA MEDICA': '%VISITA%',
                    'CORSO/ATTESTATO': '%ATTESTAT%',
                    'DOCUMENTO IDENTITÀ': '%IDENTIT%',
                    'CONTRATTO': '%CONTRATT%',
                }.get(cat, '%%')
                file_doc = db.execute(
                    "SELECT * FROM documenti_dipendente WHERE utente_id=? AND tipo_doc LIKE ? ORDER BY caricato_il DESC LIMIT 1",
                    (doc['assegnato_a'], tipo_like)).fetchone()

            # Ultimo fallback: qualsiasi file del dipendente non ancora aggiunto
            if not file_doc:
                tutti = db.execute(
                    "SELECT * FROM documenti_dipendente WHERE utente_id=? ORDER BY caricato_il DESC",
                    (doc['assegnato_a'],)).fetchall()
                for fd in tutti:
                    p = os.path.join(get_upload_path(doc['assegnato_a']), fd['nome_file'])
                    if p not in path_aggiunti:
                        file_doc = fd; break

            if not file_doc: continue
            path = os.path.join(get_upload_path(doc['assegnato_a']), file_doc['nome_file'])
            if not os.path.exists(path) or path in path_aggiunti: continue

            dest = doc['dest_nome'] or 'Generale'
            # Nome nel ZIP = titolo documento + estensione originale
            _, ext = os.path.splitext(file_doc['nome_originale'])
            safe_titolo = doc['titolo'][:60].replace('/', '-').replace('\\', '-')
            arcname = f"{dest}/{safe_titolo}{ext}"

            zf.write(path, arcname=arcname)
            path_aggiunti.add(path)
            aggiunti += 1
    db.close()
    if aggiunti == 0:
        flash('Nessun file fisico trovato per i documenti selezionati.', 'error')
        return redirect(url_for('documenti'))
    buf.seek(0)
    from datetime import date as _d
    fname = f"documenti_{_d.today().isoformat()}.zip"
    return Response(buf, mimetype='application/zip',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})

@app.route('/scadenze')
@login_required
def scadenze():
    db=get_db()
    # Documenti generici
    docs1=db.execute("""SELECT d.titolo, d.categoria, d.data_scadenza,
        u.nome||' '||u.cognome as dest_nome,
        CAST(julianday(d.data_scadenza)-julianday('now') AS INTEGER) as days_left
        FROM documenti d LEFT JOIN utenti u ON u.id=d.assegnato_a
        WHERE d.data_scadenza IS NOT NULL ORDER BY d.data_scadenza""").fetchall()
    # Documenti caricati dalla sezione Dipendenti
    docs2=db.execute("""SELECT dd.nome_originale as titolo, dd.tipo_doc as categoria, dd.data_scadenza,
        u.nome||' '||u.cognome as dest_nome,
        CAST(julianday(dd.data_scadenza)-julianday('now') AS INTEGER) as days_left
        FROM documenti_dipendente dd JOIN utenti u ON u.id=dd.utente_id
        WHERE dd.data_scadenza IS NOT NULL ORDER BY dd.data_scadenza""").fetchall()
    db.close()
    docs_list=[dict(d) for d in list(docs1)+list(docs2)]
    docs_list.sort(key=lambda d: d['data_scadenza'])
    scadute=sum(1 for d in docs_list if d['days_left']<0)
    imminenti=sum(1 for d in docs_list if 0<=d['days_left']<=30)
    ok=sum(1 for d in docs_list if d['days_left']>30)
    SCAD_TMPL="""<div class="stats-grid" style="grid-template-columns:repeat(3,1fr);margin-bottom:20px">
      <div class="stat-card"><div class="stat-icon red"><i class="fa fa-exclamation-triangle"></i></div><div><div class="stat-val">{{ s.scadute }}</div><div class="stat-lbl">Scadute</div></div></div>
      <div class="stat-card"><div class="stat-icon amber"><i class="fa fa-clock"></i></div><div><div class="stat-val">{{ s.imminenti }}</div><div class="stat-lbl">Entro 30 giorni</div></div></div>
      <div class="stat-card"><div class="stat-icon green"><i class="fa fa-check-circle"></i></div><div><div class="stat-val">{{ s.ok }}</div><div class="stat-lbl">In regola</div></div></div>
    </div>
    <div class="card"><div class="table-wrap"><table>
      <thead><tr><th>Documento</th><th>Categoria</th><th>Assegnato</th><th>Scadenza</th><th>Stato</th><th>Giorni</th></tr></thead>
      <tbody>{% for d in docs %}
      <tr><td><strong>{{ d.titolo }}</strong></td><td><span class="tag">{{ d.categoria or '–' }}</span></td>
      <td>{{ d.dest_nome or 'Tutti' }}</td><td style="font-family:monospace;font-size:12px">{{ d.data_scadenza }}</td>
      <td>{% if d.days_left<0 %}<span class="badge badge-red">Scaduto</span>
          {% elif d.days_left<=7 %}<span class="badge badge-red">Urgente</span>
          {% elif d.days_left<=30 %}<span class="badge badge-amber">Imminente</span>
          {% else %}<span class="badge badge-green">OK</span>{% endif %}</td>
      <td><strong style="color:{{ 'var(--danger)' if d.days_left<0 else 'var(--warning)' if d.days_left<=30 else 'var(--success)' }}">{{ d.days_left }}gg</strong></td>
      </tr>{% else %}<tr><td colspan="6"><div class="empty-state"><i class="fa fa-calendar-check"></i><p>Nessuna scadenza</p></div></td></tr>{% endfor %}
      </tbody></table></div></div>"""
    return render_page(SCAD_TMPL,page_title='Scadenze',active='scadenze',docs=docs_list,s={'scadute':scadute,'imminenti':imminenti,'ok':ok})

# ══════════════════════════════════════════════════════════
#  CALENDARIO
# ══════════════════════════════════════════════════════════

TIPO_COLORI_MAP = {
    'montaggio':  '#16a34a',
    'fiera':      '#2196F3',
    'smontaggio': '#e63946',
    'altro':      '#8b5cf6',
}
TIPO_LABEL_MAP = {
    'montaggio':  '🔧 Montaggio',
    'fiera':      '🏛️ Fiera',
    'smontaggio': '🔩 Smontaggio',
    'altro':      '📌 Altro',
}

CAL_TMPL = """
<style>
.cal-event{cursor:pointer;transition:.15s}
.cal-event:hover{opacity:.85;transform:scale(1.02)}
.ev-badge{display:inline-flex;align-items:center;gap:5px;padding:3px 10px;border-radius:99px;font-size:11px;font-weight:700;color:#fff}
.team-chip{display:inline-flex;align-items:center;gap:5px;background:#f1f5f9;border-radius:99px;padding:4px 12px;font-size:12px;font-weight:600;margin:3px;color:#1e293b}
</style>

<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;flex-wrap:wrap;gap:10px">
  <div class="cal-nav">
    <a href="/calendario?m={{ prev_m }}&y={{ prev_y }}" class="btn btn-secondary btn-sm"><i class="fa fa-chevron-left"></i></a>
    <span>{{ mese_nome }} {{ anno }}</span>
    <a href="/calendario?m={{ next_m }}&y={{ next_y }}" class="btn btn-secondary btn-sm"><i class="fa fa-chevron-right"></i></a>
    <a href="/calendario" class="btn btn-secondary btn-sm">Oggi</a>
  </div>
  <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
    <!-- Legenda tipi -->
    <span class="ev-badge" style="background:#16a34a">🔧 Montaggio</span>
    <span class="ev-badge" style="background:#2196F3">🏛️ Fiera</span>
    <span class="ev-badge" style="background:#0f4c81">🔩 Smontaggio</span>
    <span class="ev-badge" style="background:#8b5cf6">📌 Altro</span>
    {% if session.ruolo=='admin' %}
    <a href="/calendario/nuovo" class="btn btn-primary btn-sm"><i class="fa fa-plus"></i> Nuovo evento</a>
    <a href="/calendario/pdf?m={{ m }}&y={{ anno }}" target="_blank" class="btn btn-sm" style="background:#0f172a;color:#fff"><i class="fa fa-file-pdf"></i> PDF Calendario</a>
    {% endif %}
  </div>
</div>

<!-- Griglia calendario -->
<div class="card"><div class="card-body" style="padding:14px">
  <div class="cal-header-grid">{% for g in ['Lun','Mar','Mer','Gio','Ven','Sab','Dom'] %}<div class="cal-dow">{{ g }}</div>{% endfor %}</div>
  <div class="cal-grid">{% for cell in calendario %}
    <div class="cal-day {{ 'today' if cell.today }} {{ 'other-month' if not cell.cur_month }}">
      <div class="cal-day-num">{{ cell.day }}</div>
      {% for ev in cell.eventi %}
      <a href="/calendario/{{ ev.id }}" style="text-decoration:none;display:block">
        <div class="cal-event" style="background:{{ ev.colore }};color:#fff;border-radius:4px;padding:2px 5px;font-size:10px;font-weight:600;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">
          {{ ev.titolo }}
        </div>
      </a>
      {% endfor %}
    </div>{% endfor %}
  </div>
</div></div>

<!-- Lista eventi del mese -->
<div class="card" style="margin-top:20px">
  <div class="card-header"><h3><i class="fa fa-list" style="color:var(--accent2)"></i> Lavori di {{ mese_nome }} {{ anno }}</h3></div>
  <div class="table-wrap"><table>
    <thead><tr><th>Tipo</th><th>Evento / Luogo</th><th>Dal</th><th>Al</th><th>Team</th>{% if session.ruolo=='admin' %}<th></th>{% endif %}</tr></thead>
    <tbody>{% for e in eventi_mese %}
    <tr>
      <td><span class="ev-badge" style="background:{{ e.colore }}">{{ e.tipo_label }}</span></td>
      <td>
        <a href="/calendario/{{ e.id }}" style="font-weight:700;color:var(--text);text-decoration:none">{{ e.titolo }}</a>
        {% if e.luogo %}<div style="font-size:11px;color:var(--text-light);margin-top:2px"><i class="fa fa-map-marker-alt"></i> {{ e.luogo }}</div>{% endif %}
      </td>
      <td style="font-family:monospace;font-size:12px">{{ e.data_inizio }}</td>
      <td style="font-family:monospace;font-size:12px">{{ e.data_fine or e.data_inizio }}</td>
      <td>
        {% for d in e.team %}
        <span class="team-chip">
          <span style="width:20px;height:20px;border-radius:50%;background:var(--accent2);display:inline-flex;align-items:center;justify-content:center;color:#fff;font-size:9px;font-weight:700">{{ d.iniziali }}</span>
          {{ d.nome }}
        </span>
        {% else %}<span style="color:var(--text-light);font-size:12px">—</span>{% endfor %}
      </td>
      {% if session.ruolo=='admin' %}
      <td style="display:flex;gap:6px">
        <a href="/calendario/{{ e.id }}" class="btn btn-sm btn-secondary"><i class="fa fa-eye"></i></a>
        <a href="/calendario/{{ e.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare?')"><i class="fa fa-trash"></i></a>
      </td>
      {% endif %}
    </tr>
    {% else %}<tr><td colspan="6"><div class="empty-state"><i class="fa fa-calendar"></i><p>Nessun evento questo mese.</p></div></td></tr>{% endfor %}
    </tbody>
  </table></div>
</div>
"""

CAL_DETAIL_TMPL = """
<style>
.ev-badge{display:inline-flex;align-items:center;gap:5px;padding:4px 11px;border-radius:99px;font-size:11px;font-weight:700;color:#fff}
.pass-btn{background:linear-gradient(135deg,#1e293b,#0f172a);color:#fff;border:none;padding:10px 20px;border-radius:10px;font-size:13px;font-weight:700;cursor:pointer;display:inline-flex;align-items:center;gap:8px;text-decoration:none;transition:all .2s}
.pass-btn:hover{background:linear-gradient(135deg,#334155,#1e293b);transform:translateY(-1px);box-shadow:0 4px 12px rgba(0,0,0,.2)}
.doc-chip{display:inline-flex;align-items:center;gap:6px;background:#f1f5f9;border:1px solid var(--border);border-radius:8px;padding:5px 10px;font-size:12px;font-weight:600;margin:3px}
.doc-chip.ok{background:#dcfce7;border-color:#bbf7d0;color:#16a34a}
.doc-chip.missing{background:#fee2e2;border-color:#fecaca;color:#dc2626}
.veicolo-card{background:#f8fafc;border:1px solid var(--border);border-radius:10px;padding:14px 16px;margin-bottom:8px;display:flex;align-items:center;justify-content:space-between;gap:12px}
</style>

<div style="margin-bottom:16px;display:flex;gap:8px;flex-wrap:wrap;align-items:center">
  <a href="/calendario" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Calendario</a>
  {% if session.ruolo=='admin' %}
  <a href="/calendario/{{ ev.id }}/modifica" class="btn btn-secondary btn-sm"><i class="fa fa-pen"></i> Modifica</a>
  <a href="/calendario/{{ ev.id }}/elimina" class="btn btn-danger btn-sm" onclick="return confirm('Eliminare?')"><i class="fa fa-trash"></i></a>
  {% endif %}
  <!-- BOTTONE GENERA PASS -->
  {% if team or veicoli_evento %}
  <a href="/calendario/{{ ev.id }}/pacchetto-pass" target="_blank" class="pass-btn" style="margin-left:auto">
    <i class="fa fa-id-card"></i> Genera Pacchetto Pass
    <span style="background:rgba(255,255,255,.15);border-radius:99px;padding:2px 8px;font-size:11px">{{ team|length }} pers. · {{ veicoli_evento|length }} mezzi</span>
  </a>
  {% endif %}
</div>

{% if prev_collegato %}
<div style="background:linear-gradient(90deg,#eff6ff,#f0fdf4);border:1px solid #bfdbfe;border-radius:12px;padding:16px 20px;margin-bottom:20px;display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap">
  <div style="display:flex;align-items:center;gap:14px">
    <div style="width:42px;height:42px;background:#2563eb;border-radius:10px;display:flex;align-items:center;justify-content:center;flex-shrink:0">
      <i class="fa fa-file-invoice" style="color:#fff;font-size:18px"></i>
    </div>
    <div>
      <div style="font-size:11px;color:#3b82f6;font-weight:700;text-transform:uppercase;letter-spacing:.6px">Preventivo collegato</div>
      <div style="font-weight:700;font-size:15px;color:#1e293b">N.{{ prev_collegato.numero }} — {{ prev_collegato.oggetto or '' }}</div>
      <div style="font-size:12px;color:#64748b">
        Cliente: <strong>{{ prev_collegato.cliente_nome or '–' }}</strong>
        {% if prev_collegato.totale_finale %} &nbsp;·&nbsp; Totale: <strong>€ {{ "%.2f"|format(prev_collegato.totale_finale) }}</strong>{% endif %}
        {% if prev_collegato.data_inizio_lavoro %} &nbsp;·&nbsp; 📅 {{ prev_collegato.data_inizio_lavoro }}{% if prev_collegato.data_fine_lavoro %} → {{ prev_collegato.data_fine_lavoro }}{% endif %}{% endif %}
      </div>
    </div>
  </div>
  <a href="/preventivi/{{ prev_collegato.id }}/modifica" class="btn btn-secondary btn-sm"><i class="fa fa-external-link-alt"></i> Apri preventivo</a>
</div>
{% endif %}

<!-- Riga 1: info + team -->
<div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:20px">
  <!-- Info evento -->
  <div class="card">
    <div class="card-header" style="background:{{ ev.colore }}12;border-left:4px solid {{ ev.colore }}">
      <h3 style="color:{{ ev.colore }}"><i class="fa fa-calendar-day"></i> {{ ev.titolo }}</h3>
      <span class="ev-badge" style="background:{{ ev.colore }}">{{ ev.tipo_label }}</span>
    </div>
    <div class="card-body">
      <table style="width:100%;font-size:13.5px"><tbody>
        {% if ev.luogo %}<tr><td style="color:var(--text-light);padding:6px 0;width:40%"><i class="fa fa-map-marker-alt"></i> Luogo</td><td><strong>{{ ev.luogo }}</strong></td></tr>{% endif %}
        <tr><td style="color:var(--text-light);padding:6px 0">📅 Inizio</td><td><strong>{{ ev.data_inizio }}</strong></td></tr>
        <tr><td style="color:var(--text-light);padding:6px 0">📅 Fine</td><td><strong>{{ ev.data_fine or ev.data_inizio }}</strong></td></tr>
        {% if ev.descrizione %}<tr><td style="color:var(--text-light);padding:6px 0">📝 Note</td><td>{{ ev.descrizione }}</td></tr>{% endif %}
      </tbody></table>
    </div>
  </div>

  <!-- Team assegnato -->
  <div class="card">
    <div class="card-header">
      <h3><i class="fa fa-users" style="color:var(--accent2)"></i> Team ({{ team|length }})</h3>
      {% if session.ruolo=='admin' %}
      <button onclick="document.getElementById('modal-team').style.display='flex'" class="btn btn-sm btn-primary"><i class="fa fa-user-plus"></i> Assegna</button>
      {% endif %}
    </div>
    <div class="card-body" style="padding:14px">
      {% if team %}
      {% for ruolo_k, membri in team_per_ruolo.items() %}
      <div style="margin-bottom:12px">
        <div style="font-size:10px;font-weight:700;color:var(--text-light);text-transform:uppercase;letter-spacing:.7px;margin-bottom:7px">{{ ruolo_k }}</div>
        {% for d in membri %}
        <div style="display:flex;align-items:center;justify-content:space-between;padding:8px 10px;background:#f8fafc;border-radius:9px;margin-bottom:4px">
          <div style="display:flex;align-items:center;gap:10px">
            <div style="width:32px;height:32px;border-radius:50%;background:var(--accent2);display:flex;align-items:center;justify-content:center;color:#fff;font-size:12px;font-weight:700">{{ d.iniziali }}</div>
            <div>
              <div style="font-weight:600;font-size:13px">{{ d.nome }}</div>
              <div style="font-size:11px;color:var(--text-light);display:flex;gap:6px;align-items:center;flex-wrap:wrap">
                {% if d.ha_ci %}<span class="doc-chip ok" style="padding:2px 7px;font-size:10px"><i class="fa fa-id-card"></i> CI ✓</span>
                {% else %}<span class="doc-chip missing" style="padding:2px 7px;font-size:10px"><i class="fa fa-id-card"></i> CI mancante</span>{% endif %}
                {% if d.data_da or d.data_a %}
                <span style="background:#dbeafe;color:#1d4ed8;border-radius:5px;padding:2px 7px;font-size:10px;font-weight:600">
                  <i class="fa fa-calendar-days"></i>
                  {{ d.data_da or ev.data_inizio }} → {{ d.data_a or ev.data_fine or ev.data_inizio }}
                </span>
                {% else %}
                <span style="background:#f1f5f9;color:#64748b;border-radius:5px;padding:2px 7px;font-size:10px">tutto l'evento</span>
                {% endif %}
              </div>
            </div>
          </div>
          {% if session.ruolo=='admin' %}
          <div style="display:flex;gap:4px">
            <a href="/calendario/{{ ev.id }}/assegna-periodo/{{ d.assegn_id }}?da={{ d.data_da or ev.data_inizio }}&a={{ d.data_a or (ev.data_fine or ev.data_inizio) }}"
               class="btn btn-sm btn-secondary" title="Modifica periodo"><i class="fa fa-calendar-pen"></i></a>
            <a href="/calendario/{{ ev.id }}/rimuovi/{{ d.assegn_id }}" class="btn btn-sm btn-danger" onclick="return confirm('Rimuovere?')"><i class="fa fa-times"></i></a>
          </div>
          {% endif %}
        </div>
        {% endfor %}
      </div>
      {% endfor %}
      {% else %}
      <div class="empty-state" style="padding:24px"><i class="fa fa-users"></i><p>Nessun dipendente assegnato.</p></div>
      {% endif %}
    </div>
  </div>
</div>

<!-- Riga 2: veicoli -->
<div class="card" style="margin-bottom:20px">
  <div class="card-header">
    <h3><i class="fa fa-truck" style="color:var(--warning)"></i> Mezzi assegnati ({{ veicoli_evento|length }})</h3>
    {% if session.ruolo=='admin' %}
    <button onclick="document.getElementById('modal-veicolo').style.display='flex'" class="btn btn-sm btn-primary"><i class="fa fa-plus"></i> Aggiungi mezzo</button>
    {% endif %}
  </div>
  <div class="card-body" style="padding:14px">
    {% if veicoli_evento %}
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:10px">
    {% for v in veicoli_evento %}
    <div class="veicolo-card">
      <div style="display:flex;align-items:center;gap:12px;flex:1">
        <div style="width:42px;height:42px;background:#fef3c7;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0">
          {{ '🚐' if v.tipo=='Furgone' else ('🚛' if v.tipo in ('Camion','Autocarro') else '🚗') }}
        </div>
        <div>
          <div style="font-weight:800;font-size:16px;font-family:monospace;letter-spacing:2px;color:var(--accent2)">{{ v.targa }}</div>
          <div style="font-size:12px;color:var(--text-light);margin-top:1px">{{ v.marca or '' }} {{ v.modello or '' }}{% if v.tipo %} · {{ v.tipo }}{% endif %}</div>
          <div style="margin-top:5px;display:flex;gap:5px;flex-wrap:wrap">
            {% if v.ha_libretto %}<span class="doc-chip ok" style="padding:2px 7px;font-size:10px"><i class="fa fa-file-alt"></i> Libretto ✓</span>
            {% else %}<span class="doc-chip missing" style="padding:2px 7px;font-size:10px"><i class="fa fa-file-alt"></i> Libretto mancante</span>{% endif %}
          </div>
        </div>
      </div>
      {% if session.ruolo=='admin' %}
      <a href="/calendario/{{ ev.id }}/rimuovi-veicolo/{{ v.ev_veicolo_id }}" class="btn btn-sm btn-danger" onclick="return confirm('Rimuovere mezzo?')"><i class="fa fa-times"></i></a>
      {% endif %}
    </div>
    {% endfor %}
    </div>
    {% else %}
    <div class="empty-state" style="padding:24px"><i class="fa fa-truck"></i><p>Nessun mezzo assegnato.<br><small>Aggiungi i mezzi che andranno alla fiera.</small></p></div>
    {% endif %}
  </div>
</div>

<!-- ── Disegni / Progetti ── -->
<div class="card" style="margin-bottom:20px">
  <div class="card-header">
    <h3><i class="fa fa-drafting-compass" style="color:#6d28d9"></i> Disegni / Progetti</h3>
    {% if session.ruolo=='admin' %}
    <button onclick="document.getElementById('modal-disegno').style.display='flex'" class="btn btn-sm btn-primary">
      <i class="fa fa-upload"></i> Carica
    </button>
    {% endif %}
  </div>
  <div class="card-body" style="padding:14px">
    {% set disegni = files_evento | selectattr('categoria','equalto','disegno') | list %}
    {% if disegni %}
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:10px">
    {% for f in disegni %}
    <div style="background:#f8fafc;border:1px solid var(--border);border-radius:10px;padding:12px 14px;display:flex;align-items:center;gap:12px">
      <div style="width:40px;height:40px;background:#ede9fe;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:18px;flex-shrink:0">
        {{ '📐' if f.nome_originale.lower().endswith(('.dwg','.dxf')) else ('🖼️' if f.nome_originale.lower().endswith(('.jpg','.jpeg','.png','.gif')) else '📄') }}
      </div>
      <div style="flex:1;min-width:0">
        <div style="font-weight:600;font-size:13px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{{ f.nome_originale }}</div>
        {% if f.note %}<div style="font-size:11px;color:var(--text-muted)">{{ f.note }}</div>{% endif %}
        <div style="font-size:10px;color:var(--text-muted);margin-top:2px">{{ (f.dimensione/1024)|round(0)|int }} KB · {{ f.caricato_il[:10] }}</div>
      </div>
      <div style="display:flex;flex-direction:column;gap:4px">
        <a href="/calendario/{{ ev.id }}/files/{{ f.id }}/scarica" class="btn btn-sm btn-blue" title="Scarica"><i class="fa fa-download"></i></a>
        {% if session.ruolo=='admin' %}
        <a href="/calendario/{{ ev.id }}/files/{{ f.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare?')" title="Elimina"><i class="fa fa-trash"></i></a>
        {% endif %}
      </div>
    </div>
    {% endfor %}
    </div>
    {% else %}
    <div class="empty-state" style="padding:24px"><i class="fa fa-drafting-compass"></i><p>Nessun disegno caricato.</p></div>
    {% endif %}
  </div>
</div>

<!-- ── Pass ── -->
<div class="card" style="margin-bottom:20px">
  <div class="card-header">
    <h3><i class="fa fa-id-badge" style="color:#0891b2"></i> Pass</h3>
    {% if session.ruolo=='admin' %}
    <button onclick="document.getElementById('modal-pass').style.display='flex'" class="btn btn-sm btn-primary">
      <i class="fa fa-upload"></i> Carica
    </button>
    {% endif %}
  </div>
  <div class="card-body" style="padding:14px">
    {% set pass_files = files_evento | selectattr('categoria','equalto','pass') | list %}
    {% if pass_files %}
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:10px">
    {% for f in pass_files %}
    <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:10px;padding:12px 14px;display:flex;align-items:center;gap:12px">
      <div style="width:40px;height:40px;background:#e0f2fe;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0">🎫</div>
      <div style="flex:1;min-width:0">
        <div style="font-weight:600;font-size:13px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{{ f.nome_originale }}</div>
        {% if f.dipendente_id %}
          {% for m in team %}{% if m.uid == f.dipendente_id %}
          <div style="font-size:11px;font-weight:700;color:#0369a1;margin-top:3px">
            <i class="fa fa-user"></i> {{ m.nome }} {{ m.cognome }}
          </div>
          {% endif %}{% endfor %}
        {% else %}
          <div style="font-size:11px;color:#64748b;margin-top:3px"><i class="fa fa-users"></i> Visibile a tutti</div>
        {% endif %}
        {% if f.note %}<div style="font-size:11px;color:var(--text-muted)">{{ f.note }}</div>{% endif %}
        <div style="font-size:10px;color:var(--text-muted);margin-top:2px">{{ (f.dimensione/1024)|round(0)|int }} KB · {{ f.caricato_il[:10] }}</div>
      </div>
      <div style="display:flex;flex-direction:column;gap:4px">
        <a href="/calendario/{{ ev.id }}/files/{{ f.id }}/scarica" class="btn btn-sm btn-blue" title="Scarica"><i class="fa fa-download"></i></a>
        {% if session.ruolo=='admin' %}
        <a href="/calendario/{{ ev.id }}/files/{{ f.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare?')" title="Elimina"><i class="fa fa-trash"></i></a>
        {% endif %}
      </div>
    </div>
    {% endfor %}
    </div>
    {% else %}
    <div class="empty-state" style="padding:24px"><i class="fa fa-id-badge"></i><p>Nessun pass caricato.<br><small>Carica i pass inviati dal cliente.</small></p></div>
    {% endif %}
  </div>
</div>

{% if session.ruolo=='admin' %}
<!-- Avviso documenti mancanti -->
{% set n_ci_mancanti = team | selectattr('ha_ci', 'equalto', False) | list | length %}
{% set n_lib_mancanti = veicoli_evento | selectattr('ha_libretto', 'equalto', False) | list | length %}
{% if n_ci_mancanti > 0 or n_lib_mancanti > 0 %}
<div class="alert alert-error" style="margin-bottom:20px">
  <i class="fa fa-exclamation-triangle"></i>
  <div>
    <strong>Documenti mancanti per il pacchetto pass!</strong>
    {% if n_ci_mancanti > 0 %}<div>• {{ n_ci_mancanti }} dipendente/i senza carta d'identità caricata</div>{% endif %}
    {% if n_lib_mancanti > 0 %}<div>• {{ n_lib_mancanti }} veicolo/i senza libretto caricato</div>{% endif %}
    <div style="margin-top:6px;font-size:12px">Carica i documenti mancanti nelle schede dipendenti/veicoli prima di generare il pacchetto pass.</div>
  </div>
</div>
{% endif %}

<!-- Modal assegna dipendente -->
<div id="modal-team" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9999;align-items:center;justify-content:center">
  <div class="card" style="width:460px;max-width:95vw">
    <div class="card-header">
      <h3><i class="fa fa-user-plus" style="color:var(--accent2)"></i> Assegna dipendente</h3>
      <button onclick="document.getElementById('modal-team').style.display='none'" class="btn btn-secondary btn-sm"><i class="fa fa-times"></i></button>
    </div>
    <div class="card-body">
      <form method="POST" action="/calendario/{{ ev.id }}/assegna">
        <div class="form-group">
          <label>Dipendente</label>
          <select name="dipendente_id" required>
            <option value="">— Seleziona —</option>
            {% for d in dipendenti_disponibili %}<option value="{{ d.id }}">{{ d.nome }} {{ d.cognome }}</option>{% endfor %}
          </select>
        </div>
        <div class="form-group">
          <label>Ruolo / Servizio</label>
          <select name="ruolo">
            <option value="Montaggio">🔧 Montaggio</option>
            <option value="Assistenza fiera">🏛️ Assistenza fiera</option>
            <option value="Smontaggio">🔩 Smontaggio</option>
            <option value="Logistica">🚚 Logistica</option>
            <option value="Supervisore">👷 Supervisore</option>
            <option value="Generico">📌 Generico</option>
          </select>
        </div>
        <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:10px;padding:14px;margin-bottom:12px">
          <div style="font-size:12px;font-weight:700;color:#0369a1;margin-bottom:10px"><i class="fa fa-calendar-days"></i> Periodo assegnazione <span style="font-weight:400;color:#64748b">(lascia vuoto = tutto l'evento)</span></div>
          <div class="form-row">
            <div class="form-group" style="margin-bottom:0">
              <label>Dal</label>
              <input type="date" name="data_da" min="{{ ev.data_inizio }}" max="{{ ev.data_fine or ev.data_inizio }}" value="{{ ev.data_inizio }}">
            </div>
            <div class="form-group" style="margin-bottom:0">
              <label>Al</label>
              <input type="date" name="data_a" min="{{ ev.data_inizio }}" max="{{ ev.data_fine or ev.data_inizio }}" value="{{ ev.data_fine or ev.data_inizio }}">
            </div>
          </div>
        </div>
        <div style="display:flex;gap:10px;justify-content:flex-end">
          <button type="button" onclick="document.getElementById('modal-team').style.display='none'" class="btn btn-secondary">Annulla</button>
          <button type="submit" class="btn btn-primary"><i class="fa fa-check"></i> Assegna</button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Modal carica disegno -->
<div id="modal-disegno" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9999;align-items:center;justify-content:center">
  <div class="card" style="width:480px;max-width:95vw">
    <div class="card-header">
      <h3><i class="fa fa-drafting-compass" style="color:#6d28d9"></i> Carica Disegno/Progetto</h3>
      <button onclick="document.getElementById('modal-disegno').style.display='none'" class="btn btn-secondary btn-sm"><i class="fa fa-times"></i></button>
    </div>
    <div class="card-body">
      <form method="POST" action="/calendario/{{ ev.id }}/files/upload" enctype="multipart/form-data">
        <input type="hidden" name="categoria" value="disegno">
        <div class="form-group">
          <label>File (PDF, DWG, DXF, immagini...)</label>
          <input type="file" name="file" required accept=".pdf,.dwg,.dxf,.png,.jpg,.jpeg,.gif,.webp,.zip,.rar">
        </div>
        <div class="form-group">
          <label>Note (opzionale)</label>
          <input name="note" placeholder="Es. Pianta piano terra, vista frontale...">
        </div>
        <div style="display:flex;gap:10px;justify-content:flex-end">
          <button type="button" onclick="document.getElementById('modal-disegno').style.display='none'" class="btn btn-secondary">Annulla</button>
          <button type="submit" class="btn btn-primary"><i class="fa fa-upload"></i> Carica</button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Modal carica pass -->
<div id="modal-pass" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9999;align-items:center;justify-content:center">
  <div class="card" style="width:480px;max-width:95vw">
    <div class="card-header">
      <h3><i class="fa fa-id-badge" style="color:#0891b2"></i> Carica Pass</h3>
      <button onclick="document.getElementById('modal-pass').style.display='none'" class="btn btn-secondary btn-sm"><i class="fa fa-times"></i></button>
    </div>
    <div class="card-body">
      <form method="POST" action="/calendario/{{ ev.id }}/files/upload" enctype="multipart/form-data">
        <input type="hidden" name="categoria" value="pass">
        <div class="form-group">
          <label>File pass (PDF, immagine...)</label>
          <input type="file" name="file" required accept=".pdf,.png,.jpg,.jpeg,.gif,.webp">
        </div>
        <div class="form-group">
          <label><i class="fa fa-user" style="color:#0891b2"></i> Assegna a dipendente <span style="color:#94a3b8;font-weight:400;font-size:12px">(solo quel dipendente lo vedrà sul mobile)</span></label>
          <select name="dipendente_id" class="form-control">
            <option value="">— Visibile a tutti —</option>
            {% for m in team %}
            <option value="{{ m.uid }}">{{ m.nome }} {{ m.cognome }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="form-group">
          <label>Note (opzionale)</label>
          <input name="note" placeholder="Es. Pass ingresso area espositiva...">
        </div>
        <div style="display:flex;gap:10px;justify-content:flex-end">
          <button type="button" onclick="document.getElementById('modal-pass').style.display='none'" class="btn btn-secondary">Annulla</button>
          <button type="submit" class="btn btn-primary"><i class="fa fa-upload"></i> Carica</button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Modal assegna veicolo -->
<div id="modal-veicolo" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9999;align-items:center;justify-content:center">
  <div class="card" style="width:440px;max-width:95vw">
    <div class="card-header">
      <h3><i class="fa fa-truck" style="color:var(--warning)"></i> Aggiungi mezzo</h3>
      <button onclick="document.getElementById('modal-veicolo').style.display='none'" class="btn btn-secondary btn-sm"><i class="fa fa-times"></i></button>
    </div>
    <div class="card-body">
      <form method="POST" action="/calendario/{{ ev.id }}/assegna-veicolo">
        <div class="form-group">
          <label>Veicolo</label>
          <select name="veicolo_id" required>
            <option value="">— Seleziona mezzo —</option>
            {% for v in veicoli_disponibili %}
            <option value="{{ v.id }}">{{ v.targa }} — {{ v.marca or '' }} {{ v.modello or '' }} ({{ v.tipo or '' }})</option>
            {% endfor %}
          </select>
        </div>
        <div class="form-group">
          <label>Note (opzionale)</label>
          <input name="note" placeholder="Es. porta materiali montaggio">
        </div>
        <div style="display:flex;gap:10px;justify-content:flex-end">
          <button type="button" onclick="document.getElementById('modal-veicolo').style.display='none'" class="btn btn-secondary">Annulla</button>
          <button type="submit" class="btn btn-primary"><i class="fa fa-check"></i> Aggiungi</button>
        </div>
      </form>
    </div>
  </div>
</div>
{% endif %}


<!-- Spese extra evento -->
<div class="card" style="margin-top:20px">
  <div class="card-header" style="display:flex;align-items:center;justify-content:space-between">
    <h3><i class="fa fa-receipt" style="color:#f59e0b"></i> Spese Extra</h3>
    {% if session.ruolo=='admin' %}
    <button onclick="document.getElementById('modal-spesa').style.display='flex'" class="btn btn-sm" style="background:#f59e0b;color:#fff"><i class="fa fa-plus"></i> Aggiungi</button>
    {% endif %}
  </div>
  <div class="card-body" style="padding:0">
    {% if spese_evento %}
    <table style="width:100%;font-size:13px">
      <thead><tr style="background:#fef3c7">
        <th style="padding:10px 14px;text-align:left">Categoria</th>
        <th style="padding:10px 14px;text-align:left">Descrizione</th>
        <th style="padding:10px 14px;text-align:left">Data</th>
        <th style="padding:10px 14px;text-align:right">Importo</th>
        {% if session.ruolo=='admin' %}<th style="padding:10px 14px"></th>{% endif %}
      </tr></thead>
      <tbody>
      {% for s in spese_evento %}
      <tr style="border-bottom:1px solid #fef3c7">
        <td style="padding:9px 14px"><span style="background:#fef3c7;color:#92400e;padding:3px 9px;border-radius:99px;font-size:11px;font-weight:700">{{ s.categoria }}</span></td>
        <td style="padding:9px 14px">{{ s.descrizione or '–' }}</td>
        <td style="padding:9px 14px;color:#64748b;font-size:12px">{{ s.data or '–' }}</td>
        <td style="padding:9px 14px;text-align:right;font-weight:700">€ {{ "%.2f"|format(s.importo) }}</td>
        {% if session.ruolo=='admin' %}
        <td style="padding:9px 14px">
          <a href="/calendario/{{ ev.id }}/spese/{{ s.id }}/elimina" onclick="return confirm('Eliminare?')" style="color:#dc2626;font-size:12px"><i class="fa fa-trash"></i></a>
        </td>{% endif %}
      </tr>
      {% endfor %}
      </tbody>
      <tfoot><tr style="background:#fef3c7">
        <td colspan="3" style="padding:10px 14px;font-weight:700;text-align:right">TOTALE SPESE EXTRA</td>
        <td style="padding:10px 14px;text-align:right;font-weight:700;color:#d97706">€ {{ "%.2f"|format(spese_evento|sum(attribute='importo')) }}</td>
        {% if session.ruolo=='admin' %}<td></td>{% endif %}
      </tr></tfoot>
    </table>
    {% else %}
    <div style="padding:20px;text-align:center;color:#94a3b8;font-size:13px"><i class="fa fa-receipt" style="margin-right:6px"></i>Nessuna spesa extra registrata</div>
    {% endif %}
  </div>
</div>

<!-- Modal aggiungi spesa -->
<div id="modal-spesa" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:9999;align-items:center;justify-content:center" onclick="if(event.target===this)this.style.display='none'">
  <div style="background:#fff;border-radius:14px;padding:28px;width:420px;max-width:96vw">
    <h3 style="margin:0 0 18px;font-size:16px"><i class="fa fa-receipt" style="color:#f59e0b;margin-right:8px"></i>Nuova spesa extra</h3>
    <form method="POST" action="/calendario/{{ ev.id }}/spese/aggiungi">
      <div class="form-group">
        <label>Categoria</label>
        <select name="categoria" required>
          <option value="Hotel">🏨 Hotel</option>
          <option value="Gasolio">⛽ Gasolio</option>
          <option value="Autostrada">🛣️ Autostrada</option>
          <option value="Aereo">✈️ Aereo</option>
          <option value="Treno">🚂 Treno</option>
          <option value="Cibo">🍽️ Cibo</option>
          <option value="Noleggio furgone">🚐 Noleggio furgone</option>
          <option value="Noleggio attrezzatura">🔧 Noleggio attrezzatura</option>
          <option value="Parcheggio">🅿️ Parcheggio</option>
          <option value="Altro">📦 Altro</option>
        </select>
      </div>
      <div class="form-group"><label>Descrizione</label><input name="descrizione" placeholder="Es. Hotel Novotel 2 notti"></div>
      <div class="form-row">
        <div class="form-group"><label>Importo €</label><input type="number" name="importo" step="0.01" min="0" required placeholder="0.00"></div>
        <div class="form-group"><label>Data</label><input type="date" name="data" value="{{ ev.data_inizio }}"></div>
      </div>
      <div class="form-group"><label>Nota</label><input name="nota" placeholder="Facoltativa"></div>
      <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:12px">
        <button type="button" onclick="document.getElementById('modal-spesa').style.display='none'" class="btn btn-secondary">Annulla</button>
        <button type="submit" class="btn btn-primary" style="background:#f59e0b;border-color:#f59e0b"><i class="fa fa-save"></i> Salva</button>
      </div>
    </form>
  </div>
</div>"""

CAL_FORM_TMPL = """
<div class="card" style="max-width:640px;margin:0 auto">
  <div class="card-header"><h3>{{ 'Modifica' if ev else 'Nuovo' }} Evento</h3></div>
  <div class="card-body">
    <form method="POST">
      <div class="form-row">
        <div class="form-group">
          <label>Titolo *</label>
          <input name="titolo" required value="{{ ev.titolo if ev else request.args.get('titolo','') }}" placeholder="Es. Fiera del Mobile Milano">
        </div>
        <div class="form-group">
          <label>Tipo servizio</label>
          <select name="tipo">
            <option value="montaggio"  {{ 'selected' if ev and ev.tipo=='montaggio' }}>🔧 Montaggio</option>
            <option value="fiera"      {{ 'selected' if ev and ev.tipo=='fiera' }}>🏛️ Fiera</option>
            <option value="smontaggio" {{ 'selected' if ev and ev.tipo=='smontaggio' }}>🔩 Smontaggio</option>
            <option value="altro"      {{ 'selected' if ev and ev.tipo=='altro' }}>📌 Altro</option>
          </select>
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>Luogo / Sede fiera</label>
          <input name="luogo" value="{{ ev.luogo if ev else request.args.get('luogo','') }}" placeholder="Es. Fiera di Verona — Pad. 4">
        </div>
        <div class="form-group">
          <label>Cantiere collegato (opzionale)</label>
          <select name="cantiere_id">
            <option value="">— Nessuna fiera —</option>
            {% for c in cantieri %}
            <option value="{{ c.id }}" {{ 'selected' if ev and ev.cantiere_id==c.id }}>{{ c.nome }}{% if c.indirizzo %} — {{ c.indirizzo }}{% endif %}</option>
            {% endfor %}
          </select>
        </div>
      </div>
      <div class="form-row">
        <div class="form-group"><label>Data inizio *</label><input type="date" name="data_inizio" required value="{{ ev.data_inizio if ev else request.args.get('data_inizio','') }}"></div>
        <div class="form-group"><label>Data fine</label><input type="date" name="data_fine" value="{{ ev.data_fine if ev else request.args.get('data_fine','') }}"></div>
      </div>
      <div class="form-group"><label>Note / Descrizione</label><textarea name="descrizione" rows="2" placeholder="Istruzioni, materiali, contatti...">{{ ev.descrizione if ev else '' }}</textarea></div>

      <!-- Assegna dipendenti subito -->
      <div class="form-group">
        <label>Assegna dipendenti (opzionale)</label>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;max-height:200px;overflow-y:auto;padding:10px;background:#f8fafc;border-radius:8px;border:1px solid var(--border)">
          {% for d in dipendenti %}
          <label style="display:flex;align-items:center;gap:8px;cursor:pointer;padding:6px;border-radius:6px;font-size:13px;text-transform:none;letter-spacing:0;color:var(--text);font-weight:500" onmouseover="this.style.background='#f1f5f9'" onmouseout="this.style.background=''">
            <input type="checkbox" name="dipendenti_ids" value="{{ d.id }}" style="width:16px;height:16px;accent-color:var(--accent2)">
            <div style="width:26px;height:26px;border-radius:50%;background:var(--accent2);display:inline-flex;align-items:center;justify-content:center;color:#fff;font-size:10px;font-weight:700;flex-shrink:0">{{ (d.nome[0]+d.cognome[0])|upper }}</div>
            {{ d.nome }} {{ d.cognome }}
          </label>
          {% endfor %}
        </div>
      </div>
      <div class="form-group">
        <label>Ruolo per i dipendenti selezionati</label>
        <select name="ruolo_default">
          <option value="Montaggio">🔧 Montaggio</option>
          <option value="Assistenza fiera">🏛️ Assistenza fiera</option>
          <option value="Smontaggio">🔩 Smontaggio</option>
          <option value="Logistica">🚚 Logistica</option>
          <option value="Supervisore">👷 Supervisore</option>
          <option value="Generico">📌 Generico</option>
        </select>
      </div>

      {% if preventivi_acc %}
      <div class="form-group" style="background:#eef4ff;border:1px solid #bfdbfe;border-radius:10px;padding:14px;margin-bottom:12px">
        <label style="color:#1d4ed8;font-weight:700"><i class="fa fa-file-invoice" style="margin-right:6px"></i>Collega a preventivo confermato (opzionale)</label>
        <select name="preventivo_id" id="prev_sel" onchange="caricaDatiPreventivo(this)" data-presel="{{ request.args.get('prev_id','') }}" data-tipo="{{ request.args.get('tipo','') }}">
          <option value="">— Nessun preventivo collegato —</option>
          {% for p in preventivi_acc %}
          <option value="{{ p.id }}"
            data-titolo="{{ p.oggetto or '' }}"
            data-inizio="{{ p.data_inizio_lavoro or '' }}"
            data-fine="{{ p.data_fine_lavoro or '' }}"
            data-luogo="{{ p.luogo_lavoro or '' }}"
            {{ 'selected' if ev and ev.preventivo_id == p.id }}>
            N.{{ p.numero }} — {{ p.cliente_nome or '' }} — {{ p.oggetto or '' }}
            {% if p.data_inizio_lavoro %} ({{ p.data_inizio_lavoro }}){% endif %}
          </option>
          {% endfor %}
        </select>
        <small style="color:#3b82f6;margin-top:4px;display:block">Selezionando un preventivo, titolo, date e luogo vengono compilati automaticamente.</small>
      </div>
      {% endif %}

      <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:8px">
        <a href="/calendario" class="btn btn-secondary">Annulla</a>
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva evento</button>
      </div>
    </form>
  </div>
</div>
<script>
function caricaDatiPreventivo(sel) {
  var opt = sel.options[sel.selectedIndex];
  if (!opt.value) return;
  var titolo = opt.dataset.titolo;
  var inizio = opt.dataset.inizio;
  var fine   = opt.dataset.fine;
  var luogo  = opt.dataset.luogo;
  if (titolo) document.querySelector('[name=titolo]').value = titolo;
  if (inizio) document.querySelector('[name=data_inizio]').value = inizio;
  if (fine)   document.querySelector('[name=data_fine]').value   = fine;
  if (luogo)  document.querySelector('[name=luogo]').value       = luogo;
}
document.addEventListener('DOMContentLoaded', function() {
  var sel = document.getElementById('prev_sel');
  if (!sel) return;
  var presel = sel.dataset.presel;
  if (!presel) return;
  for (var i = 0; i < sel.options.length; i++) {
    if (sel.options[i].value == presel) {
      sel.selectedIndex = i;
      caricaDatiPreventivo(sel);
      break;
    }
  }
  // pre-select tipo se passato via URL
  var tipoArg = sel.dataset.tipo;
  if (tipoArg) {
    var tipoSel = document.querySelector('[name=tipo]');
    if (tipoSel) tipoSel.value = tipoArg;
  }
});
</script>"""


@app.route('/calendario')
@login_required
def calendario():
    import calendar as cal_mod
    mesi_it=['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno','Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']
    today=date.today(); m=int(request.args.get('m',today.month)); y=int(request.args.get('y',today.year))
    db=get_db()
    mese_str = f'{y:04d}-{m:02d}'
    mese_inizio = f'{y:04d}-{m:02d}-01'
    import calendar as _cal2
    ultimo_giorno = _cal2.monthrange(y,m)[1]
    mese_fine = f'{y:04d}-{m:02d}-{ultimo_giorno:02d}'

    # Colore automatico dal tipo
    eventi_raw = db.execute(
        "SELECT * FROM eventi WHERE data_inizio <= ? AND COALESCE(data_fine, data_inizio) >= ? ORDER BY data_inizio",
        (mese_fine, mese_inizio)
    ).fetchall()

    # Arricchisci con colore tipo
    def _enrich(e):
        ed = dict(e)
        tipo = ed.get('tipo') or 'altro'
        ed['colore'] = TIPO_COLORI_MAP.get(tipo, ed.get('colore') or '#8b5cf6')
        ed['tipo_label'] = TIPO_LABEL_MAP.get(tipo, '📌 Altro')
        return ed

    eventi_raw = [_enrich(e) for e in eventi_raw]

    # Lista mese con team
    eventi_mese_raw = db.execute(
        "SELECT * FROM eventi WHERE strftime('%Y-%m',data_inizio)=? ORDER BY data_inizio", (mese_str,)
    ).fetchall()
    eventi_mese = []
    for e in eventi_mese_raw:
        ed = _enrich(e)
        team_raw = db.execute("""
            SELECT u.nome, u.cognome, ed.id as assegn_id, ed.ruolo
            FROM eventi_dipendenti ed JOIN utenti u ON u.id=ed.dipendente_id
            WHERE ed.evento_id=? ORDER BY ed.ruolo, u.nome
        """, (e['id'],)).fetchall()
        ed['team'] = [dict(t, iniziali=(t['nome'][0]+t['cognome'][0]).upper()) for t in team_raw]
        eventi_mese.append(ed)

    # Mappa giorno → eventi
    ev_map = {}
    for e in eventi_raw:
        d_start = datetime.strptime(e['data_inizio'], '%Y-%m-%d').date()
        d_end   = datetime.strptime(e['data_fine'] or e['data_inizio'], '%Y-%m-%d').date()
        cur = d_start
        while cur <= d_end:
            if cur.year == y and cur.month == m:
                ev_map.setdefault(cur.day, []).append(e)
            cur += timedelta(days=1)

    first_wd,days_in=cal_mod.monthrange(y,m)
    prev_year,prev_month=(y,m-1) if m>1 else (y-1,12)
    _,days_prev=cal_mod.monthrange(prev_year,prev_month)
    cells=[{'day':days_prev-first_wd+1+i,'cur_month':False,'today':False,'eventi':[]} for i in range(first_wd)]
    for d in range(1,days_in+1):
        cells.append({'day':d,'cur_month':True,'today':(d==today.day and m==today.month and y==today.year),'eventi':ev_map.get(d,[])})
    for d in range(1,43-len(cells)):
        cells.append({'day':d,'cur_month':False,'today':False,'eventi':[]})
    prev_m,prev_y=(m-1,y) if m>1 else (12,y-1); next_m,next_y=(m+1,y) if m<12 else (1,y+1)
    db.close()
    return render_page(CAL_TMPL,page_title='Calendario',active='calendario',
        calendario=cells,mese_nome=mesi_it[m],anno=y,m=m,y=y,
        prev_m=prev_m,prev_y=prev_y,next_m=next_m,next_y=next_y,eventi_mese=eventi_mese)


@app.route('/calendario/pdf')
@admin_required
def calendario_pdf():
    import calendar as _cal
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io as _io

    today = date.today()
    # Genera sempre 12 mesi a partire dal mese corrente
    start_m = int(request.args.get('m', today.month))
    start_y = int(request.args.get('y', today.year))

    az_nome = get_setting('azienda', 'Accesso Fiere S.r.l.')

    db = get_db()
    # Carica tutti gli eventi dei prossimi 12 mesi
    mesi_it = ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
               'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']
    giorni_it = ['Lun','Mar','Mer','Gio','Ven','Sab','Dom']

    # Colori tipi
    CMAP = {
        'montaggio':  colors.HexColor('#16a34a'),
        'fiera':      colors.HexColor('#2196F3'),
        'smontaggio': colors.HexColor('#e63946'),
        'altro':      colors.HexColor('#8b5cf6'),
    }
    LMAP = {
        'montaggio':  'Montaggio',
        'fiera':      'Fiera/Evento',
        'smontaggio': 'Smontaggio',
        'altro':      'Altro',
    }

    BLU   = colors.HexColor('#0f172a')
    WHITE = colors.white
    GRAY  = colors.HexColor('#f8fafc')
    GRAY2 = colors.HexColor('#e2e8f0')
    WEEKEND = colors.HexColor('#f1f5f9')
    OGGI_C  = colors.HexColor('#fef9c3')

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)

    def P(txt, font='Helvetica', size=8, color=BLU, align=TA_LEFT, bg=None):
        s = ParagraphStyle('x', fontName=font, fontSize=size, textColor=color,
                           alignment=align, leading=size*1.35,
                           backColor=bg if bg else colors.white)
        return Paragraph(str(txt), s)

    story = []

    for i in range(12):
        m = ((start_m - 1 + i) % 12) + 1
        y = start_y + ((start_m - 1 + i) // 12)

        mese_inizio = f'{y:04d}-{m:02d}-01'
        ultimo_g    = _cal.monthrange(y, m)[1]
        mese_fine   = f'{y:04d}-{m:02d}-{ultimo_g:02d}'

        eventi_raw = db.execute(
            "SELECT * FROM eventi WHERE data_inizio <= ? AND COALESCE(data_fine,data_inizio) >= ? ORDER BY data_inizio",
            (mese_fine, mese_inizio)
        ).fetchall()

        # Mappa giorno → lista eventi
        ev_map = {}
        for e in eventi_raw:
            ev = dict(e)
            tipo = ev.get('tipo') or 'altro'
            ev['colore_rl'] = CMAP.get(tipo, colors.HexColor('#8b5cf6'))
            ev['label']     = LMAP.get(tipo, 'Altro')
            try:
                d1 = date.fromisoformat(ev['data_inizio'])
                d2 = date.fromisoformat(ev.get('data_fine') or ev['data_inizio'])
            except:
                d1 = d2 = date(y, m, 1)
            for dd in range((d2 - d1).days + 1):
                g = (d1 + timedelta(days=dd)).day
                if date(y, m, 1) <= d1 + timedelta(days=dd) <= date(y, m, ultimo_g):
                    ev_map.setdefault(g, []).append(ev)

        # ── Intestazione mese ──
        hdr = Table([[
            P(f'{mesi_it[m].upper()} {y}', 'Helvetica-Bold', 14, BLU, TA_LEFT),
            P(az_nome, 'Helvetica', 9, colors.HexColor('#64748b'), TA_LEFT),
            # Legenda inline
            P('', 'Helvetica', 8, BLU, TA_LEFT),
        ]], colWidths=[80*mm, 80*mm, None])
        hdr.setStyle(TableStyle([
            ('LINEBELOW', (0,0), (-1,0), 1.5, BLU),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(hdr)
        story.append(Spacer(1, 2*mm))

        # ── Legenda ──
        legenda_data = [[
            P('■', 'Helvetica-Bold', 11, CMAP['montaggio'], TA_CENTER),
            P('Montaggio', 'Helvetica', 8, BLU),
            P('■', 'Helvetica-Bold', 11, CMAP['fiera'], TA_CENTER),
            P('Fiera/Evento', 'Helvetica', 8, BLU),
            P('■', 'Helvetica-Bold', 11, CMAP['smontaggio'], TA_CENTER),
            P('Smontaggio', 'Helvetica', 8, BLU),
            P('■', 'Helvetica-Bold', 11, CMAP['altro'], TA_CENTER),
            P('Altro', 'Helvetica', 8, BLU),
        ]]
        leg = Table(legenda_data, colWidths=[6*mm,28*mm,6*mm,30*mm,6*mm,28*mm,6*mm,20*mm])
        leg.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('LEFTPADDING',(0,0),(-1,-1),2),
            ('RIGHTPADDING',(0,0),(-1,-1),2),
        ]))
        story.append(leg)
        story.append(Spacer(1, 3*mm))

        # ── Griglia calendario — 7 colonne ──
        W = 267*mm  # larghezza utile landscape A4
        col_w = [W/7]*7

        # Header giorni
        header_row = []
        for gi, gn in enumerate(giorni_it):
            bg = colors.HexColor('#e2e8f0') if gi >= 5 else BLU
            fc = BLU if gi >= 5 else WHITE
            header_row.append(P(gn, 'Helvetica-Bold', 8, fc, TA_CENTER))

        # Prima cella del mese
        first_dow = _cal.monthrange(y, m)[0]  # 0=Lun
        rows = [header_row]
        week = [''] * first_dow
        for g in range(1, ultimo_g + 1):
            week.append(g)
            if len(week) == 7:
                rows.append(week)
                week = []
        if week:
            week += [''] * (7 - len(week))
            rows.append(week)

        # Costruisci tabella celle
        cell_height = 18*mm
        tdata = [rows[0]]  # header già fatto sopra
        tstyles = [
            ('BACKGROUND', (0,0), (-1,0), BLU),
            ('TEXTCOLOR',  (0,0), (-1,0), WHITE),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',     (0,0), (-1,-1), 'TOP'),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,0), 8),
            ('BOX',        (0,0), (-1,-1), 0.5, GRAY2),
            ('INNERGRID',  (0,0), (-1,-1), 0.3, GRAY2),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ]

        for ri, week_row in enumerate(rows[1:], 1):
            row_cells = []
            for ci, cell in enumerate(week_row):
                if cell == '':
                    row_cells.append('')
                    continue
                g = int(cell)
                is_weekend = ci >= 5
                is_today   = (g == today.day and m == today.month and y == today.year)

                # Colore sfondo cella
                if is_today:
                    tstyles.append(('BACKGROUND', (ci,ri), (ci,ri), OGGI_C))
                elif is_weekend:
                    tstyles.append(('BACKGROUND', (ci,ri), (ci,ri), WEEKEND))

                # Numero giorno + eventi
                evs = ev_map.get(g, [])
                lines = [P(str(g), 'Helvetica-Bold', 8,
                           colors.HexColor('#64748b') if is_weekend else BLU, TA_LEFT)]
                for ev in evs[:3]:  # max 3 eventi per cella
                    lines.append(P(ev['titolo'], 'Helvetica-Bold', 7, WHITE, TA_LEFT, bg=ev['colore_rl']))
                if len(evs) > 3:
                    lines.append(P(f'+{len(evs)-3} altri', 'Helvetica', 6,
                                   colors.HexColor('#64748b'), TA_LEFT))

                from reportlab.platypus import KeepInFrame
                kif = KeepInFrame(col_w[ci]-6, cell_height-6, lines, mode='shrink')
                row_cells.append(kif)

            tdata.append(row_cells)
            row_heights = [7*mm] + [cell_height] * (len(tdata)-1)

        tab = Table(tdata, colWidths=col_w, rowHeights=row_heights)
        tab.setStyle(TableStyle(tstyles))
        story.append(tab)

        if i < 11:
            story.append(PageBreak())

    db.close()
    doc.build(story)
    buf.seek(0)
    return Response(buf.read(), mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="Calendario_{start_y}.pdf"'})


@app.route('/calendario/nuovo', methods=['GET','POST'])
@admin_required
def evento_nuovo():
    db = get_db()
    if request.method == 'POST':
        tipo = request.form.get('tipo','altro')
        colore = TIPO_COLORI_MAP.get(tipo, '#8b5cf6')
        cantiere_id = request.form.get('cantiere_id') or None
        prev_id = request.form.get('preventivo_id') or None
        cur = db.execute(
            "INSERT INTO eventi (titolo,descrizione,data_inizio,data_fine,colore,tipo,luogo,cantiere_id,creato_da,preventivo_id) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (request.form['titolo'], request.form.get('descrizione'),
             request.form['data_inizio'],
             request.form.get('data_fine') or request.form['data_inizio'],
             colore, tipo, request.form.get('luogo',''), cantiere_id, session['user_id'], prev_id)
        )
        eid = cur.lastrowid
        # Assegna dipendenti selezionati
        ruolo_def = request.form.get('ruolo_default','Generico')
        for did in request.form.getlist('dipendenti_ids'):
            try:
                db.execute("INSERT INTO eventi_dipendenti (evento_id,dipendente_id,ruolo) VALUES (?,?,?)", (eid, int(did), ruolo_def))
            except: pass
        safe_commit(db); db.close()
        flash('Evento aggiunto!','success')
        return redirect(url_for('evento_dettaglio', eid=eid))
    dipendenti  = db.execute("SELECT id,nome,cognome FROM utenti WHERE ruolo='dipendente' AND attivo=1 ORDER BY cognome,nome").fetchall()
    cantieri    = db.execute("SELECT id,nome,indirizzo FROM cantieri WHERE attivo=1 ORDER BY nome").fetchall()
    preventivi_acc = db.execute(
        "SELECT id,numero,oggetto,cliente_nome,data_inizio_lavoro,data_fine_lavoro,luogo_lavoro "
        "FROM preventivi WHERE stato='accettato' ORDER BY creato_il DESC").fetchall()
    db.close()
    return render_page(CAL_FORM_TMPL, page_title='Nuovo Evento', active='calendario',
        ev=None, dipendenti=dipendenti, cantieri=cantieri, preventivi_acc=preventivi_acc)


@app.route('/calendario/<int:eid>')
@login_required
def evento_dettaglio(eid):
    db = get_db()
    ev_raw = db.execute("SELECT * FROM eventi WHERE id=?", (eid,)).fetchone()
    if not ev_raw:
        db.close(); return redirect(url_for('calendario'))
    ev = dict(ev_raw)
    tipo = ev.get('tipo') or 'altro'
    ev['colore'] = TIPO_COLORI_MAP.get(tipo, ev.get('colore') or '#8b5cf6')
    ev['tipo_label'] = TIPO_LABEL_MAP.get(tipo, '📌 Altro')

    # ── Team ──
    team_raw = db.execute("""
        SELECT u.id as uid, u.nome, u.cognome, ed.id as assegn_id, ed.ruolo,
               ed.data_da, ed.data_a
        FROM eventi_dipendenti ed JOIN utenti u ON u.id=ed.dipendente_id
        WHERE ed.evento_id=? ORDER BY ed.ruolo, u.nome
    """, (eid,)).fetchall()

    team = []
    for t in team_raw:
        td = dict(t)
        td['iniziali'] = (t['nome'][0]+t['cognome'][0]).upper()
        # Controlla se ha carta d'identità caricata
        ci = db.execute("""SELECT id FROM documenti_dipendente
            WHERE utente_id=? AND (tipo_doc LIKE '%identit%' OR tipo_doc LIKE '%CI%' OR tipo_doc LIKE '%carta%' OR tipo_doc LIKE '%passaporto%')
            LIMIT 1""", (t['uid'],)).fetchone()
        td['ha_ci'] = ci is not None
        team.append(td)

    team_per_ruolo = {}
    for t in team:
        team_per_ruolo.setdefault(t['ruolo'], []).append(t)

    # ── Dipendenti disponibili ──
    assegnati_ids = [row['dipendente_id'] for row in
        db.execute("SELECT dipendente_id FROM eventi_dipendenti WHERE evento_id=?", (eid,)).fetchall()]
    dipendenti_disponibili = db.execute(
        "SELECT id,nome,cognome FROM utenti WHERE ruolo='dipendente' AND attivo=1 AND id NOT IN ({}) ORDER BY cognome,nome".format(
            ','.join(['?']*len(assegnati_ids)) if assegnati_ids else '0'),
        assegnati_ids if assegnati_ids else []).fetchall()

    # ── Veicoli assegnati ──
    veicoli_raw = db.execute("""
        SELECT v.*, ev.id as ev_veicolo_id, ev.note as ev_note
        FROM eventi_veicoli ev JOIN veicoli v ON v.id=ev.veicolo_id
        WHERE ev.evento_id=? ORDER BY v.targa
    """, (eid,)).fetchall()

    veicoli_evento = []
    for v in veicoli_raw:
        vd = dict(v)
        lib = db.execute("""SELECT id FROM documenti_veicoli
            WHERE veicolo_id=? AND (tipo_doc LIKE '%libretto%' OR tipo_doc LIKE '%circolaz%' OR tipo_doc LIKE '%Libretto%')
            LIMIT 1""", (v['id'],)).fetchone()
        vd['ha_libretto'] = lib is not None
        veicoli_evento.append(vd)

    # ── Veicoli disponibili (non già assegnati) ──
    assegnati_vid = [v['id'] for v in veicoli_raw]
    veicoli_disponibili = db.execute(
        "SELECT id,targa,marca,modello,tipo FROM veicoli WHERE attivo=1 AND id NOT IN ({}) ORDER BY targa".format(
            ','.join(['?']*len(assegnati_vid)) if assegnati_vid else '0'),
        assegnati_vid if assegnati_vid else []).fetchall()

    # ── Files (disegni + pass) ──
    files_evento = db.execute(
        "SELECT * FROM eventi_files WHERE evento_id=? ORDER BY categoria, caricato_il DESC", (eid,)
    ).fetchall()

    # ── Cantiere ──
    cantiere = None
    if ev.get('cantiere_id'):
        cantiere = db.execute("SELECT * FROM cantieri WHERE id=?", (ev['cantiere_id'],)).fetchone()

    # ── Preventivo collegato ──
    prev_collegato = None
    if ev.get('preventivo_id'):
        prev_collegato = db.execute("SELECT * FROM preventivi WHERE id=?", (ev['preventivo_id'],)).fetchone()

    # ── Spese extra ──
    spese_evento = db.execute(
        "SELECT * FROM spese_evento WHERE evento_id=? ORDER BY data, id", (eid,)).fetchall()

    db.close()
    return render_page(CAL_DETAIL_TMPL, page_title=ev['titolo'], active='calendario',
        ev=ev, team=team, team_per_ruolo=team_per_ruolo,
        dipendenti_disponibili=dipendenti_disponibili,
        veicoli_evento=veicoli_evento, veicoli_disponibili=veicoli_disponibili,
        files_evento=files_evento, cantiere=cantiere, prev_collegato=prev_collegato,
        spese_evento=spese_evento)


@app.route('/calendario/<int:eid>/spese/aggiungi', methods=['POST'])
@admin_required
def evento_spesa_aggiungi(eid):
    db = get_db()
    db.execute("""INSERT INTO spese_evento (evento_id,categoria,descrizione,importo,data,nota,creato_da)
        VALUES (?,?,?,?,?,?,?)""", (
        eid,
        request.form.get('categoria','Altro'),
        request.form.get('descrizione',''),
        float(request.form.get('importo',0) or 0),
        request.form.get('data',''),
        request.form.get('nota',''),
        session.get('user_id'),
    ))
    safe_commit(db); db.close()
    flash('Spesa aggiunta!','success')
    return redirect(url_for('evento_dettaglio', eid=eid))


@app.route('/calendario/<int:eid>/spese/<int:sid>/elimina')
@admin_required
def evento_spesa_elimina(eid, sid):
    db = get_db()
    db.execute("DELETE FROM spese_evento WHERE id=? AND evento_id=?", (sid, eid))
    safe_commit(db); db.close()
    flash('Spesa eliminata.','success')
    return redirect(url_for('evento_dettaglio', eid=eid))


@app.route('/calendario/<int:eid>/modifica', methods=['GET','POST'])
@admin_required
def evento_modifica(eid):
    db = get_db()
    if request.method == 'POST':
        tipo = request.form.get('tipo','altro')
        colore = TIPO_COLORI_MAP.get(tipo, '#8b5cf6')
        cantiere_id = request.form.get('cantiere_id') or None
        db.execute("""UPDATE eventi SET titolo=?,descrizione=?,data_inizio=?,data_fine=?,
                      colore=?,tipo=?,luogo=?,cantiere_id=? WHERE id=?""",
            (request.form['titolo'], request.form.get('descrizione'),
             request.form['data_inizio'], request.form.get('data_fine') or request.form['data_inizio'],
             colore, tipo, request.form.get('luogo',''), cantiere_id, eid))
        # Aggiungi nuovi dipendenti selezionati
        ruolo_def = request.form.get('ruolo_default','Generico')
        for did in request.form.getlist('dipendenti_ids'):
            try:
                db.execute("INSERT OR IGNORE INTO eventi_dipendenti (evento_id,dipendente_id,ruolo) VALUES (?,?,?)", (eid, int(did), ruolo_def))
            except: pass
        safe_commit(db); db.close()
        flash('Evento aggiornato!','success')
        return redirect(url_for('evento_dettaglio', eid=eid))
    ev = dict(db.execute("SELECT * FROM eventi WHERE id=?", (eid,)).fetchone())
    dipendenti  = db.execute("SELECT id,nome,cognome FROM utenti WHERE ruolo='dipendente' AND attivo=1 ORDER BY cognome,nome").fetchall()
    cantieri    = db.execute("SELECT id,nome,indirizzo FROM cantieri WHERE attivo=1 ORDER BY nome").fetchall()
    preventivi_acc = db.execute(
        "SELECT id,numero,oggetto,cliente_nome,data_inizio_lavoro,data_fine_lavoro,luogo_lavoro "
        "FROM preventivi WHERE stato='accettato' ORDER BY creato_il DESC").fetchall()
    db.close()
    return render_page(CAL_FORM_TMPL, page_title='Modifica Evento', active='calendario',
        ev=ev, dipendenti=dipendenti, cantieri=cantieri, preventivi_acc=preventivi_acc)


@app.route('/calendario/<int:eid>/assegna', methods=['POST'])
@admin_required
def evento_assegna(eid):
    did    = request.form.get('dipendente_id')
    ruolo  = request.form.get('ruolo', 'Generico')
    data_da = request.form.get('data_da') or None
    data_a  = request.form.get('data_a')  or None
    if did:
        db = get_db()
        try:
            ex_assegn = db.execute(
                "SELECT id FROM eventi_dipendenti WHERE evento_id=? AND dipendente_id=?",
                (eid, int(did))).fetchone()
            if ex_assegn:
                db.execute("UPDATE eventi_dipendenti SET ruolo=?,data_da=?,data_a=? WHERE id=?",
                           (ruolo, data_da, data_a, ex_assegn['id']))
            else:
                db.execute(
                    "INSERT INTO eventi_dipendenti (evento_id,dipendente_id,ruolo,data_da,data_a) VALUES (?,?,?,?,?)",
                    (eid, int(did), ruolo, data_da, data_a))
            safe_commit(db)
            flash('Dipendente assegnato!', 'success')

            # ── Notifica email al dipendente ──────────────────────────
            ev  = db.execute("SELECT * FROM eventi WHERE id=?", (eid,)).fetchone()
            dip = db.execute("SELECT nome, cognome, email FROM utenti WHERE id=?", (int(did),)).fetchone()
            azienda = get_setting('azienda', 'La tua azienda')

            if ev and dip and dip['email']:
                data_inizio = ev['data_inizio'] or ''
                data_fine   = ev['data_fine']   or data_inizio
                luogo       = ev['luogo']        or '—'
                titolo      = ev['titolo']       or 'Nuovo lavoro'
                app_url     = request.host_url.rstrip('/')
                link        = f"{app_url}/mobile/calendario"

                corpo = f"""
<div style="font-family:Arial,sans-serif;max-width:580px;margin:0 auto;background:#f8fafc;padding:24px;border-radius:12px">
  <div style="background:linear-gradient(135deg,#1e3a5f,#0f172a);border-radius:10px;padding:24px;text-align:center;margin-bottom:24px">
    <div style="font-size:28px;margin-bottom:8px">📋</div>
    <h1 style="color:#fff;font-size:20px;margin:0">Sei stato assegnato a un nuovo lavoro</h1>
    <p style="color:rgba(255,255,255,.6);font-size:13px;margin:6px 0 0">{azienda}</p>
  </div>

  <p style="font-size:15px;color:#1e293b">Ciao <strong>{dip['nome']} {dip['cognome']}</strong>,</p>
  <p style="font-size:14px;color:#475569">Sei stato inserito nel seguente lavoro:</p>

  <div style="background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:20px;margin:16px 0">
    <table style="width:100%;border-collapse:collapse">
      <tr>
        <td style="padding:8px 0;color:#64748b;font-size:13px;width:40%"><strong>📌 Lavoro</strong></td>
        <td style="padding:8px 0;color:#1e293b;font-size:14px;font-weight:700">{titolo}</td>
      </tr>
      <tr style="border-top:1px solid #f1f5f9">
        <td style="padding:8px 0;color:#64748b;font-size:13px"><strong>👤 Ruolo</strong></td>
        <td style="padding:8px 0;color:#1e293b;font-size:14px">{ruolo}</td>
      </tr>
      <tr style="border-top:1px solid #f1f5f9">
        <td style="padding:8px 0;color:#64748b;font-size:13px"><strong>📅 Dal</strong></td>
        <td style="padding:8px 0;color:#1e293b;font-size:14px">{data_inizio}</td>
      </tr>
      <tr style="border-top:1px solid #f1f5f9">
        <td style="padding:8px 0;color:#64748b;font-size:13px"><strong>📅 Al</strong></td>
        <td style="padding:8px 0;color:#1e293b;font-size:14px">{data_fine}</td>
      </tr>
      <tr style="border-top:1px solid #f1f5f9">
        <td style="padding:8px 0;color:#64748b;font-size:13px"><strong>📍 Luogo</strong></td>
        <td style="padding:8px 0;color:#1e293b;font-size:14px">{luogo}</td>
      </tr>
    </table>
  </div>

  <div style="text-align:center;margin:24px 0">
    <a href="{link}" style="background:linear-gradient(135deg,#2563eb,#1d4ed8);color:#fff;text-decoration:none;padding:14px 28px;border-radius:10px;font-weight:700;font-size:15px;display:inline-block">
      📱 Apri Calendario Lavori
    </a>
  </div>

  <p style="font-size:12px;color:#94a3b8;text-align:center;margin-top:24px">
    {azienda} · Notifica automatica
  </p>
</div>"""

                # Invio diretto (sincrono) — il thread separato perde il Flask app context
                ok = send_email(dip['email'], f"[{azienda}] Nuovo lavoro assegnato: {titolo}", corpo)
                if ok:
                    flash(f'Email inviata a {dip["email"]}', 'success')
                else:
                    flash('Assegnato ma email non inviata — controlla impostazioni SMTP.', 'error')
            else:
                if ev and dip and not dip['email']:
                    flash('Dipendente assegnato (nessuna email configurata sul profilo).', 'success')

        except Exception as ex:
            flash(f'Errore assegnazione: {ex}', 'error')
        db.close()
    return redirect(url_for('evento_dettaglio', eid=eid))


@app.route('/calendario/<int:eid>/assegna-periodo/<int:assegn_id>', methods=['GET','POST'])
@admin_required
def evento_assegna_periodo(eid, assegn_id):
    db = get_db()
    ev = dict(db.execute("SELECT * FROM eventi WHERE id=?", (eid,)).fetchone())
    row = db.execute("""SELECT ed.*, u.nome, u.cognome
                        FROM eventi_dipendenti ed JOIN utenti u ON u.id=ed.dipendente_id
                        WHERE ed.id=?""", (assegn_id,)).fetchone()
    if not row:
        db.close(); flash('Assegnazione non trovata.','error'); return redirect(url_for('evento_dettaglio', eid=eid))

    if request.method == 'POST':
        da = request.form.get('data_da') or None
        a  = request.form.get('data_a')  or None
        db.execute("UPDATE eventi_dipendenti SET data_da=?, data_a=? WHERE id=?", (da, a, assegn_id))
        safe_commit(db); db.close()
        flash(f'Periodo aggiornato per {row["nome"]} {row["cognome"]}.','success')
        return redirect(url_for('evento_dettaglio', eid=eid))

    # GET - pagina con form
    db.close()
    return render_page(f"""
<div class="card" style="max-width:480px;margin:0 auto">
  <div class="card-header">
    <h3><i class="fa fa-calendar-pen" style="color:var(--accent)"></i> Modifica periodo — {row['nome']} {row['cognome']}</h3>
  </div>
  <div class="card-body">
    <p style="color:var(--text-light);font-size:13px;margin-bottom:16px">
      Evento: <strong>{ev['titolo']}</strong> ({ev['data_inizio']} → {ev.get('data_fine') or ev['data_inizio']})
    </p>
    <form method="POST">
      <div class="form-row">
        <div class="form-group">
          <label>Dal <span style="color:#64748b;font-weight:400">(incluso)</span></label>
          <input type="date" name="data_da"
            min="{ev['data_inizio']}" max="{ev.get('data_fine') or ev['data_inizio']}"
            value="{row['data_da'] or ev['data_inizio']}" required>
        </div>
        <div class="form-group">
          <label>Al <span style="color:#64748b;font-weight:400">(incluso)</span></label>
          <input type="date" name="data_a"
            min="{ev['data_inizio']}" max="{ev.get('data_fine') or ev['data_inizio']}"
            value="{row['data_a'] or ev.get('data_fine') or ev['data_inizio']}" required>
        </div>
      </div>
      <div style="background:#fef3c7;border-radius:8px;padding:10px;font-size:12px;color:#92400e;margin-bottom:16px">
        <i class="fa fa-info-circle"></i> Solo le timbrature in questo periodo verranno conteggiate nei costi dell'evento.
      </div>
      <div style="display:flex;gap:10px;justify-content:flex-end">
        <a href="/calendario/{eid}" class="btn btn-secondary">Annulla</a>
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva</button>
      </div>
    </form>
  </div>
</div>""", page_title='Modifica Periodo', active='calendario')


@app.route('/calendario/<int:eid>/rimuovi/<int:assegn_id>')
@admin_required
def evento_rimuovi_dipendente(eid, assegn_id):
    db = get_db()
    db.execute("DELETE FROM eventi_dipendenti WHERE id=?", (assegn_id,))
    safe_commit(db); db.close()
    flash('Dipendente rimosso.','success')
    return redirect(url_for('evento_dettaglio', eid=eid))


@app.route('/calendario/<int:eid>/elimina')
@admin_required
def evento_elimina(eid):
    db = get_db()
    db.execute("DELETE FROM eventi_dipendenti WHERE evento_id=?", (eid,))
    db.execute("DELETE FROM eventi WHERE id=?", (eid,))
    safe_commit(db); db.close()
    flash('Evento eliminato.','success')
    return redirect(url_for('calendario'))



@app.route('/calendario/<int:eid>/assegna-veicolo', methods=['POST'])
@admin_required
def evento_assegna_veicolo(eid):
    vid = request.form.get('veicolo_id')
    note = request.form.get('note','')
    if vid:
        db = get_db()
        try:
            db.execute("INSERT INTO eventi_veicoli (evento_id,veicolo_id,note) VALUES (?,?,?)",
                       (eid, int(vid), note))
            safe_commit(db)
            flash('Mezzo aggiunto!', 'success')
        except Exception as e:
            flash('Errore: ' + str(e), 'error')
        db.close()
    return redirect(url_for('evento_dettaglio', eid=eid))


@app.route('/calendario/<int:eid>/rimuovi-veicolo/<int:ev_vid>')
@admin_required
def evento_rimuovi_veicolo(eid, ev_vid):
    db = get_db()
    db.execute("DELETE FROM eventi_veicoli WHERE id=?", (ev_vid,))
    safe_commit(db); db.close()
    flash('Mezzo rimosso.', 'success')
    return redirect(url_for('evento_dettaglio', eid=eid))


@app.route('/calendario/<int:eid>/pacchetto-pass')
@admin_required
def evento_pacchetto_pass(eid):
    import base64 as _b64
    db = get_db()

    ev = dict(db.execute("SELECT * FROM eventi WHERE id=?", (eid,)).fetchone())
    tipo = ev.get('tipo') or 'altro'
    ev['colore'] = TIPO_COLORI_MAP.get(tipo, '#8b5cf6')
    ev['tipo_label'] = TIPO_LABEL_MAP.get(tipo, '📌 Altro')

    # ── Dipendenti con i loro docs identità ──
    team_raw = db.execute("""
        SELECT u.id as uid, u.nome, u.cognome, u.titolo, u.telefono, ed.ruolo
        FROM eventi_dipendenti ed JOIN utenti u ON u.id=ed.dipendente_id
        WHERE ed.evento_id=? ORDER BY ed.ruolo, u.nome
    """, (eid,)).fetchall()

    persone = []
    for t in team_raw:
        td = dict(t)
        td['iniziali'] = (t['nome'][0]+t['cognome'][0]).upper()

        # Cerca carta d'identità o passaporto
        docs_ci = db.execute("""
            SELECT * FROM documenti_dipendente
            WHERE utente_id=?
              AND (tipo_doc LIKE '%identit%' OR tipo_doc LIKE '%CI%'
                   OR tipo_doc LIKE '%carta%' OR tipo_doc LIKE '%passaporto%')
            ORDER BY caricato_il DESC
        """, (t['uid'],)).fetchall()

        td['docs_ci'] = []
        for doc in docs_ci:
            doc_path = os.path.join(get_upload_path(t['uid']), doc['nome_file'])
            img_b64 = None
            is_pdf  = doc['nome_file'].lower().endswith('.pdf')
            if os.path.exists(doc_path) and not is_pdf:
                try:
                    with open(doc_path, 'rb') as f:
                        raw = f.read()
                    ext = doc['nome_file'].rsplit('.',1)[-1].lower()
                    mime = {'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png','gif':'image/gif','webp':'image/webp'}.get(ext,'image/jpeg')
                    img_b64 = f"data:{mime};base64,{_b64.b64encode(raw).decode()}"
                except: pass
            td['docs_ci'].append({
                'nome': doc['nome_originale'],
                'scadenza': doc['data_scadenza'] or '',
                'img_b64': img_b64,
                'is_pdf':  is_pdf,
                'nome_file': doc['nome_file'],
                'uid': t['uid'],
                'doc_id': doc['id'],
            })
        persone.append(td)

    # ── Veicoli con libretto ──
    veicoli_raw = db.execute("""
        SELECT v.*, ev.note as ev_note
        FROM eventi_veicoli ev JOIN veicoli v ON v.id=ev.veicolo_id
        WHERE ev.evento_id=? ORDER BY v.targa
    """, (eid,)).fetchall()

    veicoli = []
    for v in veicoli_raw:
        vd = dict(v)
        docs_lib = db.execute("""
            SELECT * FROM documenti_veicoli
            WHERE veicolo_id=?
              AND (tipo_doc LIKE '%libretto%' OR tipo_doc LIKE '%circolaz%'
                   OR tipo_doc LIKE '%Libretto%' OR tipo_doc LIKE '%assicuraz%')
            ORDER BY caricato_il DESC
        """, (v['id'],)).fetchall()

        vd['docs_lib'] = []
        for doc in docs_lib:
            doc_path = os.path.join(get_upload_path_veicolo(v['id']), doc['nome_file'])
            img_b64 = None
            is_pdf  = doc['nome_file'].lower().endswith('.pdf')
            if os.path.exists(doc_path) and not is_pdf:
                try:
                    with open(doc_path, 'rb') as f:
                        raw = f.read()
                    ext = doc['nome_file'].rsplit('.',1)[-1].lower()
                    mime = {'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png'}.get(ext,'image/jpeg')
                    img_b64 = f"data:{mime};base64,{_b64.b64encode(raw).decode()}"
                except: pass
            vd['docs_lib'].append({
                'nome': doc['nome_originale'],
                'tipo': doc['tipo_doc'],
                'scadenza': doc['data_scadenza'] or '',
                'img_b64': img_b64,
                'is_pdf':  is_pdf,
                'doc_id':  doc['id'],
            })
        veicoli.append(vd)

    db.close()
    html = render_template_string(PASS_TMPL,
        ev=ev, persone=persone, veicoli=veicoli,
        azienda=get_setting('azienda','Azienda'), oggi=date.today().isoformat())
    return html



@app.route('/calendario/<int:eid>/scarica-zip')
@admin_required
def evento_scarica_zip(eid):
    import zipfile, io
    db = get_db()
    ev = db.execute("SELECT * FROM eventi WHERE id=?", (eid,)).fetchone()
    if not ev:
        db.close(); flash("Evento non trovato.","error"); return redirect(url_for('calendario'))

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:

        # ── Documenti identità dipendenti ──
        team_raw = db.execute("""
            SELECT u.id as uid, u.nome, u.cognome
            FROM eventi_dipendenti ed JOIN utenti u ON u.id=ed.dipendente_id
            WHERE ed.evento_id=? ORDER BY u.cognome
        """, (eid,)).fetchall()

        for t in team_raw:
            docs = db.execute("""
                SELECT * FROM documenti_dipendente
                WHERE utente_id=?
                  AND (tipo_doc LIKE '%identit%' OR tipo_doc LIKE '%CI%'
                       OR tipo_doc LIKE '%carta%' OR tipo_doc LIKE '%passaporto%'
                       OR tipo_doc LIKE '%patente%')
                ORDER BY caricato_il DESC
            """, (t['uid'],)).fetchall()

            folder = f"dipendenti/{t['cognome']}_{t['nome']}"
            for doc in docs:
                fpath = os.path.join(get_upload_path(t['uid']), doc['nome_file'])
                if os.path.exists(fpath):
                    ext = doc['nome_originale'].rsplit('.',1)[-1] if '.' in doc['nome_originale'] else doc['nome_file'].rsplit('.',1)[-1]
                    arcname = f"{folder}/{doc['tipo_doc']}_{doc['nome_originale']}"
                    zf.write(fpath, arcname)

        # ── Documenti veicoli ──
        veicoli_raw = db.execute("""
            SELECT v.* FROM eventi_veicoli ev JOIN veicoli v ON v.id=ev.veicolo_id
            WHERE ev.evento_id=? ORDER BY v.targa
        """, (eid,)).fetchall()

        for v in veicoli_raw:
            docs = db.execute("""
                SELECT * FROM documenti_veicoli WHERE veicolo_id=?
                ORDER BY tipo_doc, caricato_il DESC
            """, (v['id'],)).fetchall()

            folder = f"veicoli/{v['targa']}"
            for doc in docs:
                fpath = os.path.join(get_upload_path_veicolo(v['id']), doc['nome_file'])
                if os.path.exists(fpath):
                    arcname = f"{folder}/{doc['tipo_doc']}_{doc['nome_originale']}"
                    zf.write(fpath, arcname)

    db.close()
    zip_buf.seek(0)
    nome_ev = ev['titolo'].replace(' ','_').replace('/','_')[:40]
    return Response(
        zip_buf.getvalue(),
        mimetype='application/zip',
        headers={'Content-Disposition': f'attachment; filename=pass_{nome_ev}.zip'}
    )


PASS_TMPL = """<!DOCTYPE html>
<html lang="it"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Pacchetto Pass — {{ ev.titolo }}</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#1a1a2e;font-size:13px}
.page-header{background:linear-gradient(135deg,#1e2d40,#2d4a6e);color:#fff;padding:28px 36px;display:flex;align-items:center;justify-content:space-between}
.page-header h1{font-size:22px;font-weight:800;letter-spacing:-.3px}
.page-header .sub{font-size:13px;opacity:.7;margin-top:4px}
.badge-tipo{display:inline-block;padding:4px 12px;border-radius:99px;font-size:11px;font-weight:700;color:#fff;background:{{ ev.colore }};margin-top:6px}
.print-btn{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);color:#fff;padding:10px 20px;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600;text-decoration:none}
.print-btn:hover{background:rgba(255,255,255,.25)}
.content{max-width:900px;margin:28px auto;padding:0 20px 60px}
.section-title{font-size:11px;font-weight:800;color:#5a6172;text-transform:uppercase;letter-spacing:1.2px;margin:28px 0 12px;display:flex;align-items:center;gap:8px}
.section-title::after{content:'';flex:1;height:1px;background:#dde1e8}

/* Persona card */
.persona-card{background:#fff;border-radius:14px;box-shadow:0 2px 8px rgba(0,0,0,.07);margin-bottom:16px;overflow:hidden;border:1px solid #e2e8f0}
.persona-header{background:linear-gradient(90deg,#f8fafc,#fff);padding:14px 20px;display:flex;align-items:center;gap:14px;border-bottom:1px solid #f0f2f5}
.avatar{width:44px;height:44px;border-radius:50%;background:linear-gradient(135deg,#5b8fb9,#2d4a6e);display:flex;align-items:center;justify-content:center;color:#fff;font-size:16px;font-weight:800;flex-shrink:0}
.persona-nome{font-size:17px;font-weight:800;color:#1a1a2e}
.persona-ruolo{font-size:11px;color:#8e8e93;font-weight:600;margin-top:2px}
.persona-meta{font-size:12px;color:#5a6172;margin-top:3px}
.docs-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:14px;padding:16px 20px}
.doc-item{border:1px solid #e2e8f0;border-radius:10px;overflow:hidden}
.doc-item-header{background:#f8fafc;padding:8px 12px;display:flex;align-items:center;justify-content:space-between}
.doc-nome{font-size:11px;font-weight:700;color:#5a6172}
.doc-scad{font-size:10px;color:#e74c3c;font-weight:600}
.doc-img{width:100%;aspect-ratio:3/2;object-fit:cover;display:block;background:#f0f2f5}
.doc-pdf{background:#fff5eb;padding:24px;text-align:center;color:#e07a5f}
.doc-pdf i{font-size:32px;display:block;margin-bottom:6px}
.doc-pdf small{font-size:11px;color:#8e8e93;display:block;margin-top:4px}
.no-doc{padding:20px;text-align:center;color:#adb5bd;font-size:12px}
.no-doc i{font-size:28px;display:block;margin-bottom:6px;opacity:.4}

/* Veicolo card */
.veicolo-card{background:#fff;border-radius:14px;box-shadow:0 2px 8px rgba(0,0,0,.07);margin-bottom:16px;overflow:hidden;border:1px solid #e2e8f0}
.veicolo-header{background:linear-gradient(90deg,#fffbeb,#fff);padding:14px 20px;display:flex;align-items:center;gap:14px;border-bottom:1px solid #f0f2f5}
.targa-box{background:#1a1a2e;color:#fff;font-family:monospace;font-size:20px;font-weight:800;letter-spacing:4px;padding:8px 18px;border-radius:8px;border:3px solid #fbbf24}
.veicolo-info{font-size:13px;color:#5a6172;margin-top:3px}

/* Riepilogo finale */
.summary-box{background:#fff;border-radius:14px;padding:20px 24px;border:1px solid #e2e8f0;box-shadow:0 2px 8px rgba(0,0,0,.05)}
.summary-row{display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid #f0f2f5;font-size:13px}
.summary-row:last-child{border-bottom:none}
.summary-ok{color:#16a34a;font-weight:700}.summary-missing{color:#dc2626;font-weight:700}

@media print{
  body{background:#fff}
  .page-header{-webkit-print-color-adjust:exact;print-color-adjust:exact}
  .print-btn{display:none}
  .persona-card,.veicolo-card{break-inside:avoid}
  .doc-img{max-height:180px}
}
</style>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
</head>
<body>

<div class="page-header">
  <div>
    <div style="font-size:12px;opacity:.6;font-weight:600;letter-spacing:1px;text-transform:uppercase">{{ azienda }} · Pacchetto Pass</div>
    <h1>{{ ev.titolo }}</h1>
    {% if ev.luogo %}<div class="sub"><i class="fa fa-map-marker-alt"></i> {{ ev.luogo }}</div>{% endif %}
    <div class="sub" style="margin-top:4px">
      📅 {{ ev.data_inizio }} → {{ ev.data_fine or ev.data_inizio }}
      &nbsp;·&nbsp; Generato il {{ oggi }}
    </div>
    <span class="badge-tipo">{{ ev.tipo_label }}</span>
  </div>
  <div style="display:flex;gap:10px;align-items:center">
    <a href="/calendario/{{ ev.id }}/scarica-zip" class="print-btn" style="background:rgba(34,197,94,.25);border-color:rgba(34,197,94,.5)">
      <i class="fa fa-file-zipper"></i> Scarica tutto ZIP
    </a>
    <button onclick="window.print()" class="print-btn"><i class="fa fa-print"></i> Stampa / PDF</button>
  </div>
</div>

<div class="content">

  <!-- RIEPILOGO -->
  <div class="section-title"><i class="fa fa-clipboard-check"></i> Riepilogo</div>
  <div class="summary-box">
    <div class="summary-row">
      <span><i class="fa fa-users" style="color:#5b8fb9;margin-right:8px"></i>Personale assegnato</span>
      <strong>{{ persone|length }} persone</strong>
    </div>
    {% for p in persone %}
    <div class="summary-row">
      <span style="padding-left:20px;color:#5a6172">{{ p.nome }} {{ p.cognome }} <em style="font-size:11px">({{ p.ruolo }})</em></span>
      {% if p.docs_ci %}<span class="summary-ok"><i class="fa fa-check"></i> CI caricata ({{ p.docs_ci|length }})</span>
      {% else %}<span class="summary-missing"><i class="fa fa-exclamation-triangle"></i> CI mancante</span>{% endif %}
    </div>
    {% endfor %}
    <div class="summary-row" style="margin-top:8px">
      <span><i class="fa fa-truck" style="color:#f59e0b;margin-right:8px"></i>Mezzi assegnati</span>
      <strong>{{ veicoli|length }} veicoli</strong>
    </div>
    {% for v in veicoli %}
    <div class="summary-row">
      <span style="padding-left:20px;font-family:monospace;font-weight:800;color:#2d4a6e">{{ v.targa }}</span>
      {% if v.docs_lib %}<span class="summary-ok"><i class="fa fa-check"></i> Docs caricati ({{ v.docs_lib|length }})</span>
      {% else %}<span class="summary-missing"><i class="fa fa-exclamation-triangle"></i> Libretto mancante</span>{% endif %}
    </div>
    {% endfor %}
  </div>

  <!-- PERSONE -->
  <div class="section-title"><i class="fa fa-id-card"></i> Documenti Personale</div>
  {% for p in persone %}
  <div class="persona-card">
    <div class="persona-header">
      <div class="avatar">{{ p.iniziali }}</div>
      <div>
        <div class="persona-nome">{{ p.nome }} {{ p.cognome }}</div>
        <div class="persona-ruolo">{{ p.ruolo }}</div>
        {% if p.titolo %}<div class="persona-meta">{{ p.titolo }}</div>{% endif %}
      </div>
      <div style="margin-left:auto;text-align:right">
        {% if p.docs_ci %}
        <span style="background:#dcfce7;color:#16a34a;padding:4px 10px;border-radius:99px;font-size:11px;font-weight:700"><i class="fa fa-check"></i> {{ p.docs_ci|length }} doc</span>
        {% else %}
        <span style="background:#fee2e2;color:#dc2626;padding:4px 10px;border-radius:99px;font-size:11px;font-weight:700"><i class="fa fa-times"></i> Nessun documento</span>
        {% endif %}
      </div>
    </div>
    {% if p.docs_ci %}
    <div class="docs-grid">
      {% for doc in p.docs_ci %}
      <div class="doc-item">
        <div class="doc-item-header">
          <span class="doc-nome"><i class="fa fa-id-card"></i> {{ doc.nome }}</span>
          {% if doc.scadenza %}<span class="doc-scad">Scad. {{ doc.scadenza }}</span>{% endif %}
        </div>
        {% if doc.img_b64 %}
          <img src="{{ doc.img_b64 }}" class="doc-img" alt="{{ doc.nome }}">
          <div style="text-align:center;padding:6px">
            <a href="/dipendenti/{{ doc.uid }}/documenti/{{ doc.doc_id }}/scarica"
               style="background:#f1f5f9;color:#1e293b;padding:5px 14px;border-radius:6px;font-size:11px;font-weight:700;text-decoration:none;display:inline-block">
              <i class="fa fa-download"></i> Scarica
            </a>
          </div>
        {% elif doc.is_pdf %}
          <div class="doc-pdf">
            <i class="fa fa-file-pdf" style="color:#e74c3c"></i>
            <strong style="font-size:12px">PDF</strong>
            <small>{{ doc.nome }}</small>
            <div style="display:flex;gap:6px;justify-content:center;margin-top:8px">
              <a href="/dipendenti/{{ doc.uid }}/documenti/{{ doc.doc_id }}/anteprima" target="_blank"
                 style="background:#fee2e2;color:#e74c3c;padding:5px 12px;border-radius:6px;font-size:11px;font-weight:700;text-decoration:none">
                <i class="fa fa-eye"></i> Apri
              </a>
              <a href="/dipendenti/{{ doc.uid }}/documenti/{{ doc.doc_id }}/scarica" target="_blank"
                 style="background:#fee2e2;color:#e74c3c;padding:5px 12px;border-radius:6px;font-size:11px;font-weight:700;text-decoration:none">
                <i class="fa fa-download"></i> Scarica
              </a>
            </div>
          </div>
        {% else %}
          <div class="no-doc"><i class="fa fa-image"></i>File non disponibile</div>
        {% endif %}
      </div>
      {% endfor %}
    </div>
    {% else %}
    <div class="no-doc" style="padding:28px">
      <i class="fa fa-exclamation-circle" style="font-size:28px;color:#fca5a5;display:block;margin-bottom:8px"></i>
      <strong style="color:#dc2626">Carta d'identità non caricata</strong>
      <div style="margin-top:4px;font-size:11px">Carica il documento nella scheda dipendente</div>
    </div>
    {% endif %}
  </div>
  {% endfor %}

  <!-- VEICOLI -->
  <div class="section-title"><i class="fa fa-truck"></i> Documenti Mezzi</div>
  {% for v in veicoli %}
  <div class="veicolo-card">
    <div class="veicolo-header">
      <div style="font-size:28px">{{ '🚐' if v.tipo=='Furgone' else ('🚛' if v.tipo in ('Camion','Autocarro') else '🚗') }}</div>
      <div>
        <div class="targa-box">{{ v.targa }}</div>
      </div>
      <div style="margin-left:8px">
        <div style="font-size:16px;font-weight:800">{{ v.marca or '' }} {{ v.modello or '' }}</div>
        <div class="veicolo-info">{{ v.tipo or '' }}{% if v.anno %} · Anno {{ v.anno }}{% endif %}</div>
      </div>
      <div style="margin-left:auto;text-align:right">
        {% if v.docs_lib %}
        <span style="background:#dcfce7;color:#16a34a;padding:4px 10px;border-radius:99px;font-size:11px;font-weight:700"><i class="fa fa-check"></i> {{ v.docs_lib|length }} doc</span>
        {% else %}
        <span style="background:#fee2e2;color:#dc2626;padding:4px 10px;border-radius:99px;font-size:11px;font-weight:700"><i class="fa fa-times"></i> Nessun documento</span>
        {% endif %}
      </div>
    </div>
    {% if v.docs_lib %}
    <div class="docs-grid">
      {% for doc in v.docs_lib %}
      <div class="doc-item">
        <div class="doc-item-header">
          <span class="doc-nome"><i class="fa fa-file-alt"></i> {{ doc.tipo }} · {{ doc.nome }}</span>
          {% if doc.scadenza %}<span class="doc-scad">Scad. {{ doc.scadenza }}</span>{% endif %}
        </div>
        {% if doc.img_b64 %}
          <img src="{{ doc.img_b64 }}" class="doc-img" alt="{{ doc.nome }}">
          <div style="text-align:center;padding:6px">
            <a href="/dipendenti/{{ doc.uid }}/documenti/{{ doc.doc_id }}/scarica"
               style="background:#f1f5f9;color:#1e293b;padding:5px 14px;border-radius:6px;font-size:11px;font-weight:700;text-decoration:none;display:inline-block">
              <i class="fa fa-download"></i> Scarica
            </a>
          </div>
        {% elif doc.is_pdf %}
          <div class="doc-pdf">
            <i class="fa fa-file-pdf" style="color:#e74c3c"></i>
            <strong style="font-size:12px">PDF</strong>
            <small>{{ doc.nome }}</small>
          </div>
        {% else %}
          <div class="no-doc"><i class="fa fa-image"></i>File non disponibile</div>
        {% endif %}
      </div>
      {% endfor %}
    </div>
    {% else %}
    <div class="no-doc" style="padding:28px">
      <i class="fa fa-exclamation-circle" style="font-size:28px;color:#fca5a5;display:block;margin-bottom:8px"></i>
      <strong style="color:#dc2626">Libretto / documenti non caricati</strong>
    </div>
    {% endif %}
  </div>
  {% endfor %}

</div>
</body>
</html>"""


# ── Evento Files: upload / download / delete ─────────────

ALLOWED_EXT_EVENTI = {'pdf','png','jpg','jpeg','gif','webp','bmp','dwg','dxf','zip','rar','7z'}

@app.route('/calendario/<int:eid>/files/upload', methods=['POST'])
@admin_required
def evento_file_upload(eid):
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Nessun file selezionato.','error')
        return redirect(url_for('evento_dettaglio', eid=eid))
    ext = f.filename.rsplit('.',1)[-1].lower() if '.' in f.filename else ''
    if ext not in ALLOWED_EXT_EVENTI:
        flash('Tipo file non permesso.','error')
        return redirect(url_for('evento_dettaglio', eid=eid))
    nome_file = datetime.now().strftime('%Y%m%d_%H%M%S_') + secure_filename(f.filename)
    path = os.path.join(get_upload_path_evento(eid), nome_file)
    f.save(path)
    dim  = os.path.getsize(path)
    cat  = request.form.get('categoria','disegno')
    note = request.form.get('note','')
    dip_id = request.form.get('dipendente_id') or None
    if dip_id:
        try: dip_id = int(dip_id)
        except: dip_id = None
    db   = get_db()
    db.execute(
        "INSERT INTO eventi_files (evento_id,categoria,nome_originale,nome_file,dimensione,note,caricato_da,dipendente_id) VALUES (?,?,?,?,?,?,?,?)",
        (eid, cat, f.filename, nome_file, dim, note, session['user_id'], dip_id)
    )
    safe_commit(db); db.close()
    label = 'Disegno' if cat=='disegno' else 'Pass'
    flash(f'{label} caricato!','success')
    return redirect(url_for('evento_dettaglio', eid=eid))


@app.route('/calendario/<int:eid>/files/<int:fid>/scarica')
@login_required
def evento_file_scarica(eid, fid):
    db  = get_db()
    row = db.execute("SELECT * FROM eventi_files WHERE id=? AND evento_id=?", (fid, eid)).fetchone()
    if not row:
        db.close(); flash('File non trovato.','error'); return redirect(url_for('calendario'))
    # Controlla che il dipendente sia assegnato all'evento (o sia admin)
    if session.get('ruolo') != 'admin':
        ass = db.execute(
            "SELECT id FROM eventi_dipendenti WHERE evento_id=? AND dipendente_id=?",
            (eid, session['user_id'])).fetchone()
        if not ass:
            db.close(); flash('Accesso non autorizzato.','error'); return redirect(url_for('mobile'))
    db.close()
    path = os.path.join(get_upload_path_evento(eid), row['nome_file'])
    if not os.path.exists(path):
        flash('File non trovato sul server.','error'); return redirect(url_for('evento_dettaglio', eid=eid))
    return send_file(path, as_attachment=True, download_name=row['nome_originale'])


@app.route('/calendario/<int:eid>/files/<int:fid>/elimina')
@admin_required
def evento_file_elimina(eid, fid):
    db  = get_db()
    row = db.execute("SELECT * FROM eventi_files WHERE id=? AND evento_id=?", (fid, eid)).fetchone()
    if row:
        path = os.path.join(get_upload_path_evento(eid), row['nome_file'])
        if os.path.exists(path):
            os.remove(path)
        db.execute("DELETE FROM eventi_files WHERE id=?", (fid,))
        safe_commit(db)
        flash('File eliminato.','success')
    db.close()
    return redirect(url_for('evento_dettaglio', eid=eid))


# ── Mobile: Calendario dipendente ────────────────────────

MOBILE_CAL_DIPENDENTE_TMPL = """<!DOCTYPE html>
<html lang="{{ lang }}" dir="{{ t.dir }}">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>{{ t.work_calendar }}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{background:#0f172a;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;min-height:100vh;color:#fff;direction:{{ t.dir }}}
.header{background:linear-gradient(135deg,#1e3a5f,#0f172a);padding:20px 20px 16px;border-bottom:1px solid rgba(255,255,255,.08)}
.header-top{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px}
.back-btn{font-size:13px;color:rgba(255,255,255,.6);text-decoration:none;padding:7px 13px;border:1px solid rgba(255,255,255,.15);border-radius:10px;display:flex;align-items:center;gap:6px}
.page-title{font-size:20px;font-weight:800;color:#fff}
.page-sub{font-size:13px;color:rgba(255,255,255,.45);margin-top:3px}
.lang-switcher{display:flex;gap:6px;margin-top:10px;flex-wrap:wrap}
.lang-pill{background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.15);border-radius:20px;padding:4px 10px;font-size:12px;color:rgba(255,255,255,.5);cursor:pointer;text-decoration:none}
.lang-pill.active{background:rgba(59,130,246,.3);border-color:#3b82f6;color:#93c5fd}
.content{padding:16px;max-width:480px;margin:0 auto;display:flex;flex-direction:column;gap:14px;padding-bottom:32px}
.month-nav{display:flex;align-items:center;justify-content:space-between;background:#1e293b;border-radius:14px;padding:12px 16px}
.month-nav a{color:rgba(255,255,255,.5);text-decoration:none;padding:6px 12px;border-radius:8px;font-size:13px}
.month-nav a:active{background:rgba(255,255,255,.1)}
.month-label{font-size:16px;font-weight:700;color:#fff}
.ev-card{border-radius:16px;overflow:hidden;border:1px solid rgba(255,255,255,.08)}
.ev-header{padding:14px 16px;display:flex;align-items:center;gap:12px}
.ev-titolo{font-size:16px;font-weight:800;color:#fff;line-height:1.2}
.ev-luogo{font-size:12px;color:rgba(255,255,255,.5);margin-top:3px;display:flex;align-items:center;gap:5px}
.ev-date-row{display:flex;gap:8px;padding:0 16px 14px}
.ev-date-chip{background:rgba(255,255,255,.07);border-radius:8px;padding:7px 12px;font-size:12px;color:rgba(255,255,255,.6);flex:1;text-align:center}
.ev-date-chip strong{display:block;font-size:14px;color:#fff;font-weight:700;margin-bottom:1px}
.my-role-banner{margin:0 16px 14px;border-radius:10px;padding:10px 14px;display:flex;align-items:center;gap:10px;font-size:13px;font-weight:700}
.team-section{background:rgba(255,255,255,.04);border-top:1px solid rgba(255,255,255,.07);padding:12px 16px}
.team-title{font-size:10px;font-weight:700;color:rgba(255,255,255,.35);text-transform:uppercase;letter-spacing:1px;margin-bottom:10px}
.team-group-label{font-size:10px;color:rgba(255,255,255,.3);text-transform:uppercase;letter-spacing:.8px;margin:8px 0 6px;font-weight:600}
.team-member{display:flex;align-items:center;gap:10px;padding:6px 0;border-bottom:1px solid rgba(255,255,255,.05)}
.team-member:last-child{border-bottom:none}
.avatar-sm{width:30px;height:30px;border-radius:50%;background:linear-gradient(135deg,#3b82f6,#1d4ed8);display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:#fff;flex-shrink:0}
.avatar-me{background:linear-gradient(135deg,#22c55e,#16a34a)}
.member-name{font-size:13px;color:rgba(255,255,255,.8);font-weight:500}
.file-link{display:flex;align-items:center;gap:10px;padding:10px;border-radius:10px;margin-bottom:8px;text-decoration:none}
.empty{text-align:center;padding:40px 20px;color:rgba(255,255,255,.3)}
.empty i{font-size:40px;margin-bottom:12px;display:block;opacity:.4}
.tipo-montaggio{background:linear-gradient(135deg,#14532d,#16a34a)}
.tipo-fiera{background:linear-gradient(135deg,#1e3a5f,#2196F3)}
.tipo-smontaggio{background:linear-gradient(135deg,#7f1d1d,#e63946)}
.tipo-altro{background:linear-gradient(135deg,#4c1d95,#8b5cf6)}
.role-montaggio{background:rgba(22,163,74,.15);color:#16a34a;border:1px solid rgba(22,163,74,.3)}
.role-fiera{background:rgba(33,150,243,.15);color:#60a5fa;border:1px solid rgba(33,150,243,.3)}
.role-smontaggio{background:rgba(15,76,129,.15);color:#f87171;border:1px solid rgba(230,57,70,.3)}
.role-altro{background:rgba(139,92,246,.15);color:#a78bfa;border:1px solid rgba(139,92,246,.3)}
</style>
</head>
<body>
<div class="header">
  <div class="header-top">
    <a href="/mobile" class="back-btn"><i class="fa fa-arrow-{{ 'right' if t.dir=='rtl' else 'left' }}"></i> {{ t.back_btn }}</a>
    <div style="font-size:11px;color:rgba(255,255,255,.35);font-weight:600">{{ t.assigned_work }}</div>
  </div>
  <div class="page-title"><i class="fa fa-hard-hat" style="font-size:18px;margin-{{ 'left' if t.dir=='rtl' else 'right' }}:8px;opacity:.8"></i>{{ t.work_calendar }}</div>
  <div class="page-sub">{{ t.work_calendar_sub }}</div>
  <!-- Cambio lingua -->
  <div class="lang-switcher">
    {% for code, l in langs.items() %}
    <form method="POST" action="/set-lang" style="display:inline">
      <input type="hidden" name="lang" value="{{ code }}">
      <input type="hidden" name="next" value="/mobile/calendario">
      <button type="submit" class="lang-pill {{ 'active' if current_lang==code }}" style="border:none;cursor:pointer">{{ l.flag }}</button>
    </form>
    {% endfor %}
  </div>
</div>

<div class="content">
  <div class="month-nav">
    <a href="/mobile/calendario?m={{ prev_m }}&y={{ prev_y }}"><i class="fa fa-chevron-{{ 'right' if t.dir=='rtl' else 'left' }}"></i></a>
    <span class="month-label">{{ mese_nome }} {{ anno }}</span>
    <a href="/mobile/calendario?m={{ next_m }}&y={{ next_y }}"><i class="fa fa-chevron-{{ 'left' if t.dir=='rtl' else 'right' }}"></i></a>
  </div>

  {% if eventi %}
  {% for ev in eventi %}
  {% set tipo = ev.tipo or 'altro' %}
  <div class="ev-card">
    <div class="ev-header tipo-{{ tipo }}">
      <div>
        <div style="font-size:10px;font-weight:700;color:rgba(255,255,255,.6);text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">{{ ev.tipo_label }}</div>
        <div class="ev-titolo">{{ ev.titolo }}</div>
        {% if ev.luogo %}<div class="ev-luogo"><i class="fa fa-map-marker-alt"></i> {{ ev.luogo }}</div>{% endif %}
      </div>
    </div>

    <div class="ev-date-row">
      <div class="ev-date-chip"><strong>{{ ev.data_inizio }}</strong>{{ t.start }}</div>
      <div style="display:flex;align-items:center;color:rgba(255,255,255,.2);font-size:16px">→</div>
      <div class="ev-date-chip"><strong>{{ ev.data_fine or ev.data_inizio }}</strong>{{ t.end }}</div>
      <div class="ev-date-chip"><strong>{{ ev.giorni }}{{ t.days_unit }}</strong>{{ t.duration }}</div>
    </div>

    <div class="my-role-banner role-{{ tipo }}">
      <i class="fa fa-user-check" style="font-size:16px"></i>
      <div>
        <div style="font-size:10px;opacity:.7;font-weight:600;letter-spacing:.5px">{{ t.your_role }}</div>
        <div style="font-size:14px">{{ ev.mio_ruolo }}</div>
      </div>
    </div>

    {% if ev.team %}
    <div class="team-section">
      <div class="team-title"><i class="fa fa-users"></i> {{ t.full_team }}</div>
      {% for ruolo_k, membri in ev.team_per_ruolo.items() %}
      <div class="team-group-label">{{ ruolo_k }}</div>
      {% for m in membri %}
      <div class="team-member">
        <div class="avatar-sm {{ 'avatar-me' if m.sono_io }}">{{ m.iniziali }}</div>
        <div style="flex:1">
          <div class="member-name">{{ m.nome }} {{ m.cognome }}{% if m.sono_io %} ✓{% endif %}</div>
          {% if m.data_da or m.data_a %}
          <div style="font-size:10px;color:rgba(255,255,255,.4);margin-top:2px">
            <i class="fa fa-calendar-days"></i> {{ m.data_da or '' }} → {{ m.data_a or '' }}
          </div>
          {% endif %}
        </div>
      </div>
      {% endfor %}
      {% endfor %}
    </div>
    {% endif %}

    {% if ev.disegni %}
    <div class="team-section">
      <div class="team-title"><i class="fa fa-drafting-compass" style="color:#a78bfa"></i> {{ t.drawings }}</div>
      {% for f in ev.disegni %}
      <a href="/calendario/{{ ev.id }}/files/{{ f.id }}/scarica" class="file-link" style="background:rgba(139,92,246,.1);border:1px solid rgba(139,92,246,.2)">
        <div style="width:36px;height:36px;background:rgba(139,92,246,.2);border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:18px;flex-shrink:0">📐</div>
        <div style="flex:1;min-width:0">
          <div style="color:#c4b5fd;font-weight:700;font-size:13px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{{ f.nome_originale }}</div>
          {% if f.note %}<div style="color:rgba(255,255,255,.4);font-size:11px">{{ f.note }}</div>{% endif %}
        </div>
        <div style="color:#a78bfa;font-size:16px"><i class="fa fa-download"></i></div>
      </a>
      {% endfor %}
    </div>
    {% endif %}

    {% if ev.pass_files %}
    <div class="team-section">
      <div class="team-title"><i class="fa fa-id-badge" style="color:#22d3ee"></i> {{ t.passes }}</div>
      {% for f in ev.pass_files %}
      <a href="/calendario/{{ ev.id }}/files/{{ f.id }}/scarica" class="file-link" style="background:rgba(8,145,178,.15);border:1px solid rgba(8,145,178,.35);border-radius:12px;margin-bottom:8px">
        <div style="width:40px;height:40px;background:rgba(8,145,178,.25);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0">🎫</div>
        <div style="flex:1;min-width:0">
          <div style="font-size:10px;font-weight:700;color:#22d3ee;text-transform:uppercase;letter-spacing:.8px;margin-bottom:2px">
            {% if f.dipendente_id %}Il tuo pass{% else %}Pass evento{% endif %}
          </div>
          <div style="color:#e0f2fe;font-weight:700;font-size:13px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{{ f.nome_originale }}</div>
          {% if f.note %}<div style="color:rgba(255,255,255,.45);font-size:11px">{{ f.note }}</div>{% endif %}
        </div>
        <div style="color:#22d3ee;font-size:18px;padding-right:4px"><i class="fa fa-download"></i></div>
      </a>
      {% endfor %}
    </div>
    {% endif %}

  </div>
  {% endfor %}

  {% else %}
  <div class="empty">
    <i class="fa fa-calendar-xmark"></i>
    <p>{{ t.no_events }} <strong>{{ mese_nome }} {{ anno }}</strong>.</p>
  </div>
  {% endif %}
</div>
</body>
</html>"""


# ── Mobile: Calendario dipendente route ──────────────────

@app.route('/mobile/calendario')
@login_required
def mobile_calendario():
    import calendar as _cal
    today = date.today()
    m = int(request.args.get('m', today.month))
    y = int(request.args.get('y', today.year))
    uid = session['user_id']

    db = get_db()
    mese_inizio = f'{y:04d}-{m:02d}-01'
    ultimo_giorno = _cal.monthrange(y,m)[1]
    mese_fine = f'{y:04d}-{m:02d}-{ultimo_giorno:02d}'

    eventi_raw = db.execute("""
        SELECT e.*, ed.ruolo as mio_ruolo
        FROM eventi e
        JOIN eventi_dipendenti ed ON ed.evento_id = e.id
        WHERE ed.dipendente_id = ?
          AND e.data_inizio <= ? AND COALESCE(e.data_fine, e.data_inizio) >= ?
          AND COALESCE(e.data_fine, e.data_inizio) >= ?
        ORDER BY e.data_inizio
    """, (uid, mese_fine, mese_inizio, today.isoformat())).fetchall()

    eventi = []
    for e in eventi_raw:
        ed = dict(e)
        tipo = ed.get('tipo') or 'altro'
        ed['colore']     = TIPO_COLORI_MAP.get(tipo, '#8b5cf6')
        ed['tipo_label'] = TIPO_LABEL_MAP.get(tipo, '📌 Altro')
        try:
            d1 = datetime.strptime(ed['data_inizio'], '%Y-%m-%d').date()
            d2 = datetime.strptime(ed['data_fine'] or ed['data_inizio'], '%Y-%m-%d').date()
            ed['giorni'] = (d2 - d1).days + 1
        except:
            ed['giorni'] = 1

        team_raw = db.execute("""
            SELECT u.id, u.nome, u.cognome, ed2.ruolo, ed2.data_da, ed2.data_a
            FROM eventi_dipendenti ed2 JOIN utenti u ON u.id=ed2.dipendente_id
            WHERE ed2.evento_id=? ORDER BY ed2.ruolo, u.nome
        """, (ed['id'],)).fetchall()

        team_per_ruolo = {}
        for t in team_raw:
            membro = dict(t)
            membro['iniziali'] = (t['nome'][0]+t['cognome'][0]).upper()
            membro['sono_io'] = (t['id'] == uid)
            membro['data_da'] = t['data_da']
            membro['data_a']  = t['data_a']
            team_per_ruolo.setdefault(t['ruolo'], []).append(membro)

        ed['team'] = [m for members in team_per_ruolo.values() for m in members]
        ed['team_per_ruolo'] = team_per_ruolo

        files_raw = db.execute(
            "SELECT * FROM eventi_files WHERE evento_id=? ORDER BY caricato_il DESC", (ed['id'],)
        ).fetchall()
        ed['disegni']    = [dict(f) for f in files_raw if f['categoria'] == 'disegno']
        # Pass: mostra solo quello assegnato al dipendente, o quelli generici (dipendente_id IS NULL)
        ed['pass_files'] = [dict(f) for f in files_raw
                            if f['categoria'] == 'pass'
                            and (f['dipendente_id'] is None or f['dipendente_id'] == uid)]

        eventi.append(ed)

    db.close()

    t    = get_lang()
    lang = session.get('lang','it')
    mese_nome = t['months'][m]
    prev_m, prev_y = (m-1, y) if m > 1 else (12, y-1)
    next_m, next_y = (m+1, y) if m < 12 else (1, y+1)
    nome = session.get('nome','')

    return render_template_string(MOBILE_CAL_DIPENDENTE_TMPL,
        eventi=eventi, mese_nome=mese_nome, anno=y, m=m, y=y,
        prev_m=prev_m, prev_y=prev_y, next_m=next_m, next_y=next_y,
        nome=nome, t=t, lang=lang, langs=LANGS, current_lang=lang)

# ══════════════════════════════════════════════════════════
#  VEICOLI
# ══════════════════════════════════════════════════════════

TIPI_DOC_VEICOLO = ['Assicurazione','Revisione','Bollo','Manutenzione','Tagliando','Carta di circolazione','Collaudo','Altro']

VEICOLI_TMPL = """
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px">
  <div></div>
  {% if session.ruolo==\'admin\' %}
  <a href="/veicoli/nuovo" class="btn btn-primary"><i class="fa fa-plus"></i> Nuovo Veicolo</a>
  {% endif %}
</div>
<div class="grid-3" style="margin-bottom:24px">
  <div class="stat-card"><div class="stat-icon" style="background:#e3f2fd"><i class="fa fa-car" style="color:#2196F3"></i></div><div class="stat-info"><div class="stat-value">{{ veicoli|length }}</div><div class="stat-label">Veicoli totali</div></div></div>
  <div class="stat-card"><div class="stat-icon" style="background:#fce4ec"><i class="fa fa-exclamation-triangle" style="color:#0f4c81"></i></div><div class="stat-info"><div class="stat-value">{{ scaduti }}</div><div class="stat-label">Documenti scaduti</div></div></div>
  <div class="stat-card"><div class="stat-icon" style="background:#fff8e1"><i class="fa fa-clock" style="color:#f59e0b"></i></div><div class="stat-info"><div class="stat-value">{{ in_scadenza }}</div><div class="stat-label">In scadenza (30gg)</div></div></div>
</div>
<div class="card">
  <div class="table-wrap"><table>
    <thead><tr><th>Targa</th><th>Veicolo</th><th>Tipo</th><th>Documenti</th><th>Prossima scadenza</th><th>Stato</th><th></th></tr></thead>
    <tbody>{% for v in veicoli %}
    <tr>
      <td><strong style="font-family:monospace;font-size:14px;color:var(--accent2)">{{ v.targa }}</strong></td>
      <td>{{ v.marca or \'\' }} {{ v.modello or \'\' }}{% if v.anno %} <span style="color:var(--text-light);font-size:12px">({{ v.anno }})</span>{% endif %}</td>
      <td><span class="tag">{{ v.tipo }}</span></td>
      <td>
        <a href="/veicoli/{{ v.id }}/documenti" style="text-decoration:none">
          <span style="background:#e3f2fd;color:#2196F3;padding:3px 10px;border-radius:99px;font-size:12px;font-weight:700">
            <i class="fa fa-folder"></i> {{ v.n_docs }}
          </span>
        </a>
      </td>
      <td style="font-family:monospace;font-size:12px">{{ v.prossima_scadenza or \'–\' }}</td>
      <td>{% if v.days_left is none %}<span class="badge badge-gray">–</span>
          {% elif v.days_left < 0 %}<span class="badge badge-red">⚠ Scaduto</span>
          {% elif v.days_left <= 30 %}<span class="badge badge-amber">{{ v.days_left }}gg</span>
          {% else %}<span class="badge badge-green">OK</span>{% endif %}</td>
      <td style="display:flex;gap:6px">
        <a href="/veicoli/{{ v.id }}/documenti" class="btn btn-primary btn-sm" title="Documenti"><i class="fa fa-folder-open"></i></a>
        {% if session.ruolo==\'admin\' %}
        <a href="/veicoli/{{ v.id }}/modifica" class="btn btn-secondary btn-sm"><i class="fa fa-pen"></i></a>
        <a href="/veicoli/{{ v.id }}/elimina" class="btn btn-danger btn-sm" onclick="return confirm(\'Eliminare il veicolo e tutti i suoi documenti?\')"><i class="fa fa-trash"></i></a>
        {% endif %}
      </td>
    </tr>{% else %}
    <tr><td colspan="7"><div class="empty-state"><i class="fa fa-car"></i><p>Nessun veicolo. <a href="/veicoli/nuovo">Aggiungi il primo!</a></p></div></td></tr>
    {% endfor %}</tbody>
  </table></div>
</div>
"""

VEICOLO_FORM_TMPL = """
<div class="card" style="max-width:600px;margin:0 auto">
  <div class="card-header"><h3>{{ \'Modifica\' if veicolo else \'Nuovo\' }} Veicolo</h3></div>
  <div class="card-body"><form method="POST">
    <div class="form-row">
      <div class="form-group"><label>Targa *</label><input name="targa" value="{{ veicolo.targa if veicolo else \'\' }}" required style="text-transform:uppercase" placeholder="ES123AB"></div>
      <div class="form-group"><label>Tipo</label>
        <select name="tipo">
          {% for t in [\'Auto\',\'Furgone\',\'Camion\',\'Moto\',\'Rimorchio\',\'Altro\'] %}
          <option value="{{ t }}" {{ \'selected\' if veicolo and veicolo.tipo==t }}>{{ t }}</option>
          {% endfor %}
        </select>
      </div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Marca</label><input name="marca" value="{{ veicolo.marca if veicolo else \'\' }}" placeholder="Es. Fiat"></div>
      <div class="form-group"><label>Modello</label><input name="modello" value="{{ veicolo.modello if veicolo else \'\' }}" placeholder="Es. Ducato"></div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Anno</label><input type="number" name="anno" value="{{ veicolo.anno if veicolo else \'\' }}" placeholder="2020"></div>
      <div class="form-group"><label>Colore</label><input name="colore" value="{{ veicolo.colore if veicolo else \'\' }}" placeholder="Bianco"></div>
    </div>
    <div class="form-group"><label>Note</label><textarea name="note">{{ veicolo.note if veicolo else \'\' }}</textarea></div>
    <div style="display:flex;gap:10px;justify-content:flex-end">
      <a href="/veicoli" class="btn btn-secondary">Annulla</a>
      <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva</button>
    </div>
  </form></div>
</div>
"""

VEICOLO_DETAIL_TMPL = """
{% macro scad_badge(days) %}
  {% if days == 9999 %}<span style="background:#f1f5f9;color:#94a3b8;padding:3px 10px;border-radius:99px;font-size:12px">Non impostata</span>
  {% elif days < 0 %}<span style="background:#fef2f2;color:#dc2626;padding:3px 10px;border-radius:99px;font-size:12px;font-weight:700">&#x26D4; Scaduta</span>
  {% elif days <= 30 %}<span style="background:#fffbeb;color:#d97706;padding:3px 10px;border-radius:99px;font-size:12px;font-weight:700">&#x26A0;&#xFE0F; {{ days }} giorni</span>
  {% else %}<span style="background:#f0fdf4;color:#16a34a;padding:3px 10px;border-radius:99px;font-size:12px;font-weight:700">&#x2705; {{ days }} giorni</span>
  {% endif %}
{% endmacro %}
<div style="margin-bottom:16px;display:flex;gap:8px;align-items:center">
  <a href="/veicoli" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Torna ai veicoli</a>
  <a href="/veicoli/{{ v.id }}/modifica" class="btn btn-secondary btn-sm"><i class="fa fa-pen"></i> Modifica</a>
  <a href="/veicoli/{{ v.id }}/documenti" class="btn btn-primary btn-sm"><i class="fa fa-folder-open"></i> Documenti ({{ n_docs }})</a>
</div>
<div class="card" style="margin-bottom:20px">
  <div class="card-header" style="background:linear-gradient(135deg,#0f172a,#1e293b);color:#fff">
    <h3 style="color:#fff;margin:0;display:flex;align-items:center;gap:12px;flex-wrap:wrap">
      <i class="fa fa-car" style="color:#60a5fa"></i>
      <span style="font-family:monospace;font-size:22px;letter-spacing:2px">{{ v.targa }}</span>
      <span style="font-size:14px;font-weight:400;color:#94a3b8">{{ v.marca or "" }} {{ v.modello or "" }}{% if v.anno %} ({{ v.anno }}){% endif %}</span>
      <span style="margin-left:auto;background:#1e40af;color:#93c5fd;padding:3px 12px;border-radius:99px;font-size:13px">{{ v.tipo }}</span>
    </h3>
  </div>
  <div class="card-body">
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:16px">
      <div><div style="font-size:11px;color:#94a3b8;text-transform:uppercase;font-weight:600;margin-bottom:4px">Marca</div><div style="font-weight:600">{{ v.marca or "–" }}</div></div>
      <div><div style="font-size:11px;color:#94a3b8;text-transform:uppercase;font-weight:600;margin-bottom:4px">Modello</div><div style="font-weight:600">{{ v.modello or "–" }}</div></div>
      <div><div style="font-size:11px;color:#94a3b8;text-transform:uppercase;font-weight:600;margin-bottom:4px">Anno</div><div style="font-weight:600">{{ v.anno or "–" }}</div></div>
      <div><div style="font-size:11px;color:#94a3b8;text-transform:uppercase;font-weight:600;margin-bottom:4px">Colore</div><div style="font-weight:600">{{ v.colore or "–" }}</div></div>
    </div>
    {% if v.note %}<div style="margin-top:12px;padding:10px;background:#f8fafc;border-radius:8px;font-size:13px;color:#64748b">{{ v.note }}</div>{% endif %}
  </div>
</div>
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-calendar-exclamation" style="color:#f59e0b"></i> Scadenze</h3></div>
  <div class="card-body">
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px">
      <div style="padding:14px;border:1px solid var(--border);border-radius:10px">
        <div style="font-size:12px;color:#64748b;font-weight:600;margin-bottom:8px"><i class="fa fa-shield-halved" style="color:#3b82f6"></i> Assicurazione</div>
        <div style="font-family:monospace;font-size:13px;margin-bottom:6px">{{ v.scad_assicurazione or "–" }}</div>
        {{ scad_badge(v.dl_ass) }}
        {% if v.note_assicurazione %}<div style="font-size:11px;color:#94a3b8;margin-top:6px">{{ v.note_assicurazione }}</div>{% endif %}
      </div>
      <div style="padding:14px;border:1px solid var(--border);border-radius:10px">
        <div style="font-size:12px;color:#64748b;font-weight:600;margin-bottom:8px"><i class="fa fa-wrench" style="color:#8b5cf6"></i> Revisione</div>
        <div style="font-family:monospace;font-size:13px;margin-bottom:6px">{{ v.scad_revisione or "–" }}</div>
        {{ scad_badge(v.dl_rev) }}
        {% if v.note_revisione %}<div style="font-size:11px;color:#94a3b8;margin-top:6px">{{ v.note_revisione }}</div>{% endif %}
      </div>
      <div style="padding:14px;border:1px solid var(--border);border-radius:10px">
        <div style="font-size:12px;color:#64748b;font-weight:600;margin-bottom:8px"><i class="fa fa-receipt" style="color:#f59e0b"></i> Bollo</div>
        <div style="font-family:monospace;font-size:13px;margin-bottom:6px">{{ v.scad_bollo or "–" }}</div>
        {{ scad_badge(v.dl_bol) }}
        {% if v.note_bollo %}<div style="font-size:11px;color:#94a3b8;margin-top:6px">{{ v.note_bollo }}</div>{% endif %}
      </div>
      <div style="padding:14px;border:1px solid var(--border);border-radius:10px">
        <div style="font-size:12px;color:#64748b;font-weight:600;margin-bottom:8px"><i class="fa fa-oil-can" style="color:#16a34a"></i> Tagliando</div>
        <div style="font-family:monospace;font-size:13px;margin-bottom:6px">{{ v.scad_tagliando or "–" }}</div>
        {{ scad_badge(v.dl_tag) }}
        {% if v.note_tagliando %}<div style="font-size:11px;color:#94a3b8;margin-top:6px">{{ v.note_tagliando }}</div>{% endif %}
      </div>
    </div>
    {% if session.ruolo=="admin" %}
    <form method="POST" action="/veicoli/{{ v.id }}/scadenze" style="margin-top:20px;padding-top:16px;border-top:1px solid var(--border)">
      <h4 style="margin-bottom:12px;font-size:13px;color:#64748b;text-transform:uppercase">Aggiorna scadenze</h4>
      <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:14px">
        <div><label style="font-size:12px;font-weight:600">Assicurazione</label>
          <input type="date" name="scad_assicurazione" value="{{ v.scad_assicurazione or \'\' }}" class="form-control" style="margin-bottom:4px">
          <input type="text" name="note_assicurazione" value="{{ v.note_assicurazione or \'\' }}" class="form-control" placeholder="Note..." style="font-size:12px">
        </div>
        <div><label style="font-size:12px;font-weight:600">Revisione</label>
          <input type="date" name="scad_revisione" value="{{ v.scad_revisione or \'\' }}" class="form-control" style="margin-bottom:4px">
          <input type="text" name="note_revisione" value="{{ v.note_revisione or \'\' }}" class="form-control" placeholder="Note..." style="font-size:12px">
        </div>
        <div><label style="font-size:12px;font-weight:600">Bollo</label>
          <input type="date" name="scad_bollo" value="{{ v.scad_bollo or \'\' }}" class="form-control" style="margin-bottom:4px">
          <input type="text" name="note_bollo" value="{{ v.note_bollo or \'\' }}" class="form-control" placeholder="Note..." style="font-size:12px">
        </div>
        <div><label style="font-size:12px;font-weight:600">Tagliando</label>
          <input type="date" name="scad_tagliando" value="{{ v.scad_tagliando or \'\' }}" class="form-control" style="margin-bottom:4px">
          <input type="text" name="note_tagliando" value="{{ v.note_tagliando or \'\' }}" class="form-control" placeholder="Note..." style="font-size:12px">
        </div>
      </div>
      <div style="display:flex;justify-content:flex-end;margin-top:12px">
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva scadenze</button>
      </div>
    </form>
    {% endif %}
  </div>
</div>
"""


VEICOLO_DOCS_TMPL = """
<div style="margin-bottom:16px">
  <a href="/veicoli" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Torna ai veicoli</a>
</div>

<!-- Info veicolo -->
<div class="card" style="margin-bottom:20px;background:linear-gradient(135deg,#0f172a,#0f4c81);color:#fff">
  <div class="card-body" style="display:flex;align-items:center;gap:24px">
    <div style="font-size:48px;opacity:.8"><i class="fa fa-car"></i></div>
    <div>
      <div style="font-size:28px;font-weight:800;font-family:monospace;letter-spacing:2px">{{ veicolo.targa }}</div>
      <div style="font-size:16px;opacity:.8">{{ veicolo.marca or \'\' }} {{ veicolo.modello or \'\' }} {% if veicolo.anno %}({{ veicolo.anno }}){% endif %}</div>
      <div style="margin-top:6px"><span style="background:rgba(255,255,255,.2);padding:3px 10px;border-radius:99px;font-size:12px">{{ veicolo.tipo }}</span>
      {% if veicolo.colore %}<span style="background:rgba(255,255,255,.2);padding:3px 10px;border-radius:99px;font-size:12px;margin-left:6px">{{ veicolo.colore }}</span>{% endif %}</div>
    </div>
  </div>
</div>

<!-- Carica nuovo documento -->
{% if session.ruolo==\'admin\' %}
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-upload" style="color:var(--accent2)"></i> Carica documento</h3></div>
  <div class="card-body">
    <form method="POST" action="/veicoli/{{ veicolo.id }}/documenti/upload" enctype="multipart/form-data">
      <input type="hidden" name="tipo_doc" value="Auto">
      <div class="form-row-4">
        <div class="form-group" style="grid-column:span 2">
          <label>File * (PDF, JPG, PNG — max 20MB)</label>
          <input type="file" name="file" accept=".pdf,.jpg,.jpeg,.png,.gif,.webp,.bmp" required>
        </div>
        <div class="form-group">
          <label>Data scadenza <span style="color:var(--text-light);font-size:11px">(l\'AI la trova da sola)</span></label>
          <input type="date" name="data_scadenza">
        </div>
        <div class="form-group"><label>Note</label><input name="note" placeholder="Descrizione..."></div>
      </div>
      <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap">
        <button type="submit" name="analizza" value="1" class="btn" style="background:linear-gradient(135deg,#7c3aed,#4f46e5);color:#fff;font-size:15px;padding:10px 22px"
          {{ \'disabled\' if not ai_ok }}>
          <i class="fa fa-robot"></i> Carica + Analizza con AI ✨
        </button>
        <button type="submit" class="btn btn-secondary"><i class="fa fa-upload"></i> Carica senza analisi</button>
      </div>
    </form>
  </div>
</div>

<!-- Risultato AI -->
{% if ai_result %}
<div class="card" style="margin-bottom:20px;border:2px solid #7c3aed">
  <div class="card-header" style="background:linear-gradient(135deg,#7c3aed22,#4f46e522)">
    <h3><i class="fa fa-robot" style="color:#7c3aed"></i> Risultato analisi AI ✨</h3>
    <form method="POST" action="/veicoli/{{ veicolo.id }}/applica-ai">
      <input type="hidden" name="dati_json" value=\'{{ ai_result.dati_json }}\'>
      <button type="submit" class="btn btn-sm" style="background:#7c3aed;color:#fff"><i class="fa fa-check"></i> Applica dati</button>
    </form>
  </div>
  <div class="card-body">
    <p style="font-size:13px;color:var(--text-light);margin-bottom:14px"><strong>Tipo rilevato:</strong> {{ ai_result.tipo_doc }}</p>
    <p style="font-size:13px;color:var(--text-light);margin-bottom:14px">{{ ai_result.testo }}</p>
    {% if ai_result.dati_estratti %}
    {% set LABELS = {\'data_scadenza\':\'Scadenza\',\'tipo_doc\':\'Tipo documento\',\'compagnia\':\'Compagnia\',\'numero_polizza\':\'N° Polizza\',\'massimale\':\'Massimale\',\'data_revisione\':\'Data revisione\',\'km_manutenzione\':\'Km manutenzione\',\'data_manutenzione\':\'Data manutenzione\',\'officina\':\'Officina\',\'descrizione\':\'Descrizione\'} %}
    <div class="form-row" style="font-size:13px">
      {% for k,v in ai_result.dati_estratti.items() %}
      {% if v %}<div><span style="color:var(--text-light)">{{ LABELS.get(k,k) }}:</span> <strong>{{ v }}</strong></div>{% endif %}
      {% endfor %}
    </div>
    {% endif %}
  </div>
</div>
{% endif %}
{% endif %}

<!-- Lista documenti -->
<div class="card">
  <div class="card-header">
    <h3><i class="fa fa-folder-open" style="color:var(--accent2)"></i> Documenti ({{ docs|length }})</h3>
  </div>
  {% if docs %}
  <div class="table-wrap"><table>
    <thead><tr><th>Tipo</th><th>File</th><th>Note</th><th>Scadenza</th><th>Stato</th><th></th></tr></thead>
    <tbody>{% for d in docs %}
    <tr>
      <td><span class="badge badge-purple">{{ d.tipo_doc }}</span></td>
      <td>
        {% if d.nome_originale.lower().endswith(\'.pdf\') %}
          <i class="fa fa-file-pdf" style="color:#0f4c81"></i>
        {% else %}
          <i class="fa fa-file-image" style="color:#2196F3"></i>
        {% endif %}
        <strong>{{ d.nome_originale }}</strong>
        <div style="font-size:10px;color:var(--text-light)">{{ (d.dimensione//1024)|string }}KB · {{ d.caricato_il[:10] }}</div>
      </td>
      <td style="font-size:12px;color:var(--text-light)">{{ d.note or \'–\' }}</td>
      <td style="font-family:monospace;font-size:12px">{{ d.data_scadenza or \'–\' }}</td>
      <td>{% if not d.data_scadenza %}<span class="badge badge-gray">–</span>
          {% elif d.days_left < 0 %}<span class="badge badge-red">Scaduto</span>
          {% elif d.days_left <= 7 %}<span class="badge badge-red">⚠ {{ d.days_left }}gg</span>
          {% elif d.days_left <= 30 %}<span class="badge badge-amber">{{ d.days_left }}gg</span>
          {% else %}<span class="badge badge-green">OK</span>{% endif %}</td>
      <td style="display:flex;gap:6px">
        <a href="/veicoli/{{ veicolo.id }}/documenti/{{ d.id }}/scarica" class="btn btn-secondary btn-sm" title="Scarica"><i class="fa fa-download"></i></a>
        <a href="/veicoli/{{ veicolo.id }}/documenti/{{ d.id }}/anteprima" class="btn btn-secondary btn-sm" title="Anteprima" target="_blank"><i class="fa fa-eye"></i></a>
        {% if session.ruolo==\'admin\' %}
        <a href="/veicoli/{{ veicolo.id }}/documenti/{{ d.id }}/elimina" class="btn btn-danger btn-sm" onclick="return confirm(\'Eliminare?\')"><i class="fa fa-trash"></i></a>
        {% endif %}
      </td>
    </tr>{% else %}
    <tr><td colspan="6"><div class="empty-state"><i class="fa fa-folder-open"></i><p>Nessun documento caricato</p></div></td></tr>
    {% endfor %}</tbody>
  </table></div>
  {% else %}
  <div class="empty-state"><i class="fa fa-folder-open"></i><p>Carica il primo documento per questo veicolo</p></div>
  {% endif %}
</div>
"""

def _days_left(scad_str):
    if not scad_str:
        return None
    try:
        from datetime import date as _date
        scad = _date.fromisoformat(scad_str)
        return (scad - _date.today()).days
    except Exception:
        return None

@app.route('/veicoli')
@login_required
def veicoli():
    db = get_db()
    vs = db.execute("SELECT * FROM veicoli WHERE attivo=1 ORDER BY targa").fetchall()
    veicoli_list = []
    scaduti = in_scadenza = 0
    for v in vs:
        vd = dict(v)
        campi_scad = ['scad_assicurazione', 'scad_revisione', 'scad_bollo', 'scad_tagliando']
        min_dl = None
        min_data = None
        contato = False
        for campo in campi_scad:
            dl = _days_left(v[campo])
            vd['dl_' + campo] = dl if dl is not None else 9999
            if dl is not None:
                if min_dl is None or dl < min_dl:
                    min_dl = dl
                    min_data = v[campo]
                if not contato:
                    if dl < 0:
                        scaduti += 1
                        contato = True
                    elif dl <= 30:
                        in_scadenza += 1
                        contato = True
        vd['days_left'] = min_dl
        vd['prossima_scadenza'] = min_data
        vd['n_docs'] = db.execute(
            "SELECT COUNT(*) FROM documenti_veicoli WHERE veicolo_id=?", (v['id'],)).fetchone()[0]
        veicoli_list.append(vd)
    db.close()
    from datetime import date
    return render_page(VEICOLI_TMPL, page_title='Veicoli Aziendali', active='veicoli',
                       veicoli=veicoli_list, scaduti=scaduti, in_scadenza=in_scadenza,
                       today_str=date.today().isoformat())


@app.route('/veicoli/nuovo', methods=['GET', 'POST'])
@admin_required
def veicolo_nuovo():
    if request.method == 'POST':
        db = get_db()
        db.execute(
            "INSERT INTO veicoli (targa,marca,modello,tipo,anno,colore,note) VALUES (?,?,?,?,?,?,?)",
            (request.form['targa'].upper(), request.form.get('marca'), request.form.get('modello'),
             request.form.get('tipo', 'Auto'), request.form.get('anno') or None,
             request.form.get('colore'), request.form.get('note')))
        safe_commit(db)
        vid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.close()
        flash('Veicolo aggiunto!', 'success')
        return redirect(url_for('veicolo_detail', vid=vid))
    # Render a simple form reusing VEICOLO_DETAIL logic via a mini template
    form_html = """
<div class="card" style="max-width:600px;margin:0 auto">
  <div class="card-header"><h3>Nuovo Veicolo</h3></div>
  <div class="card-body"><form method="POST">
    <div class="form-row">
      <div class="form-group"><label>Targa *</label><input name="targa" required placeholder="ES123AB"></div>
      <div class="form-group"><label>Tipo</label><select name="tipo">{% for t in ['Auto','Furgone','Camion','Moto','Rimorchio','Altro'] %}<option>{{ t }}</option>{% endfor %}</select></div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Marca</label><input name="marca" placeholder="Es. Ford"></div>
      <div class="form-group"><label>Modello</label><input name="modello" placeholder="Es. Transit"></div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Anno immatricolazione</label><input type="number" name="anno" placeholder="2020"></div>
      <div class="form-group"><label>Colore</label><input name="colore" placeholder="Bianco"></div>
    </div>
    <div class="form-group"><label>Note</label><input name="note"></div>
    <div style="display:flex;gap:10px;justify-content:flex-end">
      <a href="/veicoli" class="btn btn-secondary">Annulla</a>
      <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Crea veicolo</button>
    </div>
  </form></div>
</div>"""
    return render_page(form_html, page_title='Nuovo Veicolo', active='veicoli')


@app.route('/veicoli/<int:vid>')
@login_required
def veicolo_detail(vid):
    db = get_db()
    v = db.execute("SELECT * FROM veicoli WHERE id=?", (vid,)).fetchone()
    if not v:
        return redirect(url_for('veicoli'))
    n_docs = db.execute(
        "SELECT COUNT(*) FROM documenti_veicoli WHERE veicolo_id=?", (vid,)).fetchone()[0]
    db.close()
    vd = dict(v)
    vd['dl_ass'] = _days_left(v['scad_assicurazione'])
    vd['dl_rev'] = _days_left(v['scad_revisione'])
    vd['dl_bol'] = _days_left(v['scad_bollo'])
    vd['dl_tag'] = _days_left(v['scad_tagliando'])
    for k in ['dl_ass', 'dl_rev', 'dl_bol', 'dl_tag']:
        if vd[k] is None:
            vd[k] = 9999
    return render_page(VEICOLO_DETAIL_TMPL, page_title='Veicolo ' + v['targa'],
                       active='veicoli', v=vd, n_docs=n_docs)


@app.route('/veicoli/<int:vid>/modifica', methods=['GET', 'POST'])
@admin_required
def veicolo_modifica(vid):
    db = get_db()
    v = db.execute("SELECT * FROM veicoli WHERE id=?", (vid,)).fetchone()
    if not v:
        db.close()
        return redirect(url_for('veicoli'))
    if request.method == 'POST':
        db.execute(
            "UPDATE veicoli SET targa=?,marca=?,modello=?,tipo=?,anno=?,colore=?,note=? WHERE id=?",
            (request.form['targa'].upper(), request.form.get('marca'), request.form.get('modello'),
             request.form.get('tipo', 'Auto'), request.form.get('anno') or None,
             request.form.get('colore'), request.form.get('note'), vid))
        safe_commit(db); db.close()
        flash('Anagrafica aggiornata!', 'success')
        return redirect(url_for('veicolo_detail', vid=vid))
    v = dict(v)
    db.close()
    form_html = """
<div class="card" style="max-width:600px;margin:0 auto">
  <div class="card-header"><h3><i class="fa fa-pen" style="color:#3b82f6"></i> Modifica Veicolo — {{ v.targa }}</h3></div>
  <div class="card-body"><form method="POST">
    <div class="form-row">
      <div class="form-group"><label>Targa *</label><input name="targa" required value="{{ v.targa }}" style="text-transform:uppercase"></div>
      <div class="form-group"><label>Tipo</label><select name="tipo">{% for t in ['Auto','Furgone','Camion','Moto','Rimorchio','Altro'] %}<option {{ 'selected' if v.tipo==t }}>{{ t }}</option>{% endfor %}</select></div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Marca</label><input name="marca" value="{{ v.marca or '' }}"></div>
      <div class="form-group"><label>Modello</label><input name="modello" value="{{ v.modello or '' }}"></div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Anno immatricolazione</label><input type="number" name="anno" value="{{ v.anno or '' }}"></div>
      <div class="form-group"><label>Colore</label><input name="colore" value="{{ v.colore or '' }}"></div>
    </div>
    <div class="form-group"><label>Note</label><input name="note" value="{{ v.note or '' }}"></div>
    <div style="display:flex;gap:10px;justify-content:flex-end">
      <a href="/veicoli/{{ v.id }}" class="btn btn-secondary">Annulla</a>
      <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva modifiche</button>
    </div>
  </form></div>
</div>"""
    return render_page(form_html, page_title='Modifica ' + v['targa'], active='veicoli', v=v)


@app.route('/veicoli/<int:vid>/scadenze', methods=['POST'])
@admin_required
def veicolo_salva_scadenze(vid):
    db = get_db()
    db.execute(
        """UPDATE veicoli SET
           scad_assicurazione=?, note_assicurazione=?,
           scad_revisione=?,    note_revisione=?,
           scad_bollo=?,        note_bollo=?,
           scad_tagliando=?,    note_tagliando=?
           WHERE id=?""",
        (request.form.get('scad_assicurazione') or None,
         request.form.get('note_assicurazione') or None,
         request.form.get('scad_revisione') or None,
         request.form.get('note_revisione') or None,
         request.form.get('scad_bollo') or None,
         request.form.get('note_bollo') or None,
         request.form.get('scad_tagliando') or None,
         request.form.get('note_tagliando') or None,
         vid))
    safe_commit(db); db.close()
    flash('Scadenze aggiornate!', 'success')
    return redirect(url_for('veicolo_detail', vid=vid))


@app.route('/veicoli/<int:vid>/elimina')
@admin_required
def veicolo_elimina(vid):
    db = get_db()
    docs = db.execute("SELECT * FROM documenti_veicoli WHERE veicolo_id=?", (vid,)).fetchall()
    for d in docs:
        path = os.path.join(get_upload_path_veicolo(vid), d['nome_file'])
        if os.path.exists(path):
            os.remove(path)
    db.execute("DELETE FROM documenti_veicoli WHERE veicolo_id=?", (vid,))
    db.execute("DELETE FROM veicoli WHERE id=?", (vid,))
    safe_commit(db); db.close()
    flash('Veicolo eliminato.', 'success')
    return redirect(url_for('veicoli'))


@app.route('/veicoli/<int:vid>/documenti')
@login_required
def veicolo_documenti(vid):
    db = get_db()
    v = db.execute("SELECT * FROM veicoli WHERE id=?", (vid,)).fetchone()
    if not v:
        return redirect(url_for('veicoli'))
    docs = db.execute(
        "SELECT * FROM documenti_veicoli WHERE veicolo_id=? ORDER BY caricato_il DESC",
        (vid,)).fetchall()
    ai_result = session.pop('ai_result_veicolo', None)
    db.close()
    return render_page(VEICOLO_DOCS_TMPL, page_title='Documenti ' + v['targa'],
                       active='veicoli', veicolo=v,
                       docs=[dict(d) for d in docs],
                       ai_result=ai_result, ai_ok=AI_OK)


@app.route('/veicoli/<int:vid>/documenti/upload', methods=['POST'])
@admin_required
def veicolo_upload(vid):
    f = request.files.get('file')
    if not f or not allowed_file(f.filename):
        flash('File non valido.', 'error')
        return redirect(url_for('veicolo_documenti', vid=vid))
    nome_file = datetime.now().strftime('%Y%m%d_%H%M%S') + '_' + secure_filename(f.filename)
    path = os.path.join(get_upload_path_veicolo(vid), nome_file)
    f.save(path)
    dim = os.path.getsize(path)
    tipo_doc = request.form.get('tipo_doc', 'Altro')
    note = request.form.get('note', '')
    analizza = request.form.get('analizza')
    db = get_db()
    doc_id = db.execute(
        "INSERT INTO documenti_veicoli (veicolo_id,nome_originale,nome_file,tipo_doc,dimensione,note) VALUES (?,?,?,?,?,?)",
        (vid, f.filename, nome_file, tipo_doc, dim, note)).lastrowid
    safe_commit(db); db.close()
    flash('Documento caricato!', 'success')
    if analizza and AI_OK:
        ai_result = analizza_documento_veicolo_ai(path, f.filename, vid, doc_id)
        tipo_ai = ai_result.get('tipo_doc', '').strip()
        if tipo_ai and tipo_ai not in ('Altro', ''):
            db3 = get_db()
            db3.execute("UPDATE documenti_veicoli SET tipo_doc=? WHERE id=?", (tipo_ai, doc_id))
            db3.commit(); db3.close()
            flash(f'AI ha rilevato: {tipo_ai}', 'success')
        session['ai_result_veicolo'] = ai_result
    return redirect(url_for('veicolo_documenti', vid=vid))


@app.route('/veicoli/<int:vid>/applica-ai', methods=['POST'])
@admin_required
def veicolo_applica_ai(vid):
    try:
        dati = json.loads(request.form.get('dati_json', '{}'))
    except Exception:
        dati = {}
    # Aggiorna le scadenze direttamente nell'anagrafica del veicolo
    tipo = str(dati.get('tipo_documento', '')).upper()
    scad = (dati.get('data_scadenza') or dati.get('data_revisione') or
            dati.get('data_manutenzione') or '')
    note_val = str(dati.get('compagnia') or dati.get('officina') or dati.get('descrizione') or '')
    db = get_db()
    if scad:
        if 'ASSICURAZIONE' in tipo:
            db.execute("UPDATE veicoli SET scad_assicurazione=?, note_assicurazione=? WHERE id=?",
                       (scad, note_val or None, vid))
        elif 'REVISIONE' in tipo:
            db.execute("UPDATE veicoli SET scad_revisione=?, note_revisione=? WHERE id=?",
                       (scad, note_val or None, vid))
        elif 'BOLLO' in tipo:
            db.execute("UPDATE veicoli SET scad_bollo=? WHERE id=?", (scad, vid))
        elif any(k in tipo for k in ['TAGLIANDO', 'MANUTENZIONE', 'COLLAUDO']):
            db.execute("UPDATE veicoli SET scad_tagliando=?, note_tagliando=? WHERE id=?",
                       (scad, note_val or None, vid))
        safe_commit(db)
        flash(f'Scadenza {scad} aggiornata nell\'anagrafica veicolo!', 'success')
    else:
        flash('Nessuna scadenza rilevata dall\'AI — controlla i dati estratti.', 'error')
    db.close()
    return redirect(url_for('veicolo_documenti', vid=vid))


@app.route('/veicoli/<int:vid>/documenti/<int:did>/scarica')
@login_required
def veicolo_scarica(vid, did):
    db = get_db()
    doc = db.execute(
        "SELECT * FROM documenti_veicoli WHERE id=? AND veicolo_id=?", (did, vid)).fetchone()
    db.close()
    if not doc:
        return "Non trovato", 404
    return send_file(os.path.join(get_upload_path_veicolo(vid), doc['nome_file']),
                     as_attachment=True, download_name=doc['nome_originale'])


@app.route('/veicoli/<int:vid>/documenti/<int:did>/anteprima')
@login_required
def veicolo_anteprima(vid, did):
    db = get_db()
    doc = db.execute(
        "SELECT * FROM documenti_veicoli WHERE id=? AND veicolo_id=?", (did, vid)).fetchone()
    db.close()
    if not doc:
        return "Non trovato", 404
    path = os.path.join(get_upload_path_veicolo(vid), doc['nome_file'])
    mt = mimetypes.guess_type(path)[0] or 'application/octet-stream'
    return send_file(path, mimetype=mt)


@app.route('/veicoli/<int:vid>/documenti/<int:did>/elimina')
@admin_required
def veicolo_elimina_doc(vid, did):
    db = get_db()
    doc = db.execute(
        "SELECT * FROM documenti_veicoli WHERE id=? AND veicolo_id=?", (did, vid)).fetchone()
    if doc:
        path = os.path.join(get_upload_path_veicolo(vid), doc['nome_file'])
        if os.path.exists(path):
            os.remove(path)
        db.execute("DELETE FROM documenti_veicoli WHERE id=?", (did,))
        safe_commit(db)
        flash('Documento eliminato.', 'success')
    db.close()
    return redirect(url_for('veicolo_documenti', vid=vid))


def analizza_documento_veicolo_ai(file_path, nome_file, vid, doc_id):
    api_key = get_setting('anthropic_api_key', '')
    if not api_key or not AI_OK:
        return {'tipo_doc': 'Altro', 'testo': 'API Key non configurata.',
                'dati_estratti': {}, 'dati_json': '{}'}
    try:
        import urllib.request
        proxies = urllib.request.getproxies()
        proxy_url = proxies.get('https') or proxies.get('http')
        http_client = None
        if proxy_url:
            import httpx
            http_client = httpx.Client(proxy=proxy_url, verify=False)
        client = anthropic.Anthropic(
            api_key=api_key, **({"http_client": http_client} if http_client else {}))
        ext = nome_file.rsplit('.', 1)[1].lower()
        with open(file_path, 'rb') as fp:
            b64 = base64.standard_b64encode(fp.read()).decode()
        if ext == 'pdf':
            cb = {"type": "document",
                  "source": {"type": "base64", "media_type": "application/pdf", "data": b64}}
        else:
            mt_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                      'png': 'image/png', 'gif': 'image/gif', 'webp': 'image/webp'}
            cb = {"type": "image",
                  "source": {"type": "base64",
                              "media_type": mt_map.get(ext, 'image/jpeg'), "data": b64}}
        prompt = (
            "Sei un esperto di documenti veicoli italiani. Rispondi SOLO con JSON valido:\n"
            '{"tipo_documento":"Assicurazione|Revisione|Bollo|Tagliando|Carta di circolazione|Collaudo|Altro",'
            '"sommario":"breve descrizione 1 riga",'
            '"dati_estratti":{"tipo_documento":"","data_scadenza":"YYYY-MM-DD",'
            '"compagnia":"","numero_polizza":"","massimale":"",'
            '"data_revisione":"YYYY-MM-DD","km_manutenzione":"",'
            '"officina":"","targa":"","descrizione":""}}\n'
            "IMPORTANTE: data_scadenza = data fino a cui il documento e' valido. "
            "Date in formato YYYY-MM-DD. Stringa vuota se non presente."
        )
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=1024,
            messages=[{"role": "user", "content": [cb, {"type": "text", "text": prompt}]}])
        raw = resp.content[0].text.strip()
        if '```' in raw:
            raw = raw.split('```')[1].lstrip('json').strip()
        ob = raw.count('{')
        cb2 = raw.count('}')
        if ob > cb2:
            raw += '}' * (ob - cb2)
        data = json.loads(raw)
        dati = {k: str(v) for k, v in data.get('dati_estratti', {}).items()
                if v and str(v).strip()}
        tipo = str(data.get('tipo_documento', 'Altro'))
        return {'tipo_doc': tipo, 'testo': data.get('sommario', 'Analisi completata.'),
                'dati_estratti': dati, 'dati_json': json.dumps(dati, ensure_ascii=False)}
    except Exception as e:
        return {'tipo_doc': 'Altro', 'testo': f'Errore AI: {e}',
                'dati_estratti': {}, 'dati_json': '{}'}


# ══════════════════════════════════════════════════════════
#  FATTURAZIONE
# ══════════════════════════════════════════════════════════

import os as _os

FATT_UPLOAD_DIR_KEY = 'uploads_fatture'

def get_fatt_upload_path():
    p = os.path.join(DATA_DIR, 'uploads_fatture')
    os.makedirs(p, exist_ok=True)
    return p

def get_docs_azienda_path():
    p = os.path.join(DATA_DIR, 'uploads_docs_azienda')
    os.makedirs(p, exist_ok=True)
    return p

# ══════════════════════════════════════════════════════════════
#  DOCUMENTI AZIENDA
# ══════════════════════════════════════════════════════════════

DOCS_AZ_CATEGORIE = [
    'DURC', 'Visura camerale', 'Assicurazione', 'Certificato antimafia',
    'Licenza', 'Contratto', 'Autorizzazione', 'Certificazione',
    'Patente', 'Iscrizione albo', 'Altro'
]

DOCS_AZ_TMPL = """
{% macro doc_card(d) %}
{% if d.stato_scad == 'scaduto' %}{% set ic='#fef2f2' %}{% set ic2='#dc2626' %}{% set ic3='fa-file-circle-xmark' %}
{% elif d.stato_scad == 'warning' %}{% set ic='#fffbeb' %}{% set ic2='#f59e0b' %}{% set ic3='fa-file-circle-exclamation' %}
{% elif d.stato_scad == 'ok' %}{% set ic='#f0fdf4' %}{% set ic2='#16a34a' %}{% set ic3='fa-file-circle-check' %}
{% else %}{% set ic='#f8fafc' %}{% set ic2='#94a3b8' %}{% set ic3='fa-file' %}{% endif %}
<label class="doc-card" style="cursor:pointer" for="chk-{{ d.id }}">
  <input type="checkbox" name="ids" value="{{ d.id }}" id="chk-{{ d.id }}" class="doc-chk" style="width:18px;height:18px;accent-color:#7c3aed;flex-shrink:0" onchange="aggiornaZip()">
  <div class="doc-icon" style="background:{{ ic }};color:{{ ic2 }}"><i class="fa {{ ic3 }}"></i></div>
  <div class="doc-info">
    <div class="doc-title">{{ d.titolo }}</div>
    <div class="doc-cat">
      <span style="background:#f1f5f9;padding:1px 8px;border-radius:10px">{{ d.categoria or 'Altro' }}</span>
      {% if d.data_emissione %}<span style="margin-left:6px;color:#94a3b8">emesso {{ d.data_emissione }}</span>{% endif %}
      {% if not d.file_nome %}<span style="margin-left:6px;color:#cbd5e1;font-size:11px"><i class="fa fa-file-slash"></i> nessun file</span>{% endif %}
    </div>
    {% if d.data_scadenza %}
    <div class="doc-scad scad-{{ d.stato_scad }}">
      {% if d.stato_scad == 'scaduto' %}⛔ Scaduto il {{ d.data_scadenza }}
      {% elif d.stato_scad == 'warning' %}⚠️ Scade il {{ d.data_scadenza }} ({{ d.giorni_alla_scad }} giorni)
      {% else %}✅ Scade il {{ d.data_scadenza }} ({{ d.giorni_alla_scad }} giorni)
      {% endif %}
    </div>
    {% else %}<div class="doc-scad scad-none"><i class="fa fa-infinity" style="font-size:11px"></i> Nessuna scadenza</div>{% endif %}
    {% if d.note %}<div style="font-size:12px;color:#64748b;margin-top:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{{ d.note }}</div>{% endif %}
  </div>
  <div style="display:flex;flex-direction:column;gap:6px;flex-shrink:0" onclick="event.stopPropagation()">
    {% if d.file_nome %}
    <a href="/documenti-azienda/{{ d.id }}/scarica" class="btn btn-sm btn-secondary" title="Scarica"><i class="fa fa-download"></i></a>
    {% endif %}
    <a href="/documenti-azienda/{{ d.id }}/modifica" class="btn btn-sm btn-secondary" title="Modifica"><i class="fa fa-pen"></i></a>
    <a href="/documenti-azienda/{{ d.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare?')" title="Elimina"><i class="fa fa-trash"></i></a>
  </div>
</label>
{% endmacro %}

<style>
.doc-card{background:#fff;border:1px solid var(--border);border-radius:12px;padding:16px 18px;display:flex;align-items:center;gap:14px;transition:.15s;width:100%;box-sizing:border-box;text-decoration:none}
.doc-card:hover{box-shadow:0 2px 10px rgba(0,0,0,.07)}
.doc-card:has(.doc-chk:checked){border-color:#7c3aed;background:#faf5ff;box-shadow:0 0 0 2px #ede9fe}
.doc-icon{width:44px;height:44px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:19px;flex-shrink:0}
.doc-info{flex:1;min-width:0}
.doc-title{font-weight:700;font-size:14px;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.doc-cat{font-size:12px;color:#64748b;margin-top:2px}
.doc-scad{font-size:12px;font-weight:600;margin-top:5px}
.scad-ok{color:#16a34a}.scad-warning{color:#f59e0b}.scad-scaduto{color:#dc2626}.scad-none{color:#94a3b8}
.doc-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(360px,1fr));gap:14px;margin-top:14px}
.section-header{display:flex;align-items:center;gap:10px;margin:24px 0 8px;padding-bottom:8px;border-bottom:2px solid var(--border)}
.section-header h3{margin:0;font-size:15px}
#zip-bar{position:fixed;bottom:24px;left:50%;transform:translateX(-50%);background:#1e293b;color:#fff;border-radius:14px;padding:12px 22px;display:flex;align-items:center;gap:14px;box-shadow:0 8px 30px rgba(0,0,0,.25);z-index:999;transition:.2s;opacity:0;pointer-events:none;font-size:14px}
#zip-bar.visible{opacity:1;pointer-events:auto}
</style>

<!-- Toolbar FUORI dal form zip -->
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;flex-wrap:wrap;gap:10px">
  <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap">
    <form method="GET" style="display:inline-flex;gap:8px;margin:0">
      <select name="cat" onchange="this.form.submit()" style="padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px">
        <option value="">Tutte le categorie</option>
        {% for c in categorie %}<option value="{{ c }}" {{ 'selected' if filtro_cat==c }}>{{ c }}</option>{% endfor %}
      </select>
    </form>
    <button type="button" onclick="selezionaTutti()" class="btn btn-secondary btn-sm"><i class="fa fa-check-double"></i> Seleziona tutti</button>
    <button type="button" onclick="deselezionaTutti()" class="btn btn-secondary btn-sm"><i class="fa fa-xmark"></i> Deseleziona</button>
  </div>
  <a href="/documenti-azienda/nuovo" class="btn btn-primary"><i class="fa fa-plus"></i> Nuovo documento</a>
</div>

{% if scaduti or scad_imminenti %}
<div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px">
  {% if scaduti %}
  <div style="background:#fef2f2;border:1px solid #fecaca;border-radius:10px;padding:12px 18px;display:flex;align-items:center;gap:10px">
    <i class="fa fa-circle-exclamation" style="color:#dc2626;font-size:20px"></i>
    <div><div style="font-weight:700;color:#dc2626">{{ scaduti }} document{% if scaduti>1 %}i scaduti{% else %}o scaduto{% endif %}</div>
    <div style="font-size:12px;color:#b91c1c">Rinnovo urgente</div></div>
  </div>
  {% endif %}
  {% if scad_imminenti %}
  <div style="background:#fffbeb;border:1px solid #fde68a;border-radius:10px;padding:12px 18px;display:flex;align-items:center;gap:10px">
    <i class="fa fa-triangle-exclamation" style="color:#f59e0b;font-size:20px"></i>
    <div><div style="font-weight:700;color:#92400e">{{ scad_imminenti }} document{% if scad_imminenti>1 %}i in scadenza{% else %}o in scadenza{% endif %}</div>
    <div style="font-size:12px;color:#b45309">Entro 30 giorni</div></div>
  </div>
  {% endif %}
</div>
{% endif %}

<!-- Form ZIP — contiene SOLO le checkbox, nessun form annidato -->
<form id="form-zip" method="POST" action="/documenti-azienda/scarica-zip">

{% if not docs %}
<div class="empty-state"><i class="fa fa-building-columns"></i><p>Nessun documento aziendale. <a href="/documenti-azienda/nuovo">Aggiungi il primo</a></p></div>
{% else %}

{% set docs_scaduti = docs|selectattr('stato_scad','equalto','scaduto')|list %}
{% if docs_scaduti %}
<div class="section-header"><i class="fa fa-circle-exclamation" style="color:#dc2626"></i><h3 style="color:#dc2626">Scaduti</h3><span style="background:#fef2f2;color:#dc2626;border-radius:20px;padding:1px 10px;font-size:12px;font-weight:600">{{ docs_scaduti|length }}</span></div>
<div class="doc-grid">{% for d in docs_scaduti %}{{ doc_card(d) }}{% endfor %}</div>
{% endif %}

{% set docs_warn = docs|selectattr('stato_scad','equalto','warning')|list %}
{% if docs_warn %}
<div class="section-header"><i class="fa fa-triangle-exclamation" style="color:#f59e0b"></i><h3 style="color:#92400e">In scadenza</h3><span style="background:#fffbeb;color:#92400e;border-radius:20px;padding:1px 10px;font-size:12px;font-weight:600">{{ docs_warn|length }}</span></div>
<div class="doc-grid">{% for d in docs_warn %}{{ doc_card(d) }}{% endfor %}</div>
{% endif %}

{% set docs_ok = docs|selectattr('stato_scad','equalto','ok')|list %}
{% if docs_ok %}
<div class="section-header"><i class="fa fa-circle-check" style="color:#16a34a"></i><h3 style="color:#15803d">Validi</h3><span style="background:#dcfce7;color:#15803d;border-radius:20px;padding:1px 10px;font-size:12px;font-weight:600">{{ docs_ok|length }}</span></div>
<div class="doc-grid">{% for d in docs_ok %}{{ doc_card(d) }}{% endfor %}</div>
{% endif %}

{% set docs_none = docs|selectattr('stato_scad','equalto','none')|list %}
{% if docs_none %}
<div class="section-header"><i class="fa fa-infinity" style="color:#94a3b8"></i><h3 style="color:#64748b">Senza scadenza</h3><span style="background:#f1f5f9;color:#64748b;border-radius:20px;padding:1px 10px;font-size:12px;font-weight:600">{{ docs_none|length }}</span></div>
<div class="doc-grid">{% for d in docs_none %}{{ doc_card(d) }}{% endfor %}</div>
{% endif %}

{% endif %}
</form>

<!-- Barra ZIP flottante -->
<div id="zip-bar">
  <i class="fa fa-file-zipper" style="font-size:18px;color:#a78bfa"></i>
  <span id="zip-count">0 documenti selezionati</span>
  <button type="button" onclick="scaricaZip()" style="background:#7c3aed;color:#fff;border:none;border-radius:8px;padding:8px 18px;font-size:13px;font-weight:700;cursor:pointer"><i class="fa fa-download"></i> Scarica ZIP</button>
  <button type="button" onclick="deselezionaTutti()" style="background:rgba(255,255,255,.15);color:#fff;border:none;border-radius:8px;padding:8px 12px;font-size:13px;cursor:pointer"><i class="fa fa-xmark"></i></button>
</div>

<script>
function aggiornaZip(){
  const sel = document.querySelectorAll('.doc-chk:checked').length;
  const bar = document.getElementById('zip-bar');
  const cnt = document.getElementById('zip-count');
  cnt.textContent = sel + ' document' + (sel===1?'o':'i') + ' selezionat' + (sel===1?'o':'i');
  bar.classList.toggle('visible', sel > 0);
}
function selezionaTutti(){
  document.querySelectorAll('.doc-chk').forEach(c=>c.checked=true);
  aggiornaZip();
}
function deselezionaTutti(){
  document.querySelectorAll('.doc-chk').forEach(c=>c.checked=false);
  aggiornaZip();
}
function scaricaZip(){
  document.getElementById('form-zip').submit();
}
</script>
"""

DOCS_AZ_FORM_TMPL = """
{% if ai_error %}
<div style="background:#fffbeb;border:1px solid #fde68a;border-radius:10px;padding:14px 18px;margin-bottom:20px;display:flex;align-items:flex-start;gap:12px">
  <i class="fa fa-triangle-exclamation" style="color:#f59e0b;font-size:20px;margin-top:2px;flex-shrink:0"></i>
  <div>
    <div style="font-weight:700;color:#92400e;margin-bottom:4px">L'AI non è riuscita a leggere il documento</div>
    <div style="font-size:13px;color:#78350f">Il file è stato caricato correttamente e verrà salvato. Compila i campi manualmente qui sotto.</div>
  </div>
</div>
{% elif ai_result %}
<div style="background:#f5f3ff;border:1px solid #c4b5fd;border-radius:10px;padding:14px 18px;margin-bottom:20px;display:flex;align-items:flex-start;gap:12px">
  <i class="fa fa-wand-magic-sparkles" style="color:#7c3aed;font-size:20px;margin-top:2px;flex-shrink:0"></i>
  <div>
    <div style="font-weight:700;color:#5b21b6;margin-bottom:6px">Dati estratti dall'AI ✨</div>
    {% if ai_result.get('sommario') %}<div style="font-size:13px;color:#6b7280;margin-bottom:8px">{{ ai_result.get('sommario') }}</div>{% endif %}
    <div style="display:flex;flex-wrap:wrap;gap:6px">
      {% for k,v in ai_result.items() %}{% if v and k not in ('sommario','_error') %}
      <span style="background:#ede9fe;color:#5b21b6;padding:3px 10px;border-radius:6px;font-size:12px"><strong>{{ k }}:</strong> {{ v }}</span>
      {% endif %}{% endfor %}
    </div>
    <div style="font-size:12px;color:#7c3aed;margin-top:8px"><i class="fa fa-info-circle"></i> I campi sono stati pre-compilati, verifica e modifica se necessario.</div>
  </div>
</div>
{% endif %}

<!-- Box upload PDF -->
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-file-arrow-up" style="color:#7c3aed"></i> Carica PDF (analisi AI automatica)</h3></div>
  <div class="card-body">
    {% if file_nome_esistente %}
    <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;padding:10px 14px;margin-bottom:12px;display:flex;align-items:center;gap:10px">
      <i class="fa fa-file-pdf" style="color:#16a34a;font-size:18px"></i>
      <div style="flex:1">
        <div style="font-weight:600;color:#15803d;font-size:13px">File caricato: {{ file_nome_esistente }}</div>
        <div style="font-size:11px;color:#64748b">Il file verrà salvato insieme al documento</div>
      </div>
      <a href="{{ form_action }}?rimuovi_file=1" style="color:#94a3b8;font-size:12px;text-decoration:none" title="Rimuovi file"><i class="fa fa-xmark"></i> rimuovi</a>
    </div>
    {% endif %}
    <form method="POST" enctype="multipart/form-data" action="{{ form_action }}?upload=1" id="form-upload">
      <div style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap">
        <div class="form-group" style="flex:1;margin:0">
          <label>{{ 'Sostituisci PDF' if file_nome_esistente else 'Carica PDF' }}</label>
          <input type="file" name="file_doc" accept=".pdf" style="padding:8px" onchange="this.form.submit()">
        </div>
        <div style="font-size:12px;color:#94a3b8;padding-bottom:8px"><i class="fa fa-robot"></i> L'AI proverà a leggere titolo, categoria e scadenza</div>
      </div>
    </form>
  </div>
</div>

<!-- Form dati -->
<form method="POST" action="{{ form_action }}" enctype="multipart/form-data">
  {% if file_nome_esistente %}<input type="hidden" name="file_nome_esistente" value="{{ file_nome_esistente }}">{% endif %}
  {% if file_path_esistente %}<input type="hidden" name="file_path_esistente" value="{{ file_path_esistente }}">{% endif %}
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px">
    <div class="card">
      <div class="card-header"><h3><i class="fa fa-file-lines"></i> Dati documento</h3></div>
      <div class="card-body">
        <div class="form-group">
          <label>Titolo *</label>
          <input type="text" name="titolo" value="{{ doc.titolo if doc else ai_pre.get('titolo','') }}" required class="form-control" placeholder="es. DURC Marzo 2026">
        </div>
        <div class="form-group">
          <label>Categoria</label>
          <select name="categoria" class="form-control">
            {% for c in categorie %}
            <option value="{{ c }}" {{ 'selected' if (doc and doc.categoria==c) or ai_pre.get('categoria','')==c }}>{{ c }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="form-group">
          <label>Note</label>
          <textarea name="note" class="form-control" rows="3" placeholder="Note aggiuntive...">{{ doc.note if doc else '' }}</textarea>
        </div>
        <div class="form-group">
          <label>Avviso scadenza (giorni prima)</label>
          <input type="number" name="avviso_giorni" value="{{ doc.avviso_giorni if doc else 30 }}" min="0" max="365" class="form-control">
        </div>
      </div>
    </div>
    <div class="card">
      <div class="card-header"><h3><i class="fa fa-calendar-days"></i> Date</h3></div>
      <div class="card-body">
        <div class="form-group">
          <label>Data emissione</label>
          <input type="date" name="data_emissione" value="{{ doc.data_emissione if doc else ai_pre.get('data_emissione','') }}" class="form-control">
        </div>
        <div class="form-group">
          <label>Data scadenza <span style="font-size:11px;color:#94a3b8">(lascia vuoto se non scade)</span></label>
          <input type="date" name="data_scadenza" value="{{ doc.data_scadenza if doc else ai_pre.get('data_scadenza','') }}" class="form-control" style="border-color:#f59e0b">
        </div>
        {% if doc and doc.file_nome and not file_nome_esistente %}
        <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;padding:10px;font-size:13px;margin-bottom:12px">
          <i class="fa fa-paperclip" style="color:#16a34a"></i> File attuale: <strong>{{ doc.file_nome }}</strong>
          <div style="font-size:11px;color:#64748b;margin-top:4px">Carica un nuovo file sopra per sostituirlo</div>
        </div>
        {% endif %}
      </div>
    </div>
  </div>
  <div style="display:flex;justify-content:flex-end;gap:12px;margin-top:20px">
    <a href="/documenti-azienda" class="btn btn-secondary">Annulla</a>
    <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva documento</button>
  </div>
</form>
"""


def _analizza_doc_azienda_ai(file_path):
    """Estrae dati da PDF documento aziendale con AI."""
    api_key = get_setting('anthropic_api_key', '')
    if not api_key:
        return None
    try:
        import anthropic, base64 as _b64
        client = anthropic.Anthropic(api_key=api_key)
        with open(file_path, 'rb') as f:
            pdf_data = _b64.standard_b64encode(f.read()).decode()
        cb = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_data}}
        prompt = """Analizza questo documento aziendale e rispondi SOLO con JSON valido, senza markdown.
Estrai: titolo (nome del documento), categoria (una tra: DURC, Visura camerale, Assicurazione, Certificato antimafia, Licenza, Contratto, Autorizzazione, Certificazione, Patente, Iscrizione albo, Altro), data_emissione (YYYY-MM-DD), data_scadenza (YYYY-MM-DD), sommario (1 frase descrittiva).
Formato esatto: {"titolo":"","categoria":"","data_emissione":"","data_scadenza":"","sommario":""}
Se un campo non esiste metti "". Date in formato YYYY-MM-DD."""
        resp = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=512,
                                      messages=[{"role": "user", "content": [cb, {"type": "text", "text": prompt}]}])
        raw = resp.content[0].text.strip().strip('`').lstrip('json').strip()
        return json.loads(raw)
    except Exception as e:
        return {'_error': str(e)}


def _doc_az_stato(data_scadenza, avviso_giorni=30):
    if not data_scadenza:
        return 'none', None
    try:
        scad = date.fromisoformat(data_scadenza)
        oggi = date.today()
        diff = (scad - oggi).days
        if diff < 0:
            return 'scaduto', diff
        elif diff <= avviso_giorni:
            return 'warning', diff
        else:
            return 'ok', diff
    except:
        return 'none', None


@app.route('/documenti-azienda')
@admin_required
def documenti_azienda():
    db = get_db()
    filtro_cat = request.args.get('cat', '')
    sql = "SELECT * FROM documenti_azienda"
    params = []
    if filtro_cat:
        sql += " WHERE categoria=?"
        params.append(filtro_cat)
    sql += " ORDER BY CASE WHEN data_scadenza IS NULL THEN 1 ELSE 0 END, data_scadenza ASC"
    rows = db.execute(sql, params).fetchall()
    db.close()

    docs = []
    scaduti = scad_imminenti = 0
    for r in rows:
        d = dict(r)
        stato, giorni = _doc_az_stato(d.get('data_scadenza'), d.get('avviso_giorni') or 30)
        d['stato_scad'] = stato
        d['giorni_alla_scad'] = giorni
        if stato == 'scaduto': scaduti += 1
        elif stato == 'warning': scad_imminenti += 1
        docs.append(d)

    return render_page(DOCS_AZ_TMPL, page_title='Documenti Azienda', active='documenti_azienda',
        docs=docs, categorie=DOCS_AZ_CATEGORIE, filtro_cat=filtro_cat,
        scaduti=scaduti, scad_imminenti=scad_imminenti)


@app.route('/documenti-azienda/nuovo', methods=['GET', 'POST'])
@admin_required
def documenti_azienda_nuovo():
    # Rimuovi file da session se richiesto
    if request.args.get('rimuovi_file') == '1':
        session.pop('docs_az_ai', None)
        session.pop('docs_az_file_nome', None)
        session.pop('docs_az_file_path', None)
        return redirect(url_for('documenti_azienda_nuovo'))

    ai_result = session.pop('docs_az_ai', None)
    ai_error = session.pop('docs_az_ai_error', False)
    ai_pre = ai_result or {}
    file_nome_esistente = session.get('docs_az_file_nome', '')
    file_path_esistente = session.get('docs_az_file_path', '')

    if request.method == 'POST':
        # Upload + analisi AI
        if request.args.get('upload') == '1':
            f_file = request.files.get('file_doc')
            if f_file and f_file.filename:
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                fn = f"{ts}_{secure_filename(f_file.filename)}"
                fp = os.path.join(get_docs_azienda_path(), fn)
                f_file.save(fp)
                ai_data = _analizza_doc_azienda_ai(fp)
                if ai_data and '_error' not in ai_data and any(ai_data.values()):
                    session['docs_az_ai'] = ai_data
                    session['docs_az_ai_error'] = False
                else:
                    session['docs_az_ai'] = {}
                    session['docs_az_ai_error'] = True  # file tenuto, AI fallita
                session['docs_az_file_nome'] = fn
                session['docs_az_file_path'] = fp
            return redirect(url_for('documenti_azienda_nuovo'))

        # Salvataggio definitivo
        f = request.form
        fn = f.get('file_nome_esistente') or ''
        fp2 = f.get('file_path_esistente') or ''
        f_file2 = request.files.get('file_doc')
        if f_file2 and f_file2.filename:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            fn = f"{ts}_{secure_filename(f_file2.filename)}"
            fp2 = os.path.join(get_docs_azienda_path(), fn)
            f_file2.save(fp2)
        session.pop('docs_az_ai', None)
        session.pop('docs_az_ai_error', None)
        session.pop('docs_az_file_nome', None)
        session.pop('docs_az_file_path', None)
        db = get_db()
        db.execute("""INSERT INTO documenti_azienda
            (titolo, categoria, note, data_emissione, data_scadenza, avviso_giorni, file_nome, file_path)
            VALUES (?,?,?,?,?,?,?,?)""", (
            f.get('titolo',''), f.get('categoria',''), f.get('note',''),
            f.get('data_emissione') or None, f.get('data_scadenza') or None,
            int(f.get('avviso_giorni') or 30), fn or None, fp2 or None))
        safe_commit(db); db.close()
        flash('Documento salvato!', 'success')
        return redirect(url_for('documenti_azienda'))

    return render_page(DOCS_AZ_FORM_TMPL, page_title='Nuovo documento azienda', active='documenti_azienda',
        doc=None, ai_result=ai_result if ai_result else None,
        ai_error=ai_error, ai_pre=ai_pre,
        file_nome_esistente=file_nome_esistente,
        file_path_esistente=file_path_esistente,
        form_action='/documenti-azienda/nuovo', categorie=DOCS_AZ_CATEGORIE)


@app.route('/documenti-azienda/<int:did>/modifica', methods=['GET', 'POST'])
@admin_required
def documenti_azienda_modifica(did):
    # Rimuovi file da session se richiesto
    if request.args.get('rimuovi_file') == '1':
        session.pop('docs_az_ai', None)
        session.pop('docs_az_file_nome', None)
        session.pop('docs_az_file_path', None)
        return redirect(url_for('documenti_azienda_modifica', did=did))

    db = get_db()
    doc = db.execute("SELECT * FROM documenti_azienda WHERE id=?", (did,)).fetchone()
    if not doc:
        db.close(); flash('Non trovato.', 'error'); return redirect(url_for('documenti_azienda'))

    if request.method == 'POST':
        if request.args.get('upload') == '1':
            f_file = request.files.get('file_doc')
            if f_file and f_file.filename:
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                fn = f"{ts}_{secure_filename(f_file.filename)}"
                fp = os.path.join(get_docs_azienda_path(), fn)
                f_file.save(fp)
                ai_data = _analizza_doc_azienda_ai(fp)
                if ai_data and '_error' not in ai_data and any(ai_data.values()):
                    session['docs_az_ai'] = ai_data
                    session['docs_az_ai_error'] = False
                else:
                    session['docs_az_ai'] = {}
                    session['docs_az_ai_error'] = True
                session['docs_az_file_nome'] = fn
                session['docs_az_file_path'] = fp
            db.close()
            return redirect(url_for('documenti_azienda_modifica', did=did))

        f = request.form
        fn = doc['file_nome']
        fp2 = doc['file_path']
        f_file2 = request.files.get('file_doc')
        if f_file2 and f_file2.filename:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            fn = f"{ts}_{secure_filename(f_file2.filename)}"
            fp2 = os.path.join(get_docs_azienda_path(), fn)
            f_file2.save(fp2)
        elif f.get('file_nome_esistente'):
            fn = f.get('file_nome_esistente')
            fp2 = f.get('file_path_esistente') or fp2
        session.pop('docs_az_ai', None)
        session.pop('docs_az_ai_error', None)
        session.pop('docs_az_file_nome', None)
        session.pop('docs_az_file_path', None)
        db.execute("""UPDATE documenti_azienda SET
            titolo=?, categoria=?, note=?, data_emissione=?, data_scadenza=?,
            avviso_giorni=?, file_nome=?, file_path=?, aggiornato_il=datetime('now')
            WHERE id=?""", (
            f.get('titolo',''), f.get('categoria',''), f.get('note',''),
            f.get('data_emissione') or None, f.get('data_scadenza') or None,
            int(f.get('avviso_giorni') or 30), fn, fp2, did))
        safe_commit(db); db.close()
        flash('Documento aggiornato!', 'success')
        return redirect(url_for('documenti_azienda'))

    ai_result = session.pop('docs_az_ai', None)
    ai_error = session.pop('docs_az_ai_error', False)
    ai_pre = ai_result or {}
    file_nome_esistente = session.get('docs_az_file_nome', '')
    file_path_esistente = session.get('docs_az_file_path', '')
    db.close()
    return render_page(DOCS_AZ_FORM_TMPL, page_title='Modifica documento', active='documenti_azienda',
        doc=dict(doc), ai_result=ai_result if ai_result else None,
        ai_error=ai_error, ai_pre=ai_pre,
        file_nome_esistente=file_nome_esistente,
        file_path_esistente=file_path_esistente,
        form_action=f'/documenti-azienda/{did}/modifica', categorie=DOCS_AZ_CATEGORIE)


@app.route('/documenti-azienda/<int:did>/scarica')
@admin_required
def documenti_azienda_scarica(did):
    db = get_db()
    doc = db.execute("SELECT * FROM documenti_azienda WHERE id=?", (did,)).fetchone()
    db.close()
    if not doc or not doc['file_path'] or not os.path.exists(doc['file_path']):
        flash('File non trovato.', 'error')
        return redirect(url_for('documenti_azienda'))
    return send_file(doc['file_path'], as_attachment=True, download_name=doc['file_nome'])


@app.route('/documenti-azienda/<int:did>/elimina')
@admin_required
def documenti_azienda_elimina(did):
    db = get_db()
    db.execute("DELETE FROM documenti_azienda WHERE id=?", (did,))
    safe_commit(db); db.close()
    flash('Documento eliminato.', 'success')
    return redirect(url_for('documenti_azienda'))


@app.route('/documenti-azienda/scarica-zip', methods=['POST'])
@admin_required
def documenti_azienda_zip():
    ids = request.form.getlist('ids')
    if not ids:
        flash('Nessun documento selezionato.', 'error')
        return redirect(url_for('documenti_azienda'))
    db = get_db()
    rows = db.execute(
        f"SELECT * FROM documenti_azienda WHERE id IN ({','.join('?'*len(ids))})", ids
    ).fetchall()
    db.close()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        nomi_usati = {}
        for r in rows:
            if not r['file_path'] or not os.path.exists(r['file_path']):
                continue
            nome_base = r['titolo'] or r['file_nome'] or f"doc_{r['id']}"
            # Sanifica nome file
            nome_base = "".join(c if c.isalnum() or c in ' ._-' else '_' for c in nome_base).strip()
            ext = os.path.splitext(r['file_nome'] or '.pdf')[1]
            nome = nome_base + ext
            # Evita duplicati
            if nome in nomi_usati:
                nomi_usati[nome] += 1
                nome = f"{nome_base}_{nomi_usati[nome]}{ext}"
            else:
                nomi_usati[nome] = 1
            zf.write(r['file_path'], arcname=nome)

    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"documenti_azienda_{date.today().isoformat()}.zip",
                     mimetype='application/zip')


# ── Template lista fatture ─────────────────────────────────
FATT_LIST_TMPL = """
<style>
.stat-fatt{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:24px}
.sf{background:#fff;border-radius:12px;padding:16px;border:1px solid var(--border);text-align:center}
.sf .val{font-size:26px;font-weight:800;margin-bottom:4px}
.sf .lbl{font-size:12px;color:var(--text-light)}
.sf.rosso .val{color:#dc2626}.sf.verde .val{color:#16a34a}.sf.amber .val{color:#d97706}.sf.blu .val{color:#2563eb}
.badge-fatt{padding:4px 10px;border-radius:99px;font-size:11px;font-weight:700}
.bf-da{background:#fef3c7;color:#92400e}.bf-parz{background:#dbeafe;color:#1e40af}
.bf-pag{background:#dcfce7;color:#15803d}.bf-scad{background:#fee2e2;color:#991b1b}
.bf-annul{background:#f1f5f9;color:#64748b}
.progress-bar{background:#e2e8f0;border-radius:99px;height:6px;overflow:hidden;margin-top:4px}
.progress-fill{background:linear-gradient(90deg,#2563eb,#22c55e);height:100%;border-radius:99px;transition:.3s}
</style>

<!-- Statistiche -->
<div class="stat-fatt">
  <div class="sf rosso">
    <div class="val">€ {{ "%.0f"|format(tot_da_pagare) }}</div>
    <div class="lbl">📋 Da incassare</div>
  </div>
  <div class="sf amber">
    <div class="val">€ {{ "%.0f"|format(tot_parziale) }}</div>
    <div class="lbl">⏳ Parzialmente pagate</div>
  </div>
  <div class="sf verde">
    <div class="val">€ {{ "%.0f"|format(tot_pagato) }}</div>
    <div class="lbl">✅ Incassato totale</div>
  </div>
  <div class="sf blu">
    <div class="val">{{ n_scadute }}</div>
    <div class="lbl">⚠️ Scadute</div>
  </div>
</div>

<!-- Grafici -->
<div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:24px">
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-chart-pie" style="color:var(--accent2)"></i> Stato fatture</h3></div>
    <div class="card-body" style="display:flex;justify-content:center">
      <canvas id="chartStato" width="260" height="260"></canvas>
    </div>
  </div>
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-chart-bar" style="color:var(--accent2)"></i> Incassi mensili (ultimi 6 mesi)</h3></div>
    <div class="card-body">
      <canvas id="chartMensile" width="400" height="260"></canvas>
    </div>
  </div>
</div>

<!-- Top clienti -->
{% if top_clienti %}
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-trophy" style="color:#f59e0b"></i> Clienti più puntuali</h3></div>
  <div class="card-body">
    {% for c in top_clienti %}
    <div style="display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid var(--border)">
      <div>
        <strong>{{ c.nome }}</strong>
        <div style="font-size:12px;color:var(--text-light)">{{ c.n_pagate }} fatture pagate su {{ c.n_totale }}</div>
      </div>
      <div style="text-align:right">
        <div style="color:#16a34a;font-weight:700">{{ "%.0f"|format(c.perc) }}%</div>
        <div class="progress-bar" style="width:80px">
          <div class="progress-fill" style="width:{{ c.perc }}%"></div>
        </div>
      </div>
    </div>
    {% endfor %}
  </div>
</div>
{% endif %}

<!-- Azioni + filtri -->
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;flex-wrap:wrap;gap:10px">
  <div style="display:flex;gap:10px;flex-wrap:wrap">
    <a href="/fatturazione/nuova" class="btn btn-primary"><i class="fa fa-plus"></i> Nuova fattura</a>
    <a href="/fatturazione/clienti" class="btn btn-secondary"><i class="fa fa-users"></i> Clienti</a>
  </div>
  <form style="display:flex;gap:10px;flex-wrap:wrap" method="GET">
    <select name="stato" onchange="this.form.submit()" style="padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px">
      <option value="">Tutti gli stati</option>
      <option value="da_emettere" {{ 'selected' if filtro_stato=='da_emettere' }}>⏳ Da emettere</option>
      <option value="da_pagare" {{ 'selected' if filtro_stato=='da_pagare' }}>Da pagare</option>
      <option value="parziale" {{ 'selected' if filtro_stato=='parziale' }}>Parzialmente pagate</option>
      <option value="pagata" {{ 'selected' if filtro_stato=='pagata' }}>Pagate</option>
      <option value="scaduta" {{ 'selected' if filtro_stato=='scaduta' }}>Scadute</option>
    </select>
    <input type="text" name="q" value="{{ q or '' }}" placeholder="Cerca cliente/numero..." style="padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px;width:200px">
    <button type="submit" class="btn btn-secondary btn-sm"><i class="fa fa-search"></i></button>
  </form>
</div>


<!-- ═══ MACRO RIGA FATTURA ═══ -->
{% macro riga_fattura(f) %}
{% set has_rate = f.rate|length > 0 %}
{% set all_paid = has_rate and f.rate|selectattr('stato','ne','pagata')|list|length == 0 %}
<tr class="fatt-row" style="border-bottom:{% if has_rate and not all_paid %}none{% else %}1px solid var(--border){% endif %}">
  <td style="text-align:center;color:#94a3b8;font-size:11px">
    {% if has_rate and all_paid %}<span style="color:#16a34a;font-size:13px">✅</span>{% endif %}
  </td>
  <td><strong>{{ f.numero }}</strong></td>
  <td>{{ f.cliente_nome or '–' }}</td>
  <td style="font-family:monospace;font-size:12px">{{ f.data_emissione or '–' }}</td>
  <td style="font-family:monospace;font-size:12px;{{ 'color:#dc2626;font-weight:700' if f.scadenza_scaduta and f.stato_display != 'pagata' else '' }}">
    {{ f.scadenza_display or '–' }}{% if f.scadenza_scaduta and f.stato_display != 'pagata' %} ⚠️{% endif %}
  </td>
  <td><strong>€ {{ "%.2f"|format(f.importo_totale) }}</strong></td>
  <td style="color:#16a34a;font-weight:600">€ {{ "%.2f"|format(f.pagato) }}</td>
  <td style="{{ 'color:#dc2626;font-weight:600' if f.residuo > 0 else 'color:#64748b' }}">
    {% if f.residuo > 0 %}€ {{ "%.2f"|format(f.residuo) }}{% else %}–{% endif %}
  </td>
  <td>
    {% if f.stato_display == 'da_emettere' %}<span class="badge-fatt" style="background:#f3e8ff;color:#7c3aed">⏳ Da emettere</span>
    {% elif f.stato_display == 'pagata' %}<span class="badge-fatt bf-pag">✅ Pagata</span>
    {% elif f.stato_display == 'parziale' %}<span class="badge-fatt bf-parz">⏳ Parziale</span>
    {% elif f.stato_display == 'scaduta' %}<span class="badge-fatt bf-scad">⚠️ Scaduta</span>
    {% elif f.stato_display == 'annullata' %}<span class="badge-fatt bf-annul">✗ Annullata</span>
    {% else %}<span class="badge-fatt bf-da">📋 Da pagare</span>{% endif %}
  </td>
  <td style="display:flex;gap:6px;align-items:center">
    {% if f.stato_display == 'da_emettere' %}
    <label title="Allega PDF fattura reale" style="background:#7c3aed;color:#fff;border-radius:6px;padding:5px 8px;font-size:12px;cursor:pointer;display:flex;align-items:center;gap:4px;white-space:nowrap">
      <i class="fa fa-file-arrow-up"></i> Allega PDF
      <input type="file" accept=".pdf" style="display:none" onchange="allegaPdf(this, {{ f.id }})">
    </label>
    {% endif %}
    {% if f.file_nome %}
    <a href="/fatturazione/{{ f.id }}/file" class="btn btn-sm" style="background:#f0fdf4;color:#16a34a;border:1px solid #86efac" title="Scarica PDF fattura allegata"><i class="fa fa-file-pdf"></i></a>
    {% endif %}
    <a href="/fatturazione/{{ f.id }}" class="btn btn-sm btn-secondary" title="Dettaglio"><i class="fa fa-eye"></i></a>
    <a href="/fatturazione/{{ f.id }}/modifica" class="btn btn-sm btn-secondary" title="Modifica"><i class="fa fa-pen"></i></a>
    <a href="/fatturazione/{{ f.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare la fattura {{ f.numero }}?')" title="Elimina"><i class="fa fa-trash"></i></a>
  </td>
</tr>
{% if f.stato_display == 'da_emettere' %}
<tr id="upload-row-{{ f.id }}" style="display:none;border-bottom:2px solid #7c3aed">
  <td colspan="10" style="padding:8px 16px;background:#faf5ff">
    <div id="upload-progress-{{ f.id }}" style="font-size:13px;color:#7c3aed;display:flex;align-items:center;gap:8px">
      <i class="fa fa-spinner fa-spin"></i> Caricamento e analisi AI in corso...
    </div>
  </td>
</tr>
{% endif %}
{% if has_rate and not all_paid %}
<tr style="border-bottom:2px solid #e2e8f0">
  <td colspan="10" style="padding:0;background:#f8fafc">
    <table style="width:100%;border-collapse:collapse;font-size:12px">
      <thead>
        <tr style="background:#f1f5f9;color:#64748b">
          <th style="padding:5px 10px 5px 40px;text-align:left;font-weight:600">RATA</th>
          <th style="padding:5px 10px;text-align:right;font-weight:600">IMPORTO</th>
          <th style="padding:5px 10px;text-align:center;font-weight:600">SCADENZA</th>
          <th style="padding:5px 10px;text-align:center;font-weight:600">PAGATA IL</th>
          <th style="padding:5px 10px;text-align:center;font-weight:600">STATO</th>
          <th style="padding:5px 10px;text-align:left;font-weight:600">NOTE</th>
          <th style="padding:5px 10px"></th>
        </tr>
      </thead>
      <tbody>
      {% for r in f.rate %}{% if r.stato != 'pagata' %}
      <tr style="border-top:1px solid #e2e8f0;background:{% if r.scaduta %}#fef2f2{% else %}#fff{% endif %}">
        <td style="padding:7px 10px 7px 40px;font-weight:600">Rata {{ r.numero_rata }}</td>
        <td style="padding:7px 10px;text-align:right;font-weight:700">€ {{ "%.2f"|format(r.importo) }}</td>
        <td style="padding:7px 10px;text-align:center;font-family:monospace;{{ 'color:#dc2626;font-weight:700' if r.scaduta else '' }}">
          {{ r.data_scadenza or '–' }}{% if r.scaduta %} ⚠️{% endif %}
        </td>
        <td style="padding:7px 10px;text-align:center;color:#94a3b8">–</td>
        <td style="padding:7px 10px;text-align:center">
          {% if r.scaduta %}<span class="badge-fatt bf-scad" style="font-size:11px">⚠️ Scaduta</span>
          {% else %}<span class="badge-fatt bf-da" style="font-size:11px">📋 Da pagare</span>{% endif %}
        </td>
        <td style="padding:7px 10px;color:#94a3b8">{{ r.note or '–' }}</td>
        <td style="padding:7px 10px">
          <form method="POST" action="/fatturazione/rata/{{ r.id }}/paga" style="display:flex;gap:4px;align-items:center">
            <input type="date" name="data_pag" value="{{ today }}" style="padding:3px 6px;font-size:11px;border:1px solid var(--border);border-radius:6px;width:110px">
            <button type="submit" class="btn btn-sm" style="background:#dcfce7;color:#16a34a;border:1px solid #bbf7d0;font-size:11px;padding:4px 8px"><i class="fa fa-check"></i> Pagata</button>
          </form>
        </td>
      </tr>
      {% endif %}{% endfor %}
      </tbody>
    </table>
  </td>
</tr>
{% endif %}
{% endmacro %}

<!-- ═══ THEAD COMUNE ═══ -->
{% macro thead() %}
<thead><tr style="background:#0f172a;color:#fff">
  <th style="width:24px;padding:10px 8px"></th>
  <th style="padding:10px 8px;font-size:12px">N°</th>
  <th style="padding:10px 8px;font-size:12px">CLIENTE</th>
  <th style="padding:10px 8px;font-size:12px">DATA</th>
  <th style="padding:10px 8px;font-size:12px">SCADENZA</th>
  <th style="padding:10px 8px;font-size:12px">IMPORTO</th>
  <th style="padding:10px 8px;font-size:12px">PAGATO</th>
  <th style="padding:10px 8px;font-size:12px">RESIDUO</th>
  <th style="padding:10px 8px;font-size:12px">STATO</th>
  <th style="padding:10px 8px"></th>
</tr></thead>
{% endmacro %}

<!-- ═══ SEZIONE 1: Da pagare con rate ═══ -->
{% set con_rate = fatture | selectattr('rate') | selectattr('stato_display', 'ne', 'pagata') | list %}
{% set con_rate = con_rate | selectattr('rate') | list %}
{% set con_rate_attive = [] %}
{% for f in fatture %}
  {% if f.rate|length > 0 and f.stato_display != 'pagata' %}
    {% if con_rate_attive.append(f) %}{% endif %}
  {% endif %}
{% endfor %}

{% if con_rate_attive %}
<div style="margin-bottom:6px;display:flex;align-items:center;gap:10px">
  <h3 style="margin:0;font-size:15px;color:#1e293b"><i class="fa fa-list-check" style="color:#f59e0b;margin-right:6px"></i>Da pagare — con rate</h3>
  <span style="background:#fef3c7;color:#92400e;border-radius:20px;padding:2px 10px;font-size:12px;font-weight:600">{{ con_rate_attive|length }}</span>
</div>
<div class="card" style="margin-bottom:24px;border-top:3px solid #f59e0b">
  <div class="table-wrap"><table style="width:100%;border-collapse:collapse">
    {{ thead() }}
    <tbody>
    {% for f in con_rate_attive %}{{ riga_fattura(f) }}{% endfor %}
    </tbody>
  </table></div>
</div>
{% endif %}

<!-- ═══ SEZIONE 2: Da pagare / Da emettere senza rate ═══ -->
{% set senza_rate_attive = [] %}
{% for f in fatture %}
  {% if f.rate|length == 0 and f.stato_display not in ['pagata','annullata'] %}
    {% if senza_rate_attive.append(f) %}{% endif %}
  {% endif %}
{% endfor %}

{% if senza_rate_attive %}
<div style="margin-bottom:6px;display:flex;align-items:center;gap:10px">
  <h3 style="margin:0;font-size:15px;color:#1e293b"><i class="fa fa-file-invoice-dollar" style="color:#3b82f6;margin-right:6px"></i>Da pagare — senza rate</h3>
  <span style="background:#dbeafe;color:#1e40af;border-radius:20px;padding:2px 10px;font-size:12px;font-weight:600">{{ senza_rate_attive|length }}</span>
</div>
<div class="card" style="margin-bottom:24px;border-top:3px solid #3b82f6">
  <div class="table-wrap"><table style="width:100%;border-collapse:collapse">
    {{ thead() }}
    <tbody>
    {% for f in senza_rate_attive %}{{ riga_fattura(f) }}{% endfor %}
    </tbody>
  </table></div>
</div>
{% endif %}

<!-- ═══ SEZIONE 3: Pagate ═══ -->
{% set pagate = [] %}
{% for f in fatture %}
  {% if f.stato_display in ['pagata','annullata'] %}
    {% if pagate.append(f) %}{% endif %}
  {% endif %}
{% endfor %}

{% if pagate %}
<div style="margin-bottom:6px;display:flex;align-items:center;gap:10px">
  <button onclick="togglePagate()" id="btn-pagate" style="background:none;border:none;cursor:pointer;display:flex;align-items:center;gap:8px;padding:0">
    <h3 style="margin:0;font-size:15px;color:#64748b"><i class="fa fa-circle-check" style="color:#16a34a;margin-right:6px"></i>Completamente pagate</h3>
    <span style="background:#dcfce7;color:#15803d;border-radius:20px;padding:2px 10px;font-size:12px;font-weight:600">{{ pagate|length }}</span>
    <span id="ico-pagate" style="color:#94a3b8;font-size:12px">▶ mostra</span>
  </button>
</div>
<div id="sezione-pagate" style="display:none;margin-bottom:24px">
  <div class="card" style="border-top:3px solid #16a34a">
    <div class="table-wrap"><table style="width:100%;border-collapse:collapse">
      {{ thead() }}
      <tbody>
      {% for f in pagate %}{{ riga_fattura(f) }}{% endfor %}
      </tbody>
    </table></div>
  </div>
</div>
{% endif %}

{% if not con_rate_attive and not senza_rate_attive and not pagate %}
<div class="card"><div class="empty-state"><i class="fa fa-file-invoice"></i><p>Nessuna fattura trovata.</p></div></div>
{% endif %}

<script>
function togglePagate() {
  const s = document.getElementById('sezione-pagate');
  const ico = document.getElementById('ico-pagate');
  const open = s.style.display !== 'none';
  s.style.display = open ? 'none' : 'block';
  ico.textContent = open ? '▶ mostra' : '▼ nascondi';
}
function allegaPdf(input, fid) {
  if (!input.files || !input.files[0]) return;
  const row = document.getElementById('upload-row-' + fid);
  if (row) row.style.display = 'table-row';
  const fd = new FormData();
  fd.append('file_fattura', input.files[0]);
  fetch('/fatturazione/' + fid + '/allega-emessa', { method: 'POST', body: fd })
    .then(() => { window.location.reload(); })
    .catch(() => { window.location.reload(); });
}
</script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<script>
// Grafico stato
const ctx1 = document.getElementById('chartStato').getContext('2d');
new Chart(ctx1, {
  type: 'doughnut',
  data: {
    labels: ['Da pagare','Parziale','Pagate','Scadute'],
    datasets: [{
      data: [{{ n_da_pagare }}, {{ n_parziale }}, {{ n_pagate }}, {{ n_scadute_c }}],
      backgroundColor: ['#fbbf24','#60a5fa','#4ade80','#f87171'],
      borderWidth: 2, borderColor: '#fff'
    }]
  },
  options: { plugins: { legend: { position: 'bottom' } }, cutout: '65%' }
});

// Grafico mensile
const ctx2 = document.getElementById('chartMensile').getContext('2d');
new Chart(ctx2, {
  type: 'bar',
  data: {
    labels: {{ mesi_labels | tojson }},
    datasets: [
      { label: 'Fatturato', data: {{ mesi_fatturato | tojson }}, backgroundColor: 'rgba(37,99,235,.2)', borderColor: '#2563eb', borderWidth: 2, borderRadius: 6 },
      { label: 'Incassato', data: {{ mesi_incassato | tojson }}, backgroundColor: 'rgba(34,197,94,.2)', borderColor: '#22c55e', borderWidth: 2, borderRadius: 6 }
    ]
  },
  options: { responsive: true, plugins: { legend: { position: 'top' } }, scales: { y: { beginAtZero: true, ticks: { callback: v => '€'+v } } } }
});
</script>
"""

# ── Template dettaglio fattura ─────────────────────────────
FATT_DETAIL_TMPL = """
<div style="margin-bottom:16px;display:flex;gap:10px">
  <a href="/fatturazione" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Fatture</a>
  <a href="/fatturazione/{{ f.id }}/modifica" class="btn btn-secondary btn-sm"><i class="fa fa-pen"></i> Modifica</a>
  {% if f.file_nome %}
  <a href="/fatturazione/{{ f.id }}/file" class="btn btn-secondary btn-sm"><i class="fa fa-download"></i> Scarica PDF</a>
  {% endif %}
</div>

<div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:20px">
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-file-invoice-dollar" style="color:var(--accent2)"></i> Fattura {{ f.numero }}</h3></div>
    <div class="card-body">
      <table style="width:100%;font-size:14px"><tbody>
        <tr><td style="color:var(--text-light);padding:5px 0;width:45%">Cliente</td><td><strong>{{ f.cliente_nome }}</strong></td></tr>
        <tr><td style="color:var(--text-light);padding:5px 0">Data emissione</td><td>{{ f.data_emissione }}</td></tr>
        <tr><td style="color:var(--text-light);padding:5px 0">Scadenza</td><td>{{ f.data_scadenza }}</td></tr>
        <tr><td style="color:var(--text-light);padding:5px 0">Condizioni</td><td>{{ f.condizioni_pagamento }}</td></tr>
        <tr><td style="color:var(--text-light);padding:5px 0">Descrizione</td><td>{{ f.descrizione or '–' }}</td></tr>
      </tbody></table>
    </div>
  </div>
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-euro-sign" style="color:var(--accent2)"></i> Importi</h3></div>
    <div class="card-body">
      <table style="width:100%;font-size:14px"><tbody>
        <tr><td style="color:var(--text-light);padding:5px 0;width:45%">Imponibile</td><td>€ {{ "%.2f"|format(f.imponibile or 0) }}</td></tr>
        <tr><td style="color:var(--text-light);padding:5px 0">IVA ({{ f.iva_perc }}%)</td><td>€ {{ "%.2f"|format(f.iva_importo or 0) }}</td></tr>
        <tr><td style="color:var(--text-light);padding:5px 0"><strong>Totale</strong></td><td><strong style="font-size:20px">€ {{ "%.2f"|format(f.importo_totale) }}</strong></td></tr>
        <tr><td style="color:var(--text-light);padding:5px 0">Pagato</td><td style="color:#16a34a;font-weight:700">€ {{ "%.2f"|format(pagato_tot) }}</td></tr>
        <tr><td style="color:var(--text-light);padding:5px 0">Residuo</td><td style="color:{{ '#dc2626' if residuo > 0 else '#16a34a' }};font-weight:700">€ {{ "%.2f"|format(residuo) }}</td></tr>
      </tbody></table>
    </div>
  </div>
</div>

<!-- Rate / Piano pagamenti -->
<div class="card" style="margin-bottom:20px">
  <div class="card-header">
    <h3><i class="fa fa-list-check" style="color:var(--accent2)"></i> Piano di pagamento</h3>
    <a href="/fatturazione/{{ f.id }}/aggiungi-rata" class="btn btn-sm btn-primary"><i class="fa fa-plus"></i> Aggiungi rata</a>
  </div>
  <div class="table-wrap"><table>
    <thead><tr><th>Rata</th><th>Importo</th><th>Scadenza</th><th>Data pagamento</th><th>Stato</th><th>Note</th><th></th></tr></thead>
    <tbody>
    {% for r in rate %}
    <tr style="{{ 'background:#f0fdf4' if r.stato=='pagata' else ('background:#fef2f2' if r.scaduta else '') }}">
      <td><strong>Rata {{ r.numero_rata }}</strong></td>
      <td><strong>€ {{ "%.2f"|format(r.importo) }}</strong></td>
      <td style="font-family:monospace">{{ r.data_scadenza or '–' }}</td>
      <td style="font-family:monospace;color:#16a34a">{{ r.data_pagamento or '–' }}</td>
      <td>
        {% if r.stato == 'pagata' %}<span class="badge badge-green">✅ Pagata</span>
        {% elif r.stato == 'scaduta' or r.scaduta %}<span class="badge badge-red">⚠️ Scaduta</span>
        {% else %}<span class="badge badge-amber">⏳ Da pagare</span>{% endif %}
      </td>
      <td style="font-size:12px;color:var(--text-light)">{{ r.note or '–' }}</td>
      <td style="display:flex;gap:6px">
        {% if r.stato != 'pagata' %}
        <form method="POST" action="/fatturazione/rata/{{ r.id }}/paga">
          <input type="date" name="data_pag" value="{{ today }}" style="padding:4px 8px;font-size:11px;border:1px solid var(--border);border-radius:6px;width:130px">
          <button type="submit" class="btn btn-sm" style="background:#dcfce7;color:#16a34a;border:1px solid #bbf7d0;margin-top:4px"><i class="fa fa-check"></i> Segna pagata</button>
        </form>
        {% else %}
        <form method="POST" action="/fatturazione/rata/{{ r.id }}/annulla-paga">
          <button type="submit" class="btn btn-sm btn-secondary" style="font-size:11px"><i class="fa fa-undo"></i> Annulla</button>
        </form>
        {% endif %}
        <a href="/fatturazione/rata/{{ r.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare rata?')"><i class="fa fa-trash"></i></a>
      </td>
    </tr>
    {% else %}
    <tr><td colspan="7"><div class="empty-state"><i class="fa fa-list"></i><p>Nessuna rata. <a href="/fatturazione/{{ f.id }}/aggiungi-rata">Aggiungi rata</a></p></div></td></tr>
    {% endfor %}
    </tbody>
  </table></div>
</div>

{% if f.note %}
<div class="card"><div class="card-body"><strong>Note:</strong> {{ f.note }}</div></div>
{% endif %}

{% if f.stato == 'da_emettere' %}
<div class="card" style="margin-top:20px;border:2px solid #7c3aed">
  <div class="card-header" style="background:linear-gradient(135deg,#f5f3ff,#ede9fe)">
    <h3 style="color:#7c3aed"><i class="fa fa-file-arrow-up"></i> Allega fattura emessa</h3>
    <div style="font-size:13px;color:#6b7280;margin-top:4px">Carica il PDF della fattura reale: l'AI aggiornerà automaticamente i dati su questa riga</div>
  </div>
  <div class="card-body">
    {% if ai_allegato %}
    <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;padding:12px;margin-bottom:16px;font-size:13px">
      <strong style="color:#15803d"><i class="fa fa-check-circle"></i> AI ha aggiornato la fattura!</strong>
      <div style="margin-top:6px;color:#374151">{{ ai_allegato }}</div>
    </div>
    {% endif %}
    <form method="POST" action="/fatturazione/{{ f.id }}/allega-emessa" enctype="multipart/form-data">
      <div style="display:flex;gap:14px;align-items:flex-end;flex-wrap:wrap">
        <div class="form-group" style="flex:1;margin:0">
          <label>PDF fattura emessa *</label>
          <input type="file" name="file_fattura" accept=".pdf" required style="padding:8px">
        </div>
        <div style="padding-bottom:4px">
          <button type="submit" class="btn btn-primary" style="background:#7c3aed;border-color:#7c3aed">
            <i class="fa fa-wand-magic-sparkles"></i> Allega e aggiorna con AI
          </button>
        </div>
      </div>
      <div style="margin-top:10px;font-size:12px;color:#9ca3af">
        <i class="fa fa-info-circle"></i> L'AI leggerà il PDF e aggiornerà: numero, date, importi, cliente. La riga esistente verrà sovrascritta, non duplicata.
      </div>
    </form>
  </div>
</div>
{% endif %}
"""

# ── Template form fattura ──────────────────────────────────
FATT_FORM_TMPL = """
<div style="margin-bottom:16px"><a href="/fatturazione" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Torna alle fatture</a></div>

{% if ai_result %}
<div class="card" style="margin-bottom:20px;border-left:4px solid #7c3aed">
  <div class="card-header" style="background:linear-gradient(135deg,#f5f3ff,#ede9fe)">
    <h3 style="color:#7c3aed"><i class="fa fa-wand-magic-sparkles"></i> Dati estratti dall'AI ✨</h3>
  </div>
  <div class="card-body">
    <p style="font-size:13px;color:#6b7280;margin-bottom:10px">{{ ai_result.sommario }}</p>
    <div style="display:flex;flex-wrap:wrap;gap:8px">
      {% for k,v in ai_result.dati.items() %}{% if v %}
      <span style="background:#f3f4f6;padding:4px 10px;border-radius:6px;font-size:12px"><strong>{{ k }}:</strong> {{ v }}</span>
      {% endif %}{% endfor %}
    </div>
    <p style="font-size:12px;color:#9ca3af;margin-top:10px">I campi sono stati pre-compilati. Verificali e correggi se necessario.</p>
  </div>
</div>
{% endif %}

<form method="POST" action="{{ form_action }}" enctype="multipart/form-data">

<!-- Upload + AI -->
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-file-pdf" style="color:#dc2626"></i> PDF Fattura</h3></div>
  <div class="card-body">
    <div style="display:flex;gap:14px;align-items:flex-end;flex-wrap:wrap">
      <div class="form-group" style="flex:1;margin:0">
        <label>Carica PDF fattura</label>
        <input type="file" name="file_fattura" accept=".pdf" style="padding:8px">
      </div>
      <div style="display:flex;align-items:center;gap:8px;padding-bottom:4px">
        <input type="checkbox" name="analizza_ai" id="ai-chk" value="1" {{ 'checked' if ai_ok }} style="width:18px;height:18px;accent-color:#7c3aed">
        <label for="ai-chk" style="cursor:pointer;margin:0"><i class="fa fa-wand-magic-sparkles" style="color:#7c3aed"></i> Analizza con AI</label>
      </div>
    </div>
    {% if fattura and fattura.file_nome %}
    <div style="margin-top:8px;font-size:12px;color:var(--text-light)"><i class="fa fa-paperclip"></i> File attuale: {{ fattura.file_nome }}</div>
    {% endif %}
  </div>
</div>

<!-- Dati fattura -->
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-file-invoice-dollar" style="color:var(--accent2)"></i> Dati fattura</h3></div>
  <div class="card-body">
    <div class="form-row">
      <div class="form-group">
        <label>Numero fattura *</label>
        <input name="numero" value="{{ fattura.numero or '' }}" required placeholder="Es. 2024/001">
      </div>
      <div class="form-group">
        <label>Data emissione</label>
        <input type="date" name="data_emissione" value="{{ fattura.data_emissione or today }}">
      </div>
    </div>
    <div class="form-row">
      <div class="form-group">
        <label>Cliente</label>
        <select name="cliente_id" id="cliente-sel" onchange="onClienteChange()">
          <option value="">— Seleziona o scrivi sotto —</option>
          {% for c in clienti %}
          <option value="{{ c.id }}" {{ 'selected' if fattura.cliente_id == c.id }}>{{ c.nome }}</option>
          {% endfor %}
        </select>
      </div>
      <div class="form-group">
        <label>Nome cliente (se nuovo)</label>
        <input name="cliente_nome" id="cliente-nome" value="{{ fattura.cliente_nome or '' }}" placeholder="Inserisci se non in lista">
      </div>
    </div>
    <div class="form-group">
      <label>Descrizione lavori</label>
      <textarea name="descrizione" rows="2" placeholder="Es. Montaggio stand Fiera Milano - pad. 5">{{ fattura.descrizione or '' }}</textarea>
    </div>
  </div>
</div>

<!-- Importi -->
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-euro-sign" style="color:var(--accent2)"></i> Importi</h3></div>
  <div class="card-body">
    <div class="form-row">
      <div class="form-group">
        <label>Imponibile lordo (€) <span style="font-size:11px;color:var(--text-light)">— prima dello sconto</span></label>
        <input type="number" name="imponibile_lordo" id="imponibile-lordo" value="{{ fattura.imponibile_lordo or fattura.imponibile or '' }}" step="0.01" min="0" placeholder="1000.00" oninput="calcolaImporti()">
      </div>
      <div class="form-group">
        <label>Sconto (€) <span style="font-size:11px;color:var(--text-light)">— opzionale</span></label>
        <input type="number" name="sconto_importo" id="sconto-importo" value="{{ fattura.sconto_importo or '' }}" step="0.01" min="0" placeholder="0.00" oninput="calcolaImporti()" style="border-color:#f59e0b">
      </div>
    </div>
    <div class="form-row">
      <div class="form-group">
        <label>Imponibile netto (€) <span style="font-size:11px;color:#16a34a;font-weight:700">— base per IVA</span></label>
        <input type="number" name="imponibile" id="imponibile" value="{{ fattura.imponibile or '' }}" step="0.01" min="0" placeholder="1000.00" oninput="calcolaImporti()" style="font-weight:700;color:#16a34a;border-color:#16a34a">
      </div>
      <div class="form-group">
        <label>IVA (%)</label>
        <select name="iva_perc" id="iva-perc" onchange="calcolaImporti()">
          <option value="0" {{ 'selected' if fattura.iva_perc==0 }}>0% (esente)</option>
          <option value="4" {{ 'selected' if fattura.iva_perc==4 }}>4%</option>
          <option value="10" {{ 'selected' if fattura.iva_perc==10 }}>10%</option>
          <option value="22" {{ 'selected' if (not fattura.iva_perc) or fattura.iva_perc==22 }}>22%</option>
        </select>
      </div>
      <div class="form-group">
        <label>Totale fattura (€)</label>
        <input type="number" name="importo_totale" id="importo-totale" value="{{ fattura.importo_totale or '' }}" step="0.01" min="0" placeholder="1220.00" style="font-weight:700;font-size:16px">
      </div>
    </div>
    <div id="riepilogo-importi" style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;padding:10px 14px;font-size:13px;display:none;margin-top:4px">
      <span id="riepilogo-testo"></span>
    </div>
  </div>
</div>

<!-- Pagamento -->
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-calendar-check" style="color:var(--accent2)"></i> Condizioni di pagamento</h3></div>
  <div class="card-body">
    <div class="form-row">
      <div class="form-group">
        <label>Modalità</label>
        <select name="condizioni_pagamento" id="cond-pag" onchange="onCondizioniChange()">
          <option value="immediato" {{ 'selected' if fattura.condizioni_pagamento=='immediato' }}>Immediato</option>
          <option value="30gg" {{ 'selected' if (not fattura.condizioni_pagamento) or fattura.condizioni_pagamento=='30gg' }}>30 giorni</option>
          <option value="60gg" {{ 'selected' if fattura.condizioni_pagamento=='60gg' }}>60 giorni</option>
          <option value="90gg" {{ 'selected' if fattura.condizioni_pagamento=='90gg' }}>90 giorni</option>
          <option value="50_50" {{ 'selected' if fattura.condizioni_pagamento=='50_50' }}>50% subito + 50% a fine lavori</option>
          <option value="50_30_20" {{ 'selected' if fattura.condizioni_pagamento=='50_30_20' }}>50% a 30gg + 50% a 60gg</option>
          <option value="manuale" {{ 'selected' if fattura.condizioni_pagamento=='manuale' }}>Personalizzato (data manuale)</option>
        </select>
      </div>
      <div class="form-group" id="scad-manual">
        <label>Data scadenza</label>
        <input type="date" name="data_scadenza" value="{{ fattura.data_scadenza or '' }}">
      </div>
    </div>
    <div id="info-cond" style="font-size:12px;color:var(--text-light);margin-top:4px"></div>

    <!-- Genera rate automatiche -->
    <div style="margin-top:12px;padding:12px;background:var(--bg-secondary);border-radius:8px">
      <label style="font-size:13px;font-weight:600">
        <input type="checkbox" name="genera_rate" value="1" id="chk-rate" style="accent-color:var(--accent2)">
        Genera piano di pagamento automaticamente dopo il salvataggio
      </label>
    </div>
  </div>
</div>

<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-sticky-note"></i> Note</h3></div>
  <div class="card-body">
    <textarea name="note" rows="2" placeholder="Note interne sulla fattura...">{{ fattura.note or '' }}</textarea>
  </div>
</div>

<div style="display:flex;gap:12px;justify-content:flex-end">
  <a href="/fatturazione" class="btn btn-secondary">Annulla</a>
  <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva fattura</button>
</div>
</form>

<script>
function calcolaImporti() {
  const lordo  = parseFloat(document.getElementById('imponibile-lordo').value) || 0;
  const sconto = parseFloat(document.getElementById('sconto-importo').value)   || 0;
  const iva    = parseFloat(document.getElementById('iva-perc').value)          || 0;

  // Se cambia il lordo o lo sconto → aggiorna il netto
  const netto = Math.max(0, lordo - sconto);
  if (lordo > 0) {
    document.getElementById('imponibile').value = netto.toFixed(2);
  }

  const nettoUsato = parseFloat(document.getElementById('imponibile').value) || 0;
  const tot = nettoUsato * (1 + iva/100);
  if (nettoUsato > 0) document.getElementById('importo-totale').value = tot.toFixed(2);

  // Riepilogo visivo
  const ris = document.getElementById('riepilogo-importi');
  const txt = document.getElementById('riepilogo-testo');
  if (lordo > 0 && sconto > 0) {
    ris.style.display = 'block';
    txt.innerHTML = `<strong>Lordo:</strong> € ${lordo.toFixed(2)} &nbsp;−&nbsp; <strong>Sconto:</strong> € ${sconto.toFixed(2)} &nbsp;=&nbsp; <strong style="color:#16a34a">Imponibile netto: € ${netto.toFixed(2)}</strong> &nbsp;+&nbsp; IVA ${iva}% &nbsp;=&nbsp; <strong>Totale: € ${tot.toFixed(2)}</strong>`;
  } else {
    ris.style.display = 'none';
  }
}

function onCondizioniChange() {
  const v = document.getElementById('cond-pag').value;
  const info = document.getElementById('info-cond');
  const scadManual = document.getElementById('scad-manual');
  const msgs = {
    '30gg': 'La scadenza sarà calcolata automaticamente a 30 giorni dalla data di emissione.',
    '60gg': 'La scadenza sarà calcolata automaticamente a 60 giorni.',
    '90gg': 'La scadenza sarà calcolata automaticamente a 90 giorni.',
    'immediato': 'Pagamento immediato alla data di emissione.',
    '50_50': 'Rata 1: 50% subito. Rata 2: 50% a fine lavori (da definire).',
    '50_30_20': 'Rata 1: 50% a 30 giorni. Rata 2: 50% a 60 giorni.',
    'manuale': 'Inserisci manualmente la data di scadenza.'
  };
  info.textContent = msgs[v] || '';
  scadManual.style.display = (v === 'manuale') ? 'block' : 'none';
}

function onClienteChange() {
  const sel = document.getElementById('cliente-sel');
  const opt = sel.options[sel.selectedIndex];
  if (sel.value) document.getElementById('cliente-nome').value = opt.text;
}

onCondizioniChange();
</script>
"""

# ── Template aggiungi rata ─────────────────────────────────
FATT_RATA_TMPL = """
<div style="margin-bottom:16px"><a href="/fatturazione/{{ fid }}" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Torna alla fattura</a></div>
<div class="card" style="max-width:600px">
  <div class="card-header"><h3><i class="fa fa-plus-circle" style="color:var(--accent2)"></i> Aggiungi rata — Fattura {{ f_numero }}</h3></div>
  <div class="card-body">
    <form method="POST" action="/fatturazione/{{ fid }}/aggiungi-rata">
      <div class="form-row">
        <div class="form-group">
          <label>Numero rata</label>
          <input type="number" name="numero_rata" value="{{ next_num }}" min="1" required>
        </div>
        <div class="form-group">
          <label>Importo (€) *</label>
          <input type="number" name="importo" step="0.01" min="0" required placeholder="Es. 610.00">
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>Data scadenza</label>
          <input type="date" name="data_scadenza" value="{{ today }}">
        </div>
        <div class="form-group">
          <label>Stato</label>
          <select name="stato">
            <option value="da_pagare">Da pagare</option>
            <option value="pagata">Già pagata</option>
          </select>
        </div>
      </div>
      <div class="form-group">
        <label>Data pagamento (se già pagata)</label>
        <input type="date" name="data_pagamento" value="{{ today }}">
      </div>
      <div class="form-group"><label>Note</label><input name="note" placeholder="Es. acconto iniziale"></div>
      <div style="display:flex;gap:10px;justify-content:flex-end">
        <a href="/fatturazione/{{ fid }}" class="btn btn-secondary">Annulla</a>
        <button type="submit" class="btn btn-primary"><i class="fa fa-plus"></i> Aggiungi rata</button>
      </div>
    </form>
  </div>
</div>
"""

# ── Template clienti ───────────────────────────────────────
CLIENTI_TMPL = """
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px">
  <a href="/fatturazione" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Fatture</a>
  <button onclick="document.getElementById('modal-cliente').style.display='flex'" class="btn btn-primary"><i class="fa fa-plus"></i> Nuovo cliente</button>
</div>
<div class="card">
  <div class="table-wrap"><table>
    <thead><tr><th>Nome</th><th>P.IVA</th><th>Email</th><th>Telefono</th><th>Fatture</th><th>Totale</th><th></th></tr></thead>
    <tbody>{% for c in clienti %}
    <tr>
      <td><strong>{{ c.nome }}</strong></td>
      <td style="font-family:monospace;font-size:12px">{{ c.partita_iva or '–' }}</td>
      <td style="font-size:13px">{{ c.email or '–' }}</td>
      <td style="font-size:13px">{{ c.telefono or '–' }}</td>
      <td><span class="badge badge-blue">{{ c.n_fatture }}</span></td>
      <td><strong>€ {{ "%.0f"|format(c.tot_fatturato) }}</strong></td>
      <td style="display:flex;gap:6px">
        <a href="/fatturazione/clienti/{{ c.id }}/modifica" class="btn btn-sm btn-secondary"><i class="fa fa-pen"></i></a>
        <a href="/fatturazione/clienti/{{ c.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare cliente?')"><i class="fa fa-trash"></i></a>
      </td>
    </tr>{% else %}
    <tr><td colspan="7"><div class="empty-state"><i class="fa fa-users"></i><p>Nessun cliente.</p></div></td></tr>
    {% endfor %}</tbody>
  </table></div>
</div>

<!-- Modal nuovo cliente -->
<div id="modal-cliente" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9999;align-items:center;justify-content:center">
  <div class="card" style="width:480px;max-width:95vw">
    <div class="card-header">
      <h3><i class="fa fa-user-plus" style="color:var(--accent2)"></i> Nuovo cliente</h3>
      <button onclick="document.getElementById('modal-cliente').style.display='none'" class="btn btn-secondary btn-sm"><i class="fa fa-times"></i></button>
    </div>
    <div class="card-body">
      <form method="POST" action="/fatturazione/clienti/nuovo">
        <div class="form-row">
          <div class="form-group" style="grid-column:span 2"><label>Nome / Ragione sociale *</label><input name="nome" required placeholder="Es. StandHouse srl"></div>
        </div>
        <div class="form-row">
          <div class="form-group"><label>P.IVA / C.F.</label><input name="partita_iva" placeholder="Es. 01234567890"></div>
          <div class="form-group"><label>Email</label><input type="email" name="email" placeholder="info@cliente.it"></div>
        </div>
        <div class="form-row">
          <div class="form-group"><label>Telefono</label><input name="telefono" placeholder="+39 333..."></div>
          <div class="form-group"><label>Indirizzo</label><input name="indirizzo" placeholder="Via..."></div>
        </div>
        <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:8px">
          <button type="button" onclick="document.getElementById('modal-cliente').style.display='none'" class="btn btn-secondary">Annulla</button>
          <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva cliente</button>
        </div>
      </form>
    </div>
  </div>
</div>
"""


# ── Helper: calcola scadenza automatica ───────────────────
def _calc_scadenza(data_emissione, condizioni):
    from datetime import datetime as _dt, timedelta as _td
    try:
        d = _dt.strptime(data_emissione, '%Y-%m-%d').date()
    except Exception:
        d = date.today()
    mapping = {'immediato': 0, '30gg': 30, '60gg': 60, '90gg': 90}
    giorni = mapping.get(condizioni, 30)
    return (d + _td(days=giorni)).isoformat()


def _genera_rate(fattura_id, importo_totale, condizioni, data_emissione, db):
    """Genera le rate automatiche in base alle condizioni."""
    from datetime import datetime as _dt, timedelta as _td
    try:
        d = _dt.strptime(data_emissione, '%Y-%m-%d').date()
    except Exception:
        d = date.today()

    # Cancella rate esistenti
    db.execute("DELETE FROM rate_fattura WHERE fattura_id=?", (fattura_id,))

    if condizioni == '50_50':
        # 50% subito + 50% a fine lavori (60gg come default)
        db.execute("INSERT INTO rate_fattura (fattura_id,numero_rata,importo,data_scadenza,stato) VALUES (?,?,?,?,?)",
                   (fattura_id, 1, round(importo_totale * 0.5, 2), d.isoformat(), 'da_pagare'))
        db.execute("INSERT INTO rate_fattura (fattura_id,numero_rata,importo,data_scadenza,stato,note) VALUES (?,?,?,?,?,?)",
                   (fattura_id, 2, round(importo_totale * 0.5, 2), (d + _td(days=60)).isoformat(), 'da_pagare', 'Saldo a fine lavori'))

    elif condizioni == '50_30_20':
        # 50% a 30gg + 50% a 60gg
        db.execute("INSERT INTO rate_fattura (fattura_id,numero_rata,importo,data_scadenza,stato,note) VALUES (?,?,?,?,?,?)",
                   (fattura_id, 1, round(importo_totale * 0.5, 2), (d + _td(days=30)).isoformat(), 'da_pagare', '50% a 30 giorni'))
        db.execute("INSERT INTO rate_fattura (fattura_id,numero_rata,importo,data_scadenza,stato,note) VALUES (?,?,?,?,?,?)",
                   (fattura_id, 2, round(importo_totale * 0.5, 2), (d + _td(days=60)).isoformat(), 'da_pagare', '50% a 60 giorni'))

    elif condizioni in ('30gg', '60gg', '90gg', 'immediato'):
        # Unica rata
        giorni = {'immediato': 0, '30gg': 30, '60gg': 60, '90gg': 90}[condizioni]
        db.execute("INSERT INTO rate_fattura (fattura_id,numero_rata,importo,data_scadenza,stato) VALUES (?,?,?,?,?)",
                   (fattura_id, 1, importo_totale, (d + _td(days=giorni)).isoformat(), 'da_pagare'))


def _aggiorna_stato_fattura(fattura_id, db):
    """Ricalcola lo stato della fattura in base alle rate."""
    rate = db.execute("SELECT * FROM rate_fattura WHERE fattura_id=?", (fattura_id,)).fetchall()
    if not rate:
        return
    tot = sum(r['importo'] or 0 for r in rate)
    pagate = sum(r['importo'] or 0 for r in rate if r['stato'] == 'pagata')
    if pagate >= tot:
        nuovo_stato = 'pagata'
    elif pagate > 0:
        nuovo_stato = 'parziale'
    else:
        # Scaduta solo se almeno UNA rata non pagata ha scadenza passata
        oggi_s = date.today().isoformat()
        scadute = [r for r in rate if r['stato'] != 'pagata' and r['data_scadenza'] and r['data_scadenza'] < oggi_s]
        nuovo_stato = 'scaduta' if scadute else 'da_pagare'
    db.execute("UPDATE fatture SET stato=? WHERE id=?", (nuovo_stato, fattura_id))


def _analizza_fattura_ai(file_path):
    """Estrae dati da PDF fattura con AI."""
    api_key = get_setting('anthropic_api_key', '')
    if not api_key:
        return None
    try:
        import anthropic, base64
        client = anthropic.Anthropic(api_key=api_key)
        with open(file_path, 'rb') as f:
            pdf_data = base64.standard_b64encode(f.read()).decode()
        cb = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_data}}
        prompt = """Analizza questa fattura e rispondi SOLO con JSON valido.
Estrai: numero_fattura, data_emissione (YYYY-MM-DD), data_scadenza (YYYY-MM-DD), 
cliente_nome, cliente_piva, descrizione, imponibile (solo numero), iva_perc (solo numero), importo_totale (solo numero), condizioni_pagamento.
Per condizioni_pagamento usa: immediato, 30gg, 60gg, 90gg, 50_50, 50_30_20, o manuale.
Formato: {"numero_fattura":"","data_emissione":"","data_scadenza":"","cliente_nome":"","cliente_piva":"","descrizione":"","imponibile":"","iva_perc":"22","importo_totale":"","condizioni_pagamento":"30gg","sommario":"breve descrizione"}
Se un campo non esiste metti "". Date in formato YYYY-MM-DD."""
        resp = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=1024,
                                      messages=[{"role": "user", "content": [cb, {"type": "text", "text": prompt}]}])
        raw = resp.content[0].text.strip()
        if '```' in raw:
            raw = raw.split('```')[1].lstrip('json').strip()
        data = json.loads(raw)
        return data
    except Exception as e:
        return {'_error': str(e)}


# ══════════════════════════════════════════════════════════
#  ROUTES FATTURAZIONE
# ══════════════════════════════════════════════════════════

@app.route('/fatturazione')
@admin_required
def fatturazione():
    db = get_db()
    filtro_stato = request.args.get('stato', '')
    q = request.args.get('q', '').strip()

    sql = """SELECT f.*, 
             COALESCE((SELECT SUM(r.importo) FROM rate_fattura r WHERE r.fattura_id=f.id AND r.stato='pagata'),0) as pagato
             FROM fatture f WHERE 1=1"""
    params = []
    if filtro_stato:
        sql += " AND f.stato=?"
        params.append(filtro_stato)
    if q:
        sql += " AND (f.numero LIKE ? OR f.cliente_nome LIKE ?)"
        params += [f'%{q}%', f'%{q}%']
    sql += " ORDER BY f.data_emissione DESC"

    fatture_raw = db.execute(sql, params).fetchall()
    oggi_s = date.today().isoformat()

    fatture = []
    tot_da_pagare = tot_parziale = tot_pagato = 0
    n_da_pagare = n_parziale = n_pagate = n_scadute_c = 0

    for f in fatture_raw:
        fd = dict(f)
        fd['residuo'] = (fd['importo_totale'] or 0) - (fd['pagato'] or 0)

        # Scaduta = solo se ci sono RATE non pagate con scadenza passata
        # (non usare data_scadenza della fattura che può essere la data di emissione)
        rate_scadute = db.execute("""
            SELECT COUNT(*) as n FROM rate_fattura
            WHERE fattura_id=? AND stato != 'pagata'
              AND data_scadenza IS NOT NULL AND data_scadenza < ?
        """, (fd['id'], oggi_s)).fetchone()
        ha_rate = db.execute("SELECT COUNT(*) as n FROM rate_fattura WHERE fattura_id=?", (fd['id'],)).fetchone()

        if ha_rate and ha_rate['n'] > 0:
            # Se ha rate: scaduta solo se almeno una rata non pagata è scaduta
            fd['scaduta'] = bool(rate_scadute and rate_scadute['n'] > 0 and fd['stato'] != 'pagata')
        else:
            # Senza rate: usa data_scadenza della fattura
            fd['scaduta'] = bool(fd.get('data_scadenza') and fd['data_scadenza'] < oggi_s and fd['stato'] != 'pagata')

        if fd['stato'] == 'da_emettere':
            fd['stato_display'] = 'da_emettere'
        else:
            fd['stato_display'] = 'scaduta' if fd['scaduta'] and fd['stato'] != 'pagata' else fd['stato']

        # Carica rate per mostrare sotto-righe nella lista
        rate_rows = db.execute(
            "SELECT * FROM rate_fattura WHERE fattura_id=? ORDER BY numero_rata", (fd['id'],)
        ).fetchall()
        fd['rate'] = []
        for r in rate_rows:
            rd = dict(r)
            rd['scaduta'] = bool(rd.get('data_scadenza') and rd['data_scadenza'] < oggi_s and rd['stato'] != 'pagata')
            fd['rate'].append(rd)

        # Scadenza display = prima rata non pagata (o data fattura se nessuna rata)
        prima_da_pagare = next((r for r in fd['rate'] if r['stato'] != 'pagata'), None)
        fd['scadenza_display'] = prima_da_pagare['data_scadenza'] if prima_da_pagare else fd.get('data_scadenza')
        fd['scadenza_scaduta'] = prima_da_pagare['scaduta'] if prima_da_pagare else fd['scaduta']

        # Rate espanse: solo se almeno una rata non è pagata
        fd['rate_espanse'] = any(r['stato'] != 'pagata' for r in fd['rate'])

        fatture.append(fd)
        s = fd['stato']
        if s == 'pagata':
            tot_pagato += fd['importo_totale'] or 0
            n_pagate += 1
        elif s == 'parziale':
            tot_parziale += fd['residuo']
            tot_pagato += fd['pagato']
            n_parziale += 1
        else:
            tot_da_pagare += fd['importo_totale'] or 0
            n_da_pagare += 1
        if fd['scaduta'] and s != 'pagata':
            n_scadute_c += 1

    # Statistiche clienti
    top_raw = db.execute("""
        SELECT cliente_nome as nome,
               COUNT(*) as n_totale,
               SUM(CASE WHEN stato='pagata' THEN 1 ELSE 0 END) as n_pagate
        FROM fatture WHERE cliente_nome IS NOT NULL AND cliente_nome != ''
        GROUP BY cliente_nome HAVING n_totale >= 1
        ORDER BY (n_pagate*1.0/n_totale) DESC LIMIT 5""").fetchall()
    top_clienti = [dict(r, perc=round(r['n_pagate']/r['n_totale']*100, 0)) for r in top_raw]

    # Dati grafico mensile ultimi 6 mesi
    from datetime import datetime as _dt
    mesi_labels, mesi_fatturato, mesi_incassato = [], [], []
    mesi_nomi_short = ['Gen','Feb','Mar','Apr','Mag','Giu','Lug','Ago','Set','Ott','Nov','Dic']
    for i in range(5, -1, -1):
        from datetime import timedelta as _td
        d = date.today().replace(day=1)
        for _ in range(i):
            d = (d - _td(days=1)).replace(day=1)
        m_str = d.strftime('%Y-%m')
        row_f = db.execute("SELECT SUM(importo_totale) FROM fatture WHERE substr(data_emissione,1,7)=?", (m_str,)).fetchone()
        row_p = db.execute("""SELECT SUM(r.importo) FROM rate_fattura r 
                              JOIN fatture f ON f.id=r.fattura_id 
                              WHERE r.stato='pagata' AND substr(r.data_pagamento,1,7)=?""", (m_str,)).fetchone()
        mesi_labels.append(mesi_nomi_short[d.month-1])
        mesi_fatturato.append(round(row_f[0] or 0, 2))
        mesi_incassato.append(round(row_p[0] or 0, 2))

    db.close()
    return render_page(FATT_LIST_TMPL, page_title='Fatturazione', active='fatturazione',
        fatture=fatture, filtro_stato=filtro_stato, q=q,
        tot_da_pagare=tot_da_pagare, tot_parziale=tot_parziale, tot_pagato=tot_pagato,
        n_da_pagare=n_da_pagare, n_parziale=n_parziale, n_pagate=n_pagate, n_scadute=n_scadute_c, n_scadute_c=n_scadute_c,
        top_clienti=top_clienti, today=date.today().isoformat(),
        mesi_labels=mesi_labels, mesi_fatturato=mesi_fatturato, mesi_incassato=mesi_incassato)


@app.route('/fatturazione/nuova', methods=['GET', 'POST'])
@admin_required
def fatturazione_nuova():
    if request.method == 'POST':
        return _salva_fattura(None)
    db = get_db()
    clienti = db.execute("SELECT id, nome FROM clienti ORDER BY nome").fetchall()
    db.close()
    ai_result = session.pop('fatt_ai_result', None)
    pre = session.pop('fatt_pre', {})
    return render_page(FATT_FORM_TMPL, page_title='Nuova fattura', active='fatturazione',
        form_action='/fatturazione/nuova', fattura=pre, clienti=clienti,
        ai_result=ai_result, ai_ok=bool(get_setting('anthropic_api_key','')),
        today=date.today().isoformat())


@app.route('/fatturazione/<int:fid>/modifica', methods=['GET', 'POST'])
@admin_required
def fatturazione_modifica(fid):
    if request.method == 'POST':
        return _salva_fattura(fid)
    db = get_db()
    f = db.execute("SELECT * FROM fatture WHERE id=?", (fid,)).fetchone()
    clienti = db.execute("SELECT id, nome FROM clienti ORDER BY nome").fetchall()
    db.close()
    if not f:
        return redirect(url_for('fatturazione'))
    return render_page(FATT_FORM_TMPL, page_title='Modifica fattura', active='fatturazione',
        form_action=f'/fatturazione/{fid}/modifica', fattura=dict(f), clienti=clienti,
        ai_result=None, ai_ok=bool(get_setting('anthropic_api_key','')),
        today=date.today().isoformat())


def _salva_fattura(fid):
    numero = request.form.get('numero', '').strip()
    cliente_id = request.form.get('cliente_id') or None
    cliente_nome = request.form.get('cliente_nome', '').strip()
    data_em = request.form.get('data_emissione') or date.today().isoformat()
    data_scad = request.form.get('data_scadenza') or None
    cond = request.form.get('condizioni_pagamento', '30gg')
    imponibile = float(request.form.get('imponibile') or 0)
    imponibile_lordo = float(request.form.get('imponibile_lordo') or imponibile)
    sconto_importo   = float(request.form.get('sconto_importo') or 0)
    iva_perc = float(request.form.get('iva_perc') or 22)
    importo_tot = float(request.form.get('importo_totale') or 0)
    iva_imp = round(importo_tot - imponibile, 2)
    descrizione = request.form.get('descrizione', '').strip()
    note = request.form.get('note', '').strip()
    genera_rate = request.form.get('genera_rate') == '1'

    # Se condizioni non è manuale, calcola scadenza
    if cond != 'manuale' and cond not in ('50_50', '50_30_20'):
        data_scad = _calc_scadenza(data_em, cond)

    # Se cliente_id → recupera nome
    db = get_db()
    if cliente_id:
        c = db.execute("SELECT nome FROM clienti WHERE id=?", (cliente_id,)).fetchone()
        if c:
            cliente_nome = c['nome']

    # Gestisci upload file
    file_nome = None
    file_path_db = None
    f_file = request.files.get('file_fattura')
    if f_file and f_file.filename:
        ext = f_file.filename.rsplit('.', 1)[-1].lower()
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_nome = f"{ts}_{secure_filename(f_file.filename)}"
        fp = os.path.join(get_fatt_upload_path(), file_nome)
        f_file.save(fp)
        file_path_db = fp

        # AI analysis
        if request.form.get('analizza_ai') == '1':
            ai_data = _analizza_fattura_ai(fp)
            if ai_data and '_error' not in ai_data:
                session['fatt_ai_result'] = {
                    'sommario': ai_data.get('sommario', 'Analisi completata.'),
                    'dati': {k: v for k, v in ai_data.items() if k != 'sommario' and v}
                }
                # Pre-fill form with AI data
                session['fatt_pre'] = {
                    'numero': ai_data.get('numero_fattura', numero) or numero,
                    'data_emissione': ai_data.get('data_emissione', data_em) or data_em,
                    'data_scadenza': ai_data.get('data_scadenza', ''),
                    'cliente_nome': ai_data.get('cliente_nome', cliente_nome) or cliente_nome,
                    'descrizione': ai_data.get('descrizione', ''),
                    'imponibile': ai_data.get('imponibile', ''),
                    'iva_perc': ai_data.get('iva_perc', 22),
                    'importo_totale': ai_data.get('importo_totale', ''),
                    'condizioni_pagamento': ai_data.get('condizioni_pagamento', '30gg'),
                }
                db.close()
                flash('✨ AI ha analizzato la fattura. Verifica i dati pre-compilati.', 'success')
                return redirect(url_for('fatturazione_nuova'))

    if fid:
        # Mantieni file esistente se non caricato nuovo
        if not file_nome:
            old = db.execute("SELECT file_nome, file_path FROM fatture WHERE id=?", (fid,)).fetchone()
            if old:
                file_nome = old['file_nome']
                file_path_db = old['file_path']
        db.execute("""UPDATE fatture SET numero=?,cliente_id=?,cliente_nome=?,data_emissione=?,
            data_scadenza=?,condizioni_pagamento=?,imponibile=?,imponibile_lordo=?,sconto_importo=?,
            iva_perc=?,iva_importo=?,
            importo_totale=?,descrizione=?,note=?,file_nome=?,file_path=? WHERE id=?""",
            (numero, cliente_id, cliente_nome, data_em, data_scad, cond,
             imponibile, imponibile_lordo, sconto_importo,
             iva_perc, iva_imp, importo_tot, descrizione, note, file_nome, file_path_db, fid))
        saved_id = fid
    else:
        cur = db.execute("""INSERT INTO fatture (numero,cliente_id,cliente_nome,data_emissione,
            data_scadenza,condizioni_pagamento,imponibile,imponibile_lordo,sconto_importo,
            iva_perc,iva_importo,importo_totale,
            descrizione,note,file_nome,file_path,stato) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'da_pagare')""",
            (numero, cliente_id, cliente_nome, data_em, data_scad, cond,
             imponibile, imponibile_lordo, sconto_importo,
             iva_perc, iva_imp, importo_tot, descrizione, note, file_nome, file_path_db))
        saved_id = cur.lastrowid

    if genera_rate and importo_tot > 0:
        _genera_rate(saved_id, importo_tot, cond, data_em, db)
        _aggiorna_stato_fattura(saved_id, db)

    safe_commit(db)
    db.close()
    flash(f'Fattura {numero} salvata!', 'success')
    return redirect(url_for('fatturazione_dettaglio', fid=saved_id))


@app.route('/fatturazione/<int:fid>')
@admin_required
def fatturazione_dettaglio(fid):
    db = get_db()
    f = db.execute("SELECT * FROM fatture WHERE id=?", (fid,)).fetchone()
    if not f:
        return redirect(url_for('fatturazione'))
    rate_raw = db.execute("SELECT * FROM rate_fattura WHERE fattura_id=? ORDER BY numero_rata", (fid,)).fetchall()
    db.close()

    oggi_s = date.today().isoformat()
    rate = []
    for r in rate_raw:
        rd = dict(r)
        rd['scaduta'] = bool(rd.get('data_scadenza') and rd['data_scadenza'] < oggi_s and rd['stato'] != 'pagata')
        rate.append(rd)

    pagato_tot = sum(r['importo'] for r in rate if r['stato'] == 'pagata')
    residuo = (f['importo_totale'] or 0) - pagato_tot

    return render_page(FATT_DETAIL_TMPL, page_title=f'Fattura {f["numero"]}', active='fatturazione',
        f=dict(f), rate=rate, pagato_tot=pagato_tot, residuo=residuo, today=date.today().isoformat(),
        ai_allegato=session.pop('fatt_allegato_msg', None))


@app.route('/fatturazione/<int:fid>/allega-emessa', methods=['POST'])
@admin_required
def fatturazione_allega_emessa(fid):
    """Allega PDF fattura reale a una riga offline: AI legge e aggiorna la stessa riga."""
    db = get_db()
    f = db.execute("SELECT * FROM fatture WHERE id=?", (fid,)).fetchone()
    if not f:
        db.close()
        flash('Fattura non trovata.', 'error')
        return redirect(url_for('fatturazione'))

    f_file = request.files.get('file_fattura')
    if not f_file or not f_file.filename:
        db.close()
        flash('Nessun file caricato.', 'error')
        return redirect(url_for('fatturazione_dettaglio', fid=fid))

    # Salva file
    ext = f_file.filename.rsplit('.', 1)[-1].lower()
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_nome = f"{ts}_{secure_filename(f_file.filename)}"
    fp = os.path.join(get_fatt_upload_path(), file_nome)
    f_file.save(fp)

    # Analisi AI
    ai_data = _analizza_fattura_ai(fp)
    msg_parts = []

    if ai_data and '_error' not in ai_data:
        # Aggiorna campi dalla AI, mantieni quelli già validi se l'AI non li trova
        numero       = ai_data.get('numero_fattura') or f['numero']
        data_em      = ai_data.get('data_emissione') or f['data_emissione']
        data_scad    = ai_data.get('data_scadenza')  or f['data_scadenza']
        cliente_nome = ai_data.get('cliente_nome')   or f['cliente_nome']
        descrizione  = ai_data.get('descrizione')    or f['descrizione']
        cond         = ai_data.get('condizioni_pagamento') or f['condizioni_pagamento'] or '30gg'

        try: imponibile  = float(ai_data.get('imponibile')      or f['imponibile']      or 0)
        except: imponibile = float(f['imponibile'] or 0)
        try: iva_perc    = float(ai_data.get('iva_perc')         or f['iva_perc']        or 22)
        except: iva_perc = float(f['iva_perc'] or 22)
        try: importo_tot = float(ai_data.get('importo_totale')   or f['importo_totale']  or 0)
        except: importo_tot = float(f['importo_totale'] or 0)

        if importo_tot == 0 and imponibile > 0:
            importo_tot = round(imponibile * (1 + iva_perc/100), 2)
        iva_imp = round(importo_tot - imponibile, 2)

        db.execute("""UPDATE fatture SET
            numero=?, data_emissione=?, data_scadenza=?, cliente_nome=?,
            descrizione=?, condizioni_pagamento=?,
            imponibile=?, iva_perc=?, iva_importo=?, importo_totale=?,
            file_nome=?, file_path=?, stato='da_pagare'
            WHERE id=?""",
            (numero, data_em, data_scad, cliente_nome,
             descrizione, cond,
             imponibile, iva_perc, iva_imp, importo_tot,
             file_nome, fp, fid))
        safe_commit(db)

        msg_parts.append(f"Numero: {numero}")
        if data_em: msg_parts.append(f"Data: {data_em}")
        if cliente_nome: msg_parts.append(f"Cliente: {cliente_nome}")
        msg_parts.append(f"Totale: €{importo_tot:.2f}")
        msg_parts.append("Stato → Da pagare")
        session['fatt_allegato_msg'] = " · ".join(msg_parts)
    else:
        # AI non disponibile: salva solo il file e cambia stato
        db.execute("UPDATE fatture SET file_nome=?, file_path=?, stato='da_pagare' WHERE id=?",
                   (file_nome, fp, fid))
        safe_commit(db)
        session['fatt_allegato_msg'] = "File allegato. Stato aggiornato a 'Da pagare'. (AI non disponibile: verifica i dati manualmente)"

    db.close()
    return redirect(url_for('fatturazione_dettaglio', fid=fid))


@app.route('/fatturazione/<int:fid>/elimina')
@admin_required
def fatturazione_elimina(fid):
    db = get_db()
    db.execute("DELETE FROM rate_fattura WHERE fattura_id=?", (fid,))
    db.execute("DELETE FROM fatture WHERE id=?", (fid,))
    safe_commit(db)
    db.close()
    flash('Fattura eliminata.', 'success')
    return redirect(url_for('fatturazione'))


@app.route('/fatturazione/<int:fid>/aggiungi-rata', methods=['GET', 'POST'])
@admin_required
def fatturazione_aggiungi_rata(fid):
    db = get_db()
    f = db.execute("SELECT * FROM fatture WHERE id=?", (fid,)).fetchone()
    if not f:
        return redirect(url_for('fatturazione'))
    if request.method == 'POST':
        n_rata = int(request.form.get('numero_rata', 1))
        importo = float(request.form.get('importo') or 0)
        data_scad = request.form.get('data_scadenza') or None
        stato = request.form.get('stato', 'da_pagare')
        data_pag = request.form.get('data_pagamento') or None
        note = request.form.get('note', '')
        if stato != 'pagata':
            data_pag = None
        db.execute("INSERT INTO rate_fattura (fattura_id,numero_rata,importo,data_scadenza,stato,data_pagamento,note) VALUES (?,?,?,?,?,?,?)",
                   (fid, n_rata, importo, data_scad, stato, data_pag, note))
        _aggiorna_stato_fattura(fid, db)
        safe_commit(db)
        db.close()
        flash('Rata aggiunta!', 'success')
        return redirect(url_for('fatturazione_dettaglio', fid=fid))
    next_num = db.execute("SELECT COALESCE(MAX(numero_rata),0)+1 FROM rate_fattura WHERE fattura_id=?", (fid,)).fetchone()[0]
    db.close()
    return render_page(FATT_RATA_TMPL, page_title='Aggiungi rata', active='fatturazione',
        fid=fid, f_numero=f['numero'], next_num=next_num, today=date.today().isoformat())


@app.route('/fatturazione/rata/<int:rid>/paga', methods=['POST'])
@admin_required
def fatturazione_paga_rata(rid):
    data_pag = request.form.get('data_pag') or date.today().isoformat()
    db = get_db()
    r = db.execute("SELECT * FROM rate_fattura WHERE id=?", (rid,)).fetchone()
    if r:
        db.execute("UPDATE rate_fattura SET stato='pagata', data_pagamento=? WHERE id=?", (data_pag, rid))
        _aggiorna_stato_fattura(r['fattura_id'], db)
        safe_commit(db)
        flash('✅ Rata segnata come pagata!', 'success')
        fid = r['fattura_id']
    db.close()
    return redirect(url_for('fatturazione_dettaglio', fid=fid))


@app.route('/fatturazione/rata/<int:rid>/annulla-paga', methods=['POST'])
@admin_required
def fatturazione_annulla_paga_rata(rid):
    db = get_db()
    r = db.execute("SELECT * FROM rate_fattura WHERE id=?", (rid,)).fetchone()
    if r:
        db.execute("UPDATE rate_fattura SET stato='da_pagare', data_pagamento=NULL WHERE id=?", (rid,))
        _aggiorna_stato_fattura(r['fattura_id'], db)
        safe_commit(db)
        fid = r['fattura_id']
    db.close()
    return redirect(url_for('fatturazione_dettaglio', fid=fid))


@app.route('/fatturazione/rata/<int:rid>/elimina')
@admin_required
def fatturazione_elimina_rata(rid):
    db = get_db()
    r = db.execute("SELECT * FROM rate_fattura WHERE id=?", (rid,)).fetchone()
    if r:
        fid = r['fattura_id']
        db.execute("DELETE FROM rate_fattura WHERE id=?", (rid,))
        _aggiorna_stato_fattura(fid, db)
        safe_commit(db)
    db.close()
    return redirect(url_for('fatturazione_dettaglio', fid=fid if r else 0))


@app.route('/fatturazione/<int:fid>/file')
@admin_required
def fatturazione_file(fid):
    db = get_db()
    f = db.execute("SELECT * FROM fatture WHERE id=?", (fid,)).fetchone()
    db.close()
    if not f or not f['file_nome']:
        flash('Nessun PDF allegato a questa fattura.', 'error')
        return redirect(url_for('fatturazione'))
    # Controlla se il file esiste fisicamente
    fp = f['file_path']
    if not fp or not os.path.exists(fp):
        # File perso (es. dopo deploy Railway) — mostra messaggio con link per riallegare
        flash(f'⚠️ Il file "{f["file_nome"]}" non è più disponibile sul server (potrebbe essere andato perso dopo un aggiornamento). Riallega il PDF dalla pagina di modifica.', 'error')
        return redirect(url_for('fatturazione_dettaglio', fid=fid))
    from flask import send_file
    nome = f['file_nome'] if f['file_nome'] else f"Fattura_{f['numero'] or fid}.pdf"
    return send_file(fp, as_attachment=True, download_name=nome)


# ── CLIENTI ───────────────────────────────────────────────

@app.route('/fatturazione/clienti')
@admin_required
def fatturazione_clienti():
    db = get_db()
    clienti_raw = db.execute("""
        SELECT c.*,
               COUNT(f.id) as n_fatture,
               COALESCE(SUM(f.importo_totale),0) as tot_fatturato
        FROM clienti c LEFT JOIN fatture f ON f.cliente_id=c.id
        GROUP BY c.id ORDER BY c.nome""").fetchall()
    db.close()
    return render_page(CLIENTI_TMPL, page_title='Clienti', active='fatturazione',
        clienti=[dict(c) for c in clienti_raw])


@app.route('/fatturazione/clienti/nuovo', methods=['POST'])
@admin_required
def fatturazione_nuovo_cliente():
    db = get_db()
    db.execute("INSERT INTO clienti (nome,partita_iva,email,telefono,indirizzo) VALUES (?,?,?,?,?)",
               (request.form.get('nome',''), request.form.get('partita_iva',''),
                request.form.get('email',''), request.form.get('telefono',''),
                request.form.get('indirizzo','')))
    safe_commit(db)
    db.close()
    flash('Cliente aggiunto!', 'success')
    return redirect(url_for('fatturazione_clienti'))


@app.route('/fatturazione/clienti/<int:cid>/elimina')
@admin_required
def fatturazione_elimina_cliente(cid):
    db = get_db()
    db.execute("DELETE FROM clienti WHERE id=?", (cid,))
    safe_commit(db)
    db.close()
    flash('Cliente eliminato.', 'success')
    return redirect(url_for('fatturazione_clienti'))


# ══════════════════════════════════════════════════════════
#  MOBILE — TIMBRATURA DIPENDENTI
# ══════════════════════════════════════════════════════════

MOBILE_TMPL = """<!DOCTYPE html>
<html lang="{{ lang }}" dir="{{ t.dir }}">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<title>{{ t.day_label }}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{background:#0f172a;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;min-height:100vh;display:flex;flex-direction:column}
.header{background:linear-gradient(135deg,#1e3a5f,#0f172a);padding:20px 20px 16px;border-bottom:1px solid rgba(255,255,255,.08)}
.header-top{display:flex;justify-content:space-between;align-items:center;margin-bottom:4px}
.logo{font-size:13px;font-weight:700;color:rgba(255,255,255,.5);letter-spacing:1px;text-transform:uppercase}
.logout-btn{font-size:12px;color:rgba(255,255,255,.4);text-decoration:none;padding:6px 10px;border:1px solid rgba(255,255,255,.15);border-radius:8px}
.user-name{font-size:22px;font-weight:800;color:#fff}
.user-sub{font-size:13px;color:rgba(255,255,255,.5);margin-top:2px}
.content{flex:1;padding:20px;display:flex;flex-direction:column;gap:16px;max-width:480px;margin:0 auto;width:100%}
.card{background:#1e293b;border-radius:16px;padding:20px;border:1px solid rgba(255,255,255,.07)}
.card-title{font-size:11px;font-weight:700;color:rgba(255,255,255,.4);text-transform:uppercase;letter-spacing:1px;margin-bottom:14px}
.field-label{font-size:13px;color:rgba(255,255,255,.6);margin-bottom:7px;display:block;font-weight:500}
input,select,textarea{width:100%;background:#0f172a;border:1.5px solid rgba(255,255,255,.12);border-radius:12px;color:#fff;font-size:16px;padding:14px 16px;outline:none;-webkit-appearance:none;appearance:none;font-family:inherit}
input:focus,select:focus{border-color:#3b82f6;box-shadow:0 0 0 3px rgba(59,130,246,.15)}
select{background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%2364748b' d='M6 8L1 3h10z'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 14px center;padding-right:40px}
select option{background:#1e293b}
.ore-row{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.ore-input-wrap{position:relative}
.ore-suffix{position:absolute;right:14px;top:50%;transform:translateY(-50%);color:rgba(255,255,255,.3);font-size:14px;pointer-events:none}
.quick-btns{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}
.quick-btn{background:rgba(59,130,246,.15);border:1px solid rgba(59,130,246,.3);color:#93c5fd;border-radius:8px;padding:7px 12px;font-size:13px;cursor:pointer;font-family:inherit}
.quick-btn:active{background:rgba(59,130,246,.3)}
.submit-btn{width:100%;background:linear-gradient(135deg,#2563eb,#1d4ed8);color:#fff;border:none;border-radius:14px;padding:18px;font-size:17px;font-weight:700;cursor:pointer;letter-spacing:.3px;display:flex;align-items:center;justify-content:center;gap:10px}
.submit-btn:active{opacity:.85;transform:scale(.99)}
.submit-btn:disabled{background:#334155;color:rgba(255,255,255,.4);cursor:not-allowed}
.flash{padding:14px 16px;border-radius:12px;font-size:14px;font-weight:600;display:flex;align-items:center;gap:10px}
.flash.success{background:rgba(34,197,94,.15);border:1px solid rgba(34,197,94,.3);color:#86efac}
.flash.error{background:rgba(239,68,68,.15);border:1px solid rgba(239,68,68,.3);color:#fca5a5}
.storico{margin-top:4px}
.storico-item{background:#0f172a;border-radius:10px;padding:12px 14px;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center}
.storico-data{font-size:13px;color:rgba(255,255,255,.5)}
.storico-ore{font-size:20px;font-weight:800;color:#3b82f6}
.storico-cantiere{font-size:12px;color:rgba(255,255,255,.4);margin-top:2px}
.storico-pausa{font-size:11px;color:rgba(255,255,255,.3)}
.date-display{font-size:28px;font-weight:800;color:#fff;text-align:center;padding:8px 0 4px}
.date-sub{font-size:13px;color:rgba(255,255,255,.4);text-align:center;margin-bottom:4px}
.ore-nette-preview{background:rgba(59,130,246,.1);border:1px solid rgba(59,130,246,.2);border-radius:10px;padding:10px 14px;text-align:center;font-size:13px;color:#93c5fd;margin-top:10px;display:none}
.switch-link{text-align:center;padding:8px 0 20px;font-size:13px;color:rgba(255,255,255,.3)}
.switch-link a{color:#60a5fa;text-decoration:none}
</style>
</head>
<body>
<div class="header">
  <div class="header-top">
    <div class="logo"><i class="fa fa-building"></i> {{ azienda }}</div>
    <a href="/logout" class="logout-btn"><i class="fa fa-sign-out-alt"></i> {{ t.logout }}</a>
  </div>
  <div class="user-name">{{ t.ciao }}, {{ nome }}!</div>
  <!-- Lang switcher -->
  <div style="display:flex;gap:6px;margin-top:10px;flex-wrap:wrap">
    {% for code, l in langs.items() %}
    <form method="POST" action="/set-lang" style="display:inline">
      <input type="hidden" name="lang" value="{{ code }}">
      <input type="hidden" name="next" value="/mobile">
      <button type="submit" style="background:rgba(255,255,255,{{ '.25' if current_lang==code else '.08' }});border:1px solid rgba(255,255,255,{{ '.4' if current_lang==code else '.15' }});border-radius:20px;padding:4px 10px;font-size:13px;color:rgba(255,255,255,{{ '.9' if current_lang==code else '.5' }});cursor:pointer">{{ l.flag }}</button>
    </form>
    {% endfor %}
  </div>
  <div class="user-sub"><i class="fa fa-circle" style="color:#22c55e;font-size:8px"></i> {{ ruolo }}</div>
</div>

<div class="content">
  {% if flash_msg %}
  <div class="flash {{ flash_type }}">
    <i class="fa {{ 'fa-check-circle' if flash_type=='success' else 'fa-exclamation-circle' }}"></i>
    {{ flash_msg }}
  </div>
  {% endif %}

  <!-- Form inserimento ore -->
  <form method="POST" action="/mobile/inserisci" id="form-ore">
    <div class="card">
      <div class="card-title"><i class="fa fa-calendar-day"></i> {{ t.day_label }}</div>
      <div class="date-display">{{ oggi_display }}</div>
      <div class="date-sub">{{ oggi_giorno }}</div>
      <input type="date" name="data" id="data-input" value="{{ oggi }}" style="margin-top:10px" onchange="aggiornaData(this.value)">
    </div>

    <div class="card">
      <div class="card-title"><i class="fa fa-store"></i> {{ t.site }}</div>
      <select name="cantiere_id" required>
        <option value="">{{ t.select_site }}</option>
        {% for c in cantieri %}
        <option value="{{ c.id }}">{{ c.nome }}</option>
        {% endfor %}
      </select>
    </div>

    <div class="card">
      <div class="card-title"><i class="fa fa-clock"></i> {{ t.work_hours }}</div>
      <div class="ore-row">
        <div>
          <label class="field-label">{{ t.total_hours }}</label>
          <div class="ore-input-wrap">
            <input type="number" name="ore" id="ore-input" min="0" max="24" step="0.5"
                   placeholder="8" inputmode="decimal" onchange="calcolaOreNette()" oninput="calcolaOreNette()">
            <span class="ore-suffix">h</span>
          </div>
          <div class="quick-btns">
            {% for h in [6,7,8,9,10] %}
            <button type="button" class="quick-btn" onclick="setOre({{ h }})">{{ h }}h</button>
            {% endfor %}
          </div>
        </div>
        <div>
          <label class="field-label">{{ t.break }}</label>
          <div class="ore-input-wrap">
            <select name="pausa" id="pausa-input" onchange="calcolaOreNette()">
              <option value="0">{{ t.no_break }}</option>
              <option value="0.5">30 min</option>
              <option value="1" selected>1 ora</option>
              <option value="1.5">1h 30m</option>
              <option value="2">2 ore</option>
            </select>
          </div>
        </div>
      </div>
      <div id="ore-nette-preview" class="ore-nette-preview"></div>
    </div>

    <div class="card">
      <div class="card-title"><i class="fa fa-comment"></i> {{ t.notes }}</div>
      <textarea name="note" rows="2" placeholder="{{ t.note_ph }}" style="resize:none"></textarea>
    </div>

    <button type="submit" class="submit-btn" id="submit-btn" onclick="this.disabled=true;this.innerHTML='<i class=\'fa fa-spinner fa-spin\'></i> Invio...';this.closest('form').submit()">
      <i class="fa fa-check-circle"></i> {{ t.submit_hours }}
    </button>
  </form>

  <!-- Storico ultimi 5 giorni -->
  {% if storico %}
  <div class="card">
    <div class="card-title"><i class="fa fa-history"></i> {{ t.last_entries }}</div>
    <div class="storico">
      {% for p in storico %}
      <div class="storico-item">
        <div>
          <div class="storico-data">{{ p.data }}</div>
          <div class="storico-cantiere">{{ p.cantiere_nome or t.no_site }}</div>
          {% if p.pausa_ore %}<div class="storico-pausa">Pausa: {{ p.pausa_ore }}h</div>{% endif %}
        </div>
        <div style="text-align:right">
          <div class="storico-ore">{{ p.ore_totali or '–' }}h</div>
          {% if p.ore_nette %}<div style="font-size:11px;color:rgba(255,255,255,.3)">{{ t.net_hours_lbl }}: {{ p.ore_nette }}{{ t.hours_suffix }}</div>{% endif %}
        </div>
      </div>
      {% endfor %}
    </div>
  </div>
  {% endif %}

  <!-- Navigation buttons -->
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
    <a href="/mobile/mie-ore" style="background:linear-gradient(135deg,#1e40af,#3b82f6);color:#fff;border:none;border-radius:14px;padding:16px;font-size:14px;font-weight:700;text-align:center;text-decoration:none;display:flex;flex-direction:column;align-items:center;gap:6px">
      <i class="fa fa-chart-bar" style="font-size:22px"></i>
      {{ t.my_hours_title }}
    </a>
    <a href="/mobile/calendario" style="background:linear-gradient(135deg,#92400e,#f59e0b);color:#fff;border:none;border-radius:14px;padding:16px;font-size:14px;font-weight:700;text-align:center;text-decoration:none;display:flex;flex-direction:column;align-items:center;gap:6px">
      <i class="fa fa-hard-hat" style="font-size:22px"></i>
      {{ t.work_calendar_btn }}
    </a>
  </div>
  <a href="/mobile/spese" style="background:linear-gradient(135deg,#064e3b,#10b981);color:#fff;border:none;border-radius:14px;padding:16px;font-size:14px;font-weight:700;text-align:center;text-decoration:none;display:flex;align-items:center;justify-content:center;gap:10px">
    <i class="fa fa-receipt" style="font-size:22px"></i>
    {{ t.spese_btn }}
  </a>
  <div class="switch-link">
    <a href="/dashboard"><i class="fa fa-desktop"></i> {{ t.desktop_ver }}</a>
  </div>
</div>

<script>
const MESI = ['gen','feb','mar','apr','mag','giu','lug','ago','set','ott','nov','dic'];
const GIORNI = {{ t.days | tojson }};

function aggiornaData(val) {
  if (!val) return;
  const d = new Date(val + 'T00:00:00');
  document.querySelector('.date-display').textContent =
    d.getDate() + ' ' + MESI[d.getMonth()] + ' ' + d.getFullYear();
  document.querySelector('.date-sub').textContent = GIORNI[d.getDay()];
}

function setOre(h) {
  document.getElementById('ore-input').value = h;
  calcolaOreNette();
}

function calcolaOreNette() {
  const ore   = parseFloat(document.getElementById('ore-input').value) || 0;
  const pausa = parseFloat(document.getElementById('pausa-input').value) || 0;
  const nette = Math.max(0, ore - pausa);
  const el = document.getElementById('ore-nette-preview');
  if (ore > 0) {
    el.style.display = 'block';
    el.innerHTML = '<i class="fa fa-info-circle"></i> Ore nette: <strong>' + nette.toFixed(1) + 'h</strong>' +
      (pausa > 0 ? ' (' + ore + 'h - ' + pausa + 'h pausa)' : '');
  } else {
    el.style.display = 'none';
  }
}

// Prevent double submit
document.getElementById('form-ore').addEventListener('submit', function() {
  document.getElementById('submit-btn').disabled = true;
  document.getElementById('submit-btn').innerHTML = '<i class="fa fa-spinner fa-spin"></i> {{ t.sending }}';
});
</script>
</body>
</html>"""


@app.route('/mobile')
@login_required
def mobile():
    db = get_db()
    uid = session['user_id']
    cantieri = db.execute(
        "SELECT id, nome FROM cantieri WHERE attivo=1 ORDER BY nome").fetchall()
    storico_raw = db.execute(
        """SELECT p.*, c.nome as cantiere_nome
           FROM presenze p LEFT JOIN cantieri c ON c.id=p.cantiere_id
           WHERE p.utente_id=? ORDER BY p.data DESC, p.id DESC LIMIT 7""",
        (uid,)).fetchall()
    db.close()

    storico = []
    for p in storico_raw:
        row = dict(p)
        ore = row.get('ore_totali') or 0
        # Pausa non salvata direttamente — mostra solo ore_totali
        row['ore_nette'] = None
        row['pausa_ore'] = None
        storico.append(row)

    from datetime import date
    oggi = date.today()
    giorni_it = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato','Domenica']
    mesi_it = ['gennaio','febbraio','marzo','aprile','maggio','giugno',
               'luglio','agosto','settembre','ottobre','novembre','dicembre']

    msgs = session.get('_flashes', [])
    flash_msg = flash_type = None
    if msgs:
        flash_type, flash_msg = msgs[-1]
        session['_flashes'] = []

    azienda = get_setting('nome_azienda', 'Accesso Fiere')
    t    = get_lang()
    lang = session.get('lang', 'it')
    return render_template_string(
        MOBILE_TMPL,
        nome=session.get('nome', ''),
        ruolo='Admin' if session.get('ruolo') == 'admin' else 'Dipendente',
        azienda=azienda,
        cantieri=cantieri,
        storico=storico,
        oggi=oggi.isoformat(),
        oggi_display=f"{oggi.day} {t['months'][oggi.month]} {oggi.year}",
        oggi_giorno=t['days'][oggi.weekday()],
        flash_msg=flash_msg,
        flash_type=flash_type or 'success',
        t=t, lang=lang, langs=LANGS, current_lang=lang
    )


@app.route('/mobile/inserisci', methods=['POST'])
@login_required
def mobile_inserisci():
    uid = session['user_id']
    data = request.form.get('data') or date.today().isoformat()
    cantiere_id = request.form.get('cantiere_id') or None
    ore_str = request.form.get('ore', '').strip()
    pausa_str = request.form.get('pausa', '0')
    note = request.form.get('note', '').strip()

    try:
        ore = float(ore_str) if ore_str else 0
        pausa = float(pausa_str) if pausa_str else 0
        ore_nette = max(0, ore - pausa)
    except ValueError:
        flash('Inserisci un valore valido per le ore.', 'error')
        return redirect(url_for('mobile'))

    if ore <= 0:
        flash('Inserisci le ore lavorate.', 'error')
        return redirect(url_for('mobile'))

    if not cantiere_id:
        flash('Seleziona la fiera.', 'error')
        return redirect(url_for('mobile'))

    nota_completa = f"Pausa: {pausa}h — Ore nette: {ore_nette}h"
    if note:
        nota_completa += f" | {note}"

    db = get_db()
    # Tutti i dipendenti (anche supervisori) mandano la richiesta — l'admin approva sempre
    # Permetti più richieste per lo stesso giorno se cantieri diversi
    gia_pendente = db.execute(
        "SELECT id FROM richieste_presenze WHERE utente_id=? AND data=? AND cantiere_id=? AND stato='in_attesa'",
        (uid, data, int(cantiere_id))).fetchone()
    if gia_pendente:
        flash('Hai già una richiesta in attesa per questo giorno e cantiere.', 'error')
        db.close(); return redirect(url_for('mobile'))

    # ora_entrata e ora_uscita sono NOT NULL nel DB — usiamo valori simbolici per le richieste mobile
    db.execute(
        "INSERT INTO richieste_presenze (utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note) VALUES (?,?,?,?,?,?,?)",
        (uid, data, '00:00', '00:00', ore_nette, int(cantiere_id), nota_completa))
    safe_commit(db)

    db.close()
    # Notifica email admin — in background per non bloccare la risposta
    email_admin = get_setting('email_notifiche', '')
    nome_dip = f"{session.get('nome','')} {session.get('cognome','')}".strip()
    if email_admin:
        import threading
        _base = request.host_url.rstrip('/')
        def _send():
            try:
                send_email(email_admin,
                    f'[ACCESSO FIERE] Richiesta ore da {nome_dip}',
                    f'<p><b>{nome_dip}</b> ha inviato una richiesta ore per il <b>{data}</b>.<br>'
                    f'Ore nette: {ore_nette}h — Pausa: {pausa}h<br>Note: {note or "–"}</p>'
                    f'<p><a href="{_base}/admin/richieste" style="background:#2563eb;color:#fff;padding:10px 20px;border-radius:8px;text-decoration:none;font-weight:700">Approva o rifiuta →</a></p>')
            except: pass
        threading.Thread(target=_send, daemon=True).start()
    flash(f'✅ Richiesta inviata per {data} ({ore_nette}h) — in attesa di approvazione.', 'success')
    return redirect(url_for('mobile'))


# ══════════════════════════════════════════════════════════
#  MOBILE — LE MIE ORE
# ══════════════════════════════════════════════════════════

MIE_ORE_TMPL = """<!DOCTYPE html>
<html lang="{{ lang }}" dir="{{ t.dir }}">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>{{ t.my_hours_title }}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{background:#0f172a;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;min-height:100vh}
.header{background:linear-gradient(135deg,#1e3a5f,#0f172a);padding:20px 20px 16px;border-bottom:1px solid rgba(255,255,255,.08)}
.header-top{display:flex;justify-content:space-between;align-items:center;margin-bottom:4px}
.back-btn{font-size:13px;color:rgba(255,255,255,.6);text-decoration:none;padding:8px 12px;border:1px solid rgba(255,255,255,.15);border-radius:8px}
.title{font-size:20px;font-weight:800;color:#fff;margin-top:8px}
.content{padding:16px;max-width:480px;margin:0 auto}
.card{background:#1e293b;border-radius:16px;padding:16px;border:1px solid rgba(255,255,255,.07);margin-bottom:14px}
.card-title{font-size:11px;font-weight:700;color:rgba(255,255,255,.4);text-transform:uppercase;letter-spacing:1px;margin-bottom:12px}
select,input{width:100%;background:#0f172a;border:1.5px solid rgba(255,255,255,.12);border-radius:10px;color:#fff;font-size:15px;padding:12px 14px;outline:none;-webkit-appearance:none;appearance:none;font-family:inherit}
select{background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%2364748b' d='M6 8L1 3h10z'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 14px center;padding-right:36px}
select option{background:#1e293b}
.btn-filter{width:100%;background:#2563eb;color:#fff;border:none;border-radius:10px;padding:13px;font-size:15px;font-weight:700;cursor:pointer;margin-top:10px}
.btn-pdf{width:100%;background:linear-gradient(135deg,#dc2626,#b91c1c);color:#fff;border:none;border-radius:12px;padding:15px;font-size:15px;font-weight:700;cursor:pointer;display:flex;align-items:center;justify-content:center;gap:8px;text-decoration:none}
.totale-box{background:linear-gradient(135deg,#1e3a5f,#0f172a);border-radius:14px;padding:18px;text-align:center;margin-bottom:14px;border:1px solid rgba(59,130,246,.3)}
.totale-num{font-size:42px;font-weight:900;color:#60a5fa;line-height:1}
.totale-label{font-size:13px;color:rgba(255,255,255,.5);margin-top:6px}
.totale-sub{font-size:12px;color:rgba(255,255,255,.3);margin-top:4px}
.row{background:#0f172a;border-radius:10px;padding:13px 14px;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center}
.row-data{font-size:13px;color:rgba(255,255,255,.5)}
.row-cantiere{font-size:12px;color:rgba(255,255,255,.35);margin-top:2px}
.row-ore{font-size:20px;font-weight:800;color:#60a5fa}
.badge{padding:3px 8px;border-radius:99px;font-size:10px;font-weight:700;margin-top:4px;display:inline-block}
.badge-wait{background:rgba(245,158,11,.15);color:#fbbf24}
.badge-ok{background:rgba(34,197,94,.15);color:#86efac}
.badge-no{background:rgba(239,68,68,.15);color:#fca5a5}
.empty{text-align:center;padding:40px 20px;color:rgba(255,255,255,.3)}
.empty i{font-size:40px;margin-bottom:12px;display:block}
</style>
</head>
<body>
<div class="header">
  <div class="header-top">
    <button onclick="window.location.replace('/mobile')" class="back-btn" style="cursor:pointer;background:none;border:1px solid rgba(255,255,255,.15);color:rgba(255,255,255,.6);font-size:13px;padding:8px 12px;border-radius:8px;font-family:inherit"><i class="fa fa-arrow-{{ 'right' if t.dir=='rtl' else 'left' }}"></i> {{ t.back }}</button>
    <span style="font-size:12px;color:rgba(255,255,255,.3)">{{ t.my_hours_title }}</span>
  </div>
  <div class="title">{{ nome }} {{ cognome }}</div>
  <div style="display:flex;gap:6px;margin-top:10px;flex-wrap:wrap">
    {% for code, l in langs.items() %}
    <form method="POST" action="/set-lang" style="display:inline">
      <input type="hidden" name="lang" value="{{ code }}">
      <input type="hidden" name="next" value="/mobile/mie-ore">
      <button type="submit" style="background:rgba(255,255,255,{{ '.25' if current_lang==code else '.08' }});border:1px solid rgba(255,255,255,{{ '.4' if current_lang==code else '.15' }});border-radius:20px;padding:4px 10px;font-size:13px;color:rgba(255,255,255,{{ '.9' if current_lang==code else '.5' }});cursor:pointer">{{ l.flag }}</button>
    </form>
    {% endfor %}
  </div>
</div>

<div class="content">
  <!-- Filtro mese -->
  <div class="card">
    <div class="card-title"><i class="fa fa-filter"></i> {{ t.filter_by_month }}</div>
    <form method="GET">
      <select name="mese" onchange="this.form.submit()">
        <option value="">{{ t.all_months_opt }}</option>
        {% for m in mesi_disponibili %}
        <option value="{{ m.val }}" {{ 'selected' if m.val == mese_sel }}>{{ m.label }}</option>
        {% endfor %}
      </select>
    </form>
  </div>

  <!-- Totale ore -->
  {% if righe %}
  <div class="totale-box">
    <div class="totale-num">{{ totale_ore }}h</div>
    <div class="totale-label">{{ t.approved_total }}{% if mese_sel %} — {{ mese_label }}{% endif %}</div>
    <div class="totale-sub">{{ n_approvate }} {{ t.days_worked_lbl }} · {{ n_attesa }} {{ t.pending_lbl }}</div>
  </div>

  <!-- Download PDF -->
  <a href="/mobile/mie-ore/pdf?mese={{ mese_sel }}" class="btn-pdf" style="margin-bottom:14px">
    <i class="fa fa-file-pdf"></i> {{ t.download_pdf_btn }}
  </a>

  <!-- Lista timbrature -->
  <div class="card">
    <div class="card-title"><i class="fa fa-list"></i> {{ t.timbrature }} ({{ righe|length }})</div>
    {% for r in righe %}
    <div class="row">
      <div>
        <div class="row-data">{{ r.data }}</div>
        <div class="row-cantiere">{{ r.cantiere_nome or t.no_site }}</div>
        {% if r.stato == 'in_attesa' %}<span class="badge badge-wait">⏳ {{ t.status_waiting }}</span>
        {% elif r.stato == 'approvata' %}<span class="badge badge-ok">✓ {{ t.status_approved }}</span>
        {% elif r.stato == 'rifiutata' %}<span class="badge badge-no">✗ {{ t.status_rejected }}</span>{% endif %}
      </div>
      <div style="text-align:right">
        <div class="row-ore">{{ r.ore_totali or '–' }}h</div>
      </div>
    </div>
    {% endfor %}
  </div>

  {% else %}
  <div class="empty">
    <i class="fa fa-clock"></i>
    <div>{{ t.no_entries }}{% if mese_sel %} — {{ mese_label }}{% endif %}</div>
  </div>
  {% endif %}
</div>
</body>
</html>"""


@app.route('/mobile/mie-ore')
@login_required
def mobile_mie_ore():
    uid = session['user_id']
    mese_sel = request.args.get('mese', '')
    db = get_db()

    # Mesi disponibili — leggiamo da ENTRAMBE le tabelle
    tutti = db.execute(
        """SELECT DISTINCT substr(data,1,7) as m FROM presenze WHERE utente_id=?
           UNION
           SELECT DISTINCT substr(data,1,7) as m FROM richieste_presenze WHERE utente_id=? AND stato='in_attesa'
           ORDER BY m DESC""",
        (uid, uid)).fetchall()
    mesi_nomi = ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
                 'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']
    mesi_disponibili = []
    for row in tutti:
        val = row['m']
        anno, mese = val.split('-')
        label = f"{mesi_nomi[int(mese)]} {anno}"
        mesi_disponibili.append({'val': val, 'label': label})

    # Righe approvate = presenze reali (fonte di verità)
    if mese_sel:
        pres = db.execute(
            """SELECT p.id, p.data, p.ora_entrata, p.ora_uscita, p.ore_totali, p.note,
                      c.nome as cantiere_nome, 'approvata' as stato
               FROM presenze p LEFT JOIN cantieri c ON c.id=p.cantiere_id
               WHERE p.utente_id=? AND substr(p.data,1,7)=?
               ORDER BY p.data DESC""", (uid, mese_sel)).fetchall()
        attesa = db.execute(
            """SELECT r.id, r.data, r.ora_entrata, r.ora_uscita, r.ore_totali, r.note,
                      c.nome as cantiere_nome, r.stato
               FROM richieste_presenze r LEFT JOIN cantieri c ON c.id=r.cantiere_id
               WHERE r.utente_id=? AND substr(r.data,1,7)=? AND r.stato='in_attesa'
               ORDER BY r.data DESC""", (uid, mese_sel)).fetchall()
    else:
        pres = db.execute(
            """SELECT p.id, p.data, p.ora_entrata, p.ora_uscita, p.ore_totali, p.note,
                      c.nome as cantiere_nome, 'approvata' as stato
               FROM presenze p LEFT JOIN cantieri c ON c.id=p.cantiere_id
               WHERE p.utente_id=? ORDER BY p.data DESC LIMIT 50""", (uid,)).fetchall()
        attesa = db.execute(
            """SELECT r.id, r.data, r.ora_entrata, r.ora_uscita, r.ore_totali, r.note,
                      c.nome as cantiere_nome, r.stato
               FROM richieste_presenze r LEFT JOIN cantieri c ON c.id=r.cantiere_id
               WHERE r.utente_id=? AND r.stato='in_attesa'
               ORDER BY r.data DESC LIMIT 20""", (uid,)).fetchall()
    db.close()

    righe = [dict(r) for r in pres] + [dict(r) for r in attesa]
    righe.sort(key=lambda x: x['data'], reverse=True)
    ore_approvate = sum(r['ore_totali'] or 0 for r in righe if r['stato'] == 'approvata')
    n_approvate = sum(1 for r in righe if r['stato'] == 'approvata')
    n_attesa = sum(1 for r in righe if r['stato'] == 'in_attesa')

    # Mese label
    mese_label = ''
    if mese_sel:
        anno, mese = mese_sel.split('-')
        mese_label = f"{mesi_nomi[int(mese)]} {anno}"

    t    = get_lang()
    lang = session.get('lang', 'it')
    # Usa nomi mesi tradotti
    mesi_nomi_t = t['months']
    mesi_disponibili2 = []
    for row in tutti:
        val = row['m']
        anno2, mese2 = val.split('-')
        label2 = f"{mesi_nomi_t[int(mese2)]} {anno2}"
        mesi_disponibili2.append({'val': val, 'label': label2})
    if mese_sel:
        anno3, mese3 = mese_sel.split('-')
        mese_label = f"{mesi_nomi_t[int(mese3)]} {anno3}"
    return render_template_string(
        MIE_ORE_TMPL,
        nome=session.get('nome',''),
        cognome=session.get('cognome',''),
        righe=righe,
        mesi_disponibili=mesi_disponibili2,
        mese_sel=mese_sel,
        mese_label=mese_label,
        totale_ore=round(ore_approvate, 1),
        n_approvate=n_approvate,
        n_attesa=n_attesa,
        t=t, lang=lang, langs=LANGS, current_lang=lang,
    )


@app.route('/mobile/mie-ore/pdf')
@login_required
def mobile_mie_ore_pdf():
    uid = session['user_id']
    mese_sel = request.args.get('mese', '')
    lang = session.get('lang', 'it')
    t = get_lang()
    db = get_db()
    u = db.execute("SELECT * FROM utenti WHERE id=?", (uid,)).fetchone()

    if mese_sel:
        righe = db.execute(
            """SELECT p.id, p.data, p.ora_entrata, p.ora_uscita, p.ore_totali, p.note,
                      c.nome as cantiere_nome, 'approvata' as stato
               FROM presenze p LEFT JOIN cantieri c ON c.id=p.cantiere_id
               WHERE p.utente_id=? AND substr(p.data,1,7)=?
               ORDER BY p.data ASC""", (uid, mese_sel)).fetchall()
        spese_rows = db.execute(
            """SELECT s.data, s.categoria, s.descrizione, s.importo,
                      v.targa as veicolo_targa
               FROM spese_rimborso s LEFT JOIN veicoli v ON v.id=s.veicolo_id
               WHERE s.utente_id=? AND s.stato='approvata' AND substr(s.data,1,7)=?
               ORDER BY s.data ASC""", (uid, mese_sel)).fetchall()
    else:
        righe = db.execute(
            """SELECT p.id, p.data, p.ora_entrata, p.ora_uscita, p.ore_totali, p.note,
                      c.nome as cantiere_nome, 'approvata' as stato
               FROM presenze p LEFT JOIN cantieri c ON c.id=p.cantiere_id
               WHERE p.utente_id=?
               ORDER BY p.data ASC""", (uid,)).fetchall()
        spese_rows = db.execute(
            """SELECT s.data, s.categoria, s.descrizione, s.importo,
                      v.targa as veicolo_targa
               FROM spese_rimborso s LEFT JOIN veicoli v ON v.id=s.veicolo_id
               WHERE s.utente_id=? AND s.stato='approvata'
               ORDER BY s.data ASC""", (uid,)).fetchall()

    azienda = get_setting('nome_azienda', 'Azienda')
    db.close()

    righe = [dict(r) for r in righe]
    spese_rows = [dict(s) for s in spese_rows]
    totale_ore = round(sum(r['ore_totali'] or 0 for r in righe), 1)
    totale_spese = round(sum(s['importo'] or 0 for s in spese_rows), 2)

    mesi_nomi = t.get('months', ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
                 'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre'])
    mese_label = t.get('all_months_label', 'Tutti i mesi')
    if mese_sel:
        try:
            anno, mese = mese_sel.split('-')
            mese_label = f"{mesi_nomi[int(mese)]} {anno}"
        except:
            mese_label = mese_sel

    nome_completo = f"{u['nome']} {u['cognome']}"
    oggi = date.today().strftime('%d/%m/%Y')

    # Traduzioni per il PDF
    lbl = {
        'title': t.get('report_title', 'Report Ore Lavorate'),
        'period': t.get('report_period', 'Periodo'),
        'days': t.get('report_days', 'Giorni lavorati'),
        'total_hours': t.get('report_total_hours', 'Ore totali'),
        'col_date': t.get('report_date', 'Data'),
        'col_site': t.get('report_site', 'Cantiere'),
        'col_hours': t.get('report_hours', 'Ore'),
        'col_notes': t.get('report_notes', 'Note'),
        'total_row': t.get('report_total_row', 'TOTALE'),
        'spese_title': t.get('report_spese_title', 'Spese Rimborsabili Approvate'),
        'spese_cat': t.get('report_spese_cat', 'Categoria'),
        'spese_desc': t.get('report_spese_desc', 'Descrizione'),
        'spese_vehicle': t.get('report_spese_vehicle', 'Veicolo'),
        'spese_amount': t.get('report_spese_amount', 'Importo'),
        'spese_total': t.get('report_spese_total', 'TOTALE RIMBORSI'),
        'footer': t.get('report_footer', 'Report generato automaticamente da'),
        'employee': 'Dipendente' if lang=='it' else ('Employé' if lang=='fr' else ('Employee' if lang=='en' else 'الموظف')),
    }
    rtl = 'rtl' if lang == 'ar' else 'ltr'

    html = f"""<!DOCTYPE html>
<html dir="{rtl}"><head><meta charset="UTF-8">
<style>
body{{font-family:Arial,sans-serif;font-size:12px;color:#1e293b;margin:0;padding:30px;direction:{rtl}}}
h1{{font-size:20px;font-weight:800;color:#0f172a;margin-bottom:4px}}
h2{{font-size:15px;font-weight:700;color:#0f172a;margin:28px 0 12px;padding-bottom:6px;border-bottom:2px solid #10b981}}
.sub{{color:#64748b;font-size:13px;margin-bottom:24px}}
.info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:24px;background:#f8fafc;padding:16px;border-radius:8px}}
.info-item label{{font-size:10px;color:#94a3b8;text-transform:uppercase;letter-spacing:.5px;display:block}}
.info-item span{{font-size:14px;font-weight:700;color:#0f172a}}
table{{width:100%;border-collapse:collapse;margin-bottom:16px}}
thead th{{background:#0f172a;color:#fff;padding:10px 12px;text-align:{'right' if rtl=='rtl' else 'left'};font-size:11px;text-transform:uppercase;letter-spacing:.5px}}
tbody tr:nth-child(even){{background:#f8fafc}}
tbody td{{padding:9px 12px;border-bottom:1px solid #e2e8f0;font-size:12px}}
.total-row{{background:#0f172a!important;color:#fff;font-weight:700}}
.total-row td{{color:#fff!important;padding:12px;font-size:14px}}
.spese-table thead th{{background:#065f46}}
.spese-total-row{{background:#065f46!important}}
.spese-total-row td{{color:#fff!important;padding:12px;font-size:14px}}
.footer{{margin-top:30px;padding-top:16px;border-top:1px solid #e2e8f0;color:#94a3b8;font-size:11px;text-align:center}}
</style></head>
<body>
<h1>{lbl['title']}</h1>
<div class="sub">{azienda} — {oggi}</div>
<div class="info-grid">
  <div class="info-item"><label>{lbl['employee']}</label><span>{nome_completo}</span></div>
  <div class="info-item"><label>{lbl['period']}</label><span>{mese_label}</span></div>
  <div class="info-item"><label>{lbl['days']}</label><span>{len(righe)}</span></div>
  <div class="info-item"><label>{lbl['total_hours']}</label><span>{totale_ore}h</span></div>
</div>
<table>
<thead><tr><th>{lbl['col_date']}</th><th>{lbl['col_site']}</th><th>{lbl['col_hours']}</th><th>{lbl['col_notes']}</th></tr></thead>
<tbody>"""

    for r in righe:
        note = r.get('note') or ''
        if ' | ' in note:
            note = note.split(' | ', 1)[1]
        elif note.startswith('Pausa:'):
            note = ''
        html += f"<tr><td>{r['data']}</td><td>{r['cantiere_nome'] or '–'}</td><td><strong>{r['ore_totali'] or 0}h</strong></td><td>{note}</td></tr>"

    html += f"""<tr class="total-row"><td colspan="2">{lbl['total_row']}</td><td>{totale_ore}h</td><td>{len(righe)} {t.get('days_worked_lbl','giorni lavorati')}</td></tr>
</tbody></table>"""

    # Sezione spese se presenti
    if spese_rows:
        html += f"""<h2><i>💶</i> {lbl['spese_title']}</h2>
<table class="spese-table">
<thead><tr><th>{lbl['col_date']}</th><th>{lbl['spese_cat']}</th><th>{lbl['spese_desc']}</th><th>{lbl['spese_vehicle']}</th><th style="text-align:right">{lbl['spese_amount']}</th></tr></thead>
<tbody>"""
        for s in spese_rows:
            html += f"<tr><td>{s['data']}</td><td>{s['categoria']}</td><td>{s['descrizione'] or '–'}</td><td>{s['veicolo_targa'] or '–'}</td><td style='text-align:right;font-weight:700'>€ {s['importo']:.2f}</td></tr>"
        html += f"""<tr class="spese-total-row"><td colspan="4">{lbl['spese_total']}</td><td style="text-align:right">€ {totale_spese:.2f}</td></tr>
</tbody></table>"""

    html += f"""<div class="footer">{lbl['footer']} {azienda} — {oggi}</div>
</body></html>"""

    try:
        from weasyprint import HTML as WH
        pdf_bytes = WH(string=html).write_pdf()
        filename = f"ore_{nome_completo.replace(' ','_')}_{mese_sel or 'tutti'}.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Exception:
        html_print = html.replace('</body>', '<script>window.onload=function(){window.print()}</script></body>')
        return Response(html_print, mimetype='text/html')


# ══════════════════════════════════════════════════════════
#  POS — PIANO OPERATIVO DI SICUREZZA
# ══════════════════════════════════════════════════════════

POS_LIST_TMPL = """
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px">
  <div></div>
  <a href="/pos/nuovo" class="btn btn-primary"><i class="fa fa-plus"></i> Nuovo POS</a>
</div>
<div class="card">
  <div class="card-header"><h3><i class="fa fa-file-shield" style="color:var(--accent2)"></i> POS generati</h3></div>
  {% if pos_list %}
  <div class="table-wrap"><table>
    <thead><tr><th>Titolo</th><th>Fiera</th><th>Committente</th><th>Data inizio</th><th>Data fine</th><th>Lavoratori</th><th>Creato il</th><th></th></tr></thead>
    <tbody>{% for p in pos_list %}
    <tr>
      <td><strong>{{ p.titolo }}</strong></td>
      <td>{{ p.cantiere_nome or '–' }}<br><small style="color:var(--text-light)">{{ p.cantiere_indirizzo or '' }}</small></td>
      <td>{{ p.committente or '–' }}</td>
      <td style="font-family:monospace">{{ p.data_inizio or '–' }}</td>
      <td style="font-family:monospace">{{ p.data_fine or '–' }}</td>
      <td><span class="badge badge-purple">{{ p.num_lavoratori }} pers.</span></td>
      <td style="font-size:12px;color:var(--text-light)">{{ p.creato_il[:10] }}</td>
      <td style="display:flex;gap:6px">
        <a href="/pos/{{ p.id }}/pdf" class="btn btn-sm btn-danger" title="Scarica PDF"><i class="fa fa-file-pdf"></i> PDF</a>
        <a href="/pos/{{ p.id }}/modifica" class="btn btn-sm btn-secondary" title="Modifica"><i class="fa fa-pen"></i></a>
        <a href="/pos/{{ p.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare il POS?')"><i class="fa fa-trash"></i></a>
      </td>
    </tr>{% endfor %}</tbody>
  </table></div>
  {% else %}
  <div class="empty-state"><i class="fa fa-file-shield"></i><p>Nessun POS generato. <a href="/pos/nuovo">Crea il primo!</a></p></div>
  {% endif %}
</div>
"""

POS_FORM_TMPL = """
<div style="margin-bottom:16px"><a href="/pos" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Torna ai POS</a></div>
<form method="POST" action="{{ form_action }}">
<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-info-circle" style="color:var(--accent2)"></i> Dati cantiere</h3></div>
  <div class="card-body">
    <div class="form-row">
      <div class="form-group" style="grid-column:span 2">
        <label>Titolo POS *</label>
        <input name="titolo" value="{{ pos.titolo or '' }}" required placeholder="Es. POS MEC SPE 2026">
      </div>
      <div class="form-group"><label>Data documento</label><input type="date" name="data_documento" value="{{ pos.data_documento or today }}"></div>
    </div>
    <div class="form-row">
      <div class="form-group" style="grid-column:span 2">
        <label>Nome cantiere *</label>
        <input name="cantiere_nome" value="{{ pos.cantiere_nome or '' }}" required placeholder="Es. MEC SPE (Fiera Bologna) 2026">
      </div>
    </div>
    <div class="form-row">
      <div class="form-group" style="grid-column:span 3">
        <label>Indirizzo cantiere *</label>
        <input name="cantiere_indirizzo" value="{{ pos.cantiere_indirizzo or '' }}" required placeholder="Es. Piazza Costituzione 6, 40127 Bologna">
      </div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Committente</label><input name="committente" value="{{ pos.committente or '' }}" placeholder="Es. StandHouse srl"></div>
      <div class="form-group"><label>Responsabile dei lavori</label><input name="responsabile_lavori" value="{{ pos.responsabile_lavori or datore_lavoro }}" placeholder="Es. FATHI RACHID"></div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Data inizio cantiere</label><input type="date" name="data_inizio" value="{{ pos.data_inizio or '' }}"></div>
      <div class="form-group"><label>Data fine presunta</label><input type="date" name="data_fine" value="{{ pos.data_fine or '' }}"></div>
    </div>
    <div class="form-group">
      <label>Descrizione opere</label>
      <textarea name="opere" rows="2" placeholder="Es. Montaggio e smontaggio di stand/allestimenti temporanei">{{ pos.opere or '' }}</textarea>
    </div>
    <div class="form-group">
      <label>Note aggiuntive cantiere</label>
      <textarea name="note_cantiere" rows="2" placeholder="Informazioni specifiche della fiera...">{{ pos.note_cantiere or '' }}</textarea>
    </div>
  </div>
</div>

<div class="card" style="margin-bottom:20px">
  <div class="card-header"><h3><i class="fa fa-users" style="color:var(--accent2)"></i> Personale presente in cantiere</h3></div>
  <div class="card-body">
    <p style="font-size:13px;color:var(--text-light);margin-bottom:16px">Seleziona i dipendenti che figureranno nel POS. Il titolare viene aggiunto automaticamente.</p>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:10px">
      {% for d in dipendenti %}
      <label style="display:flex;align-items:center;gap:10px;background:var(--bg-secondary);padding:12px;border-radius:10px;cursor:pointer;border:2px solid transparent" 
             id="label-{{ d.id }}"
             onclick="toggleDip({{ d.id }})">
        <input type="checkbox" name="personale" value="{{ d.id }}" id="chk-{{ d.id }}"
               {{ 'checked' if d.id|string in (pos.personale_selezionati or []) }}
               style="width:18px;height:18px;accent-color:var(--accent2)">
        <div>
          <div style="font-weight:600;font-size:14px">{{ d.cognome }} {{ d.nome }}</div>
          <div style="font-size:12px;color:var(--text-light)">{{ d.mansione or d.titolo or 'Operaio' }}</div>
        </div>
      </label>
      {% endfor %}
    </div>
  </div>
</div>

<div style="display:flex;gap:12px;justify-content:flex-end">
  <a href="/pos" class="btn btn-secondary">Annulla</a>
  <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva POS</button>
  <button type="submit" name="genera_pdf" value="1" class="btn btn-danger"><i class="fa fa-file-pdf"></i> Salva e genera PDF</button>
</div>
</form>

<script>
function toggleDip(id) {
  const chk = document.getElementById('chk-' + id);
  const lbl = document.getElementById('label-' + id);
  lbl.style.borderColor = chk.checked ? 'var(--accent2)' : 'transparent';
}
document.addEventListener('DOMContentLoaded', function() {
  document.querySelectorAll('input[type=checkbox]').forEach(function(chk) {
    if (chk.checked) document.getElementById('label-' + chk.value).style.borderColor = 'var(--accent2)';
  });
});
</script>
"""


def _pos_pdf_html(pos_data, azienda_data, personale):
    """Genera HTML completo per il POS da convertire in PDF."""
    az = azienda_data
    p  = pos_data

    nome_az      = az.get('azienda', 'Accesso Fiere S.r.l.')
    sede_legale  = az.get('sede_legale', '')
    piva         = az.get('partita_iva', '')
    inail        = az.get('codice_inail', '')
    inps         = az.get('codice_inps', '')
    cciaa        = az.get('cciaa', '')
    rspp         = az.get('rspp', '')
    medico       = az.get('medico_competente', '')
    attivita     = az.get('attivita_azienda', 'Montaggio e smontaggio di Stand Fieristici')
    datore       = p.get('responsabile_lavori', rspp or '')

    cantiere_nome    = p.get('cantiere_nome', '')
    cantiere_ind     = p.get('cantiere_indirizzo', '')
    committente      = p.get('committente', '')
    data_inizio      = p.get('data_inizio', '')
    data_fine        = p.get('data_fine', '')
    data_doc         = p.get('data_documento', date.today().isoformat())
    opere            = p.get('opere', 'Montaggio e smontaggio di stand/allestimenti temporanei.')
    note_cantiere    = p.get('note_cantiere', '')

    # Formatta date in italiano
    def fmt_date(d):
        if not d:
            return '–'
        try:
            from datetime import datetime as _dt
            return _dt.strptime(d, '%Y-%m-%d').strftime('%d/%m/%Y')
        except Exception:
            return d

    # Intestazione tabelle personale
    righe_personale = ''
    for idx, dip in enumerate(personale):
        mansione = dip.get('mansione') or dip.get('titolo') or 'Operaio'
        nome_completo = f"{dip.get('cognome','').upper()} {dip.get('nome','').upper()}"
        righe_personale += f'<tr><td>{nome_completo}</td><td>{mansione}</td></tr>'

    num_lavoratori = len(personale)

    html = f"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<style>
  @page {{ margin: 2cm 2cm 2.5cm 2cm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a1a1a; line-height: 1.5; }}

  /* Header ripetuto su ogni pagina */
  .page-header {{ display: flex; justify-content: space-between; align-items: center;
    border-bottom: 2px solid #1e3a5f; padding-bottom: 8px; margin-bottom: 16px; }}
  .page-header .az {{ font-weight: 800; font-size: 13px; color: #1e3a5f; }}
  .page-header .doc-title {{ font-size: 11px; color: #666; text-align: center; }}
  .page-header .cantiere-short {{ font-size: 10px; color: #666; text-align: right; }}

  h1 {{ font-size: 22px; font-weight: 900; color: #1e3a5f; text-align: center; margin: 20px 0 8px; }}
  h2 {{ font-size: 18px; font-weight: 800; color: #1e3a5f; text-align: center; margin-bottom: 30px; }}
  h3 {{ font-size: 13px; font-weight: 700; color: #1e3a5f; margin: 20px 0 8px; border-bottom: 1px solid #e2e8f0; padding-bottom: 4px; }}
  h4 {{ font-size: 12px; font-weight: 700; margin: 14px 0 6px; }}
  p {{ margin-bottom: 8px; }}
  ul {{ margin: 6px 0 10px 20px; }}
  li {{ margin-bottom: 4px; }}

  .cover {{ page-break-after: always; display: flex; flex-direction: column; align-items: center;
    justify-content: center; min-height: 22cm; text-align: center; }}
  .cover-logo {{ font-size: 32px; font-weight: 900; color: #1e3a5f; letter-spacing: 2px; margin-bottom: 8px; }}
  .cover-subtitle {{ font-size: 16px; color: #475569; margin-bottom: 40px; }}
  .cover-box {{ background: #f8fafc; border: 2px solid #1e3a5f; border-radius: 12px;
    padding: 30px 40px; margin: 20px 0; width: 100%; max-width: 500px; }}
  .cover-box .label {{ font-size: 10px; color: #94a3b8; text-transform: uppercase; letter-spacing: 1px; }}
  .cover-box .value {{ font-size: 14px; font-weight: 700; color: #1e3a5f; margin-bottom: 14px; }}
  .cover-dlgs {{ font-size: 13px; color: #64748b; margin-top: 20px; font-style: italic; }}
  .cover-date {{ font-size: 12px; color: #94a3b8; margin-top: 10px; }}

  table {{ width: 100%; border-collapse: collapse; margin: 12px 0; font-size: 11px; }}
  thead th {{ background: #1e3a5f; color: #fff; padding: 8px 10px; text-align: left; }}
  tbody td {{ padding: 7px 10px; border-bottom: 1px solid #e2e8f0; }}
  tbody tr:nth-child(even) {{ background: #f8fafc; }}
  .table-2col td:first-child {{ font-weight: 600; color: #475569; width: 45%; background: #f8fafc; }}

  .section {{ page-break-inside: avoid; margin-bottom: 20px; }}
  .alert-box {{ background: #fef3c7; border-left: 4px solid #f59e0b; padding: 10px 14px;
    margin: 10px 0; border-radius: 4px; font-size: 11px; }}
  .firma-box {{ margin-top: 40px; display: flex; justify-content: space-between; }}
  .firma-item {{ text-align: center; width: 45%; }}
  .firma-line {{ border-top: 1px solid #1a1a1a; margin-top: 40px; padding-top: 4px; font-size: 10px; color: #666; }}

  .chapter {{ page-break-before: always; }}
  .page-num {{ text-align: right; font-size: 9px; color: #94a3b8; margin-top: 4px; }}
</style>
</head>
<body>

<!-- ═══════════════════════════════════════════════
     COPERTINA
════════════════════════════════════════════════ -->
<div class="cover">
  <div class="cover-logo">{nome_az}</div>
  <div class="cover-subtitle">PIANO OPERATIVO DI SICUREZZA</div>
  <div class="cover-dlgs">(D.Lgs. 81/08)</div>

  <div class="cover-box">
    <div class="label">Fiera</div>
    <div class="value">{cantiere_nome}</div>
    <div class="label">Indirizzo</div>
    <div class="value">{cantiere_ind}</div>
    <div class="label">Opere</div>
    <div class="value">{opere}</div>
  </div>

  <div class="cover-date">DATA {fmt_date(data_doc)}</div>
</div>

<!-- ═══════════════════════════════════════════════
     INTESTAZIONE RIPETUTA
════════════════════════════════════════════════ -->
<div class="page-header">
  <div class="az">{nome_az}</div>
  <div class="doc-title">PIANO OPERATIVO DI SICUREZZA</div>
  <div class="cantiere-short">Cantiere {cantiere_ind}<br>presso {cantiere_nome}<br>Rev. 00</div>
</div>

<!-- ═══════════════════════════════════════════════
     1. GENERALITÀ
════════════════════════════════════════════════ -->
<div class="section">
<h3>1.0 GENERALITÀ</h3>
<h4>1.1 PRESENTAZIONE DEL PIANO</h4>
<p>Il presente Piano di Sicurezza Allestimento Fieristico è stato redatto ai sensi del D.Lgs 81/08 e rappresenta il documento di pianificazione delle misure di sicurezza della fiera dalla Ditta. Il presente Piano è da considerarsi come Piano di dettaglio e da includersi nel più generale Piano di Sicurezza e Coordinamento. Il Piano contiene quanto richiesto dal D.Lgs 81/08 e precisamente:</p>
<ul>
  <li>I dati identificativi dell'impresa esecutrice, che comprendono: il nominativo del datore di lavoro, gli indirizzi ed i riferimenti telefonici della sede legale e degli uffici di cantiere; i nominativi degli addetti al pronto soccorso, antincendio ed evacuazione dei lavoratori.</li>
  <li>Le specifiche mansioni, inerenti la sicurezza, svolte in cantiere da ogni figura nominata allo scopo dall'impresa esecutrice.</li>
  <li>La descrizione dell'attività di cantiere, delle modalità organizzative e dei turni di lavoro.</li>
  <li>L'elenco dei ponteggi, dei ponti su ruote a torre e di altre opere provvisionali di notevole importanza, delle macchine e degli impianti utilizzati nel cantiere.</li>
  <li>L'elenco delle sostanze e preparati pericolosi utilizzati nel cantiere con le relative schede di sicurezza.</li>
  <li>L'esito del rapporto di valutazione del rumore.</li>
  <li>L'individuazione delle misure preventive e protettive, integrative rispetto a quelle contenute nel Piano di Sicurezza e Coordinamento quando previsto, adottate in relazione ai rischi connessi alle proprie lavorazioni in cantiere.</li>
  <li>Le procedure complementari di dettaglio, richieste dal Piano di Sicurezza e Coordinamento quando previsto.</li>
  <li>L'elenco dei dispositivi di protezione individuale forniti ai lavoratori occupati in cantiere.</li>
  <li>La documentazione in merito all'informazione ed alla formazione fornite ai lavoratori occupati in cantiere.</li>
</ul>
</div>

<!-- ═══════════════════════════════════════════════
     1.3 ANAGRAFICA
════════════════════════════════════════════════ -->
<div class="section chapter">
<div class="page-header">
  <div class="az">{nome_az}</div>
  <div class="doc-title">PIANO OPERATIVO DI SICUREZZA</div>
  <div class="cantiere-short">Cantiere {cantiere_ind}<br>presso {cantiere_nome}<br>Rev. 00</div>
</div>
<h3>1.3 ANAGRAFICA</h3>
<table class="table-2col">
  <tbody>
    <tr><td>Nome e sede legale</td><td><strong>{nome_az}</strong> {sede_legale}</td></tr>
    <tr><td>Attività</td><td>{attivita}</td></tr>
    <tr><td>Posizione INAIL</td><td>Codice ditta n. {inail}</td></tr>
    <tr><td>Posizione INPS</td><td>n. {inps}</td></tr>
    <tr><td>Iscrizione C.C.I.A.A.</td><td>{cciaa}</td></tr>
    <tr><td>Partita IVA / Cod. Fisc.</td><td>{piva}</td></tr>
    <tr><td>Legale Rappresentante</td><td>Sig. {datore}</td></tr>
    <tr><td>Responsabile Servizio Prevenzione e Protezione</td><td>Sig. {rspp or datore}</td></tr>
    <tr><td>Medico Competente</td><td>{medico or 'Da nominare'}</td></tr>
    <tr><td>Direttore Tecnico Cantiere</td><td>Sig. {datore}</td></tr>
    <tr><td>Rappresentante dei lavoratori per la sicurezza</td><td>Sig. {datore}</td></tr>
    <tr><td>Addetti Prevenzione Incendi</td><td>Sig. {datore}</td></tr>
    <tr><td>Addetti Pronto Soccorso</td><td>Sig. {datore}</td></tr>
  </tbody>
</table>

<h3>1.4 DATI IDENTIFICATIVI DI CANTIERE</h3>
<table class="table-2col">
  <tbody>
    <tr><td>Fiera</td><td><strong>{cantiere_nome}</strong><br>{cantiere_ind}</td></tr>
    <tr><td>Committente</td><td>{committente}</td></tr>
    <tr><td>Responsabile dei Lavori</td><td>{datore}</td></tr>
    <tr><td>Coordinatore Sicurezza in Fase di Progettazione</td><td>{datore}</td></tr>
    <tr><td>Coordinatore Sicurezza in Fase di Esecuzione</td><td>{datore}</td></tr>
    <tr><td>Direttore dei Lavori</td><td>{datore}</td></tr>
    <tr><td>Data inizio Cantiere</td><td>{fmt_date(data_inizio)}</td></tr>
    <tr><td>Data fine presunta Cantiere</td><td>{fmt_date(data_fine)}</td></tr>
    <tr><td>Numero uomini/giorni</td><td>{num_lavoratori} operai</td></tr>
  </tbody>
</table>

<h3>1.5 PERSONALE PRESENTE IN CANTIERE</h3>
<table>
  <thead><tr><th>Nominativo</th><th>Mansione</th></tr></thead>
  <tbody>
    <tr><td>{datore.upper()}</td><td>Titolare</td></tr>
    {righe_personale}
  </tbody>
</table>
{"<div class='alert-box'><strong>Note cantiere:</strong> " + note_cantiere + "</div>" if note_cantiere else ""}
</div>

<!-- ═══════════════════════════════════════════════
     2. RELAZIONE TECNICA
════════════════════════════════════════════════ -->
<div class="section chapter">
<div class="page-header">
  <div class="az">{nome_az}</div>
  <div class="doc-title">PIANO OPERATIVO DI SICUREZZA</div>
  <div class="cantiere-short">Cantiere {cantiere_ind}<br>presso {cantiere_nome}<br>Rev. 00</div>
</div>
<h3>2.0 RELAZIONE TECNICA</h3>
<h4>2.1 DESCRIZIONE DEI LAVORI</h4>
<p>Nel Cantiere oggetto del presente Piano di Sicurezza Allestimento Fieristico, {nome_az} eseguirà i seguenti lavori:</p>
<ul><li>{opere}</li></ul>
<p>Durante le lavorazioni il personale dovrà obbligatoriamente indossare i Dispositivi di Protezione Individuale (DPI) quali: scarpe antinfortunistiche, guanti da lavoro, elmetto, dispositivi per la protezione degli occhi ove necessario.</p>

<h4>2.2 PROGRAMMA TEMPORALE DEI LAVORI</h4>
<p>Le lavorazioni avranno presumibilmente inizio il giorno <strong>{fmt_date(data_inizio)}</strong> e fine il giorno <strong>{fmt_date(data_fine)}</strong>. Le lavorazioni verranno eseguite da {nome_az}.</p>
<p>Il piano di esecuzione dei lavori prevede una ripartizione programmata delle attività nel periodo di validità contrattuale, con l'obiettivo di completare le lavorazioni entro i tempi indicati dalla Committente.</p>

<h4>2.3 MANSIONI INERENTI ALLA SICUREZZA</h4>
<p>Il Datore di lavoro deve assolvere ai seguenti obblighi:</p>
<ul>
  <li>Programmare la prevenzione in tutti i suoi aspetti organizzativi ed attuativi.</li>
  <li>Designare dirigenti e preposti in grado di dare attuazione a quanto previsto dalle norme di Legge e da quelle dei POS.</li>
  <li>Designare i lavoratori incaricati di attuare le misure di Pronto Soccorso, Salvataggio, Prevenzione Incendi e gestione dell'Emergenza.</li>
  <li>Designare il Responsabile del Servizio di Prevenzione e Protezione.</li>
  <li>Provvedere alla sorveglianza sanitaria del personale dipendente.</li>
  <li>Disporre affinché venga assicurato il pieno rispetto di tutte le norme di Legge e di buona tecnica in materia di igiene e sicurezza.</li>
  <li>Fornire ai lavoratori attrezzature di lavoro e dispositivi di protezione individuale rispondenti alle norme in materia di tutela della sicurezza e della salute dei lavoratori.</li>
  <li>Garantire ai lavoratori la necessaria informazione sui rischi generali e particolari e sulle procedure da osservare in caso di emergenza e pronto soccorso.</li>
</ul>

<h4>2.4 ORGANIZZAZIONE DI CANTIERE E TURNI DI LAVORO</h4>
<p>{nome_az} non eseguirà, per conto di questa opera, turni di lavoro particolari. Il personale opererà negli orari concordati con la Committente. Deve conservare in cantiere copia del Piano di Sicurezza e Coordinamento, copia del POS e copia della Notifica Preliminare inviata agli Organi di Vigilanza.</p>

<h4>2.5 ATTREZZATURE E SOSTANZE UTILIZZATE</h4>
<table>
  <thead><tr><th>Macchine / Attrezzature</th><th>Presenti in cantiere</th><th>Libretto istruzioni</th><th>Verifiche periodiche</th></tr></thead>
  <tbody>
    <tr><td>Attrezzatura manuale</td><td>&#10003;</td><td>–</td><td>–</td></tr>
    <tr><td>Utensili vari</td><td>&#10003;</td><td>&#10003;</td><td>&#10003;</td></tr>
  </tbody>
</table>
</div>

<!-- ═══════════════════════════════════════════════
     3. VALUTAZIONE DEI RISCHI
════════════════════════════════════════════════ -->
<div class="section chapter">
<div class="page-header">
  <div class="az">{nome_az}</div>
  <div class="doc-title">PIANO OPERATIVO DI SICUREZZA</div>
  <div class="cantiere-short">Cantiere {cantiere_ind}<br>presso {cantiere_nome}<br>Rev. 00</div>
</div>
<h3>3.0 VALUTAZIONE DEI RISCHI</h3>
<h4>3.1 DETTAGLI DELL'ANALISI ESEGUITA</h4>
<p>La presente valutazione dei rischi ha lo scopo di consentire al Datore di Lavoro di identificare tutti i rischi presenti durante le suddette lavorazioni, e di prendere i necessari provvedimenti per la sicurezza e la salute dei lavoratori. I principali rischi individuati sono:</p>
<ul>
  <li>Caduta di persone dall'alto durante l'impiego delle scale a mano.</li>
  <li>Investimento per caduta di materiali dall'alto durante la fase di carico, scarico, movimentazione e sollevamento dei materiali e durante le operazioni di montaggio e smontaggio delle opere provvisionali.</li>
  <li>Strappi muscolari per l'irrazionale maneggio e sollevamento manuale dei materiali.</li>
  <li>Punture per l'eventuale presenza di punte o chiodi sulle vie di transito o per il maneggio di materiali scheggiabili e/o sfaldabili.</li>
  <li>Danno all'apparato uditivo da rumore provocato da macchinari ed utensili utilizzati in cantiere.</li>
  <li>Folgorazione per contatti diretti ed indiretti in conseguenza dell'utilizzo di macchine ed apparecchiature elettriche.</li>
  <li>Danno per contatto o per inalazione di sostanze pericolose o nocive alla salute dei lavoratori.</li>
  <li>Rischi vari per lavorazioni disagevoli dovuti a motivi vari (presenza di acqua, umidità, spazi ristretti, posizioni scomode ecc.).</li>
  <li>Rischi derivanti dalle avverse condizioni meteorologiche (vento, pioggia, gelo, fulmini ecc.).</li>
  <li>Rischi derivanti dalla postura e da fattori microclimatici.</li>
</ul>

<h4>3.2 ESPOSIZIONE AL RUMORE</h4>
<p>Nel cantiere sono presenti attrezzature che producono un livello sonoro tale che un lavoratore risulta esposto ad una esposizione quotidiana personale pari a LEP &gt; 80 dB(A). È obbligatorio che i lavoratori indossino gli otoprotettori se il livello di rumore emesso risulta superiore agli 80 dB(A).</p>

<h4>3.3 DISPOSITIVI DI PROTEZIONE INDIVIDUALE (D.P.I.)</h4>
<p>I lavoratori devono utilizzare sempre ed in modo appropriato i dispositivi di protezione individuale che gli sono messi a disposizione. I D.P.I. forniti sono:</p>
<table>
  <thead><tr><th>Zona da proteggere</th><th>D.P.I.</th></tr></thead>
  <tbody>
    <tr><td>Testa</td><td>Elmetti di protezione contro urti e caduta di materiale</td></tr>
    <tr><td>Occhi</td><td>Occhiali di sicurezza con lenti specifiche contro schegge e spruzzi</td></tr>
    <tr><td>Mani</td><td>Guanti di protezione in tela/cuoio/Kevlar contro tagli, abrasioni, punture</td></tr>
    <tr><td>Vie respiratorie</td><td>Mascherine con opportuno grado di protezione</td></tr>
    <tr><td>Piedi</td><td>Scarponcini/stivali di sicurezza con puntale antischiacciamento e suola antiforo</td></tr>
    <tr><td>Udito</td><td>Cuffie antirumore e/o auricolari monouso</td></tr>
  </tbody>
</table>
</div>

<!-- ═══════════════════════════════════════════════
     FIRME
════════════════════════════════════════════════ -->
<div class="section" style="margin-top:40px">
  <div class="firma-box">
    <div class="firma-item">
      <div class="firma-line">Il Datore di Lavoro<br><strong>{datore}</strong></div>
    </div>
    <div class="firma-item">
      <div class="firma-line">Il Responsabile S.P.P.<br><strong>{rspp or datore}</strong></div>
    </div>
  </div>
  <p style="text-align:center;margin-top:30px;font-size:10px;color:#94a3b8">
    Documento generato il {fmt_date(data_doc)} — {nome_az} — D.Lgs. 81/08
  </p>
</div>

</body>
</html>"""
    return html


# ══════════════════════════════════════════════════════════════
#  SPESE RIMBORSABILI — Mobile + Admin
# ══════════════════════════════════════════════════════════════

CATEGORIE_SPESA = ['Carburante', 'Pranzo / Cena', 'Pedaggi / Autostrada',
                   'Parcheggio', 'Trasporto / Taxi', 'Materiali', 'Altro']

def get_spese_upload_path():
    p = os.path.join(DATA_DIR, 'uploads_spese')
    os.makedirs(p, exist_ok=True)
    return p

MOBILE_SPESE_TMPL = """<!DOCTYPE html>
<html lang="{{ lang }}" dir="{{ t.dir }}">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>{{ t.spese_title }}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{background:#0f172a;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;min-height:100vh}
.header{background:linear-gradient(135deg,#064e3b,#065f46);padding:18px 20px 14px;display:flex;align-items:center;gap:12px}
.header a{color:rgba(255,255,255,.6);text-decoration:none;font-size:20px}
.header h1{font-size:18px;font-weight:800;color:#fff;flex:1}
.content{padding:16px;display:flex;flex-direction:column;gap:14px;max-width:480px;margin:0 auto}
.card{background:#1e293b;border-radius:16px;padding:18px;border:1px solid rgba(255,255,255,.07)}
.card-title{font-size:11px;font-weight:700;color:rgba(255,255,255,.4);text-transform:uppercase;letter-spacing:1px;margin-bottom:14px;display:flex;align-items:center;gap:8px}
label{font-size:13px;color:rgba(255,255,255,.6);margin-bottom:7px;display:block;font-weight:500}
input,select,textarea{width:100%;background:#0f172a;border:1.5px solid rgba(255,255,255,.12);border-radius:12px;color:#fff;font-size:16px;padding:13px 16px;outline:none;-webkit-appearance:none;appearance:none;font-family:inherit;margin-bottom:12px}
input:focus,select:focus,textarea:focus{border-color:#10b981;box-shadow:0 0 0 3px rgba(16,185,129,.15)}
select option{background:#1e293b}
.foto-area{border:2px dashed rgba(16,185,129,.4);border-radius:14px;padding:24px 16px;text-align:center;cursor:pointer;background:rgba(16,185,129,.05);transition:.2s}
.foto-area:active{background:rgba(16,185,129,.12)}
.foto-area i{font-size:36px;color:#10b981;margin-bottom:10px;display:block}
.foto-area p{color:rgba(255,255,255,.5);font-size:13px;line-height:1.5;white-space:pre-line}
.foto-area.has-foto{border-color:#10b981;background:rgba(16,185,129,.1)}
#foto-preview{width:100%;border-radius:10px;margin-top:10px;display:none;max-height:200px;object-fit:cover}
.submit-btn{width:100%;background:linear-gradient(135deg,#065f46,#10b981);color:#fff;border:none;border-radius:14px;padding:18px;font-size:17px;font-weight:700;cursor:pointer;display:flex;align-items:center;justify-content:center;gap:10px}
.submit-btn:disabled{background:#334155;color:rgba(255,255,255,.4)}
.spesa-item{background:#0f172a;border-radius:12px;padding:13px 14px;margin-bottom:8px}
.spesa-top{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:4px}
.spesa-cat{font-size:12px;font-weight:700;color:#10b981}
.spesa-importo{font-size:18px;font-weight:800;color:#fff}
.spesa-data{font-size:12px;color:rgba(255,255,255,.4)}
.spesa-desc{font-size:12px;color:rgba(255,255,255,.5);margin-top:3px}
.stato-badge{display:inline-block;padding:2px 10px;border-radius:99px;font-size:11px;font-weight:700}
.stato-in_attesa{background:rgba(251,191,36,.15);color:#fbbf24}
.stato-approvata{background:rgba(34,197,94,.15);color:#22c55e}
.stato-rifiutata{background:rgba(239,68,68,.15);color:#ef4444}
.flash{padding:13px 15px;border-radius:12px;font-size:13px;font-weight:600;display:flex;align-items:center;gap:10px}
.flash.success{background:rgba(34,197,94,.15);border:1px solid rgba(34,197,94,.3);color:#86efac}
.flash.error{background:rgba(239,68,68,.15);border:1px solid rgba(239,68,68,.3);color:#fca5a5}
.req-badge{color:#ef4444;font-size:11px;margin-left:4px}
</style>
</head>
<body>
<div class="header">
  <a href="/mobile"><i class="fa fa-arrow-left"></i></a>
  <h1><i class="fa fa-receipt"></i> {{ t.spese_title }}</h1>
</div>
<div class="content">
  {% with messages = get_flashed_messages(with_categories=true) %}
  {% for cat, msg in messages %}
  <div class="flash {{ cat }}"><i class="fa fa-{{ 'check-circle' if cat=='success' else 'exclamation-circle' }}"></i> {{ msg }}</div>
  {% endfor %}{% endwith %}

  <!-- Form nuova spesa -->
  <form method="POST" action="/mobile/spese/inserisci" enctype="multipart/form-data" id="form-spesa">
    <div class="card">
      <div class="card-title"><i class="fa fa-plus-circle" style="color:#10b981"></i> {{ t.spese_new }}</div>

      <label>{{ t.spese_date }} *</label>
      <input type="date" name="data" value="{{ oggi }}" required>

      <label>{{ t.spese_cat }} *</label>
      <select name="categoria" required>
        {% for c in categorie %}<option>{{ c }}</option>{% endfor %}
      </select>

      <label>{{ t.spese_amount }} *</label>
      <input type="number" name="importo" step="0.01" min="0.01" placeholder="0.00" required inputmode="decimal">

      <label>{{ t.spese_desc }}</label>
      <textarea name="descrizione" rows="2" placeholder="..."></textarea>

      {% if veicoli %}
      <label>{{ t.spese_vehicle }}</label>
      <select name="veicolo_id">
        <option value="">{{ t.spese_no_vehicle }}</option>
        {% for v in veicoli %}<option value="{{ v.id }}">{{ v.targa }} – {{ v.marca or '' }} {{ v.modello or '' }}</option>{% endfor %}
      </select>
      {% endif %}

      <!-- Foto OBBLIGATORIA -->
      <label>{{ t.spese_photo_label }} <span class="req-badge">* {{ t.spese_photo_req }}</span></label>
      <input type="file" name="foto" id="foto-input" accept="image/*,application/pdf" capture="environment" required style="display:none" onchange="anteprimaFoto(this)">
      <div class="foto-area" id="foto-area" onclick="document.getElementById('foto-input').click()">
        <i class="fa fa-camera" id="foto-icon"></i>
        <p id="foto-testo">{{ t.spese_photo_hint }}</p>
        <img id="foto-preview" alt="">
      </div>
    </div>

    <button type="submit" class="submit-btn" id="submit-btn" disabled>
      <i class="fa fa-paper-plane"></i> {{ t.spese_submit }}
    </button>
  </form>

  <!-- Storico spese -->
  {% if spese %}
  <div class="card">
    <div class="card-title"><i class="fa fa-clock-rotate-left" style="color:#94a3b8"></i> {{ t.spese_history }}</div>
    {% for s in spese %}
    <div class="spesa-item">
      <div class="spesa-top">
        <div>
          <div class="spesa-cat">{{ s.categoria }}</div>
          <div class="spesa-data">{{ s.data }}</div>
        </div>
        <div style="text-align:right">
          <div class="spesa-importo">€ {{ "%.2f"|format(s.importo) }}</div>
          <span class="stato-badge stato-{{ s.stato }}">
            {% if s.stato=='approvata' %}{{ t.spese_approved }}{% elif s.stato=='rifiutata' %}{{ t.spese_rejected }}{% else %}{{ t.spese_waiting }}{% endif %}
          </span>
        </div>
      </div>
      {% if s.descrizione %}<div class="spesa-desc">{{ s.descrizione }}</div>{% endif %}
      {% if s.note_admin %}<div style="font-size:11px;color:#f59e0b;margin-top:4px"><i class="fa fa-comment"></i> {{ t.spese_note_admin }} {{ s.note_admin }}</div>{% endif %}
    </div>
    {% endfor %}
  </div>
  {% endif %}
</div>

<script>
const SENDING_TXT = '{{ t.spese_sending }}';
function anteprimaFoto(input) {
  const area = document.getElementById('foto-area');
  const preview = document.getElementById('foto-preview');
  const icon = document.getElementById('foto-icon');
  const testo = document.getElementById('foto-testo');
  const btn = document.getElementById('submit-btn');
  if (input.files && input.files[0]) {
    const f = input.files[0];
    area.classList.add('has-foto');
    testo.textContent = f.name;
    icon.className = 'fa fa-check-circle';
    icon.style.color = '#10b981';
    btn.disabled = false;
    if (f.type.startsWith('image/')) {
      const reader = new FileReader();
      reader.onload = e => { preview.src = e.target.result; preview.style.display = 'block'; };
      reader.readAsDataURL(f);
    } else {
      preview.style.display = 'none';
    }
  }
}
document.getElementById('form-spesa').addEventListener('submit', function() {
  document.getElementById('submit-btn').disabled = true;
  document.getElementById('submit-btn').innerHTML = '<i class="fa fa-spinner fa-spin"></i> ' + SENDING_TXT;
});
</script>
</body></html>
"""


@app.route('/mobile/spese')
@login_required
def mobile_spese():
    uid = session['user_id']
    db = get_db()
    spese = db.execute(
        "SELECT * FROM spese_rimborso WHERE utente_id=? ORDER BY data DESC, creato_il DESC LIMIT 20",
        (uid,)).fetchall()
    veicoli = db.execute("SELECT id,targa,marca,modello FROM veicoli ORDER BY targa").fetchall()
    db.close()
    lang = session.get('lang', 'it')
    t = get_lang()
    return render_template_string(MOBILE_SPESE_TMPL,
        oggi=date.today().isoformat(),
        categorie=CATEGORIE_SPESA,
        spese=[dict(s) for s in spese],
        veicoli=veicoli,
        t=t, lang=lang, langs=LANGS, current_lang=lang,
        azienda=get_setting('nome_azienda', 'Azienda'))


@app.route('/mobile/spese/inserisci', methods=['POST'])
@login_required
def mobile_spese_inserisci():
    uid = session['user_id']
    f = request.form
    foto = request.files.get('foto')

    if not foto or not foto.filename:
        t = get_lang()
        flash(t.get('spese_error_photo', 'La foto dello scontrino è obbligatoria.'), 'error')
        return redirect(url_for('mobile_spese'))

    # Salva foto
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    ext = os.path.splitext(secure_filename(foto.filename))[1] or '.jpg'
    fn = f"{uid}_{ts}{ext}"
    fp = os.path.join(get_spese_upload_path(), fn)
    foto.save(fp)

    db = get_db()
    db.execute("""INSERT INTO spese_rimborso
        (utente_id, data, categoria, descrizione, importo, veicolo_id, foto_nome, foto_path, stato)
        VALUES (?,?,?,?,?,?,?,?,?)""", (
        uid,
        f.get('data') or date.today().isoformat(),
        f.get('categoria', 'Altro'),
        f.get('descrizione', ''),
        float(f.get('importo') or 0),
        f.get('veicolo_id') or None,
        fn, fp, 'in_attesa'))
    safe_commit(db); db.close()

    # Notifica email admin
    email_admin = get_setting('email_notifiche', '')
    nome_dip = f"{session.get('nome','')} {session.get('cognome','')}".strip()
    cat = f.get('categoria','Altro')
    importo = float(f.get('importo') or 0)
    data_spesa = f.get('data') or date.today().isoformat()
    if email_admin:
        import threading
        _base = request.host_url.rstrip('/')
        # Salva app_url per uso futuro nei thread senza contesto request
        try:
            _db2 = get_db()
            _db2.execute("INSERT OR REPLACE INTO impostazioni (chiave,valore) VALUES ('app_url',?)", (_base,))
            _db2.commit(); _db2.close()
        except: pass
        def _send_spesa():
            try:
                send_email(email_admin,
                    f'[ACCESSO FIERE] Nuova richiesta rimborso da {nome_dip}',
                    f'<p><b>{nome_dip}</b> ha inviato una richiesta di rimborso spese:</p>'
                    f'<table style="border-collapse:collapse;font-family:sans-serif">'
                    f'<tr><td style="padding:6px 12px;color:#64748b">Data</td><td style="padding:6px 12px"><b>{data_spesa}</b></td></tr>'
                    f'<tr><td style="padding:6px 12px;color:#64748b">Categoria</td><td style="padding:6px 12px"><b>{cat}</b></td></tr>'
                    f'<tr><td style="padding:6px 12px;color:#64748b">Importo</td><td style="padding:6px 12px;font-size:18px;color:#059669"><b>€ {importo:.2f}</b></td></tr>'
                    f'<tr><td style="padding:6px 12px;color:#64748b">Descrizione</td><td style="padding:6px 12px">{f.get("descrizione","–")}</td></tr>'
                    f'</table>'
                    f'<p style="margin-top:16px"><a href="{_base}/admin/spese?stato=in_attesa" '
                    f'style="background:#059669;color:#fff;padding:10px 20px;border-radius:8px;text-decoration:none;font-weight:700">'
                    f'Approva o rifiuta →</a></p>')
            except: pass
        threading.Thread(target=_send_spesa, daemon=True).start()

    t = get_lang()
    flash(t.get('spese_success', 'Spesa inviata!'), 'success')
    return redirect(url_for('mobile_spese'))


@app.route('/mobile/spese/foto/<fn>')
@login_required
def mobile_spese_foto(fn):
    """Serve foto scontrino — solo se appartiene all'utente o admin."""
    db = get_db()
    spesa = db.execute("SELECT * FROM spese_rimborso WHERE foto_nome=?", (fn,)).fetchone()
    db.close()
    if not spesa:
        return 'Not found', 404
    if session.get('ruolo') != 'admin' and spesa['utente_id'] != session.get('user_id'):
        return 'Forbidden', 403
    fp = os.path.join(get_spese_upload_path(), fn)
    if not os.path.exists(fp):
        return 'Not found', 404
    mt = mimetypes.guess_type(fn)[0] or 'image/jpeg'
    return send_file(fp, mimetype=mt)


@app.route('/pos')

@admin_required
def pos_list():
    db = get_db()
    pos_list = db.execute(
        "SELECT * FROM pos_documenti ORDER BY creato_il DESC").fetchall()
    db.close()
    return render_page(POS_LIST_TMPL, page_title='POS - Piani Operativi di Sicurezza',
                       active='pos', pos_list=[dict(p) for p in pos_list])


@app.route('/pos/nuovo', methods=['GET', 'POST'])
@admin_required
def pos_nuovo():
    if request.method == 'POST':
        return _salva_pos(None)
    db = get_db()
    dipendenti = db.execute(
        "SELECT id, nome, cognome, mansione, titolo FROM utenti WHERE attivo=1 AND ruolo='dipendente' ORDER BY cognome").fetchall()
    db.close()
    az_keys = ['azienda', 'sede_legale', 'partita_iva', 'codice_inail', 'codice_inps',
               'cciaa', 'rspp', 'medico_competente', 'attivita_azienda']
    datore = get_setting('rspp', '') or get_setting('azienda', '')
    return render_page(POS_FORM_TMPL, page_title='Nuovo POS', active='pos',
                       form_action='/pos/nuovo', pos={},
                       dipendenti=[dict(d) for d in dipendenti],
                       datore_lavoro=datore, today=date.today().isoformat())


@app.route('/pos/<int:pid>/modifica', methods=['GET', 'POST'])
@admin_required
def pos_modifica(pid):
    if request.method == 'POST':
        return _salva_pos(pid)
    db = get_db()
    p = db.execute("SELECT * FROM pos_documenti WHERE id=?", (pid,)).fetchone()
    if not p:
        return redirect(url_for('pos_list'))
    dipendenti = db.execute(
        "SELECT id, nome, cognome, mansione, titolo FROM utenti WHERE attivo=1 AND ruolo='dipendente' ORDER BY cognome").fetchall()
    db.close()
    pd = dict(p)
    # Ricostruisci lista ID selezionati
    try:
        import json as _json
        personale = _json.loads(pd.get('personale_json') or '[]')
        pd['personale_selezionati'] = [str(x.get('id', x)) if isinstance(x, dict) else str(x) for x in personale]
    except Exception:
        pd['personale_selezionati'] = []
    datore = get_setting('rspp', '') or get_setting('azienda', '')
    return render_page(POS_FORM_TMPL, page_title='Modifica POS', active='pos',
                       form_action=f'/pos/{pid}/modifica', pos=pd,
                       dipendenti=[dict(d) for d in dipendenti],
                       datore_lavoro=datore, today=date.today().isoformat())


def _salva_pos(pid):
    """Salva o aggiorna un POS, opzionalmente genera PDF."""
    personale_ids = request.form.getlist('personale')
    db = get_db()
    # Recupera dati personale selezionato
    personale = []
    if personale_ids:
        placeholders = ','.join('?' * len(personale_ids))
        dipendenti_sel = db.execute(
            f"SELECT id, nome, cognome, mansione, titolo FROM utenti WHERE id IN ({placeholders})",
            personale_ids).fetchall()
        personale = [dict(d) for d in dipendenti_sel]
    personale_json = json.dumps(personale, ensure_ascii=False)

    titolo    = request.form.get('titolo', '').strip()
    data_doc  = request.form.get('data_documento') or date.today().isoformat()

    fields = {
        'titolo':             titolo,
        'cantiere_nome':      request.form.get('cantiere_nome', '').strip(),
        'cantiere_indirizzo': request.form.get('cantiere_indirizzo', '').strip(),
        'committente':        request.form.get('committente', '').strip(),
        'responsabile_lavori':request.form.get('responsabile_lavori', '').strip(),
        'data_inizio':        request.form.get('data_inizio') or None,
        'data_fine':          request.form.get('data_fine') or None,
        'data_documento':     data_doc,
        'opere':              request.form.get('opere', '').strip(),
        'note_cantiere':      request.form.get('note_cantiere', '').strip(),
        'num_lavoratori':     len(personale),
        'personale_json':     personale_json,
        'creato_da':          session.get('user_id'),
    }

    if pid:
        db.execute("""UPDATE pos_documenti SET titolo=?,cantiere_nome=?,cantiere_indirizzo=?,
            committente=?,responsabile_lavori=?,data_inizio=?,data_fine=?,data_documento=?,
            opere=?,note_cantiere=?,num_lavoratori=?,personale_json=? WHERE id=?""",
            (fields['titolo'], fields['cantiere_nome'], fields['cantiere_indirizzo'],
             fields['committente'], fields['responsabile_lavori'], fields['data_inizio'],
             fields['data_fine'], fields['data_documento'], fields['opere'],
             fields['note_cantiere'], fields['num_lavoratori'], fields['personale_json'], pid))
        saved_id = pid
    else:
        cur = db.execute("""INSERT INTO pos_documenti
            (titolo,cantiere_nome,cantiere_indirizzo,committente,responsabile_lavori,
             data_inizio,data_fine,data_documento,opere,note_cantiere,num_lavoratori,
             personale_json,creato_da)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (fields['titolo'], fields['cantiere_nome'], fields['cantiere_indirizzo'],
             fields['committente'], fields['responsabile_lavori'], fields['data_inizio'],
             fields['data_fine'], fields['data_documento'], fields['opere'],
             fields['note_cantiere'], fields['num_lavoratori'], fields['personale_json'],
             fields['creato_da']))
        saved_id = cur.lastrowid
    safe_commit(db)
    db.close()
    flash('POS salvato!', 'success')

    if request.form.get('genera_pdf'):
        return redirect(url_for('pos_pdf', pid=saved_id))
    return redirect(url_for('pos_list'))


@app.route('/pos/<int:pid>/pdf')
@admin_required
def pos_pdf(pid):
    db = get_db()
    p = db.execute("SELECT * FROM pos_documenti WHERE id=?", (pid,)).fetchone()
    if not p:
        return redirect(url_for('pos_list'))
    db.close()

    pd = dict(p)
    try:
        personale = json.loads(pd.get('personale_json') or '[]')
    except Exception:
        personale = []

    az_keys = ['azienda', 'sede_legale', 'partita_iva', 'codice_inail', 'codice_inps',
               'cciaa', 'rspp', 'medico_competente', 'attivita_azienda']
    az_data = {k: get_setting(k, '') for k in az_keys}

    html = _pos_pdf_html(pd, az_data, personale)

    # Prova WeasyPrint
    try:
        from weasyprint import HTML as WH
        pdf_bytes = WH(string=html).write_pdf()
        fname = f"POS_{pd['titolo'].replace(' ', '_')}.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{fname}"'})
    except Exception:
        # Fallback HTML stampabile
        html_print = html.replace('</body>',
            '<script>window.onload=function(){window.print()}</script></body>')
        return Response(html_print, mimetype='text/html')


@app.route('/pos/<int:pid>/elimina')
@admin_required
def pos_elimina(pid):
    db = get_db()
    db.execute("DELETE FROM pos_documenti WHERE id=?", (pid,))
    safe_commit(db)
    db.close()
    flash('POS eliminato.', 'success')
    return redirect(url_for('pos_list'))





@app.route('/admin/test-email')
@admin_required
def test_email():
    """Route di test per verificare la configurazione SMTP."""
    to = get_setting('email_notifiche', '')
    user = get_setting('smtp_user', '')
    if not to:
        to = user  # manda a se stesso se non c'e' email notifiche
    if not to:
        flash('Configura smtp_user nelle Impostazioni prima di testare.', 'error')
        return redirect(url_for('impostazioni'))
    result = send_email(
        to,
        'Test SMTP — Gestionale',
        '<p>Se ricevi questa mail, la configurazione SMTP funziona correttamente! ✅</p>'
    )
    if result:
        flash(f'Email di test inviata a {to}!', 'success')
    else:
        flash('Invio fallito — controlla i log del server per i dettagli.', 'error')
    return redirect(url_for('impostazioni'))


# ══════════════════════════════════════════════════════════
#  IMPOSTAZIONI
# ══════════════════════════════════════════════════════════

IMP_TMPL = """
<div class="grid-2">
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-building"></i> Azienda</h3></div>
    <div class="card-body">
      <form method="POST" action="/admin/impostazioni/azienda">
        <div class="form-group"><label>Nome azienda</label><input name="azienda" value="{{ cfg.azienda }}"></div>
        <div class="form-group"><label>Sede legale</label><input name="sede_legale" value="{{ cfg.sede_legale }}"></div>
        <div class="form-group"><label>Partita IVA</label><input name="partita_iva" value="{{ cfg.partita_iva }}"></div>
        <div class="form-group"><label>Attività aziendale</label><input name="attivita_azienda" value="{{ cfg.attivita_azienda }}"></div>
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva</button>
      </form>
    </div>
  </div>
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-key"></i> Cambia password admin</h3></div>
    <div class="card-body">
      <form method="POST" action="/admin/impostazioni/password">
        <div class="form-group"><label>Nuova password</label><input type="password" name="new_password" required></div>
        <div class="form-group"><label>Conferma password</label><input type="password" name="confirm_password" required></div>
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Aggiorna</button>
      </form>
    </div>
  </div>
</div>
<div class="card" style="margin-top:20px">
  <div class="card-header"><h3><i class="fa fa-envelope"></i> Email notifiche</h3></div>
  <div class="card-body">
    <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:12px 16px;margin-bottom:16px;font-size:13px;color:#1e40af">
      <strong>⚡ Railway blocca SMTP — usa Brevo (gratuito, 300 mail/giorno):</strong><br>
      1. Registrati su <strong>brevo.com</strong> → <em>SMTP &amp; API → API Keys</em> → crea e copia la chiave<br>
      2. Nel campo <em>Metodo</em> scrivi esattamente: <strong>brevo</strong><br>
      3. In <em>Email mittente</em> metti la tua email verificata su Brevo<br>
      4. In <em>API Key</em> incolla la chiave copiata → Salva → Test
    </div>
    <form method="POST" action="/admin/impostazioni/email">
      <div class="form-row">
        <div class="form-group">
          <label>Metodo (<code>brevo</code> per Railway)</label>
          <input name="smtp_host" value="{{ cfg.smtp_host }}" placeholder="brevo">
        </div>
        <div class="form-group">
          <label>Email mittente (verificata su Brevo)</label>
          <input name="smtp_user" value="{{ cfg.smtp_user }}" placeholder="info@tuaazienda.it">
        </div>
      </div>
      <div class="form-group">
        <label>API Key Brevo</label>
        <input type="password" name="smtp_pass" placeholder="invariata se vuoto" autocomplete="new-password">
      </div>
      <div class="form-group" style="background:#fefce8;border:1px solid #fde047;border-radius:8px;padding:14px">
        <label style="color:#92400e;font-weight:700"><i class="fa fa-bell"></i> Email destinatario notifiche <span style="font-weight:400;font-size:12px">(dove ricevi le notifiche)</span></label>
        <input name="email_notifiche" value="{{ cfg.email_notifiche }}" placeholder="admin@tuaazienda.it" style="margin-top:6px">
        <div style="font-size:12px;color:#92400e;margin-top:6px">⚠️ Senza questo campo non arrivano notifiche di timbrature, rimborsi spese e scadenze.</div>
      </div>
      <div class="form-group">
        <label>URL pubblico del sito <span style="font-weight:400;font-size:12px;color:var(--text-light)">(usato nei link delle email, es: https://web-production-xxxx.up.railway.app)</span></label>
        <input name="app_url" value="{{ cfg.app_url }}" placeholder="https://web-production-xxxx.up.railway.app">
      </div>
      <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva</button>
    </form>
    <div style="margin-top:14px;padding-top:14px;border-top:1px solid var(--border)">
      <a href="/admin/test-email" class="btn btn-secondary btn-sm"><i class="fa fa-paper-plane"></i> Invia email di test</a>
      <span style="font-size:12px;color:var(--text-muted);margin-left:10px">Verifica la configurazione</span>
    </div>
  </div>
</div>
<div class="card" style="margin-top:20px">
  <div class="card-header"><h3><i class="fa fa-robot"></i> API Anthropic AI</h3></div>
  <div class="card-body">
    <form method="POST" action="/admin/impostazioni/ai">
      <div class="form-group"><label>API Key Anthropic</label><input name="anthropic_api_key" value="{{ cfg.anthropic_api_key }}" placeholder="sk-ant-..."></div>
      <div style="margin-top:10px">
        {% if ai_ok %}<span class="badge badge-green">AI attiva</span>
        {% else %}<span class="badge badge-red">AI non configurata</span>{% endif %}
      </div>
      <button type="submit" class="btn btn-primary" style="margin-top:12px"><i class="fa fa-save"></i> Salva API Key</button>
    </form>
  </div>
</div>

<!-- Google OAuth -->
<div class="card" style="margin-top:20px">
  <div class="card-header">
    <h3><svg width="18" height="18" viewBox="0 0 48 48" style="vertical-align:middle;margin-right:8px"><path fill="#4285F4" d="M45.12 24.5c0-1.56-.14-3.06-.4-4.5H24v8.51h11.84c-.51 2.75-2.06 5.08-4.39 6.64v5.52h7.11c4.16-3.83 6.56-9.47 6.56-16.17z"/><path fill="#34A853" d="M24 46c5.94 0 10.92-1.97 14.56-5.33l-7.11-5.52c-1.97 1.32-4.49 2.1-7.45 2.1-5.73 0-10.58-3.87-12.31-9.07H4.34v5.7C7.96 41.07 15.4 46 24 46z"/><path fill="#FBBC05" d="M11.69 28.18C11.25 26.86 11 25.45 11 24s.25-2.86.69-4.18v-5.7H4.34C2.85 17.09 2 20.45 2 24c0 3.55.85 6.91 2.34 9.88l7.35-5.7z"/><path fill="#EA4335" d="M24 10.75c3.23 0 6.13 1.11 8.41 3.29l6.31-6.31C34.91 4.18 29.93 2 24 2 15.4 2 7.96 6.93 4.34 14.12l7.35 5.7z"/></svg>Login con Google (OAuth)</h3>
    <small style="color:var(--text-light)">Permette ai dipendenti di registrarsi e accedere con il loro account Gmail</small>
  </div>
  <div class="card-body">
    <form method="POST" action="/admin/impostazioni/google">
      <div class="form-group"><label>Google Client ID</label><input name="google_client_id" value="{{ cfg.google_client_id }}" placeholder="xxxxxxxxxx-xxx.apps.googleusercontent.com"></div>
      <div class="form-group"><label>Google Client Secret</label><input name="google_client_secret" value="{{ cfg.google_client_secret }}" placeholder="GOCSPX-..."></div>
      <div style="background:#f8fafc;border-radius:8px;padding:12px;margin-bottom:12px;font-size:12px;color:var(--text-light);line-height:1.7">
        <strong>Come ottenere le credenziali:</strong><br>
        1. Vai su <a href="https://console.cloud.google.com" target="_blank" style="color:var(--accent2)">console.cloud.google.com</a><br>
        2. Crea progetto → <em>API & Servizi → Credenziali → Crea credenziali → ID client OAuth 2.0</em><br>
        3. Tipo applicazione: <strong>Applicazione web</strong><br>
        4. URI di reindirizzamento autorizzato: <code style="background:#e2e8f0;padding:2px 6px;border-radius:4px">{{ request.host_url }}auth/google/callback</code><br>
        5. Copia Client ID e Client Secret qui sopra
      </div>
      <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva credenziali Google</button>
    </form>
  </div>
</div>
"""

@app.route('/admin/impostazioni')
@admin_required
def impostazioni():
    db = get_db()
    keys = ['azienda','sede_legale','partita_iva','attivita_azienda','email_notifiche',
            'smtp_host','smtp_port','smtp_user','smtp_pass','anthropic_api_key',
            'google_client_id','google_client_secret','app_url']
    cfg = {k: get_setting(k,'') for k in keys}
    db.close()
    return render_page(IMP_TMPL, page_title='Impostazioni', active='impostazioni', cfg=cfg, ai_ok=AI_OK)


@app.route('/admin/impostazioni/azienda', methods=['POST'])
@admin_required
def impostazioni_azienda():
    db = get_db()
    for k in ['azienda','sede_legale','partita_iva','attivita_azienda']:
        db.execute("INSERT OR REPLACE INTO impostazioni (chiave,valore) VALUES (?,?)",
                   (k, request.form.get(k,'')))
    safe_commit(db); db.close()
    flash('Dati azienda salvati!','success')
    return redirect(url_for('impostazioni'))


@app.route('/admin/impostazioni/password', methods=['POST'])
@admin_required
def impostazioni_password():
    np = request.form.get('new_password','')
    cp = request.form.get('confirm_password','')
    if np != cp:
        flash('Le password non corrispondono.','error')
    elif len(np) < 6:
        flash('Password troppo corta (min. 6 caratteri).','error')
    else:
        db = get_db()
        h  = hashlib.sha256(np.encode()).hexdigest()
        db.execute("UPDATE utenti SET password=? WHERE ruolo='admin'", (h,))
        safe_commit(db); db.close()
        flash('Password aggiornata!','success')
    return redirect(url_for('impostazioni'))


@app.route('/admin/impostazioni/email', methods=['POST'])
@admin_required
def impostazioni_email():
    db = get_db()
    for k in ['email_notifiche','smtp_host','smtp_port','smtp_user','app_url']:
        db.execute("INSERT OR REPLACE INTO impostazioni (chiave,valore) VALUES (?,?)",
                   (k, request.form.get(k,'')))
    pw = request.form.get('smtp_pass','')
    if pw:
        db.execute("INSERT OR REPLACE INTO impostazioni (chiave,valore) VALUES (?,?)",('smtp_pass', pw))
    safe_commit(db); db.close()
    flash('Impostazioni email salvate!','success')
    return redirect(url_for('impostazioni'))


@app.route('/admin/impostazioni/ai', methods=['POST'])
@admin_required
def impostazioni_ai():
    key = request.form.get('anthropic_api_key','').strip()
    db  = get_db()
    db.execute("INSERT OR REPLACE INTO impostazioni (chiave,valore) VALUES (?,?)",('anthropic_api_key', key))
    safe_commit(db); db.close()
    flash('API Key salvata!','success')
    return redirect(url_for('impostazioni'))


@app.route('/admin/impostazioni/google', methods=['POST'])
@admin_required
def impostazioni_google():
    db = get_db()
    for k in ['google_client_id','google_client_secret']:
        db.execute("INSERT OR REPLACE INTO impostazioni (chiave,valore) VALUES (?,?)",
                   (k, request.form.get(k,'').strip()))
    safe_commit(db); db.close()
    flash('Credenziali Google salvate!','success')
    return redirect(url_for('impostazioni'))


# ══════════════════════════════════════════════════════════
#  DIPENDENTI — DOCUMENTI
# ══════════════════════════════════════════════════════════

DOC_DIP_TMPL = """
<div style="margin-bottom:16px">
  <a href="/dipendenti" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Dipendenti</a>
</div>
<div class="card">
  <div class="card-header">
    <h3><i class="fa fa-file-alt" style="color:var(--accent2)"></i> Documenti — {{ dip.nome }} {{ dip.cognome }}</h3>
  </div>
  <div class="card-body">
    <form method="POST" action="/dipendenti/{{ dip.id }}/documenti/upload" enctype="multipart/form-data" style="background:#f8fafc;padding:16px;border-radius:10px;margin-bottom:20px;border:1px solid var(--border)">
      <div class="form-row">
        <div class="form-group">
          <label>File *</label>
          <input type="file" name="file" required accept=".pdf,.png,.jpg,.jpeg,.gif,.webp">
        </div>
        <div class="form-group">
          <label>Tipo documento</label>
          <select name="tipo_doc">
            <option>Carta identità</option><option>Passaporto</option><option>Patente</option>
            <option>Contratto</option><option>UNILAV</option><option>Corso/Attestato</option>
            <option>Altro</option><option value="Auto">🤖 Analisi AI automatica</option>
          </select>
        </div>
      </div>
      <div class="form-row">
        <div class="form-group"><label>Note</label><input name="note" placeholder="Opzionale"></div>
        <div class="form-group"><label>Scadenza</label><input type="date" name="data_scadenza"></div>
      </div>
      <div style="display:flex;align-items:center;gap:14px">
        <label style="display:flex;align-items:center;gap:6px;font-size:13px;cursor:pointer">
          <input type="checkbox" name="analizza" value="1"> Analizza con AI
        </label>
        <button type="submit" class="btn btn-primary"><i class="fa fa-upload"></i> Carica</button>
      </div>
    </form>
    {% if docs %}
    <table class="table">
      <thead><tr><th>Tipo</th><th>File</th><th>Note</th><th>Scadenza</th><th>Data</th><th>Azioni</th></tr></thead>
      <tbody>
      {% for d in docs %}
      <tr>
        <td><span class="badge badge-purple">{{ d.tipo_doc }}</span></td>
        <td>{{ d.nome_originale }}</td>
        <td>{{ d.note or '—' }}</td>
        <td>{{ d.data_scadenza or '—' }}</td>
        <td>{{ d.caricato_il[:10] }}</td>
        <td style="display:flex;gap:4px">
          <a href="/dipendenti/{{ dip.id }}/documenti/{{ d.id }}/anteprima" class="btn btn-sm btn-secondary" target="_blank"><i class="fa fa-eye"></i></a>
          <a href="/dipendenti/{{ dip.id }}/documenti/{{ d.id }}/scarica" class="btn btn-sm btn-blue"><i class="fa fa-download"></i></a>
          <a href="/dipendenti/{{ dip.id }}/documenti/{{ d.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare?')"><i class="fa fa-trash"></i></a>
        </td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    {% if ai_result %}
    <div class="alert alert-success" style="margin-top:16px">
      <strong>AI ha rilevato:</strong> {{ ai_result.tipo_doc }} — {{ ai_result.testo }}
    </div>
    {% endif %}
    {% else %}
    <div class="empty-state"><i class="fa fa-file"></i><p>Nessun documento caricato.</p></div>
    {% endif %}
  </div>
</div>
"""

@app.route('/dipendenti/<int:uid>/documenti', methods=['GET'])
@admin_required
def dip_documenti(uid):
    db  = get_db()
    dip = db.execute("SELECT * FROM utenti WHERE id=?", (uid,)).fetchone()
    if not dip:
        db.close(); flash("Dipendente non trovato.","error"); return redirect(url_for("dipendenti"))
    docs = db.execute("SELECT * FROM documenti_dipendente WHERE utente_id=? ORDER BY caricato_il DESC", (uid,)).fetchall()
    db.close()
    ai_result = session.pop('ai_result_dip', None)
    return render_page(DOC_DIP_TMPL, page_title=f'Documenti – {dip["nome"]}',
                       active='dipendenti', dip=dip, docs=docs, ai_result=ai_result, AI_OK=AI_OK)


@app.route('/dipendenti/<int:uid>/documenti/upload', methods=['POST'])
@admin_required
def dip_doc_upload(uid):
    f = request.files.get('file')
    if not f or not allowed_file(f.filename):
        flash('File non valido.','error'); return redirect(url_for('dip_documenti', uid=uid))
    nome_file = datetime.now().strftime('%Y%m%d_%H%M%S_') + secure_filename(f.filename)
    path = os.path.join(get_upload_path(uid), nome_file)
    f.save(path)
    dim      = os.path.getsize(path)
    tipo_doc = request.form.get('tipo_doc','Altro')
    note     = request.form.get('note','')
    data_sc  = request.form.get('data_scadenza','') or None
    analizza = request.form.get('analizza')
    db = get_db()
    doc_id = db.execute(
        "INSERT INTO documenti_dipendente (utente_id,nome_originale,nome_file,tipo_doc,dimensione,note,data_scadenza) VALUES (?,?,?,?,?,?,?)",
        (uid, f.filename, nome_file, tipo_doc, dim, note, data_sc)).lastrowid
    safe_commit(db); db.close()
    flash(f'Documento "{f.filename}" caricato!','success')
    if analizza and AI_OK:
        ai_result = analizza_documento_ai(path, f.filename, tipo_doc, uid)
        updates = {}
        if ai_result.get('tipo_doc') and ai_result['tipo_doc'] not in ('Altro', ''):
            updates['tipo_doc'] = ai_result['tipo_doc'][:50]
        if ai_result.get('data_scadenza') and not data_sc:
            updates['data_scadenza'] = ai_result['data_scadenza']
        if ai_result.get('note') and not note:
            updates['note'] = ai_result['note'][:200]
        if updates:
            db2 = get_db()
            set_clause = ', '.join(f"{k}=?" for k in updates)
            db2.execute(f"UPDATE documenti_dipendente SET {set_clause} WHERE id=?",
                        list(updates.values()) + [doc_id])
            db2.commit(); db2.close()
            msg = f'AI: {ai_result.get("tipo_doc","")}'
            if ai_result.get('data_scadenza'): msg += f' · scadenza {ai_result["data_scadenza"]}'
            flash(msg, 'success')
        elif '_error' not in ai_result:
            flash('AI analizzata, nessun dato aggiuntivo rilevato.', 'info')
        session['ai_result_dip'] = ai_result
    return redirect(url_for('dip_documenti', uid=uid))


@app.route('/dipendenti/<int:uid>/documenti/<int:did>/scarica')
@admin_required
def dip_doc_scarica(uid, did):
    db  = get_db()
    doc = db.execute("SELECT * FROM documenti_dipendente WHERE id=? AND utente_id=?", (did,uid)).fetchone()
    db.close()
    if not doc: flash('Non trovato.','error'); return redirect(url_for('dip_documenti',uid=uid))
    path = os.path.join(get_upload_path(uid), doc['nome_file'])
    return send_file(path, as_attachment=True, download_name=doc['nome_originale'])


@app.route('/dipendenti/<int:uid>/documenti/<int:did>/anteprima')
@admin_required
def dip_doc_anteprima(uid, did):
    db  = get_db()
    doc = db.execute("SELECT * FROM documenti_dipendente WHERE id=? AND utente_id=?", (did,uid)).fetchone()
    db.close()
    if not doc: flash('Non trovato.','error'); return redirect(url_for('dip_documenti',uid=uid))
    path = os.path.join(get_upload_path(uid), doc['nome_file'])
    mt   = mimetypes.guess_type(doc['nome_file'])[0] or 'application/octet-stream'
    return send_file(path, mimetype=mt)


@app.route('/dipendenti/<int:uid>/documenti/<int:did>/elimina')
@admin_required
def dip_doc_elimina(uid, did):
    db  = get_db()
    doc = db.execute("SELECT * FROM documenti_dipendente WHERE id=? AND utente_id=?", (did,uid)).fetchone()
    if doc:
        path = os.path.join(get_upload_path(uid), doc['nome_file'])
        if os.path.exists(path): os.remove(path)
        db.execute("DELETE FROM documenti_dipendente WHERE id=?", (did,))
        safe_commit(db)
        flash('Documento eliminato.','success')
    db.close()
    return redirect(url_for('dip_documenti', uid=uid))


# ══════════════════════════════════════════════════════════
# ==========================================================
#  PREVENTIVI
# ==========================================================

import base64 as _b64, io as _io, json as _json

LOGO_FHORIZON_B64 = 'iVBORw0KGgoAAAANSUhEUgAAAK4AAABDCAYAAAAF1j6yAAAzn0lEQVR4nO29eZxcV3Xv+11773Oqulrdas2zLGuwLWMb4QHPA8Zgg81oAiEMviQQSEg+vCS8hNzcJPeTz0teeLm5N7mZXgJcZhLHBjPEgI1tjPGA8SxbNhqseR661a2u7qo6Z+/1/tinqqsHSSXL5BO/q/X5tLpUXbXHtdde67eGI9/ctld5qaQC2PhSj9OMgrzkjo5GAkgHfSuWo39GVVttiAgiE9sUJo5+Yp8iL+/sWu1rs/+fQ9uTaPI8X+4+lcg2OmEckZOm6ttM2Z47xn52MJqjvH4F0kTGa///VPv8cjPqKToxcvIKZ7iXg04x4SuP3NSCuENqY/rj3dbhFG+cognULjBOVIC6TvlpYrsy6S/a9q+M+8zLLdSb7cnk7k/RK4GUk1alXacfDITYmUYF3rSY0xejGOMcLxbRqFYrARUQFQQZx3Ta8W/Fi2BUMQJBhBxwIRBbHSNFp2zjFP3HoOZeiE61R53vVMeMO541hSZbRrhAx/6qgAaQ8Uza/jNpEsf5rQoaPE491GvxinEO40pj142MfXaqNk7Rfxw62j6fyG51rOKKggmKKV4H4k87qSqo4jRg1BPEE1BET459RMCZwE9/dA8ffNN1fOC6a/jSp/+cpK3fdkjrFP3/nzqWuCYUOqwo2nZUtMBnm9inoLiQ442Qq40SUSUK5ZPhX4XhQwfo37UVUeHgjm2IBpCTMi9P0SuUOpe4CLP6enEyXqds/4+IYIxhbl8PqG+ZiicNuamgWMYUlkCQU/bY/850AjquMqOrzPCRKg1VAmA0SlODEjQgBkoIM3q6ZM+RIQ0oFh9ZTQyGgABeFNFosKlI1AVUMYUBFwjR0xUUxOIBJFD0iooQTB77DdrSnTPxSGEIUkh6W+gz/phH1CDeY6RlexLvjuZP89+JazL5Chnnb5OisRDa/q6tb4ooEnKsCt5YcjFYjUMPbZ6meFxlrEcFkbgWofiA4AkSPyEKqiF6AYv/N+HI5ohDMTeE4uZqDr5tBiL4Yj1t8XZetBHHKQT1KEI0NZp2TTEGFFEtPGXNz2jTbVZIzVCst5liNY9OHTOuBVJFEkVDwQQmRIaUACMGchsX0noI2oUXCyEjGEcultRHHbjuFKuQEMiNpy6ANVGHzi1ChZodBdsgzTNKwVI1wqgAYgtm9PQ36jQqZRI1dGU5iVV8yFpMq2LIbHPDx3NuixnVYBS2bPgZ29Y/z/IzVrL4tGUkaZmAIMYBAZGmqlR8X5XQ5PI2u9gYU7i3tWUE2Dakw2to2bO1+jA7Nv6M3Zu2c9YlV9C3dBkExec5ogFnBPDkGjDi4uZqPOSCx5gEVQsmoJqRSGFPeFBjaKgiNiFerAHRgOQZDkMwEZ1RLY5FwUxxoM1DE5nNBMF5IQDeWdRaUCWEHBMCRjQeNgEwrXEaIGhOoKlKBlQDVuP+o6BGyFCQpsXSGXXMuFOSKCqBukvYUq1yz1NPM6Orh+SK1+gXH3uCoVIFNEc0dpN6wWMYkYAL4IKSaaCOJ0cJongjZKL0jRxhZaXMa5YvZcac+fzjDx/l0IatGKOkISdoyoPb9/Pwtsc5Y8UKLp6/mCVpCWMESyDxSjCBUQfBQClPJw6egGAIuLzB7Z/5Wx783rcBw9wFC1mweDHd3d0458A0dfWxb2sE+iYvidhiaQLtqHbkh2hAZlnGyPAw27Zu4vDBA4gm3PLbv8/bP/yrgKEkns/8zX9ny3Nr6UoTktRgkwTnUpxNMTZFrKJB8XlkdO8b+EadvJZRywJnX3QxH/i1X6caFDXxIO16cT3/8H//GV3OkqSGJEmwLsG5BJemGIm3QAhKlmfkWY42GuS1jDyDBo7f+OM/YvqCBYh15MNH+Is//i/UjgxSTi1pyWCcw7kSzqbYpARGCb6B95488/g8kGUjZLU6tUbgre//T1xw9evxGOwJiNyOGPdo1rrRgvkAGkq9mtFfG+ZgyHl0/352JRUwgqhDCbggBHV4JyR5TjdKJXGUFMoacdmS9axZ0Ms1qy+AAN/buIEfPfIESg8q09CkCw1VmDaLoaSXTfXACxs38eP1mzhjZh+vXrGUM+fMYRaGbu+p1HK8BW/GgmjaZoYhEGpVdmxaj8nrGBUObt/CgR1bI6RnpECtC0laqDUh5Khmk9cEEyWVsWBMvAYlqkgUvBw0ICGABKyCwbP52adJNFA3gqqnZA2jQ4fZ13+Igf79ZKPDEJr3hAXxoMSRSUIQQ1Lpom/2XKbPmoumjqp6cptgiFe7NYIJgQO793J48ABH+g+ieaO4vQ3jcaKoRhiX0ts3k95Z85gxfxG5jQfBaxx/YoX9/fvZ1X+QwwN7yWu1lmoTxwngo2ooFmwXld4yfTNm0jNjHra4EfQE3e7yb1uOHx2mqpQ1sHLeXNm+/5AOGFDi1R/HJVQTS8MkVFRYPSflz+59mjte3ElmLU4TMuo4tYgkBOrMNoEPXPkaVs5IqfhAX5ows7sbZw0jtQYHRkYZDJY9RzI27NrP/Zs2QrVO7d5vkux5kfd/+u/4zp4jbKyN0kia6kNOl8+Yn6S8asliLlq6hKVdFbp8jhrfQj6MMYWuqbgQGN69k4++8yZGDu3FiODFMnfpadzw9ney8pzz6JrWA4ANUWWSAN/42hf54XfvGL+YFJeySbjl13+TNZddRWYcgQxDRGaqh4d45smn+MG3bmewfw82BKwK0xev5LPf/wFZWsFqwOUZiVHE5+zeso0//q1fY/fmDWgI+II1DHE8KglLz3oVf/w//5ZZCxYjJqWBIUsciCkQoYiDpz7HemWkVuObX/xHvvIPfw0+J0SLYdxsrDF85Pf+iBvf/V4kLZOLJRcb1UMEpx4bcpwG8BlPPPQAf/ap/4PG0CCRXWM7TdPai+WaG9/Bxz/1B0ybPgPFkYmjbgwYgz0B2LRjVUG1aViMUS5C7iDBU9KcUjZKj1fK9MmZyxeo3bqDulG8KiFxeK8IdaxWmV22vLqnlzNdSqUnYTR4Dg4OciTkNBBKITCLjN4ey7KzF7J4UR933vME/b3zGT24m0H1mERJ6wGTpZjc0HCOURK2Z8rWrVv54eYNnD1zFq9dtJjzF84jMYITAQ2FxqsIgd27dzFcrWGkhHdw3U1v5WO/+59JZs2hIY4gkSGNgvGKU6XnvnsKzbVtjVAwhiCGmUuXc9r5l5G5FMgAjYyvsObaN/PWd/8if/l//QHPPPQA4pWB/kPs3beXmUtWkYvgU8eoerBlZi4/h9UXXcH2TRvGMRYFqwVgzRVXMX3RChomxRS2lvOCFs4hL4aGkXiEjIGkl0tueCu3fuFz1IcHi51tswOMIemqcNUNbycvzWwpPTYIxgiZUTKxNMSCKCbp4sxLXsfcpavYtu6JtuCVOE4PYC1X3PBmQt8CDosjEYl/MQWu+nI7IMYFQ7S9bzWQ+AZqPSrxVAUrBCvMNDDTC4aUusnBG0res1jqvGXpXH7t4gtZMb2L0azOlv39bO8/wmBuCZqSesg1ITNdCJZKnrNg9w4Of+3v2Xbnv7L3uWf5+ic/xvkDO7hi0SwQpWEMiieYgDcQJKVqpvH44RqffXYD//W+h/ja8xt54cgI1cI4kswSQoOtG9ZCfRSVhKtvupnf+K9/ip09n4Z1YD0WH+WGWNQYMAEjnpbnsPVToAHaiPq/ESQ0cKokKhiRAhERepYt4/f/6m848zUXohh8VmXb5p+RFBLMaMAWJqUaT9/MPsZwjCbTmgIIyOjunQ4iGAmo8a0x2uJTiSpJAIfBAoYGpWkVknKlZd0L8fqHgAQlqfSQVioRDShOrreBIL5AiAJGpFCPwKSWnu6eCYwS0OY4CZR7+7DWkhgFiciH0RPAZQs6KfTeqEGCRdWhzWsES99oSlkS8tRjGWG+r3N57wifuGglv3bVVaSlGXz9qU08PTTCTq1RTTxIRldoMDe1LJs9gzmVbmwGiXdI3fOXv/+7bH76UaQ+iPGj7Nu0nlv/6Pc51wbmlnMyN0pm8yiDClgIichCzTh2escPN+/hH+5/lL956Gnu3DvEdgP1JOWFZ56lpDlLF8/iQ7/9W2SVHnJSTCiRZF3Y4Ca5qo8uG8Z73Se6u5sImQRLedoMfudP/hTXVUHywM7nf4YpAtlFx+S5iFAqlY/aqRghTScYnzL+5VTjcM5inRt3HFp7CxgxWOembGiqNk1zHBPc7u0HzlpXoCKTv38idFKoghINL9FoYwsQxHO4lCGmxkqX88YVKzh74Tx2HOnnzp+9yI7BDWgeOH1aig1K2UO5lDC9t0ylK2FkZJRDg0OMZEqwJTCQa+DAji1YbWAJeA0ED43qMPngABWboCjGpzjxxfXYPk4lE0+eGBpi6a+O8uxTa1mQWM6bPZ0NQ1CvzOGSd7yHmfMW4IMikgEGMabAVCeqBcdbnGN/wmJo4Ji36kyueNNb+OE3bmfrxs2EEBBnWzhus7MmBHq0vow5wa1vMtdRjCJtu0U6bU716O016eWKfT45OKwJiwePFY/zOV3O0G1hzaw+qhecz11btvK99c9RDdNQLVFLLJRGKVVgaSVlaamHmnoGRkY5MDSEF/A4vHXRGg85WTbC2975Tv75y1+ikWURe7WW1a+9mEXzFzNn7xCJr2LERcvb+HGjNKKYAjpXBRWDt8IOTdm7Z4A5b3g7p7/6Yhrnn8v2hmFGV0KiOdZkeK8YaxAtYK4JsNixqQ10n2rtRPCa8Ma3vwf1gRVr1kRQv81hIUWfx9ruJlItRWBTp8whIpjC0JKWA2Jiy52TNJ1JxxroCYzvWDQ1445TrIu32n7aPSxBciyBijPMmT6dHmPFNSzrBkb4H488xl4zjRDmY01gmta5aHYPV644k+WlEhbPpsEBch/AOAJdsWsVrIAJnsRn/I+//AtGDh7i45/6A7Zs20ZWq7N02ekM7tvF//zkJ3nPn/8Vzx8+xCHrMSFqcM3RjyGpSQytVMUUOq6XOtXUULNlurpO555Dg/zkgUc4a95sLlu2iFU9vfRiweeoGZu3TuFFG7dQ43hb2/0TY++aPDJMcKy58FLOfO355E7QLLrOpuSjY5BOetH2xsS+ix9jDGKkNZ/2M/Yf3Z0+mXFVER0DXIp4MAoTRQMG0QxLoFuE3u6UtFwmy3MODlc5mAddPm+eHDbCaCqYbIRV01MuO30RK2fMYNOhQb7x7GZ66yN85NrL6DYBsYI0nYnBIJKgUiOgbH1xEz/6zr+Rj4yw6txX85E/+CNUYfPTT/O7H/oLfLXBvp/cz5qzz+GBA0M0jI2u1AkbIDrm4gxQ6GmKDRr1YKsYSRlV5dDeAzy5ZzeLK92sWXwaFyxayJw0payepKmqYBDSwl5uAj9FfyIEE1nchDDm+m1fZokuTjFKXSCIiy50E+G6oBMxi/EkbfCV4silMJAmfC4QceRxQkiiEWlVEVHylunUNr7Wuh1jEBNINRxbRZp8hl4yTWJcFciMKYK/C/+SQsnnlDWTCnVNKiV6Kt04LxysZVQPDkdQXQxWBLENljQavG/eacxZvoSh0VF+vHUHdz62jgFTpmZLnJdUMCFK3XaPvFHBaINgQHzOv3z2H8mHD9Mz/zTe+K5307AJBmHleWu4/PqbeOCOb3Dnlz7DL3/+yzx/qMah3KCmkCA6Nic1xcrR5rZFIg6q2oL7jBq8JBwxhhero+x67jkeev5nLJs3h1cvP42z58xktrfxUJsGJrRLp+iJCxp/F1bAUZa+8M1LHI9tspyZbChNpknimwh5BfyU13DRtjYZVFA1iFhEzCRDrv2+7VTyd8qQP0cdtxlYAlY9RgOJKJUEyollzuxZsq9e0739g2gDRkxccqdKitJTSulyhoWzZ1Bbv5P/9cAT7DI1RqQbTaaBddGDVsBWphUQE/sWPBhFgmP72qd46M5voyIsufo6mDG74Dols453/qdf59Ef3Ev/1g1suPObrLnweh7cPUBdTIwek7EZHY+MMdG4UBP97gh1CzmOmknZ2T/ET/qfYGk55dK5CxiwFYzpxoYGSk5uNErLvPDWh4gVB1HG8kVePmoqK80jn/qcpBUX0Cb4VJCWniNFAI8UCHY0PoFWRsJLH8+/L01iXFGl5BUJnq4koXdaN5VSSu49o2rZcWBAB1CQlFQCZR2lkiZM7+qiK03QRoOsITx9uMo3d+5hLyWML5OqEqROTh0BDifK1nyEui1hjUULg0SsI/Pghkb4/N/8NT6rkyxcybQLL6VqHNN8jtVAwzqWrjqLK978Zu699Uvc99Uv8/Fr38K6/YHdEN28xXxsODbrNKWAUFyr2oTSEnKn4IUgCaPWsDFrsHfrVnpXncO8j3ySw48/Sn3jc7jB/TifxYAhBSGmLwXTFAMvJ+sqzYu9iaI+/uMfMtTfj3VlnHMkaYq1Fqwtgmk0xgvkOT73+HpOvdpPdXCgGO9U8W8nRlF6v9xHdGqa0jhz4pkzdwYGYWS0xuEDA4TgqSwsoSHgjCcRZWZ3hXJ3N7lXhqoj7B0cJjGWrq5KjBILSkmVxMdrbDT1qBg0JOyr1vm7ex5kmilhrW3pdCIC9Zzy+hdY95MHIUmZ9bq3caSrF4zjxZ/+hF1bN3P1zTeDTXj3Rz7CI/f/gKG9B1h7521c+Iab+cHm3dQk4opOo36bdwKgCAWsp1iNQThBoiFlVDFeQBJGbMpIdxdy7hqmrV7NrIP7yJ9dR/+TT8C+F9G8TmjlZ8Tww58LSYz8dMCTP3mIJx5+aHJaSsvi0nE3m1GDWE/QvGXPNHX1VwI5GBPzTQs8V8P+/kFyDVFTM9AlgZScmRVLX6WCGEujWmf3oRoN7wnWoqaEimBEKYcYzjeSeoL12MzircV4Q24TJGTUxVETIclqGJ8Xuqihp1Fn+313YrM6dukyes44jbRxGFMbYsPzz/KTH9/PFTddT0LCorlzeeONb+X2//UF7v3aV/idN74Vt3IeIzRIVBAchyXhJzv2kYfo3W+YnHIeCKLkRoAcg+BJUVHSEDCqZIUqE4rA9XIe8eLgAj25kIdAlqQMzFlA17WLmHfJFfS8+AwH161jNO0hx2K9xiSNn4sgiizn8aw4ew2rzltD4hKcdTG8svhEKzswKD54QgiELOfIkUEevud7NEarL8tV30zdOokWpnhv6oVzkz4kQm4Er9GFm6CUrWNGdxelxJE4x+6hKiN5wAZDbgzexaVxGkiKONiAI8ehWBLfjPc0lHNhNM3pK8OHz76AJaUSn/n0n7J704bihrYM7T9AtnM9YhLSrMTuf/kWLvF8+l9vZXTXdg73H+K//MqvocYgKCPVIdKuEvXBQb7wsV+hb8kiGkmC9Y5gEt736U/z2NYdeCnH9CITuGTRAuYkjoYxWMnIxfDU1oP09VRYPr2HMoGaVcCyaWCA5wYHQS1BAkEa8OxTjDxyTzS+VBkW4QjKQQOSCYGMYBoYNVHP/HkwbmEEqsA1N72Tt33ww4QQwa0m4zZZV2jyVAyt9KJUDx7gmccfIx+tFuH+L53i9E6W/U+Accdgj6YuFq91Zw29aYm+rgrOewarQ1S7PIeOjDJqBG+arr22u0k5pp4U100xAaZLyor5M+gNo/zWp34DzTMUwecN/vijH+bw7sBZF17M6z/5+3z9hRdZYLv5yNUX8tRtX+bxR3/K7336r8nSlBByUvF899av8sW/+SsGh3bziY//IfunL+RAnlA3yvO1UYJNqIvixFPxNa5dcRore1OMWqwGDjvDrv0HWT2nzHVnLafsPZkFEyzf2rCVFwYOIyE61SUP+AN7aLzw2JgHqvhtMahapjUOY6RG5hKSPDkGuvDykDFCpjE+ounqbsLtzZT9sSE0syReuQmmTlWI0VIBI1Apl+grl3AiHKnX2DNwiLoPOFFmKOqtK7xbtkBuJipVR6cmxikKaRBKmZIaS6lnTqFXKnd//Vb2Pf8sxnXx/t/8BL3nrmHu3gHUlWHBHGaffw5ndJXQRfMphwILJefN734vd337DvZs28Rtt32L8z/6e9y5aSsj5RI29NMwlmDjXG3ISfOMJHgS77Ah4KwDm2G0TqINXGjEttVgQx4ZQEBNNPRsSBBfAtr97YEYCWYBh4QSNqTx/Z8j30rh5BAxeBOPyDhYbKKzL4QIA3KSN3sbjTl6/n3IRRchTOvqYvq0LvJGjcGhIepZTkOEzBhyl1AKWXHdFcClCYWHZ+waFDn60E0BPucGvBADh1VQEnIRTMjIjgxx2+c+h+TCmiuu4NwLL2Y0KfHrV12DJzBbExZf/AYuu+j1aACvnlwcnpTuWQt453tv4e///E944t4HuOK9H2PmtDIDmcVqQi4Z4g2OaKiNJiUa0kVuLEayyG55QtAKWehGpExmIqSfmwQV8E5ivhyCmhw1daDw0bcwYwEcuTggwYUEpPHvtKcG08Jsm+9N1jubwfDNypRRGjfxw5fOycd1+b6MFHVchZFajfpoFfUZtcSSO4MRiwQh8QaneRGnrjG/UWMVmSk3pF1lKO6ruCRF4pwo3sS0HyVHxeAk5/vfuY0dW9dju3v48G/9DmCxIWPmNIMLSlDPqCsj6rHUSZQirdIScNzw9nfzvdtvZfv6F7jvC3/LG//zH/LVZzYzKhVUIfFR9DTShO8/+RT3uUAQi6FOwxrW1zJ2793Li/39JOqpuRjHemioAUU+nTYxUjUxJ6iNhBir6wkxNkKyAqueevEnLZ0e5XWHJChmkgg9SkM68W969M92Su3G2c9ZA3EWBSN4jcEtWIdRSKHAMwvXKKYAsuOF2Ew/bGaZtlNCBFbUGFyIVq0v4i+N2ni1FdLbS8w1Gz00wG2f/QISAldcdy3LznoVVZOgJjKB4LCA1bzVS5AiI0EbMVGvdzo3f+gj/Pc//D3WPngfb9r6QZZXUp5uZFi1pAFyF1Cbc86yxcydXiKQ4NRTNZb+JzeyZOY0Llm0mK4sY8QpdWt4bPtetu3ei9EC3Gq5Ltuzd9u2vvA6miIbOkyBhk12Arf/ZSJ1xgXSMfMVKUniC4lrkCkz6DqniLod2019fOocNhzD5Ztivnn9F+bEeKV+0n+mJOXo1RnbN0zFAhYXAt++9Z/Zt3snrlzivR/5VTLASEyLdmHymJrxtvHXmK5y2RtuZNnq8wj1Ee74+7/mimXL6c4hc4aGFUQDaeZZOr2P83sqXFqpcHG5mwunz2BWyFiWGNb09XL+jOlcOL2Xi3t6WZSWI45b9B/k+Jr92GgnTPzoHPvvTzLpxcvTXNv/f15T/fkg4yJjagGtHMFITX95iNUTBBjcs4N/+5evAoEb3vVuFp+5mmBjHQYbwHZQ/aPJvGbaDN7/8U+QuoQXn/wp2cafcU5vD04gs9ErlgaLIHgXMWa1nob11FIlszl5EqjbDG9yMB5jFB0/CY43oJfPWtfjtqeTXnTQqiohhGO2G0LnDTabaQ/JnOpDLxeS8XNh3CBCHjxeQwHPTP05USX1Gf/2lS9weO8uKr193HzLh/GSEE3hyC2hQwRfRFBjuOCKKznz/NcSQuA7/+9f8frT5jKtNooRjzEWLwmHguFgrcyerMRu7zhUUySUGMlKHBgV9jUsezPL/obhiC9SkqJrDxtA/FE2qBimz/OiJNXJyxxVpZFPzihup7zor7P24u8QAnmeT/0hMS1nRScU7T0t2jv6dZsdZx6d0gkFkrfLmXEyZ8ILBWqNBnnwqCRjnnptGq4a/eV5g707X+T7X/9XjHpuetd7mLV4ORGsia5XlWb9g86uHRNyTFLiPb/6m/zRE2vZtu4J9j52P2fPP4vHjhxBxTGijtse/inWKA0roDUQ4Yiv8LPabh7auRPjG9QdoJbMB7xzWB8PmwsK2VHco4XFPjQ4GDOG25ZmctDP0Wc0jgVVGegfaDrKxv21ubaH+/sREzO84lqPFxhK+/fiGKvDVbJGY9K+NvvMsoyRkRG6u3vHBt/W57gpaKzvMHxkeNL42haHkeEqTWT5ZI70CUhcxQuaEzAhhgE2JPq8y7mguacuihcYDUHXHuhnWB1lD4rD+gTnSzSMxWGojzpe7B/k1q99FTLPlde/gde99UaaQIW0PD4Ta9Acm4KBhhjOuewyLnjdNRiBuz/7d7xh2Qy6RGkAqTccVGGXKPtU2S8J+9Uxajw1DRxU2G8SBoJjQA3DIiReYvVUhbLPCSYHMSDNmmbF0QrRsn78Rz+itm8vJh+lK/O4XAkhsqxRX5i4E1ZYBdQRCoMWzRHfYHj3DjY8vRZRW6zNGOsoMYlzwxMPM3xoBxpGEd9AQl5czbFaTQBycoI2MFrDVY/wwoM/oVEdbYsmbia8xzn50SpPPnI/0qiS5BnWCwQHIUGCAQ8SAuIzqFfZ9tyT7Nu9nZanDsZ2Txyo8NwDd5NVD6GhjoQ8GrxBCHpil39HdRUAnK9z+oI57DnQTy3XGFguSm4zckpsr+c88Px6rBd+6bJX8du338U+Nw0JntxKkZlQWNgBEhEWH9zF1r/7byybN4sj9SqnX3AJn/qT/ycmDJ7AJNqvSE/MpLWasePZp/nt978HP1rjw3/1OR7qXcj6IyMYIlqQhHD0ftrEipdYvFpJEKP01o6Q3fMdqt/7Z0DHwaS24F2MY/qcucxfdhpz553Oa19/PZffcENRQ80X4ZzjN8sXsi/JG3z/9n9l3aM/Yf/unezZuYPB/kNIyNGQR6jR0MrojmWflOmz5rFgyTLmLlrCBddcy+U33khuy0hwOM0Z3LWD2z/zGfbt3MbOnS9yaM8e8loVpMjqUFpysBnM7rq6mbNwCUuWrWTWwiW896MfozJrNooQRoa59TP/yK6NG9m3ayv79mxjZHAo5vRqUdC78H4EATGCsZYZcxaxaOlyZs9fyHW/8B5WnH8+Xi0lOi/D1LmqIIKIobe7QnceABvrW5GQk5I3qiwxCZVSmUUm5V3LljHiSlhtxLJKRrBFbOgohi2HB+hK5jDjjW8iDBxA8hqXv/ktBGMx2rk3biLFAnxKwHL66rN53Vvfzs4NGzl71Sp+tm0/K1yOipJLincFbtUmH+IhiK/jFsbfLgRUc7wJTEst9rRl2Kuuo1Qq4ZxtqQiNLKdRb1Crj1Cr1agOH+GF/Y8zc8Fcrrz+ukJ1iodrkhWuhW8H4bmnn+H5Z5+hUkqYtWABi1asoFKpYG2CTVwMWA8eHwL1ep3aaI3GSJUjQ0McOPQc5ZnzuOqmt+Pb4J2Bw/088vCPKVkoV8osf9Wr6K5USEppLMVkXAHFKrV6nUa9Tm10mEY9Y9uWzWzeto2b3v9eps2dQ8iFkOU89eijjPQfoFRyLF6xirRUpqtcwbkS1iWIBNT7GBZbr1Gr1WhUhzm4fy+7du3hnMuuZOWa87FFzbWO2bFTiWtCTlfisEXKSQxviYzoQgTrcxtIfSAh0KArepkkA3UkQUi8YoJQM4YDeYOasyS554kHfsTClcuYs3AJqSa48UH5x6Vxz8sq0mq8EZwGagf38fnP/D0f+T9/l8O2m7rkpAqp7+oYuWyYvAgSEhomUPKB6Wpw0h7MUvQfwBc4itGAb9TxeRWblLFdPQSxhR46OUZYgyDegeTUGlXyvE6aOFJJUGNQY0Ga3rG8OflCTitBM0KeE+qKkRJpV3f05BmPNwEf6oxWBykZS4kUm6TRNSymyMBo02M1Sk18hihkuafqA+n0HgwJSRbbHa4OYq2SJJZE4hhVEoSI3KjksdlChVJVchr4LMPXla7yNMSVCly5c87tmHFbcJBAs9xbRKkMqQ9kNsJfiQ9kLpDkJYQcbzPEd5EbIUgoCkPHdHaCI20oe7ZvY+Hy0whiSb1plfrslNoZ16gnIGQmXiZJ1mDd2qdYcfbZlNJp5K6GVaGUVxi1eZHKPiZxixbHv6eK8/GduhVcEKyHhm3DkONAYgVLF4Nc8IpTUNNAMVBUgdEiv84UDzdsUtCxIkiCxBKqEkMrfQwyw2isiNNym7eRUHjiQ/SQ5CZi4RKK2IUCizYav5+LIkawzXFD07SLs9dmUZOo9QYxeBECHhOi+iRN9UKi0PAmHgSjsepNw4aWQ0ZCLDsas0KKNkMguDg5p+PX41g0JeOOcwRKIKCF4TBWn9VYE4M1Aog4MjGEoo5qMJ7ER2+MNwEJSQHcR6aNmbYxtC5KJh+ls4luVC8mVg7Uou4uUU0JEv07RqNrNSY/mkkSVxG8xFR0i6LeAw5jhIZpoNZgQooNEbkY408BmgXqQixcB0Udm0h5wUhxI4vkwGZWroIRi1cf7WYxoAaVvFi3OK9Q9GGYqOOGIhYibrILkTEaxsR1o0hwLBwyuTX4YkNcrlHGKVEw2MJRolp4K4EgJFjU+1hCqSlsUEweyMUWxUhopeEHiniGEIpiJpHxgoTI0BqZlGJ/QzGemDtIjHcWWrUptJib1VjTTG2gYTK8OBKfkmjkGSUyvh4lAGZysiRjp64pZg2R2cRH5lIVgvetZMrIXPG6URSrlmCassOgZizaU4p/QlEDAImRZlB4o4xEaxPQIoopnlgT3chFGSM9iljWwuBpfg8k1nOVGBCeiC2A9QZNrLjJdK26HwJipbgZmlIlUgI0I2oMwsSgEh8UR3wGhhoTnSyqEdNWDzYpRjf5ojMa68oGo6h4QrEyNmZyxgNcPAzU+phzbY3gKZ5GpIrY4spXJpc2MkpOXkCSYwW5g8YRGdFYEDrGTxXH1dI8282H1EgAJ661oWNqdBsCVLwvPhploh6lKLxtDN5r3BcMrlhLi8f5DDVCjgG1UzItHMM4U6KFvGX9RnZv2065yG2qi1Dqm845r3kNO7dvZ/fm9UUtKkPNK5XePi64+DKO2mMbNf3jMVDGFEZQjoxU2bRlM7v27GZ0pIYTYcGi+Sxbvpzu6X2oJARxhZbdGQViVcRnHn+SI/37SYoiwzmgJl7aC+fNZcnpy0kq3XibFCnoJ5AzVui0T/z0YchGMWLwGBrGserM1cyZO+8460Fk0MI+FBHWr3uWQ3t2Ii4FU4rwlvdIGIVGjm8EfJJw+Q3XIy7pXMOSmOXhUIz3+Eadbdu2sGvPHkZHRjAKC+fOY/Hpy5k+a3a8wSRWDBOETi91UeGFtWsZPLSXhBzyGt09fZx90aVUXYKKMLRnHxvXraWkAReUqjWcfvarWTRnMUez2CYnSzImeYLCksVLGN5zgG9+5cuIgUve+AbWXHox1joWLVpEPnSQr/zTPyI+cOFVV3PBJZd0unQ0i9RDlDQ2r/Pkww/w4L13s+JVr+Lyq68hsQkD+w/x3e99k2/fdiuXXnoZl7/+ekylrykWjt9PwREBw+ozz+Lhu17kgXu/jwq865YPsXDZ6Tz20IN89dt30DdzJje//xYWrlxNEInWVvNq6KAj5yznnXM2t33p82zZtIk0LfO+3/gEs+fMnmTITb0mTcZVNPfcf98P6e2yXHzZ1XRN68MoaGOUe+/6BtvWb0Yo8fq3vQMxx3rM9nhS4pWeqGLzjHVPP8m9d9/FouXLufKaa0jKZQb7+7nvu3fxnW/eznnnr+HaG24kndaLx2GMLTC/DvoysHLlcjbVBvnGP38Bl9exSUrdNzj7iqvIccyaOZNp553Lnbd9jdDwvOUDH8T1zingwqkdPVOuZHOLggHXVWbWvLmxNpXm9M3qo9xdQRDSpMScufNJBFINzJ4xk65pvSewgGPBKjavc++3buXOW7/M9BkzufHt76Bn3jwqM+dw2qrVfPBDH6anq4sH7/out33+s+TVoUKVYSyu9KixoFqEmwrl7gpz5s1GQx1rDb0zZ9I7dy5XXHstlXKJIwf2cP/dd6OhsNZPAKIBAZvQ09tHz7RenELZWebMno2zluNm0WpLU8EgjA5XSUsl3vFLt7B09bnMWryU+Qvn8eKGZ9m5cT1BPOe/7jIuuPoqMOkknflYwwwoeM/9d93FN772FbocvOVt72L6vEWkfbNYsGwVv3TLh1kws48nfnwfX/vcP1E/MhSVHN9ZUqUSs5yTrpTlZ6zCuISZ8+YT/Ch33XEra++/n0rmcdYybeZM+mbNo2/WXHp6p8f1OsbiH4Vxi4dOAN4K4myMnS2qzoDGeNygWJtgioKYYlxhdnR2YTVr9hkNbHj2aR69725cXue811yAJileHN44vDFM65vB1de8DvE5m55/jkfuv7dVev14zzlrSTFAjaBGMUWeXHQpm2gJm1ggI01KhV6rUeJ2eBQjYmaiVLLROG1Z3ZhorLU+OdX3x8o7qUJXdy/v/sX343pmkNkSQQyPPPIQP7rne+DhnPMv4tq3vp28XIrlTzs8ZHGZAts3b+Th++8l5A3OPe88bLmClwQ1KZk4ku4+rrz6Gqz37Nu8ifu+d2eMkjOdW/+emEjrRSBJufl9t7DynHPRPOOH37yDh7//PfAZwVhISuASxKREPf3oB2RK4yxIwKpBgwXjQQOmCJpZ/9wLDNcKTUeV6mA/6puVYCL80impxMW26tmwbi3io5HXO70vujfVRO3e5uTAkuWnYxKH5oEXnnqSq264CUkmPtthKmqm+CiiefF0wRgPMTowwOCuXTx4713UR45w2upXc/3bbo7Zyu0YYEe9KBoCwVhEwFLEARgXD2kEII4ad6EGvIT4wlhwCaXeEqqBxNdZ98RPuffbX0fynEVnnseN7/ogxvUWa9/MsujAtpA41s3PP4OODoJNKM+dR25CNP0UEEMugRkrzsDYEmk2ypafraVWHyEpTe/weQ3NclI2Gl8aSHr6uPF9v8oPvv111j78EA/ccyej2QjXvukmjAYUTwgWMRYrWRG3MrmzY3rOpM0sMUYIAWbPmsWKFSsiTOE91YFenpTxnvcTgGChqPOlWcAFi0igVhtBCQQToRZPRDKm9fRhjCNIA+8bJ3qPt8YW9zfixc+sfZoXvn47fvgw559xOm/95Q+ilb5o1bacLS+ll3Y+moy3TvktjRXLtXmVm5hLYT1sX/cMd9/6RXy9waJlZ/POWz6M7Z4WM51VC0hKO1p8QbBqyHMFsUhQ8uGRohCgiTMWQVSYVq6Q2ASyBnkAEyLM1ekmx3tNIjxW9G66pnP9O36J7tI0HnngXh790b3UhoZJrEOstB5iI+Ho2dHH3ZPWFRziFTF33jyWLDuNJctO47Tly1m4aBGgUdIqnJDfruXzEZavOgNv4iORdmzZhCWABkLhZzJiGR4eIfNKrrBq9Wps0vmVNWlOIWI+l11+OTe95S0kacr6Dev51rfvoN4YjQdHQ+sm6XQ27Xp2K4uleHG8sMN4qKK4MARQj9GM/Vs38+1/+RJhdIjZs+fwjvf9MmnvTIIxEHJM8GOPe+psASAIp69ajaRdqCo7X9yI+CIpVG3EnDVQHx4mBCW3CYuXr6SUljtHFIq+DBGTbzohAila6uGqG9/BtTfciDPCs48/ytrHfoot0u3HnCFTc24HjNscRcQ8gwYwQkaIQSEagXYtnAUnIgSbWQWKcO5FF7Hq/NdQN7DuiccZGRggCQGn8aF11mdseOF5ch+Yu+R0Ln/Dm2g+yeH4xln7fIqHl2g0DVWE8y++lDe/6xfJkm6eeuhB7vyXryCjR3AhxzWdDB1SjF+Nj3Ka2O/xxhekcJwAhoALOf27tvP1L/wd1aHDlPvm8O4P/Qoz5s6ODJ7luNDg7jvvYN/eHSdUUE6NYdlZq1l9wUUEY9jwwjr692zH+YihW42P5dr64gZqwdM9bx7XvPnG6ME7kVgSBfUB9QFrDBoCScijcykpc/Hrb+T1N/0CLinj8xpeozfTqOCOYV5Mybja+nz898jwcKyUjaE6MkLmPRTenGq1Su6jd742OgodBh7DmNEURAjllLd94H1c+cY3IlnGF//pn9i7bQc0ciTL2fjs09x/7w94zQUX8qGP/jpdM+YVqT8drV3T5EFVGRg4HOvCauBIdYRcHK+++Epues8t9DjL848+whf+/m/Zsv558lqMJut0PgDBe0ZroyiFj79ajeM4zgFQClcu0XU73D/A7V/+IiOHdmHLJa6/+b1UZs1hZGSI2uAgA3t38t07bmfX9k3MnTsDOVpW5qR+JEbRpSlvvvlmrrvxzQQC//qlL7Jj8wakUYPGKNs2PM/dd9/FGeeeywc++rEYXHOCgSRSxDzs37+P0ZFRDh44gPU1nDZQY2m4Mhdc+Qbe8gu/SNpVKnR8LR7id/SOJrl8FcglMp9TYcNz69jy/HrwDdQEvBhmzV/AxZddyvbNW9mw9lmyei0WWBND36zZXPy6N0zyKB1tCWPnTaRCseo5cmAPz617nj379mOdJUlL2EoXrzr7HObPm4+xCUEszcvkeAwRCHjxOIUnH36UAzt3YTUQTEBdwpLlK3nVea8mqDCwcwtPPv44hwYGEGOZPr2Xq667ga6evuNPJwSy2jAPP/gg1cOHW27prNzNBa99LfPnzz+mVGwKmBiPoKx9+mk2vvA8ZYleOJeWCmwzyuQ8z8l8xjlrXs2KVWcQjOvINRDGdJiIFwdP9cgQm55dy65du0jSMs5ZymnKsrPOZv7iJdHFb4qqPNDZ/qpitcFjP/0pB/fvQ0MOWCrTZ3HplVe2kBw0IOrZuP4FDvb3c/GlVyLiCj/s1Fj9lIwbClMrFnpTrBTl3aVgFevw2mS7Zpp6tCrilXgSlVvUQ95oXefGGEJQvAUNJj78omUyFYWNO2DcXHOcGCQUFclDLMWkhXfLG0GMjTG6IRBCKEroQ65FZNZxxx6ZwJoYYNMMp/bWkBc3kbVHb6ellQH4KHkEaT1csD3lPyaQRmy6lbypdOTkaF+vcfq3jKkBESEqHkzI2PPhTohUUZ/HQtuM5X2ojDlLmgdZW9BXoekf5zad0nNmx9aH4IoSxcYWjUULn6ZOKU0MtXBwc3JJGYqBtIssNJ8Pa+KzIMQXcdcnYNK25iQ4cdGnb+PTIFDFFX7/Zi3dUEgyEYvaaHA2cdVOtkwBcSlZaFZZbxbvDCe06REPpuWwa8+5a4IUuch4Ziiw9ZdMhUtXbFKENMYx0HzI90sp9CFCsK74rjbjkab2uomgIbTqFB+PjhmrEANamgsorWuiaf0SxgpQNHFKEE4wC2PSBGIGQbNKQJQEpjmGl9YoiLSqFjbf8q0DENUUiGX1izCoKMVEYphgh/2EEHenMP0AisjUjkbZGl+QsWDzpnRtUlTnxpYj1nCQl7Q+7bWB4yPOxs81yMkV7JPCLd8MI4jMO0UQPSDNauwddHgUBwRtiX0xykilGRwdsdWWZ614jGHr+lGQIu38pZEiMkFyTHygw0tpVZv+q2LxoEiPL8IYpXlM/BQMcHRYZvyn4kGeuIYt0XncQRa/ZEziBvSoRpcgMf62NR9e8uFWaIvoa+tDT65GqilUgHb+UE6+APSUEjd6xZqvi9/aPJmFhBVP2yPkTmoQk/pvbXS7WjCReU7sWmxGkrWHKJpWmmBxLzZdsidzQNpG26GRP2GcTQPt2EMx2ia12iTaSdMEhjr5ZmVsEhN4atLnWqLl+DSljnvMi22CJDHNeM2X4MWaunmBjiHuTttk3CaPvZw45qln3vHM2qLVOpPRk7oHGK/SHC8dRDoX6MfrOj4h/uUVQmPCQI8zxs6ZlhP65Ck6Rf+B6BTjnqJXJJ1i3FP0iqSTfJbvKTpFLydNZU1MrRifYtxT9B+IOmfcU6rCKXpF0inGPUWvSPr/ALx2mDBy5hbUAAAAAElFTkSuQmCC'

VOCI_PREDEFINITE = [
    ('Posa moquette', 'POSA DIURNA (AREE E PREALLESTITI) MOQUETTE TRIS SPAGNA', 1.4, 'MQ'),
    ('Posa moquette', 'POSA NOTTURNA (CORSIE CENTRATE) MOQUETTE TRIS SPAGNA', 1.6, 'MQ'),
    ('Posa moquette', 'POSA NOTTURNA (CORSIE VUOTO X PIENO) MOQUETTE TRIS SPAGNA', 1.8, 'MQ'),
    ('Posa moquette', 'POSA MOQUETTE RAPID/BOUCLE', 1.9, 'MQ'),
    ('Nastro', 'POSA NASTRO BIADESIVO LATERALE PER NASTRATURA TOTALE CORSIE CENTRATE', 0.4, 'MQ'),
    ('Nastro', 'POSA NASTRO CARTA (PER PROTEZIONE PAVIMENTO DA NASTRO BIADESIVO)', 0.4, 'M/L'),
    ('Altri pavimenti', 'POSA PVC', 1.3, 'MQ'),
    ('Altri pavimenti', 'POSA POLIETILENE VERDE X COPERTURA MOQUETTE', 0.7, 'M/L'),
    ('Altri pavimenti', 'POSA ERBA SINTETICA', 2.5, 'MQ'),
    ('Altri pavimenti', 'POSA PARQUET', 12.0, 'MQ'),
    ('Altri pavimenti', 'POSA SISAL', 1.8, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STAND COMPLETO', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STRUTTURA IN ALLUMINIO', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO PARETI / PANNELLI STAND', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STAND 2 LATI APERTI', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STAND 3 LATI APERTI (PENISOLA)', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STAND 4 LATI APERTI (ISOLA)', 0.0, 'MQ'),
    ('Montaggio stand', 'POSA COPERTURA / FASCIA SUPERIORE', 0.0, 'MQ'),
    ('Montaggio stand', 'POSA SOFFITTO / CONTROSOFFITTO', 0.0, 'MQ'),
    ('Montaggio stand', 'INSTALLAZIONE ILLUMINAZIONE STAND', 0.0, 'NR'),
    ('Montaggio stand', 'POSA GRAFICA / BANNER / TOTEM', 0.0, 'NR'),
    ('Montaggio stand', 'INSTALLAZIONE ARREDI E ACCESSORI', 0.0, 'NR'),
    ('Smontaggio stand', 'SMONTAGGIO STAND COMPLETO', 0.0, 'MQ'),
    ('Smontaggio stand', 'SMONTAGGIO STRUTTURA E PANNELLI', 0.0, 'MQ'),
    ('Smontaggio stand', 'SMONTAGGIO ILLUMINAZIONE E CABLAGGI', 0.0, 'NR'),
    ('Smontaggio stand', 'SMONTAGGIO ARREDI E IMBALLAGGIO', 0.0, 'NR'),
    ('Smontaggio stand', 'RITIRO MATERIALE CON FURGONE / CAMION', 0.0, 'NR'),
    ('Assistenza evento', 'ASSISTENZA DURANTE FIERA (GIORNATA INTERA)', 0.0, 'GG'),
    ('Assistenza evento', 'ASSISTENZA DURANTE FIERA (MEZZA GIORNATA)', 0.0, 'NR'),
    ('Assistenza evento', 'PRESIDIO / REPERIBILITA\'  STAND', 0.0, 'GG'),
    ('Assistenza evento', 'SOPRALLUOGO PREVENTIVO IN SEDE FIERA', 0.0, 'NR'),
    ('Assistenza evento', 'COORDINAMENTO ALLESTITORI IN LOCO', 0.0, 'GG'),
    ('Pulizie', 'PULIZIA STAND POST-MONTAGGIO', 0.0, 'MQ'),
    ('Pulizie', 'PULIZIA FINALE POST-FIERA', 0.0, 'MQ'),
    ('Pulizie', 'RACCOLTA E SMALTIMENTO IMBALLAGGI', 0.0, 'NR'),
    ('Trasporti', 'TRASPORTO MATERIALI CON FURGONE (GIORNATA)', 0.0, 'GG'),
    ('Trasporti', 'NOLEGGIO FURGONE / VAN (GIORNATA)', 0.0, 'GG'),
    ('Trasporti', 'SCARICO / CARICO MATERIALI IN FIERA', 0.0, 'NR'),
    ('Rimozione', 'RIMOZIONE MOQUETTE', 0.6, 'MQ'),
    ('Rimozione', 'RIMOZIONE PVC', 0.8, 'MQ'),
    ('Rimozione', 'RIMOZIONE ERBA SINTETICA', 0.8, 'MQ'),
    ('Rimozione', 'RIMOZIONE PARQUET', 2.0, 'MQ'),
    ('Rimozione', 'RIMOZIONE NASTRI E BIADESIVI', 0.3, 'MQ'),
]


def genera_pdf_preventivo(dati):
    """Genera PDF preventivo elegante Accesso Fiere, ritorna bytes."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, HRFlowable, KeepTogether)
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.lib.utils import ImageReader
    except ImportError:
        return None

    BLU   = colors.HexColor('#0f172a')
    ACC   = colors.HexColor('#2563eb')
    GRAY  = colors.HexColor('#f8fafc')
    GRAY2 = colors.HexColor('#e2e8f0')
    TEXTC = colors.HexColor('#1e293b')
    TEXTL = colors.HexColor('#64748b')
    WHITE = colors.white
    LBLUE = colors.HexColor('#eef4ff')
    W, H  = A4
    HEADER_H = 36 * mm
    FOOTER_H = 14 * mm

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=HEADER_H + 8*mm,
        bottomMargin=FOOTER_H + 6*mm)

    def hf(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(BLU)
        canvas.rect(0, H - HEADER_H, W, HEADER_H, fill=1, stroke=0)
        canvas.setFillColor(ACC)
        canvas.rect(0, H - HEADER_H - 1.5*mm, W, 1.5*mm, fill=1, stroke=0)
        try:
            img = ImageReader(_io.BytesIO(_b64.b64decode(LOGO_FHORIZON_B64)))
            lw, lh = 38*mm, 15*mm
            canvas.drawImage(img, W - 18*mm - lw, H - HEADER_H + (HEADER_H - lh)/2,
                             width=lw, height=lh, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
        canvas.setFillColor(WHITE)
        canvas.setFont('Helvetica-Bold', 11)
        canvas.drawString(18*mm, H - 12*mm, dati.get('az_nome', 'F. HORIZON SRLS'))
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#93c5fd'))
        canvas.drawString(18*mm, H - 18*mm, dati.get('az_indirizzo', ''))
        canvas.drawString(18*mm, H - 24*mm,
            'P.IVA: ' + dati.get('az_piva', '') + '   |   ' + dati.get('az_email', ''))
        canvas.setFillColor(colors.HexColor('#f1f5f9'))
        canvas.rect(0, 0, W, FOOTER_H, fill=1, stroke=0)
        canvas.setFillColor(colors.HexColor('#1e3a5f'))
        canvas.rect(0, FOOTER_H - 1*mm, W, 1*mm, fill=1, stroke=0)
        canvas.setFillColor(TEXTL)
        canvas.setFont('Helvetica', 7.5)
        canvas.drawString(18*mm, 5*mm,
            dati.get('az_nome', '') + '  ·  ' + dati.get('az_indirizzo', '') +
            '  ·  P.IVA ' + dati.get('az_piva', ''))
        canvas.drawRightString(W - 18*mm, 5*mm, 'Pag. ' + str(canvas.getPageNumber()))
        canvas.restoreState()

    def S(n, **kw): return ParagraphStyle(n, **kw)
    def P(txt, fn='Helvetica', fs=9, tc=None, al=TA_LEFT):
        tc = tc or TEXTC
        return Paragraph(str(txt), S('_', fontName=fn, fontSize=fs, textColor=tc,
                                      alignment=al, leading=fs * 1.4))

    story = []

    story.append(Table([
        [P('N° ' + dati.get('numero', '---'), 'Helvetica-Bold', 26, BLU)],
        [P(dati.get('oggetto', ''), 'Helvetica', 10, TEXTL)],
    ], colWidths=[W - 36*mm], style=TableStyle([
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING',    (0,0), (-1,-1), 0)])))
    story.append(Spacer(1, 5*mm))

    story.append(Table([
        [P('DATA EMISSIONE', 'Helvetica', 8, TEXTL),
         P('VALIDA FINO AL', 'Helvetica', 8, TEXTL),
         P('')],
        [P(dati.get('data', ''), 'Helvetica-Bold', 11, TEXTC),
         P(dati.get('validita', '30 giorni'), 'Helvetica-Bold', 11, TEXTC),
         P('')],
    ], colWidths=[55*mm, 60*mm, None],
       style=TableStyle([('BOTTOMPADDING', (0,0), (-1,-1), 2),
                          ('TOPPADDING',    (0,0), (-1,-1), 1)])))
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width='100%', thickness=1, color=GRAY2))
    story.append(Spacer(1, 6*mm))

    story.append(P('DESTINATARIO', 'Helvetica-Bold', 8, ACC))
    story.append(Spacer(1, 3*mm))
    cli = dati.get('cliente', {})
    cli_t = Table([
        [P('Azienda / Nome', 'Helvetica', 7.5, TEXTL),
         P('P.IVA / C.F.',   'Helvetica', 7.5, TEXTL),
         P('Indirizzo',      'Helvetica', 7.5, TEXTL)],
        [P(cli.get('nome', '---'),      'Helvetica-Bold', 11, TEXTC),
         P(cli.get('piva', '---'),      'Helvetica',       9, TEXTC),
         P(cli.get('indirizzo', '---'), 'Helvetica',       9, TEXTC)],
    ], colWidths=[72*mm, 50*mm, None])
    cli_t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), GRAY),
        ('BACKGROUND',    (0,1), (-1,1), LBLUE),
        ('TOPPADDING',    (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('LEFTPADDING',   (0,0), (-1,-1), 10),
        ('RIGHTPADDING',  (0,0), (-1,-1), 8),
        ('BOX',           (0,0), (-1,-1), 0.8, colors.HexColor('#cbd5e1')),
        ('INNERGRID',     (0,0), (-1,-1), 0.4, GRAY2),
        ('LINEBELOW',     (0,0), (-1,0),  1,   ACC),
    ]))
    story.append(cli_t)
    story.append(Spacer(1, 7*mm))

    if dati.get('descrizione'):
        story.append(P('OGGETTO DEL SERVIZIO', 'Helvetica-Bold', 8, ACC))
        story.append(Spacer(1, 3*mm))
        story.append(P(dati['descrizione'], 'Helvetica', 9, TEXTL))
        story.append(Spacer(1, 6*mm))

    story.append(P('VOCI DI PREVENTIVO', 'Helvetica-Bold', 8, ACC))
    story.append(Spacer(1, 3*mm))

    tdata = [[
        P('DESCRIZIONE / DATE SERVIZIO', 'Helvetica-Bold', 8.5, WHITE, TA_LEFT),
        P("QTA'",        'Helvetica-Bold', 8.5, WHITE, TA_CENTER),
        P('U.M.',        'Helvetica-Bold', 8.5, WHITE, TA_CENTER),
        P('P. UNIT.',    'Helvetica-Bold', 8.5, WHITE, TA_RIGHT),
        P('TOTALE',      'Helvetica-Bold', 8.5, WHITE, TA_RIGHT),
    ]]
    totale_imp = 0.0
    for v in dati.get('voci', []):
        qty = float(v.get('qty') or 0)
        pu  = float(v.get('pu')  or 0)
        tot = qty * pu
        totale_imp += tot
        desc_txt = v.get('desc', '')
        di = v.get('data_inizio') or ''
        df = v.get('data_fine') or ''
        if di and df and di != df:
            date_txt = f'dal {di}  al {df}'
        elif di:
            date_txt = f'dal {di}'
        else:
            date_txt = ''
        from reportlab.platypus import Paragraph as _Par
        from reportlab.lib.styles import getSampleStyleSheet as _gss
        _styles = _gss()
        desc_style = ParagraphStyle('desc', fontName='Helvetica', fontSize=8.5, textColor=TEXTC)
        date_style = ParagraphStyle('date', fontName='Helvetica-Oblique', fontSize=7.5, textColor=TEXTL)
        if date_txt:
            desc_cell = [_Par(desc_txt, desc_style), _Par(date_txt, date_style)]
        else:
            desc_cell = _Par(desc_txt, desc_style)
        tdata.append([
            desc_cell,
            P('{:g}'.format(qty) if qty else '—', 'Helvetica', 8.5, TEXTL, TA_CENTER),
            P(v.get('um', ''), 'Helvetica', 8.5, TEXTL, TA_CENTER),
            P('€ {:,.2f}'.format(pu) if pu else '—', 'Helvetica', 8.5, TEXTC, TA_RIGHT),
            P('€ {:,.2f}'.format(tot) if (qty and pu) else '—',
              'Helvetica-Bold', 8.5, TEXTC, TA_RIGHT),
        ])

    vt = Table(tdata, colWidths=[94*mm, 18*mm, 16*mm, 26*mm, 30*mm], repeatRows=1)
    vt.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  BLU),
        ('TOPPADDING',    (0,0),  (-1,0),  9),
        ('BOTTOMPADDING', (0,0),  (-1,0),  9),
        ('TOPPADDING',    (0,1),  (-1,-1), 8),
        ('BOTTOMPADDING', (0,1),  (-1,-1), 8),
        ('LEFTPADDING',   (0,0),  (-1,-1), 8),
        ('RIGHTPADDING',  (0,0),  (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, colors.HexColor('#f8fbff')]),
        ('LINEBELOW',     (0,0),  (-1,-2), 0.3, GRAY2),
        ('LINEBELOW',     (0,-1), (-1,-1), 1.5, ACC),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
    ]))
    story.append(vt)
    story.append(Spacer(1, 5*mm))

    iva_pct = float(dati.get('iva') or 0)
    iva_val = totale_imp * iva_pct / 100
    tot_fin = totale_imp + iva_val
    tot_rows = []
    if iva_pct > 0:
        tot_rows += [
            [P('Imponibile', 'Helvetica', 9, TEXTL, TA_RIGHT),
             P('€ {:,.2f}'.format(totale_imp), 'Helvetica', 10, TEXTC, TA_RIGHT)],
            [P('IVA {:.0f}%'.format(iva_pct), 'Helvetica', 9, TEXTL, TA_RIGHT),
             P('€ {:,.2f}'.format(iva_val), 'Helvetica', 10, TEXTC, TA_RIGHT)],
        ]
    tot_rows.append([
        P('TOTALE PREVENTIVO', 'Helvetica-Bold', 10.5, WHITE, TA_RIGHT),
        P('€ {:,.2f}'.format(tot_fin), 'Helvetica-Bold', 14, WHITE, TA_RIGHT),
    ])
    tt = Table(tot_rows, colWidths=[78*mm, 46*mm], hAlign='RIGHT')
    last = len(tot_rows) - 1
    tt.setStyle(TableStyle([
        ('TOPPADDING',    (0,0),    (-1,-1), 6),
        ('BOTTOMPADDING', (0,0),    (-1,-1), 6),
        ('LEFTPADDING',   (0,0),    (-1,-1), 12),
        ('RIGHTPADDING',  (0,0),    (-1,-1), 12),
        ('LINEBELOW',     (0,0),    (-1,-2), 0.4, GRAY2),
        ('BACKGROUND',    (0,last), (-1,last), ACC),
        ('TOPPADDING',    (0,last), (-1,last), 11),
        ('BOTTOMPADDING', (0,last), (-1,last), 11),
    ]))
    story.append(KeepTogether([tt]))
    story.append(Spacer(1, 10*mm))

    note_pag = []
    if dati.get('note'):
        note_pag.append(HRFlowable(width='100%', thickness=0.5, color=GRAY2))
        note_pag.append(Spacer(1, 5*mm))
        note_pag.append(P('NOTE', 'Helvetica-Bold', 8, ACC))
        note_pag.append(Spacer(1, 3*mm))
        note_pag.append(P(dati['note'], 'Helvetica-Oblique', 8.5, TEXTL))
        note_pag.append(Spacer(1, 6*mm))

    note_pag.append(HRFlowable(width='100%', thickness=0.5, color=GRAY2))
    note_pag.append(Spacer(1, 5*mm))
    note_pag.append(P("MODALITA' DI PAGAMENTO", 'Helvetica-Bold', 8, ACC))
    note_pag.append(Spacer(1, 3*mm))
    for linea in dati.get('pagamento_lista', []):
        if linea.strip():
            note_pag.append(P('•  ' + linea.strip(), 'Helvetica', 9, TEXTC))
            note_pag.append(Spacer(1, 2*mm))
    story.extend(note_pag)
    story.append(Spacer(1, 12*mm))

    firma = Table([[
        P('Per accettazione\n\nData ___________________', 'Helvetica', 9, TEXTL),
        P('Firma e timbro cliente\n\n\n______________________', 'Helvetica', 9, TEXTL, TA_CENTER),
        P(dati.get('az_nome', '') + '\n\n\n______________________', 'Helvetica', 9, TEXTL, TA_CENTER),
    ]], colWidths=[None, None, None])
    firma.setStyle(TableStyle([
        ('BOX',           (0,0), (-1,-1), 0.8, colors.HexColor('#cbd5e1')),
        ('INNERGRID',     (0,0), (-1,-1), 0.4, GRAY2),
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#fafbfc')),
        ('TOPPADDING',    (0,0), (-1,-1), 14),
        ('BOTTOMPADDING', (0,0), (-1,-1), 14),
        ('LEFTPADDING',   (0,0), (-1,-1), 14),
        ('RIGHTPADDING',  (0,0), (-1,-1), 14),
    ]))
    story.append(KeepTogether([firma]))

    doc.build(story, onFirstPage=hf, onLaterPages=hf)
    return buf.getvalue()

PREV_LIST_TMPL = """
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px">
  <div>
    <h2 style="margin:0"><i class="fa fa-file-invoice" style="color:var(--accent)"></i> Preventivi</h2>
    <p style="color:var(--text-light);margin:4px 0 0;font-size:14px">Crea e gestisci i preventivi per i clienti</p>
  </div>
  <a href="/preventivi/nuovo" class="btn btn-primary"><i class="fa fa-plus"></i> Nuovo preventivo</a>
</div>
<div class="card">
  <div class="card-body" style="padding:0">
    {% if preventivi %}
    <table class="table">
      <thead><tr>
        <th>Numero</th><th>Cliente</th><th>Oggetto</th><th>Data</th>
        <th style="text-align:right">Totale</th><th>Stato</th><th>Azioni</th>
      </tr></thead>
      <tbody>
      {% for p in preventivi %}
      <tr>
        <td><strong>{{ p.numero }}</strong></td>
        <td>{{ p.cliente_nome or '---' }}</td>
        <td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{{ p.oggetto or '---' }}</td>
        <td>{{ p.data_emissione }}</td>
        <td style="text-align:right;font-weight:700">{{ "%.2f"|format(p.totale_finale or 0) }} EUR</td>
        <td>
          {% if p.stato=='bozza' %}<span class="badge badge-yellow">Bozza</span>
          {% elif p.stato=='inviato' %}<span class="badge badge-blue">Inviato</span>
          {% elif p.stato=='accettato' %}<span class="badge badge-green">Accettato</span>
          {% elif p.stato=='rifiutato' %}<span class="badge badge-red">Rifiutato</span>
          {% else %}<span class="badge">{{ p.stato }}</span>{% endif %}
        </td>
        <td style="display:flex;gap:4px">
          <form method="POST" action="/preventivi/{{ p.id }}/stato" style="display:inline">
            <select name="stato" onchange="this.form.submit()" title="Cambia stato rapidamente"
              style="padding:4px 6px;border-radius:6px;font-size:12px;cursor:pointer;border:1px solid var(--border);font-weight:600;
              background:{% if p.stato=='accettato' %}#dcfce7{% elif p.stato=='rifiutato' %}#fee2e2{% elif p.stato=='inviato' %}#dbeafe{% else %}#f1f5f9{% endif %};
              color:{% if p.stato=='accettato' %}#15803d{% elif p.stato=='rifiutato' %}#dc2626{% elif p.stato=='inviato' %}#1d4ed8{% else %}#475569{% endif %}">
              {% for s in ['bozza','inviato','accettato','rifiutato'] %}
              <option value="{{ s }}" {{ 'selected' if p.stato==s }}>{{ s.capitalize() }}</option>
              {% endfor %}
            </select>
          </form>
          <a href="/preventivi/{{ p.id }}/pdf" class="btn btn-sm btn-danger" target="_blank" title="Scarica PDF"><i class="fa fa-file-pdf"></i></a>
          <a href="/preventivi/{{ p.id }}/duplica" class="btn btn-sm btn-secondary" title="Duplica"><i class="fa fa-copy"></i></a>
          <a href="/preventivi/{{ p.id }}/modifica" class="btn btn-sm btn-secondary" title="Modifica"><i class="fa fa-pen"></i></a>
          {% if p.stato == 'accettato' %}<a href="/preventivi/{{ p.id }}/crea-evento" class="btn btn-sm" style="background:#16a34a;color:#fff" title="Crea evento"><i class="fa fa-calendar-plus"></i></a>{% endif %}
          <a href="/preventivi/{{ p.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare il preventivo?')" title="Elimina"><i class="fa fa-trash"></i></a>
        </td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    {% else %}
    <div class="empty-state">
      <i class="fa fa-file-invoice"></i>
      <p>Nessun preventivo ancora.<br><a href="/preventivi/nuovo">Crea il primo preventivo</a></p>
    </div>
    {% endif %}
  </div>
</div>
"""
PREV_FORM_TMPL = """
<div style="margin-bottom:16px">
  <a href="/preventivi" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Preventivi</a>
</div>
<form method="POST" id="form-prev">
<div class="grid-2" style="gap:16px">

  <div style="display:flex;flex-direction:column;gap:16px">
    <div class="card">
      <div class="card-header"><h3><i class="fa fa-info-circle" style="color:var(--accent)"></i> Dati preventivo</h3></div>
      <div class="card-body">
        <div class="form-row">
          <div class="form-group">
            <label>Numero *</label>
            <input name="numero" value="{{ prev.numero if prev else num_sugg }}" required placeholder="Es. 001/2025">
          </div>
          <div class="form-group">
            <label>Data emissione *</label>
            <input type="date" name="data_emissione" value="{{ prev.data_emissione if prev else oggi }}" required>
          </div>
        </div>
        <div class="form-row">
          <div class="form-group">
            <label>Validita</label>
            <input name="validita" value="{{ prev.validita if prev else '30 giorni' }}">
          </div>
          <div class="form-group">
            <label>Stato</label>
            <select name="stato">
              {% for s in ['bozza','inviato','accettato','rifiutato'] %}
              <option value="{{ s }}" {{ 'selected' if (prev and prev.stato==s) or (not prev and s=='bozza') }}>{{ s.capitalize() }}</option>
              {% endfor %}
            </select>
          </div>
        </div>
        <div class="form-group">
          <label>Oggetto</label>
          <input name="oggetto" value="{{ prev.oggetto if prev else '' }}" placeholder="Es. Preventivo per servizio di posa e rimozione pavimentazioni">
        </div>
        <div class="form-group">
          <label>Descrizione servizio</label>
          <textarea name="descrizione_servizio" rows="3" placeholder="Descrivi il servizio offerto...">{{ prev.descrizione_servizio if prev else '' }}</textarea>
        </div>
        <div style="margin-top:16px;padding:14px 16px;background:#f0fdf4;border:1px solid #86efac;border-radius:10px">
          <div style="font-size:12px;font-weight:700;color:#16a34a;margin-bottom:10px;text-transform:uppercase;letter-spacing:.5px">
            <i class="fa fa-calendar-days"></i> Date Evento / Fiera
            <span style="font-weight:400;color:#64748b;text-transform:none;letter-spacing:0;margin-left:6px">→ appariranno sul PDF e serviranno per creare l'evento sul calendario</span>
          </div>
          <div class="form-row">
            <div class="form-group">
              <label style="color:#16a34a;font-weight:600">📅 Data inizio evento</label>
              <input type="date" name="data_inizio_lavoro" value="{{ prev.data_inizio_lavoro if prev else '' }}" style="border-color:#86efac">
            </div>
            <div class="form-group">
              <label style="color:#16a34a;font-weight:600">📅 Data fine evento</label>
              <input type="date" name="data_fine_lavoro" value="{{ prev.data_fine_lavoro if prev else '' }}" style="border-color:#86efac">
            </div>
          </div>
          <div class="form-group" style="margin-bottom:0">
            <label>📍 Sede / Luogo evento</label>
            <input name="luogo_lavoro" value="{{ prev.luogo_lavoro if prev else '' }}" placeholder="Es. Fiera di Bologna — Padiglione 36">
          </div>
          {% if prev and prev.stato == 'accettato' %}
          <div style="margin-top:10px;padding:8px 12px;background:#dcfce7;border-radius:8px;font-size:13px;color:#15803d">
            <i class="fa fa-calendar-plus"></i> <strong>Preventivo accettato!</strong>
            <a href="/preventivi/{{ prev.id }}/crea-evento" style="color:#15803d;font-weight:700;margin-left:8px;text-decoration:underline">Crea evento sul calendario →</a>
          </div>
          {% endif %}
        </div>
      </div>
    </div>


    <div class="card">
      <div class="card-header"><h3><i class="fa fa-building" style="color:var(--accent)"></i> Cliente</h3></div>
      <div class="card-body">
        <div class="form-group">
          <label>Seleziona da anagrafica</label>
          <select onchange="caricaCliente(this.value)">
            <option value="">--- oppure compila manualmente ---</option>
            {% for c in clienti %}
            <option value="{{ c.id }}" {{ 'selected' if prev and prev.cliente_id==c.id }}>
              {{ c.nome }}{% if c.partita_iva %} ({{ c.partita_iva }}){% endif %}
            </option>
            {% endfor %}
          </select>
        </div>
        <input type="hidden" name="cliente_id" id="inp-cli-id" value="{{ prev.cliente_id if prev else '' }}">
        <div class="form-row">
          <div class="form-group">
            <label>Nome / Azienda *</label>
            <input name="cliente_nome" id="cli_nome" value="{{ prev.cliente_nome if prev else '' }}" required placeholder="Es. Fiera Milano SPA">
          </div>
          <div class="form-group">
            <label>P.IVA / C.F.</label>
            <input name="cliente_piva" id="cli_piva" value="{{ prev.cliente_piva if prev else '' }}" placeholder="12345678901">
          </div>
        </div>
        <div class="form-group">
          <label>Indirizzo</label>
          <input name="cliente_indirizzo" id="cli_ind" value="{{ prev.cliente_indirizzo if prev else '' }}" placeholder="Es. Via Milano 1, 20100 Milano">
        </div>
      </div>
    </div>
  </div>

  <div style="display:flex;flex-direction:column;gap:16px">
    <div class="card">
      <div class="card-header"><h3><i class="fa fa-percent" style="color:var(--accent)"></i> IVA e totali</h3></div>
      <div class="card-body">
        <div class="form-group">
          <label>Aliquota IVA</label>
          <select name="iva" id="sel-iva" onchange="ricalcola()">
            <option value="0">Esente (0%)</option>
            <option value="4">4%</option>
            <option value="10">10%</option>
            <option value="22" {{ 'selected' if not prev or (prev.iva|int)==22 }}>22%</option>
          </select>
        </div>
        <div style="background:#f8fafc;border:1px solid var(--border);border-radius:10px;padding:16px;margin-top:8px">
          <div style="display:flex;justify-content:space-between;margin-bottom:8px;font-size:14px">
            <span style="color:var(--text-light)">Imponibile:</span><strong id="tot-imp">0.00</strong>
          </div>
          <div id="tot-sconto" style="display:flex;justify-content:space-between;margin-bottom:8px;font-size:13px;color:#dc2626;font-weight:700"></div>
          <div style="display:flex;justify-content:space-between;margin-bottom:8px;font-size:14px">
            <span style="color:var(--text-light)">IVA:</span><strong id="tot-iva">0.00</strong>
          </div>
          <div style="display:flex;justify-content:space-between;padding-top:10px;border-top:2px solid var(--accent);font-size:17px">
            <strong>TOTALE:</strong><strong id="tot-fin" style="color:var(--accent)">0.00</strong>
          </div>
        </div>
        <input type="hidden" name="totale_finale" id="inp-tot">
        <input type="hidden" name="imponibile" id="inp-imp">
      </div>
    </div>

    <div class="card">
      <div class="card-header"><h3><i class="fa fa-credit-card" style="color:var(--accent)"></i> Modalita di pagamento</h3></div>
      <div class="card-body">
        <div class="form-group">
          <label>Condizioni (una per riga)</label>
          <textarea name="pagamento" rows="3" placeholder="50% alla firma&#10;50% al termine del montaggio">{{ prev.pagamento if prev else '50% alla firma del presente preventivo\n50% al termine del montaggio' }}</textarea>
        </div>
      </div>
    </div>

    <div class="card">
      <div class="card-header"><h3><i class="fa fa-sticky-note" style="color:var(--accent)"></i> Note</h3></div>
      <div class="card-body">
        <textarea name="note" rows="3" placeholder="Note aggiuntive...">{{ prev.note if prev else "I prezzi indicati sono da intendersi per singola unita' di misura. Il preventivo e' valido 30 giorni dalla data di emissione." }}</textarea>
      </div>
    </div>
  </div>
</div>

<div class="card" style="margin-top:16px">
  <div class="card-header"><h3><i class="fa fa-list" style="color:var(--accent)"></i> Voci di preventivo</h3></div>
  <div class="card-body">
    <div style="background:#f0f7ff;border:1px solid #bfdbfe;border-radius:10px;padding:14px;margin-bottom:16px">
      <div style="font-size:12px;font-weight:700;color:#1e40af;margin-bottom:10px;text-transform:uppercase;letter-spacing:.5px">
        <i class="fa fa-bolt"></i> Aggiungi voce predefinita
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
        <select id="sel-pred" style="flex:1;min-width:280px">
          <option value="">--- Seleziona voce predefinita ---</option>
          {% set ns = namespace(cat='') %}
          {% for cat, desc, prezzo, um in voci_pred %}
            {% if cat != ns.cat %}
              {% if ns.cat %}</optgroup>{% endif %}
              <optgroup label="{{ cat }}">
              {% set ns.cat = cat %}
            {% endif %}
            <option value="{{ desc }}|||{{ prezzo }}|||{{ um }}">{{ desc }} -- {{ prezzo }}/{{ um }}</option>
          {% endfor %}
          {% if ns.cat %}</optgroup>{% endif %}
        </select>
        <button type="button" class="btn btn-primary btn-sm" onclick="aggiungiPred()"><i class="fa fa-plus"></i> Aggiungi</button>
        <button type="button" class="btn btn-secondary btn-sm" onclick="aggiungiLibera()"><i class="fa fa-pen"></i> Riga libera</button>
      </div>
    </div>
    <div style="overflow-x:auto">
      <table style="width:100%;border-collapse:collapse" id="tab-voci">
        <thead>
          <tr style="background:#0f172a;color:#fff">
            <th style="padding:10px 8px;text-align:left;font-size:12px;width:28%">DESCRIZIONE</th>
            <th style="padding:10px 8px;text-align:center;font-size:12px;width:8%">DAL</th>
            <th style="padding:10px 8px;text-align:center;font-size:12px;width:8%">AL</th>
            <th style="padding:10px 8px;text-align:center;font-size:12px;width:7%">QTA</th>
            <th style="padding:10px 8px;text-align:center;font-size:12px;width:6%">U.M.</th>
            <th style="padding:10px 8px;text-align:right;font-size:12px;width:10%">P.UNIT.</th>
            <th style="padding:10px 8px;text-align:right;font-size:12px;width:11%;background:#1e3a5f">IMPORTO <span style="font-size:9px;opacity:.7">(modificabile)</span></th>
            <th style="padding:10px 8px;text-align:right;font-size:12px;width:10%;background:#7f1d1d">SCONTO</th>
            <th style="width:4%"></th>
          </tr>
        </thead>
        <tbody id="voci-body">
          {% for v in prev_voci %}
          {% set lordo = (v.quantita or 0)*(v.prezzo_unitario or 0) %}
          {% set imp_mod = v.importo_modificato if v.importo_modificato is not none else lordo %}
          {% set sconto_r = lordo - imp_mod %}
          <tr class="vr" style="border-bottom:1px solid #e2e8f0">
            <td style="padding:4px"><input name="desc[]" value="{{ v.descrizione }}" class="ti" required></td>
            <td style="padding:4px"><input name="vdata_inizio[]" type="date" value="{{ v.data_inizio or '' }}" class="ti" style="font-size:11px"></td>
            <td style="padding:4px"><input name="vdata_fine[]" type="date" value="{{ v.data_fine or '' }}" class="ti" style="font-size:11px"></td>
            <td style="padding:4px"><input name="qty[]" type="number" value="{{ v.quantita }}" step="0.001" min="0" class="ti tr" oninput="rc(this)"></td>
            <td style="padding:4px"><input name="um[]" value="{{ v.unita_misura }}" class="ti tc" placeholder="MQ"></td>
            <td style="padding:4px"><input name="pu[]" type="number" value="{{ v.prezzo_unitario }}" step="0.001" min="0" class="ti tr" oninput="rc(this)"></td>
            <td style="padding:4px"><input name="importo_mod[]" type="number" value="{{ '%.2f'|format(imp_mod) }}" step="0.01" min="0" class="ti tr td-imp" oninput="rcImp(this)" style="font-weight:700;color:#1e3a5f;background:#eff6ff;border-color:#93c5fd"></td>
            <td style="padding:8px;text-align:right;font-size:13px;font-weight:600" class="td-sc" style="color:{% if sconto_r > 0 %}#dc2626{% else %}#94a3b8{% endif %}">
              {% if sconto_r > 0 %}<span style="color:#dc2626">-{{ '%.2f'|format(sconto_r) }}</span>{% else %}<span style="color:#94a3b8">—</span>{% endif %}
            </td>
            <td style="padding:4px;text-align:center"><button type="button" onclick="delRow(this)" style="background:#fee2e2;border:none;border-radius:6px;padding:6px 10px;cursor:pointer;color:#dc2626"><i class="fa fa-times"></i></button></td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
    <div id="empty-voci" style="text-align:center;padding:30px;color:var(--text-light);{% if prev_voci %}display:none{% endif %}">
      <i class="fa fa-list" style="font-size:30px;margin-bottom:8px;display:block;opacity:.4"></i>
      Nessuna voce. Usa i pulsanti sopra per aggiungere.
    </div>
  </div>
</div>

<div style="display:flex;justify-content:flex-end;gap:12px;margin-top:20px;padding-bottom:20px">
  <a href="/preventivi" class="btn btn-secondary">Annulla</a>
  <button type="submit" name="azione" value="salva" class="btn btn-secondary"><i class="fa fa-save"></i> Salva bozza</button>
  <button type="submit" name="azione" value="pdf" class="btn btn-primary"><i class="fa fa-file-pdf"></i> Salva e scarica PDF</button>
</div>
</form>

<style>
.ti{width:100%;border:1px solid transparent;background:transparent;font-size:13px;padding:7px 6px;border-radius:6px;outline:none;font-family:inherit;color:#1e293b}
.ti:focus{border-color:#3b82f6;background:#eff6ff}
.tr{text-align:right}.tc{text-align:center}
.vr:hover{background:#f8faff}.vr:nth-child(even){background:#fafafa}
</style>
<script>
const CLI = {{ clienti_json|safe }};
function caricaCliente(id){
  if(!id)return;const c=CLI[id];if(!c)return;
  document.getElementById('cli_nome').value=c.nome||'';
  document.getElementById('cli_piva').value=c.piva||'';
  document.getElementById('cli_ind').value=c.indirizzo||'';
  document.getElementById('inp-cli-id').value=id;
}
function mkRow(desc,qty,um,pu,di,df){
  const b=document.getElementById('voci-body');
  const tr=document.createElement('tr');tr.className='vr';tr.style.borderBottom='1px solid #e2e8f0';
  tr.innerHTML=
    '<td style="padding:4px"><input name="desc[]" value="'+desc+'" class="ti" required placeholder="Descrizione servizio"></td>'+
    '<td style="padding:4px"><input name="vdata_inizio[]" type="date" value="'+(di||'')+'" class="ti" style="font-size:11px"></td>'+
    '<td style="padding:4px"><input name="vdata_fine[]" type="date" value="'+(df||'')+'" class="ti" style="font-size:11px"></td>'+
    '<td style="padding:4px"><input name="qty[]" type="number" value="'+qty+'" step="0.001" min="0" class="ti tr" oninput="rc(this)" placeholder="0"></td>'+
    '<td style="padding:4px"><input name="um[]" value="'+um+'" class="ti tc" placeholder="MQ"></td>'+
    '<td style="padding:4px"><input name="pu[]" type="number" value="'+pu+'" step="0.001" min="0" class="ti tr" oninput="rc(this)" placeholder="0.00"></td>'+
    '<td style="padding:4px"><input name="importo_mod[]" type="number" value="0.00" step="0.01" min="0" class="ti tr td-imp" oninput="rcImp(this)" style="font-weight:700;color:#1e3a5f;background:#eff6ff;border-color:#93c5fd"></td>'+
    '<td style="padding:8px;text-align:right;font-size:13px;font-weight:600" class="td-sc"><span style="color:#94a3b8">—</span></td>'+
    '<td style="padding:4px;text-align:center"><button type="button" onclick="delRow(this)" style="background:#fee2e2;border:none;border-radius:6px;padding:6px 10px;cursor:pointer;color:#dc2626"><i class="fa fa-times"></i></button></td>';
  b.appendChild(tr);document.getElementById('empty-voci').style.display='none';
  rc(tr.querySelector('[name="qty[]"]'));tr.querySelector('input[name="desc[]"]').focus();
}
function aggiungiPred(){
  const s=document.getElementById('sel-pred');if(!s.value)return;
  const p=s.value.split('|||');mkRow(p[0],'',p[2],p[1]);s.value='';
}
function aggiungiLibera(){mkRow('','','','');}
function delRow(btn){
  btn.closest('tr').remove();ricalcola();
  if(!document.querySelector('.vr'))document.getElementById('empty-voci').style.display='';
}
function rc(inp){
  // Quando cambio qty o pu: ricalcolo il lordo e aggiorno l'importo modificato
  const tr=inp.closest('tr');
  const q=parseFloat(tr.querySelector('input[name="qty[]"]').value)||0;
  const p=parseFloat(tr.querySelector('input[name="pu[]"]').value)||0;
  const lordo=q*p;
  const impInp=tr.querySelector('.td-imp');
  // Aggiorna importo modificato solo se non era già cambiato manualmente
  if(!impInp.dataset.manuale) impInp.value=lordo.toFixed(2);
  aggiornaSconto(tr, lordo);
  ricalcola();
}
function rcImp(inp){
  // Quando cambio l'importo manualmente: calcola sconto
  inp.dataset.manuale='1';
  const tr=inp.closest('tr');
  const q=parseFloat(tr.querySelector('input[name="qty[]"]').value)||0;
  const p=parseFloat(tr.querySelector('input[name="pu[]"]').value)||0;
  const lordo=q*p;
  aggiornaSconto(tr, lordo);
  ricalcola();
}
function aggiornaSconto(tr, lordo){
  const imp=parseFloat(tr.querySelector('.td-imp').value)||0;
  const sc=lordo-imp;
  const td=tr.querySelector('.td-sc');
  if(sc>0.005){
    td.innerHTML='<span style="color:#dc2626;font-weight:700">-'+sc.toFixed(2)+'</span>';
  } else if(sc<-0.005){
    td.innerHTML='<span style="color:#16a34a;font-weight:700">+'+Math.abs(sc).toFixed(2)+'</span>';
  } else {
    td.innerHTML='<span style="color:#94a3b8">—</span>';
  }
}
function ricalcola(){
  let imp=0,scTot=0;
  document.querySelectorAll('.vr').forEach(tr=>{
    const q=parseFloat(tr.querySelector('input[name="qty[]"]')?.value)||0;
    const p=parseFloat(tr.querySelector('input[name="pu[]"]')?.value)||0;
    const lordo=q*p;
    const mod=parseFloat(tr.querySelector('.td-imp')?.value)||lordo;
    imp+=mod;
    scTot+=lordo-mod;
  });
  const ip=parseFloat(document.getElementById('sel-iva').value)||0;
  const iv=imp*ip/100,tot=imp+iv;
  document.getElementById('tot-imp').textContent=imp.toFixed(2);
  document.getElementById('tot-iva').textContent=iv.toFixed(2);
  document.getElementById('tot-fin').textContent=tot.toFixed(2);
  document.getElementById('inp-tot').value=tot.toFixed(2);
  document.getElementById('inp-imp').value=imp.toFixed(2);
  // Mostra sconto totale se presente
  const scEl=document.getElementById('tot-sconto');
  if(scEl){scEl.textContent=scTot>0.01?'Sconto totale: -€'+scTot.toFixed(2):'';}
}
document.addEventListener('DOMContentLoaded',()=>{
  {% if prev and prev.iva is not none %}document.getElementById('sel-iva').value='{{ prev.iva|int }}';{% endif %}
  ricalcola();
});
</script>
"""

@app.route('/preventivi')
@admin_required
def preventivi():
    db = get_db()
    prevs = db.execute("SELECT * FROM preventivi ORDER BY creato_il DESC").fetchall()
    db.close()
    return render_page(PREV_LIST_TMPL, page_title='Preventivi', active='preventivi', preventivi=prevs)


@app.route('/preventivi/nuovo', methods=['GET','POST'])
@admin_required
def preventivo_nuovo():
    db = get_db()
    if request.method == 'POST':
        pid = _salva_preventivo(db, None)
        db.close()
        if request.form.get('azione') == 'pdf':
            return redirect(url_for('preventivo_pdf', pid=pid))
        flash('Preventivo salvato!', 'success')
        return redirect(url_for('preventivi'))
    clienti = db.execute("SELECT * FROM clienti ORDER BY nome").fetchall()
    ultimo  = db.execute("SELECT numero FROM preventivi ORDER BY id DESC LIMIT 1").fetchone()
    db.close()
    num_sugg = _next_numero(ultimo['numero'] if ultimo else None)
    return render_page(PREV_FORM_TMPL, page_title='Nuovo Preventivo', active='preventivi',
        prev=None, prev_voci=[], clienti=clienti, num_sugg=num_sugg,
        oggi=date.today().isoformat(), voci_pred=VOCI_PREDEFINITE,
        clienti_json=_clienti_json(clienti))


@app.route('/preventivi/<int:pid>/modifica', methods=['GET','POST'])
@admin_required
def preventivo_modifica(pid):
    db = get_db()
    prev = db.execute("SELECT * FROM preventivi WHERE id=?", (pid,)).fetchone()
    if not prev:
        db.close(); flash('Non trovato.','error'); return redirect(url_for('preventivi'))
    if request.method == 'POST':
        _salva_preventivo(db, pid)
        db.close()
        if request.form.get('azione') == 'pdf':
            return redirect(url_for('preventivo_pdf', pid=pid))
        flash('Preventivo aggiornato!', 'success')
        return redirect(url_for('preventivi'))
    voci    = db.execute("SELECT * FROM preventivi_voci WHERE preventivo_id=? ORDER BY ordine",(pid,)).fetchall()
    clienti = db.execute("SELECT * FROM clienti ORDER BY nome").fetchall()
    db.close()
    return render_page(PREV_FORM_TMPL, page_title='Modifica Preventivo', active='preventivi',
        prev=prev, prev_voci=voci, clienti=clienti, num_sugg=prev['numero'],
        oggi=date.today().isoformat(), voci_pred=VOCI_PREDEFINITE,
        clienti_json=_clienti_json(clienti))


@app.route('/preventivi/<int:pid>/pdf')
@admin_required
def preventivo_pdf(pid):
    db   = get_db()
    prev = db.execute("SELECT * FROM preventivi WHERE id=?", (pid,)).fetchone()
    if not prev:
        db.close(); flash('Non trovato.','error'); return redirect(url_for('preventivi'))
    voci = db.execute("SELECT * FROM preventivi_voci WHERE preventivo_id=? ORDER BY ordine",(pid,)).fetchall()
    db.close()
    dati = {
        'az_nome':      get_setting('azienda','F. HORIZON SRLS'),
        'az_indirizzo': get_setting('sede_legale','Via Luciano Manara 35, 37045 Legnago (VR)'),
        'az_piva':      get_setting('partita_iva','05029320230'),
        'az_email':     get_setting('smtp_user','info.fhorizon@gmail.com'),
        'numero':       prev['numero'],
        'data':         prev['data_emissione'] or '',
        'validita':     prev['validita'] or '30 giorni',
        'oggetto':      prev['oggetto'] or '',
        'descrizione':  prev['descrizione_servizio'] or '',
        'data_inizio_lavoro': prev['data_inizio_lavoro'] or '',
        'data_fine_lavoro':   prev['data_fine_lavoro'] or '',
        'luogo_lavoro':       prev['luogo_lavoro'] or '',
        'cliente': {
            'nome':      prev['cliente_nome'] or '---',
            'piva':      prev['cliente_piva'] or '---',
            'indirizzo': prev['cliente_indirizzo'] or '---',
        },
        'voci': [{'desc': v['descrizione'], 'qty': v['quantita'],
                  'um': v['unita_misura'], 'pu': v['prezzo_unitario'],
                  'data_inizio': v['data_inizio'] or '', 'data_fine': v['data_fine'] or ''} for v in voci],
        'iva':            prev['iva'] or 22,
        'note':           prev['note'] or '',
        'pagamento_lista':[l for l in (prev['pagamento'] or '').split('\n') if l.strip()],
    }
    pdf_bytes = genera_pdf_preventivo(dati)
    if not pdf_bytes:
        flash('Errore generazione PDF.','error')
        return redirect(url_for('preventivi'))
    nome_file = 'Preventivo_' + prev['numero'].replace('/','_') + '.pdf'
    return Response(pdf_bytes, mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment; filename="' + nome_file + '"'})


@app.route('/preventivi/<int:pid>/duplica')
@admin_required
def preventivo_duplica(pid):
    db   = get_db()
    prev = db.execute("SELECT * FROM preventivi WHERE id=?", (pid,)).fetchone()
    if not prev:
        db.close(); flash('Non trovato.','error'); return redirect(url_for('preventivi'))
    voci   = db.execute("SELECT * FROM preventivi_voci WHERE preventivo_id=? ORDER BY ordine",(pid,)).fetchall()
    ultimo = db.execute("SELECT numero FROM preventivi ORDER BY id DESC LIMIT 1").fetchone()
    nn = _next_numero(ultimo['numero'] if ultimo else None)
    new_id = db.execute(
        """INSERT INTO preventivi (numero,data_emissione,validita,stato,cliente_id,cliente_nome,
           cliente_piva,cliente_indirizzo,oggetto,descrizione_servizio,iva,imponibile,
           totale_finale,note,pagamento)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (nn, date.today().isoformat(), prev['validita'], 'bozza',
         prev['cliente_id'], prev['cliente_nome'], prev['cliente_piva'], prev['cliente_indirizzo'],
         prev['oggetto'], prev['descrizione_servizio'], prev['iva'], prev['imponibile'],
         prev['totale_finale'], prev['note'], prev['pagamento'])).lastrowid
    for i, v in enumerate(voci):
        db.execute("INSERT INTO preventivi_voci (preventivo_id,descrizione,quantita,unita_misura,prezzo_unitario,totale,ordine) VALUES (?,?,?,?,?,?,?)",
                   (new_id, v['descrizione'], v['quantita'], v['unita_misura'], v['prezzo_unitario'], v['totale'], i))
    safe_commit(db); db.close()
    flash('Preventivo duplicato come ' + nn + '!', 'success')
    return redirect(url_for('preventivo_modifica', pid=new_id))


def _rileva_tipo_voce(desc):
    d = (desc or '').upper()
    if 'SMONTAGGIO' in d:                                                 return 'smontaggio'
    if 'MONTAGGIO' in d:                                                  return 'montaggio'
    if any(k in d for k in ('ASSISTENZA','PRESIDIO','REPERIBILITA','FIERA','COORDINAMENTO')): return 'fiera'
    return 'altro'


@app.route('/preventivi/<int:pid>/crea-evento')
@admin_required
def preventivo_crea_evento(pid):
    db = get_db()
    prev = db.execute("SELECT * FROM preventivi WHERE id=?", (pid,)).fetchone()
    if not prev:
        db.close(); flash('Preventivo non trovato.','error')
        return redirect(url_for('preventivi'))
    if prev['stato'] != 'accettato':
        db.close(); flash('Solo i preventivi accettati possono generare un evento.','warning')
        return redirect(url_for('preventivi'))
    voci_raw = db.execute(
        "SELECT id, descrizione, quantita, unita_misura, data_inizio, data_fine FROM preventivi_voci WHERE preventivo_id=? ORDER BY ordine",
        (pid,)).fetchall()
    db.close()

    # Raggruppa voci per tipo — una card per tipo con date min/max
    from collections import defaultdict
    gruppi = defaultdict(lambda: {'voci': [], 'date_inizio': [], 'date_fine': []})
    for v in voci_raw:
        tipo = _rileva_tipo_voce(v['descrizione'])
        gruppi[tipo]['voci'].append(v['descrizione'])
        if v['data_inizio']: gruppi[tipo]['date_inizio'].append(v['data_inizio'])
        if v['data_fine']:   gruppi[tipo]['date_fine'].append(v['data_fine'])

    # Se non ci sono voci con date, metti il preventivo come unico gruppo
    if not gruppi:
        gruppi['altro']['voci'].append(prev['oggetto'] or '')

    eventi_suggeriti = []
    tipo_labels = {'montaggio':'🔧 Montaggio','smontaggio':'🔩 Smontaggio','fiera':'🏛️ Fiera/Assistenza','altro':'📌 Servizio'}
    tipo_colori  = {'montaggio':'#16a34a','smontaggio':'#e63946','fiera':'#2196F3','altro':'#8b5cf6'}
    for tipo, g in gruppi.items():
        d_inizio = min(g['date_inizio']) if g['date_inizio'] else (prev['data_inizio_lavoro'] or '')
        d_fine   = max(g['date_fine'])   if g['date_fine']   else (prev['data_fine_lavoro'] or d_inizio)
        titolo_sug = f"{tipo_labels.get(tipo,'Servizio')} — {prev['oggetto'] or prev['numero']}"
        eventi_suggeriti.append({
            'tipo':     tipo,
            'label':    tipo_labels.get(tipo,'Servizio'),
            'colore':   tipo_colori.get(tipo,'#8b5cf6'),
            'titolo':   titolo_sug,
            'd_inizio': d_inizio,
            'd_fine':   d_fine,
            'voci':     g['voci'],
        })
    # Ordine: montaggio → fiera → smontaggio → altro
    _ord = {'montaggio':0,'fiera':1,'smontaggio':2,'altro':3}
    eventi_suggeriti.sort(key=lambda x: _ord.get(x['tipo'],9))

    # Se il preventivo ha date evento/fiera, assicurati che ci sia sempre un evento Fiera
    # con quelle date (ha priorità sulle voci)
    d_inizio_fiera = prev['data_inizio_lavoro'] or ''
    d_fine_fiera   = prev['data_fine_lavoro']   or d_inizio_fiera
    luogo_fiera    = prev['luogo_lavoro']        or ''
    if d_inizio_fiera:
        # Aggiorna o inserisci evento fiera con le date del preventivo
        fiera_esistente = next((e for e in eventi_suggeriti if e['tipo'] == 'fiera'), None)
        titolo_fiera = f"🏛️ Fiera — {prev['oggetto'] or prev['numero']}"
        if luogo_fiera:
            titolo_fiera = f"🏛️ {luogo_fiera}"
        if fiera_esistente:
            fiera_esistente['d_inizio'] = d_inizio_fiera
            fiera_esistente['d_fine']   = d_fine_fiera
            fiera_esistente['titolo']   = titolo_fiera
        else:
            eventi_suggeriti.insert(0, {
                'tipo':     'fiera',
                'label':    '🏛️ Fiera/Evento',
                'colore':   '#2196F3',
                'titolo':   titolo_fiera,
                'd_inizio': d_inizio_fiera,
                'd_fine':   d_fine_fiera,
                'voci':     [],
            })
        eventi_suggeriti.sort(key=lambda x: _ord.get(x['tipo'],9))

    from urllib.parse import urlencode
    tmpl = """
<div style="max-width:700px;margin:0 auto">
  <div style="margin-bottom:18px">
    <a href="/preventivi" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Preventivi</a>
  </div>
  <div class="card">
    <div class="card-header">
      <h3><i class="fa fa-calendar-plus" style="color:var(--accent2)"></i> Crea eventi dal preventivo {{ prev.numero }}</h3>
      <div style="color:var(--text-light);font-size:13px;margin-top:4px">{{ prev.cliente_nome }} — {{ prev.oggetto }}</div>
    </div>
    <div class="card-body">
      <p style="color:var(--text-light);margin-bottom:20px">Abbiamo rilevato <strong>{{ eventi|length }} tipo/i di servizio</strong> nelle voci del preventivo. Clicca su un pulsante per creare il relativo evento sul calendario, con le date già compilate.</p>
      {% for e in eventi %}
      <div style="background:#fff;border:2px solid {{ e.colore }};border-radius:12px;padding:18px;margin-bottom:14px;display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap">
        <div>
          <div style="font-size:18px;font-weight:800;color:{{ e.colore }};margin-bottom:4px">{{ e.label }}</div>
          <div style="font-size:13px;color:var(--text);margin-bottom:6px">{{ e.titolo }}</div>
          {% if e.d_inizio %}
          <div style="font-size:12px;color:var(--text-light)"><i class="fa fa-calendar" style="margin-right:4px"></i>
            {{ e.d_inizio }}{% if e.d_fine and e.d_fine != e.d_inizio %} → {{ e.d_fine }}{% endif %}
          </div>
          {% endif %}
          <div style="margin-top:6px;font-size:11px;color:#94a3b8">
            {% for v in e.voci %}<span style="background:#f1f5f9;padding:2px 7px;border-radius:4px;margin-right:4px">{{ v }}</span>{% endfor %}
          </div>
        </div>
        <a href="/calendario/nuovo?{{ e.url }}" class="btn btn-primary" style="background:{{ e.colore }};border-color:{{ e.colore }};white-space:nowrap">
          <i class="fa fa-calendar-plus"></i> Crea evento
        </a>
      </div>
      {% endfor %}
    </div>
  </div>
</div>"""

    # Aggiungi url a ciascun evento
    for e in eventi_suggeriti:
        p = {k: v for k, v in {
            'prev_id':     pid,
            'titolo':      e['titolo'],
            'tipo':        e['tipo'],
            'data_inizio': e['d_inizio'],
            'data_fine':   e['d_fine'],
            'luogo':       prev['luogo_lavoro'] or '',
        }.items() if v}
        e['url'] = urlencode(p)

    return render_page(tmpl, page_title='Crea eventi dal preventivo',
        active='preventivi', prev=prev, eventi=eventi_suggeriti)


@app.route('/preventivi/<int:pid>/stato', methods=['POST'])
@admin_required
def preventivo_stato(pid):
    nuovo_stato = request.form.get('stato','bozza')
    db = get_db()
    prev_row = db.execute("SELECT stato FROM preventivi WHERE id=?", (pid,)).fetchone()
    stato_precedente = prev_row['stato'] if prev_row else None
    db.execute("UPDATE preventivi SET stato=? WHERE id=?", (nuovo_stato, pid))
    safe_commit(db)
    if nuovo_stato == 'accettato' and stato_precedente != 'accettato':
        _crea_fatture_da_preventivo(db, pid)
    db.close()
    flash(f'Stato aggiornato: {nuovo_stato.capitalize()}', 'success')
    return redirect(url_for('preventivi'))


@app.route('/preventivi/<int:pid>/elimina')
@admin_required
def preventivo_elimina(pid):
    db = get_db()
    # Elimina fatture collegate (e relative rate)
    fatture_ids = [r[0] for r in db.execute(
        "SELECT id FROM fatture WHERE note LIKE ?", (f'%preventivo:{pid}%',)).fetchall()]
    for fid in fatture_ids:
        db.execute("DELETE FROM rate_fattura WHERE fattura_id=?", (fid,))
    if fatture_ids:
        db.execute("DELETE FROM fatture WHERE id IN ({})".format(','.join('?'*len(fatture_ids))), fatture_ids)
    db.execute("DELETE FROM preventivi_voci WHERE preventivo_id=?", (pid,))
    db.execute("DELETE FROM preventivi WHERE id=?", (pid,))
    safe_commit(db); db.close()
    flash('Preventivo eliminato.', 'success')
    return redirect(url_for('preventivi'))


def _next_numero(last_num):
    anno = date.today().year
    if last_num and '/' in str(last_num):
        try:
            n = int(str(last_num).split('/')[0]) + 1
            return '{:03d}/{}'.format(n, anno)
        except Exception:
            pass
    return '001/{}'.format(anno)


def _next_numero_fattura(db):
    """Genera il prossimo numero fattura offline."""
    row = db.execute("SELECT numero FROM fatture ORDER BY id DESC LIMIT 1").fetchone()
    anno = date.today().year
    if row and row['numero']:
        try:
            parti = str(row['numero']).split('/')
            n = int(parti[0]) + 1
            return '{:03d}/{}'.format(n, anno)
        except Exception:
            pass
    return '001/{}'.format(anno)


def _parse_condizioni_pagamento(pagamento_text, totale, iva_perc, data_inizio):
    """
    Analizza il testo delle condizioni di pagamento.
    Gestisce: testo su una riga con più % separate da virgola/punto,
    testo su righe separate, o testo senza percentuali.
    """
    import re as _re
    from datetime import date as _date, timedelta as _td

    testo = (pagamento_text or '').strip()
    if not testo:
        return [{'descrizione': 'Pagamento', 'percentuale': 100,
                 'importo': totale, 'data_scadenza': _date.today().isoformat()}]

    # Trova tutte le occorrenze di "NN%" nel testo completo
    matches = list(_re.finditer(r'(\d+)\s*%', testo))

    if not matches:
        # Nessuna percentuale: unica fattura per il totale
        return [{'descrizione': testo, 'percentuale': 100,
                 'importo': totale, 'data_scadenza': _date.today().isoformat()}]

    oggi = _date.today()
    rate = []

    for m in matches:
        pct = float(m.group(1))
        # Prendi il contesto intorno al match (50 caratteri dopo) per determinare la scadenza
        ctx = testo[max(0, m.start()-30):min(len(testo), m.end()+60)].lower()

        if any(k in ctx for k in ['firma', 'acconto', 'anticipo', 'prima', 'inizio']):
            scadenza = oggi.isoformat()
        elif any(k in ctx for k in ['consegna', 'saldo', 'fine', 'termine', 'completamento']):
            scadenza = data_inizio or (oggi + _td(days=30)).isoformat()
        elif any(k in ctx for k in ['evento', 'fiera', 'lavoro', 'montaggio']):
            scadenza = data_inizio or (oggi + _td(days=7)).isoformat()
        else:
            scadenza = (oggi + _td(days=30)).isoformat()

        # Descrizione: pezzo di testo vicino al match
        start_ctx = max(0, m.start() - 5)
        end_ctx   = min(len(testo), m.end() + 50)
        desc_raw  = testo[start_ctx:end_ctx].strip().strip(',').strip()

        importo = round(totale * pct / 100, 2)
        rate.append({'descrizione': desc_raw, 'percentuale': pct,
                     'importo': importo, 'data_scadenza': scadenza})

    # Correggi arrotondamenti sull'ultima rata
    somma = sum(r['importo'] for r in rate)
    if abs(somma - totale) > 0.01:
        rate[-1]['importo'] = round(totale - sum(r['importo'] for r in rate[:-1]), 2)

    return rate


def _crea_fatture_da_preventivo(db, preventivo_id):
    """Crea automaticamente fatture offline quando un preventivo viene accettato."""
    try:
        prev = db.execute("SELECT * FROM preventivi WHERE id=?", (preventivo_id,)).fetchone()
        if not prev:
            return

        totale   = float(prev['totale_finale'] or 0)
        iva_perc = float(prev['iva'] or 22)
        pagamento = prev['pagamento'] or ''
        data_fine = prev['data_fine_lavoro'] or prev['data_inizio_lavoro']

        rate = _parse_condizioni_pagamento(pagamento, totale, iva_perc, data_fine)

        # ── Protezione anti-duplicazione ─────────────────────────
        # 1) Trova TUTTE le fatture collegate a questo preventivo
        #    (sia via preventivo_id che via nota legacy)
        esistenti = db.execute("""
            SELECT id, stato, file_nome FROM fatture
            WHERE preventivo_id=? OR note LIKE ?
        """, (preventivo_id, f'%preventivo:{preventivo_id}%')).fetchall()

        # 2) Conta le fatture "reali" (già emesse = hanno file o stato != da_emettere)
        #    → vanno preservate sempre
        reali = [r for r in esistenti
                 if r['stato'] != 'da_emettere' or (r['file_nome'] and r['file_nome'].strip())]

        # 3) Se le fatture reali coprono già tutte le rate → non fare nulla
        if len(reali) >= len(rate):
            return

        # 4) Cancella SOLO le offline pure (da_emettere senza file)
        for r in esistenti:
            if r['stato'] == 'da_emettere' and not (r['file_nome'] and r['file_nome'].strip()):
                db.execute("DELETE FROM rate_fattura WHERE fattura_id=?", (r['id'],))
                db.execute("DELETE FROM fatture WHERE id=?", (r['id'],))
        safe_commit(db)

        # 5) Crea solo le rate mancanti (totale rate - reali già presenti)
        n_da_creare = len(rate) - len(reali)
        for i, rata in enumerate(rate[:n_da_creare], len(reali) + 1):
            imp_rata = round(rata['importo'] / (1 + iva_perc/100), 2)
            iva_imp  = round(rata['importo'] - imp_rata, 2)
            numero   = _next_numero_fattura(db)
            desc     = f"{prev['oggetto'] or 'Lavoro'}"
            if len(rate) > 1:
                desc += f" — Rata {i}/{len(rate)}: {rata['descrizione']}"
            else:
                desc += f" — {rata['descrizione']}"

            db.execute("""INSERT INTO fatture
                (numero, cliente_id, cliente_nome, data_emissione, data_scadenza,
                 importo_totale, imponibile, iva_perc, iva_importo,
                 descrizione, note, condizioni_pagamento, stato, preventivo_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                numero,
                prev['cliente_id'],
                prev['cliente_nome'] or '',
                date.today().isoformat(),
                rata['data_scadenza'],
                rata['importo'],
                imp_rata,
                iva_perc,
                iva_imp,
                desc,
                f"Generata automaticamente da preventivo:{preventivo_id} — N.{prev['numero']}",
                rata['descrizione'],
                'da_emettere',
                preventivo_id,
            ))
            safe_commit(db)
    except Exception as _exc:
        import traceback
        print(f"[ERRORE _crea_fatture_da_preventivo] {_exc}")
        traceback.print_exc()


def _clienti_json(clienti_rows):
    d = {}
    for c in clienti_rows:
        d[str(c['id'])] = {
            'nome':       c['nome'],
            'piva':       c['partita_iva'] or '',
            'indirizzo':  c['indirizzo'] or '',
        }
    return _json.dumps(d)


def _salva_preventivo(db, pid):
    f     = request.form
    descs = f.getlist('desc[]')
    qtys  = f.getlist('qty[]')
    ums   = f.getlist('um[]')
    pus   = f.getlist('pu[]')
    imp   = float(f.get('imponibile', 0) or 0)
    tot   = float(f.get('totale_finale', 0) or 0)
    stato_nuovo = f.get('stato', 'bozza')
    fields = (
        f.get('numero',''), f.get('data_emissione',''), f.get('validita','30 giorni'),
        stato_nuovo, f.get('cliente_id') or None,
        f.get('cliente_nome',''), f.get('cliente_piva',''), f.get('cliente_indirizzo',''),
        f.get('oggetto',''), f.get('descrizione_servizio',''),
        f.get('data_inizio_lavoro','') or None, f.get('data_fine_lavoro','') or None,
        f.get('luogo_lavoro',''),
        int(f.get('iva',22) or 22), imp, tot,
        f.get('note',''), f.get('pagamento',''),
    )
    stato_precedente = None
    if pid:
        row = db.execute("SELECT stato FROM preventivi WHERE id=?", (pid,)).fetchone()
        stato_precedente = row['stato'] if row else None
        db.execute("""UPDATE preventivi SET numero=?,data_emissione=?,validita=?,stato=?,
            cliente_id=?,cliente_nome=?,cliente_piva=?,cliente_indirizzo=?,oggetto=?,
            descrizione_servizio=?,data_inizio_lavoro=?,data_fine_lavoro=?,luogo_lavoro=?,
            iva=?,imponibile=?,totale_finale=?,note=?,pagamento=?
            WHERE id=?""", fields + (pid,))
        db.execute("DELETE FROM preventivi_voci WHERE preventivo_id=?", (pid,))
        new_id = pid
    else:
        new_id = db.execute("""INSERT INTO preventivi
            (numero,data_emissione,validita,stato,cliente_id,cliente_nome,cliente_piva,
             cliente_indirizzo,oggetto,descrizione_servizio,data_inizio_lavoro,data_fine_lavoro,
             luogo_lavoro,iva,imponibile,totale_finale,note,pagamento)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", fields).lastrowid
    vdata_inizio_list = f.getlist('vdata_inizio[]')
    vdata_fine_list   = f.getlist('vdata_fine[]')
    importo_mod_list  = f.getlist('importo_mod[]')
    for i, desc in enumerate(descs):
        if not desc.strip(): continue
        qty = float(qtys[i]) if i < len(qtys) and qtys[i] else 0.0
        pu  = float(pus[i])  if i < len(pus)  and pus[i]  else 0.0
        um  = ums[i] if i < len(ums) else ''
        vdi = vdata_inizio_list[i] if i < len(vdata_inizio_list) else ''
        vdf = vdata_fine_list[i]   if i < len(vdata_fine_list)   else ''
        lordo = qty * pu
        imp_mod = float(importo_mod_list[i]) if i < len(importo_mod_list) and importo_mod_list[i] else lordo
        sconto_r = round(lordo - imp_mod, 4)
        db.execute("""INSERT INTO preventivi_voci
            (preventivo_id,descrizione,quantita,unita_misura,prezzo_unitario,totale,ordine,data_inizio,data_fine,importo_modificato,sconto_riga)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""", (new_id, desc, qty, um, pu, imp_mod, i, vdi or None, vdf or None, imp_mod, sconto_r))
    safe_commit(db)
    # Crea fatture offline SOLO quando lo stato passa a 'accettato' per la prima volta
    # (non ad ogni salvataggio successivo)
    if stato_nuovo == 'accettato' and stato_precedente != 'accettato':
        _crea_fatture_da_preventivo(db, new_id)
    return new_id


# ==========================================================
#  PREVENTIVI
# ==========================================================

import base64 as _b64, io as _io, json as _json

LOGO_FHORIZON_B64 = 'iVBORw0KGgoAAAANSUhEUgAAAK4AAABDCAYAAAAF1j6yAAAzn0lEQVR4nO29eZxcV3Xv+11773Oqulrdas2zLGuwLWMb4QHPA8Zgg81oAiEMviQQSEg+vCS8hNzcJPeTz0teeLm5N7mZXgJcZhLHBjPEgI1tjPGA8SxbNhqseR661a2u7qo6Z+/1/tinqqsHSSXL5BO/q/X5tLpUXbXHtdde67eGI9/ctld5qaQC2PhSj9OMgrzkjo5GAkgHfSuWo39GVVttiAgiE9sUJo5+Yp8iL+/sWu1rs/+fQ9uTaPI8X+4+lcg2OmEckZOm6ttM2Z47xn52MJqjvH4F0kTGa///VPv8cjPqKToxcvIKZ7iXg04x4SuP3NSCuENqY/rj3dbhFG+cognULjBOVIC6TvlpYrsy6S/a9q+M+8zLLdSb7cnk7k/RK4GUk1alXacfDITYmUYF3rSY0xejGOMcLxbRqFYrARUQFQQZx3Ta8W/Fi2BUMQJBhBxwIRBbHSNFp2zjFP3HoOZeiE61R53vVMeMO541hSZbRrhAx/6qgAaQ8Uza/jNpEsf5rQoaPE491GvxinEO40pj142MfXaqNk7Rfxw62j6fyG51rOKKggmKKV4H4k87qSqo4jRg1BPEE1BET459RMCZwE9/dA8ffNN1fOC6a/jSp/+cpK3fdkjrFP3/nzqWuCYUOqwo2nZUtMBnm9inoLiQ442Qq40SUSUK5ZPhX4XhQwfo37UVUeHgjm2IBpCTMi9P0SuUOpe4CLP6enEyXqds/4+IYIxhbl8PqG+ZiicNuamgWMYUlkCQU/bY/850AjquMqOrzPCRKg1VAmA0SlODEjQgBkoIM3q6ZM+RIQ0oFh9ZTQyGgABeFNFosKlI1AVUMYUBFwjR0xUUxOIBJFD0iooQTB77DdrSnTPxSGEIUkh6W+gz/phH1CDeY6RlexLvjuZP89+JazL5Chnnb5OisRDa/q6tb4ooEnKsCt5YcjFYjUMPbZ6meFxlrEcFkbgWofiA4AkSPyEKqiF6AYv/N+HI5ohDMTeE4uZqDr5tBiL4Yj1t8XZetBHHKQT1KEI0NZp2TTEGFFEtPGXNz2jTbVZIzVCst5liNY9OHTOuBVJFEkVDwQQmRIaUACMGchsX0noI2oUXCyEjGEcultRHHbjuFKuQEMiNpy6ANVGHzi1ChZodBdsgzTNKwVI1wqgAYgtm9PQ36jQqZRI1dGU5iVV8yFpMq2LIbHPDx3NuixnVYBS2bPgZ29Y/z/IzVrL4tGUkaZmAIMYBAZGmqlR8X5XQ5PI2u9gYU7i3tWUE2Dakw2to2bO1+jA7Nv6M3Zu2c9YlV9C3dBkExec5ogFnBPDkGjDi4uZqPOSCx5gEVQsmoJqRSGFPeFBjaKgiNiFerAHRgOQZDkMwEZ1RLY5FwUxxoM1DE5nNBMF5IQDeWdRaUCWEHBMCRjQeNgEwrXEaIGhOoKlKBlQDVuP+o6BGyFCQpsXSGXXMuFOSKCqBukvYUq1yz1NPM6Orh+SK1+gXH3uCoVIFNEc0dpN6wWMYkYAL4IKSaaCOJ0cJongjZKL0jRxhZaXMa5YvZcac+fzjDx/l0IatGKOkISdoyoPb9/Pwtsc5Y8UKLp6/mCVpCWMESyDxSjCBUQfBQClPJw6egGAIuLzB7Z/5Wx783rcBw9wFC1mweDHd3d0458A0dfWxb2sE+iYvidhiaQLtqHbkh2hAZlnGyPAw27Zu4vDBA4gm3PLbv8/bP/yrgKEkns/8zX9ny3Nr6UoTktRgkwTnUpxNMTZFrKJB8XlkdO8b+EadvJZRywJnX3QxH/i1X6caFDXxIO16cT3/8H//GV3OkqSGJEmwLsG5BJemGIm3QAhKlmfkWY42GuS1jDyDBo7f+OM/YvqCBYh15MNH+Is//i/UjgxSTi1pyWCcw7kSzqbYpARGCb6B95488/g8kGUjZLU6tUbgre//T1xw9evxGOwJiNyOGPdo1rrRgvkAGkq9mtFfG+ZgyHl0/352JRUwgqhDCbggBHV4JyR5TjdKJXGUFMoacdmS9axZ0Ms1qy+AAN/buIEfPfIESg8q09CkCw1VmDaLoaSXTfXACxs38eP1mzhjZh+vXrGUM+fMYRaGbu+p1HK8BW/GgmjaZoYhEGpVdmxaj8nrGBUObt/CgR1bI6RnpECtC0laqDUh5Khmk9cEEyWVsWBMvAYlqkgUvBw0ICGABKyCwbP52adJNFA3gqqnZA2jQ4fZ13+Igf79ZKPDEJr3hAXxoMSRSUIQQ1Lpom/2XKbPmoumjqp6cptgiFe7NYIJgQO793J48ABH+g+ieaO4vQ3jcaKoRhiX0ts3k95Z85gxfxG5jQfBaxx/YoX9/fvZ1X+QwwN7yWu1lmoTxwngo2ooFmwXld4yfTNm0jNjHra4EfQE3e7yb1uOHx2mqpQ1sHLeXNm+/5AOGFDi1R/HJVQTS8MkVFRYPSflz+59mjte3ElmLU4TMuo4tYgkBOrMNoEPXPkaVs5IqfhAX5ows7sbZw0jtQYHRkYZDJY9RzI27NrP/Zs2QrVO7d5vkux5kfd/+u/4zp4jbKyN0kia6kNOl8+Yn6S8asliLlq6hKVdFbp8jhrfQj6MMYWuqbgQGN69k4++8yZGDu3FiODFMnfpadzw9ney8pzz6JrWA4ANUWWSAN/42hf54XfvGL+YFJeySbjl13+TNZddRWYcgQxDRGaqh4d45smn+MG3bmewfw82BKwK0xev5LPf/wFZWsFqwOUZiVHE5+zeso0//q1fY/fmDWgI+II1DHE8KglLz3oVf/w//5ZZCxYjJqWBIUsciCkQoYiDpz7HemWkVuObX/xHvvIPfw0+J0SLYdxsrDF85Pf+iBvf/V4kLZOLJRcb1UMEpx4bcpwG8BlPPPQAf/ap/4PG0CCRXWM7TdPai+WaG9/Bxz/1B0ybPgPFkYmjbgwYgz0B2LRjVUG1aViMUS5C7iDBU9KcUjZKj1fK9MmZyxeo3bqDulG8KiFxeK8IdaxWmV22vLqnlzNdSqUnYTR4Dg4OciTkNBBKITCLjN4ey7KzF7J4UR933vME/b3zGT24m0H1mERJ6wGTpZjc0HCOURK2Z8rWrVv54eYNnD1zFq9dtJjzF84jMYITAQ2FxqsIgd27dzFcrWGkhHdw3U1v5WO/+59JZs2hIY4gkSGNgvGKU6XnvnsKzbVtjVAwhiCGmUuXc9r5l5G5FMgAjYyvsObaN/PWd/8if/l//QHPPPQA4pWB/kPs3beXmUtWkYvgU8eoerBlZi4/h9UXXcH2TRvGMRYFqwVgzRVXMX3RChomxRS2lvOCFs4hL4aGkXiEjIGkl0tueCu3fuFz1IcHi51tswOMIemqcNUNbycvzWwpPTYIxgiZUTKxNMSCKCbp4sxLXsfcpavYtu6JtuCVOE4PYC1X3PBmQt8CDosjEYl/MQWu+nI7IMYFQ7S9bzWQ+AZqPSrxVAUrBCvMNDDTC4aUusnBG0res1jqvGXpXH7t4gtZMb2L0azOlv39bO8/wmBuCZqSesg1ITNdCJZKnrNg9w4Of+3v2Xbnv7L3uWf5+ic/xvkDO7hi0SwQpWEMiieYgDcQJKVqpvH44RqffXYD//W+h/ja8xt54cgI1cI4kswSQoOtG9ZCfRSVhKtvupnf+K9/ip09n4Z1YD0WH+WGWNQYMAEjnpbnsPVToAHaiPq/ESQ0cKokKhiRAhERepYt4/f/6m848zUXohh8VmXb5p+RFBLMaMAWJqUaT9/MPsZwjCbTmgIIyOjunQ4iGAmo8a0x2uJTiSpJAIfBAoYGpWkVknKlZd0L8fqHgAQlqfSQVioRDShOrreBIL5AiAJGpFCPwKSWnu6eCYwS0OY4CZR7+7DWkhgFiciH0RPAZQs6KfTeqEGCRdWhzWsES99oSlkS8tRjGWG+r3N57wifuGglv3bVVaSlGXz9qU08PTTCTq1RTTxIRldoMDe1LJs9gzmVbmwGiXdI3fOXv/+7bH76UaQ+iPGj7Nu0nlv/6Pc51wbmlnMyN0pm8yiDClgIichCzTh2escPN+/hH+5/lL956Gnu3DvEdgP1JOWFZ56lpDlLF8/iQ7/9W2SVHnJSTCiRZF3Y4Ca5qo8uG8Z73Se6u5sImQRLedoMfudP/hTXVUHywM7nf4YpAtlFx+S5iFAqlY/aqRghTScYnzL+5VTjcM5inRt3HFp7CxgxWOembGiqNk1zHBPc7u0HzlpXoCKTv38idFKoghINL9FoYwsQxHO4lCGmxkqX88YVKzh74Tx2HOnnzp+9yI7BDWgeOH1aig1K2UO5lDC9t0ylK2FkZJRDg0OMZEqwJTCQa+DAji1YbWAJeA0ED43qMPngABWboCjGpzjxxfXYPk4lE0+eGBpi6a+O8uxTa1mQWM6bPZ0NQ1CvzOGSd7yHmfMW4IMikgEGMabAVCeqBcdbnGN/wmJo4Ji36kyueNNb+OE3bmfrxs2EEBBnWzhus7MmBHq0vow5wa1vMtdRjCJtu0U6bU716O016eWKfT45OKwJiwePFY/zOV3O0G1hzaw+qhecz11btvK99c9RDdNQLVFLLJRGKVVgaSVlaamHmnoGRkY5MDSEF/A4vHXRGg85WTbC2975Tv75y1+ikWURe7WW1a+9mEXzFzNn7xCJr2LERcvb+HGjNKKYAjpXBRWDt8IOTdm7Z4A5b3g7p7/6Yhrnn8v2hmFGV0KiOdZkeK8YaxAtYK4JsNixqQ10n2rtRPCa8Ma3vwf1gRVr1kRQv81hIUWfx9ruJlItRWBTp8whIpjC0JKWA2Jiy52TNJ1JxxroCYzvWDQ1445TrIu32n7aPSxBciyBijPMmT6dHmPFNSzrBkb4H488xl4zjRDmY01gmta5aHYPV644k+WlEhbPpsEBch/AOAJdsWsVrIAJnsRn/I+//AtGDh7i45/6A7Zs20ZWq7N02ekM7tvF//zkJ3nPn/8Vzx8+xCHrMSFqcM3RjyGpSQytVMUUOq6XOtXUULNlurpO555Dg/zkgUc4a95sLlu2iFU9vfRiweeoGZu3TuFFG7dQ43hb2/0TY++aPDJMcKy58FLOfO355E7QLLrOpuSjY5BOetH2xsS+ix9jDGKkNZ/2M/Yf3Z0+mXFVER0DXIp4MAoTRQMG0QxLoFuE3u6UtFwmy3MODlc5mAddPm+eHDbCaCqYbIRV01MuO30RK2fMYNOhQb7x7GZ66yN85NrL6DYBsYI0nYnBIJKgUiOgbH1xEz/6zr+Rj4yw6txX85E/+CNUYfPTT/O7H/oLfLXBvp/cz5qzz+GBA0M0jI2u1AkbIDrm4gxQ6GmKDRr1YKsYSRlV5dDeAzy5ZzeLK92sWXwaFyxayJw0payepKmqYBDSwl5uAj9FfyIEE1nchDDm+m1fZokuTjFKXSCIiy50E+G6oBMxi/EkbfCV4silMJAmfC4QceRxQkiiEWlVEVHylunUNr7Wuh1jEBNINRxbRZp8hl4yTWJcFciMKYK/C/+SQsnnlDWTCnVNKiV6Kt04LxysZVQPDkdQXQxWBLENljQavG/eacxZvoSh0VF+vHUHdz62jgFTpmZLnJdUMCFK3XaPvFHBaINgQHzOv3z2H8mHD9Mz/zTe+K5307AJBmHleWu4/PqbeOCOb3Dnlz7DL3/+yzx/qMah3KCmkCA6Nic1xcrR5rZFIg6q2oL7jBq8JBwxhhero+x67jkeev5nLJs3h1cvP42z58xktrfxUJsGJrRLp+iJCxp/F1bAUZa+8M1LHI9tspyZbChNpknimwh5BfyU13DRtjYZVFA1iFhEzCRDrv2+7VTyd8qQP0cdtxlYAlY9RgOJKJUEyollzuxZsq9e0739g2gDRkxccqdKitJTSulyhoWzZ1Bbv5P/9cAT7DI1RqQbTaaBddGDVsBWphUQE/sWPBhFgmP72qd46M5voyIsufo6mDG74Dols453/qdf59Ef3Ev/1g1suPObrLnweh7cPUBdTIwek7EZHY+MMdG4UBP97gh1CzmOmknZ2T/ET/qfYGk55dK5CxiwFYzpxoYGSk5uNErLvPDWh4gVB1HG8kVePmoqK80jn/qcpBUX0Cb4VJCWniNFAI8UCHY0PoFWRsJLH8+/L01iXFGl5BUJnq4koXdaN5VSSu49o2rZcWBAB1CQlFQCZR2lkiZM7+qiK03QRoOsITx9uMo3d+5hLyWML5OqEqROTh0BDifK1nyEui1hjUULg0SsI/Pghkb4/N/8NT6rkyxcybQLL6VqHNN8jtVAwzqWrjqLK978Zu699Uvc99Uv8/Fr38K6/YHdEN28xXxsODbrNKWAUFyr2oTSEnKn4IUgCaPWsDFrsHfrVnpXncO8j3ySw48/Sn3jc7jB/TifxYAhBSGmLwXTFAMvJ+sqzYu9iaI+/uMfMtTfj3VlnHMkaYq1Fqwtgmk0xgvkOT73+HpOvdpPdXCgGO9U8W8nRlF6v9xHdGqa0jhz4pkzdwYGYWS0xuEDA4TgqSwsoSHgjCcRZWZ3hXJ3N7lXhqoj7B0cJjGWrq5KjBILSkmVxMdrbDT1qBg0JOyr1vm7ex5kmilhrW3pdCIC9Zzy+hdY95MHIUmZ9bq3caSrF4zjxZ/+hF1bN3P1zTeDTXj3Rz7CI/f/gKG9B1h7521c+Iab+cHm3dQk4opOo36bdwKgCAWsp1iNQThBoiFlVDFeQBJGbMpIdxdy7hqmrV7NrIP7yJ9dR/+TT8C+F9G8TmjlZ8Tww58LSYz8dMCTP3mIJx5+aHJaSsvi0nE3m1GDWE/QvGXPNHX1VwI5GBPzTQs8V8P+/kFyDVFTM9AlgZScmRVLX6WCGEujWmf3oRoN7wnWoqaEimBEKYcYzjeSeoL12MzircV4Q24TJGTUxVETIclqGJ8Xuqihp1Fn+313YrM6dukyes44jbRxGFMbYsPzz/KTH9/PFTddT0LCorlzeeONb+X2//UF7v3aV/idN74Vt3IeIzRIVBAchyXhJzv2kYfo3W+YnHIeCKLkRoAcg+BJUVHSEDCqZIUqE4rA9XIe8eLgAj25kIdAlqQMzFlA17WLmHfJFfS8+AwH161jNO0hx2K9xiSNn4sgiizn8aw4ew2rzltD4hKcdTG8svhEKzswKD54QgiELOfIkUEevud7NEarL8tV30zdOokWpnhv6oVzkz4kQm4Er9GFm6CUrWNGdxelxJE4x+6hKiN5wAZDbgzexaVxGkiKONiAI8ehWBLfjPc0lHNhNM3pK8OHz76AJaUSn/n0n7J704bihrYM7T9AtnM9YhLSrMTuf/kWLvF8+l9vZXTXdg73H+K//MqvocYgKCPVIdKuEvXBQb7wsV+hb8kiGkmC9Y5gEt736U/z2NYdeCnH9CITuGTRAuYkjoYxWMnIxfDU1oP09VRYPr2HMoGaVcCyaWCA5wYHQS1BAkEa8OxTjDxyTzS+VBkW4QjKQQOSCYGMYBoYNVHP/HkwbmEEqsA1N72Tt33ww4QQwa0m4zZZV2jyVAyt9KJUDx7gmccfIx+tFuH+L53i9E6W/U+Accdgj6YuFq91Zw29aYm+rgrOewarQ1S7PIeOjDJqBG+arr22u0k5pp4U100xAaZLyor5M+gNo/zWp34DzTMUwecN/vijH+bw7sBZF17M6z/5+3z9hRdZYLv5yNUX8tRtX+bxR3/K7336r8nSlBByUvF899av8sW/+SsGh3bziY//IfunL+RAnlA3yvO1UYJNqIvixFPxNa5dcRore1OMWqwGDjvDrv0HWT2nzHVnLafsPZkFEyzf2rCVFwYOIyE61SUP+AN7aLzw2JgHqvhtMahapjUOY6RG5hKSPDkGuvDykDFCpjE+ounqbsLtzZT9sSE0syReuQmmTlWI0VIBI1Apl+grl3AiHKnX2DNwiLoPOFFmKOqtK7xbtkBuJipVR6cmxikKaRBKmZIaS6lnTqFXKnd//Vb2Pf8sxnXx/t/8BL3nrmHu3gHUlWHBHGaffw5ndJXQRfMphwILJefN734vd337DvZs28Rtt32L8z/6e9y5aSsj5RI29NMwlmDjXG3ISfOMJHgS77Ah4KwDm2G0TqINXGjEttVgQx4ZQEBNNPRsSBBfAtr97YEYCWYBh4QSNqTx/Z8j30rh5BAxeBOPyDhYbKKzL4QIA3KSN3sbjTl6/n3IRRchTOvqYvq0LvJGjcGhIepZTkOEzBhyl1AKWXHdFcClCYWHZ+waFDn60E0BPucGvBADh1VQEnIRTMjIjgxx2+c+h+TCmiuu4NwLL2Y0KfHrV12DJzBbExZf/AYuu+j1aACvnlwcnpTuWQt453tv4e///E944t4HuOK9H2PmtDIDmcVqQi4Z4g2OaKiNJiUa0kVuLEayyG55QtAKWehGpExmIqSfmwQV8E5ivhyCmhw1daDw0bcwYwEcuTggwYUEpPHvtKcG08Jsm+9N1jubwfDNypRRGjfxw5fOycd1+b6MFHVchZFajfpoFfUZtcSSO4MRiwQh8QaneRGnrjG/UWMVmSk3pF1lKO6ruCRF4pwo3sS0HyVHxeAk5/vfuY0dW9dju3v48G/9DmCxIWPmNIMLSlDPqCsj6rHUSZQirdIScNzw9nfzvdtvZfv6F7jvC3/LG//zH/LVZzYzKhVUIfFR9DTShO8/+RT3uUAQi6FOwxrW1zJ2793Li/39JOqpuRjHemioAUU+nTYxUjUxJ6iNhBir6wkxNkKyAqueevEnLZ0e5XWHJChmkgg9SkM68W969M92Su3G2c9ZA3EWBSN4jcEtWIdRSKHAMwvXKKYAsuOF2Ew/bGaZtlNCBFbUGFyIVq0v4i+N2ni1FdLbS8w1Gz00wG2f/QISAldcdy3LznoVVZOgJjKB4LCA1bzVS5AiI0EbMVGvdzo3f+gj/Pc//D3WPngfb9r6QZZXUp5uZFi1pAFyF1Cbc86yxcydXiKQ4NRTNZb+JzeyZOY0Llm0mK4sY8QpdWt4bPtetu3ei9EC3Gq5Ltuzd9u2vvA6miIbOkyBhk12Arf/ZSJ1xgXSMfMVKUniC4lrkCkz6DqniLod2019fOocNhzD5Ztivnn9F+bEeKV+0n+mJOXo1RnbN0zFAhYXAt++9Z/Zt3snrlzivR/5VTLASEyLdmHymJrxtvHXmK5y2RtuZNnq8wj1Ee74+7/mimXL6c4hc4aGFUQDaeZZOr2P83sqXFqpcHG5mwunz2BWyFiWGNb09XL+jOlcOL2Xi3t6WZSWI45b9B/k+Jr92GgnTPzoHPvvTzLpxcvTXNv/f15T/fkg4yJjagGtHMFITX95iNUTBBjcs4N/+5evAoEb3vVuFp+5mmBjHQYbwHZQ/aPJvGbaDN7/8U+QuoQXn/wp2cafcU5vD04gs9ErlgaLIHgXMWa1nob11FIlszl5EqjbDG9yMB5jFB0/CY43oJfPWtfjtqeTXnTQqiohhGO2G0LnDTabaQ/JnOpDLxeS8XNh3CBCHjxeQwHPTP05USX1Gf/2lS9weO8uKr193HzLh/GSEE3hyC2hQwRfRFBjuOCKKznz/NcSQuA7/+9f8frT5jKtNooRjzEWLwmHguFgrcyerMRu7zhUUySUGMlKHBgV9jUsezPL/obhiC9SkqJrDxtA/FE2qBimz/OiJNXJyxxVpZFPzihup7zor7P24u8QAnmeT/0hMS1nRScU7T0t2jv6dZsdZx6d0gkFkrfLmXEyZ8ILBWqNBnnwqCRjnnptGq4a/eV5g707X+T7X/9XjHpuetd7mLV4ORGsia5XlWb9g86uHRNyTFLiPb/6m/zRE2vZtu4J9j52P2fPP4vHjhxBxTGijtse/inWKA0roDUQ4Yiv8LPabh7auRPjG9QdoJbMB7xzWB8PmwsK2VHco4XFPjQ4GDOG25ZmctDP0Wc0jgVVGegfaDrKxv21ubaH+/sREzO84lqPFxhK+/fiGKvDVbJGY9K+NvvMsoyRkRG6u3vHBt/W57gpaKzvMHxkeNL42haHkeEqTWT5ZI70CUhcxQuaEzAhhgE2JPq8y7mguacuihcYDUHXHuhnWB1lD4rD+gTnSzSMxWGojzpe7B/k1q99FTLPlde/gde99UaaQIW0PD4Ta9Acm4KBhhjOuewyLnjdNRiBuz/7d7xh2Qy6RGkAqTccVGGXKPtU2S8J+9Uxajw1DRxU2G8SBoJjQA3DIiReYvVUhbLPCSYHMSDNmmbF0QrRsn78Rz+itm8vJh+lK/O4XAkhsqxRX5i4E1ZYBdQRCoMWzRHfYHj3DjY8vRZRW6zNGOsoMYlzwxMPM3xoBxpGEd9AQl5czbFaTQBycoI2MFrDVY/wwoM/oVEdbYsmbia8xzn50SpPPnI/0qiS5BnWCwQHIUGCAQ8SAuIzqFfZ9tyT7Nu9nZanDsZ2Txyo8NwDd5NVD6GhjoQ8GrxBCHpil39HdRUAnK9z+oI57DnQTy3XGFguSm4zckpsr+c88Px6rBd+6bJX8du338U+Nw0JntxKkZlQWNgBEhEWH9zF1r/7byybN4sj9SqnX3AJn/qT/ycmDJ7AJNqvSE/MpLWasePZp/nt978HP1rjw3/1OR7qXcj6IyMYIlqQhHD0ftrEipdYvFpJEKP01o6Q3fMdqt/7Z0DHwaS24F2MY/qcucxfdhpz553Oa19/PZffcENRQ80X4ZzjN8sXsi/JG3z/9n9l3aM/Yf/unezZuYPB/kNIyNGQR6jR0MrojmWflOmz5rFgyTLmLlrCBddcy+U33khuy0hwOM0Z3LWD2z/zGfbt3MbOnS9yaM8e8loVpMjqUFpysBnM7rq6mbNwCUuWrWTWwiW896MfozJrNooQRoa59TP/yK6NG9m3ayv79mxjZHAo5vRqUdC78H4EATGCsZYZcxaxaOlyZs9fyHW/8B5WnH8+Xi0lOi/D1LmqIIKIobe7QnceABvrW5GQk5I3qiwxCZVSmUUm5V3LljHiSlhtxLJKRrBFbOgohi2HB+hK5jDjjW8iDBxA8hqXv/ktBGMx2rk3biLFAnxKwHL66rN53Vvfzs4NGzl71Sp+tm0/K1yOipJLincFbtUmH+IhiK/jFsbfLgRUc7wJTEst9rRl2Kuuo1Qq4ZxtqQiNLKdRb1Crj1Cr1agOH+GF/Y8zc8Fcrrz+ukJ1iodrkhWuhW8H4bmnn+H5Z5+hUkqYtWABi1asoFKpYG2CTVwMWA8eHwL1ep3aaI3GSJUjQ0McOPQc5ZnzuOqmt+Pb4J2Bw/088vCPKVkoV8osf9Wr6K5USEppLMVkXAHFKrV6nUa9Tm10mEY9Y9uWzWzeto2b3v9eps2dQ8iFkOU89eijjPQfoFRyLF6xirRUpqtcwbkS1iWIBNT7GBZbr1Gr1WhUhzm4fy+7du3hnMuuZOWa87FFzbWO2bFTiWtCTlfisEXKSQxviYzoQgTrcxtIfSAh0KArepkkA3UkQUi8YoJQM4YDeYOasyS554kHfsTClcuYs3AJqSa48UH5x6Vxz8sq0mq8EZwGagf38fnP/D0f+T9/l8O2m7rkpAqp7+oYuWyYvAgSEhomUPKB6Wpw0h7MUvQfwBc4itGAb9TxeRWblLFdPQSxhR46OUZYgyDegeTUGlXyvE6aOFJJUGNQY0Ga3rG8OflCTitBM0KeE+qKkRJpV3f05BmPNwEf6oxWBykZS4kUm6TRNSymyMBo02M1Sk18hihkuafqA+n0HgwJSRbbHa4OYq2SJJZE4hhVEoSI3KjksdlChVJVchr4LMPXla7yNMSVCly5c87tmHFbcJBAs9xbRKkMqQ9kNsJfiQ9kLpDkJYQcbzPEd5EbIUgoCkPHdHaCI20oe7ZvY+Hy0whiSb1plfrslNoZ16gnIGQmXiZJ1mDd2qdYcfbZlNJp5K6GVaGUVxi1eZHKPiZxixbHv6eK8/GduhVcEKyHhm3DkONAYgVLF4Nc8IpTUNNAMVBUgdEiv84UDzdsUtCxIkiCxBKqEkMrfQwyw2isiNNym7eRUHjiQ/SQ5CZi4RKK2IUCizYav5+LIkawzXFD07SLs9dmUZOo9QYxeBECHhOi+iRN9UKi0PAmHgSjsepNw4aWQ0ZCLDsas0KKNkMguDg5p+PX41g0JeOOcwRKIKCF4TBWn9VYE4M1Aog4MjGEoo5qMJ7ER2+MNwEJSQHcR6aNmbYxtC5KJh+ls4luVC8mVg7Uou4uUU0JEv07RqNrNSY/mkkSVxG8xFR0i6LeAw5jhIZpoNZgQooNEbkY408BmgXqQixcB0Udm0h5wUhxI4vkwGZWroIRi1cf7WYxoAaVvFi3OK9Q9GGYqOOGIhYibrILkTEaxsR1o0hwLBwyuTX4YkNcrlHGKVEw2MJRolp4K4EgJFjU+1hCqSlsUEweyMUWxUhopeEHiniGEIpiJpHxgoTI0BqZlGJ/QzGemDtIjHcWWrUptJib1VjTTG2gYTK8OBKfkmjkGSUyvh4lAGZysiRjp64pZg2R2cRH5lIVgvetZMrIXPG6URSrlmCassOgZizaU4p/QlEDAImRZlB4o4xEaxPQIoopnlgT3chFGSM9iljWwuBpfg8k1nOVGBCeiC2A9QZNrLjJdK26HwJipbgZmlIlUgI0I2oMwsSgEh8UR3wGhhoTnSyqEdNWDzYpRjf5ojMa68oGo6h4QrEyNmZyxgNcPAzU+phzbY3gKZ5GpIrY4spXJpc2MkpOXkCSYwW5g8YRGdFYEDrGTxXH1dI8282H1EgAJ661oWNqdBsCVLwvPhploh6lKLxtDN5r3BcMrlhLi8f5DDVCjgG1UzItHMM4U6KFvGX9RnZv2065yG2qi1Dqm845r3kNO7dvZ/fm9UUtKkPNK5XePi64+DKO2mMbNf3jMVDGFEZQjoxU2bRlM7v27GZ0pIYTYcGi+Sxbvpzu6X2oJARxhZbdGQViVcRnHn+SI/37SYoiwzmgJl7aC+fNZcnpy0kq3XibFCnoJ5AzVui0T/z0YchGMWLwGBrGserM1cyZO+8460Fk0MI+FBHWr3uWQ3t2Ii4FU4rwlvdIGIVGjm8EfJJw+Q3XIy7pXMOSmOXhUIz3+Eadbdu2sGvPHkZHRjAKC+fOY/Hpy5k+a3a8wSRWDBOETi91UeGFtWsZPLSXhBzyGt09fZx90aVUXYKKMLRnHxvXraWkAReUqjWcfvarWTRnMUez2CYnSzImeYLCksVLGN5zgG9+5cuIgUve+AbWXHox1joWLVpEPnSQr/zTPyI+cOFVV3PBJZd0unQ0i9RDlDQ2r/Pkww/w4L13s+JVr+Lyq68hsQkD+w/x3e99k2/fdiuXXnoZl7/+ekylrykWjt9PwREBw+ozz+Lhu17kgXu/jwq865YPsXDZ6Tz20IN89dt30DdzJje//xYWrlxNEInWVvNq6KAj5yznnXM2t33p82zZtIk0LfO+3/gEs+fMnmTITb0mTcZVNPfcf98P6e2yXHzZ1XRN68MoaGOUe+/6BtvWb0Yo8fq3vQMxx3rM9nhS4pWeqGLzjHVPP8m9d9/FouXLufKaa0jKZQb7+7nvu3fxnW/eznnnr+HaG24kndaLx2GMLTC/DvoysHLlcjbVBvnGP38Bl9exSUrdNzj7iqvIccyaOZNp553Lnbd9jdDwvOUDH8T1zingwqkdPVOuZHOLggHXVWbWvLmxNpXm9M3qo9xdQRDSpMScufNJBFINzJ4xk65pvSewgGPBKjavc++3buXOW7/M9BkzufHt76Bn3jwqM+dw2qrVfPBDH6anq4sH7/out33+s+TVoUKVYSyu9KixoFqEmwrl7gpz5s1GQx1rDb0zZ9I7dy5XXHstlXKJIwf2cP/dd6OhsNZPAKIBAZvQ09tHz7RenELZWebMno2zluNm0WpLU8EgjA5XSUsl3vFLt7B09bnMWryU+Qvn8eKGZ9m5cT1BPOe/7jIuuPoqMOkknflYwwwoeM/9d93FN772FbocvOVt72L6vEWkfbNYsGwVv3TLh1kws48nfnwfX/vcP1E/MhSVHN9ZUqUSs5yTrpTlZ6zCuISZ8+YT/Ch33XEra++/n0rmcdYybeZM+mbNo2/WXHp6p8f1OsbiH4Vxi4dOAN4K4myMnS2qzoDGeNygWJtgioKYYlxhdnR2YTVr9hkNbHj2aR69725cXue811yAJileHN44vDFM65vB1de8DvE5m55/jkfuv7dVev14zzlrSTFAjaBGMUWeXHQpm2gJm1ggI01KhV6rUeJ2eBQjYmaiVLLROG1Z3ZhorLU+OdX3x8o7qUJXdy/v/sX343pmkNkSQQyPPPIQP7rne+DhnPMv4tq3vp28XIrlTzs8ZHGZAts3b+Th++8l5A3OPe88bLmClwQ1KZk4ku4+rrz6Gqz37Nu8ifu+d2eMkjOdW/+emEjrRSBJufl9t7DynHPRPOOH37yDh7//PfAZwVhISuASxKREPf3oB2RK4yxIwKpBgwXjQQOmCJpZ/9wLDNcKTUeV6mA/6puVYCL80impxMW26tmwbi3io5HXO70vujfVRO3e5uTAkuWnYxKH5oEXnnqSq264CUkmPtthKmqm+CiiefF0wRgPMTowwOCuXTx4713UR45w2upXc/3bbo7Zyu0YYEe9KBoCwVhEwFLEARgXD2kEII4ad6EGvIT4wlhwCaXeEqqBxNdZ98RPuffbX0fynEVnnseN7/ogxvUWa9/MsujAtpA41s3PP4OODoJNKM+dR25CNP0UEEMugRkrzsDYEmk2ypafraVWHyEpTe/weQ3NclI2Gl8aSHr6uPF9v8oPvv111j78EA/ccyej2QjXvukmjAYUTwgWMRYrWRG3MrmzY3rOpM0sMUYIAWbPmsWKFSsiTOE91YFenpTxnvcTgGChqPOlWcAFi0igVhtBCQQToRZPRDKm9fRhjCNIA+8bJ3qPt8YW9zfixc+sfZoXvn47fvgw559xOm/95Q+ilb5o1bacLS+ll3Y+moy3TvktjRXLtXmVm5hLYT1sX/cMd9/6RXy9waJlZ/POWz6M7Z4WM51VC0hKO1p8QbBqyHMFsUhQ8uGRohCgiTMWQVSYVq6Q2ASyBnkAEyLM1ekmx3tNIjxW9G66pnP9O36J7tI0HnngXh790b3UhoZJrEOstB5iI+Ho2dHH3ZPWFRziFTF33jyWLDuNJctO47Tly1m4aBGgUdIqnJDfruXzEZavOgNv4iORdmzZhCWABkLhZzJiGR4eIfNKrrBq9Wps0vmVNWlOIWI+l11+OTe95S0kacr6Dev51rfvoN4YjQdHQ+sm6XQ27Xp2K4uleHG8sMN4qKK4MARQj9GM/Vs38+1/+RJhdIjZs+fwjvf9MmnvTIIxEHJM8GOPe+psASAIp69ajaRdqCo7X9yI+CIpVG3EnDVQHx4mBCW3CYuXr6SUljtHFIq+DBGTbzohAila6uGqG9/BtTfciDPCs48/ytrHfoot0u3HnCFTc24HjNscRcQ8gwYwQkaIQSEagXYtnAUnIgSbWQWKcO5FF7Hq/NdQN7DuiccZGRggCQGn8aF11mdseOF5ch+Yu+R0Ln/Dm2g+yeH4xln7fIqHl2g0DVWE8y++lDe/6xfJkm6eeuhB7vyXryCjR3AhxzWdDB1SjF+Nj3Ka2O/xxhekcJwAhoALOf27tvP1L/wd1aHDlPvm8O4P/Qoz5s6ODJ7luNDg7jvvYN/eHSdUUE6NYdlZq1l9wUUEY9jwwjr692zH+YihW42P5dr64gZqwdM9bx7XvPnG6ME7kVgSBfUB9QFrDBoCScijcykpc/Hrb+T1N/0CLinj8xpeozfTqOCOYV5Mybja+nz898jwcKyUjaE6MkLmPRTenGq1Su6jd742OgodBh7DmNEURAjllLd94H1c+cY3IlnGF//pn9i7bQc0ciTL2fjs09x/7w94zQUX8qGP/jpdM+YVqT8drV3T5EFVGRg4HOvCauBIdYRcHK+++Epues8t9DjL848+whf+/m/Zsv558lqMJut0PgDBe0ZroyiFj79ajeM4zgFQClcu0XU73D/A7V/+IiOHdmHLJa6/+b1UZs1hZGSI2uAgA3t38t07bmfX9k3MnTsDOVpW5qR+JEbRpSlvvvlmrrvxzQQC//qlL7Jj8wakUYPGKNs2PM/dd9/FGeeeywc++rEYXHOCgSRSxDzs37+P0ZFRDh44gPU1nDZQY2m4Mhdc+Qbe8gu/SNpVKnR8LR7id/SOJrl8FcglMp9TYcNz69jy/HrwDdQEvBhmzV/AxZddyvbNW9mw9lmyei0WWBND36zZXPy6N0zyKB1tCWPnTaRCseo5cmAPz617nj379mOdJUlL2EoXrzr7HObPm4+xCUEszcvkeAwRCHjxOIUnH36UAzt3YTUQTEBdwpLlK3nVea8mqDCwcwtPPv44hwYGEGOZPr2Xq667ga6evuNPJwSy2jAPP/gg1cOHW27prNzNBa99LfPnzz+mVGwKmBiPoKx9+mk2vvA8ZYleOJeWCmwzyuQ8z8l8xjlrXs2KVWcQjOvINRDGdJiIFwdP9cgQm55dy65du0jSMs5ZymnKsrPOZv7iJdHFb4qqPNDZ/qpitcFjP/0pB/fvQ0MOWCrTZ3HplVe2kBw0IOrZuP4FDvb3c/GlVyLiCj/s1Fj9lIwbClMrFnpTrBTl3aVgFevw2mS7Zpp6tCrilXgSlVvUQ95oXefGGEJQvAUNJj78omUyFYWNO2DcXHOcGCQUFclDLMWkhXfLG0GMjTG6IRBCKEroQ65FZNZxxx6ZwJoYYNMMp/bWkBc3kbVHb6ellQH4KHkEaT1csD3lPyaQRmy6lbypdOTkaF+vcfq3jKkBESEqHkzI2PPhTohUUZ/HQtuM5X2ojDlLmgdZW9BXoekf5zad0nNmx9aH4IoSxcYWjUULn6ZOKU0MtXBwc3JJGYqBtIssNJ8Pa+KzIMQXcdcnYNK25iQ4cdGnb+PTIFDFFX7/Zi3dUEgyEYvaaHA2cdVOtkwBcSlZaFZZbxbvDCe06REPpuWwa8+5a4IUuch4Ziiw9ZdMhUtXbFKENMYx0HzI90sp9CFCsK74rjbjkab2uomgIbTqFB+PjhmrEANamgsorWuiaf0SxgpQNHFKEE4wC2PSBGIGQbNKQJQEpjmGl9YoiLSqFjbf8q0DENUUiGX1izCoKMVEYphgh/2EEHenMP0AisjUjkbZGl+QsWDzpnRtUlTnxpYj1nCQl7Q+7bWB4yPOxs81yMkV7JPCLd8MI4jMO0UQPSDNauwddHgUBwRtiX0xykilGRwdsdWWZ614jGHr+lGQIu38pZEiMkFyTHygw0tpVZv+q2LxoEiPL8IYpXlM/BQMcHRYZvyn4kGeuIYt0XncQRa/ZEziBvSoRpcgMf62NR9e8uFWaIvoa+tDT65GqilUgHb+UE6+APSUEjd6xZqvi9/aPJmFhBVP2yPkTmoQk/pvbXS7WjCReU7sWmxGkrWHKJpWmmBxLzZdsidzQNpG26GRP2GcTQPt2EMx2ia12iTaSdMEhjr5ZmVsEhN4atLnWqLl+DSljnvMi22CJDHNeM2X4MWaunmBjiHuTttk3CaPvZw45qln3vHM2qLVOpPRk7oHGK/SHC8dRDoX6MfrOj4h/uUVQmPCQI8zxs6ZlhP65Ck6Rf+B6BTjnqJXJJ1i3FP0iqSTfJbvKTpFLydNZU1MrRifYtxT9B+IOmfcU6rCKXpF0inGPUWvSPr/ALx2mDBy5hbUAAAAAElFTkSuQmCC'

VOCI_PREDEFINITE = [
    ('Posa moquette', 'POSA DIURNA (AREE E PREALLESTITI) MOQUETTE TRIS SPAGNA', 1.4, 'MQ'),
    ('Posa moquette', 'POSA NOTTURNA (CORSIE CENTRATE) MOQUETTE TRIS SPAGNA', 1.6, 'MQ'),
    ('Posa moquette', 'POSA NOTTURNA (CORSIE VUOTO X PIENO) MOQUETTE TRIS SPAGNA', 1.8, 'MQ'),
    ('Posa moquette', 'POSA MOQUETTE RAPID/BOUCLE', 1.9, 'MQ'),
    ('Nastro', 'POSA NASTRO BIADESIVO LATERALE PER NASTRATURA TOTALE CORSIE CENTRATE', 0.4, 'MQ'),
    ('Nastro', 'POSA NASTRO CARTA (PER PROTEZIONE PAVIMENTO DA NASTRO BIADESIVO)', 0.4, 'M/L'),
    ('Altri pavimenti', 'POSA PVC', 1.3, 'MQ'),
    ('Altri pavimenti', 'POSA POLIETILENE VERDE X COPERTURA MOQUETTE', 0.7, 'M/L'),
    ('Altri pavimenti', 'POSA ERBA SINTETICA', 2.5, 'MQ'),
    ('Altri pavimenti', 'POSA PARQUET', 12.0, 'MQ'),
    ('Altri pavimenti', 'POSA SISAL', 1.8, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STAND COMPLETO', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STRUTTURA IN ALLUMINIO', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO PARETI / PANNELLI STAND', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STAND 2 LATI APERTI', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STAND 3 LATI APERTI (PENISOLA)', 0.0, 'MQ'),
    ('Montaggio stand', 'MONTAGGIO STAND 4 LATI APERTI (ISOLA)', 0.0, 'MQ'),
    ('Montaggio stand', 'POSA COPERTURA / FASCIA SUPERIORE', 0.0, 'MQ'),
    ('Montaggio stand', 'POSA SOFFITTO / CONTROSOFFITTO', 0.0, 'MQ'),
    ('Montaggio stand', 'INSTALLAZIONE ILLUMINAZIONE STAND', 0.0, 'NR'),
    ('Montaggio stand', 'POSA GRAFICA / BANNER / TOTEM', 0.0, 'NR'),
    ('Montaggio stand', 'INSTALLAZIONE ARREDI E ACCESSORI', 0.0, 'NR'),
    ('Smontaggio stand', 'SMONTAGGIO STAND COMPLETO', 0.0, 'MQ'),
    ('Smontaggio stand', 'SMONTAGGIO STRUTTURA E PANNELLI', 0.0, 'MQ'),
    ('Smontaggio stand', 'SMONTAGGIO ILLUMINAZIONE E CABLAGGI', 0.0, 'NR'),
    ('Smontaggio stand', 'SMONTAGGIO ARREDI E IMBALLAGGIO', 0.0, 'NR'),
    ('Smontaggio stand', 'RITIRO MATERIALE CON FURGONE / CAMION', 0.0, 'NR'),
    ('Assistenza evento', 'ASSISTENZA DURANTE FIERA (GIORNATA INTERA)', 0.0, 'GG'),
    ('Assistenza evento', 'ASSISTENZA DURANTE FIERA (MEZZA GIORNATA)', 0.0, 'NR'),
    ('Assistenza evento', 'PRESIDIO / REPERIBILITA\'  STAND', 0.0, 'GG'),
    ('Assistenza evento', 'SOPRALLUOGO PREVENTIVO IN SEDE FIERA', 0.0, 'NR'),
    ('Assistenza evento', 'COORDINAMENTO ALLESTITORI IN LOCO', 0.0, 'GG'),
    ('Pulizie', 'PULIZIA STAND POST-MONTAGGIO', 0.0, 'MQ'),
    ('Pulizie', 'PULIZIA FINALE POST-FIERA', 0.0, 'MQ'),
    ('Pulizie', 'RACCOLTA E SMALTIMENTO IMBALLAGGI', 0.0, 'NR'),
    ('Trasporti', 'TRASPORTO MATERIALI CON FURGONE (GIORNATA)', 0.0, 'GG'),
    ('Trasporti', 'NOLEGGIO FURGONE / VAN (GIORNATA)', 0.0, 'GG'),
    ('Trasporti', 'SCARICO / CARICO MATERIALI IN FIERA', 0.0, 'NR'),
    ('Rimozione', 'RIMOZIONE MOQUETTE', 0.6, 'MQ'),
    ('Rimozione', 'RIMOZIONE PVC', 0.8, 'MQ'),
    ('Rimozione', 'RIMOZIONE ERBA SINTETICA', 0.8, 'MQ'),
    ('Rimozione', 'RIMOZIONE PARQUET', 2.0, 'MQ'),
    ('Rimozione', 'RIMOZIONE NASTRI E BIADESIVI', 0.3, 'MQ'),
]


@app.route('/clienti/ai-estrai', methods=['POST'])
@admin_required
def cliente_ai_estrai():
    """Estrae dati cliente da PDF/immagine usando Claude server-side."""
    from flask import jsonify
    import tempfile, os, base64 as _b64
    api_key = get_setting('anthropic_api_key', '')
    if not api_key or not AI_OK:
        return jsonify({'error': 'Chiave API Anthropic non configurata nelle impostazioni'}), 400
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Nessun file ricevuto'}), 400
    try:
        import anthropic
        data = file.read()
        b64  = _b64.standard_b64encode(data).decode()
        mt   = file.content_type or 'application/pdf'
        # Normalize media type
        if mt not in ('application/pdf','image/png','image/jpeg','image/gif','image/webp'):
            mt = 'application/pdf' if file.filename.lower().endswith('.pdf') else 'image/jpeg'
        is_pdf = mt == 'application/pdf'
        content_block = {
            "type": "document" if is_pdf else "image",
            "source": {"type": "base64", "media_type": mt, "data": b64}
        }
        azienda = get_setting('azienda_nome', 'Accesso Fiere')
        prompt = f"""Analizza questo documento e estrai i dati dell'AZIENDA CLIENTE.
Il cliente è l'azienda che NON è "{azienda}" (la nostra azienda).
Cerca nella sezione "cedente/prestatore", "cliente", "destinatario" o "cessionario/committente".
Rispondi SOLO con JSON valido, zero testo aggiuntivo:
{{"nome":"ragione sociale","partita_iva":"solo numeri o null","codice_fiscale":"o null","indirizzo":"via + civico o null","citta":"o null","cap":"o null","paese":"default Italia","email":"o null","telefono":"o null","referente":"nome contatto se presente o null","sito_web":"o null"}}"""
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=800,
            messages=[{"role":"user","content":[content_block,{"type":"text","text":prompt}]}]
        )
        raw = resp.content[0].text.strip()
        raw = raw.replace('```json','').replace('```','').strip()
        import json as _json
        result = _json.loads(raw)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': f'Errore AI: {str(e)}'}), 500


@app.route('/clienti')
@login_required
def clienti_lista():
    db = get_db()
    rows = db.execute("SELECT * FROM clienti ORDER BY nome").fetchall()
    clienti = []
    for r in rows:
        cd = dict(r)
        cd['n_fatture'] = db.execute("SELECT COUNT(*) FROM fatture WHERE cliente_id=?", (r['id'],)).fetchone()[0]
        tot = db.execute("SELECT COALESCE(SUM(importo_totale),0) FROM fatture WHERE cliente_id=?", (r['id'],)).fetchone()[0]
        cd['tot_fatturato'] = tot or 0.0
        clienti.append(cd)
    db.close()
    return render_page(CLIENTI_TMPL, page_title='Clienti', active='clienti', clienti=clienti)


@app.route('/clienti/nuovo', methods=['GET','POST'])
@admin_required
def cliente_nuovo():
    db = get_db()
    if request.method == 'POST':
        db.execute("""INSERT INTO clienti
            (nome,partita_iva,codice_fiscale,indirizzo,citta,cap,paese,email,telefono,sito_web,referente,note)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", (
            request.form.get('nome',''),
            request.form.get('partita_iva',''),
            request.form.get('codice_fiscale',''),
            request.form.get('indirizzo',''),
            request.form.get('citta',''),
            request.form.get('cap',''),
            request.form.get('paese','Italia'),
            request.form.get('email',''),
            request.form.get('telefono',''),
            request.form.get('sito_web',''),
            request.form.get('referente',''),
            request.form.get('note',''),
        ))
        safe_commit(db); db.close()
        flash('Cliente aggiunto!','success')
        return redirect(url_for('clienti_lista'))
    db.close()
    return render_page(CLIENTE_FORM_TMPL, page_title='Nuovo Cliente', active='clienti', c=None)


@app.route('/clienti/<int:cid>/modifica', methods=['GET','POST'])
@admin_required
def cliente_modifica(cid):
    db = get_db()
    c = db.execute("SELECT * FROM clienti WHERE id=?", (cid,)).fetchone()
    if not c:
        db.close(); flash('Non trovato.','error'); return redirect(url_for('clienti_lista'))
    if request.method == 'POST':
        db.execute("""UPDATE clienti SET nome=?,partita_iva=?,codice_fiscale=?,indirizzo=?,
            citta=?,cap=?,paese=?,email=?,telefono=?,sito_web=?,referente=?,note=? WHERE id=?""", (
            request.form.get('nome',''),
            request.form.get('partita_iva',''),
            request.form.get('codice_fiscale',''),
            request.form.get('indirizzo',''),
            request.form.get('citta',''),
            request.form.get('cap',''),
            request.form.get('paese','Italia'),
            request.form.get('email',''),
            request.form.get('telefono',''),
            request.form.get('sito_web',''),
            request.form.get('referente',''),
            request.form.get('note',''),
            cid,
        ))
        safe_commit(db); db.close()
        flash('Cliente aggiornato!','success')
        return redirect(url_for('clienti_lista'))
    db.close()
    return render_page(CLIENTE_FORM_TMPL, page_title='Modifica Cliente', active='clienti', c=c)


@app.route('/clienti/<int:cid>/elimina')
@admin_required
def cliente_elimina(cid):
    db = get_db()
    db.execute("DELETE FROM clienti WHERE id=?", (cid,))
    safe_commit(db); db.close()
    flash('Cliente eliminato.','success')
    return redirect(url_for('clienti_lista'))



# ══════════════════════════════════════════════════════
#  CEDOLINI
# ══════════════════════════════════════════════════════

CEDOLINI_TMPL = """
<style>
.ced-card{background:#fff;border:1px solid var(--border);border-radius:12px;padding:20px;margin-bottom:12px}
.ced-badge{display:inline-block;padding:4px 12px;border-radius:99px;font-size:11px;font-weight:700}
.chart-bar{height:18px;border-radius:4px;background:linear-gradient(90deg,#6366f1,#8b5cf6);min-width:4px;transition:width .4s}
</style>

<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;flex-wrap:wrap;gap:10px">
  <h2 style="margin:0"><i class="fa fa-money-bill-wave" style="color:#10b981"></i> Cedolini & Costi Personale</h2>
  <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
    <a href="/cedolini/foglio-presenze" class="btn btn-secondary btn-sm"><i class="fa fa-file-alt"></i> Foglio presenze</a>
    <a href="/admin/detrazioni" class="btn btn-sm" style="background:#fef2f2;color:#dc2626;border:1px solid #fecaca"><i class="fa fa-minus-circle"></i> Detrazioni & Anticipi</a>
    <form method="GET" style="display:flex;gap:8px;align-items:center">
      <select name="mese" onchange="this.form.submit()" style="padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px">
        {% for m in range(1,13) %}
        <option value="{{ m }}" {{ 'selected' if m==mese_sel }}>{{ ['','Gen','Feb','Mar','Apr','Mag','Giu','Lug','Ago','Set','Ott','Nov','Dic'][m] }}</option>
        {% endfor %}
      </select>
      <select name="anno" onchange="this.form.submit()" style="padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px">
        {% for a in anni %}
        <option value="{{ a }}" {{ 'selected' if a==anno_sel }}>{{ a }}</option>
        {% endfor %}
      </select>
    </form>
  </div>
</div>

<!-- KPI row -->
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:24px">
  <div class="card" style="padding:18px;text-align:center;border-top:3px solid #10b981">
    <div style="font-size:22px;font-weight:800;color:#10b981">€ {{ "%.0f"|format(totale_mese) }}</div>
    <div style="font-size:11px;color:#64748b;margin-top:4px;text-transform:uppercase;letter-spacing:.5px">Totale netto mese</div>
  </div>
  <div class="card" style="padding:18px;text-align:center;border-top:3px solid #6366f1">
    <div style="font-size:22px;font-weight:800;color:#6366f1">{{ "%.1f"|format(ore_totali_mese) }}h</div>
    <div style="font-size:11px;color:#64748b;margin-top:4px;text-transform:uppercase;letter-spacing:.5px">Ore lavorate</div>
  </div>
  <div class="card" style="padding:18px;text-align:center;border-top:3px solid #f59e0b">
    <div style="font-size:22px;font-weight:800;color:#f59e0b">{{ dipendenti_attivi }}</div>
    <div style="font-size:11px;color:#64748b;margin-top:4px;text-transform:uppercase;letter-spacing:.5px">Dipendenti attivi</div>
  </div>
  <div class="card" style="padding:18px;text-align:center;border-top:3px solid #ec4899">
    <div style="font-size:22px;font-weight:800;color:#ec4899">€ {{ "%.0f"|format(totale_spese_extra) }}</div>
    <div style="font-size:11px;color:#64748b;margin-top:4px;text-transform:uppercase;letter-spacing:.5px">Spese extra eventi</div>
  </div>
</div>

<!-- Cedolini dipendenti -->
<div class="card" style="margin-bottom:24px">
  <div class="card-header" style="display:flex;justify-content:space-between;align-items:center">
    <h3><i class="fa fa-users" style="color:#6366f1"></i> Cedolini Dipendenti — {{ ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno','Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre'][mese_sel] }} {{ anno_sel }}</h3>
    {% if session.ruolo=='admin' %}
    <a href="/cedolini/impostazioni-paga" class="btn btn-sm btn-secondary"><i class="fa fa-gear"></i> Paga oraria</a>
    {% endif %}
  </div>
  <div class="card-body" style="padding:0">
    {% if cedolini %}
    <table style="width:100%;border-collapse:collapse;font-size:13.5px">
      <thead><tr style="background:#f8fafc;border-bottom:2px solid var(--border)">
        <th style="padding:12px 16px;text-align:left">Dipendente</th>
        <th style="padding:12px 16px;text-align:right">Ore lavorate</th>
        <th style="padding:12px 16px;text-align:right">Retribuzione</th>
        <th style="padding:12px 16px;text-align:right">Totale netto</th>
        <th style="padding:12px 16px;text-align:right">Rimborsi</th>
        <th style="padding:12px 16px;text-align:right;color:#dc2626">Detrazioni</th>
        <th style="padding:12px 16px;text-align:right;color:#10b981">Da pagare</th>
        <th style="padding:12px 16px;text-align:left">Dettaglio</th>
      </tr></thead>
      <tbody>
      {% for d in cedolini %}
      <tr style="border-bottom:1px solid var(--border)">
        <td style="padding:12px 16px">
          <div style="display:flex;align-items:center;gap:10px">
            <div style="width:34px;height:34px;border-radius:50%;background:{{ d.colore }};display:flex;align-items:center;justify-content:center;color:#fff;font-size:12px;font-weight:700;flex-shrink:0">{{ d.iniziali }}</div>
            <div>
              <div style="font-weight:700">{{ d.nome }} {{ d.cognome }}</div>
              <div style="font-size:11px;color:#64748b">{{ d.mansione or '' }}</div>
            </div>
          </div>
        </td>
        <td style="padding:12px 16px;text-align:right;font-family:monospace;font-weight:700">{{ "%.1f"|format(d.ore) }}h</td>
        <td style="padding:12px 16px;text-align:right;font-size:12px;color:#64748b">
          {% if d.paga_base %}
            <span title="Paga base">€{{ "%.0f"|format(d.paga_base) }}/mese</span>
            {% if d.ore_soglia %}<span style="color:#94a3b8"> + {{ "%.0f"|format(d.ore_soglia) }}h soglia</span>{% endif %}
            {% if d.paga_oraria %}<br>+ € {{ "%.2f"|format(d.paga_oraria) }}/h extra</br>{% endif %}
          {% elif d.paga_oraria %}€ {{ "%.2f"|format(d.paga_oraria) }}/h
          {% else %}<span style="color:#dc2626">⚠ Non impostata</span>{% endif %}
        </td>
        <td style="padding:12px 16px;text-align:right;font-weight:800;font-size:16px;color:#10b981">
          {% if d.paga_oraria %}€ {{ "%.2f"|format(d.totale) }}{% else %}–{% endif %}
        </td>
        <td style="padding:12px 16px;text-align:right">
          {% if d.spese_rimborso > 0 %}
          <a href="/admin/spese?uid={{ d.id }}&stato=approvata" style="color:#f59e0b;font-weight:700;text-decoration:none">€ {{ "%.2f"|format(d.spese_rimborso) }}</a>
          {% else %}<span style="color:#94a3b8">–</span>{% endif %}
        </td>
        <td style="padding:12px 16px;text-align:right">
          {% if d.detrazioni > 0 %}
          <a href="/admin/detrazioni?uid={{ d.id }}&mese={{ prefisso }}"
             style="color:#dc2626;font-weight:700;text-decoration:none"
             title="{% for x in d.detr_dettaglio %}{{ x.descrizione }}: €{{ '%.2f'|format(x.importo) }}&#10;{% endfor %}">
            − € {{ "%.2f"|format(d.detrazioni) }}
          </a>
          {% else %}<span style="color:#94a3b8">–</span>{% endif %}
        </td>
        <td style="padding:12px 16px;text-align:right;font-weight:800;font-size:15px;color:#059669">
          {% if d.paga_oraria %}€ {{ "%.2f"|format(d.totale_con_spese) }}{% else %}–{% endif %}
        </td>
        <td style="padding:12px 16px">
          <div style="width:140px;background:#f1f5f9;border-radius:6px;overflow:hidden;position:relative">
            {% if max_ore > 0 %}
            <div class="chart-bar" style="width:{{ (d.ore/max_ore*100)|int }}%">&nbsp;</div>
            {% endif %}
          </div>
          <div style="font-size:11px;color:#94a3b8;margin-top:3px">{{ d.giorni_lavorati }} giorni</div>
        </td>
      </tr>
      {% endfor %}
      </tbody>
      <tfoot><tr style="background:#f0fdf4;font-weight:800">
        <td style="padding:12px 16px">TOTALE</td>
        <td style="padding:12px 16px;text-align:right;font-family:monospace">{{ "%.1f"|format(ore_totali_mese) }}h</td>
        <td></td>
        <td style="padding:12px 16px;text-align:right;color:#10b981;font-size:16px">€ {{ "%.2f"|format(totale_mese) }}</td>
        <td style="padding:12px 16px;text-align:right;color:#f59e0b">€ {{ "%.2f"|format(cedolini|sum(attribute='spese_rimborso')) }}</td>
        <td style="padding:12px 16px;text-align:right;color:#dc2626">− € {{ "%.2f"|format(cedolini|sum(attribute='detrazioni')) }}</td>
        <td style="padding:12px 16px;text-align:right;color:#059669;font-size:16px">€ {{ "%.2f"|format(cedolini|sum(attribute='totale_con_spese')) }}</td>
        <td></td>
      </tr></tfoot>
    </table>
    {% else %}
    <div style="padding:40px;text-align:center;color:#94a3b8">Nessuna presenza registrata per questo periodo.</div>
    {% endif %}
  </div>
</div>

{% if jolly_cedolini %}
<div class="card" style="margin-bottom:24px;border-left:4px solid #f59e0b">
  <div class="card-header" style="background:#fffbeb">
    <h3><i class="fa fa-user-clock" style="color:#d97706"></i> Lavoratori Jolly (Esterni) — {{ ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno','Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre'][mese_sel] }} {{ anno_sel }}</h3>
    <div style="display:flex;gap:8px;align-items:center">
      <span style="font-size:12px;color:#92400e;background:#fef3c7;border-radius:6px;padding:4px 10px">{{ jolly_cedolini|length }} jolly · {{ "%.1f"|format(jolly_cedolini|sum(attribute='ore')) }}h · € {{ "%.2f"|format(jolly_cedolini|sum(attribute='totale')) }}</span>
      <a href="/cedolini/jolly-paghe" class="btn btn-sm" style="background:#fef3c7;color:#92400e;border:1px solid #fde047"><i class="fa fa-coins"></i> Paga oraria jolly</a>
    </div>
  </div>
  <div class="card-body" style="padding:0">
    <table style="width:100%;border-collapse:collapse;font-size:13.5px">
      <thead><tr style="background:#fef9c3;border-bottom:2px solid #fde047">
        <th style="padding:12px 16px;text-align:left">Nome</th>
        <th style="padding:12px 16px;text-align:right">Ore totali</th>
        <th style="padding:12px 16px;text-align:center">Giorni</th>
        <th style="padding:12px 16px;text-align:right">€/h</th>
        <th style="padding:12px 16px;text-align:right;color:#d97706">Totale €</th>
        <th style="padding:12px 16px;text-align:left">Cantiere/i</th>
        <th style="padding:12px 16px;text-align:center">Azioni</th>
      </tr></thead>
      <tbody>
      {% for j in jolly_cedolini %}
      <!-- Riga riepilogo jolly -->
      <tr style="border-bottom:1px solid #fef3c7;background:#fffbeb" id="jolly-row-{{ loop.index }}">
        <td style="padding:10px 16px">
          <div style="display:flex;align-items:center;gap:10px">
            <div style="width:34px;height:34px;border-radius:50%;background:linear-gradient(135deg,#d97706,#f59e0b);display:flex;align-items:center;justify-content:center;color:#fff;font-size:12px;font-weight:700;flex-shrink:0">{{ j.iniziali }}</div>
            <div>
              <div style="font-weight:700" id="jolly-nome-{{ loop.index }}">{{ j.nome }} {{ j.cognome }}</div>
              <span style="font-size:10px;background:#fef3c7;color:#92400e;border-radius:4px;padding:1px 6px;font-weight:700">👷 JOLLY</span>
            </div>
          </div>
        </td>
        <td style="padding:10px 16px;text-align:right;font-family:monospace;font-weight:700;color:#d97706">{{ "%.1f"|format(j.ore) }}h</td>
        <td style="padding:10px 16px;text-align:center"><strong>{{ j.giorni_lavorati }}</strong></td>
        <td style="padding:10px 16px;text-align:right;font-size:12px;color:#64748b">
          {% if j.paga_oraria %}€ {{ "%.2f"|format(j.paga_oraria) }}{% else %}<a href="/cedolini/jolly-paghe" style="color:#f59e0b;font-size:11px">⚠ imposta</a>{% endif %}
        </td>
        <td style="padding:10px 16px;text-align:right;font-weight:800;font-size:15px;color:#d97706">
          {% if j.paga_oraria %}€ {{ "%.2f"|format(j.totale) }}{% else %}<span style="color:#94a3b8">–</span>{% endif %}
        </td>
        <td style="padding:10px 16px;font-size:12px">{% if j.cantieri %}<span class="tag">{{ j.cantieri }}</span>{% else %}<span style="color:#94a3b8">–</span>{% endif %}</td>
        <td style="padding:10px 16px;text-align:center">
          <div style="display:flex;gap:6px;justify-content:center">
            <button onclick="toggleJollyDettaglio({{ loop.index }})" class="btn btn-sm btn-secondary" title="Vedi/nascondi presenze"><i class="fa fa-list"></i></button>
            <button onclick="apriRinominaJolly('{{ j.nome }}','{{ j.cognome }}')" class="btn btn-sm" style="background:#fef3c7;color:#92400e;border:1px solid #fde047" title="Rinomina questo jolly in tutte le presenze del mese"><i class="fa fa-user-pen"></i> Rinomina</button>
          </div>
        </td>
      </tr>
      <!-- Righe dettaglio presenze jolly (nascoste di default) -->
      <tr id="jolly-dettaglio-{{ loop.index }}" style="display:none">
        <td colspan="5" style="padding:0;background:#fffdf0">
          <table style="width:100%;border-collapse:collapse;font-size:13px">
            <thead><tr style="background:#fef3c7">
              <th style="padding:8px 20px;text-align:left;font-size:11px;color:#92400e">DATA</th>
              <th style="padding:8px 16px;font-size:11px;color:#92400e">CANTIERE</th>
              <th style="padding:8px 16px;text-align:center;font-size:11px;color:#92400e">ENTRATA</th>
              <th style="padding:8px 16px;text-align:center;font-size:11px;color:#92400e">USCITA</th>
              <th style="padding:8px 16px;text-align:center;font-size:11px;color:#92400e">ORE</th>
              <th style="padding:8px 16px;text-align:center;font-size:11px;color:#92400e">AZIONI</th>
            </tr></thead>
            <tbody>
            {% for p in j.presenze %}
            <tr style="border-bottom:1px solid #fef9c3" id="jolly-p-{{ p.id }}">
              <td style="padding:8px 20px;font-family:monospace;font-weight:600">{{ p.data }}</td>
              <td style="padding:8px 16px;font-size:12px">{{ p.cantiere_nome or '–' }}</td>
              <td style="padding:8px 16px;text-align:center;font-family:monospace;color:#16a34a">{{ p.ora_entrata or '–' }}</td>
              <td style="padding:8px 16px;text-align:center;font-family:monospace">{{ p.ora_uscita or '–' }}</td>
              <td style="padding:8px 16px;text-align:center;font-weight:700;color:#d97706">{{ "%.1f"|format(p.ore_totali) if p.ore_totali else '–' }}h</td>
              <td style="padding:8px 16px;text-align:center">
                <div style="display:flex;gap:4px;justify-content:center">
                  <button onclick="apriModificaJolly({{ p.id }},'{{ p.data }}','{{ p.ora_entrata or '' }}','{{ p.ora_uscita or '' }}','{{ p.ore_totali or '' }}','{{ p.cantiere_id or '' }}','{{ p.note or '' }}','{{ j.nome }}','{{ j.cognome }}')"
                          class="btn btn-sm btn-secondary" title="Modifica"><i class="fa fa-pen"></i></button>
                  <button onclick="apriDuplicaJolly({{ p.id }},'{{ p.data }}','{{ p.ora_entrata or '' }}','{{ p.ora_uscita or '' }}','{{ p.ore_totali or '' }}','{{ p.cantiere_id or '' }}','{{ p.note or '' }}','{{ j.nome }}','{{ j.cognome }}')"
                          class="btn btn-sm" style="background:#eff6ff;color:#2563eb;border:1px solid #bfdbfe" title="Duplica e modifica"><i class="fa fa-copy"></i></button>
                  <a href="/presenze/{{ p.id }}/elimina" class="btn btn-sm btn-danger" onclick="return confirm('Eliminare?')" title="Elimina"><i class="fa fa-trash"></i></a>
                </div>
              </td>
            </tr>
            {% endfor %}
            </tbody>
          </table>
        </td>
      </tr>
      {% endfor %}
      </tbody>
      <tfoot><tr style="background:#fef3c7;font-weight:800">
        <td style="padding:12px 16px">TOTALE JOLLY</td>
        <td style="padding:12px 16px;text-align:right;font-family:monospace;color:#d97706">{{ "%.1f"|format(jolly_cedolini|sum(attribute='ore')) }}h</td>
        <td style="padding:12px 16px;text-align:center">{{ jolly_cedolini|sum(attribute='giorni_lavorati') }}</td>
        <td></td>
        <td style="padding:12px 16px;text-align:right;color:#d97706;font-size:15px">€ {{ "%.2f"|format(jolly_cedolini|sum(attribute='totale')) }}</td>
        <td colspan="2"></td>
      </tr></tfoot>
    </table>
  </div>
</div>

<!-- Modal modifica/duplica presenza jolly -->
<div id="modal-jolly-mod" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:1001;align-items:center;justify-content:center">
  <div class="card" style="width:480px;max-width:95vw">
    <div class="card-header" style="background:#fffbeb">
      <h3 style="color:#92400e" id="jolly-modal-titolo"><i class="fa fa-user-clock"></i> Modifica presenza Jolly</h3>
      <button onclick="document.getElementById('modal-jolly-mod').style.display='none'" style="background:none;border:none;font-size:22px;cursor:pointer">×</button>
    </div>
    <form method="POST" id="form-jolly-mod">
      <div class="card-body">
        <input type="hidden" name="pid" id="jmod-pid">
        <div style="background:#eff6ff;border-radius:8px;padding:8px 12px;font-size:12px;color:#1d4ed8;margin-bottom:12px" id="jmod-info-duplica" style="display:none">
          <i class="fa fa-copy"></i> Stai creando una <strong>nuova presenza</strong> copiando i dati — modifica ciò che serve e salva.
        </div>
        <div class="form-row">
          <div class="form-group"><label>Nome *</label><input name="nome_jolly" id="jmod-nome" required></div>
          <div class="form-group"><label>Cognome *</label><input name="cognome_jolly" id="jmod-cognome" required></div>
        </div>
        <div class="form-row">
          <div class="form-group"><label>Data *</label><input type="date" name="data" id="jmod-data" required></div>
          <div class="form-group"><label>Ore totali</label><input type="number" step="0.5" min="0.5" max="24" name="ore_dirette" id="jmod-ore" placeholder="es. 8"></div>
        </div>
        <div class="form-row">
          <div class="form-group"><label>Entrata</label><input type="time" name="ora_entrata" id="jmod-oe"></div>
          <div class="form-group"><label>Uscita</label><input type="time" name="ora_uscita" id="jmod-ou"></div>
        </div>
        <div class="form-group"><label>Fiera</label>
          <select name="cantiere_id" id="jmod-cant">
            <option value="">— nessuno —</option>
            {% for c in cantieri_list %}<option value="{{ c.id }}">{{ c.nome }}</option>{% endfor %}
          </select>
        </div>
        <div class="form-group"><label>Note</label><input name="note" id="jmod-note" placeholder="Opzionale"></div>
      </div>
      <div style="display:flex;gap:10px;justify-content:flex-end;padding:0 20px 16px">
        <button type="button" onclick="document.getElementById('modal-jolly-mod').style.display='none'" class="btn btn-secondary">Annulla</button>
        <button type="submit" class="btn btn-primary" id="jmod-btn-submit"><i class="fa fa-save"></i> Salva</button>
      </div>
    </form>
  </div>
</div>

<!-- Modal rinomina jolly -->
<div id="modal-jolly-rinomina" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:1001;align-items:center;justify-content:center">
  <div class="card" style="width:420px;max-width:95vw">
    <div class="card-header" style="background:#fffbeb">
      <h3 style="color:#92400e"><i class="fa fa-user-pen"></i> Rinomina Jolly</h3>
      <button onclick="document.getElementById('modal-jolly-rinomina').style.display='none'" style="background:none;border:none;font-size:22px;cursor:pointer">×</button>
    </div>
    <form method="POST" action="/presenze/jolly/rinomina">
      <div class="card-body">
        <div style="background:#fef3c7;border-radius:8px;padding:10px;font-size:12px;color:#92400e;margin-bottom:14px">
          <i class="fa fa-info-circle"></i> Aggiorna nome e cognome in <strong>tutte</strong> le presenze di questo jolly nel mese selezionato.
        </div>
        <input type="hidden" name="nome_old" id="jrin-nome-old">
        <input type="hidden" name="cognome_old" id="jrin-cognome-old">
        <input type="hidden" name="mese" value="{{ prefisso }}">
        <div class="form-row">
          <div class="form-group"><label>Nuovo Nome *</label><input name="nome_new" id="jrin-nome-new" required></div>
          <div class="form-group"><label>Nuovo Cognome *</label><input name="cognome_new" id="jrin-cognome-new" required></div>
        </div>
      </div>
      <div style="display:flex;gap:10px;justify-content:flex-end;padding:0 20px 16px">
        <button type="button" onclick="document.getElementById('modal-jolly-rinomina').style.display='none'" class="btn btn-secondary">Annulla</button>
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Rinomina</button>
      </div>
    </form>
  </div>
</div>

<script>
function toggleJollyDettaglio(idx) {
  var r = document.getElementById('jolly-dettaglio-' + idx);
  r.style.display = r.style.display === 'none' ? 'table-row' : 'none';
}
function _popolaJollyModal(pid, data, oe, ou, ore, cid, note, nome, cognome) {
  document.getElementById('jmod-pid').value = pid;
  document.getElementById('jmod-data').value = data;
  document.getElementById('jmod-nome').value = nome;
  document.getElementById('jmod-cognome').value = cognome;
  document.getElementById('jmod-oe').value = oe ? oe.substring(0,5) : '';
  document.getElementById('jmod-ou').value = ou ? ou.substring(0,5) : '';
  document.getElementById('jmod-ore').value = ore ? parseFloat(ore).toFixed(1) : '';
  document.getElementById('jmod-note').value = note;
  var sel = document.getElementById('jmod-cant');
  for (var i=0;i<sel.options.length;i++) sel.options[i].selected=(sel.options[i].value==cid);
}
function apriModificaJolly(pid, data, oe, ou, ore, cid, note, nome, cognome) {
  _popolaJollyModal(pid, data, oe, ou, ore, cid, note, nome, cognome);
  document.getElementById('form-jolly-mod').action = '/presenze/jolly/modifica';
  document.getElementById('jolly-modal-titolo').innerHTML = '<i class="fa fa-user-clock"></i> Modifica presenza Jolly';
  document.getElementById('jmod-info-duplica').style.display = 'none';
  document.getElementById('jmod-btn-submit').innerHTML = '<i class="fa fa-save"></i> Salva modifiche';
  document.getElementById('jmod-btn-submit').className = 'btn btn-primary';
  document.getElementById('modal-jolly-mod').style.display='flex';
}
function apriDuplicaJolly(pid, data, oe, ou, ore, cid, note, nome, cognome) {
  _popolaJollyModal(pid, data, oe, ou, ore, cid, note, nome, cognome);
  // Suggerisci giorno successivo
  try {
    var d = new Date(data + 'T00:00:00');
    d.setDate(d.getDate() + 1);
    document.getElementById('jmod-data').value = d.toISOString().split('T')[0];
  } catch(e) {}
  document.getElementById('form-jolly-mod').action = '/presenze/jolly/duplica-salva';
  document.getElementById('jolly-modal-titolo').innerHTML = '<i class="fa fa-copy"></i> Duplica presenza Jolly';
  document.getElementById('jmod-info-duplica').style.display = 'block';
  document.getElementById('jmod-btn-submit').innerHTML = '<i class="fa fa-copy"></i> Crea nuova presenza';
  document.getElementById('jmod-btn-submit').className = 'btn' ;
  document.getElementById('jmod-btn-submit').style.cssText = 'background:linear-gradient(135deg,#1d4ed8,#3b82f6);color:#fff';
  document.getElementById('modal-jolly-mod').style.display='flex';
}
function apriRinominaJolly(nome, cognome) {
  document.getElementById('jrin-nome-old').value = nome;
  document.getElementById('jrin-cognome-old').value = cognome;
  document.getElementById('jrin-nome-new').value = nome;
  document.getElementById('jrin-cognome-new').value = cognome;
  document.getElementById('modal-jolly-rinomina').style.display='flex';
}
</script>
{% endif %}

<!-- Grafici -->
<div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:24px">
  <!-- Grafico mensile ultimi 6 mesi -->
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-chart-bar" style="color:#6366f1"></i> Costi ultimi 6 mesi</h3></div>
    <div class="card-body">
      {% if storico_mesi %}
      {% set max_v = storico_mesi|map(attribute='totale')|max %}
      {% for sm in storico_mesi %}
      <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px">
        <div style="width:50px;font-size:12px;color:#64748b;text-align:right;flex-shrink:0">{{ sm.etichetta }}</div>
        <div style="flex:1;background:#f1f5f9;border-radius:6px;overflow:hidden">
          <div style="height:22px;border-radius:6px;background:linear-gradient(90deg,#6366f1,#8b5cf6);width:{{ ((sm.totale/max_v*100) if max_v > 0 else 0)|int }}%;min-width:{% if sm.totale > 0 %}4{% else %}0{% endif %}px"></div>
        </div>
        <div style="width:70px;font-size:12px;font-weight:700;text-align:right">€ {{ "%.0f"|format(sm.totale) }}</div>
      </div>
      {% endfor %}
      {% else %}
      <p style="color:#94a3b8;text-align:center">Dati insufficienti</p>
      {% endif %}
    </div>
  </div>

  <!-- Grafico per cantiere/evento -->
  <div class="card">
    <div class="card-header"><h3><i class="fa fa-layer-group" style="color:#f59e0b"></i> Costi per cantiere/evento (mese)</h3></div>
    <div class="card-body">
      {% if costi_per_cantiere %}
      {% set max_c = costi_per_cantiere|map(attribute='totale')|max %}
      {% for cc in costi_per_cantiere %}
      <div style="margin-bottom:12px">
        <div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:4px">
          <span style="font-weight:600;color:#1e293b">{{ cc.nome }}</span>
          <span style="font-weight:700;color:#f59e0b">€ {{ "%.0f"|format(cc.totale) }} / {{ "%.1f"|format(cc.ore) }}h</span>
        </div>
        <div style="background:#f1f5f9;border-radius:6px;overflow:hidden">
          <div style="height:18px;border-radius:6px;background:linear-gradient(90deg,#f59e0b,#fbbf24);width:{{ ((cc.totale/max_c*100) if max_c > 0 else 0)|int }}%;min-width:{% if cc.totale > 0 %}4{% else %}0{% endif %}px"></div>
        </div>
      </div>
      {% endfor %}
      {% else %}
      <p style="color:#94a3b8;text-align:center">Nessuna fiera con ore registrate.</p>
      {% endif %}
    </div>
  </div>
</div>

<!-- Spese extra per evento nel mese -->
{% if spese_per_evento %}
<div class="card">
  <div class="card-header"><h3><i class="fa fa-receipt" style="color:#ec4899"></i> Spese extra per evento — {{ ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno','Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre'][mese_sel] }} {{ anno_sel }}</h3></div>
  <div class="card-body" style="padding:0">
    <!-- Barre grafiche per evento -->
    {% set max_tot = spese_per_evento|map(attribute='totale')|max %}
    {% for ev in spese_per_evento %}
    <div style="padding:16px 20px;border-bottom:1px solid #fce7f3">
      <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:10px;flex-wrap:wrap;margin-bottom:10px">
        <div>
          <a href="/calendario/{{ ev.id }}" style="font-weight:700;font-size:14px;color:var(--accent2);text-decoration:none">{{ ev.titolo }}</a>
          {% if ev.cantiere %}<span style="margin-left:8px;background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:99px;font-size:11px;font-weight:700">{{ ev.cantiere }}</span>{% endif %}
          {% if ev.tipo %}<span style="margin-left:4px;background:#f1f5f9;color:#64748b;padding:2px 8px;border-radius:99px;font-size:11px">{{ ev.tipo }}</span>{% endif %}
          <div style="font-size:12px;color:#94a3b8;margin-top:3px">
            {{ ev.data_inizio }}{% if ev.data_fine and ev.data_fine != ev.data_inizio %} → {{ ev.data_fine }}{% endif %}
            {% if ev.team %}· {{ ev.team|length }} persone · {{ ev.ore_dip }}h totali{% endif %}
          </div>
          {% if ev.team %}
          <div style="margin-top:5px;display:flex;flex-wrap:wrap;gap:5px">
            {% for t in ev.team %}
            <span style="background:#eef2ff;color:#4338ca;padding:2px 8px;border-radius:99px;font-size:11px">{{ t.nome }} {{ t.ore }}h</span>
            {% endfor %}
          </div>
          {% endif %}
        </div>
        <div style="text-align:right;flex-shrink:0">
          <div style="font-size:20px;font-weight:800">€ {{ "%.0f"|format(ev.totale) }}</div>
          <div style="font-size:11px;color:#94a3b8;margin-top:2px">
            <span style="color:#6366f1">€{{ "%.0f"|format(ev.costo_dip) }} dip.</span>
            {% if ev.spese_extra > 0 %} + <span style="color:#ec4899">€{{ "%.0f"|format(ev.spese_extra) }} spese</span>{% endif %}
          </div>
        </div>
      </div>
      <!-- Barra proporzionale -->
      <div style="background:#f1f5f9;border-radius:6px;overflow:hidden;height:10px">
        <div style="height:10px;border-radius:6px;background:linear-gradient(90deg,#6366f1,#ec4899);width:{{ ((ev.totale/max_tot*100) if max_tot>0 else 0)|int }}%"></div>
      </div>
    </div>
    {% endfor %}
  </div>
</div>
{% endif %}
"""

PAGA_FORM_TMPL = """
<div class="card" style="max-width:760px;margin:0 auto">
  <div class="card-header"><h3><i class="fa fa-gear" style="color:#6366f1"></i> Impostazioni Retribuzione Dipendenti</h3></div>
  <div class="card-body">
    <div style="background:#eef4ff;border:1px solid #bfdbfe;border-radius:10px;padding:14px;margin-bottom:20px;font-size:13px;color:#1d4ed8">
      <strong>Logica calcolo:</strong> Se impostata una <em>paga base mensile</em>, quella viene applicata fino alle ore soglia.
      Le ore eccedenti vengono moltiplicate per la <em>paga oraria</em> e sommate alla base.
      Se non c'è paga base, si usa solo la tariffa oraria × ore totali.
    </div>
    <form method="POST">
      <table style="width:100%;border-collapse:collapse;font-size:13px">
        <thead><tr style="background:#f8fafc;border-bottom:2px solid var(--border)">
          <th style="padding:10px 14px;text-align:left">Dipendente</th>
          <th style="padding:10px 14px;text-align:center">Paga base netta<br><small style="font-weight:400;color:#94a3b8">€/mese fisso</small></th>
          <th style="padding:10px 14px;text-align:center">Ore soglia<br><small style="font-weight:400;color:#94a3b8">incluse nella base</small></th>
          <th style="padding:10px 14px;text-align:center">Paga oraria netta<br><small style="font-weight:400;color:#94a3b8">€/h extra soglia</small></th>
        </tr></thead>
        <tbody>
        {% for d in dipendenti %}
        <tr style="border-bottom:1px solid var(--border)">
          <td style="padding:10px 14px">
            <div style="display:flex;align-items:center;gap:10px">
              <div style="width:32px;height:32px;border-radius:50%;background:#6366f1;display:flex;align-items:center;justify-content:center;color:#fff;font-size:11px;font-weight:700;flex-shrink:0">{{ (d.nome[0]+d.cognome[0])|upper }}</div>
              <div style="font-weight:600">{{ d.nome }} {{ d.cognome }}<div style="font-size:11px;color:#94a3b8">{{ d.mansione or '' }}</div></div>
            </div>
          </td>
          <td style="padding:10px 14px;text-align:center">
            <div style="display:flex;align-items:center;gap:4px;justify-content:center">
              <span style="color:#64748b">€</span>
              <input type="number" name="base_{{ d.id }}" step="0.01" min="0" value="{{ d.paga_base_netta or '' }}" placeholder="es. 800" style="width:100px;padding:6px 8px;border:1px solid var(--border);border-radius:8px;font-size:13px;text-align:right">
            </div>
          </td>
          <td style="padding:10px 14px;text-align:center">
            <div style="display:flex;align-items:center;gap:4px;justify-content:center">
              <input type="number" name="soglia_{{ d.id }}" step="0.5" min="0" value="{{ d.ore_soglia_mensile or '' }}" placeholder="es. 160" style="width:80px;padding:6px 8px;border:1px solid var(--border);border-radius:8px;font-size:13px;text-align:right">
              <span style="color:#64748b">h</span>
            </div>
          </td>
          <td style="padding:10px 14px;text-align:center">
            <div style="display:flex;align-items:center;gap:4px;justify-content:center">
              <span style="color:#64748b">€</span>
              <input type="number" name="paga_{{ d.id }}" step="0.01" min="0" value="{{ d.paga_oraria_netta or '' }}" placeholder="es. 10" style="width:80px;padding:6px 8px;border:1px solid var(--border);border-radius:8px;font-size:13px;text-align:right">
              <span style="color:#64748b">/h</span>
            </div>
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
      <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:20px">
        <a href="/cedolini" class="btn btn-secondary">Annulla</a>
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva</button>
      </div>
    </form>
  </div>
</div>
"""

COLORI_DIP = ['#6366f1','#ec4899','#f59e0b','#10b981','#3b82f6','#8b5cf6','#06b6d4','#f97316']

ADMIN_SPESE_TMPL = """
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;flex-wrap:wrap;gap:10px">
  <form method="GET" style="display:flex;gap:8px;flex-wrap:wrap">
    <select name="stato" onchange="this.form.submit()" style="padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px">
      <option value="">Tutti gli stati</option>
      <option value="in_attesa" {{ 'selected' if filtro_stato=='in_attesa' }}>In attesa</option>
      <option value="approvata" {{ 'selected' if filtro_stato=='approvata' }}>Approvate</option>
      <option value="rifiutata" {{ 'selected' if filtro_stato=='rifiutata' }}>Rifiutate</option>
    </select>
    <select name="uid" onchange="this.form.submit()" style="padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px">
      <option value="">Tutti i dipendenti</option>
      {% for d in dipendenti %}<option value="{{ d.id }}" {{ 'selected' if filtro_uid==d.id|string }}>{{ d.nome }} {{ d.cognome }}</option>{% endfor %}
    </select>
  </form>
  <div style="display:flex;gap:10px;align-items:center">
    {% if totale_in_attesa > 0 %}
    <span style="background:#fffbeb;color:#92400e;border:1px solid #fde68a;border-radius:8px;padding:6px 14px;font-size:13px;font-weight:700">
      <i class="fa fa-clock"></i> € {{ "%.2f"|format(totale_in_attesa) }} da approvare
    </span>
    {% endif %}
    <a href="/admin/spese/nuova" class="btn btn-primary"><i class="fa fa-plus"></i> Nuova spesa</a>
  </div>
</div>

{% if not spese %}
<div class="card"><div class="empty-state"><i class="fa fa-receipt"></i><p>Nessuna spesa trovata.</p></div></div>
{% else %}
<div class="card" style="padding:0">
  <table style="width:100%;border-collapse:collapse">
    <thead><tr style="background:#0f172a;color:#fff">
      <th style="padding:10px 14px;font-size:12px;text-align:left">DIPENDENTE</th>
      <th style="padding:10px 14px;font-size:12px">DATA</th>
      <th style="padding:10px 14px;font-size:12px">CATEGORIA</th>
      <th style="padding:10px 14px;font-size:12px">DESCRIZIONE</th>
      <th style="padding:10px 14px;font-size:12px;text-align:right">IMPORTO</th>
      <th style="padding:10px 14px;font-size:12px">SCONTRINO</th>
      <th style="padding:10px 14px;font-size:12px">STATO</th>
      <th style="padding:10px 14px;font-size:12px"></th>
    </tr></thead>
    <tbody>
    {% for s in spese %}
    <tr style="border-bottom:1px solid var(--border);{% if s.stato=='in_attesa' %}background:#fffdf0{% endif %}">
      <td style="padding:10px 14px">
        <div style="font-weight:600;font-size:13px">{{ s.nome }} {{ s.cognome }}</div>
      </td>
      <td style="padding:10px 14px;font-family:monospace;font-size:12px">{{ s.data }}</td>
      <td style="padding:10px 14px">
        <span style="background:#f1f5f9;color:#475569;padding:2px 8px;border-radius:99px;font-size:12px;font-weight:600">{{ s.categoria }}</span>
      </td>
      <td style="padding:10px 14px;font-size:13px;color:#64748b;max-width:180px">{{ s.descrizione or '–' }}</td>
      <td style="padding:10px 14px;text-align:right;font-weight:800;font-size:15px">€ {{ "%.2f"|format(s.importo) }}</td>
      <td style="padding:10px 14px">
        {% if s.foto_nome %}
        <a href="/admin/spese/foto/{{ s.foto_nome }}" target="_blank"
           style="display:inline-flex;align-items:center;gap:5px;background:#eff6ff;color:#2563eb;border:1px solid #bfdbfe;border-radius:8px;padding:5px 10px;font-size:12px;text-decoration:none">
          <i class="fa fa-image"></i> Scontrino
        </a>
        {% else %}–{% endif %}
      </td>
      <td style="padding:10px 14px">
        {% if s.stato == 'in_attesa' %}<span style="background:#fffbeb;color:#d97706;padding:3px 10px;border-radius:99px;font-size:12px;font-weight:700">⏳ In attesa</span>
        {% elif s.stato == 'approvata' %}<span style="background:#f0fdf4;color:#15803d;padding:3px 10px;border-radius:99px;font-size:12px;font-weight:700">✅ Approvata</span>
        {% else %}<span style="background:#fef2f2;color:#dc2626;padding:3px 10px;border-radius:99px;font-size:12px;font-weight:700">✗ Rifiutata</span>{% endif %}
        {% if s.note_admin %}<div style="font-size:11px;color:#94a3b8;margin-top:3px">{{ s.note_admin }}</div>{% endif %}
      </td>
      <td style="padding:10px 14px">
        <div style="display:flex;gap:5px;flex-wrap:wrap;align-items:center">
          {% if s.stato == 'in_attesa' %}
          <a href="/admin/spese/{{ s.id }}/approva" style="background:#f0fdf4;color:#15803d;border:1px solid #86efac;border-radius:7px;padding:5px 10px;font-size:12px;font-weight:700;text-decoration:none" title="Approva"><i class="fa fa-check"></i></a>
          <a href="/admin/spese/{{ s.id }}/rifiuta-prompt" style="background:#fef2f2;color:#dc2626;border:1px solid #fecaca;border-radius:7px;padding:5px 10px;font-size:12px;font-weight:700;text-decoration:none" title="Rifiuta"><i class="fa fa-xmark"></i></a>
          {% endif %}
          <a href="/admin/spese/{{ s.id }}/modifica-form" style="background:#eff6ff;color:#2563eb;border:1px solid #bfdbfe;border-radius:7px;padding:5px 10px;font-size:12px;font-weight:700;text-decoration:none" title="Modifica"><i class="fa fa-pen"></i></a>
          <a href="/admin/spese/{{ s.id }}/elimina-confirm" style="background:#fef2f2;color:#dc2626;border:1px solid #fecaca;border-radius:7px;padding:5px 10px;font-size:12px;text-decoration:none" title="Elimina"><i class="fa fa-trash"></i></a>
        </div>
      </td>
    </tr>
    {% endfor %}
    </tbody>
    {% if totale_filtrato > 0 %}
    <tfoot><tr style="background:#f8fafc;font-weight:800">
      <td colspan="4" style="padding:10px 14px;font-size:13px">TOTALE ({{ spese|length }} spese)</td>
      <td style="padding:10px 14px;text-align:right;font-size:15px;color:#10b981">€ {{ "%.2f"|format(totale_filtrato) }}</td>
      <td colspan="3"></td>
    </tr></tfoot>
    {% endif %}
  </table>
</div>
{% endif %}
"""


@app.route('/admin/spese')
@admin_required
def admin_spese():
    db = get_db()
    filtro_stato = request.args.get('stato', '')
    filtro_uid = request.args.get('uid', '')
    sql = """SELECT s.*, u.nome, u.cognome FROM spese_rimborso s
             JOIN utenti u ON u.id=s.utente_id WHERE 1=1"""
    params = []
    if filtro_stato:
        sql += " AND s.stato=?"; params.append(filtro_stato)
    if filtro_uid:
        sql += " AND s.utente_id=?"; params.append(filtro_uid)
    sql += " ORDER BY CASE s.stato WHEN 'in_attesa' THEN 0 ELSE 1 END, s.data DESC"
    spese = [dict(r) for r in db.execute(sql, params).fetchall()]
    dipendenti = db.execute("SELECT id,nome,cognome FROM utenti WHERE ruolo='dipendente' AND attivo=1 ORDER BY cognome").fetchall()
    totale_in_attesa = sum(s['importo'] for s in spese if s['stato']=='in_attesa')
    totale_filtrato = sum(s['importo'] for s in spese)
    db.close()
    return render_page(ADMIN_SPESE_TMPL, page_title='Spese Rimborsabili', active='admin_spese',
        spese=spese, dipendenti=dipendenti,
        filtro_stato=filtro_stato, filtro_uid=filtro_uid,
        totale_in_attesa=totale_in_attesa, totale_filtrato=totale_filtrato,
        categorie_spesa=CATEGORIE_SPESA)


@app.route('/admin/spese/nuova', methods=['GET','POST'])
@admin_required
def admin_spesa_nuova():
    db = get_db()
    dipendenti = db.execute(
        "SELECT id,nome,cognome FROM utenti WHERE attivo=1 AND ruolo='dipendente' ORDER BY cognome,nome"
    ).fetchall()
    veicoli = db.execute("SELECT id,targa,marca,modello FROM veicoli WHERE attivo=1 ORDER BY targa").fetchall()

    if request.method == 'POST':
        uid       = int(request.form.get('utente_id'))
        data      = request.form.get('data') or date.today().isoformat()
        categoria = request.form.get('categoria','Altro')
        descrizione = request.form.get('descrizione','')
        importo   = float(request.form.get('importo') or 0)
        veicolo_id = request.form.get('veicolo_id') or None
        stato     = request.form.get('stato','approvata')  # admin inserisce direttamente come approvata
        note_admin = request.form.get('note_admin','')

        # Foto opzionale
        foto_nome = foto_path = None
        f = request.files.get('foto')
        if f and f.filename:
            ext = f.filename.rsplit('.',1)[-1].lower()
            if ext in {'jpg','jpeg','png','gif','webp','pdf'}:
                foto_nome = datetime.now().strftime('%Y%m%d_%H%M%S_') + secure_filename(f.filename)
                foto_path = os.path.join(get_spese_upload_path(), foto_nome)
                f.save(foto_path)

        db.execute(
            "INSERT INTO spese_rimborso (utente_id,data,categoria,descrizione,importo,veicolo_id,foto_nome,foto_path,stato,note_admin) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (uid, data, categoria, descrizione, importo, veicolo_id, foto_nome, foto_path, stato, note_admin or None)
        )
        safe_commit(db); db.close()
        flash(f'Spesa inserita e {stato}!', 'success')
        return redirect(url_for('admin_spese'))

    db.close()
    tmpl = f"""
<div style="margin-bottom:16px">
  <a href="/admin/spese" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Spese</a>
</div>
<div class="card" style="max-width:580px;margin:0 auto">
  <div class="card-header">
    <h3><i class="fa fa-plus" style="color:var(--accent)"></i> Nuova Rimborso Spesa</h3>
    <div style="font-size:12px;color:var(--text-light);margin-top:2px">Inserimento manuale da admin — viene approvata direttamente</div>
  </div>
  <div class="card-body">
    <form method="POST" enctype="multipart/form-data">
      <div class="form-row">
        <div class="form-group">
          <label>Dipendente *</label>
          <select name="utente_id" required>
            <option value="">— Seleziona —</option>
            {''.join(f'<option value="{d["id"]}">{d["nome"]} {d["cognome"]}</option>' for d in dipendenti)}
          </select>
        </div>
        <div class="form-group">
          <label>Data *</label>
          <input type="date" name="data" value="{date.today().isoformat()}" required>
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>Categoria *</label>
          <select name="categoria" required>
            {''.join(f'<option value="{c}">{c}</option>' for c in CATEGORIE_SPESA)}
          </select>
        </div>
        <div class="form-group">
          <label>Importo (€) *</label>
          <input type="number" step="0.01" min="0.01" name="importo" required
                 placeholder="es. 25.50" style="font-size:18px;font-weight:700;color:var(--accent)">
        </div>
      </div>
      <div class="form-group">
        <label>Descrizione</label>
        <input name="descrizione" placeholder="es. Pranzo trasferta Milano">
      </div>
      <div class="form-group">
        <label>Veicolo (per rimborsi carburante)</label>
        <select name="veicolo_id">
          <option value="">— Nessuno —</option>
          {''.join(f'<option value="{v["id"]}">{v["targa"]} — {v["marca"]} {v["modello"]}</option>' for v in veicoli)}
        </select>
      </div>
      <div class="form-group">
        <label>Stato</label>
        <select name="stato">
          <option value="approvata" selected>✅ Approvata</option>
          <option value="in_attesa">⏳ In attesa</option>
          <option value="rifiutata">❌ Rifiutata</option>
        </select>
      </div>
      <div class="form-group">
        <label>Scontrino / Documento <span style="color:var(--text-light);font-weight:400;font-size:12px">(opzionale)</span></label>
        <input type="file" name="foto" accept=".jpg,.jpeg,.png,.gif,.webp,.pdf">
      </div>
      <div class="form-group">
        <label>Note admin</label>
        <input name="note_admin" placeholder="Note interne (opzionale)">
      </div>
      <div style="display:flex;gap:10px;justify-content:flex-end;margin-top:8px">
        <a href="/admin/spese" class="btn btn-secondary">Annulla</a>
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva spesa</button>
      </div>
    </form>
  </div>
</div>"""
    return render_page(tmpl, page_title='Nuova Spesa', active='admin_spese')


TIPI_DETRAZIONE = [
    ('anticipo', '💵 Anticipo stipendio'),
    ('acquisto', '🛒 Acquisto per dipendente'),
    ('multa',    '⚠️ Multa / penale'),
    ('prestito', '🤝 Restituzione prestito'),
    ('altro',    '📌 Altro'),
]

@app.route('/admin/detrazioni')
@admin_required
def admin_detrazioni():
    db = get_db()
    uid_f  = request.args.get('uid', '')
    mese_f = request.args.get('mese', '')
    q = """SELECT d.*, u.nome, u.cognome
           FROM detrazioni_dipendente d JOIN utenti u ON u.id=d.utente_id
           WHERE 1=1"""
    params = []
    if uid_f:  q += " AND d.utente_id=?"; params.append(int(uid_f))
    if mese_f: q += " AND d.mese_competenza=?"; params.append(mese_f)
    q += " ORDER BY d.data DESC"
    detrazioni = db.execute(q, params).fetchall()
    dipendenti = db.execute("SELECT id,nome,cognome FROM utenti WHERE attivo=1 AND ruolo='dipendente' ORDER BY cognome").fetchall()
    totale = sum(d['importo'] for d in detrazioni)
    db.close()

    tipi_map = dict(TIPI_DETRAZIONE)
    tmpl = f"""
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;flex-wrap:wrap;gap:10px">
  <form method="GET" style="display:flex;gap:8px;flex-wrap:wrap">
    <select name="uid" onchange="this.form.submit()" style="padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px">
      <option value="">Tutti i dipendenti</option>
      {chr(10).join(f'<option value="{d["id"]}" {"selected" if uid_f==str(d["id"]) else ""}>{d["nome"]} {d["cognome"]}</option>' for d in dipendenti)}
    </select>
    <input type="month" name="mese" value="{mese_f}" style="padding:7px 12px;border:1px solid var(--border);border-radius:8px;font-size:13px" onchange="this.form.submit()" placeholder="Filtra per mese">
    <a href="/admin/detrazioni" class="btn btn-secondary btn-sm">Reset</a>
  </form>
  <div style="display:flex;gap:10px;align-items:center">
    {f'<span style="background:#fef2f2;color:#dc2626;border:1px solid #fecaca;border-radius:8px;padding:6px 14px;font-size:13px;font-weight:700"><i class=\'fa fa-minus-circle\'></i> € {totale:.2f} totale detrazioni</span>' if detrazioni else ''}
    <a href="/admin/detrazioni/nuova" class="btn btn-primary"><i class="fa fa-plus"></i> Nuova detrazione</a>
  </div>
</div>

{'''<div class="card"><div class="empty-state"><i class="fa fa-coins"></i><p>Nessuna detrazione trovata.</p></div></div>''' if not detrazioni else f'''
<div class="card" style="padding:0">
  <table style="width:100%;border-collapse:collapse">
    <thead><tr style="background:#0f172a;color:#fff">
      <th style="padding:10px 14px;font-size:12px">DIPENDENTE</th>
      <th style="padding:10px 14px;font-size:12px">DATA</th>
      <th style="padding:10px 14px;font-size:12px">TIPO</th>
      <th style="padding:10px 14px;font-size:12px">DESCRIZIONE</th>
      <th style="padding:10px 14px;font-size:12px">MESE CEDOLINO</th>
      <th style="padding:10px 14px;font-size:12px;text-align:right">IMPORTO</th>
      <th style="padding:10px 14px;font-size:12px"></th>
    </tr></thead>
    <tbody>
    {"".join(f'''<tr style="border-bottom:1px solid var(--border)">
      <td style="padding:10px 14px"><span class="avatar-sm">{d["nome"][0]}{d["cognome"][0]}</span><strong>{d["nome"]} {d["cognome"]}</strong></td>
      <td style="padding:10px 14px;font-family:monospace;font-size:12px">{d["data"]}</td>
      <td style="padding:10px 14px"><span style="background:#fef2f2;color:#dc2626;padding:2px 8px;border-radius:99px;font-size:12px;font-weight:600">{tipi_map.get(d["tipo"], d["tipo"])}</span></td>
      <td style="padding:10px 14px;font-size:13px;color:#64748b">{d["descrizione"] or "–"}</td>
      <td style="padding:10px 14px;font-size:12px;font-family:monospace">{d["mese_competenza"] or "–"}</td>
      <td style="padding:10px 14px;text-align:right;font-weight:800;font-size:15px;color:#dc2626">− € {d["importo"]:.2f}</td>
      <td style="padding:10px 14px">
        <a href="/admin/detrazioni/{d["id"]}/elimina" class="btn btn-danger btn-sm" onclick="return confirm(\'Eliminare questa detrazione?\')"><i class="fa fa-trash"></i></a>
      </td>
    </tr>''' for d in detrazioni)}
    </tbody>
  </table>
</div>'''}
"""
    return render_page(tmpl, page_title='Detrazioni & Anticipi', active='cedolini',
                       dipendenti=dipendenti)


@app.route('/admin/detrazioni/nuova', methods=['GET','POST'])
@admin_required
def admin_detrazione_nuova():
    db = get_db()
    dipendenti = db.execute(
        "SELECT id,nome,cognome FROM utenti WHERE attivo=1 AND ruolo='dipendente' ORDER BY cognome,nome"
    ).fetchall()

    if request.method == 'POST':
        uid         = int(request.form.get('utente_id'))
        data        = request.form.get('data') or date.today().isoformat()
        tipo        = request.form.get('tipo','altro')
        descrizione = request.form.get('descrizione','')
        importo     = float(request.form.get('importo') or 0)
        mese_comp   = request.form.get('mese_competenza') or None
        note        = request.form.get('note','')
        db.execute(
            "INSERT INTO detrazioni_dipendente (utente_id,data,tipo,descrizione,importo,mese_competenza,note) VALUES (?,?,?,?,?,?,?)",
            (uid, data, tipo, descrizione, importo, mese_comp, note)
        )
        safe_commit(db); db.close()
        flash('Detrazione inserita!', 'success')
        return redirect(url_for('admin_detrazioni'))
    db.close()

    oggi = date.today()
    mese_default = oggi.strftime('%Y-%m')
    tipi_options = ''.join(f'<option value="{k}">{v}</option>' for k,v in TIPI_DETRAZIONE)
    dip_options  = ''.join(f'<option value="{d["id"]}">{d["nome"]} {d["cognome"]}</option>' for d in dipendenti)

    tmpl = f"""
<div style="margin-bottom:16px">
  <a href="/admin/detrazioni" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Detrazioni</a>
</div>
<div class="card" style="max-width:580px;margin:0 auto">
  <div class="card-header">
    <h3><i class="fa fa-minus-circle" style="color:#dc2626"></i> Nuova Detrazione / Anticipo</h3>
    <div style="font-size:12px;color:var(--text-light);margin-top:2px">Verrà scalata automaticamente dal cedolino del mese indicato</div>
  </div>
  <div class="card-body">
    <form method="POST">
      <div class="form-row">
        <div class="form-group">
          <label>Dipendente *</label>
          <select name="utente_id" required>
            <option value="">— Seleziona —</option>
            {dip_options}
          </select>
        </div>
        <div class="form-group">
          <label>Data operazione *</label>
          <input type="date" name="data" value="{oggi.isoformat()}" required>
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>Tipo *</label>
          <select name="tipo" required>{tipi_options}</select>
        </div>
        <div class="form-group">
          <label>Importo da detrarre (€) *</label>
          <input type="number" step="0.01" min="0.01" name="importo" required
                 placeholder="es. 150.00" style="font-size:18px;font-weight:800;color:#dc2626">
        </div>
      </div>
      <div class="form-group">
        <label>Descrizione *</label>
        <input name="descrizione" required placeholder="es. Anticipo stipendio Marzo, Ordine Amazon cuffie, Multa parcheggio...">
      </div>
      <div class="form-group">
        <label>Mese cedolino di competenza <span style="font-weight:400;color:var(--text-light);font-size:12px">(mese dal cui stipendio scalare)</span></label>
        <input type="month" name="mese_competenza" value="{mese_default}">
      </div>
      <div class="form-group">
        <label>Note interne</label>
        <textarea name="note" placeholder="Note per uso interno (non visibili al dipendente)" rows="2" style="resize:vertical"></textarea>
      </div>
      <div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:10px 14px;font-size:12px;color:#dc2626;margin-bottom:16px">
        <i class="fa fa-info-circle"></i> L'importo verrà sottratto automaticamente dal <strong>Totale netto</strong> nel cedolino del mese selezionato.
      </div>
      <div style="display:flex;gap:10px;justify-content:flex-end">
        <a href="/admin/detrazioni" class="btn btn-secondary">Annulla</a>
        <button type="submit" class="btn btn-danger"><i class="fa fa-save"></i> Salva detrazione</button>
      </div>
    </form>
  </div>
</div>"""
    return render_page(tmpl, page_title='Nuova Detrazione', active='cedolini')


@app.route('/admin/detrazioni/<int:did>/elimina')
@admin_required
def admin_detrazione_elimina(did):
    db = get_db()
    db.execute("DELETE FROM detrazioni_dipendente WHERE id=?", (did,))
    safe_commit(db); db.close()
    flash('Detrazione eliminata.', 'success')
    return redirect(url_for('admin_detrazioni'))


@app.route('/admin/spese/<int:sid>/approva')
@admin_required
def admin_spesa_approva(sid):
    db = get_db()
    db.execute("UPDATE spese_rimborso SET stato='approvata', note_admin=NULL WHERE id=?", (sid,))
    safe_commit(db); db.close()
    flash('Spesa approvata.', 'success')
    return redirect(url_for('admin_spese'))


@app.route('/admin/spese/<int:sid>/rifiuta-prompt', methods=['GET','POST'])
@admin_required
def admin_spesa_rifiuta_prompt(sid):
    db = get_db()
    spesa = db.execute("""SELECT s.*,u.nome,u.cognome FROM spese_rimborso s
        JOIN utenti u ON u.id=s.utente_id WHERE s.id=?""", (sid,)).fetchone()
    if not spesa:
        db.close(); flash('Spesa non trovata.','error'); return redirect(url_for('admin_spese'))
    if request.method == 'POST':
        note = request.form.get('note_admin','')
        db.execute("UPDATE spese_rimborso SET stato='rifiutata', note_admin=? WHERE id=?", (note, sid))
        safe_commit(db); db.close()
        flash('Spesa rifiutata.', 'success')
        return redirect(url_for('admin_spese'))
    db.close()
    tmpl = f"""
<div class="card" style="max-width:500px;margin:40px auto">
  <div class="card-header"><h3><i class="fa fa-xmark-circle" style="color:#dc2626"></i> Rifiuta spesa</h3></div>
  <div class="card-body">
    <p style="margin-bottom:16px;color:#64748b">
      Spesa di <strong>{spesa['nome']} {spesa['cognome']}</strong> —
      {spesa['categoria']} — <strong>€ {spesa['importo']:.2f}</strong> del {spesa['data']}
    </p>
    <form method="POST">
      <div class="form-group">
        <label>Motivo del rifiuto *</label>
        <textarea name="note_admin" class="form-control" rows="3" required placeholder="Specifica il motivo..."></textarea>
      </div>
      <div style="display:flex;gap:10px;margin-top:16px">
        <a href="/admin/spese" class="btn btn-secondary">Annulla</a>
        <button type="submit" class="btn btn-danger"><i class="fa fa-xmark"></i> Conferma rifiuto</button>
      </div>
    </form>
  </div>
</div>"""
    return render_page(tmpl, page_title='Rifiuta spesa', active='admin_spese')


@app.route('/admin/spese/<int:sid>/modifica-form', methods=['GET','POST'])
@admin_required
def admin_spesa_modifica_form(sid):
    db = get_db()
    spesa = db.execute("""SELECT s.*,u.nome,u.cognome FROM spese_rimborso s
        JOIN utenti u ON u.id=s.utente_id WHERE s.id=?""", (sid,)).fetchone()
    if not spesa:
        db.close(); flash('Spesa non trovata.','error'); return redirect(url_for('admin_spese'))
    if request.method == 'POST':
        db.execute("""UPDATE spese_rimborso SET data=?,categoria=?,importo=?,descrizione=?,stato=?
            WHERE id=?""", (
            request.form.get('data'),
            request.form.get('categoria'),
            float(request.form.get('importo') or 0),
            request.form.get('descrizione',''),
            request.form.get('stato','in_attesa'),
            sid))
        safe_commit(db); db.close()
        flash('Spesa aggiornata.', 'success')
        return redirect(url_for('admin_spese'))
    cats_options = ''.join(
        f'<option value="{c}" {"selected" if c==spesa["categoria"] else ""}>{c}</option>'
        for c in CATEGORIE_SPESA)
    stati = [('in_attesa','⏳ In attesa'),('approvata','✅ Approvata'),('rifiutata','✗ Rifiutata')]
    stati_options = ''.join(
        f'<option value="{v}" {"selected" if v==spesa["stato"] else ""}>{l}</option>'
        for v,l in stati)
    db.close()
    tmpl = f"""
<div class="card" style="max-width:540px;margin:40px auto">
  <div class="card-header"><h3><i class="fa fa-pen" style="color:#2563eb"></i> Modifica spesa</h3></div>
  <div class="card-body">
    <p style="margin-bottom:20px;color:#64748b">
      Dipendente: <strong>{spesa['nome']} {spesa['cognome']}</strong>
    </p>
    <form method="POST">
      <div class="form-row">
        <div class="form-group">
          <label>Data *</label>
          <input type="date" name="data" class="form-control" value="{spesa['data']}" required>
        </div>
        <div class="form-group">
          <label>Importo (€) *</label>
          <input type="number" name="importo" class="form-control" step="0.01" min="0.01" value="{spesa['importo']}" required>
        </div>
      </div>
      <div class="form-group">
        <label>Categoria *</label>
        <select name="categoria" class="form-control" required>{cats_options}</select>
      </div>
      <div class="form-group">
        <label>Descrizione</label>
        <textarea name="descrizione" class="form-control" rows="2">{spesa['descrizione'] or ''}</textarea>
      </div>
      <div class="form-group">
        <label>Stato</label>
        <select name="stato" class="form-control">{stati_options}</select>
      </div>
      <div style="display:flex;gap:10px;margin-top:20px">
        <a href="/admin/spese" class="btn btn-secondary"><i class="fa fa-arrow-left"></i> Annulla</a>
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva modifiche</button>
      </div>
    </form>
  </div>
</div>"""
    return render_page(tmpl, page_title='Modifica spesa', active='admin_spese')


@app.route('/admin/spese/<int:sid>/elimina-confirm')
@admin_required
def admin_spesa_elimina_confirm(sid):
    db = get_db()
    spesa = db.execute("""SELECT s.*,u.nome,u.cognome FROM spese_rimborso s
        JOIN utenti u ON u.id=s.utente_id WHERE s.id=?""", (sid,)).fetchone()
    if not spesa:
        db.close(); flash('Spesa non trovata.','error'); return redirect(url_for('admin_spese'))
    db.close()
    tmpl = f"""
<div class="card" style="max-width:500px;margin:40px auto;border-left:4px solid #dc2626">
  <div class="card-header"><h3><i class="fa fa-trash" style="color:#dc2626"></i> Elimina spesa</h3></div>
  <div class="card-body">
    <p style="margin-bottom:20px">
      Stai per eliminare definitivamente la spesa di
      <strong>{spesa['nome']} {spesa['cognome']}</strong>:<br>
      <span style="font-size:15px;color:#dc2626;font-weight:700">
        {spesa['categoria']} — € {spesa['importo']:.2f} del {spesa['data']}
      </span>
    </p>
    <p style="color:#94a3b8;font-size:13px;margin-bottom:20px">⚠️ L'operazione non è reversibile. La spesa verrà rimossa anche da cedolini e report.</p>
    <div style="display:flex;gap:10px">
      <a href="/admin/spese" class="btn btn-secondary"><i class="fa fa-arrow-left"></i> Annulla</a>
      <a href="/admin/spese/{sid}/elimina-esegui" class="btn btn-danger"><i class="fa fa-trash"></i> Sì, elimina</a>
    </div>
  </div>
</div>"""
    return render_page(tmpl, page_title='Elimina spesa', active='admin_spese')


@app.route('/admin/spese/<int:sid>/elimina-esegui')
@admin_required
def admin_spesa_elimina_esegui(sid):
    db = get_db()
    spesa = db.execute("SELECT foto_path FROM spese_rimborso WHERE id=?", (sid,)).fetchone()
    if spesa and spesa['foto_path'] and os.path.exists(spesa['foto_path']):
        try: os.remove(spesa['foto_path'])
        except: pass
    db.execute("DELETE FROM spese_rimborso WHERE id=?", (sid,))
    safe_commit(db); db.close()
    flash('Spesa eliminata definitivamente.', 'success')
    return redirect(url_for('admin_spese'))




@app.route('/admin/spese/foto/<fn>')
@admin_required
def admin_spese_foto(fn):
    fp = os.path.join(get_spese_upload_path(), fn)
    if not os.path.exists(fp):
        return 'Not found', 404
    mt = mimetypes.guess_type(fn)[0] or 'image/jpeg'
    return send_file(fp, mimetype=mt)


@app.route('/cedolini')
@admin_required
def cedolini():
    from datetime import date as _d, datetime as _dt, timedelta as _td
    db  = get_db()
    oggi = _d.today()
    mese_sel = int(request.args.get('mese', oggi.month))
    anno_sel = int(request.args.get('anno', oggi.year))
    anni = list(range(oggi.year - 2, oggi.year + 1))
    prefisso = f'{anno_sel}-{mese_sel:02d}'

    def calc_paga(ore, paga_base, ore_soglia, paga_oraria):
        """Calcola totale netto: base fissa + ore extra × tariffa oraria."""
        if paga_base and paga_base > 0:
            ore_extra = max(0, ore - (ore_soglia or 0))
            return round(paga_base + ore_extra * (paga_oraria or 0), 2)
        return round(ore * (paga_oraria or 0), 2)

    # Dipendenti come dict (fix sqlite3.Row.get)
    dips_raw = db.execute(
        "SELECT * FROM utenti WHERE ruolo='dipendente' AND attivo=1 ORDER BY nome").fetchall()
    dips = [dict(d) for d in dips_raw]

    cedolini_list = []
    ore_totali_mese = 0.0
    totale_mese = 0.0
    for i, d in enumerate(dips):
        row = db.execute(
            "SELECT COALESCE(SUM(ore_totali),0) as ore, COUNT(DISTINCT data) as giorni "
            "FROM presenze WHERE utente_id=? AND data LIKE ?",
            (d['id'], prefisso + '%')).fetchone()
        ore    = float(row['ore'] or 0)
        giorni = row['giorni'] or 0
        paga_base    = float(d.get('paga_base_netta') or 0)
        ore_soglia   = float(d.get('ore_soglia_mensile') or 0)
        paga_oraria  = float(d.get('paga_oraria_netta') or 0)
        totale = calc_paga(ore, paga_base, ore_soglia, paga_oraria)
        # Spese rimborsabili approvate del mese
        spese_mese = db.execute(
            "SELECT COALESCE(SUM(importo),0) as tot FROM spese_rimborso WHERE utente_id=? AND stato='approvata' AND substr(data,1,7)=?",
            (d['id'], prefisso)).fetchone()['tot'] or 0.0
        # Detrazioni del mese (anticipi, acquisti, ecc.)
        detr_mese = db.execute(
            "SELECT COALESCE(SUM(importo),0) as tot FROM detrazioni_dipendente WHERE utente_id=? AND mese_competenza=?",
            (d['id'], prefisso)).fetchone()['tot'] or 0.0
        detr_dettaglio = db.execute(
            "SELECT tipo,descrizione,importo FROM detrazioni_dipendente WHERE utente_id=? AND mese_competenza=? ORDER BY data",
            (d['id'], prefisso)).fetchall()
        totale_netto_finale = round(totale + spese_mese - detr_mese, 2)
        ore_totali_mese += ore
        totale_mese += totale
        cedolini_list.append({
            'id': d['id'], 'nome': d['nome'], 'cognome': d['cognome'],
            'mansione': d.get('mansione') or '',
            'paga_oraria': paga_oraria, 'paga_base': paga_base, 'ore_soglia': ore_soglia,
            'ore': ore, 'giorni_lavorati': giorni,
            'totale': totale,
            'spese_rimborso': round(spese_mese, 2),
            'detrazioni': round(detr_mese, 2),
            'detr_dettaglio': [dict(x) for x in detr_dettaglio],
            'totale_con_spese': totale_netto_finale,
            'iniziali': (d['nome'][0]+d['cognome'][0]).upper(),
            'colore': COLORI_DIP[i % len(COLORI_DIP)],
        })
    max_ore = max((c['ore'] for c in cedolini_list), default=0)

    # Jolly del mese: raggruppa per nome+cognome
    jolly_cedolini = []
    jolly_rows = db.execute("""
        SELECT nome_jolly, cognome_jolly,
               COUNT(DISTINCT data) as giorni,
               COALESCE(SUM(ore_totali),0) as ore,
               GROUP_CONCAT(DISTINCT cantieri.nome) as cantieri_nomi
        FROM presenze
        LEFT JOIN cantieri ON cantieri.id=presenze.cantiere_id
        WHERE nome_jolly IS NOT NULL AND data LIKE ?
        GROUP BY nome_jolly, cognome_jolly
        ORDER BY cognome_jolly
    """, (prefisso+'%',)).fetchall()
    for j in jolly_rows:
        ore_j = float(j['ore'] or 0)
        if ore_j > 0 or j['giorni'] > 0:
            ore_totali_mese += ore_j
            # Recupera le singole presenze di questo jolly per edit/duplica
            presenze_jolly = db.execute("""
                SELECT p.id, p.data, p.ora_entrata, p.ora_uscita, p.ore_totali, p.note,
                       p.cantiere_id, c.nome as cantiere_nome
                FROM presenze p LEFT JOIN cantieri c ON c.id=p.cantiere_id
                WHERE p.nome_jolly=? AND p.cognome_jolly=? AND p.data LIKE ?
                ORDER BY p.data
            """, (j['nome_jolly'], j['cognome_jolly'], prefisso+'%')).fetchall()
            # Recupera paga oraria
            paga_row = db.execute(
                "SELECT paga_oraria FROM jolly_paghe WHERE nome_jolly=? AND cognome_jolly=?",
                (j['nome_jolly'], j['cognome_jolly'])).fetchone()
            paga_j = float(paga_row['paga_oraria']) if paga_row else 0.0
            totale_j = round(ore_j * paga_j, 2)
            jolly_cedolini.append({
                'nome': j['nome_jolly'], 'cognome': j['cognome_jolly'],
                'iniziali': (j['nome_jolly'][0]+j['cognome_jolly'][0]).upper(),
                'ore': round(ore_j, 1),
                'giorni_lavorati': j['giorni'],
                'cantieri': j['cantieri_nomi'] or '',
                'presenze': [dict(p) for p in presenze_jolly],
                'paga_oraria': paga_j,
                'totale': totale_j,
            })

    # Storico ultimi 6 mesi
    storico_mesi = []
    m, a = mese_sel, anno_sel
    for _ in range(6):
        pref = f'{a}-{m:02d}'
        tot = 0.0
        for d in dips:
            row = db.execute(
                "SELECT COALESCE(SUM(ore_totali),0) as ore FROM presenze WHERE utente_id=? AND data LIKE ?",
                (d['id'], pref+'%')).fetchone()
            ore = float(row['ore'] or 0)
            tot += calc_paga(ore, float(d.get('paga_base_netta') or 0),
                             float(d.get('ore_soglia_mensile') or 0),
                             float(d.get('paga_oraria_netta') or 0))
        storico_mesi.insert(0, {
            'etichetta': f"{['','Gen','Feb','Mar','Apr','Mag','Giu','Lug','Ago','Set','Ott','Nov','Dic'][m]}{str(a)[2:]}",
            'totale': round(tot, 2),
        })
        m -= 1
        if m == 0: m = 12; a -= 1

    # Costi per cantiere nel mese
    cantieri_rows = db.execute("SELECT id,nome FROM cantieri ORDER BY nome").fetchall()
    costi_per_cantiere = []
    for cant in cantieri_rows:
        ore_cant = 0.0; tot_cant = 0.0
        for d in dips:
            row = db.execute(
                "SELECT COALESCE(SUM(ore_totali),0) as ore FROM presenze "
                "WHERE utente_id=? AND cantiere_id=? AND data LIKE ?",
                (d['id'], cant['id'], prefisso+'%')).fetchone()
            ore_d = float(row['ore'] or 0)
            ore_cant += ore_d
            tot_cant += calc_paga(ore_d, float(d.get('paga_base_netta') or 0),
                                  float(d.get('ore_soglia_mensile') or 0),
                                  float(d.get('paga_oraria_netta') or 0))
        if ore_cant > 0:
            costi_per_cantiere.append({'nome': cant['nome'], 'ore': ore_cant, 'totale': round(tot_cant,2)})
    costi_per_cantiere.sort(key=lambda x: x['totale'], reverse=True)

    # ── Grafico spese per evento ──
    # Prende TUTTI gli eventi che si sovrappongono almeno in parte al mese selezionato
    # e per ognuno calcola ore dipendenti dalle timbrature nelle date dell'evento
    import calendar as _cal
    mese_start = _d(anno_sel, mese_sel, 1)
    mese_end   = _d(anno_sel, mese_sel, _cal.monthrange(anno_sel, mese_sel)[1])

    tutti_eventi = db.execute(
        "SELECT e.*, c.nome as cantiere_nome FROM eventi e "
        "LEFT JOIN cantieri c ON e.cantiere_id = c.id "
        "WHERE e.data_inizio IS NOT NULL ORDER BY e.data_inizio").fetchall()

    spese_per_evento = []
    totale_spese_extra = 0.0

    for ev in tutti_eventi:
        ev = dict(ev)
        try:
            ev_start_str = ev['data_inizio'][:10]  # prendi solo la parte data YYYY-MM-DD
            ev_end_str   = (ev.get('data_fine') or ev['data_inizio'])[:10]
            ev_start = _d.fromisoformat(ev_start_str)
            ev_end   = _d.fromisoformat(ev_end_str)
        except Exception:
            continue

        # Controlla se l'evento si sovrappone al mese selezionato
        if ev_end < mese_start or ev_start > mese_end:
            continue

        # Genera lista di date dell'evento
        delta = (ev_end - ev_start).days + 1
        date_evento = [(ev_start + _td(days=i)).isoformat() for i in range(delta)]

        # Spese extra (da tabella spese_evento)
        spese = db.execute(
            "SELECT COALESCE(SUM(importo),0) as tot FROM spese_evento WHERE evento_id=?",
            (ev['id'],)).fetchone()['tot']
        totale_spese_extra += float(spese or 0)

        # Ore dipendenti: tutte le timbrature sul cantiere dell'evento nelle date dell'evento
        # Non serve più l'assegnazione — conta chi ha timbrato su quel cantiere
        costo_dip  = 0.0
        ore_ev     = 0.0
        dip_detail = []

        if ev.get('cantiere_id'):
            # Prendi tutte le presenze sul cantiere nelle date dell'evento
            placeholders = ','.join('?' for _ in date_evento)
            rows_cantiere = db.execute(f"""
                SELECT p.utente_id, p.nome_jolly, p.cognome_jolly,
                       p.data, COALESCE(p.ore_totali,0) as ore,
                       u.nome, u.cognome,
                       u.paga_oraria_netta, u.paga_base_netta, u.ore_soglia_mensile
                FROM presenze p
                LEFT JOIN utenti u ON u.id=p.utente_id
                WHERE p.cantiere_id=? AND p.data IN ({placeholders})
            """, [ev['cantiere_id']] + date_evento).fetchall()

            # Raggruppa per dipendente/jolly
            agg = {}
            for r in rows_cantiere:
                if r['nome_jolly']:
                    key = f"jolly_{r['nome_jolly']}_{r['cognome_jolly']}"
                    if key not in agg:
                        agg[key] = {'nome': f"👷 {r['nome_jolly']} {r['cognome_jolly']}",
                                    'ore': 0.0, 'paga': 0.0, 'is_jolly': True,
                                    'jolly_nome': r['nome_jolly'], 'jolly_cognome': r['cognome_jolly']}
                    agg[key]['ore'] += float(r['ore'])
                elif r['utente_id']:
                    key = f"dip_{r['utente_id']}"
                    if key not in agg:
                        agg[key] = {'nome': f"{r['nome']} {r['cognome']}",
                                    'ore': 0.0, 'paga': float(r['paga_oraria_netta'] or 0),
                                    'is_jolly': False}
                    agg[key]['ore'] += float(r['ore'])

            for key, d in agg.items():
                if d['ore'] <= 0: continue
                if d['is_jolly']:
                    # Recupera paga jolly dalla tabella jolly_paghe
                    jp = db.execute("SELECT paga_oraria FROM jolly_paghe WHERE nome_jolly=? AND cognome_jolly=?",
                                    (d['jolly_nome'], d['jolly_cognome'])).fetchone()
                    paga_h = float(jp['paga_oraria']) if jp else 0.0
                else:
                    paga_h = d['paga']
                ore_ev    += d['ore']
                costo_dip += d['ore'] * paga_h
                dip_detail.append({'nome': d['nome'], 'ore': round(d['ore'], 1)})
        else:
            # Nessuna fiera: nessun costo automatico (non si sa dove ha lavorato)
            pass

        totale_ev = round(float(spese or 0) + costo_dip, 2)

        # Mostra anche eventi con solo spese extra registrate
        if totale_ev > 0 or float(spese or 0) > 0:
            spese_per_evento.append({
                'id': ev['id'], 'titolo': ev['titolo'],
                'tipo': ev.get('tipo') or '',
                'cantiere': ev.get('cantiere_nome') or '',
                'data_inizio': ev_start_str,
                'data_fine': ev_end_str,
                'ore_dip': round(ore_ev, 1),
                'costo_dip': round(costo_dip, 2),
                'spese_extra': round(float(spese or 0), 2),
                'totale': totale_ev,
                'team': dip_detail,
            })

    spese_per_evento.sort(key=lambda x: x['totale'], reverse=True)

    cantieri_list = db.execute("SELECT id,nome FROM cantieri WHERE attivo=1 ORDER BY nome").fetchall()
    db.close()
    return render_page(CEDOLINI_TMPL, page_title='Cedolini', active='cedolini',
        mese_sel=mese_sel, anno_sel=anno_sel, anni=anni,
        cedolini=cedolini_list, max_ore=max_ore,
        ore_totali_mese=ore_totali_mese, totale_mese=totale_mese,
        dipendenti_attivi=len([d for d in cedolini_list if d['ore'] > 0]),
        totale_spese_extra=totale_spese_extra,
        storico_mesi=storico_mesi,
        costi_per_cantiere=costi_per_cantiere,
        spese_per_evento=spese_per_evento,
        jolly_cedolini=jolly_cedolini,
        prefisso=prefisso,
        cantieri_list=cantieri_list)



@app.route('/presenze/jolly/duplica-salva', methods=['POST'])
@admin_required
def presenze_jolly_duplica_salva():
    nome_jolly    = request.form.get('nome_jolly', '').strip()
    cognome_jolly = request.form.get('cognome_jolly', '').strip()
    data          = request.form.get('data', '')
    cid           = request.form.get('cantiere_id') or None
    note          = request.form.get('note', '')
    oe            = request.form.get('ora_entrata', '') or None
    ou            = request.form.get('ora_uscita', '') or None
    ore_s         = request.form.get('ore_dirette', '')
    if oe and ou:
        try:
            diff = datetime.strptime(ou, '%H:%M') - datetime.strptime(oe, '%H:%M')
            ore  = round(diff.total_seconds() / 3600, 2)
        except: ore = None
    elif ore_s:
        try: ore = round(float(ore_s), 2)
        except: ore = None
    else: ore = None
    if not nome_jolly or not cognome_jolly or not data:
        flash('Compila nome, cognome e data.', 'error')
        return redirect(request.referrer or url_for('cedolini'))
    db = get_db()
    db.execute(
        "INSERT INTO presenze (utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note,nome_jolly,cognome_jolly) VALUES (?,?,?,?,?,?,?,?,?)",
        (None, data, oe, ou, ore, cid, note, nome_jolly, cognome_jolly))
    safe_commit(db); db.close()
    flash(f'Nuova presenza jolly {nome_jolly} {cognome_jolly} del {data} creata!', 'success')
    return redirect(request.referrer or url_for('cedolini'))


@app.route('/presenze/jolly/modifica', methods=['POST'])
@admin_required
def presenze_jolly_modifica():
    pid           = int(request.form.get('pid', 0))
    nome_jolly    = request.form.get('nome_jolly', '').strip()
    cognome_jolly = request.form.get('cognome_jolly', '').strip()
    data          = request.form.get('data', '')
    cid           = request.form.get('cantiere_id') or None
    note          = request.form.get('note', '')
    oe            = request.form.get('ora_entrata', '') or None
    ou            = request.form.get('ora_uscita', '') or None
    ore_s         = request.form.get('ore_dirette', '')
    if oe and ou:
        try:
            diff = datetime.strptime(ou, '%H:%M') - datetime.strptime(oe, '%H:%M')
            ore  = round(diff.total_seconds() / 3600, 2)
        except: ore = None
    elif ore_s:
        try: ore = round(float(ore_s), 2)
        except: ore = None
    else: ore = None
    db = get_db()
    db.execute(
        "UPDATE presenze SET nome_jolly=?,cognome_jolly=?,data=?,cantiere_id=?,note=?,ora_entrata=?,ora_uscita=?,ore_totali=? WHERE id=?",
        (nome_jolly, cognome_jolly, data, cid, note, oe, ou, ore, pid))
    safe_commit(db); db.close()
    flash(f'Presenza jolly aggiornata!', 'success')
    return redirect(request.referrer or url_for('cedolini'))


@app.route('/presenze/jolly/<int:pid>/duplica')
@admin_required
def presenze_jolly_duplica(pid):
    db = get_db()
    p = db.execute("SELECT * FROM presenze WHERE id=?", (pid,)).fetchone()
    if not p:
        db.close(); flash('Presenza non trovata.', 'error'); return redirect(url_for('cedolini'))
    try:
        from datetime import date as _d, timedelta as _td
        nuova_data = (_d.fromisoformat(p['data']) + _td(days=1)).isoformat()
    except: nuova_data = p['data']
    db.execute(
        "INSERT INTO presenze (utente_id,data,ora_entrata,ora_uscita,ore_totali,cantiere_id,note,nome_jolly,cognome_jolly) VALUES (?,?,?,?,?,?,?,?,?)",
        (None, nuova_data, p['ora_entrata'], p['ora_uscita'], p['ore_totali'],
         p['cantiere_id'], p['note'], p['nome_jolly'], p['cognome_jolly']))
    safe_commit(db); db.close()
    flash(f'Presenza duplicata per {nuova_data}.', 'success')
    return redirect(request.referrer or url_for('cedolini'))


@app.route('/presenze/jolly/rinomina', methods=['POST'])
@admin_required
def presenze_jolly_rinomina():
    nome_old    = request.form.get('nome_old', '').strip()
    cognome_old = request.form.get('cognome_old', '').strip()
    nome_new    = request.form.get('nome_new', '').strip()
    cognome_new = request.form.get('cognome_new', '').strip()
    mese        = request.form.get('mese', '')
    if not nome_new or not cognome_new:
        flash('Inserisci nome e cognome.', 'error'); return redirect(request.referrer or url_for('cedolini'))
    db = get_db()
    n = db.execute(
        "UPDATE presenze SET nome_jolly=?, cognome_jolly=? WHERE nome_jolly=? AND cognome_jolly=? AND data LIKE ?",
        (nome_new, cognome_new, nome_old, cognome_old, mese + '%')).rowcount
    safe_commit(db); db.close()
    flash(f'Rinominato "{nome_old} {cognome_old}" -> "{nome_new} {cognome_new}" ({n} presenze aggiornate).', 'success')
    return redirect(request.referrer or url_for('cedolini'))


@app.route('/cedolini/jolly-paghe', methods=['GET','POST'])
@admin_required
def cedolini_jolly_paghe():
    db = get_db()

    if request.method == 'POST':
        # Salva paga per i jolly selezionati
        nomi_sel = request.form.getlist('jolly_sel')  # lista "nome|cognome"
        paga     = float(request.form.get('paga_bulk') or 0)
        n_saved  = 0
        for key in nomi_sel:
            try:
                nome, cognome = key.split('|', 1)
                db.execute("""INSERT INTO jolly_paghe (nome_jolly,cognome_jolly,paga_oraria)
                              VALUES (?,?,?)
                              ON CONFLICT(nome_jolly,cognome_jolly)
                              DO UPDATE SET paga_oraria=excluded.paga_oraria, aggiornato_il=datetime('now')""",
                           (nome, cognome, paga))
                n_saved += 1
            except: pass
        # Salva anche modifiche individuali
        for key, val in request.form.items():
            if key.startswith('paga_ind_'):
                raw = key[len('paga_ind_'):]
                try:
                    nome, cognome = raw.split('|', 1)
                    p_ind = float(val or 0)
                    db.execute("""INSERT INTO jolly_paghe (nome_jolly,cognome_jolly,paga_oraria)
                                  VALUES (?,?,?)
                                  ON CONFLICT(nome_jolly,cognome_jolly)
                                  DO UPDATE SET paga_oraria=excluded.paga_oraria, aggiornato_il=datetime('now')""",
                               (nome, cognome, p_ind))
                except: pass
        safe_commit(db)
        flash(f'Paga aggiornata per {n_saved} jolly!', 'success') if n_saved else flash('Paghe individuali aggiornate.', 'success')
        return redirect(url_for('cedolini_jolly_paghe'))

    # Carica tutti i jolly distinti dal DB presenze + le loro paghe
    jolly_dist = db.execute("""
        SELECT DISTINCT nome_jolly, cognome_jolly
        FROM presenze
        WHERE nome_jolly IS NOT NULL
        ORDER BY cognome_jolly, nome_jolly
    """).fetchall()

    paghe = {f"{r['nome_jolly']}|{r['cognome_jolly']}": float(r['paga_oraria'])
             for r in db.execute("SELECT * FROM jolly_paghe").fetchall()}
    db.close()

    tmpl = """
<div style="margin-bottom:16px">
  <a href="/cedolini" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Cedolini</a>
</div>
<div class="card" style="max-width:680px;margin:0 auto">
  <div class="card-header" style="background:#fffbeb;border-bottom:2px solid #fde047">
    <h3 style="color:#92400e"><i class="fa fa-coins"></i> Paga oraria lavoratori Jolly</h3>
    <div style="font-size:12px;color:#92400e;margin-top:2px">Seleziona uno o più jolly e imposta la stessa paga, oppure modifica singolarmente</div>
  </div>
  <div class="card-body">
    <form method="POST">
      <!-- Barra paga bulk -->
      <div style="background:#fef9c3;border:1px solid #fde047;border-radius:10px;padding:14px;margin-bottom:20px">
        <div style="font-size:13px;font-weight:700;color:#92400e;margin-bottom:10px"><i class="fa fa-users"></i> Applica stessa paga ai selezionati</div>
        <div style="display:flex;gap:10px;align-items:flex-end;flex-wrap:wrap">
          <div class="form-group" style="margin:0;flex:1;min-width:140px">
            <label style="font-size:12px">€/ora da applicare</label>
            <input type="number" step="0.01" min="0" name="paga_bulk" id="paga-bulk-inp"
                   placeholder="es. 12.50" style="font-size:18px;font-weight:700;color:#d97706">
          </div>
          <button type="submit" class="btn" style="background:linear-gradient(135deg,#d97706,#f59e0b);color:#fff;font-weight:700;padding:10px 20px">
            <i class="fa fa-save"></i> Applica ai selezionati
          </button>
        </div>
        <div style="font-size:11px;color:#92400e;margin-top:8px">
          <i class="fa fa-info-circle"></i> Spunta le caselle nella tabella sotto, poi clicca "Applica"
        </div>
      </div>

      <!-- Tabella jolly -->
      {% if jolly_dist %}
      <table style="width:100%;border-collapse:collapse">
        <thead><tr style="background:#fef9c3;border-bottom:2px solid #fde047">
          <th style="padding:10px 14px;width:36px">
            <input type="checkbox" id="chk-jolly-all" onchange="toggleTutti(this)">
          </th>
          <th style="padding:10px 14px;text-align:left">Jolly</th>
          <th style="padding:10px 14px;text-align:right">Paga attuale (€/h)</th>
          <th style="padding:10px 14px;text-align:right">Modifica singola</th>
        </tr></thead>
        <tbody>
        {% for j in jolly_dist %}
        {% set key = j.nome_jolly + '|' + j.cognome_jolly %}
        {% set paga_att = paghe.get(key, 0) %}
        <tr style="border-bottom:1px solid #fef3c7" id="jrow-{{ loop.index }}">
          <td style="padding:10px 14px">
            <input type="checkbox" name="jolly_sel" value="{{ key }}" class="chk-jolly" onchange="aggiornaBulk()">
          </td>
          <td style="padding:10px 14px">
            <div style="display:flex;align-items:center;gap:10px">
              <div style="width:32px;height:32px;border-radius:50%;background:linear-gradient(135deg,#d97706,#f59e0b);display:flex;align-items:center;justify-content:center;color:#fff;font-size:12px;font-weight:700">
                {{ j.nome_jolly[0] }}{{ j.cognome_jolly[0] }}
              </div>
              <strong>{{ j.nome_jolly }} {{ j.cognome_jolly }}</strong>
            </div>
          </td>
          <td style="padding:10px 14px;text-align:right;font-family:monospace">
            {% if paga_att %}<span style="color:#d97706;font-weight:700">€ {{ "%.2f"|format(paga_att) }}</span>
            {% else %}<span style="color:#94a3b8;font-size:12px">non impostata</span>{% endif %}
          </td>
          <td style="padding:10px 14px;text-align:right">
            <input type="number" step="0.01" min="0" name="paga_ind_{{ key }}"
                   value="{{ '%.2f'|format(paga_att) if paga_att else '' }}"
                   placeholder="€/h" style="width:100px;text-align:right;font-weight:700;color:#d97706">
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
      <div style="display:flex;justify-content:flex-end;margin-top:14px">
        <button type="submit" class="btn btn-primary"><i class="fa fa-save"></i> Salva tutte le modifiche individuali</button>
      </div>
      {% else %}
      <div class="empty-state"><i class="fa fa-user-clock"></i><p>Nessun jolly trovato.<br><small>Inserisci prima delle presenze jolly dalla sezione Presenze.</small></p></div>
      {% endif %}
    </form>
  </div>
</div>
<script>
function toggleTutti(cb) {
  document.querySelectorAll('.chk-jolly').forEach(c => {
    c.checked = cb.checked;
    document.getElementById('jrow-' + (Array.from(document.querySelectorAll('.chk-jolly')).indexOf(c)+1)).style.background = cb.checked ? '#fffbeb' : '';
  });
}
function aggiornaBulk() {
  var n = document.querySelectorAll('.chk-jolly:checked').length;
  document.querySelectorAll('.chk-jolly').forEach((c,i) => {
    document.getElementById('jrow-'+(i+1)).style.background = c.checked ? '#fffbeb' : '';
  });
  var all = document.getElementById('chk-jolly-all');
  all.indeterminate = n > 0 && n < document.querySelectorAll('.chk-jolly').length;
  all.checked = n === document.querySelectorAll('.chk-jolly').length;
}
</script>"""
    return render_page(tmpl, page_title='Paga Jolly', active='cedolini',
                       jolly_dist=jolly_dist, paghe=paghe)


@app.route('/cedolini/impostazioni-paga', methods=['GET','POST'])
@admin_required
def cedolini_paga():
    db = get_db()
    if request.method == 'POST':
        dips = db.execute("SELECT id FROM utenti WHERE ruolo='dipendente'").fetchall()
        for d in dips:
            try:
                paga   = float(request.form.get(f'paga_{d["id"]}') or 0)
                base   = float(request.form.get(f'base_{d["id"]}') or 0)
                soglia = float(request.form.get(f'soglia_{d["id"]}') or 0)
                db.execute(
                    "UPDATE utenti SET paga_oraria_netta=?, paga_base_netta=?, ore_soglia_mensile=? WHERE id=?",
                    (paga, base, soglia, d['id']))
            except: pass
        safe_commit(db); db.close()
        flash('Paghe aggiornate!','success')
        return redirect(url_for('cedolini'))
    dipendenti = db.execute(
        "SELECT * FROM utenti WHERE ruolo='dipendente' AND attivo=1 ORDER BY nome").fetchall()
    db.close()
    return render_page(PAGA_FORM_TMPL, page_title='Impostazioni Paga', active='cedolini',
        dipendenti=dipendenti)




# ══════════════════════════════════════════════════════════
#  FOGLIO PRESENZE MENSILE — PDF stile Accesso Fiere
# ══════════════════════════════════════════════════════════

FOGLIO_PRES_TMPL = """
<div style="margin-bottom:18px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px">
  <a href="/cedolini" class="btn btn-secondary btn-sm"><i class="fa fa-arrow-left"></i> Cedolini</a>
  <h2 style="margin:0"><i class="fa fa-file-alt" style="color:#0f172a"></i> Fogli Presenze</h2>
</div>

<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:20px;max-width:900px;margin:0 auto">

  <!-- Card Pulizia -->
  <div class="card">
    <div class="card-header">
      <h3><i class="fa fa-broom" style="color:#0891b2"></i> Pulizia</h3>
    </div>
    <div class="card-body">
      <form method="GET" action="/cedolini/foglio-presenze/pdf" target="_blank">
        <input type="hidden" name="cantiere_nome" value="Pulizia">
        <div class="form-row">
          <div class="form-group">
            <label>Mese *</label>
            <select name="mese" style="width:100%">
              {% for m in range(1,13) %}
              <option value="{{ m }}" {{ 'selected' if m==mese_sel }}>{{ ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno','Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre'][m] }}</option>
              {% endfor %}
            </select>
          </div>
          <div class="form-group">
            <label>Anno *</label>
            <select name="anno" style="width:100%">
              {% for a in anni %}
              <option value="{{ a }}" {{ 'selected' if a==anno_sel }}>{{ a }}</option>
              {% endfor %}
            </select>
          </div>
        </div>
        <div class="form-group">
          <label>Tariffa oraria (€/h)</label>
          <input type="number" name="tariffa" value="20" step="0.5" min="0" style="font-size:18px;font-weight:700;color:#0891b2">
        </div>
        <button type="submit" class="btn btn-primary" style="width:100%;margin-top:4px"><i class="fa fa-file-pdf"></i> Genera PDF Pulizia</button>
      </form>
    </div>
  </div>

  <!-- Card Magazzino 30group -->
  <div class="card">
    <div class="card-header">
      <h3><i class="fa fa-warehouse" style="color:#7c3aed"></i> Magazzino 30group</h3>
    </div>
    <div class="card-body">
      <form method="GET" action="/cedolini/foglio-presenze/pdf" target="_blank">
        <input type="hidden" name="cantiere_nome" value="Magazzino 30group">
        <div class="form-row">
          <div class="form-group">
            <label>Mese *</label>
            <select name="mese" style="width:100%">
              {% for m in range(1,13) %}
              <option value="{{ m }}" {{ 'selected' if m==mese_sel }}>{{ ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno','Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre'][m] }}</option>
              {% endfor %}
            </select>
          </div>
          <div class="form-group">
            <label>Anno *</label>
            <select name="anno" style="width:100%">
              {% for a in anni %}
              <option value="{{ a }}" {{ 'selected' if a==anno_sel }}>{{ a }}</option>
              {% endfor %}
            </select>
          </div>
        </div>
        <div class="form-group">
          <label>Tariffa oraria (€/h) <span style="font-size:11px;color:var(--text-light);font-weight:400">(non appare nel PDF)</span></label>
          <input type="number" name="tariffa" value="20" step="0.5" min="0" style="font-size:18px;font-weight:700;color:#7c3aed">
        </div>
        <button type="submit" class="btn" style="width:100%;margin-top:4px;background:linear-gradient(135deg,#7c3aed,#a78bfa);color:#fff"><i class="fa fa-file-pdf"></i> Genera PDF Magazzino</button>
      </form>
    </div>
  </div>

</div>
"""


@app.route('/cedolini/foglio-presenze')
@admin_required
def foglio_presenze():
    from datetime import date as _d
    oggi = _d.today()
    anni = list(range(oggi.year - 2, oggi.year + 1))
    return render_page(FOGLIO_PRES_TMPL, page_title='Fogli Presenze', active='cedolini',
        mese_sel=oggi.month, anno_sel=oggi.year, anni=anni)


@app.route('/cedolini/foglio-presenze/pdf')
@admin_required
def foglio_presenze_pdf():
    from datetime import date as _d
    import calendar as _cal
    oggi = _d.today()
    mese    = int(request.args.get('mese', oggi.month))
    anno    = int(request.args.get('anno', oggi.year))
    tariffa = float(request.args.get('tariffa') or 20)
    cantiere_nome_filter = request.args.get('cantiere_nome', 'Pulizia')

    db = get_db()
    prefisso = f'{anno}-{mese:02d}'
    mesi_it  = ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
                'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']

    # Trova la fiera per nome (LIKE, case-insensitive)
    cant_row = db.execute("SELECT id,nome FROM cantieri WHERE nome LIKE ? LIMIT 1",
                          (f'%{cantiere_nome_filter}%',)).fetchone()
    cid = cant_row['id'] if cant_row else None
    cantiere_nome = cant_row['nome'] if cant_row else cantiere_nome_filter

    if cid:
        righe = db.execute("""
            SELECT p.data,
                   COUNT(DISTINCT CASE WHEN p.utente_id IS NOT NULL THEN p.utente_id END)
                   + COUNT(CASE WHEN p.nome_jolly IS NOT NULL THEN 1 END) as n_persone,
                   SUM(p.ore_totali) as ore_tot
            FROM presenze p
            WHERE p.data LIKE ? AND p.cantiere_id=?
            GROUP BY p.data ORDER BY p.data
        """, (prefisso+'%', cid)).fetchall()
    else:
        # Se la fiera non esiste ancora, foglio vuoto
        righe = []
    db.close()

    # Costruisci dict giorno → dati
    giorni_dati = {}
    for r in righe:
        giorno = int(r['data'].split('-')[2])
        giorni_dati[giorno] = {
            'n_persone': r['n_persone'] or 0,
            'ore':       float(r['ore_tot'] or 0),
        }

    totale_ore = sum(v['ore'] for v in giorni_dati.values())
    totale_importo = round(totale_ore * tariffa, 2)
    n_giorni_mese = _cal.monthrange(anno, mese)[1]
    az_nome = get_setting('azienda', 'Accesso Fiere S.r.l.')

    # ── ReportLab PDF ──
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        import io as _io
    except ImportError:
        return "ReportLab non installato", 500

    BLU   = colors.HexColor('#0f172a')
    ACC   = colors.HexColor('#2563eb')
    WHITE = colors.white
    GRAY  = colors.HexColor('#f1f5f9')
    GRAY2 = colors.HexColor('#e2e8f0')
    TEXTC = colors.HexColor('#1e293b')
    TEXTL = colors.HexColor('#64748b')
    W, H  = A4

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm, topMargin=16*mm, bottomMargin=16*mm)

    def P(txt, font='Helvetica', size=9, color=TEXTC, align=TA_LEFT):
        s = ParagraphStyle('x', fontName=font, fontSize=size, textColor=color, alignment=align, leading=size*1.3)
        return Paragraph(str(txt), s)

    story = []

    # ── Intestazione identica al modello ──
    hdr = Table([[
        P(az_nome, 'Helvetica-Bold', 13, BLU),
        P(f'Foglio presenze — {cantiere_nome.upper()} — Mese: {mesi_it[mese].upper()} {anno}', 'Helvetica-Bold', 11, BLU, TA_RIGHT),
    ]], colWidths=[95*mm, None])
    hdr.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), colors.HexColor('#f8fafc')),
        ('LINEBELOW',(0,0),(-1,0), 1.5, ACC),
        ('TOPPADDING',(0,0),(-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
        ('LEFTPADDING',(0,0),(-1,-1), 8),
        ('RIGHTPADDING',(0,0),(-1,-1), 8),
    ]))
    story.append(hdr)
    story.append(Spacer(1, 5*mm))

    # Riga "Foglio presenze lavoratore: ___"
    dip_row = Table([[
        P('Foglio presenze lavoratore:', 'Helvetica', 9, TEXTL),
        P('_' * 50, 'Helvetica', 9, TEXTC),
    ]], colWidths=[60*mm, None])
    dip_row.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    story.append(dip_row)
    story.append(Spacer(1, 5*mm))

    # ── Tabella dati — colonne identiche al modello ──
    col_w = [18*mm, 80*mm, 28*mm, 28*mm, 30*mm]
    thead = [
        P('DATA', 'Helvetica-Bold', 9, WHITE, TA_CENTER),
        P('CANTIERE', 'Helvetica-Bold', 9, WHITE, TA_LEFT),
        P('ORE', 'Helvetica-Bold', 9, WHITE, TA_CENTER),
        P('SPESE', 'Helvetica-Bold', 9, WHITE, TA_CENTER),
        P('RIFERIMENTO', 'Helvetica-Bold', 9, WHITE, TA_CENTER),
    ]
    tdata = [thead]

    for g in range(1, n_giorni_mese + 1):
        if g in giorni_dati:
            d = giorni_dati[g]
            np_ = d['n_persone']
            ore_ = d['ore']
            persona_str = f"{np_} PERSON{'A' if np_==1 else 'E'}"
            ore_fmt = str(int(ore_)) if ore_ == int(ore_) else f"{ore_:.1f}".replace('.', ',')
            tdata.append([
                P(str(g), 'Helvetica-Bold', 9, TEXTC, TA_CENTER),
                P(persona_str, 'Helvetica', 9, TEXTC, TA_LEFT),
                P(ore_fmt, 'Helvetica-Bold', 9, TEXTC, TA_CENTER),
                P('', 'Helvetica', 9, TEXTL, TA_CENTER),
                P('', 'Helvetica', 9, TEXTL, TA_CENTER),
            ])
        else:
            tdata.append([
                P(str(g), 'Helvetica', 9, TEXTC, TA_CENTER),
                P('', 'Helvetica', 9, TEXTL, TA_LEFT),
                P('', 'Helvetica', 9, TEXTL, TA_CENTER),
                P('', 'Helvetica', 9, TEXTL, TA_CENTER),
                P('', 'Helvetica', 9, TEXTL, TA_CENTER),
            ])

    # Riga totali
    tot_ore_fmt = str(int(totale_ore)) if totale_ore == int(totale_ore) else f"{totale_ore:.1f}".replace('.', ',')
    mostra_importo = 'magazzino' not in cantiere_nome.lower()
    if mostra_importo:
        tdata.append([
            P('', 'Helvetica-Bold', 9, BLU, TA_LEFT),
            P('TOTALE ORE:', 'Helvetica-Bold', 10, BLU, TA_RIGHT),
            P(tot_ore_fmt, 'Helvetica-Bold', 10, BLU, TA_CENTER),
            P('TOTALE IMPORTO:', 'Helvetica-Bold', 9, BLU, TA_RIGHT),
            P(f'€ {totale_importo:,.0f}'.replace(',','.'), 'Helvetica-Bold', 10, BLU, TA_CENTER),
        ])
    else:
        tdata.append([
            P('', 'Helvetica-Bold', 9, BLU, TA_LEFT),
            P('TOTALE ORE:', 'Helvetica-Bold', 10, BLU, TA_RIGHT),
            P(tot_ore_fmt, 'Helvetica-Bold', 12, BLU, TA_CENTER),
            P('', 'Helvetica', 9, BLU, TA_CENTER),
            P('', 'Helvetica', 9, BLU, TA_CENTER),
        ])

    tab = Table(tdata, colWidths=col_w, repeatRows=1)
    n = len(tdata)
    row_styles = [
        ('BACKGROUND', (0,0), (-1,0), BLU),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('BOX',  (0,0), (-1,-1), 0.8, GRAY2),
        ('INNERGRID', (0,0), (-1,-1), 0.4, GRAY2),
        ('BACKGROUND', (0, n-1), (-1, n-1), colors.HexColor('#eef4ff')),
        ('LINEABOVE',  (0, n-1), (-1, n-1), 1.2, ACC),
        ('LINEBELOW',  (0, n-1), (-1, n-1), 1.2, ACC),
    ]
    # Zebra
    for i in range(1, n-1):
        g_num = i  # riga 1=giorno1
        giorno_dt = _d(anno, mese, g_num) if 1 <= g_num <= n_giorni_mese else None
        if giorno_dt and giorno_dt.weekday() >= 5:
            row_styles.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#f0f4ff')))
        elif i % 2 == 0:
            row_styles.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#fafafa')))
    tab.setStyle(TableStyle(row_styles))
    story.append(tab)

    doc.build(story)
    buf.seek(0)
    nome_file = f"FoglioPresenze_{cantiere_nome.replace(' ','_')}_{mesi_it[mese]}_{anno}.pdf"
    return Response(buf.read(), mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{nome_file}"'})


# ══════════════════════════════════════════════════════════════════════════════
#  SAAS — LANDING · REGISTRAZIONE · ABBONAMENTI · SUPERADMIN
# ══════════════════════════════════════════════════════════════════════════════

STRIPE_SECRET      = os.environ.get('STRIPE_SECRET_KEY', '')
STRIPE_WEBHOOK_SEC = os.environ.get('STRIPE_WEBHOOK_SECRET', '')
STRIPE_PRICE_BASE  = os.environ.get('STRIPE_PRICE_BASE', '')
STRIPE_PRICE_PRO   = os.environ.get('STRIPE_PRICE_PRO', '')
STRIPE_PRICE_ENT   = os.environ.get('STRIPE_PRICE_ENT', '')

SUPERADMIN_EMAIL = os.environ.get('SUPERADMIN_EMAIL', 'superadmin@gestionale.app')
SUPERADMIN_PW    = os.environ.get('SUPERADMIN_PASSWORD', 'superadmin123')

# ── Landing page ─────────────────────────────────────────────
@app.route('/landing')
def landing():
    mdb = get_master_db()
    piani = mdb.execute("SELECT * FROM piani ORDER BY prezzo_mensile").fetchall()
    mdb.close()
    return render_template_string(LANDING_TMPL, piani=piani)

# ── Registrazione nuova azienda ───────────────────────────────
@app.route('/registrati', methods=['GET','POST'])
def registrati():
    if request.method == 'POST':
        nome_az  = request.form.get('nome_azienda','').strip()
        email    = request.form.get('email','').strip().lower()
        password = request.form.get('password','')
        piano    = request.form.get('piano','base')
        if not nome_az or not email or not password:
            return render_template_string(REGISTRATI_TMPL, error='Compila tutti i campi.', piano_sel=piano)
        if len(password) < 6:
            return render_template_string(REGISTRATI_TMPL, error='Password di almeno 6 caratteri.', piano_sel=piano)
        mdb = get_master_db()
        existing = mdb.execute("SELECT id FROM aziende WHERE email_admin=?", (email,)).fetchone()
        if existing:
            mdb.close()
            return render_template_string(REGISTRATI_TMPL, error='Email già registrata.', piano_sel=piano)
        import re, time as _t
        slug = re.sub(r'[^a-z0-9]', '', nome_az.lower())[:20] + str(int(_t.time()))[-4:]
        pw_hash = hash_pw(password)
        # Trial 14 giorni
        from datetime import date as _d, timedelta as _td
        trial_fino = (_d.today() + _td(days=14)).isoformat()
        cur = mdb.execute(
            "INSERT INTO aziende (nome,email_admin,password_admin,piano,stato,trial_fino_al,slug) VALUES (?,?,?,?,?,?,?)",
            (nome_az, email, pw_hash, piano, 'trial', trial_fino, slug))
        azienda_id = cur.lastrowid
        mdb.commit(); mdb.close()
        # Crea DB tenant e inizializza schema
        session['azienda_id'] = azienda_id
        init_db()
        # Crea utente admin nel DB tenant
        db = get_db()
        db.execute(
            "INSERT INTO utenti (nome,cognome,email,password,ruolo,titolo) VALUES (?,?,?,?,?,?)",
            ('Admin', nome_az, email, pw_hash, 'admin', 'Amministratore'))
        db.execute("INSERT OR REPLACE INTO impostazioni (chiave,valore) VALUES ('azienda',?)", (nome_az,))
        safe_commit(db)
        admin_row = db.execute("SELECT id FROM utenti WHERE ruolo='admin' LIMIT 1").fetchone()
        db.close()
        # Login automatico
        session.update({
            'user_id': admin_row['id'],
            'nome': 'Admin', 'cognome': nome_az,
            'ruolo': 'admin', 'email': email,
            'azienda_id': azienda_id, 'azienda_nome': nome_az,
            'is_saas': True,
        })
        flash(f'🎉 Benvenuto! Hai 14 giorni di prova gratuita. Buon lavoro con {nome_az}!', 'success')
        return redirect(url_for('dashboard'))
    piano_sel = request.args.get('piano', 'base')
    return render_template_string(REGISTRATI_TMPL, error=None, piano_sel=piano_sel)

# ── Stripe: crea sessione checkout ───────────────────────────
@app.route('/abbonamento/checkout', methods=['POST'])
def abbonamento_checkout():
    if not STRIPE_SECRET:
        flash('Pagamenti non ancora configurati. Contatta il supporto.', 'error')
        return redirect(url_for('abbonamento_gestisci'))
    try:
        import stripe
        stripe.api_key = STRIPE_SECRET
        piano = request.form.get('piano', 'base')
        price_map = {'base': STRIPE_PRICE_BASE, 'professional': STRIPE_PRICE_PRO, 'enterprise': STRIPE_PRICE_ENT}
        price_id = price_map.get(piano)
        if not price_id:
            flash('Piano non valido.', 'error'); return redirect(url_for('abbonamento_gestisci'))
        azienda_id = session.get('azienda_id')
        mdb = get_master_db()
        az = dict(mdb.execute("SELECT * FROM aziende WHERE id=?", (azienda_id,)).fetchone())
        mdb.close()
        checkout = stripe.checkout.Session.create(
            customer_email=az['email_admin'],
            payment_method_types=['card'],
            line_items=[{'price': price_id, 'quantity': 1}],
            mode='subscription',
            success_url=request.host_url + 'abbonamento/successo?session_id={CHECKOUT_SESSION_ID}',
            cancel_url=request.host_url + 'abbonamento/gestisci',
            metadata={'azienda_id': str(azienda_id)},
        )
        return redirect(checkout.url)
    except Exception as e:
        flash(f'Errore pagamento: {e}', 'error')
        return redirect(url_for('abbonamento_gestisci'))

@app.route('/abbonamento/successo')
def abbonamento_successo():
    flash('✅ Abbonamento attivato! Grazie.', 'success')
    return redirect(url_for('abbonamento_gestisci'))

@app.route('/abbonamento/gestisci')
def abbonamento_gestisci():
    azienda_id = session.get('azienda_id')
    if not azienda_id:
        return redirect(url_for('login'))
    mdb = get_master_db()
    az = dict(mdb.execute("SELECT * FROM aziende WHERE id=?", (azienda_id,)).fetchone())
    piani = [dict(p) for p in mdb.execute("SELECT * FROM piani ORDER BY prezzo_mensile").fetchall()]
    mdb.close()
    return render_template_string(ABBONAMENTO_TMPL, az=az, piani=piani,
                                  stripe_ok=bool(STRIPE_SECRET))

# ── Stripe webhook ────────────────────────────────────────────
@app.route('/stripe/webhook', methods=['POST'])
def stripe_webhook():
    if not STRIPE_SECRET: return 'ok', 200
    try:
        import stripe
        stripe.api_key = STRIPE_SECRET
        payload = request.get_data()
        sig = request.headers.get('Stripe-Signature','')
        event = stripe.Webhook.construct_event(payload, sig, STRIPE_WEBHOOK_SEC)
        if event['type'] == 'checkout.session.completed':
            obj = event['data']['object']
            azienda_id = int(obj['metadata'].get('azienda_id', 0))
            if azienda_id:
                mdb = get_master_db()
                mdb.execute("UPDATE aziende SET stato='attivo', stripe_customer_id=?, stripe_subscription_id=? WHERE id=?",
                            (obj.get('customer'), obj.get('subscription'), azienda_id))
                mdb.commit(); mdb.close()
        elif event['type'] in ('customer.subscription.deleted','customer.subscription.paused'):
            sub = event['data']['object']
            mdb = get_master_db()
            mdb.execute("UPDATE aziende SET stato='sospeso' WHERE stripe_subscription_id=?", (sub['id'],))
            mdb.commit(); mdb.close()
    except Exception as e:
        print(f'[Stripe webhook] {e}')
    return 'ok', 200

# ── Super-Admin ───────────────────────────────────────────────
def superadmin_required(f):
    @wraps(f)
    def d(*a, **k):
        if session.get('is_superadmin') != True:
            return redirect(url_for('superadmin_login'))
        return f(*a, **k)
    return d

@app.route('/superadmin/login', methods=['GET','POST'])
def superadmin_login():
    if request.method == 'POST':
        email = request.form.get('email','')
        pw    = request.form.get('password','')
        if email == SUPERADMIN_EMAIL and pw == SUPERADMIN_PW:
            session['is_superadmin'] = True
            session['superadmin_email'] = email
            return redirect(url_for('superadmin_dashboard'))
        error = 'Credenziali non valide'
        return render_template_string(_SA_LOGIN_TMPL, error=error)
    return render_template_string(_SA_LOGIN_TMPL, error=None)

@app.route('/superadmin/logout')
def superadmin_logout():
    session.pop('is_superadmin', None)
    return redirect(url_for('superadmin_login'))

@app.route('/superadmin')
@superadmin_required
def superadmin_dashboard():
    mdb = get_master_db()
    aziende = [dict(a) for a in mdb.execute(
        "SELECT * FROM aziende ORDER BY creato_il DESC").fetchall()]
    totali = {
        'totale': len(aziende),
        'attive': sum(1 for a in aziende if a['stato']=='attivo'),
        'trial':  sum(1 for a in aziende if a['stato']=='trial'),
        'sospese':sum(1 for a in aziende if a['stato']=='sospeso'),
    }
    mdb.close()
    return render_template_string(_SA_DASHBOARD_TMPL, aziende=aziende, totali=totali)

@app.route('/superadmin/aziende/<int:aid>/sospendi')
@superadmin_required
def superadmin_sospendi(aid):
    mdb = get_master_db()
    mdb.execute("UPDATE aziende SET stato='sospeso' WHERE id=?", (aid,))
    mdb.commit(); mdb.close()
    flash('Azienda sospesa.', 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/superadmin/aziende/<int:aid>/riattiva')
@superadmin_required
def superadmin_riattiva(aid):
    mdb = get_master_db()
    mdb.execute("UPDATE aziende SET stato='attivo' WHERE id=?", (aid,))
    mdb.commit(); mdb.close()
    flash('Azienda riattivata.', 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/superadmin/aziende/<int:aid>/entra')
@superadmin_required
def superadmin_entra_azienda(aid):
    """Permette al superadmin di entrare come admin di un'azienda (supporto clienti)."""
    mdb = get_master_db()
    az = dict(mdb.execute("SELECT * FROM aziende WHERE id=?", (aid,)).fetchone())
    mdb.close()
    session_backup_superadmin = session.get('is_superadmin')
    # Imposta sessione come admin dell'azienda
    db_path = get_tenant_db_path(aid)
    if not os.path.exists(db_path):
        flash('DB azienda non ancora creato.', 'error')
        return redirect(url_for('superadmin_dashboard'))
    session['azienda_id'] = aid
    db = get_db()
    admin = db.execute("SELECT * FROM utenti WHERE ruolo='admin' LIMIT 1").fetchone()
    db.close()
    if admin:
        session.update({
            'user_id': admin['id'], 'nome': 'Admin', 'cognome': az['nome'],
            'ruolo': 'admin', 'email': az['email_admin'],
            'azienda_id': aid, 'azienda_nome': az['nome'],
            'is_saas': True, 'is_superadmin': True,
        })
    flash(f'Ora stai operando come admin di {az["nome"]}.', 'info')
    return redirect(url_for('dashboard'))

# ── Templates SaaS ────────────────────────────────────────────

LANDING_TMPL = """<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>GestionaleHR — Il Accesso Fiere per le PMI italiane</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f8fafc;color:#1e293b}
.hero{background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 50%,#0f172a 100%);padding:80px 20px;text-align:center;color:#fff}
.hero h1{font-size:clamp(32px,5vw,56px);font-weight:900;line-height:1.1;margin-bottom:20px}
.hero h1 span{color:#f59e0b}
.hero p{font-size:18px;color:rgba(255,255,255,.7);max-width:600px;margin:0 auto 32px}
.btn-hero{background:linear-gradient(135deg,#f59e0b,#d97706);color:#fff;padding:16px 36px;border-radius:50px;font-size:18px;font-weight:700;text-decoration:none;display:inline-block;transition:.2s}
.btn-hero:hover{transform:translateY(-2px);box-shadow:0 8px 30px rgba(245,158,11,.4)}
.btn-sec{background:transparent;color:#fff;border:2px solid rgba(255,255,255,.3);padding:14px 28px;border-radius:50px;font-size:16px;text-decoration:none;display:inline-block;margin-left:12px}
.features{padding:80px 20px;max-width:1100px;margin:0 auto}
.features h2{text-align:center;font-size:36px;font-weight:800;margin-bottom:48px}
.feat-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:24px}
.feat-card{background:#fff;border-radius:16px;padding:28px;border:1px solid #e2e8f0;transition:.2s}
.feat-card:hover{box-shadow:0 8px 30px rgba(0,0,0,.08);transform:translateY(-2px)}
.feat-icon{width:52px;height:52px;border-radius:14px;display:flex;align-items:center;justify-content:center;font-size:22px;margin-bottom:16px}
.feat-card h3{font-size:18px;font-weight:700;margin-bottom:8px}
.feat-card p{color:#64748b;font-size:14px;line-height:1.6}
.pricing{background:#0f172a;padding:80px 20px;color:#fff}
.pricing h2{text-align:center;font-size:36px;font-weight:800;margin-bottom:12px}
.pricing .sub{text-align:center;color:rgba(255,255,255,.5);margin-bottom:48px}
.price-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:24px;max-width:900px;margin:0 auto}
.price-card{background:#1e293b;border-radius:20px;padding:32px;border:1px solid rgba(255,255,255,.1);position:relative}
.price-card.popular{border-color:#f59e0b;transform:scale(1.03)}
.popular-badge{position:absolute;top:-14px;left:50%;transform:translateX(-50%);background:#f59e0b;color:#000;font-weight:800;font-size:12px;padding:4px 16px;border-radius:99px}
.price-card h3{font-size:20px;font-weight:700;margin-bottom:8px}
.price-card .price{font-size:42px;font-weight:900;color:#f59e0b;margin:16px 0 4px}
.price-card .price span{font-size:16px;color:rgba(255,255,255,.5)}
.price-card ul{list-style:none;margin:20px 0}
.price-card li{padding:6px 0;font-size:14px;color:rgba(255,255,255,.7);display:flex;align-items:center;gap:8px}
.price-card li i{color:#22c55e;font-size:12px}
.btn-price{display:block;text-align:center;background:linear-gradient(135deg,#f59e0b,#d97706);color:#000;padding:13px;border-radius:12px;font-weight:700;text-decoration:none;margin-top:24px}
.price-card.popular .btn-price{background:linear-gradient(135deg,#f59e0b,#ef4444)}
.trial-badge{text-align:center;color:#22c55e;font-size:13px;margin-top:8px}
.footer{background:#0f172a;padding:40px 20px;text-align:center;color:rgba(255,255,255,.3);font-size:13px;border-top:1px solid rgba(255,255,255,.08)}
.nav{display:flex;justify-content:space-between;align-items:center;padding:16px 32px;background:#0f172a;position:sticky;top:0;z-index:100}
.nav-logo{color:#fff;font-size:20px;font-weight:900;text-decoration:none}<i class="fa fa-chart-line" style="color:#f59e0b"></i>
.nav-links{display:flex;gap:20px}
.nav-links a{color:rgba(255,255,255,.6);text-decoration:none;font-size:14px}
.nav-links a.cta{background:#f59e0b;color:#000;padding:8px 20px;border-radius:8px;font-weight:700}
</style>
</head>
<body>
<nav class="nav">
  <a href="/landing" class="nav-logo"><i class="fa fa-chart-line" style="color:#f59e0b"></i> GestionaleHR</a>
  <div class="nav-links">
    <a href="#features">Funzionalità</a>
    <a href="#pricing">Prezzi</a>
    <a href="/login">Accedi</a>
    <a href="/registrati" class="cta">Prova gratis</a>
  </div>
</nav>

<div class="hero">
  <h1>Il Accesso Fiere per le<br><span>PMI italiane</span></h1>
  <p>Presenze, cedolini, fatture, cantieri, veicoli e molto altro.<br>Tutto in un'unica piattaforma semplice e potente.</p>
  <a href="/registrati" class="btn-hero">Inizia gratis — 14 giorni</a>
  <a href="/login" class="btn-sec">Accedi</a>
</div>

<div class="features" id="features">
  <h2>Tutto quello che ti serve per gestire la tua azienda</h2>
  <div class="feat-grid">
    <div class="feat-card">
      <div class="feat-icon" style="background:#eff6ff;color:#2563eb"><i class="fa fa-clock"></i></div>
      <h3>Presenze & Timbrature</h3>
      <p>Timbratura mobile con GPS, geofencing automatico, jolly esterni, approvazione richieste.</p>
    </div>
    <div class="feat-card">
      <div class="feat-icon" style="background:#f0fdf4;color:#16a34a"><i class="fa fa-money-bill-wave"></i></div>
      <h3>Cedolini & Paghe</h3>
      <p>Calcolo automatico stipendi, paga oraria, anticipi, detrazioni, export Excel professionale.</p>
    </div>
    <div class="feat-card">
      <div class="feat-icon" style="background:#fef3c7;color:#d97706"><i class="fa fa-file-invoice-dollar"></i></div>
      <h3>Fatturazione</h3>
      <p>Fatture, preventivi, rate di pagamento, scadenziario clienti con promemoria automatici.</p>
    </div>
    <div class="feat-card">
      <div class="feat-icon" style="background:#fdf4ff;color:#9333ea"><i class="fa fa-calendar-check"></i></div>
      <h3>Calendario & Fiere & Stand</h3>
      <p>Pianifica eventi, assegna team e mezzi, traccia i costi di ogni cantiere in tempo reale.</p>
    </div>
    <div class="feat-card">
      <div class="feat-icon" style="background:#fef2f2;color:#dc2626"><i class="fa fa-car"></i></div>
      <h3>Gestione Veicoli</h3>
      <p>Parco auto aziendale, scadenze assicurazione e revisione, documenti digitali.</p>
    </div>
    <div class="feat-card">
      <div class="feat-icon" style="background:#ecfdf5;color:#059669"><i class="fa fa-receipt"></i></div>
      <h3>Note Spese</h3>
      <p>Rimborsi spese con foto scontrino, approvazione admin, integrazione cedolini.</p>
    </div>
  </div>
</div>

<div class="pricing" id="pricing">
  <h2>Prezzi trasparenti, senza sorprese</h2>
  <p class="sub">14 giorni di prova gratuita — Nessuna carta di credito richiesta</p>
  <div class="price-grid">
    {% for p in piani %}
    <div class="price-card {{ 'popular' if p.nome=='Professional' }}">
      {% if p.nome=='Professional' %}<div class="popular-badge">⭐ PIÙ SCELTO</div>{% endif %}
      <h3>{{ p.nome }}</h3>
      <div class="price">€{{ p.prezzo_mensile|int }}<span>/mese</span></div>
      <ul>
        <li><i class="fa fa-check"></i> {% if p.max_dipendenti < 999 %}Fino a {{ p.max_dipendenti }} dipendenti{% else %}Dipendenti illimitati{% endif %}</li>
        <li><i class="fa fa-check"></i> Presenze e timbrature GPS</li>
        <li><i class="fa fa-check"></i> Fatturazione e preventivi</li>
        <li><i class="fa fa-check"></i> Cedolini automatici</li>
        <li><i class="fa fa-check"></i> App mobile inclusa</li>
        {% if p.nome != 'Base' %}<li><i class="fa fa-check"></i> Export Excel avanzato</li>{% endif %}
        {% if p.nome == 'Enterprise' %}<li><i class="fa fa-check"></i> Supporto prioritario</li>{% endif %}
      </ul>
      <a href="/registrati?piano={{ p.nome|lower }}" class="btn-price">Inizia gratis</a>
      <div class="trial-badge">✓ 14 giorni gratuiti</div>
    </div>
    {% endfor %}
  </div>
</div>

<div class="footer">
  <p>© 2025 GestionaleHR — Made in Italy 🇮🇹 · <a href="/privacy" style="color:rgba(255,255,255,.4)">Privacy</a> · <a href="/termini" style="color:rgba(255,255,255,.4)">Termini</a></p>
</div>
</body>
</html>"""


REGISTRATI_TMPL = """<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Registrati — GestionaleHR</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:linear-gradient(135deg,#0f172a,#0f4c81);min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px}
.card{background:#fff;border-radius:20px;padding:40px;width:100%;max-width:460px;box-shadow:0 20px 60px rgba(0,0,0,.3)}
.logo{text-align:center;font-size:22px;font-weight:900;color:#0f172a;margin-bottom:8px}<i class="fa fa-chart-line" style="color:#f59e0b"></i>
.subtitle{text-align:center;color:#64748b;font-size:14px;margin-bottom:28px}
.form-group{margin-bottom:16px}
label{font-size:13px;font-weight:600;color:#374151;display:block;margin-bottom:5px}
input,select{width:100%;padding:11px 14px;border:1.5px solid #e2e8f0;border-radius:10px;font-size:14px;outline:none;transition:.2s}
input:focus,select:focus{border-color:#f59e0b;box-shadow:0 0 0 3px rgba(245,158,11,.1)}
.btn{width:100%;padding:14px;background:linear-gradient(135deg,#f59e0b,#d97706);color:#fff;border:none;border-radius:12px;font-size:16px;font-weight:700;cursor:pointer;margin-top:8px}
.btn:hover{opacity:.9}
.error{background:#fef2f2;color:#dc2626;border:1px solid #fecaca;border-radius:8px;padding:10px 14px;font-size:13px;margin-bottom:16px}
.trial-badge{text-align:center;background:#f0fdf4;color:#16a34a;border-radius:8px;padding:10px;font-size:13px;font-weight:600;margin-bottom:20px}
.login-link{text-align:center;margin-top:20px;font-size:13px;color:#64748b}
.login-link a{color:#f59e0b;font-weight:700;text-decoration:none}
.piano-row{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:4px}
.piano-opt{border:2px solid #e2e8f0;border-radius:10px;padding:10px 8px;text-align:center;cursor:pointer;transition:.2s}
.piano-opt.sel{border-color:#f59e0b;background:#fffbeb}
.piano-opt .nome{font-weight:700;font-size:13px}
.piano-opt .prezzo{color:#f59e0b;font-weight:800;font-size:15px}
</style>
</head>
<body>
<div class="card">
  <div class="logo"><i class="fa fa-chart-line" style="color:#f59e0b"></i> GestionaleHR</div>
  <div class="subtitle">Crea il tuo account aziendale in 30 secondi</div>
  <div class="trial-badge"><i class="fa fa-gift"></i> 14 giorni gratuiti — Nessuna carta richiesta</div>
  {% if error %}<div class="error"><i class="fa fa-exclamation-circle"></i> {{ error }}</div>{% endif %}
  <form method="POST">
    <div class="form-group">
      <label>Nome azienda *</label>
      <input name="nome_azienda" placeholder="Es. Rossi Costruzioni SRL" required>
    </div>
    <div class="form-group">
      <label>Email admin *</label>
      <input type="email" name="email" placeholder="admin@tuaazienda.it" required>
    </div>
    <div class="form-group">
      <label>Password *</label>
      <input type="password" name="password" placeholder="Minimo 6 caratteri" required>
    </div>
    <div class="form-group">
      <label>Piano</label>
      <select name="piano">
        <option value="base" {{ 'selected' if piano_sel=='base' }}>Base — €29/mese (fino a 5 dip.)</option>
        <option value="professional" {{ 'selected' if piano_sel=='professional' }}>Professional — €59/mese (fino a 20 dip.)</option>
        <option value="enterprise" {{ 'selected' if piano_sel=='enterprise' }}>Enterprise — €99/mese (illimitati)</option>
      </select>
    </div>
    <button type="submit" class="btn"><i class="fa fa-rocket"></i> Crea account — inizia gratis 14 giorni</button>
  </form>
  <div class="login-link">Hai già un account? <a href="/login">Accedi</a></div>
</div>
</body>
</html>"""


ABBONAMENTO_TMPL = """
<div class="card" style="max-width:600px;margin:0 auto">
  <div class="card-header">
    <h3><i class="fa fa-credit-card" style="color:#f59e0b"></i> Il tuo abbonamento</h3>
  </div>
  <div class="card-body">
    <div style="display:flex;justify-content:space-between;align-items:center;background:#f8fafc;border-radius:10px;padding:16px;margin-bottom:20px">
      <div>
        <div style="font-size:12px;color:#64748b;text-transform:uppercase;letter-spacing:.5px">Azienda</div>
        <div style="font-size:18px;font-weight:700">{{ az.nome }}</div>
      </div>
      <div style="text-align:right">
        <div style="font-size:12px;color:#64748b">Stato</div>
        {% if az.stato=='attivo' %}<span class="badge badge-green">✓ Attivo</span>
        {% elif az.stato=='trial' %}<span class="badge badge-amber">⏳ Trial fino al {{ az.trial_fino_al }}</span>
        {% else %}<span class="badge badge-red">✗ Sospeso</span>{% endif %}
      </div>
    </div>
    <div style="font-weight:700;margin-bottom:12px">Piano attuale: <span style="color:#f59e0b;text-transform:capitalize">{{ az.piano }}</span></div>
    {% if az.stato in ('trial','sospeso') %}
    <div style="background:#fef3c7;border-radius:8px;padding:12px;margin-bottom:20px;font-size:13px;color:#92400e">
      <i class="fa fa-info-circle"></i>
      {% if az.stato=='trial' %}Attiva un abbonamento per continuare a usare il gestionale dopo il trial.
      {% else %}Il tuo abbonamento è sospeso. Rinnova per riaccedere.{% endif %}
    </div>
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px">
    {% for p in piani %}
    <form method="POST" action="/abbonamento/checkout">
      <input type="hidden" name="piano" value="{{ p.nome|lower }}">
      <div style="border:2px solid {{ '#f59e0b' if p.nome|lower==az.piano else '#e2e8f0' }};border-radius:12px;padding:16px;text-align:center">
        <div style="font-weight:700">{{ p.nome }}</div>
        <div style="font-size:24px;font-weight:800;color:#f59e0b">€{{ p.prezzo_mensile|int }}</div>
        <div style="font-size:11px;color:#64748b;margin-bottom:12px">/mese</div>
        <button type="submit" class="btn btn-primary" style="width:100%;font-size:13px">
          {{ 'Piano attuale' if p.nome|lower==az.piano else 'Attiva' }}
        </button>
      </div>
    </form>
    {% endfor %}
    </div>
    {% else %}
    <div style="background:#f0fdf4;border-radius:8px;padding:12px;color:#16a34a;font-size:13px">
      <i class="fa fa-check-circle"></i> Il tuo abbonamento è attivo. Gestiscilo dal portale Stripe.
    </div>
    {% endif %}
  </div>
</div>"""


_SA_LOGIN_TMPL = """<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>SuperAdmin</title>
<style>body{font-family:sans-serif;background:#0f172a;display:flex;align-items:center;justify-content:center;min-height:100vh}
.card{background:#1e293b;border-radius:16px;padding:40px;width:360px;color:#fff}
h2{text-align:center;margin-bottom:24px;color:#f59e0b}
input{width:100%;padding:10px;margin-bottom:12px;border:1px solid #334155;border-radius:8px;background:#0f172a;color:#fff;font-size:14px}
button{width:100%;padding:12px;background:#f59e0b;color:#000;border:none;border-radius:8px;font-weight:700;cursor:pointer}
.err{color:#ef4444;font-size:13px;margin-bottom:12px}</style>
</head><body><div class="card">
<h2>🔐 SuperAdmin</h2>
{% if error %}<div class="err">{{ error }}</div>{% endif %}
<form method="POST">
  <input type="email" name="email" placeholder="Email" required>
  <input type="password" name="password" placeholder="Password" required>
  <button type="submit">Accedi</button>
</form>
</div></body></html>"""


_SA_DASHBOARD_TMPL = """
<div style="margin-bottom:24px;display:flex;justify-content:space-between;align-items:center">
  <h2><i class="fa fa-globe" style="color:#f59e0b"></i> SuperAdmin — Aziende clienti</h2>
  <a href="/superadmin/logout" class="btn btn-secondary btn-sm"><i class="fa fa-sign-out"></i> Esci</a>
</div>
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:24px">
  <div class="card" style="padding:20px;text-align:center;border-top:3px solid #6366f1">
    <div style="font-size:28px;font-weight:800">{{ totali.totale }}</div>
    <div style="font-size:12px;color:#64748b">Aziende totali</div>
  </div>
  <div class="card" style="padding:20px;text-align:center;border-top:3px solid #22c55e">
    <div style="font-size:28px;font-weight:800;color:#16a34a">{{ totali.attive }}</div>
    <div style="font-size:12px;color:#64748b">Attive</div>
  </div>
  <div class="card" style="padding:20px;text-align:center;border-top:3px solid #f59e0b">
    <div style="font-size:28px;font-weight:800;color:#d97706">{{ totali.trial }}</div>
    <div style="font-size:12px;color:#64748b">In trial</div>
  </div>
  <div class="card" style="padding:20px;text-align:center;border-top:3px solid #ef4444">
    <div style="font-size:28px;font-weight:800;color:#dc2626">{{ totali.sospese }}</div>
    <div style="font-size:12px;color:#64748b">Sospese</div>
  </div>
</div>
<div class="card" style="padding:0">
  <table style="width:100%;border-collapse:collapse">
    <thead><tr style="background:#0f172a;color:#fff">
      <th style="padding:12px 16px;text-align:left">Azienda</th>
      <th style="padding:12px 16px">Email</th>
      <th style="padding:12px 16px">Piano</th>
      <th style="padding:12px 16px">Stato</th>
      <th style="padding:12px 16px">Registrata</th>
      <th style="padding:12px 16px">Azioni</th>
    </tr></thead>
    <tbody>
    {% for az in aziende %}
    <tr style="border-bottom:1px solid var(--border)">
      <td style="padding:12px 16px;font-weight:700">{{ az.nome }}</td>
      <td style="padding:12px 16px;font-size:13px;color:#64748b">{{ az.email_admin }}</td>
      <td style="padding:12px 16px"><span class="tag" style="text-transform:capitalize">{{ az.piano }}</span></td>
      <td style="padding:12px 16px">
        {% if az.stato=='attivo' %}<span class="badge badge-green">Attivo</span>
        {% elif az.stato=='trial' %}<span class="badge badge-amber">Trial</span>
        {% else %}<span class="badge badge-red">Sospeso</span>{% endif %}
      </td>
      <td style="padding:12px 16px;font-size:12px;font-family:monospace">{{ az.creato_il[:10] }}</td>
      <td style="padding:12px 16px">
        <div style="display:flex;gap:6px">
          <a href="/superadmin/aziende/{{ az.id }}/entra" class="btn btn-sm btn-secondary" title="Entra come admin"><i class="fa fa-sign-in"></i></a>
          {% if az.stato != 'sospeso' %}
          <a href="/superadmin/aziende/{{ az.id }}/sospendi" class="btn btn-sm btn-danger" onclick="return confirm('Sospendere?')"><i class="fa fa-ban"></i></a>
          {% else %}
          <a href="/superadmin/aziende/{{ az.id }}/riattiva" class="btn btn-sm btn-green"><i class="fa fa-check"></i></a>
          {% endif %}
        </div>
      </td>
    </tr>
    {% else %}
    <tr><td colspan="6" style="padding:40px;text-align:center;color:#64748b">Nessuna azienda registrata.</td></tr>
    {% endfor %}
    </tbody>
  </table>
</div>"""

# ── Init master DB all'avvio ──────────────────────────────────
try:
    init_master_db()
except Exception as e:
    print(f'[init_master_db] {e}')


# Esegui migrazioni DB all'avvio (sicuro, usa IF NOT EXISTS)
with app.app_context():
    try:
        init_db()
    except Exception as _e:
        print(f'[INIT_DB] Warning: {_e}')
    # Garantisce le colonne critiche anche se init_db è parzialmente fallita
    try:
        ensure_columns()
    except Exception as _e:
        print(f'[ENSURE_COLUMNS] Warning: {_e}')

# (ensure_columns viene chiamata solo all'avvio, vedi fondo file)

if __name__ == '__main__':
    debug = os.environ.get('RAILWAY_ENVIRONMENT') is None
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=debug)
